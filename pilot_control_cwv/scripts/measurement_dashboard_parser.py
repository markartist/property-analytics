from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Optional
import os
import re
from datetime import datetime

from openpyxl import load_workbook

from measurement_database import (
    ensure_measurement_tables,
    replace_measurement_workbook_rows,
    upsert_measurement_metric_records,
    upsert_measurement_raw_records,
)


MEASUREMENT_DROP_DIR = Path(
    "/Users/mark/Library/CloudStorage/OneDrive-VenterraRealty(Canada)Inc/Guest_Card_Reports"
)
MEASUREMENT_FILE_RE = re.compile(r"^Measurement_Dashboard(?:_(?P<version>[0-9.]+))?\.xlsx$", re.IGNORECASE)

# Backward-compatible fallback for imports that still reference MEASUREMENT_PATH.
MEASUREMENT_PATH = MEASUREMENT_DROP_DIR / "Measurement_Dashboard.xlsx"

PROPERTY_TAB_MAP = {
    "Calais Midtown": "Calais_Midtown",
    "Champions Green": "Champions_Green",
    "The District Universal Boulevard": "The_District",
    "The Harrison": "The_Harrison",
    "Ventana": "Ventana",
}

PROPERTY_ROW_ALIASES = {
    "Calais Midtown": "Calais",
    "Champions Green": "Champions Green",
    "The District Universal Boulevard": "District",
    "The Harrison": "Harrison",
    "Ventana": "Ventana",
    "Avasa Spring Branch": "Spring Branch",
    "Axial Buckhead": "Axial",
    "Northbridge at Millenia Lake": "Northbridge",
    "The Whitney": "Whitney",
    "Park on Wurzbach": "Wurzbach",
}

ALIAS_TO_PROPERTY = {alias: canonical for canonical, alias in PROPERTY_ROW_ALIASES.items()}
PILOT_PROPERTIES = {
    "Calais Midtown",
    "Champions Green",
    "The District Universal Boulevard",
    "The Harrison",
    "Ventana",
}
SISTER_PROPERTIES = {
    "Avasa Spring Branch",
    "Axial Buckhead",
    "Northbridge at Millenia Lake",
    "The Whitney",
    "Park on Wurzbach",
}

METRIC_HEADER_MAP = {
    "HEAP - Prev.Sites - Avg per day Traffic Volume from 2.24.26-3.24.26": {
        "metric_key": "organic_traffic_volume_baseline",
        "metric_label": "Organic Traffic Volume Baseline",
        "value_type": "count",
    },
    "HEAP - Prev Sites Organic Traffic Volume - T7": {
        "metric_key": "organic_traffic_volume_t7",
        "metric_label": "Organic Traffic Volume - T7",
        "value_type": "count",
    },
    "Organic Share - Percent of Unique Users to Total Traffic 2.24.26-3.24.26": {
        "metric_key": "organic_share_baseline",
        "metric_label": "Organic Traffic as a % of Unique Users Baseline",
        "value_type": "percent",
    },
    "HEAP - Prev.Sites T30-%": {
        "metric_key": "organic_share_baseline",
        "metric_label": "Organic Traffic as a % of Unique Users Baseline",
        "value_type": "percent",
    },
    "Organic Share - Percentage of Organic Traffic to Unique Users - T7": {
        "metric_key": "organic_share_t7",
        "metric_label": "Organic Traffic as a % of Unique Users - T7",
        "value_type": "percent",
    },
    "HEAP - Pilot Sites Today T30-%": {
        "metric_key": "organic_share_t7",
        "metric_label": "Organic Traffic as a % of Unique Users - T7",
        "value_type": "percent",
    },
    "YOY - Vol Δ to Organic Traffic Vol": {
        "metric_key": "organic_yoy_vol_delta",
        "metric_label": "YOY - Vol Δ to Organic Traffic Vol",
        "value_type": "percent_delta",
    },
    "GA4 - Prev.Sites YOY - Vol Δ": {
        "metric_key": "organic_yoy_vol_delta",
        "metric_label": "YOY - Vol Δ to Organic Traffic Vol",
        "value_type": "percent_delta",
    },
    "YOY - Vol ∆ - Percent Points to Organic Share": {
        "metric_key": "organic_share_yoy_points_delta",
        "metric_label": "YOY - Vol Δ - Percent Points to Organic Share",
        "value_type": "points",
    },
    "GA4 - Pilot Sites Today YOY-% Δ": {
        "metric_key": "organic_share_yoy_points_delta",
        "metric_label": "YOY - Vol Δ - Percent Points to Organic Share",
        "value_type": "points",
    },
    "HEAP - Organic Traffic for Pilot Sites - Notes": {
        "metric_key": "organic_traffic_notes",
        "metric_label": "Organic Traffic Notes",
        "value_type": "text",
    },
    "Total Number of Leads Generated - T7": {
        "metric_key": "total_leads_generated_t7",
        "metric_label": "Total Number of Leads Generated - T7",
        "value_type": "count",
    },
    "Number of Leads Generated - T7": {
        "metric_key": "total_leads_generated_t7",
        "metric_label": "Total Number of Leads Generated - T7",
        "value_type": "count",
    },
    "CTA - High Intent Click Rate - Clicks/Unique Users 2.24.26-3.42.26": {
        "metric_key": "high_intent_click_rate_baseline",
        "metric_label": "High Intent User Rate Baseline",
        "value_type": "percent",
    },
    "HEAP - Prev Sites - T30": {
        "metric_key": "high_intent_click_rate_baseline",
        "metric_label": "High Intent User Rate Baseline",
        "value_type": "percent",
    },
    "CTA - High Intent Click Rate - Clicks/Unique Users T7": {
        "metric_key": "high_intent_click_rate_t7",
        "metric_label": "High Intent User Rate - T7",
        "value_type": "percent",
    },
    "HEAP - Pilot Sites - Today": {
        "metric_key": "high_intent_click_rate_t7",
        "metric_label": "High Intent User Rate - T7",
        "value_type": "percent",
    },
    "CTA - High Intent Click RateBenchmark": {
        "metric_key": "high_intent_click_rate_benchmark",
        "metric_label": "High Intent Benchmark Note",
        "value_type": "text",
    },
    "HEAP - Pilot Sites - Benchmark": {
        "metric_key": "high_intent_click_rate_benchmark",
        "metric_label": "High Intent Benchmark Note",
        "value_type": "text",
    },
    "Heap - 2.24-3.24 Baseline - Tour": {
        "metric_key": "heap_tour_baseline",
        "metric_label": "Heap Tour Baseline",
        "value_type": "percent",
    },
    "Heap Click Through Rate - Tour": {
        "metric_key": "heap_tour_ctr_t7",
        "metric_label": "Heap CTR - Tour",
        "value_type": "percent",
    },
    "Heap - 2.24-3.24 Baseline - PQ": {
        "metric_key": "heap_pq_baseline",
        "metric_label": "Heap PQ Baseline",
        "value_type": "percent",
    },
    "Heap Click Through Rate - PQ": {
        "metric_key": "heap_pq_ctr_t7",
        "metric_label": "Heap CTR - PQ",
        "value_type": "percent",
    },
    "Heap - 2.24-3.24 Baseline - Apply": {
        "metric_key": "heap_apply_baseline",
        "metric_label": "Heap Apply Baseline",
        "value_type": "percent",
    },
    "Heap Click Through Rate - Apply": {
        "metric_key": "heap_apply_ctr_t7",
        "metric_label": "Heap CTR - Apply",
        "value_type": "percent",
    },
    "Heap - 2.24-3.24 Baseline - Phone": {
        "metric_key": "heap_phone_baseline",
        "metric_label": "Heap Phone Baseline",
        "value_type": "percent",
    },
    "Heap Click Through Rate - Phone": {
        "metric_key": "heap_phone_ctr_t7",
        "metric_label": "Heap CTR - Phone",
        "value_type": "percent",
    },
    "Heap - 2.24-3.24 Baseline - Contact": {
        "metric_key": "heap_contact_baseline",
        "metric_label": "Heap Contact Baseline",
        "value_type": "percent",
    },
    "Heap Click Through Rate - Contact": {
        "metric_key": "heap_contact_ctr_t7",
        "metric_label": "Heap CTR - Contact",
        "value_type": "percent",
    },
}

@dataclass
class HeapStatus:
    latest_sheet: str
    latest_date_label: str
    today_present: bool
    source_file: str


def _version_tuple(path: Path) -> tuple[int, ...]:
    match = MEASUREMENT_FILE_RE.match(path.name)
    if not match or not match.group("version"):
        return (0,)
    return tuple(int(part) for part in match.group("version").split(".") if part.isdigit()) or (0,)


def list_measurement_workbooks(drop_dir: Path = MEASUREMENT_DROP_DIR) -> list[Path]:
    if not drop_dir.exists():
        return []
    return sorted(
        [
            path
            for path in drop_dir.iterdir()
            if path.is_file()
            and not path.name.startswith("~$")
            and MEASUREMENT_FILE_RE.match(path.name)
        ],
        key=lambda path: (_version_tuple(path), path.stat().st_mtime, path.name),
    )


def resolve_measurement_path(explicit_path: Path | None = None) -> Path:
    if explicit_path is not None:
        return Path(explicit_path)

    env_path = os.environ.get("MEASUREMENT_DASHBOARD_PATH")
    if env_path:
        candidate = Path(env_path).expanduser()
        if candidate.exists():
            return candidate

    candidates = list_measurement_workbooks()
    if candidates:
        return candidates[-1]
    return MEASUREMENT_PATH


def _workbook(path: Path | None = None):
    return load_workbook(resolve_measurement_path(path), read_only=True, data_only=True)


def _parse_sheet_date(name: str, assumed_year: int | None = None) -> Optional[datetime]:
    m = re.search(r"(\d{1,2})\.(\d{1,2})\.(\d{2})", name)
    if not m:
        return None
    month, day, yy = m.groups()
    year = assumed_year or (2000 + int(yy))
    return datetime(year, int(month), int(day))


def _daily_sheets(wb) -> list[tuple[datetime, str]]:
    parsed: list[tuple[str, int, int, int]] = []
    for sheet in wb.sheetnames:
        match = re.search(r"(\d{1,2})\.(\d{1,2})\.(\d{2})", sheet)
        if not match:
            continue
        month, day, yy = match.groups()
        parsed.append((sheet, int(month), int(day), 2000 + int(yy)))

    if not parsed:
        return []

    current_year = datetime.now().year
    years = [year for _, _, _, year in parsed]
    dominant_year = max(
        set(years),
        key=lambda year: (years.count(year), year == current_year, year),
    )
    today = datetime.now().date()
    dated: dict[datetime, str] = {}
    for sheet, month, day, year in parsed:
        corrected_year = year
        candidate = datetime(dominant_year, month, day).date()
        if year != dominant_year and abs(year - dominant_year) <= 1 and candidate <= today:
            corrected_year = dominant_year
        dt = datetime(corrected_year, month, day)
        dated[dt] = sheet
    return sorted((dt, sheet) for dt, sheet in dated.items())


def _normalize_header_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").replace("∆", "Δ")
    return re.sub(r"\s+", " ", text).strip()


NORMALIZED_METRIC_HEADER_MAP = {
    _normalize_header_text(header): meta for header, meta in METRIC_HEADER_MAP.items()
}


def _row_role(raw_name: str) -> str:
    if raw_name in {"PILOT AVERAGE", "Pilot Average"}:
        return "pilot_average"
    if raw_name in {"SISTER AVERAGE", "Sister Average"}:
        return "sister_average"
    if raw_name in {"Comparative Results", "Pilot vs. Sister"}:
        return "comparative"
    canonical = ALIAS_TO_PROPERTY.get(raw_name)
    if canonical in PILOT_PROPERTIES:
        return "pilot"
    if canonical in SISTER_PROPERTIES:
        return "sister"
    return "other"


def _canonical_property_name(raw_name: str) -> Optional[str]:
    if raw_name in {"PILOT AVERAGE", "Pilot Average", "SISTER AVERAGE", "Sister Average", "Comparative Results", "Pilot vs. Sister"}:
        return raw_name
    return ALIAS_TO_PROPERTY.get(raw_name)


def _parse_numeric(value: Any, value_type: str) -> Optional[float]:
    if value is None:
        return None
    if value_type == "text":
        return None
    if isinstance(value, (int, float)):
        number = float(value)
    else:
        text = str(value).strip()
        if not text:
            return None
        cleaned = text.replace("%", "")
        cleaned = re.sub(r"[^0-9+\-\.]", "", cleaned)
        if cleaned.count(".") > 1:
            first = cleaned.find(".")
            cleaned = cleaned[: first + 1] + cleaned[first + 1 :].replace(".", "")
        if not cleaned or cleaned in {"-", ".", "-.", "+."}:
            return None
        try:
            number = float(cleaned)
        except ValueError:
            return None
        if "%" in text:
            return number / 100.0
    if value_type in {"percent", "percent_delta"} and abs(number) > 1:
        return number / 100.0
    return number


def parse_measurement_daily_records(path: Path | None = None) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    source_path = resolve_measurement_path(path)
    wb = _workbook(source_path)
    raw_records: list[dict[str, Any]] = []
    metric_records: list[dict[str, Any]] = []
    source_file = str(source_path)

    for dt, sheet in _daily_sheets(wb):
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            continue
        headers = [_normalize_header_text(value) for value in rows[0]]
        sections = rows[1]
        section_labels: list[Optional[str]] = []
        current_section: Optional[str] = None
        for value in sections:
            normalized = _normalize_header_text(value)
            if normalized:
                current_section = normalized
            section_labels.append(current_section)

        snapshot_date = dt.date().isoformat()
        for source_row, row in enumerate(rows[2:], start=3):
            raw_name = row[0]
            if not isinstance(raw_name, str):
                continue
            raw_name = raw_name.strip()
            if not raw_name:
                continue
            role = _row_role(raw_name)
            canonical_name = _canonical_property_name(raw_name)

            for idx, value in enumerate(row, start=1):
                header_raw = headers[idx - 1] if idx - 1 < len(headers) else ""
                if idx == 1 or not header_raw or value is None:
                    continue
                raw_records.append(
                    {
                        "snapshot_date": snapshot_date,
                        "source_file": source_file,
                        "source_sheet": sheet,
                        "source_row": source_row,
                        "source_column": idx,
                        "property_name_raw": raw_name,
                        "row_role": role,
                        "section_label": section_labels[idx - 1] if idx - 1 < len(section_labels) else None,
                        "header_raw": header_raw,
                        "value_raw_text": str(value),
                        "value_numeric": _parse_numeric(value, "raw"),
                    }
                )
                metric_meta = NORMALIZED_METRIC_HEADER_MAP.get(header_raw)
                if not metric_meta:
                    continue
                value_type = metric_meta["value_type"]
                metric_records.append(
                    {
                        "snapshot_date": snapshot_date,
                        "source_file": source_file,
                        "source_sheet": sheet,
                        "property_name_raw": raw_name,
                        "property_name": canonical_name,
                        "row_role": role,
                        "metric_key": metric_meta["metric_key"],
                        "metric_label": metric_meta["metric_label"],
                        "value_type": value_type,
                        "value_numeric": _parse_numeric(value, value_type),
                        "value_text": None if value_type != "text" else str(value).strip(),
                        "header_raw": header_raw,
                    }
                )
    return raw_records, metric_records


def ingest_measurement_workbook(conn, path: Path | None = None) -> tuple[int, int]:
    ensure_measurement_tables(conn)
    source_path = resolve_measurement_path(path)
    raw_records, metric_records = parse_measurement_daily_records(source_path)
    replace_measurement_workbook_rows(conn)
    raw_count = upsert_measurement_raw_records(conn, raw_records)
    metric_count = upsert_measurement_metric_records(conn, metric_records)
    return raw_count, metric_count


def latest_heap_status() -> HeapStatus:
    source_path = resolve_measurement_path()
    wb = _workbook(source_path)
    daily_sheets = _daily_sheets(wb)
    latest_dt, latest_sheet = daily_sheets[-1]
    latest_date_label = latest_dt.strftime("%-m/%-d")
    today_present = latest_dt.date() >= datetime.now().date()
    return HeapStatus(
        latest_sheet=latest_sheet,
        latest_date_label=latest_date_label,
        today_present=today_present,
        source_file=str(source_path),
    )


def _find_header_index(headers: tuple, candidates: list[str]) -> Optional[int]:
    for idx, value in enumerate(headers):
        if value in candidates:
            return idx
    return None


def _normalize_pct(value):
    if value is None:
        return None
    value = float(value)
    return value / 100.0 if value > 1 else value


def organic_heap_series(property_name: str) -> list[tuple[str, float]]:
    alias = PROPERTY_ROW_ALIASES[property_name]
    wb = _workbook()
    series: list[tuple[str, float]] = []
    for dt, sheet in _daily_sheets(wb):
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 3:
            continue
        headers = rows[1]
        current_idx = _find_header_index(
            headers,
            [
                "HEAP - Pilot Sites Today T30-%",
                "Org Unique Users Today  - %",
                "Org Traffic",
                "Org Traffic %",
            ],
        )
        if current_idx is None:
            continue
        target_row = None
        for row in rows[2:]:
            if row and row[0] == alias:
                target_row = row
                break
        if target_row is None or target_row[current_idx] is None:
            continue
        series.append((dt.strftime("%-m/%d"), _normalize_pct(target_row[current_idx])))
    return series


def organic_heap_current_and_baseline(property_name: str) -> tuple[Optional[float], Optional[float]]:
    alias = PROPERTY_ROW_ALIASES[property_name]
    wb = _workbook()
    latest_dt, latest_sheet = _daily_sheets(wb)[-1]
    ws = wb[latest_sheet]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[1]
    current_idx = _find_header_index(
        headers,
        [
            "HEAP - Pilot Sites Today T30-%",
            "Org Unique Users Today  - %",
            "Org Traffic",
            "Org Traffic %",
        ],
    )
    baseline_idx = _find_header_index(
        headers,
        [
            "HEAP - Prev.Sites T30-%",
            "Org Unique Users Baseline - %",
            "Org Unique Users Sister - %",
            "Org Traf Baseline",
        ],
    )
    current = baseline = None
    for row in rows[2:]:
        if row and row[0] == alias:
            if current_idx is not None and row[current_idx] is not None:
                current = _normalize_pct(row[current_idx])
            if baseline_idx is not None and row[baseline_idx] is not None:
                baseline = _normalize_pct(row[baseline_idx])
            break
    return current, baseline


def organic_heap_pair_today(pilot_name: str, sister_name: str) -> tuple[Optional[float], Optional[float], Optional[float], Optional[float]]:
    wb = _workbook()
    latest_dt, latest_sheet = _daily_sheets(wb)[-1]
    ws = wb[latest_sheet]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[1]
    current_idx = _find_header_index(
        headers,
        [
            "HEAP - Pilot Sites Today T30-%",
            "Org Unique Users Today  - %",
            "Org Traffic",
            "Org Traffic %",
        ],
    )
    baseline_idx = _find_header_index(
        headers,
        [
            "HEAP - Prev.Sites T30-%",
            "Org Unique Users Baseline - %",
            "Org Unique Users Sister - %",
            "Org Traf Baseline",
        ],
    )
    pilot_alias = PROPERTY_ROW_ALIASES[pilot_name]
    sister_alias = PROPERTY_ROW_ALIASES[sister_name]
    pilot_prev = pilot_today = sister_prev = sister_today = None
    for row in rows[2:]:
        if not row or row[0] is None:
            continue
        name = row[0]
        if name == pilot_alias:
            if baseline_idx is not None and row[baseline_idx] is not None:
                pilot_prev = _normalize_pct(row[baseline_idx])
            if current_idx is not None and row[current_idx] is not None:
                pilot_today = _normalize_pct(row[current_idx])
        elif name == sister_alias:
            if baseline_idx is not None and row[baseline_idx] is not None:
                sister_prev = _normalize_pct(row[baseline_idx])
            if current_idx is not None and row[current_idx] is not None:
                sister_today = _normalize_pct(row[current_idx])
    return pilot_prev, pilot_today, sister_prev, sister_today
