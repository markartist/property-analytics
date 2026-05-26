#!/usr/bin/env python3
"""
Populate PSI / GTMetrix cells in the measurement dashboard workbook.
"""

from __future__ import annotations

import argparse
import json
import sqlite3
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook

from measurement_dashboard_parser import resolve_measurement_path


ROOT = Path("/Users/mark/Property_Analytics")
DEFAULT_CONFIG_PATH = ROOT / "pilot_control_cwv" / "config" / "pilot_control_cwv_config.json"
DEFAULT_DB_PATH = ROOT / "data" / "portfolio_analytics.db"
DEFAULT_WORKBOOK_PATH = resolve_measurement_path()

PROPERTY_ROW_ALIASES = {
    "Calais Midtown": "Calais",
    "Champions Green": "Champions Green",
    "The District Universal Boulevard": "District",
    "The Harrison": "Harrison",
    "Ventana": "Ventana",
    "Avasa Spring Branch": "Spring Branch",
    "Axial Buckhead": "Axial",
    "Northbridge at Millenia Lake": "Northbridge",
    "The Whitney": "Whitney",
    "Park on Wurzbach": "Wurzbach",
}

PILOT_ROW_RANGE = range(3, 8)
CONTROL_ROW_RANGE = range(10, 15)


def load_config(path: Path) -> Dict[str, object]:
    with path.open() as fh:
        return json.load(fh)


def normalize_sheet_name(sheet_date: str) -> str:
    dt = datetime.strptime(sheet_date, "%Y-%m-%d").date()
    return f"{dt.month}.{dt.day}.{str(dt.year)[2:]}"


def latest_complete_psi_date(conn: sqlite3.Connection, cohort_keys: List[str]) -> Optional[str]:
    placeholders = ",".join("?" for _ in cohort_keys)
    row = conn.execute(
        f"""
        SELECT metric_date
        FROM pilot_control_psi_metrics
        WHERE strategy = 'mobile'
          AND cohort_key IN ({placeholders})
        GROUP BY metric_date
        HAVING COUNT(DISTINCT cohort_key) >= ?
        ORDER BY metric_date DESC
        LIMIT 1
        """,
        [*cohort_keys, len(cohort_keys)],
    ).fetchone()
    return row[0] if row else None


def latest_complete_gt_date(conn: sqlite3.Connection, property_ids: List[str]) -> Optional[str]:
    placeholders = ",".join("?" for _ in property_ids)
    row = conn.execute(
        f"""
        SELECT metric_date
        FROM gtmetrix_metrics
        WHERE property_id IN ({placeholders})
        GROUP BY metric_date
        HAVING COUNT(DISTINCT property_id) >= ?
        ORDER BY metric_date DESC
        LIMIT 1
        """,
        [*property_ids, len(property_ids)],
    ).fetchone()
    return row[0] if row else None


def latest_psi_scores(conn: sqlite3.Connection, cohort_keys: List[str]) -> Tuple[Optional[str], Dict[str, float]]:
    placeholders = ",".join("?" for _ in cohort_keys)
    metric_date = latest_complete_psi_date(conn, cohort_keys)
    if not metric_date:
        return None, {}
    rows = conn.execute(
        f"""
        SELECT cohort_key, performance_score
        FROM pilot_control_psi_metrics
        WHERE strategy = 'mobile'
          AND metric_date = ?
          AND cohort_key IN ({placeholders})
        """,
        [metric_date, *cohort_keys],
    ).fetchall()
    return metric_date, {row[0]: float(row[1]) for row in rows if row[1] is not None}


def latest_gt_scores(conn: sqlite3.Connection, property_ids: List[str]) -> Tuple[Optional[str], Dict[str, float]]:
    placeholders = ",".join("?" for _ in property_ids)
    metric_date = latest_complete_gt_date(conn, property_ids)
    if not metric_date:
        return None, {}
    rows = conn.execute(
        f"""
        SELECT property_id, pagespeed_score
        FROM gtmetrix_metrics
        WHERE metric_date = ?
          AND property_id IN ({placeholders})
        """,
        [metric_date, *property_ids],
    ).fetchall()
    return metric_date, {str(row[0]): float(row[1]) for row in rows if row[1] is not None}


def row_map(ws, rows: range) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for row_idx in rows:
        label = ws[f"A{row_idx}"].value
        if label:
            mapping[str(label).strip()] = row_idx
    return mapping


def write_scores(
    ws,
    row_mapping: Dict[str, int],
    display_to_score: Dict[str, Optional[float]],
    target_column: str,
) -> List[Tuple[str, str, Optional[float]]]:
    written: List[Tuple[str, str, Optional[float]]] = []
    for display_name, score in display_to_score.items():
        alias = PROPERTY_ROW_ALIASES[display_name]
        row_idx = row_mapping[alias]
        ws[f"{target_column}{row_idx}"] = score
        written.append((display_name, f"{target_column}{row_idx}", score))
    return written


def main() -> int:
    parser = argparse.ArgumentParser(description="Update measurement dashboard CWV cells")
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK_PATH), help="Workbook path")
    parser.add_argument(
        "--sheet-date",
        default=(date.today() - timedelta(days=1)).isoformat(),
        help="Workbook sheet date YYYY-MM-DD; defaults to yesterday",
    )
    parser.add_argument("--config", default=str(DEFAULT_CONFIG_PATH), help="Pilot/control config path")
    parser.add_argument("--db", default=str(DEFAULT_DB_PATH), help="SQLite DB path")
    parser.add_argument("--gt-test-value", type=float, help="Optional test value to write into GT cells")
    args = parser.parse_args()

    workbook_path = Path(args.workbook)
    config = load_config(Path(args.config))
    sheet_name = normalize_sheet_name(args.sheet_date)

    conn = sqlite3.connect(Path(args.db))
    cohort_rows = [row for row in config["cohorts"] if row.get("active", True)]
    psi_date, psi_scores = latest_psi_scores(conn, [row["key"] for row in cohort_rows])
    gt_date, gt_scores = latest_gt_scores(conn, [str(row["property_id"]) for row in cohort_rows])
    conn.close()

    wb = load_workbook(workbook_path)
    ws = wb[sheet_name]

    pilot_rows = row_map(ws, PILOT_ROW_RANGE)
    control_rows = row_map(ws, CONTROL_ROW_RANGE)

    pilot_scores = {
        row["display_name"]: psi_scores.get(row["key"])
        for row in cohort_rows
        if row["role"] == "pilot"
    }
    control_scores = {
        row["display_name"]: psi_scores.get(row["key"])
        for row in cohort_rows
        if row["role"] == "control"
    }
    pilot_gt_scores = {
        row["display_name"]: (args.gt_test_value if args.gt_test_value is not None else round(gt_scores.get(str(row["property_id"]))) if gt_scores.get(str(row["property_id"])) is not None else None)
        for row in cohort_rows
        if row["role"] == "pilot"
    }
    control_gt_scores = {
        row["display_name"]: (args.gt_test_value if args.gt_test_value is not None else round(gt_scores.get(str(row["property_id"]))) if gt_scores.get(str(row["property_id"])) is not None else None)
        for row in cohort_rows
        if row["role"] == "control"
    }

    written = []
    written.extend(write_scores(ws, pilot_rows, pilot_scores, "C"))
    written.extend(write_scores(ws, control_rows, control_scores, "C"))
    written.extend(write_scores(ws, pilot_rows, pilot_gt_scores, "E"))
    written.extend(write_scores(ws, control_rows, control_gt_scores, "E"))

    wb.save(workbook_path)

    print(f"Updated workbook: {workbook_path}")
    print(f"Sheet: {sheet_name}")
    print(f"PSI source date: {psi_date}")
    print(f"GT source date:  {gt_date if args.gt_test_value is None else 'test override'}")
    print("Written cells:")
    for name, cell, score in written:
        print(f"  {cell} {name}: {score}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
