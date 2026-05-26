#!/usr/bin/env python3
from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path("/Users/mark/Property_Analytics")
REPORTS = ROOT / "pilot_control_cwv" / "reports"
STAMP = datetime.now().strftime("%Y-%m-%d")
MATRIX_JSON = REPORTS / f"calais_cross_source_matrix_{STAMP}.json"
OUTPUT = REPORTS / f"Pilot_Diagnostic_Package_{STAMP}.xlsx"

NAVY = "15284B"
BAY = "294782"
SAN_MARINO = "3D66B9"
MONTE_CARLO = "7DCAC2"
WHITE_SMOKE = "F6F6F5"
QUILL_GRAY = "D6D6D2"
TERRA_COTTA = "BD4830"
BLUE_CHILL = "358279"
LIGHT_BLUE = "D9E2F3"

THIN_GRAY = Side(style="thin", color=QUILL_GRAY)
BOX_BORDER = Border(left=THIN_GRAY, right=THIN_GRAY, top=THIN_GRAY, bottom=THIN_GRAY)


def load_payload() -> dict:
    return json.loads(MATRIX_JSON.read_text())


def set_col_widths(ws, widths: dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def style_header_cell(cell, fill=NAVY, color="FFFFFF", size=20) -> None:
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = Font(name="Aptos", size=size, bold=True, color=color)
    cell.alignment = Alignment(vertical="center")


def style_section_title(cell) -> None:
    cell.font = Font(name="Aptos", size=14, bold=True, color=MONTE_CARLO)
    cell.alignment = Alignment(vertical="center")


def style_label(cell, bold=False, color=NAVY):
    cell.font = Font(name="Aptos", size=11, bold=bold, color=color)
    cell.alignment = Alignment(vertical="top", wrap_text=True)


def style_value(cell, fmt=None, color="000000", bold=False):
    cell.font = Font(name="Aptos", size=11, bold=bold, color=color)
    cell.alignment = Alignment(horizontal="right", vertical="center")
    if fmt:
        cell.number_format = fmt


def add_summary_sheet(wb: Workbook, payload: dict) -> None:
    ws = wb.active
    ws.title = "Executive Summary"
    set_col_widths(ws, {"A": 28, "B": 22, "C": 22, "D": 22, "E": 22, "F": 22})

    ws.merge_cells("A1:F2")
    ws["A1"] = "Pilot Diagnostic Executive Package"
    style_header_cell(ws["A1"], fill=NAVY, size=22)
    ws["A3"] = "Cross-source diagnostic comparison using GA4, GSC, PSI, GTMetrix, BrowserStack / EVS, and structural comparator audit."
    ws["A3"].font = Font(name="Aptos", size=12, color="5B6575")
    ws["A3"].alignment = Alignment(wrap_text=True)

    row = 5
    ws[f"A{row}"] = "Key Findings"
    style_section_title(ws[f"A{row}"])
    row += 1
    for finding in payload["findings"]:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1, value=f"• {finding}")
        ws.cell(row=row, column=1).font = Font(name="Aptos", size=11, color=NAVY)
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
        row += 2

    ws[f"A{row}"] = "Property Summary"
    style_section_title(ws[f"A{row}"])
    row += 1

    headers = [
        "Property",
        "GA4 Organic 30d",
        "GSC Clicks 30d",
        "PSI",
        "GTMetrix",
        "EVS",
    ]
    for idx, header in enumerate(headers, start=1):
        c = ws.cell(row=row, column=idx, value=header)
        c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        c.font = Font(name="Aptos", size=11, bold=True, color=NAVY)
        c.border = BOX_BORDER
    row += 1

    for report in payload["reports"]:
        values = [
            report["property_name"],
            report["ga_30d_organic"],
            report["gsc_30d_clicks"],
            report["psi_performance"],
            report["gt_score"],
            report["evs_status"],
        ]
        for idx, value in enumerate(values, start=1):
            c = ws.cell(row=row, column=idx, value=value)
            c.border = BOX_BORDER
            if idx == 1:
                style_label(c, bold=True)
            elif idx == 6:
                style_value(c, color=BLUE_CHILL if str(value).lower() == "pass" else TERRA_COTTA, bold=True)
            else:
                style_value(c)
        row += 1


def add_matrix_sheet(wb: Workbook, payload: dict) -> None:
    ws = wb.create_sheet("Cross-Source Matrix")
    set_col_widths(ws, {"A": 34, "B": 16, "C": 16, "D": 16, "E": 16, "F": 14, "G": 16, "H": 16, "I": 18})

    ws.merge_cells("A1:I2")
    ws["A1"] = "Calais vs District vs Harrison"
    style_header_cell(ws["A1"], fill=NAVY, size=20)

    headers = [
        "Property",
        "GA4 Organic 30d",
        "GA4 Organic Prior",
        "GA4 Organic Δ%",
        "GSC Clicks 30d",
        "GSC Clicks Prior",
        "GSC Clicks Δ%",
        "Apartments Links",
        "Unit Detail Sample",
    ]
    row = 4
    for idx, header in enumerate(headers, start=1):
        c = ws.cell(row=row, column=idx, value=header)
        c.fill = PatternFill("solid", fgColor=BAY)
        c.font = Font(name="Aptos", size=11, bold=True, color="FFFFFF")
        c.border = BOX_BORDER
    row += 1

    for report in payload["reports"]:
        values = [
            report["property_name"],
            report["ga_30d_organic"],
            report["ga_30d_organic_prev"],
            report["ga_30d_organic_delta_pct"],
            report["gsc_30d_clicks"],
            report["gsc_30d_clicks_prev"],
            report["gsc_30d_clicks_delta_pct"],
            report["apartments_internal_links"],
            report["unit_detail_url"],
        ]
        for idx, value in enumerate(values, start=1):
            c = ws.cell(row=row, column=idx, value=value)
            c.border = BOX_BORDER
            if idx in (1, 9):
                style_label(c, bold=(idx == 1))
            elif idx in (4, 7):
                style_value(c, fmt='0.0%;-0.0%')
            else:
                style_value(c)
        row += 1


def add_signal_sheet(wb: Workbook, payload: dict) -> None:
    ws = wb.create_sheet("Signal Detail")
    set_col_widths(ws, {"A": 34, "B": 18, "C": 18, "D": 18, "E": 18, "F": 18})
    ws.merge_cells("A1:F2")
    ws["A1"] = "Performance + Search + Traffic Signals"
    style_header_cell(ws["A1"], fill=NAVY, size=20)

    row = 4
    headers = ["Property", "PSI", "LCP", "CLS", "TTFB", "GTMetrix"]
    for idx, header in enumerate(headers, start=1):
        c = ws.cell(row=row, column=idx, value=header)
        c.fill = PatternFill("solid", fgColor=BAY)
        c.font = Font(name="Aptos", size=11, bold=True, color="FFFFFF")
        c.border = BOX_BORDER
    row += 1

    for report in payload["reports"]:
        values = [
            report["property_name"],
            report["psi_performance"],
            report["psi_lcp"],
            report["psi_cls"],
            report["psi_ttfb"],
            report["gt_score"],
        ]
        for idx, value in enumerate(values, start=1):
            c = ws.cell(row=row, column=idx, value=value)
            c.border = BOX_BORDER
            if idx == 1:
                style_label(c, bold=True)
            else:
                style_value(c)
        row += 1


def main() -> None:
    payload = load_payload()
    wb = Workbook()
    add_summary_sheet(wb, payload)
    add_matrix_sheet(wb, payload)
    add_signal_sheet(wb, payload)
    wb.save(OUTPUT)
    print(OUTPUT)


if __name__ == "__main__":
    main()
