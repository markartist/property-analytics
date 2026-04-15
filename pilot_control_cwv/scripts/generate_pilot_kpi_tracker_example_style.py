#!/usr/bin/env python3
"""
Generate a lean pilot KPI tracker workbook that visually mirrors the supplied example.

This version is presentation-first:
- keeps the example tab structure
- uses live PSI / GTMetrix / GA4 organic data where available
- leaves explicit placeholders for Heap / Snowflake / BI-fed metrics
"""

from __future__ import annotations

import json
import sqlite3
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


BASE_DIR = Path(__file__).resolve().parents[1]
DEFAULT_CONFIG_PATH = BASE_DIR / "config" / "pilot_control_cwv_config.json"
DB_PATH = Path("/Users/mark/Property_Analytics/data/portfolio_analytics.db")
OUTPUT_DIR = BASE_DIR / "reports"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


@dataclass
class CohortEntry:
    key: str
    display_name: str
    role: str
    property_id: str
    sister_key: Optional[str]


def load_entries(path: Path) -> List[CohortEntry]:
    with path.open() as fh:
        config = json.load(fh)
    rows: List[CohortEntry] = []
    for row in config.get("cohorts", []):
        if not row.get("active", True):
            continue
        rows.append(
            CohortEntry(
                key=row["key"],
                display_name=row["display_name"],
                role=row["role"],
                property_id=row.get("property_id", ""),
                sister_key=row.get("sister_key"),
            )
        )
    return rows


def latest_metrics(conn: sqlite3.Connection) -> Dict[str, dict]:
    conn.row_factory = sqlite3.Row
    query = """
    WITH latest_psi AS (
      SELECT cohort_key, MAX(metric_date) AS metric_date
      FROM pilot_control_psi_metrics
      WHERE strategy='mobile'
      GROUP BY cohort_key
    ),
    psi AS (
      SELECT p.cohort_key, p.metric_date, p.performance_score
      FROM pilot_control_psi_metrics p
      JOIN latest_psi l
        ON l.cohort_key=p.cohort_key
       AND l.metric_date=p.metric_date
      WHERE p.strategy='mobile'
    ),
    latest_gt AS (
      SELECT property_id, MAX(metric_date) AS metric_date
      FROM gtmetrix_metrics
      GROUP BY property_id
    ),
    gt AS (
      SELECT g.property_id, g.metric_date, g.pagespeed_score
      FROM gtmetrix_metrics g
      JOIN latest_gt l
        ON l.property_id=g.property_id
       AND l.metric_date=g.metric_date
    ),
    latest_ga AS (
      SELECT property_id, MAX(metric_date) AS metric_date
      FROM ga4_traffic_sources
      GROUP BY property_id
    ),
    ga AS (
      SELECT
        g.property_id,
        g.metric_date,
        SUM(CASE WHEN g.channel_group='Organic Search' THEN g.sessions ELSE 0 END) AS organic_sessions,
        SUM(g.sessions) AS total_sessions
      FROM ga4_traffic_sources g
      JOIN latest_ga l
        ON l.property_id=g.property_id
       AND l.metric_date=g.metric_date
      GROUP BY g.property_id, g.metric_date
    ),
    latest_gc AS (
      SELECT MAX(run_date) AS run_date
      FROM guest_card_metrics
    ),
    gc AS (
      SELECT property_name, run_date, gc_this_period, quotes_this_period, ipt_appt_this_period, apps_this_period
      FROM guest_card_metrics
      WHERE run_date=(SELECT run_date FROM latest_gc)
    )
    SELECT
      c.key,
      c.display_name,
      c.role,
      c.property_id,
      psi.metric_date AS psi_date,
      psi.performance_score AS psi_score,
      gt.metric_date AS gt_date,
      gt.pagespeed_score AS gt_score,
      ga.metric_date AS ga_date,
      ga.organic_sessions,
      ga.total_sessions,
      gc.run_date AS gc_date,
      gc.gc_this_period,
      gc.quotes_this_period,
      gc.ipt_appt_this_period,
      gc.apps_this_period
    FROM (
      SELECT ? AS key, ? AS display_name, ? AS role, ? AS property_id
    ) c
    LEFT JOIN psi ON psi.cohort_key=c.key
    LEFT JOIN gt ON gt.property_id=c.property_id
    LEFT JOIN ga ON ga.property_id=c.property_id
    LEFT JOIN gc ON gc.property_name=c.display_name
    """
    return query


def fetch_entry_metrics(conn: sqlite3.Connection, entries: List[CohortEntry]) -> Dict[str, dict]:
    conn.row_factory = sqlite3.Row
    sql = latest_metrics(conn)
    data: Dict[str, dict] = {}
    for entry in entries:
        row = conn.execute(sql, (entry.key, entry.display_name, entry.role, entry.property_id)).fetchone()
        data[entry.key] = dict(row) if row else {}
    return data


def latest_dates(conn: sqlite3.Connection) -> dict:
    return {
        "psi_date": conn.execute("SELECT MAX(metric_date) FROM pilot_control_psi_metrics WHERE strategy='mobile'").fetchone()[0],
        "gt_date": conn.execute("SELECT MAX(metric_date) FROM gtmetrix_metrics").fetchone()[0],
        "ga_date": conn.execute("SELECT MAX(metric_date) FROM ga4_traffic_sources").fetchone()[0],
        "gc_date": conn.execute("SELECT MAX(run_date) FROM guest_card_metrics").fetchone()[0],
    }


def apply_sheet_theme(ws) -> None:
    navy = "17375E"
    light = "DCE6F1"
    accent = "EAF2F8"
    thin = Side(style="thin", color="B8CCE4")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.sheet_view.showGridLines = False
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in range(2, ws.max_row + 1, 2):
        for cell in ws[row]:
            if cell.value not in (None, ""):
                cell.fill = PatternFill("solid", fgColor=accent)

    ws.freeze_panes = "A2"


def set_widths(ws, widths: List[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        col = ""
        n = idx
        while n:
            n, rem = divmod(n - 1, 26)
            col = chr(65 + rem) + col
        ws.column_dimensions[col].width = width


def build_exec_dashboard(ws, dates: dict) -> None:
    ws.append(
        [
            "Metric",
            "Definition",
            "Why It Matters",
            "General Expectation",
            "Current System of Record",
            "T12 Weeks Performance",
            "7-15 Days",
            "15-30 Days",
            "30-60 Days",
            "60-90 Days",
        ]
    )

    rows = [
        (
            "CWV / Technical Health",
            "Composite of PSI, GTMetrix, and future CWV pass rate.",
            "SEO and UX structural risk. If this degrades, traffic and conversion pressure usually follows.",
            "Short-term volatility is acceptable, but performance should stabilize quickly.",
            f"PSI + GTMetrix (live). Latest PSI {dates['psi_date']}, GT {dates['gt_date']}.",
            "Baseline logic still to be finalized.",
            "First 1-2 weeks may fluctuate.",
            "Continuous improvement or stabilization expected.",
            "Stabilize with no sustained red flags.",
            "Sustained technical stability as traffic scales.",
        ),
        (
            "Organic Traffic",
            "Total Organic Search sessions to pilot properties vs baseline and sister set.",
            "Structural SEO risk indicator. If URL/canonical/indexing breaks, this shows up early.",
            "Some early volatility; by day 30 should be within expectation and seasonality-adjusted trend.",
            f"GA4 Organic Search. Latest {dates['ga_date']}.",
            "T12 / YoY baseline to be finalized.",
            "Short 7-14 day volatility window possible.",
            "Within +/-5% of baseline expectation.",
            "Should reflect seasonal demand lift.",
            "If it lags market trend materially, that is structural concern.",
        ),
        (
            "High Intent User Rate (HEAP CTR)",
            "Percent of unique users generating a high-intent CTA interaction.",
            "UX efficiency risk indicator independent of traffic volume.",
            "Minor dip possible early if CTA patterns change.",
            "Heap placeholder",
            "Pending Heap feed",
            "Minor dip possible first 7-10 days.",
            "Within +/-5% of baseline.",
            "Modest lift if UX improves.",
            "Clear signal of UX strength if sustained.",
        ),
        (
            "Lead (Guest Card) to Available Unit Rate",
            "Website leads divided by available units.",
            "Revenue-pressure context signal that blends demand and inventory.",
            "Should remain directionally stable with modest seasonal lift.",
            f"BI / guest cards latest {dates['gc_date']}; inventory feed pending final alignment.",
            "Baseline pending final source alignment.",
            "Within +/-5% of baseline.",
            "Slight lift expected.",
            "Lift maintained.",
            "Lift expected in stronger seasonality.",
        ),
        (
            "Website Sales Funnel Conversions",
            "Guest card to quote / tour / application / move-in performance.",
            "Confirms lead quality did not degrade after launch.",
            "Noise is okay early; by day 30 this should stabilize or improve.",
            "Snowflake / Heap / BI placeholder",
            "Pending funnel feed",
            "Stabilization or increase expected.",
            "Watch GC/Quote and GC/App rates.",
            "Watch GC/App and GC/Lease rate.",
            "Watch GC/Move-In rate.",
        ),
    ]
    for row in rows:
        ws.append(list(row))

    apply_sheet_theme(ws)
    set_widths(ws, [28, 26, 42, 42, 28, 20, 20, 20, 20, 24])


def build_heap_sheet(ws, pilots: List[CohortEntry], metrics: Dict[str, dict], dates: dict) -> None:
    ws["B1"] = "Launch Dashboard"
    ws["B1"].font = Font(size=14, bold=True)
    ws["B3"] = "Selected Date"
    ws["C3"] = dates["ga_date"]
    ws["E3"] = "Overall Status"
    ws["F3"] = "PLACEHOLDER"

    headers = [
        "Property",
        "Unique Users",
        "Clicks",
        "CTR",
        "PP Baseline",
        "Flag",
        "Source Note",
    ]
    row = 15
    for idx, header in enumerate(headers, start=2):
        ws.cell(row=row, column=idx, value=header)
    for pilot in pilots:
        row += 1
        current = metrics.get(pilot.key, {})
        organic = current.get("organic_sessions")
        clicks = ""
        ctr = ""
        ws.cell(row=row, column=2, value=pilot.display_name)
        ws.cell(row=row, column=3, value=organic if organic is not None else "")
        ws.cell(row=row, column=4, value=clicks)
        ws.cell(row=row, column=5, value=ctr)
        ws.cell(row=row, column=6, value="Pending Heap baseline")
        ws.cell(row=row, column=7, value="PLACEHOLDER")
        ws.cell(row=row, column=8, value="Use Organic Search users as temporary proxy only if needed; final feed should come from Heap.")

    ws["B5"] = "Unique Users"
    ws["F5"] = "Clicks"
    ws["J5"] = "CTR"
    ws["B13"] = "Last 14 days trend (relative to Selected Date)"
    ws["B14"] = "This tab intentionally mirrors the supplied example, but most fields remain placeholders until Heap naming and extract logic are locked."
    set_widths(ws, [4, 28, 14, 14, 12, 16, 12, 42, 12, 12, 12, 12])


def build_operations_sheet(ws, entries: List[CohortEntry], metrics: Dict[str, dict], dates: dict) -> None:
    ws["B1"] = "Launch Dashboard"
    ws["B1"].font = Font(size=14, bold=True)
    ws["B3"] = "Selected Date"
    ws["C3"] = dates["gc_date"]
    ws["E3"] = "Overall Status"
    ws["F3"] = "PARTIAL"

    ws["B5"] = "Unique Users"
    ws["F5"] = "Website Guest Cards"
    ws["J5"] = "Conversion Rate"
    ws["N5"] = "Guest Cards"
    ws["B15"] = "Guest Cards / Available Unit"
    ws["H15"] = "Guest Cards / Application"
    ws["N15"] = "ORGANIC"
    ws["R15"] = "DIRECT"

    headers_row = 22
    headers = [
        "Property",
        "Organic Sessions",
        "Guest Cards",
        "Quotes",
        "Tours",
        "Applications",
        "GC / App",
        "Status",
        "Source Note",
    ]
    for idx, header in enumerate(headers, start=2):
        ws.cell(row=headers_row, column=idx, value=header)
    for entry in entries:
        headers_row += 1
        current = metrics.get(entry.key, {})
        guest_cards = current.get("gc_this_period")
        apps = current.get("apps_this_period")
        gc_app = ""
        if guest_cards not in (None, "", 0) and apps not in (None, ""):
            gc_app = apps / guest_cards
        status = "LIVE" if guest_cards is not None else "PLACEHOLDER"
        note = "Guest card snapshot available." if guest_cards is not None else "No guest-card row found in latest extract."
        ws.cell(row=headers_row, column=2, value=entry.display_name)
        ws.cell(row=headers_row, column=3, value=current.get("organic_sessions") or "")
        ws.cell(row=headers_row, column=4, value=guest_cards or "")
        ws.cell(row=headers_row, column=5, value=current.get("quotes_this_period") or "")
        ws.cell(row=headers_row, column=6, value=current.get("ipt_appt_this_period") or "")
        ws.cell(row=headers_row, column=7, value=apps or "")
        ws.cell(row=headers_row, column=8, value=gc_app)
        ws.cell(row=headers_row, column=9, value=status)
        ws.cell(row=headers_row, column=10, value=note)

    set_widths(ws, [4, 28, 14, 12, 12, 12, 12, 12, 12, 34])


def build_metric_measurement(ws, pilots: List[CohortEntry], by_key: Dict[str, CohortEntry], metrics: Dict[str, dict]) -> None:
    ws.append(
        [
            "Metric",
            "All Traffic",
            "Organic Traffic",
            "High Intent User Rate (HEAP CTR)",
            "Website-Sourced Lead Volume (Guest Cards)",
            "Lead (Guest Card) to Available Unit Rate",
            "Website Sales Funnel Conversions",
            "CWV - Mobile Pass Rate",
        ]
    )
    metric_rows = [
        "Baseline",
        "Rolling T12",
        "Sister Property",
        "7-15 Days",
        "Variance from Baseline",
        "15-30 Days",
        "Variance from Baseline",
        "30-45 Days",
        "Variance from Baseline",
        "Go / No-Go",
    ]
    for pilot in pilots:
        ws.append([pilot.display_name, "", "", "", "", "", "", ""])
        current = metrics.get(pilot.key, {})
        sister = by_key.get(pilot.sister_key or "")
        sister_name = sister.display_name if sister else ""
        values = {
            "Baseline": "Pending final baseline logic",
            "Rolling T12": "Pending final baseline logic",
            "Sister Property": sister_name,
            "7-15 Days": current.get("organic_sessions") or "",
            "Variance from Baseline": "",
            "15-30 Days": "",
            "30-45 Days": "",
            "Go / No-Go": "Pending",
        }
        for label in metric_rows:
            row = [label]
            if label == "Sister Property":
                row += ["", sister_name, "Heap placeholder", current.get("gc_this_period") or "", "Inventory placeholder", "Funnel placeholder", current.get("psi_score") or ""]
            elif label == "7-15 Days":
                row += [
                    current.get("total_sessions") or "",
                    current.get("organic_sessions") or "",
                    "PLACEHOLDER",
                    current.get("gc_this_period") or "",
                    "PLACEHOLDER",
                    "PLACEHOLDER",
                    current.get("psi_score") or "",
                ]
            else:
                row += [values.get(label, "")] + [""] * 6
            ws.append(row)
        ws.append([""] * 8)
    apply_sheet_theme(ws)
    set_widths(ws, [28, 14, 14, 24, 18, 18, 18, 18])


def build_glossary(ws) -> None:
    ws.append(
        [
            "Term",
            "Global Definition",
            "Venterra Definition & Application (Platform Context)",
            "7-15 Day Expectation (Dip Phase)",
            "30-45 Day Expectation (Stabilization)",
            "90 Day Expectation (Acceleration)",
            "Action if Below Expectation",
        ]
    )
    rows = [
        (
            "Core Web Vitals / Technical Health",
            "Google / web performance health indicators.",
            "Use PSI and GTMetrix now; add CWV pass rate when final source is agreed.",
            "Some volatility accepted after launch.",
            "Stabilize with no sustained degradation.",
            "Stable technical health at scale.",
            "Review image delivery, scripts, and template performance.",
        ),
        (
            "Organic Traffic",
            "Sessions from organic search acquisition.",
            "Treat Organic Search as the true organic source for this workbook.",
            "Short volatility window possible.",
            "Return to baseline-adjusted trend.",
            "Should outperform if SEO structure is healthy.",
            "Check indexing, canonicals, redirects, and crawl path changes.",
        ),
        (
            "High Intent User Rate",
            "Share of users performing high-intent CTA actions.",
            "Primary engagement KPI once Heap feed is live.",
            "Minor early dip possible.",
            "Should recover to baseline or modest lift.",
            "Sustained lift if UX improves.",
            "Check CTA prominence, friction, and event naming integrity.",
        ),
        (
            "Lead (Guest Card) to Available Unit Rate",
            "Website leads divided by available units.",
            "Revenue context metric combining website demand and inventory pressure.",
            "Directional stability expected.",
            "Slight lift possible.",
            "Lift expected if site quality and seasonality align.",
            "Verify inventory source and lead attribution.",
        ),
    ]
    for row in rows:
        ws.append(list(row))
    apply_sheet_theme(ws)
    set_widths(ws, [22, 28, 40, 28, 28, 28, 34])


def build_workbook(entries: List[CohortEntry], metrics: Dict[str, dict], dates: dict) -> Workbook:
    wb = Workbook()
    exec_ws = wb.active
    exec_ws.title = "Exec Dashboard"
    build_exec_dashboard(exec_ws, dates)

    pilots = [entry for entry in entries if entry.role == "pilot"]
    by_key = {entry.key: entry for entry in entries}

    heap_ws = wb.create_sheet("HEAP")
    build_heap_sheet(heap_ws, pilots, metrics, dates)

    ops_ws = wb.create_sheet("Operations")
    build_operations_sheet(ops_ws, entries, metrics, dates)

    mm_ws = wb.create_sheet("Metric Measurement")
    build_metric_measurement(mm_ws, pilots, by_key, metrics)

    glossary_ws = wb.create_sheet("Glossary")
    build_glossary(glossary_ws)
    return wb


def main() -> int:
    entries = load_entries(DEFAULT_CONFIG_PATH)
    with sqlite3.connect(DB_PATH) as conn:
        metrics = fetch_entry_metrics(conn, entries)
        dates = latest_dates(conn)

    workbook = build_workbook(entries, metrics, dates)
    out_path = OUTPUT_DIR / f"Pilot_KPI_Tracker_Example_Style_{date.today().isoformat()}.xlsx"
    workbook.save(out_path)
    print(f"Saved workbook: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
