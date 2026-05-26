from __future__ import annotations

import json
import sqlite3
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.axis import ChartLines
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path("/Users/mark/Property_Analytics")
DB_PATH = ROOT / "data" / "portfolio_analytics.db"
CONFIG_PATH = ROOT / "pilot_control_cwv" / "config" / "pilot_control_cwv_config.json"
OUTPUT_PATH = (
    ROOT
    / "pilot_control_cwv"
    / "reports"
    / f"CWV_Section_Prototype_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
)

PILOT_COLOR = "4473D0"
SISTER_COLOR = "7CCAC2"
BASELINE_COLOR = "A3A3A3"
FLOOR_COLOR = "F2A7A7"
CARD_FILL = "F8FAFC"
TITLE_FILL = "EAF2FF"
TEXT_BLACK = "0F172A"

PSI_BASELINE = 90
PSI_FLOOR = 60
GTM_BASELINE = 94
GTM_FLOOR = 70
REPORT_DATES = ["2026-03-26", "2026-03-27", "2026-03-28", "2026-03-29", "2026-03-30", "2026-03-31"]


def load_pairs() -> list[dict]:
    cfg = json.loads(CONFIG_PATH.read_text())
    cohorts = {c["key"]: c for c in cfg["cohorts"] if c.get("active")}
    pairs = []
    for cohort in cohorts.values():
        if cohort["role"] != "pilot":
            continue
        pairs.append({"pilot": cohort, "sister": cohorts[cohort["sister_key"]]})
    return pairs


def fetch_series(conn: sqlite3.Connection, property_id: str, kind: str) -> list[tuple[str, float]]:
    if kind == "psi":
        query = """
            SELECT metric_date, performance_score
            FROM pilot_control_psi_metrics
            WHERE property_id = ? AND strategy = 'mobile'
            ORDER BY metric_date DESC
            LIMIT 7
        """
    else:
        query = """
            SELECT metric_date, pagespeed_score
            FROM gtmetrix_metrics
            WHERE property_id = ?
            ORDER BY metric_date DESC
            LIMIT 7
        """
    rows = conn.execute(query, (property_id,)).fetchall()
    by_date = {row[0]: float(row[1]) for row in rows if row[1] is not None}
    return [(metric_date, by_date.get(metric_date)) for metric_date in REPORT_DATES]


def write_data_block(data_ws, start_row: int, label: str, pilot_name: str, sister_name: str, pilot_series, sister_series, baseline: float, floor: float) -> int:
    dates = [datetime.fromisoformat(d).strftime("%-m/%d") for d, _ in pilot_series]
    data_ws.cell(start_row, 1, label)
    data_ws.cell(start_row + 1, 1, "Baseline")
    data_ws.cell(start_row + 2, 1, "Floor")
    data_ws.cell(start_row + 3, 1, pilot_name)
    data_ws.cell(start_row + 4, 1, sister_name)

    for idx, day in enumerate(dates, start=2):
        data_ws.cell(start_row, idx, day)
        data_ws.cell(start_row + 1, idx, baseline)
        data_ws.cell(start_row + 2, idx, floor)
        data_ws.cell(start_row + 3, idx, pilot_series[idx - 2][1])
        data_ws.cell(start_row + 4, idx, sister_series[idx - 2][1])
    return start_row + 7


def style_chart(chart: LineChart, data_ws, top_row: int, bottom_row: int) -> None:
    chart.height = 2.25
    chart.width = 20.5
    chart.style = 2
    chart.legend = None
    chart.title = None
    chart.layout = Layout(
        manualLayout=ManualLayout(
            x=0.02,
            y=0.04,
            w=0.96,
            h=0.80,
            layoutTarget="inner",
        )
    )
    chart.graphical_properties = GraphicalProperties()
    chart.graphical_properties.noFill = True
    chart.graphical_properties.line.noFill = True
    chart.y_axis.delete = True
    chart.x_axis.majorTickMark = "none"
    chart.y_axis.majorTickMark = "none"
    chart.y_axis.spPr = None
    chart.y_axis.majorGridlines = None
    chart.x_axis.delete = False
    chart.x_axis.axPos = "b"
    chart.x_axis.tickLblPos = "low"
    chart.x_axis.spPr = GraphicalProperties()
    chart.x_axis.spPr.line.noFill = True
    chart.x_axis.majorGridlines = ChartLines()
    chart.x_axis.majorGridlines.spPr = GraphicalProperties()
    chart.x_axis.majorGridlines.spPr.line.noFill = True
    chart.plot_area.spPr = None

    data = Reference(data_ws, min_col=2, max_col=8, min_row=top_row, max_row=bottom_row)
    cats = Reference(data_ws, min_col=2, max_col=8, min_row=top_row - 1, max_row=top_row - 1)
    chart.add_data(data, titles_from_data=False, from_rows=True)
    chart.set_categories(cats)

    colors = [BASELINE_COLOR, FLOOR_COLOR, PILOT_COLOR, SISTER_COLOR]
    widths = [15000, 15000, 23000, 23000]
    dashes = ["sysDash", "sysDash", None, None]

    for idx, series in enumerate(chart.series):
        line = series.graphicalProperties.line
        line.solidFill = colors[idx]
        line.width = widths[idx]
        if dashes[idx]:
            line.prstDash = dashes[idx]
        marker = series.marker
        marker.symbol = "circle"
        marker.size = 3 if idx < 2 else 6
        marker.graphicalProperties.solidFill = colors[idx]
        marker.graphicalProperties.line.solidFill = colors[idx]


def add_pair_card(display_ws, data_ws, start_row: int, pair: dict, conn: sqlite3.Connection, data_row: int) -> tuple[int, int]:
    pilot = pair["pilot"]
    sister = pair["sister"]
    pair_title = f"{pilot['display_name']} vs {sister['display_name']}"

    display_ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=12)
    title_cell = display_ws.cell(start_row, 1)
    title_cell.value = CellRichText(
        TextBlock(InlineFont(rFont="Calibri", sz=20, b=True, color=PILOT_COLOR), pilot["display_name"]),
        TextBlock(InlineFont(rFont="Calibri", sz=20, b=True, color=TEXT_BLACK), " vs "),
        TextBlock(InlineFont(rFont="Calibri", sz=20, b=True, color=SISTER_COLOR), sister["display_name"]),
    )
    title_cell.font = Font(size=14, bold=True, color=TEXT_BLACK)
    title_cell.fill = PatternFill("solid", fgColor=TITLE_FILL)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    display_ws.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=12)
    psi_label = display_ws.cell(start_row + 1, 1, "PSI")
    gtm_label = display_ws.cell(start_row + 7, 1, "GTMetrix")
    for cell in (psi_label, gtm_label):
        cell.font = Font(size=10, bold=True, color=TEXT_BLACK)
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for row in range(start_row + 1, start_row + 14):
        for col in range(1, 13):
            display_ws.cell(row, col).fill = PatternFill("solid", fgColor=CARD_FILL)

    pilot_psi = fetch_series(conn, pilot["property_id"], "psi")
    sister_psi = fetch_series(conn, sister["property_id"], "psi")
    pilot_gtm = fetch_series(conn, pilot["property_id"], "gtm")
    sister_gtm = fetch_series(conn, sister["property_id"], "gtm")

    data_row = write_data_block(
        data_ws,
        data_row,
        pair_title + " PSI",
        pilot["display_name"],
        sister["display_name"],
        pilot_psi,
        sister_psi,
        PSI_BASELINE,
        PSI_FLOOR,
    )
    psi_chart = LineChart()
    style_chart(psi_chart, data_ws, data_row - 6, data_row - 3)
    display_ws.add_chart(psi_chart, f"A{start_row + 2}")

    data_row = write_data_block(
        data_ws,
        data_row,
        pair_title + " GTM",
        pilot["display_name"],
        sister["display_name"],
        pilot_gtm,
        sister_gtm,
        GTM_BASELINE,
        GTM_FLOOR,
    )
    gtm_chart = LineChart()
    style_chart(gtm_chart, data_ws, data_row - 6, data_row - 3)
    display_ws.add_chart(gtm_chart, f"A{start_row + 8}")

    return start_row + 15, data_row


def build_workbook() -> None:
    wb = Workbook()
    display_ws = wb.active
    display_ws.title = "CWV Prototype"
    data_ws = wb.create_sheet("Chart Data")
    data_ws.sheet_state = "hidden"

    display_ws.sheet_view.showGridLines = False
    display_ws.merge_cells("A1:L1")
    display_ws["A1"] = "CWV Section Prototype"
    display_ws["A1"].font = Font(size=18, bold=True, color="0F172A")
    display_ws["A1"].alignment = Alignment(horizontal="left")

    display_ws.merge_cells("A2:L2")
    display_ws["A2"] = "Clean paired charts only: pilot, sister, baseline, floor."
    display_ws["A2"].font = Font(size=10, color="64748B")

    for col in range(1, 13):
        display_ws.column_dimensions[get_column_letter(col)].width = 9

    conn = sqlite3.connect(DB_PATH)
    row = 4
    data_row = 1
    for pair in load_pairs():
        row, data_row = add_pair_card(display_ws, data_ws, row, pair, conn, data_row)
    conn.close()

    wb.save(OUTPUT_PATH)


if __name__ == "__main__":
    build_workbook()
    print(OUTPUT_PATH)
