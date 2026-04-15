#!/usr/bin/env python3
"""
Generate pilot/control GTMetrix paste-friendly report outputs.
"""

from __future__ import annotations

import argparse
import csv
import json
import sqlite3
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parents[1]
DEFAULT_CONFIG_PATH = BASE_DIR / "config" / "pilot_control_cwv_config.json"
OUTPUT_DIR = BASE_DIR / "reports"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


@dataclass
class CohortEntry:
    key: str
    display_name: str
    role: str
    property_id: Optional[str]
    sister_key: Optional[str]
    active: bool


def load_config(path: Path) -> Dict[str, object]:
    with path.open() as fh:
        return json.load(fh)


def active_entries(config: Dict[str, object]) -> List[CohortEntry]:
    rows: List[CohortEntry] = []
    for row in config.get("cohorts", []):
        if not row.get("active", True):
            continue
        rows.append(
            CohortEntry(
                key=row["key"],
                display_name=row["display_name"],
                role=row["role"],
                property_id=row.get("property_id"),
                sister_key=row.get("sister_key"),
                active=bool(row.get("active", True)),
            )
        )
    return rows


def pilot_control_pairs(entries: List[CohortEntry]) -> List[tuple[CohortEntry, Optional[CohortEntry]]]:
    pilots = [entry for entry in entries if entry.role == "pilot"]
    index = {entry.key: entry for entry in entries}
    return [(pilot, index.get(pilot.sister_key) if pilot.sister_key else None) for pilot in pilots]


def fmt_num(value: Optional[float]) -> str:
    return "" if value is None else f"{value:.1f}"


def history_rows(conn: sqlite3.Connection, property_id: str, as_of_date: str) -> List[sqlite3.Row]:
    return conn.execute(
        """
        SELECT metric_date, pagespeed_score
        FROM gtmetrix_metrics
        WHERE property_id = ?
          AND metric_date <= ?
        ORDER BY metric_date DESC
        """,
        (property_id, as_of_date),
    ).fetchall()


def prior_scores(history: List[sqlite3.Row], as_of_date: str) -> List[float]:
    return [
        row["pagespeed_score"]
        for row in history
        if row["metric_date"] < as_of_date and row["pagespeed_score"] is not None
    ]


def average_window(scores: List[float], length: int) -> Optional[float]:
    if len(scores) < length:
        return None
    window = scores[:length]
    return sum(window) / len(window)


def static_baseline(scores: List[float], days: int = 90, minimum_days: int = 30) -> Optional[float]:
    if len(scores) < minimum_days:
        return None
    window = scores[:days]
    return sum(window) / len(window)


def latest_complete_date(conn: sqlite3.Connection, entries: List[CohortEntry]) -> Optional[str]:
    expected = len([entry for entry in entries if entry.property_id])
    row = conn.execute(
        """
        SELECT metric_date
        FROM gtmetrix_metrics
        GROUP BY metric_date
        HAVING COUNT(DISTINCT property_id) >= ?
        ORDER BY metric_date DESC
        LIMIT 1
        """,
        (expected,),
    ).fetchone()
    return row[0] if row else None


def build_dataset(
    conn: sqlite3.Connection,
    entries: List[CohortEntry],
    metric_date: str,
) -> Dict[str, Dict[str, Optional[float]]]:
    dataset: Dict[str, Dict[str, Optional[float]]] = {}
    for entry in entries:
        if not entry.property_id:
            dataset[entry.key] = {"score": None, "baseline": None, "t7": None, "t30": None}
            continue

        history = history_rows(conn, entry.property_id, metric_date)
        current = next((row for row in history if row["metric_date"] == metric_date), None)
        prior = prior_scores(history, metric_date)
        dataset[entry.key] = {
            "score": current["pagespeed_score"] if current else None,
            "baseline": static_baseline(prior),
            "t7": average_window(prior, 7),
            "t30": average_window(prior, 30),
        }
    return dataset


def exact_paste_columns() -> tuple[list[tuple[str, Optional[str]]], list[tuple[str, Optional[str]]]]:
    pilot_columns = [
        ("Today", "score"),
        ("Baseline (T90 static)", "baseline"),
        ("", None),
        ("T7", "t7"),
        ("", None),
        ("T30", "t30"),
        ("", None),
        ("", None),
    ]
    control_columns = [
        ("Today", "score"),
        ("Baseline (T90 static)", "baseline"),
        ("", None),
        ("T7", "t7"),
        ("", None),
        ("T30", "t30"),
        ("", None),
    ]
    return pilot_columns, control_columns


def sheet_title_for_pilot(pilot: CohortEntry) -> str:
    custom = {"pilot_district_universal": "The District"}
    return custom.get(pilot.key, pilot.display_name)[:31]


def build_workbook(entries: List[CohortEntry], dataset: Dict[str, Dict[str, Optional[float]]], metric_date: str) -> Workbook:
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="D9E2F3")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_gray = Side(style="thin", color="D9DCE3")
    dark_gray = Side(style="thin", color="333333")
    light_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    header_border = Border(left=dark_gray, right=dark_gray, top=dark_gray, bottom=dark_gray)
    pilot_columns, control_columns = exact_paste_columns()

    first = True
    for pilot, sister in pilot_control_pairs(entries):
        ws = wb.active if first else wb.create_sheet()
        first = False
        ws.title = sheet_title_for_pilot(pilot)
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "B4"
        ws.column_dimensions["A"].width = 16
        ws.row_dimensions[1].height = 36
        ws.row_dimensions[2].height = 40
        ws.row_dimensions[3].height = 42

        total_cols = 1 + len(pilot_columns) + (len(control_columns) if sister else 0)
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=total_cols)
        top = ws.cell(row=1, column=2, value="GTMetrix SCORE")
        top.alignment = center
        top.fill = header_fill
        top.border = header_border
        top.font = Font(size=12)

        ws.cell(row=1, column=1, value=pilot.display_name).font = Font(size=16, bold=True)

        ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=1 + len(pilot_columns))
        pilot_header = ws.cell(row=2, column=2, value=pilot.display_name)
        pilot_header.alignment = center
        pilot_header.fill = header_fill
        pilot_header.border = header_border
        pilot_header.font = Font(size=12)

        if sister:
            start = 2 + len(pilot_columns)
            end = start + len(control_columns) - 1
            ws.merge_cells(start_row=2, start_column=start, end_row=2, end_column=end)
            sister_header = ws.cell(row=2, column=start, value=sister.display_name)
            sister_header.alignment = center
            sister_header.fill = header_fill
            sister_header.border = header_border
            sister_header.font = Font(size=12)

        ws.cell(row=3, column=1, value="")
        col = 2
        for label, _ in pilot_columns:
            cell = ws.cell(row=3, column=col, value=label)
            cell.alignment = center
            cell.fill = white_fill
            cell.border = light_border
            cell.font = Font(size=11)
            ws.column_dimensions[get_column_letter(col)].width = 13
            col += 1
        if sister:
            for label, _ in control_columns:
                cell = ws.cell(row=3, column=col, value=label)
                cell.alignment = center
                cell.fill = white_fill
                cell.border = light_border
                cell.font = Font(size=11)
                ws.column_dimensions[get_column_letter(col)].width = 13
                col += 1

        ws.cell(row=4, column=1, value="PRACTICE ROW").alignment = center
        ws.cell(row=4, column=1).border = light_border
        for col_num in range(2, total_cols + 1):
            ws.cell(row=4, column=col_num).border = light_border

        row_num = 5
        row = [metric_date]
        pilot_values = dataset[pilot.key]
        for _, key in pilot_columns:
            row.append("" if key is None else fmt_num(pilot_values.get(key)))
        if sister:
            sister_values = dataset[sister.key]
            for _, key in control_columns:
                row.append("" if key is None else fmt_num(sister_values.get(key)))
        for col_num, value in enumerate(row, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = center
            cell.border = light_border

    return wb


def write_copy_paste_csv(entries: List[CohortEntry], dataset: Dict[str, Dict[str, Optional[float]]], metric_date: str, out_path: Path) -> None:
    headers = [
        "Pilot Property",
        "Sister Property",
        "Date",
        "Pilot Today",
        "Pilot Baseline (T90 static)",
        "Pilot T7",
        "Pilot T30",
        "Sister Today",
        "Sister Baseline (T90 static)",
        "Sister T7",
        "Sister T30",
    ]
    with out_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(headers)
        for pilot, sister in pilot_control_pairs(entries):
            pilot_values = dataset[pilot.key]
            sister_values = dataset[sister.key] if sister else {}
            writer.writerow(
                [
                    pilot.display_name,
                    sister.display_name if sister else "",
                    metric_date,
                    fmt_num(pilot_values.get("score")),
                    fmt_num(pilot_values.get("baseline")),
                    fmt_num(pilot_values.get("t7")),
                    fmt_num(pilot_values.get("t30")),
                    fmt_num(sister_values.get("score")),
                    fmt_num(sister_values.get("baseline")),
                    fmt_num(sister_values.get("t7")),
                    fmt_num(sister_values.get("t30")),
                ]
            )


def write_per_property_tsv(entries: List[CohortEntry], dataset: Dict[str, Dict[str, Optional[float]]], metric_date: str) -> None:
    pilot_columns, control_columns = exact_paste_columns()
    for pilot, sister in pilot_control_pairs(entries):
        path = OUTPUT_DIR / f"{sheet_title_for_pilot(pilot).replace(' ', '_')}_GTMetrix_Paste_Lines.tsv"
        with path.open("w", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh, delimiter="\t")
            headers = ["Date"] + [label for label, _ in pilot_columns]
            if sister:
                headers.extend(label for label, _ in control_columns)
            writer.writerow(headers)

            row = [metric_date]
            pilot_values = dataset[pilot.key]
            for _, key in pilot_columns:
                row.append("" if key is None else fmt_num(pilot_values.get(key)))
            if sister:
                sister_values = dataset[sister.key]
                for _, key in control_columns:
                    row.append("" if key is None else fmt_num(sister_values.get(key)))
            writer.writerow(row)


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate pilot/control GTMetrix report outputs")
    parser.add_argument("--config", default=str(DEFAULT_CONFIG_PATH), help="Path to config JSON")
    parser.add_argument("--date", default=None, help="Optional metric date override")
    args = parser.parse_args()

    config = load_config(Path(args.config))
    entries = active_entries(config)
    conn = sqlite3.connect(Path(config["db_path"]))
    conn.row_factory = sqlite3.Row

    metric_date = args.date or latest_complete_date(conn, entries)
    if not metric_date:
        raise SystemExit("No complete GTMetrix cohort date found")

    dataset = build_dataset(conn, entries, metric_date)
    conn.close()

    workbook_path = OUTPUT_DIR / f"Pilot_Control_GTMetrix_Paste_Exact_{metric_date}.xlsx"
    copy_paste_path = OUTPUT_DIR / "Pilot_Control_GTMetrix_Copy_Paste.csv"

    wb = build_workbook(entries, dataset, metric_date)
    wb.save(workbook_path)
    write_copy_paste_csv(entries, dataset, metric_date, copy_paste_path)
    write_per_property_tsv(entries, dataset, metric_date)

    print(f"Saved workbook: {workbook_path}")
    print(f"Saved CSV:      {copy_paste_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
