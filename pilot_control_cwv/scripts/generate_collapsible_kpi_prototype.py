from __future__ import annotations

import csv
import json
import sqlite3
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path("/Users/mark/Property_Analytics")
DB_PATH = ROOT / "data" / "portfolio_analytics.db"
CONFIG_PATH = ROOT / "pilot_control_cwv" / "config" / "pilot_control_cwv_config.json"
BI_SERIES_PATH = ROOT / "pilot_control_cwv" / "reports" / "pilot_bi_report_series.csv"
OUTPUT_PATH = (
    ROOT
    / "pilot_control_cwv"
    / "reports"
    / f"Pilot_KPI_Collapsible_Prototype_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
)

PILOT_COLOR = "4473D0"
SISTER_COLOR = "7CCAC2"
BASELINE_COLOR = "A3A3A3"
FLOOR_COLOR = "F2A7A7"
SISTER_BASELINE_COLOR = "D6D3D1"
CARD_FILL = "F8FAFC"
TITLE_FILL = "EAF2FF"
SECTION_FILL = "EEF4FF"
TEXT_BLACK = "0F172A"
MUTED_TEXT = "64748B"

PSI_BASELINE = 90
PSI_FLOOR = 60
GTM_BASELINE = 94
GTM_FLOOR = 70
CWV_DATES = ["2026-03-26", "2026-03-27", "2026-03-28", "2026-03-29", "2026-03-30", "2026-03-31"]
BI_SECTION_KEY = "lead_to_available_unit_rate"
BI_SECTION_TITLE = "Lead (Guest Card) to Available Unit Rate"


def load_pairs() -> list[dict]:
    cfg = json.loads(CONFIG_PATH.read_text())
    cohorts = {c["key"]: c for c in cfg["cohorts"] if c.get("active")}
    pairs = []
    for cohort in cohorts.values():
        if cohort["role"] != "pilot":
            continue
        sister = cohorts[cohort["sister_key"]]
        pairs.append({"pilot": cohort, "sister": sister})
    return pairs


def fetch_cwv_series(conn: sqlite3.Connection, property_id: str, kind: str) -> list[tuple[str, float | None]]:
    if kind == "psi":
        query = """
            SELECT metric_date, performance_score
            FROM pilot_control_psi_metrics
            WHERE property_id = ? AND strategy = 'mobile'
            ORDER BY metric_date DESC
            LIMIT 7
        """
    else:
        query = """
            SELECT metric_date, pagespeed_score
            FROM gtmetrix_metrics
            WHERE property_id = ?
            ORDER BY metric_date DESC
            LIMIT 7
        """
    rows = conn.execute(query, (property_id,)).fetchall()
    by_date = {row[0]: float(row[1]) for row in rows if row[1] is not None}
    return [(metric_date, by_date.get(metric_date)) for metric_date in CWV_DATES]


def load_bi_grouped() -> dict[tuple[str, str], list[dict]]:
    with BI_SERIES_PATH.open() as f:
        rows = list(csv.DictReader(f))
    grouped: dict[tuple[str, str], list[dict]] = defaultdict(list)
    for row in rows:
        if row["section_key"] != BI_SECTION_KEY:
            continue
        grouped[(row["pilot_property_name"], row["sister_property_name"])].append(row)
    for key in grouped:
        grouped[key].sort(key=lambda r: r["snapshot_date"])
    return grouped


def fmt_num(value: float | None, is_ratio: bool = False) -> str:
    if value is None or value == "":
        return "n/a"
    return f"{value:.1f}" if not is_ratio else f"{value * 100:.1f}%"


def avg(values: list[float | None]) -> float | None:
    clean = [v for v in values if v is not None]
    if not clean:
        return None
    return sum(clean) / len(clean)


def style_chart(
    chart: LineChart,
    data_ws,
    top_row: int,
    bottom_row: int,
    num_dates: int,
    colors: list[str],
    widths: list[int],
    dashes: list[str | None],
    height: float = 2.1,
) -> None:
    chart.height = height
    chart.width = 15.8
    chart.style = 2
    chart.legend = None
    chart.title = None
    chart.layout = Layout(
        manualLayout=ManualLayout(
            x=0.01,
            y=0.04,
            w=0.97,
            h=0.80,
            layoutTarget="inner",
        )
    )
    chart.graphical_properties = GraphicalProperties()
    chart.graphical_properties.noFill = True
    chart.graphical_properties.line.noFill = True
    chart.y_axis.delete = True
    chart.x_axis.majorTickMark = "none"
    chart.y_axis.majorTickMark = "none"
    chart.y_axis.spPr = None
    chart.y_axis.majorGridlines = None
    chart.x_axis.delete = False
    chart.x_axis.axPos = "b"
    chart.x_axis.tickLblPos = "low"
    chart.x_axis.spPr = GraphicalProperties()
    chart.x_axis.spPr.line.noFill = True
    chart.x_axis.majorGridlines = ChartLines()
    chart.x_axis.majorGridlines.spPr = GraphicalProperties()
    chart.x_axis.majorGridlines.spPr.line.noFill = True
    chart.plot_area.spPr = None

    max_col = 1 + num_dates
    data = Reference(data_ws, min_col=2, max_col=max_col, min_row=top_row, max_row=bottom_row)
    cats = Reference(data_ws, min_col=2, max_col=max_col, min_row=top_row - 1, max_row=top_row - 1)
    chart.add_data(data, titles_from_data=False, from_rows=True)
    chart.set_categories(cats)

    for idx, series in enumerate(chart.series):
        line = series.graphicalProperties.line
        line.solidFill = colors[idx]
        line.width = widths[idx]
        if dashes[idx]:
            line.prstDash = dashes[idx]
        marker = series.marker
        marker.symbol = "circle"
        marker.size = 3 if idx < 2 else 6
        marker.graphicalProperties.solidFill = colors[idx]
        marker.graphicalProperties.line.solidFill = colors[idx]


def add_section_title(ws, row: int, text: str) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    cell = ws.cell(row, 1, text)
    cell.font = Font(size=14, bold=True, color=TEXT_BLACK)
    cell.fill = PatternFill("solid", fgColor=SECTION_FILL)
    cell.alignment = Alignment(horizontal="left")


def add_summary_title(ws, row: int, title: str, note: str) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    cell = ws.cell(row, 1, f"{title}  [-] Summary above, expand rows below")
    cell.font = Font(size=12, bold=True, color=TEXT_BLACK)
    cell.fill = PatternFill("solid", fgColor=TITLE_FILL)
    cell.alignment = Alignment(horizontal="left")
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=12)
    sub = ws.cell(row + 1, 1, note)
    sub.font = Font(size=9, color=MUTED_TEXT)
    sub.alignment = Alignment(horizontal="left")


def add_drawer_title(ws, row: int, title: str, note: str) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    cell = ws.cell(row, 1, f"{title}   [+] Expand details")
    cell.font = Font(size=13, bold=True, color=TEXT_BLACK)
    cell.fill = PatternFill("solid", fgColor=TITLE_FILL)
    cell.alignment = Alignment(horizontal="left")
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=12)
    sub = ws.cell(row + 1, 1, note)
    sub.font = Font(size=9, color=MUTED_TEXT)
    sub.alignment = Alignment(horizontal="left")


def add_pair_title(ws, row: int, pilot_name: str, sister_name: str) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    cell = ws.cell(row, 1)
    cell.value = CellRichText(
        TextBlock(InlineFont(rFont="Calibri", sz=16, b=True, color=PILOT_COLOR), pilot_name),
        TextBlock(InlineFont(rFont="Calibri", sz=16, b=True, color=TEXT_BLACK), " vs "),
        TextBlock(InlineFont(rFont="Calibri", sz=16, b=True, color=SISTER_COLOR), sister_name),
    )
    cell.fill = PatternFill("solid", fgColor=TITLE_FILL)
    cell.alignment = Alignment(horizontal="left")


def write_value_labels(ws, top_row: int, labels: list[tuple[str, str, str]]) -> None:
    for idx, (label, value, color) in enumerate(labels):
        row = top_row + idx
        ws.cell(row, 10, label).font = Font(size=9, bold=True, color=color)
        ws.cell(row, 10).alignment = Alignment(horizontal="left")
        ws.cell(row, 11, value).font = Font(size=9, color=TEXT_BLACK)
        ws.cell(row, 11).alignment = Alignment(horizontal="right")


def add_summary_card(
    ws,
    data_ws,
    start_row: int,
    title: str,
    dates: list[str],
    pilot_values: list[float | None],
    sister_values: list[float | None],
    baseline_values: list[float | None],
    floor_values: list[float | None] | None,
    data_row: int,
    is_ratio: bool = False,
    sister_baseline_values: list[float | None] | None = None,
) -> tuple[int, int]:
    for row in range(start_row, start_row + 7):
        for col in range(1, 13):
            ws.cell(row, col).fill = PatternFill("solid", fgColor=CARD_FILL)
    note = "Simple average of the 5 pilot properties vs the 5 sister properties"
    add_summary_title(ws, start_row, title, note)

    data_ws.cell(data_row, 1, title)
    for idx, day in enumerate(dates, start=2):
        data_ws.cell(data_row, idx, day)
        data_ws.cell(data_row + 1, idx, baseline_values[idx - 2])
        if floor_values is not None:
            data_ws.cell(data_row + 2, idx, floor_values[idx - 2])
            data_ws.cell(data_row + 3, idx, pilot_values[idx - 2])
            data_ws.cell(data_row + 4, idx, sister_values[idx - 2])
            top_row = data_row + 1
            bottom_row = data_row + 4
            colors = [BASELINE_COLOR, FLOOR_COLOR, PILOT_COLOR, SISTER_COLOR]
            widths = [12000, 12000, 23000, 23000]
            dashes = ["sysDash", "sysDash", None, None]
            label_rows = [
                ("Pilot Avg", fmt_num(pilot_values[-1], is_ratio), PILOT_COLOR),
                ("Sister Avg", fmt_num(sister_values[-1], is_ratio), SISTER_COLOR),
                ("Baseline", fmt_num(baseline_values[-1], is_ratio), BASELINE_COLOR),
                ("Floor", fmt_num(floor_values[-1], is_ratio), FLOOR_COLOR),
            ]
            next_data_row = data_row + 6
        else:
            if sister_baseline_values is not None:
                data_ws.cell(data_row + 2, idx, sister_baseline_values[idx - 2])
                data_ws.cell(data_row + 3, idx, pilot_values[idx - 2])
                data_ws.cell(data_row + 4, idx, sister_values[idx - 2])
                top_row = data_row + 1
                bottom_row = data_row + 4
                colors = [BASELINE_COLOR, SISTER_BASELINE_COLOR, PILOT_COLOR, SISTER_COLOR]
                widths = [12000, 12000, 23000, 23000]
                dashes = ["sysDash", "sysDash", None, None]
                label_rows = [
                    ("Pilot Avg", fmt_num(pilot_values[-1], is_ratio), PILOT_COLOR),
                    ("Sister Avg", fmt_num(sister_values[-1], is_ratio), SISTER_COLOR),
                    ("Pilot BL", fmt_num(baseline_values[-1], is_ratio), BASELINE_COLOR),
                    ("Sister BL", fmt_num(sister_baseline_values[-1], is_ratio), SISTER_BASELINE_COLOR),
                ]
                next_data_row = data_row + 6
            else:
                data_ws.cell(data_row + 2, idx, pilot_values[idx - 2])
                data_ws.cell(data_row + 3, idx, sister_values[idx - 2])
                top_row = data_row + 1
                bottom_row = data_row + 3
                colors = [BASELINE_COLOR, PILOT_COLOR, SISTER_COLOR]
                widths = [12000, 23000, 23000]
                dashes = ["sysDash", None, None]
                label_rows = [
                    ("Pilot Avg", fmt_num(pilot_values[-1], is_ratio), PILOT_COLOR),
                    ("Sister Avg", fmt_num(sister_values[-1], is_ratio), SISTER_COLOR),
                    ("Baseline", fmt_num(baseline_values[-1], is_ratio), BASELINE_COLOR),
                ]
                next_data_row = data_row + 5

    chart = LineChart()
    style_chart(chart, data_ws, top_row, bottom_row, len(dates), colors, widths, dashes, height=2.0)
    ws.add_chart(chart, f"A{start_row + 2}")
    write_value_labels(ws, start_row + 3, label_rows)
    return start_row + 8, next_data_row


def add_cwv_group_summary(
    ws,
    data_ws,
    start_row: int,
    psi_dates: list[str],
    psi_pilot_avg: list[float | None],
    psi_sister_avg: list[float | None],
    gtm_dates: list[str],
    gtm_pilot_avg: list[float | None],
    gtm_sister_avg: list[float | None],
    data_row: int,
) -> tuple[int, int]:
    for row in range(start_row, start_row + 14):
        for col in range(1, 13):
            ws.cell(row, col).fill = PatternFill("solid", fgColor=CARD_FILL)

    add_drawer_title(
        ws,
        start_row,
        "Core Web Vitals",
        "Two-source rollup: PSI Avg and GTMetrix Avg. Use Excel's outline controls on the left to expand the full property chart set.",
    )

    ws.merge_cells(start_row=start_row + 2, start_column=1, end_row=start_row + 2, end_column=12)
    ws.cell(start_row + 2, 1, "PSI Avg").font = Font(size=10, bold=True, color=TEXT_BLACK)
    ws.merge_cells(start_row=start_row + 8, start_column=1, end_row=start_row + 8, end_column=12)
    ws.cell(start_row + 8, 1, "GTMetrix Avg").font = Font(size=10, bold=True, color=TEXT_BLACK)

    data_ws.cell(data_row, 1, "CWV PSI Summary")
    for idx, day in enumerate(psi_dates, start=2):
        data_ws.cell(data_row, idx, day)
        data_ws.cell(data_row + 1, idx, PSI_BASELINE)
        data_ws.cell(data_row + 2, idx, PSI_FLOOR)
        data_ws.cell(data_row + 3, idx, psi_pilot_avg[idx - 2])
        data_ws.cell(data_row + 4, idx, psi_sister_avg[idx - 2])

    psi_chart = LineChart()
    style_chart(
        psi_chart,
        data_ws,
        data_row + 1,
        data_row + 4,
        len(psi_dates),
        [BASELINE_COLOR, FLOOR_COLOR, PILOT_COLOR, SISTER_COLOR],
        [12000, 12000, 23000, 23000],
        ["sysDash", "sysDash", None, None],
        height=1.55,
    )
    ws.add_chart(psi_chart, f"A{start_row + 3}")
    write_value_labels(
        ws,
        start_row + 3,
        [
            ("Pilot Avg", fmt_num(psi_pilot_avg[-1]), PILOT_COLOR),
            ("Sister Avg", fmt_num(psi_sister_avg[-1]), SISTER_COLOR),
            ("Baseline", fmt_num(PSI_BASELINE), BASELINE_COLOR),
            ("Floor", fmt_num(PSI_FLOOR), FLOOR_COLOR),
        ],
    )
    data_row += 6

    data_ws.cell(data_row, 1, "CWV GTM Summary")
    for idx, day in enumerate(gtm_dates, start=2):
        data_ws.cell(data_row, idx, day)
        data_ws.cell(data_row + 1, idx, GTM_BASELINE)
        data_ws.cell(data_row + 2, idx, GTM_FLOOR)
        data_ws.cell(data_row + 3, idx, gtm_pilot_avg[idx - 2])
        data_ws.cell(data_row + 4, idx, gtm_sister_avg[idx - 2])

    gtm_chart = LineChart()
    style_chart(
        gtm_chart,
        data_ws,
        data_row + 1,
        data_row + 4,
        len(gtm_dates),
        [BASELINE_COLOR, FLOOR_COLOR, PILOT_COLOR, SISTER_COLOR],
        [12000, 12000, 23000, 23000],
        ["sysDash", "sysDash", None, None],
        height=1.55,
    )
    ws.add_chart(gtm_chart, f"A{start_row + 9}")
    write_value_labels(
        ws,
        start_row + 9,
        [
            ("Pilot Avg", fmt_num(gtm_pilot_avg[-1]), PILOT_COLOR),
            ("Sister Avg", fmt_num(gtm_sister_avg[-1]), SISTER_COLOR),
            ("Baseline", fmt_num(GTM_BASELINE), BASELINE_COLOR),
            ("Floor", fmt_num(GTM_FLOOR), FLOOR_COLOR),
        ],
    )
    return start_row + 15, data_row + 6


def add_cwv_detail_card(
    ws,
    data_ws,
    start_row: int,
    pair: dict,
    conn: sqlite3.Connection,
    kind: str,
    data_row: int,
) -> tuple[int, int]:
    pilot = pair["pilot"]
    sister = pair["sister"]
    add_pair_title(ws, start_row, pilot["display_name"], sister["display_name"])
    for row in range(start_row + 1, start_row + 7):
        for col in range(1, 13):
            ws.cell(row, col).fill = PatternFill("solid", fgColor=CARD_FILL)

    title = "PSI" if kind == "psi" else "GTMetrix"
    ws.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=12)
    ws.cell(start_row + 1, 1, title).font = Font(size=10, bold=True, color=TEXT_BLACK)

    pilot_series = fetch_cwv_series(conn, pilot["property_id"], kind)
    sister_series = fetch_cwv_series(conn, sister["property_id"], kind)
    dates = [datetime.fromisoformat(d).strftime("%-m/%d") for d, _ in pilot_series]
    baseline = PSI_BASELINE if kind == "psi" else GTM_BASELINE
    floor = PSI_FLOOR if kind == "psi" else GTM_FLOOR

    data_ws.cell(data_row, 1, f"{title} detail")
    for idx, day in enumerate(dates, start=2):
        data_ws.cell(data_row, idx, day)
        data_ws.cell(data_row + 1, idx, baseline)
        data_ws.cell(data_row + 2, idx, floor)
        data_ws.cell(data_row + 3, idx, pilot_series[idx - 2][1])
        data_ws.cell(data_row + 4, idx, sister_series[idx - 2][1])

    chart = LineChart()
    style_chart(
        chart,
        data_ws,
        data_row + 1,
        data_row + 4,
        len(dates),
        [BASELINE_COLOR, FLOOR_COLOR, PILOT_COLOR, SISTER_COLOR],
        [12000, 12000, 23000, 23000],
        ["sysDash", "sysDash", None, None],
        height=1.7,
    )
    ws.add_chart(chart, f"A{start_row + 2}")
    write_value_labels(
        ws,
        start_row + 2,
        [
            ("Pilot", fmt_num(pilot_series[-1][1]), PILOT_COLOR),
            ("Sister", fmt_num(sister_series[-1][1]), SISTER_COLOR),
            ("BL", fmt_num(baseline), BASELINE_COLOR),
            ("Floor", fmt_num(floor), FLOOR_COLOR),
        ],
    )
    return start_row + 7, data_row + 6


def add_bi_detail_card(ws, data_ws, start_row: int, rows: list[dict], data_row: int) -> tuple[int, int]:
    pilot_name = rows[0]["pilot_property_name"]
    sister_name = rows[0]["sister_property_name"]
    add_pair_title(ws, start_row, pilot_name, sister_name)
    for row in range(start_row + 1, start_row + 7):
        for col in range(1, 13):
            ws.cell(row, col).fill = PatternFill("solid", fgColor=CARD_FILL)

    ws.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=12)
    ws.cell(start_row + 1, 1, BI_SECTION_TITLE).font = Font(size=10, bold=True, color=TEXT_BLACK)

    dates = [datetime.fromisoformat(r["snapshot_date"]).strftime("%-m/%d") for r in rows]
    data_ws.cell(data_row, 1, "BI detail")
    for idx, day in enumerate(dates, start=2):
        row = rows[idx - 2]
        data_ws.cell(data_row, idx, day)
        data_ws.cell(data_row + 1, idx, float(row["pilot_baseline_value"]) if row["pilot_baseline_value"] else None)
        data_ws.cell(data_row + 2, idx, float(row["sister_baseline_value"]) if row["sister_baseline_value"] else None)
        data_ws.cell(data_row + 3, idx, float(row["pilot_daily_value"]) if row["pilot_daily_value"] else None)
        data_ws.cell(data_row + 4, idx, float(row["sister_daily_value"]) if row["sister_daily_value"] else None)

    chart = LineChart()
    style_chart(
        chart,
        data_ws,
        data_row + 1,
        data_row + 4,
        len(dates),
        [BASELINE_COLOR, SISTER_BASELINE_COLOR, PILOT_COLOR, SISTER_COLOR],
        [12000, 12000, 23000, 23000],
        ["sysDash", "sysDash", None, None],
        height=1.7,
    )
    ws.add_chart(chart, f"A{start_row + 2}")
    latest = rows[-1]
    write_value_labels(
        ws,
        start_row + 2,
        [
            ("Pilot", fmt_num(float(latest["pilot_daily_value"]) if latest["pilot_daily_value"] else None, True), PILOT_COLOR),
            ("Sister", fmt_num(float(latest["sister_daily_value"]) if latest["sister_daily_value"] else None, True), SISTER_COLOR),
            ("Pilot BL", fmt_num(float(latest["pilot_baseline_value"]) if latest["pilot_baseline_value"] else None, True), BASELINE_COLOR),
            ("Sister BL", fmt_num(float(latest["sister_baseline_value"]) if latest["sister_baseline_value"] else None, True), SISTER_BASELINE_COLOR),
        ],
    )
    return start_row + 7, data_row + 6


def build_workbook() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Collapsible Prototype"
    data_ws = wb.create_sheet("Chart Data")
    data_ws.sheet_state = "hidden"

    ws.sheet_view.showGridLines = False
    ws.sheet_properties.outlinePr.summaryBelow = False

    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].width = 9

    ws.merge_cells("A1:L1")
    ws["A1"] = "Pilot KPI Collapsible Prototype"
    ws["A1"].font = Font(size=18, bold=True, color=TEXT_BLACK)
    ws.merge_cells("A2:L2")
    ws["A2"] = "Summary rows stay visible. Use Excel outline controls on the left to expand the five property-pair charts."
    ws["A2"].font = Font(size=10, color=MUTED_TEXT)

    pairs = load_pairs()
    bi_grouped = load_bi_grouped()
    conn = sqlite3.connect(DB_PATH)

    row = 4
    data_row = 1

    # Core Web Vitals group
    add_section_title(ws, row, "Core Web Vitals")
    row += 2
    psi_series_by_pair = {
        (p["pilot"]["display_name"], p["sister"]["display_name"]): (
            fetch_cwv_series(conn, p["pilot"]["property_id"], "psi"),
            fetch_cwv_series(conn, p["sister"]["property_id"], "psi"),
        )
        for p in pairs
    }
    gtm_series_by_pair = {
        (p["pilot"]["display_name"], p["sister"]["display_name"]): (
            fetch_cwv_series(conn, p["pilot"]["property_id"], "gtm"),
            fetch_cwv_series(conn, p["sister"]["property_id"], "gtm"),
        )
        for p in pairs
    }
    psi_pilot_avg = [avg([series[0][i][1] for series in psi_series_by_pair.values()]) for i in range(len(CWV_DATES))]
    psi_sister_avg = [avg([series[1][i][1] for series in psi_series_by_pair.values()]) for i in range(len(CWV_DATES))]
    gtm_pilot_avg = [avg([series[0][i][1] for series in gtm_series_by_pair.values()]) for i in range(len(CWV_DATES))]
    gtm_sister_avg = [avg([series[1][i][1] for series in gtm_series_by_pair.values()]) for i in range(len(CWV_DATES))]
    dates = [datetime.fromisoformat(d).strftime("%-m/%d") for d in CWV_DATES]
    row, data_row = add_cwv_group_summary(
        ws,
        data_ws,
        row,
        dates,
        psi_pilot_avg,
        psi_sister_avg,
        dates,
        gtm_pilot_avg,
        gtm_sister_avg,
        data_row,
    )
    detail_start = row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    ws.cell(row, 1, "PSI Detail").font = Font(size=11, bold=True, color=TEXT_BLACK)
    ws.cell(row, 1).fill = PatternFill("solid", fgColor=SECTION_FILL)
    row += 1
    for pair in pairs:
        row, data_row = add_cwv_detail_card(ws, data_ws, row, pair, conn, "psi", data_row)
        row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    ws.cell(row, 1, "GTMetrix Detail").font = Font(size=11, bold=True, color=TEXT_BLACK)
    ws.cell(row, 1).fill = PatternFill("solid", fgColor=SECTION_FILL)
    row += 1
    for pair in pairs:
        row, data_row = add_cwv_detail_card(ws, data_ws, row, pair, conn, "gtm", data_row)
        row += 1
    ws.row_dimensions.group(detail_start, row - 1, outline_level=1, hidden=True)
    row += 1
    conn.close()

    # BI section
    add_section_title(ws, row, BI_SECTION_TITLE)
    row += 2
    bi_dates = [datetime.fromisoformat(r["snapshot_date"]).strftime("%-m/%d") for r in next(iter(bi_grouped.values()))]
    pilot_avg = []
    sister_avg = []
    pilot_bl = []
    sister_bl = []
    for i in range(len(bi_dates)):
        pilot_avg.append(avg([float(rows[i]["pilot_daily_value"]) if rows[i]["pilot_daily_value"] else None for rows in bi_grouped.values()]))
        sister_avg.append(avg([float(rows[i]["sister_daily_value"]) if rows[i]["sister_daily_value"] else None for rows in bi_grouped.values()]))
        pilot_bl.append(avg([float(rows[i]["pilot_baseline_value"]) if rows[i]["pilot_baseline_value"] else None for rows in bi_grouped.values()]))
        sister_bl.append(avg([float(rows[i]["sister_baseline_value"]) if rows[i]["sister_baseline_value"] else None for rows in bi_grouped.values()]))
    row, data_row = add_summary_card(
        ws, data_ws, row, BI_SECTION_TITLE, bi_dates, pilot_avg, sister_avg,
        pilot_bl, None, data_row, True, sister_bl
    )
    detail_start = row
    for _, pair_rows in bi_grouped.items():
        row, data_row = add_bi_detail_card(ws, data_ws, row, pair_rows, data_row)
        row += 1
    ws.row_dimensions.group(detail_start, row - 1, outline_level=1, hidden=True)

    wb.save(OUTPUT_PATH)


if __name__ == "__main__":
    build_workbook()
    print(OUTPUT_PATH)
