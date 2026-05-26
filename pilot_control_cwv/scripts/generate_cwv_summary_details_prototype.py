from __future__ import annotations

import json
import sqlite3
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path("/Users/mark/Property_Analytics")
DB_PATH = ROOT / "data" / "portfolio_analytics.db"
CONFIG_PATH = ROOT / "pilot_control_cwv" / "config" / "pilot_control_cwv_config.json"
OUTPUT_PATH = (
    ROOT
    / "pilot_control_cwv"
    / "reports"
    / f"CWV_Summary_Details_Prototype_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
)

PILOT_COLOR = "4473D0"
SISTER_COLOR = "7CCAC2"
BASELINE_COLOR = "A3A3A3"
FLOOR_COLOR = "F2A7A7"
TITLE_FILL = "EAF2FF"
SECTION_FILL = "EEF4FF"
CARD_FILL = "F8FAFC"
TEXT_BLACK = "0F172A"
MUTED_TEXT = "64748B"
LINK_FILL = "EFF6FF"

PSI_BASELINE = 90
PSI_FLOOR = 60
GTM_BASELINE = 94
GTM_FLOOR = 70
CWV_DATES = ["2026-03-26", "2026-03-27", "2026-03-28", "2026-03-29", "2026-03-30", "2026-03-31"]


def load_pairs() -> list[dict]:
    cfg = json.loads(CONFIG_PATH.read_text())
    cohorts = {c["key"]: c for c in cfg["cohorts"] if c.get("active")}
    pairs = []
    for cohort in cohorts.values():
        if cohort["role"] != "pilot":
            continue
        sister = cohorts[cohort["sister_key"]]
        pairs.append({"pilot": cohort, "sister": sister})
    return pairs


def fetch_cwv_series(conn: sqlite3.Connection, property_id: str, kind: str) -> list[tuple[str, float | None]]:
    if kind == "psi":
        query = """
            SELECT metric_date, performance_score
            FROM pilot_control_psi_metrics
            WHERE property_id = ? AND strategy = 'mobile'
            ORDER BY metric_date DESC
            LIMIT 7
        """
    else:
        query = """
            SELECT metric_date, pagespeed_score
            FROM gtmetrix_metrics
            WHERE property_id = ?
            ORDER BY metric_date DESC
            LIMIT 7
        """
    rows = conn.execute(query, (property_id,)).fetchall()
    by_date = {row[0]: float(row[1]) for row in rows if row[1] is not None}
    return [(metric_date, by_date.get(metric_date)) for metric_date in CWV_DATES]


def avg(values: list[float | None]) -> float | None:
    clean = [v for v in values if v is not None]
    if not clean:
        return None
    return sum(clean) / len(clean)


def fmt_num(value: float | None) -> str:
    if value is None:
        return "n/a"
    return f"{value:.1f}"


def style_chart(
    chart: LineChart,
    data_ws,
    top_row: int,
    bottom_row: int,
    num_dates: int,
    colors: list[str],
    widths: list[int],
    dashes: list[str | None],
    height: float,
    width: float,
) -> None:
    chart.height = height
    chart.width = width
    chart.style = 2
    chart.legend = None
    chart.title = None
    chart.layout = Layout(
        manualLayout=ManualLayout(
            x=0.01,
            y=0.06,
            w=0.97,
            h=0.76,
            layoutTarget="inner",
        )
    )
    chart.graphical_properties = GraphicalProperties()
    chart.graphical_properties.noFill = True
    chart.graphical_properties.line.noFill = True
    chart.y_axis.delete = True
    chart.y_axis.majorGridlines = None
    chart.y_axis.spPr = None
    chart.x_axis.axPos = "b"
    chart.x_axis.tickLblPos = "low"
    chart.x_axis.majorTickMark = "none"
    chart.x_axis.spPr = GraphicalProperties()
    chart.x_axis.spPr.line.noFill = True
    chart.x_axis.majorGridlines = ChartLines()
    chart.x_axis.majorGridlines.spPr = GraphicalProperties()
    chart.x_axis.majorGridlines.spPr.line.noFill = True
    chart.plot_area.spPr = None

    max_col = 1 + num_dates
    data = Reference(data_ws, min_col=2, max_col=max_col, min_row=top_row, max_row=bottom_row)
    cats = Reference(data_ws, min_col=2, max_col=max_col, min_row=top_row - 1, max_row=top_row - 1)
    chart.add_data(data, titles_from_data=False, from_rows=True)
    chart.set_categories(cats)

    for idx, series in enumerate(chart.series):
        line = series.graphicalProperties.line
        line.solidFill = colors[idx]
        line.width = widths[idx]
        if dashes[idx]:
            line.prstDash = dashes[idx]
        marker = series.marker
        marker.symbol = "circle"
        marker.size = 3 if idx < 2 else 6
        marker.graphicalProperties.solidFill = colors[idx]
        marker.graphicalProperties.line.solidFill = colors[idx]


def add_sheet_title(ws, title: str, subtitle: str) -> None:
    ws.merge_cells("A1:L1")
    ws["A1"] = title
    ws["A1"].font = Font(size=18, bold=True, color=TEXT_BLACK)
    ws.merge_cells("A2:L2")
    ws["A2"] = subtitle
    ws["A2"].font = Font(size=10, color=MUTED_TEXT)


def add_pair_title(ws, row: int, pilot_name: str, sister_name: str) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    cell = ws.cell(row, 1)
    cell.value = CellRichText(
        TextBlock(InlineFont(rFont="Calibri", sz=15, b=True, color=PILOT_COLOR), pilot_name),
        TextBlock(InlineFont(rFont="Calibri", sz=15, b=True, color=TEXT_BLACK), " vs "),
        TextBlock(InlineFont(rFont="Calibri", sz=15, b=True, color=SISTER_COLOR), sister_name),
    )
    cell.fill = PatternFill("solid", fgColor=TITLE_FILL)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def write_value_labels(ws, top_row: int, labels: list[tuple[str, str, str]], start_col: int = 10) -> None:
    for idx, (label, value, color) in enumerate(labels):
        row = top_row + idx
        ws.cell(row, start_col, label).font = Font(size=9, bold=True, color=color)
        ws.cell(row, start_col).alignment = Alignment(horizontal="left")
        ws.cell(row, start_col + 1, value).font = Font(size=9, color=TEXT_BLACK)
        ws.cell(row, start_col + 1).alignment = Alignment(horizontal="right")


def add_summary_rollup(
    ws,
    data_ws,
    start_row: int,
    title: str,
    dates: list[str],
    pilot_values: list[float | None],
    sister_values: list[float | None],
    baseline: float,
    floor: float,
    data_row: int,
) -> int:
    for row in range(start_row, start_row + 5):
        for col in range(1, 13):
            ws.cell(row, col).fill = PatternFill("solid", fgColor=CARD_FILL)
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=12)
    ws.cell(start_row, 1, title).font = Font(size=11, bold=True, color=TEXT_BLACK)

    data_ws.cell(data_row, 1, title)
    for idx, day in enumerate(dates, start=2):
        data_ws.cell(data_row, idx, day)
        data_ws.cell(data_row + 1, idx, baseline)
        data_ws.cell(data_row + 2, idx, floor)
        data_ws.cell(data_row + 3, idx, pilot_values[idx - 2])
        data_ws.cell(data_row + 4, idx, sister_values[idx - 2])

    chart = LineChart()
    style_chart(
        chart,
        data_ws,
        data_row + 1,
        data_row + 4,
        len(dates),
        [BASELINE_COLOR, FLOOR_COLOR, PILOT_COLOR, SISTER_COLOR],
        [11000, 11000, 22000, 22000],
        ["sysDash", "sysDash", None, None],
        height=1.4,
        width=10.8,
    )
    ws.add_chart(chart, f"A{start_row + 1}")
    write_value_labels(
        ws,
        start_row + 1,
        [
            ("Pilot Avg", fmt_num(pilot_values[-1]), PILOT_COLOR),
            ("Sister Avg", fmt_num(sister_values[-1]), SISTER_COLOR),
            ("Baseline", fmt_num(baseline), BASELINE_COLOR),
            ("Floor", fmt_num(floor), FLOOR_COLOR),
        ],
    )
    return data_row + 6


def add_detail_card(
    ws,
    data_ws,
    start_row: int,
    pilot_name: str,
    sister_name: str,
    title: str,
    dates: list[str],
    pilot_values: list[float | None],
    sister_values: list[float | None],
    baseline: float,
    floor: float,
    data_row: int,
) -> int:
    add_pair_title(ws, start_row, pilot_name, sister_name)
    for row in range(start_row + 1, start_row + 7):
        for col in range(1, 13):
            ws.cell(row, col).fill = PatternFill("solid", fgColor=CARD_FILL)
    ws.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=12)
    ws.cell(start_row + 1, 1, title).font = Font(size=10, bold=True, color=TEXT_BLACK)

    data_ws.cell(data_row, 1, f"{pilot_name} {title}")
    for idx, day in enumerate(dates, start=2):
        data_ws.cell(data_row, idx, day)
        data_ws.cell(data_row + 1, idx, baseline)
        data_ws.cell(data_row + 2, idx, floor)
        data_ws.cell(data_row + 3, idx, pilot_values[idx - 2])
        data_ws.cell(data_row + 4, idx, sister_values[idx - 2])

    chart = LineChart()
    style_chart(
        chart,
        data_ws,
        data_row + 1,
        data_row + 4,
        len(dates),
        [BASELINE_COLOR, FLOOR_COLOR, PILOT_COLOR, SISTER_COLOR],
        [11000, 11000, 22000, 22000],
        ["sysDash", "sysDash", None, None],
        height=1.65,
        width=12.5,
    )
    ws.add_chart(chart, f"A{start_row + 2}")
    write_value_labels(
        ws,
        start_row + 2,
        [
            ("Pilot", fmt_num(pilot_values[-1]), PILOT_COLOR),
            ("Sister", fmt_num(sister_values[-1]), SISTER_COLOR),
            ("BL", fmt_num(baseline), BASELINE_COLOR),
            ("Floor", fmt_num(floor), FLOOR_COLOR),
        ],
    )
    return data_row + 6


def build_workbook() -> None:
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"
    detail_ws = wb.create_sheet("Details - CWV")
    data_ws = wb.create_sheet("Chart Data")
    data_ws.sheet_state = "hidden"

    for ws in (summary_ws, detail_ws):
        ws.sheet_view.showGridLines = False
        for col in range(1, 13):
            ws.column_dimensions[get_column_letter(col)].width = 9

    add_sheet_title(
        summary_ws,
        "Pilot KPI Summary",
        "High-level rollup view. Use the links in each section to jump to detailed property charts.",
    )
    add_sheet_title(
        detail_ws,
        "Core Web Vitals Details",
        "Property-pair detail for PSI and GTMetrix. Use Back to Summary links to return.",
    )

    conn = sqlite3.connect(DB_PATH)
    pairs = load_pairs()
    psi_by_pair = []
    gtm_by_pair = []
    for pair in pairs:
        psi_by_pair.append(
            (
                pair["pilot"]["display_name"],
                pair["sister"]["display_name"],
                fetch_cwv_series(conn, pair["pilot"]["property_id"], "psi"),
                fetch_cwv_series(conn, pair["sister"]["property_id"], "psi"),
            )
        )
        gtm_by_pair.append(
            (
                pair["pilot"]["display_name"],
                pair["sister"]["display_name"],
                fetch_cwv_series(conn, pair["pilot"]["property_id"], "gtm"),
                fetch_cwv_series(conn, pair["sister"]["property_id"], "gtm"),
            )
        )
    conn.close()

    dates = [datetime.fromisoformat(d).strftime("%-m/%d") for d in CWV_DATES]
    psi_pilot_avg = [avg([series[2][i][1] for series in psi_by_pair]) for i in range(len(CWV_DATES))]
    psi_sister_avg = [avg([series[3][i][1] for series in psi_by_pair]) for i in range(len(CWV_DATES))]
    gtm_pilot_avg = [avg([series[2][i][1] for series in gtm_by_pair]) for i in range(len(CWV_DATES))]
    gtm_sister_avg = [avg([series[3][i][1] for series in gtm_by_pair]) for i in range(len(CWV_DATES))]

    # Summary sheet
    row = 4
    summary_ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    summary_ws.cell(row, 1, "Core Web Vitals").font = Font(size=14, bold=True, color=TEXT_BLACK)
    summary_ws.cell(row, 1).fill = PatternFill("solid", fgColor=SECTION_FILL)
    row += 1
    for r in range(row, row + 11):
        for c in range(1, 13):
            summary_ws.cell(r, c).fill = PatternFill("solid", fgColor=CARD_FILL)
    summary_ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    summary_ws.cell(row, 1, "PSI Avg and GTMetrix Avg").font = Font(size=12, bold=True, color=TEXT_BLACK)
    summary_ws.merge_cells(start_row=row, start_column=9, end_row=row, end_column=12)
    link_cell = summary_ws.cell(row, 9, "View CWV Details")
    link_cell.font = Font(size=10, bold=True, color=PILOT_COLOR, underline="single")
    link_cell.fill = PatternFill("solid", fgColor=LINK_FILL)
    link_cell.alignment = Alignment(horizontal="center")
    link_cell.hyperlink = "#'Details - CWV'!A1"
    summary_ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=12)
    summary_ws.cell(row + 1, 1, "Average pilot line vs average sister line across the five property pairs").font = Font(size=9, color=MUTED_TEXT)

    data_row = 1
    data_row = add_summary_rollup(summary_ws, data_ws, row + 2, "PSI Avg", dates, psi_pilot_avg, psi_sister_avg, PSI_BASELINE, PSI_FLOOR, data_row)
    data_row = add_summary_rollup(summary_ws, data_ws, row + 7, "GTMetrix Avg", dates, gtm_pilot_avg, gtm_sister_avg, GTM_BASELINE, GTM_FLOOR, data_row)

    # Details sheet
    detail_row = 4
    detail_ws.merge_cells(start_row=detail_row, start_column=1, end_row=detail_row, end_column=9)
    detail_ws.cell(detail_row, 1, "Core Web Vitals").font = Font(size=14, bold=True, color=TEXT_BLACK)
    detail_ws.cell(detail_row, 1).fill = PatternFill("solid", fgColor=SECTION_FILL)
    detail_ws.merge_cells(start_row=detail_row, start_column=10, end_row=detail_row, end_column=12)
    back_cell = detail_ws.cell(detail_row, 10, "Back to Summary")
    back_cell.font = Font(size=10, bold=True, color=PILOT_COLOR, underline="single")
    back_cell.fill = PatternFill("solid", fgColor=LINK_FILL)
    back_cell.alignment = Alignment(horizontal="center")
    back_cell.hyperlink = "#Summary!A1"
    detail_row += 2

    detail_ws.merge_cells(start_row=detail_row, start_column=1, end_row=detail_row, end_column=12)
    detail_ws.cell(detail_row, 1, "PSI").font = Font(size=12, bold=True, color=TEXT_BLACK)
    detail_ws.cell(detail_row, 1).fill = PatternFill("solid", fgColor=SECTION_FILL)
    detail_row += 1
    for pilot_name, sister_name, pilot_series, sister_series in psi_by_pair:
        data_row = add_detail_card(
            detail_ws,
            data_ws,
            detail_row,
            pilot_name,
            sister_name,
            "PSI",
            dates,
            [v for _, v in pilot_series],
            [v for _, v in sister_series],
            PSI_BASELINE,
            PSI_FLOOR,
            data_row,
        )
        detail_row += 8

    detail_ws.merge_cells(start_row=detail_row, start_column=1, end_row=detail_row, end_column=12)
    detail_ws.cell(detail_row, 1, "GTMetrix").font = Font(size=12, bold=True, color=TEXT_BLACK)
    detail_ws.cell(detail_row, 1).fill = PatternFill("solid", fgColor=SECTION_FILL)
    detail_row += 1
    for pilot_name, sister_name, pilot_series, sister_series in gtm_by_pair:
        data_row = add_detail_card(
            detail_ws,
            data_ws,
            detail_row,
            pilot_name,
            sister_name,
            "GTMetrix",
            dates,
            [v for _, v in pilot_series],
            [v for _, v in sister_series],
            GTM_BASELINE,
            GTM_FLOOR,
            data_row,
        )
        detail_row += 8

    summary_ws.freeze_panes = "A4"
    detail_ws.freeze_panes = "A4"
    wb.save(OUTPUT_PATH)


if __name__ == "__main__":
    build_workbook()
    print(OUTPUT_PATH)
