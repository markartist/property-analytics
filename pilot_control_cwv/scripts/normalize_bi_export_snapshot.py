from __future__ import annotations

import argparse
import csv
import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path("/Users/mark/Property_Analytics")
CONFIG_PATH = ROOT / "pilot_control_cwv" / "config" / "pilot_control_cwv_config.json"
DEFAULT_INPUT = Path("/Users/mark/Downloads/bi-03-31-mc.xlsx")
DEFAULT_OUTPUT_DIR = ROOT / "pilot_control_cwv" / "reports"

PROPERTY_ALIASES = {
    "district": "The District Universal Boulevard",
    "harrison": "The Harrison",
    "spring branch": "Avasa Spring Branch",
    "northbridge": "Northbridge at Millenia Lake",
    "whitney": "The Whitney",
    "wurzbach": "Park on Wurzbach",
}

HEADER_RE = re.compile(
    r"^(?P<metric>[A-Za-z0-9/]+)\s+(?P<window>Yesterday|Daily Avg T7D|Daily Avg T15D|Daily Avg T30D|Daily Avg T60D|Daily Avg T90D|T7D|T15D|T30D|T60D|T90D)(?P<sister>\s+Sister)?(?P<py>\s+PY)?$"
)


def load_pairs() -> tuple[dict[str, dict], dict[str, dict]]:
    cfg = json.loads(CONFIG_PATH.read_text())
    cohorts_by_key = {c["key"]: c for c in cfg["cohorts"] if c.get("active")}
    by_name: dict[str, dict] = {}
    for cohort in cohorts_by_key.values():
        by_name[cohort["display_name"]] = cohort
    return by_name, cohorts_by_key


def canonical_property_name(raw_name: str, known_names: dict[str, dict]) -> str:
    if raw_name in known_names:
        return raw_name

    normalized = raw_name.strip().lower()
    for display_name in known_names:
        if normalized == display_name.lower():
            return display_name

    if normalized in PROPERTY_ALIASES:
        return PROPERTY_ALIASES[normalized]

    for alias, canonical in PROPERTY_ALIASES.items():
        if normalized == alias:
            return canonical

    raise KeyError(f"Unknown BI property name: {raw_name}")


def parse_header(value: str) -> dict | None:
    if value in {"Property", "Conv Src (website vs other)"} or value is None:
        return None
    match = HEADER_RE.match(str(value).strip())
    if not match:
        return None
    parsed = match.groupdict()
    window = parsed["window"]
    if window.startswith("Daily Avg "):
        window = window.replace("Daily Avg ", "") + "_DAILY_AVG"
    return {
        "metric_code": parsed["metric"],
        "window": window,
        "is_sister": bool(parsed["sister"]),
        "is_prior_year": bool(parsed["py"]),
    }


def derive_snapshot_date(input_path: Path, explicit_date: str | None) -> str:
    if explicit_date:
        return explicit_date
    match = re.search(r"Run(\d{8})", input_path.name, re.IGNORECASE)
    if match:
        raw = match.group(1)
        return f"{raw[0:4]}-{raw[4:6]}-{raw[6:8]}"
    match = re.search(r"(\d{2})-(\d{2})-mc", input_path.name)
    if match:
        month, day = match.groups()
        return f"{datetime.now().year}-{month}-{day}"
    return datetime.now().strftime("%Y-%m-%d")


def extract_bi_export_records(input_path: Path, snapshot_date: str) -> tuple[list[dict], list[dict]]:
    wb = load_workbook(input_path, read_only=False, data_only=True)
    ws = wb[wb.sheetnames[0]]
    sheet_name = ws.title
    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    parsed_headers = [parse_header(value) for value in headers]
    known_names, cohorts_by_key = load_pairs()

    raw_records: list[dict] = []
    normalized_records: list[dict] = []
    current_property_name: str | None = None
    for row_idx in range(2, ws.max_row + 1):
        property_name = ws.cell(row_idx, 1).value
        conv_source = ws.cell(row_idx, 2).value
        if property_name:
            current_property_name = str(property_name)
        if not current_property_name or not conv_source:
            continue

        canonical_name = canonical_property_name(current_property_name, known_names)
        cohort = known_names[canonical_name]
        sister_key = cohort.get("sister_key")
        sister_name = cohorts_by_key[sister_key]["display_name"] if sister_key and sister_key in cohorts_by_key else None

        for col_idx in range(3, ws.max_column + 1):
            header = parsed_headers[col_idx - 1]
            if not header:
                continue
            value = ws.cell(row_idx, col_idx).value
            if value is None:
                continue
            header_raw = str(headers[col_idx - 1]).strip()
            raw_records.append(
                {
                    "snapshot_date": snapshot_date,
                    "source_file": str(input_path),
                    "source_sheet": sheet_name,
                    "source_row": row_idx,
                    "source_column": col_idx,
                    "header_raw": header_raw,
                    "property_name_raw": current_property_name,
                    "conv_source_raw": str(conv_source),
                    "value_raw": float(value),
                }
            )
            normalized_records.append(
                {
                    "snapshot_date": snapshot_date,
                    "property_name": canonical_name,
                    "property_id": cohort["property_id"],
                    "role": cohort["role"],
                    "sister_property_name": sister_name,
                    "conv_source": conv_source,
                    "metric_code": header["metric_code"],
                    "window": header["window"],
                    "comparison_type": (
                        "sister" if header["is_sister"] else "prior_year" if header["is_prior_year"] else "current"
                    ),
                    "value": float(value),
                    "source_file": str(input_path),
                    "source_sheet": sheet_name,
                    "source_row": row_idx,
                    "source_column": col_idx,
                    "header_raw": header_raw,
                }
            )
    return raw_records, normalized_records


def normalize_bi_export(input_path: Path, snapshot_date: str) -> list[dict]:
    _, normalized_records = extract_bi_export_records(input_path, snapshot_date)
    return normalized_records


def write_csv(records: list[dict], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "snapshot_date",
        "property_name",
        "property_id",
        "role",
        "sister_property_name",
        "conv_source",
        "metric_code",
        "window",
        "comparison_type",
        "value",
        "source_file",
        "source_sheet",
        "source_row",
        "source_column",
        "header_raw",
    ]
    with output_path.open("w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(records)


def main() -> None:
    parser = argparse.ArgumentParser(description="Normalize a BI export snapshot for the pilot KPI report.")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--date", help="Snapshot date in YYYY-MM-DD format. Defaults from filename when possible.")
    parser.add_argument("--output", type=Path, help="Output CSV path.")
    args = parser.parse_args()

    snapshot_date = derive_snapshot_date(args.input, args.date)
    output_path = args.output or (DEFAULT_OUTPUT_DIR / f"pilot_bi_snapshot_normalized_{snapshot_date}.csv")

    records = normalize_bi_export(args.input, snapshot_date)
    write_csv(records, output_path)

    print(output_path)
    print(f"rows={len(records)}")


if __name__ == "__main__":
    main()
