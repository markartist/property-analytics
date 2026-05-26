from __future__ import annotations

import csv
import json
import sqlite3
import subprocess
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path("/Users/mark/Property_Analytics")
DB_PATH = ROOT / "data" / "portfolio_analytics.db"
CONFIG_PATH = ROOT / "pilot_control_cwv" / "config" / "pilot_control_cwv_config.json"
BI_SERIES_PATH = ROOT / "pilot_control_cwv" / "reports" / "pilot_bi_report_series.csv"
SNAPSHOT_DIR = ROOT / "apps" / "pilot-tracker-standalone" / "public" / "pilot-kpi" / "latest"
EXPORT_SCRIPT = ROOT / "pilot_control_cwv" / "scripts" / "export_dashboard_snapshots.py"
OUTPUT_PATH = (
    ROOT
    / "pilot_control_cwv"
    / "reports"
    / f"Pilot_KPI_Summary_Details_Full_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
)

PILOT_COLOR = "4473D0"
SISTER_COLOR = "7CCAC2"
BASELINE_COLOR = "A3A3A3"
FLOOR_COLOR = "F2A7A7"
SISTER_BASELINE_COLOR = "D6D3D1"
TITLE_FILL = "EAF2FF"
SECTION_FILL = "EEF4FF"
CARD_FILL = "F8FAFC"
TEXT_BLACK = "0F172A"
MUTED_TEXT = "64748B"
LINK_FILL = "EFF6FF"
PENDING_FILL = "FFF7ED"
PENDING_TEXT = "9A3412"

PSI_BASELINE = 90
PSI_FLOOR = 60
GTM_BASELINE = 94
GTM_FLOOR = 70
CWV_DATES = ["2026-03-26", "2026-03-27", "2026-03-28", "2026-03-29", "2026-03-30", "2026-03-31"]
SHEET_MAX_COL = 16
SUMMARY_LINK_START_COL = 13
SUMMARY_LABEL_START_COL = 12

BI_SECTION_GROUPS = {
    "Traffic & Engagement": [
        ("organic_traffic_pct", "Organic Traffic as a % of Unique Users"),
        ("organic_volume_t7", "Organic Volume #"),
        ("high_intent_rate", "High Intent User Rate"),
        ("organic_yoy_vol_delta", "Organic Volume # - YoY"),
        ("organic_share_yoy_points_delta", "Organic Share % - YoY"),
    ],
    "Funnel": [
        ("lead_to_available_unit_rate", "Lead (Guest Card) to Available Unit Rate"),
        ("website_sales_funnel_price_quote", "Website Sales Funnel - Price Quote"),
        ("website_sales_funnel_visits_schedule_tour", "Website Sales Funnel - Visits (Schedule a Tour)"),
        ("website_sales_funnel_completed_applications", "Website Sales Funnel - Completed Applications"),
        ("website_funnel_conversions_click_to_call", "Website Funnel Conversions - Click to Call / Phone"),
        ("website_funnel_conversions_contact_form", "Website Funnel Conversions - Contact Form"),
    ],
}


def load_pairs() -> list[dict]:
    cfg = json.loads(CONFIG_PATH.read_text())
    cohorts = {c["key"]: c for c in cfg["cohorts"] if c.get("active")}
    pairs = []
    for cohort in cohorts.values():
        if cohort["role"] != "pilot":
            continue
        sister = cohorts[cohort["sister_key"]]
        pairs.append({"pilot": cohort, "sister": sister})
    return pairs


def compact_name(name: str) -> str:
    replacements = {
        "The District Universal Boulevard": "District",
        "Northbridge at Millenia Lake": "Northbridge",
        "Avasa Spring Branch": "Spring Branch",
        "Park on Wurzbach": "Wurzbach",
        "Champions Green": "Champions",
        "Calais Midtown": "Calais",
        "The Harrison": "Harrison",
        "The Whitney": "Whitney",
        "Ventana": "Ventana",
        "Axial Buckhead": "Axial",
    }
    return replacements.get(name, name)


def fetch_cwv_series(conn: sqlite3.Connection, property_id: str, kind: str) -> list[tuple[str, float | None]]:
    if kind == "psi":
        query = """
            SELECT metric_date, performance_score
            FROM pilot_control_psi_metrics
            WHERE property_id = ? AND strategy = 'mobile'
            ORDER BY metric_date DESC
            LIMIT 7
        """
    else:
        query = """
            SELECT metric_date, pagespeed_score
            FROM gtmetrix_metrics
            WHERE property_id = ?
            ORDER BY metric_date DESC
            LIMIT 7
        """
    rows = conn.execute(query, (property_id,)).fetchall()
    by_date = {row[0]: float(row[1]) for row in rows if row[1] is not None}
    return [(metric_date, by_date.get(metric_date)) for metric_date in CWV_DATES]


def avg(values: list[float | None]) -> float | None:
    clean = [v for v in values if v is not None]
    if not clean:
        return None
    return sum(clean) / len(clean)


def backfill_leading(values: list[float | None]) -> list[float | None]:
    first = next((v for v in values if v is not None), None)
    if first is None:
        return values[:]
    out = values[:]
    for i, val in enumerate(out):
        if val is None:
            out[i] = first
        else:
            break
    return out


def fmt_num(value: float | None, format_name: str = "score") -> str:
    if value is None:
        return "n/a"
    if format_name == "percent":
        return f"{value * 100:.1f}%"
    if format_name == "count":
        return f"{value:.0f}"
    if format_name == "points":
        return f"{value:.1f} pts"
    return f"{value:.1f}"


def latest_non_none(values: list[float | None]) -> float | None:
    for value in reversed(values):
        if value is not None:
            return value
    return None


def style_chart(
    chart: LineChart,
    data_ws,
    top_row: int,
    bottom_row: int,
    num_dates: int,
    colors: list[str],
    widths: list[int],
    dashes: list[str | None],
    height: float,
    width: float,
) -> None:
    chart.height = height
    chart.width = width
    chart.style = 2
    chart.legend = None
    chart.title = None
    chart.layout = Layout(
        manualLayout=ManualLayout(x=0.01, y=0.04, w=0.97, h=0.80, layoutTarget="inner")
    )
    chart.graphical_properties = GraphicalProperties()
    chart.graphical_properties.noFill = True
    chart.graphical_properties.line.noFill = True
    chart.y_axis.delete = True
    chart.y_axis.majorGridlines = None
    chart.y_axis.spPr = None
    chart.y_axis.majorTickMark = "none"
    chart.x_axis.delete = False
    chart.x_axis.axPos = "b"
    chart.x_axis.tickLblPos = "low"
    chart.x_axis.majorTickMark = "none"
    chart.x_axis.spPr = GraphicalProperties()
    chart.x_axis.spPr.line.noFill = True
    chart.x_axis.majorGridlines = ChartLines()
    chart.x_axis.majorGridlines.spPr = GraphicalProperties()
    chart.x_axis.majorGridlines.spPr.line.noFill = True
    chart.plot_area.spPr = None

    max_col = 1 + num_dates
    data = Reference(data_ws, min_col=2, max_col=max_col, min_row=top_row, max_row=bottom_row)
    cats = Reference(data_ws, min_col=2, max_col=max_col, min_row=top_row - 1, max_row=top_row - 1)
    chart.add_data(data, titles_from_data=False, from_rows=True)
    chart.set_categories(cats)

    for idx, series in enumerate(chart.series):
        line = series.graphicalProperties.line
        line.solidFill = colors[idx]
        line.width = widths[idx]
        if dashes[idx]:
            line.prstDash = dashes[idx]
        marker = series.marker
        marker.symbol = "circle"
        marker.size = 3 if idx < 2 else 6
        marker.graphicalProperties.solidFill = colors[idx]
        marker.graphicalProperties.line.solidFill = colors[idx]


def add_sheet_title(ws, title: str, subtitle: str) -> None:
    ws.merge_cells(f"A1:{get_column_letter(SHEET_MAX_COL)}1")
    ws["A1"] = title
    ws["A1"].font = Font(size=18, bold=True, color=TEXT_BLACK)
    ws.merge_cells(f"A2:{get_column_letter(SHEET_MAX_COL)}2")
    ws["A2"] = subtitle
    ws["A2"].font = Font(size=10, color=MUTED_TEXT)


def add_nav_link(ws, row: int, text: str, target: str, col_start: int = SUMMARY_LINK_START_COL, col_end: int = SHEET_MAX_COL) -> None:
    ws.merge_cells(start_row=row, start_column=col_start, end_row=row, end_column=col_end)
    cell = ws.cell(row, col_start, text)
    cell.font = Font(size=10, bold=True, color=PILOT_COLOR, underline="single")
    cell.fill = PatternFill("solid", fgColor=LINK_FILL)
    cell.alignment = Alignment(horizontal="center")
    cell.hyperlink = target


def add_pair_title(ws, row: int, pilot_name: str, sister_name: str) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=SHEET_MAX_COL)
    cell = ws.cell(row, 1)
    cell.value = CellRichText(
        TextBlock(InlineFont(rFont="Calibri", sz=15, b=True, color=PILOT_COLOR), compact_name(pilot_name)),
        TextBlock(InlineFont(rFont="Calibri", sz=15, b=True, color=TEXT_BLACK), " vs "),
        TextBlock(InlineFont(rFont="Calibri", sz=15, b=True, color=SISTER_COLOR), compact_name(sister_name)),
    )
    cell.fill = PatternFill("solid", fgColor=TITLE_FILL)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def add_section_header(ws, row: int, text: str, expanded: bool = True) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=SHEET_MAX_COL)
    cue = "▼" if expanded else "▶"
    cell = ws.cell(row, 1, f"{cue} {text}")
    cell.font = Font(size=12, bold=True, color=TEXT_BLACK)
    cell.fill = PatternFill("solid", fgColor=SECTION_FILL)
    cell.alignment = Alignment(horizontal="left")


def apply_number_format(cell, value: float | None, format_name: str) -> None:
    if value is None:
        cell.value = "n/a"
        return
    cell.value = value
    if format_name == "percent":
        cell.number_format = "0.0%"
    elif format_name == "count":
        cell.number_format = "0"
    elif format_name == "points":
        cell.number_format = '0.0 "pts"'
    else:
        cell.number_format = "0.0"


def write_value_labels(
    ws,
    top_row: int,
    labels: list[tuple[str, float | None, str, str]],
    start_col: int = SUMMARY_LABEL_START_COL,
) -> None:
    for idx, (label, value, color, format_name) in enumerate(labels):
        row = top_row + idx
        ws.cell(row, start_col, label).font = Font(size=9, bold=True, color=color)
        ws.cell(row, start_col).alignment = Alignment(horizontal="left")
        value_cell = ws.cell(row, start_col + 1)
        value_cell.font = Font(size=9, color=TEXT_BLACK)
        value_cell.alignment = Alignment(horizontal="right")
        apply_number_format(value_cell, value, format_name)


def write_date_strip(ws, row: int, dates: list[str], start_col: int = 2, end_col: int = 8) -> None:
    if not dates:
        return
    display_dates = dates
    if len(dates) > 4 and (end_col - start_col) >= 8:
        mid_idx = len(dates) // 2
        display_dates = [dates[0], dates[mid_idx], dates[-1]]
    span = end_col - start_col
    if len(display_dates) == 1 or span <= 0:
        positions = [start_col]
    else:
        positions = [
            start_col + round((span * idx) / (len(display_dates) - 1))
            for idx in range(len(display_dates))
        ]
    max_used_col = end_col
    for col in range(start_col, end_col + 1):
        ws.cell(row, col).fill = PatternFill("solid", fgColor=CARD_FILL)
    for date, col in zip(display_dates, positions):
        cell = ws.cell(row, col, date)
        cell.font = Font(size=12, bold=True, color=TEXT_BLACK)
        cell.alignment = Alignment(horizontal="center")
    for col in range(start_col, end_col + 1):
        if col not in positions:
            ws.cell(row, col, "").fill = PatternFill("solid", fgColor=CARD_FILL)


def load_bi_grouped() -> dict[str, dict[tuple[str, str], list[dict]]]:
    with BI_SERIES_PATH.open() as f:
        rows = list(csv.DictReader(f))
    grouped: dict[str, dict[tuple[str, str], list[dict]]] = defaultdict(lambda: defaultdict(list))
    for row in rows:
        grouped[row["section_key"]][(row["pilot_property_name"], row["sister_property_name"])].append(row)
    for section_key in grouped:
        for pair_key in grouped[section_key]:
            grouped[section_key][pair_key].sort(key=lambda r: r["snapshot_date"])
    return grouped


def load_snapshot(name: str) -> dict:
    return json.loads((SNAPSHOT_DIR / name).read_text())


def load_snapshots() -> tuple[dict, dict, dict]:
    subprocess.run(["python3", str(EXPORT_SCRIPT)], check=True)
    return load_snapshot("cwv.json"), load_snapshot("traffic.json"), load_snapshot("funnel.json")


def metric_from_list(metrics: list[dict], metric_key: str) -> dict:
    for metric in metrics:
        if metric.get("metric_key") == metric_key:
            return metric
    raise KeyError(metric_key)


def pair_from_entries(entries: list[dict], pilot_name: str, sister_name: str) -> dict:
    for entry in entries:
        identity = entry["identity"]
        if identity["pilot"]["name"] == pilot_name and identity["sister"]["name"] == sister_name:
            return entry
    raise KeyError((pilot_name, sister_name))


def add_summary_rollup(
    ws,
    data_ws,
    start_row: int,
    title: str,
    dates: list[str],
    pilot_values: list[float | None],
    sister_values: list[float | None],
    baseline: float | None,
    floor: float | None,
    data_row: int,
    format_name: str = "score",
    sister_baseline: float | None = None,
) -> int:
    for row in range(start_row, start_row + 6):
        for col in range(1, SHEET_MAX_COL + 1):
            ws.cell(row, col).fill = PatternFill("solid", fgColor=CARD_FILL)
    ws.row_dimensions[start_row].height = 18
    ws.row_dimensions[start_row + 1].height = 62
    ws.row_dimensions[start_row + 2].height = 20
    ws.row_dimensions[start_row + 3].height = 18
    ws.row_dimensions[start_row + 4].height = 18
    ws.row_dimensions[start_row + 5].height = 18
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=SHEET_MAX_COL)
    ws.cell(start_row, 1, title).font = Font(size=11, bold=True, color=TEXT_BLACK)

    if dates:
        data_ws.cell(data_row, 1, title)
        for idx, day in enumerate(dates, start=2):
            data_ws.cell(data_row, idx, day)
            row_offset = 1
            if baseline is not None:
                data_ws.cell(data_row + row_offset, idx, baseline)
                row_offset += 1
            if sister_baseline is not None:
                data_ws.cell(data_row + row_offset, idx, sister_baseline)
                row_offset += 1
            if floor is not None:
                data_ws.cell(data_row + row_offset, idx, floor)
                row_offset += 1
            data_ws.cell(data_row + row_offset, idx, pilot_values[idx - 2])
            data_ws.cell(data_row + row_offset + 1, idx, sister_values[idx - 2])

    colors = []
    widths = []
    dashes = []
    top_row = data_row + 1
    current_bottom = data_row
    label_rows = [
        ("Pilot Avg", latest_non_none(pilot_values), PILOT_COLOR, format_name),
        ("Sister Avg", latest_non_none(sister_values), SISTER_COLOR, format_name),
    ]
    if baseline is not None:
        colors.append(BASELINE_COLOR)
        widths.append(11000)
        dashes.append("sysDash")
        current_bottom += 1
        label_rows.append(("Baseline", baseline, BASELINE_COLOR, format_name))
    if sister_baseline is not None:
        colors.append(SISTER_BASELINE_COLOR)
        widths.append(11000)
        dashes.append("sysDash")
        current_bottom += 1
        label_rows.append(("Sister BL", sister_baseline, SISTER_BASELINE_COLOR, format_name))
    if floor is not None:
        colors.append(FLOOR_COLOR)
        widths.append(11000)
        dashes.append("sysDash")
        current_bottom += 1
        label_rows.append(("Floor", floor, FLOOR_COLOR, format_name))
    colors.extend([PILOT_COLOR, SISTER_COLOR])
    widths.extend([22000, 22000])
    dashes.extend([None, None])
    current_bottom += 2

    if dates:
        chart = LineChart()
        style_chart(chart, data_ws, top_row, current_bottom, len(dates), colors, widths, dashes, height=2.35, width=18.5)
        ws.add_chart(chart, f"A{start_row + 1}")
    write_value_labels(ws, start_row + 1, label_rows[:4], start_col=SUMMARY_LABEL_START_COL)
    if len(label_rows) > 4:
        write_value_labels(ws, start_row + 5, label_rows[4:], start_col=SUMMARY_LABEL_START_COL)
    return current_bottom + 2


def add_pending_block(ws, row: int, title: str, note: str) -> int:
    for r in range(row, row + 3):
        for c in range(1, SHEET_MAX_COL + 1):
            ws.cell(r, c).fill = PatternFill("solid", fgColor=PENDING_FILL)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=SHEET_MAX_COL)
    ws.cell(row, 1, title).font = Font(size=11, bold=True, color=TEXT_BLACK)
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 2, end_column=SHEET_MAX_COL)
    cell = ws.cell(row + 1, 1, note)
    cell.font = Font(size=10, bold=True, color=PENDING_TEXT)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    return row + 4


def add_detail_card(
    ws,
    data_ws,
    start_row: int,
    pilot_name: str,
    sister_name: str,
    title: str,
    dates: list[str],
    pilot_values: list[float | None],
    sister_values: list[float | None],
    baseline: float | None,
    floor: float | None,
    data_row: int,
    format_name: str = "score",
    sister_baseline: float | None = None,
) -> int:
    add_pair_title(ws, start_row, pilot_name, sister_name)
    for row in range(start_row + 1, start_row + 8):
        for col in range(1, SHEET_MAX_COL + 1):
            ws.cell(row, col).fill = PatternFill("solid", fgColor=CARD_FILL)
    ws.row_dimensions[start_row].height = 22
    ws.row_dimensions[start_row + 1].height = 18
    ws.row_dimensions[start_row + 2].height = 30
    ws.row_dimensions[start_row + 3].height = 16
    ws.row_dimensions[start_row + 4].height = 14
    ws.row_dimensions[start_row + 5].height = 14
    ws.row_dimensions[start_row + 6].height = 14
    ws.row_dimensions[start_row + 7].height = 12
    ws.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=SHEET_MAX_COL)
    ws.cell(start_row + 1, 1, title).font = Font(size=10, bold=True, color=TEXT_BLACK)

    if dates:
        data_ws.cell(data_row, 1, f"{pilot_name} {title}")
        for idx, day in enumerate(dates, start=2):
            data_ws.cell(data_row, idx, day)
            row_offset = 1
            if baseline is not None:
                data_ws.cell(data_row + row_offset, idx, baseline)
                row_offset += 1
            if sister_baseline is not None:
                data_ws.cell(data_row + row_offset, idx, sister_baseline)
                row_offset += 1
            if floor is not None:
                data_ws.cell(data_row + row_offset, idx, floor)
                row_offset += 1
            data_ws.cell(data_row + row_offset, idx, pilot_values[idx - 2])
            data_ws.cell(data_row + row_offset + 1, idx, sister_values[idx - 2])

    colors = []
    widths = []
    dashes = []
    top_row = data_row + 1
    current_bottom = data_row
    if baseline is not None:
        colors.append(BASELINE_COLOR)
        widths.append(11000)
        dashes.append("sysDash")
        current_bottom += 1
    if sister_baseline is not None:
        colors.append(SISTER_BASELINE_COLOR)
        widths.append(11000)
        dashes.append("sysDash")
        current_bottom += 1
    if floor is not None:
        colors.append(FLOOR_COLOR)
        widths.append(11000)
        dashes.append("sysDash")
        current_bottom += 1
    colors.extend([PILOT_COLOR, SISTER_COLOR])
    widths.extend([22000, 22000])
    dashes.extend([None, None])
    current_bottom += 2

    if dates:
        chart = LineChart()
        style_chart(chart, data_ws, top_row, current_bottom, len(dates), colors, widths, dashes, height=1.8, width=14.2)
        ws.add_chart(chart, f"A{start_row + 2}")

    labels = [
        ("Pilot", latest_non_none(pilot_values), PILOT_COLOR, format_name),
        ("Sister", latest_non_none(sister_values), SISTER_COLOR, format_name),
    ]
    if baseline is not None:
        labels.append(("BL", baseline, BASELINE_COLOR, format_name))
    if sister_baseline is not None:
        labels.append(("Sister BL", sister_baseline, SISTER_BASELINE_COLOR, format_name))
    if floor is not None:
        labels.append(("Floor", floor, FLOOR_COLOR, format_name))
    write_value_labels(ws, start_row + 2, labels[:4])
    if len(labels) > 4:
        write_value_labels(ws, start_row + 6, labels[4:])
    return current_bottom + 2


def build_workbook() -> None:
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"
    cwv_ws = wb.create_sheet("Details - CWV")
    traffic_ws = wb.create_sheet("Details - Traffic")
    funnel_ws = wb.create_sheet("Details - Funnel")
    data_ws = wb.create_sheet("Chart Data")
    data_ws.sheet_state = "hidden"

    for ws in (summary_ws, cwv_ws, traffic_ws, funnel_ws):
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.outlinePr.summaryBelow = False
        for col in range(1, SHEET_MAX_COL + 1):
            ws.column_dimensions[get_column_letter(col)].width = 9
        for col in range(1, 13):
            ws.column_dimensions[get_column_letter(col)].width = 10
        for col in range(13, SHEET_MAX_COL + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12
    summary_ws.sheet_view.zoomScale = 180
    cwv_ws.sheet_view.zoomScale = 120
    traffic_ws.sheet_view.zoomScale = 120
    funnel_ws.sheet_view.zoomScale = 120

    add_sheet_title(summary_ws, "Pilot KPI Summary", "Executive rollup view. Use the section links to jump to detailed grouped sheets.")
    add_sheet_title(cwv_ws, "Core Web Vitals Details", "")
    add_sheet_title(traffic_ws, "Traffic & Engagement Details", "")
    add_sheet_title(funnel_ws, "Funnel Details", "")

    pairs = load_pairs()
    cwv_snapshot, traffic_snapshot, funnel_snapshot = load_snapshots()

    data_row = 1
    # Summary sheet sections
    row = 4
    # CWV summary
    summary_ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    summary_ws.cell(row, 1, "Core Web Vitals").font = Font(size=14, bold=True, color=TEXT_BLACK)
    summary_ws.cell(row, 1).fill = PatternFill("solid", fgColor=SECTION_FILL)
    add_nav_link(summary_ws, row, "View CWV Details", "#'Details - CWV'!A1")
    row += 1
    for r in range(row, row + 12):
        for c in range(1, SHEET_MAX_COL + 1):
            summary_ws.cell(r, c).fill = PatternFill("solid", fgColor=CARD_FILL)
    psi_rollup = cwv_snapshot["rollups"]["psi"]
    gtm_rollup = cwv_snapshot["rollups"]["gtmetrix"]
    dates = [point["label"] for point in psi_rollup["series"]]
    psi_pilot_avg = [point["pilot_value"] for point in psi_rollup["series"]]
    psi_sister_avg = [point["sister_value"] for point in psi_rollup["series"]]
    gtm_pilot_avg = [point["pilot_value"] for point in gtm_rollup["series"]]
    gtm_sister_avg = [point["sister_value"] for point in gtm_rollup["series"]]
    data_row = add_summary_rollup(summary_ws, data_ws, row, "PSI Avg", dates, psi_pilot_avg, psi_sister_avg, psi_rollup.get("baseline"), psi_rollup.get("floor"), data_row)
    data_row = add_summary_rollup(summary_ws, data_ws, row + 6, "GTMetrix Avg", dates, gtm_pilot_avg, gtm_sister_avg, gtm_rollup.get("baseline"), gtm_rollup.get("floor"), data_row)
    row += 13

    # Traffic summary
    summary_ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    summary_ws.cell(row, 1, "Traffic & Engagement").font = Font(size=14, bold=True, color=TEXT_BLACK)
    summary_ws.cell(row, 1).fill = PatternFill("solid", fgColor=SECTION_FILL)
    add_nav_link(summary_ws, row, "View Traffic Details", "#'Details - Traffic'!A1")
    row += 1
    for metric_key, section_title in BI_SECTION_GROUPS["Traffic & Engagement"]:
        metric = metric_from_list(
            traffic_snapshot["metrics"],
            {
                "organic_traffic_pct": "organic_pct_unique_users",
                "organic_volume_t7": "organic_volume_t7",
                "high_intent_rate": "high_intent_user_rate",
                "organic_yoy_vol_delta": "organic_yoy_vol_delta",
                "organic_share_yoy_points_delta": "organic_share_yoy_points_delta",
            }[metric_key],
        )
        for r in range(row, row + 6):
            for c in range(1, SHEET_MAX_COL + 1):
                summary_ws.cell(r, c).fill = PatternFill("solid", fgColor=CARD_FILL)
        data_row = add_summary_rollup(
            summary_ws,
            data_ws,
            row,
            section_title,
            [point["label"] for point in metric["series"]],
            [point["pilot_value"] for point in metric["series"]],
            [point["sister_value"] for point in metric["series"]],
            metric.get("pilot_baseline"),
            None,
            data_row,
            metric["format"],
            metric.get("sister_baseline"),
        )
        row += 6

    # Funnel summary
    summary_ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    summary_ws.cell(row, 1, "Funnel").font = Font(size=14, bold=True, color=TEXT_BLACK)
    summary_ws.cell(row, 1).fill = PatternFill("solid", fgColor=SECTION_FILL)
    add_nav_link(summary_ws, row, "View Funnel Details", "#'Details - Funnel'!A1")
    row += 1
    for metric_key, section_title in BI_SECTION_GROUPS["Funnel"]:
        metric = metric_from_list(funnel_snapshot["metrics"], metric_key)
        data_row = add_summary_rollup(
            summary_ws,
            data_ws,
            row,
            section_title,
            [point["label"] for point in metric["series"]],
            [point["pilot_value"] for point in metric["series"]],
            [point["sister_value"] for point in metric["series"]],
            metric.get("pilot_baseline"),
            None,
            data_row,
            True,
            metric.get("sister_baseline"),
        )
        row += 6

    # Details - CWV
    add_nav_link(cwv_ws, 4, "Back to Summary", "#Summary!A1")
    drow = 6
    for title, metric_key in [
        ("PSI", "psi"),
        ("GTMetrix", "gtmetrix"),
    ]:
        metric_rollup = cwv_snapshot["rollups"][metric_key]
        add_section_header(cwv_ws, drow, title, expanded=True)
        section_start = drow + 1
        drow += 1
        for pair_entry in cwv_snapshot["pairs"]:
            pilot_name = pair_entry["identity"]["pilot"]["name"]
            sister_name = pair_entry["identity"]["sister"]["name"]
            metric_data = pair_entry["metrics"][metric_key]
            data_row = add_detail_card(
                cwv_ws,
                data_ws,
                drow,
                pilot_name,
                sister_name,
                title,
                [point["label"] for point in metric_data["series"]],
                [point["pilot_value"] for point in metric_data["series"]],
                [point["sister_value"] for point in metric_data["series"]],
                metric_rollup.get("baseline"),
                metric_rollup.get("floor"),
                data_row,
                False,
            )
            drow += 8
        cwv_ws.row_dimensions.group(section_start, drow - 1, outline_level=1, hidden=False)
        drow += 1

    # Details - Traffic
    add_nav_link(traffic_ws, 4, "Back to Summary", "#Summary!A1")
    drow = 6
    traffic_metric_map = {
        "Organic Traffic as a % of Unique Users": "organic_pct_unique_users",
        "Organic Volume #": "organic_volume_t7",
        "High Intent User Rate": "high_intent_user_rate",
        "Organic Volume # - YoY": "organic_yoy_vol_delta",
        "Organic Share % - YoY": "organic_share_yoy_points_delta",
    }
    for section_title, metric_key in traffic_metric_map.items():
        metric = metric_from_list(traffic_snapshot["metrics"], metric_key)
        add_section_header(traffic_ws, drow, section_title, expanded=True)
        section_start = drow + 1
        drow += 1
        if metric["pairs"]:
            for pair in metric["pairs"]:
                data_row = add_detail_card(
                    traffic_ws,
                    data_ws,
                    drow,
                    pair["identity"]["pilot"]["name"],
                    pair["identity"]["sister"]["name"],
                    section_title,
                    [point["label"] for point in pair["series"]],
                    [point["pilot_value"] for point in pair["series"]],
                    [point["sister_value"] for point in pair["series"]],
                    pair.get("pilot_baseline"),
                    None,
                    data_row,
                    metric["format"],
                    pair.get("sister_baseline"),
                )
                drow += 8
            traffic_ws.row_dimensions.group(section_start, drow - 1, outline_level=1, hidden=False)
        else:
            data_row = add_summary_rollup(
                traffic_ws,
                data_ws,
                drow,
                f"{section_title} (Pilot Avg vs Sister Avg)",
                [point["label"] for point in metric["series"]],
                [point["pilot_value"] for point in metric["series"]],
                [point["sister_value"] for point in metric["series"]],
                metric.get("pilot_baseline"),
                None,
                data_row,
                metric["format"],
                metric.get("sister_baseline"),
            )
            drow += 6
            traffic_ws.row_dimensions.group(section_start, drow - 1, outline_level=1, hidden=False)
        drow += 1

    # Details - Funnel
    add_nav_link(funnel_ws, 4, "Back to Summary", "#Summary!A1")
    drow = 6
    for section_key, section_title in BI_SECTION_GROUPS["Funnel"]:
        metric = metric_from_list(funnel_snapshot["metrics"], section_key)
        add_section_header(funnel_ws, drow, section_title, expanded=True)
        section_start = drow + 1
        drow += 1
        for pair_rows in metric["pairs"]:
            data_row = add_detail_card(
                funnel_ws,
                data_ws,
                drow,
                pair_rows["identity"]["pilot"]["name"],
                pair_rows["identity"]["sister"]["name"],
                section_title,
                [point["label"] for point in pair_rows["series"]],
                [point["pilot_value"] for point in pair_rows["series"]],
                [point["sister_value"] for point in pair_rows["series"]],
                pair_rows.get("pilot_baseline"),
                None,
                data_row,
                "percent",
                pair_rows.get("sister_baseline"),
            )
            drow += 8
        funnel_ws.row_dimensions.group(section_start, drow - 1, outline_level=1, hidden=False)
        drow += 1

    for ws in (summary_ws, cwv_ws, traffic_ws, funnel_ws):
        ws.freeze_panes = "A4"

    wb.save(OUTPUT_PATH)


if __name__ == "__main__":
    build_workbook()
    print(OUTPUT_PATH)
