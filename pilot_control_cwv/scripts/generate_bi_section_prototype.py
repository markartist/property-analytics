from __future__ import annotations

import csv
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path("/Users/mark/Property_Analytics")
INPUT_PATH = ROOT / "pilot_control_cwv" / "reports" / "pilot_bi_report_series.csv"
OUTPUT_PATH = ROOT / "pilot_control_cwv" / "reports" / f"BI_Section_Prototype_{datetime.now().strftime('%Y-%m-%d')}.xlsx"

PILOT_COLOR = "4473D0"
SISTER_COLOR = "7CCAC2"
PILOT_BASELINE_COLOR = "A3A3A3"
SISTER_BASELINE_COLOR = "D6D3D1"
CARD_FILL = "F8FAFC"
TITLE_FILL = "EAF2FF"
SECTION_FILL = "EEF4FF"
TEXT_BLACK = "0F172A"
MUTED_TEXT = "64748B"
PENDING_FILL = "FFF7ED"

SECTION_TITLES = {
    "lead_to_available_unit_rate": "Lead (Guest Card) to Available Unit Rate",
    "website_sales_funnel_price_quote": "Website Sales Funnel - Price Quote",
    "website_sales_funnel_visits_schedule_tour": "Website Sales Funnel - Visits (Schedule a Tour)",
    "website_sales_funnel_completed_applications": "Website Sales Funnel - Completed Applications",
    "website_funnel_conversions_click_to_call": "Website Funnel Conversions - Click to Call / Phone",
    "website_funnel_conversions_contact_form": "Website Funnel Conversions - Contact Form",
}

SECTION_ORDER = list(SECTION_TITLES)


def load_rows() -> list[dict]:
    with INPUT_PATH.open() as f:
        return list(csv.DictReader(f))


def grouped_rows(rows: list[dict]) -> dict[str, dict[tuple[str, str], list[dict]]]:
    grouped: dict[str, dict[tuple[str, str], list[dict]]] = defaultdict(lambda: defaultdict(list))
    for row in rows:
        grouped[row["section_key"]][(row["pilot_property_name"], row["sister_property_name"])].append(row)
    for section_key in grouped:
        for pair_key in grouped[section_key]:
            grouped[section_key][pair_key].sort(key=lambda r: r["snapshot_date"])
    return grouped


def write_data_block(data_ws, start_row: int, pair_label: str, rows: list[dict]) -> int:
    dates = [datetime.fromisoformat(r["snapshot_date"]).strftime("%-m/%d") for r in rows]
    pilot_name = rows[0]["pilot_property_name"]
    sister_name = rows[0]["sister_property_name"]

    data_ws.cell(start_row, 1, pair_label)
    data_ws.cell(start_row + 1, 1, f"{pilot_name} Baseline")
    data_ws.cell(start_row + 2, 1, f"{sister_name} Baseline")
    data_ws.cell(start_row + 3, 1, pilot_name)
    data_ws.cell(start_row + 4, 1, sister_name)

    for idx, day in enumerate(dates, start=2):
        row = rows[idx - 2]
        data_ws.cell(start_row, idx, day)
        data_ws.cell(start_row + 1, idx, float(row["pilot_baseline_value"]) if row["pilot_baseline_value"] else None)
        data_ws.cell(start_row + 2, idx, float(row["sister_baseline_value"]) if row["sister_baseline_value"] else None)
        data_ws.cell(start_row + 3, idx, float(row["pilot_daily_value"]) if row["pilot_daily_value"] else None)
        data_ws.cell(start_row + 4, idx, float(row["sister_daily_value"]) if row["sister_daily_value"] else None)

    return start_row + 7


def style_chart(chart: LineChart, data_ws, top_row: int, bottom_row: int, num_dates: int) -> None:
    chart.height = 2.15
    chart.width = 20.5
    chart.style = 2
    chart.legend = None
    chart.title = None
    chart.layout = Layout(
        manualLayout=ManualLayout(
            x=0.02,
            y=0.04,
            w=0.96,
            h=0.80,
            layoutTarget="inner",
        )
    )
    chart.graphical_properties = GraphicalProperties()
    chart.graphical_properties.noFill = True
    chart.graphical_properties.line.noFill = True
    chart.y_axis.delete = True
    chart.x_axis.majorTickMark = "none"
    chart.y_axis.majorTickMark = "none"
    chart.y_axis.spPr = None
    chart.y_axis.majorGridlines = None
    chart.x_axis.delete = False
    chart.x_axis.axPos = "b"
    chart.x_axis.tickLblPos = "low"
    chart.x_axis.spPr = GraphicalProperties()
    chart.x_axis.spPr.line.noFill = True
    chart.x_axis.majorGridlines = ChartLines()
    chart.x_axis.majorGridlines.spPr = GraphicalProperties()
    chart.x_axis.majorGridlines.spPr.line.noFill = True
    chart.plot_area.spPr = None

    max_col = 1 + num_dates
    data = Reference(data_ws, min_col=2, max_col=max_col, min_row=top_row, max_row=bottom_row)
    cats = Reference(data_ws, min_col=2, max_col=max_col, min_row=top_row - 1, max_row=top_row - 1)
    chart.add_data(data, titles_from_data=False, from_rows=True)
    chart.set_categories(cats)

    colors = [PILOT_BASELINE_COLOR, SISTER_BASELINE_COLOR, PILOT_COLOR, SISTER_COLOR]
    widths = [12000, 12000, 23000, 23000]
    dashes = ["sysDash", "sysDash", None, None]

    for idx, series in enumerate(chart.series):
        line = series.graphicalProperties.line
        line.solidFill = colors[idx]
        line.width = widths[idx]
        if dashes[idx]:
            line.prstDash = dashes[idx]
        marker = series.marker
        marker.symbol = "circle"
        marker.size = 3 if idx < 2 else 6
        marker.graphicalProperties.solidFill = colors[idx]
        marker.graphicalProperties.line.solidFill = colors[idx]


def add_pair_chart(display_ws, data_ws, start_row: int, section_title: str, pilot_name: str, sister_name: str, rows: list[dict], data_row: int) -> tuple[int, int]:
    display_ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=12)
    title_cell = display_ws.cell(start_row, 1)
    title_cell.value = CellRichText(
        TextBlock(InlineFont(rFont="Calibri", sz=18, b=True, color=PILOT_COLOR), pilot_name),
        TextBlock(InlineFont(rFont="Calibri", sz=18, b=True, color=TEXT_BLACK), " vs "),
        TextBlock(InlineFont(rFont="Calibri", sz=18, b=True, color=SISTER_COLOR), sister_name),
    )
    title_cell.fill = PatternFill("solid", fgColor=TITLE_FILL)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    display_ws.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=12)
    metric_cell = display_ws.cell(start_row + 1, 1, section_title)
    metric_cell.font = Font(size=10, bold=True, color=TEXT_BLACK)
    metric_cell.alignment = Alignment(horizontal="left", vertical="center")

    for row in range(start_row + 1, start_row + 7):
        for col in range(1, 13):
            display_ws.cell(row, col).fill = PatternFill("solid", fgColor=CARD_FILL)

    data_row = write_data_block(data_ws, data_row, f"{pilot_name} vs {sister_name} {section_title}", rows)
    chart = LineChart()
    style_chart(chart, data_ws, data_row - 6, data_row - 3, len(rows))
    display_ws.add_chart(chart, f"A{start_row + 2}")

    return start_row + 8, data_row


def add_pending_section(display_ws, start_row: int) -> int:
    display_ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=12)
    title = display_ws.cell(start_row, 1, "High Intent User Rate")
    title.font = Font(size=14, bold=True, color=TEXT_BLACK)
    title.fill = PatternFill("solid", fgColor=SECTION_FILL)
    title.alignment = Alignment(horizontal="left")

    display_ws.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 3, end_column=12)
    note = display_ws.cell(start_row + 1, 1, "Pending Heap export")
    note.font = Font(size=12, bold=True, color="9A3412")
    note.fill = PatternFill("solid", fgColor=PENDING_FILL)
    note.alignment = Alignment(horizontal="left", vertical="center")
    return start_row + 5


def build_workbook() -> None:
    rows = load_rows()
    grouped = grouped_rows(rows)

    wb = Workbook()
    display_ws = wb.active
    display_ws.title = "BI Prototype"
    data_ws = wb.create_sheet("Chart Data")
    data_ws.sheet_state = "hidden"

    display_ws.sheet_view.showGridLines = False
    display_ws.merge_cells("A1:L1")
    display_ws["A1"] = "BI Section Prototype"
    display_ws["A1"].font = Font(size=18, bold=True, color=TEXT_BLACK)
    display_ws["A1"].alignment = Alignment(horizontal="left")

    display_ws.merge_cells("A2:L2")
    display_ws["A2"] = "Daily values use Website Conversion T7D. Baselines use the seeded T90D file."
    display_ws["A2"].font = Font(size=10, color=MUTED_TEXT)

    for col in range(1, 13):
        display_ws.column_dimensions[get_column_letter(col)].width = 9

    row = 4
    data_row = 1

    row = add_pending_section(display_ws, row)

    for section_key in SECTION_ORDER:
        display_ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
        section = display_ws.cell(row, 1, SECTION_TITLES[section_key])
        section.font = Font(size=14, bold=True, color=TEXT_BLACK)
        section.fill = PatternFill("solid", fgColor=SECTION_FILL)
        section.alignment = Alignment(horizontal="left")
        row += 2

        for (pilot_name, sister_name), pair_rows in grouped.get(section_key, {}).items():
            row, data_row = add_pair_chart(display_ws, data_ws, row, SECTION_TITLES[section_key], pilot_name, sister_name, pair_rows, data_row)
            row += 1

    wb.save(OUTPUT_PATH)


if __name__ == "__main__":
    build_workbook()
    print(OUTPUT_PATH)
