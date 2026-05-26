from __future__ import annotations

import csv
import json
import sqlite3
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from measurement_dashboard_parser import (
    organic_heap_current_and_baseline,
    organic_heap_series,
    latest_heap_status,
)

ROOT = Path("/Users/mark/Property_Analytics")
DB_PATH = ROOT / "data" / "portfolio_analytics.db"
CONFIG_PATH = ROOT / "pilot_control_cwv" / "config" / "pilot_control_cwv_config.json"
BI_SERIES_PATH = ROOT / "pilot_control_cwv" / "reports" / "pilot_bi_report_series.csv"
OUTPUT_PATH = ROOT / "pilot_control_cwv" / "reports" / f"Pilot_KPI_Combined_Prototype_{datetime.now().strftime('%Y-%m-%d')}.xlsx"

PILOT_COLOR = "4473D0"
SISTER_COLOR = "7CCAC2"
BASELINE_COLOR = "A3A3A3"
FLOOR_COLOR = "F2A7A7"
SISTER_BASELINE_COLOR = "D6D3D1"
CARD_FILL = "F8FAFC"
TITLE_FILL = "EAF2FF"
SECTION_FILL = "EEF4FF"
PENDING_FILL = "FFF7ED"
TEXT_BLACK = "0F172A"
MUTED_TEXT = "64748B"

PSI_BASELINE = 90
PSI_FLOOR = 60
GTM_BASELINE = 94
GTM_FLOOR = 70
CWV_DATES = ["2026-03-26", "2026-03-27", "2026-03-28", "2026-03-29", "2026-03-30", "2026-03-31"]

BI_SECTION_TITLES = {
    "lead_to_available_unit_rate": "Lead (Guest Card) to Available Unit Rate",
    "website_sales_funnel_price_quote": "Website Sales Funnel - Price Quote",
    "website_sales_funnel_visits_schedule_tour": "Website Sales Funnel - Visits (Schedule a Tour)",
    "website_sales_funnel_completed_applications": "Website Sales Funnel - Completed Applications",
    "website_funnel_conversions_click_to_call": "Website Funnel Conversions - Click to Call / Phone",
    "website_funnel_conversions_contact_form": "Website Funnel Conversions - Contact Form",
}
BI_SECTION_ORDER = list(BI_SECTION_TITLES)


def load_pairs() -> list[dict]:
    cfg = json.loads(CONFIG_PATH.read_text())
    cohorts = {c["key"]: c for c in cfg["cohorts"] if c.get("active")}
    pairs = []
    for cohort in cohorts.values():
        if cohort["role"] != "pilot":
            continue
        sister = cohorts[cohort["sister_key"]]
        pairs.append({"pilot": cohort, "sister": sister})
    return pairs


def fetch_cwv_series(conn: sqlite3.Connection, property_id: str, kind: str) -> list[tuple[str, float | None]]:
    if kind == "psi":
        query = """
            SELECT metric_date, performance_score
            FROM pilot_control_psi_metrics
            WHERE property_id = ? AND strategy = 'mobile'
            ORDER BY metric_date DESC
            LIMIT 7
        """
    else:
        query = """
            SELECT metric_date, pagespeed_score
            FROM gtmetrix_metrics
            WHERE property_id = ?
            ORDER BY metric_date DESC
            LIMIT 7
        """
    rows = conn.execute(query, (property_id,)).fetchall()
    by_date = {row[0]: float(row[1]) for row in rows if row[1] is not None}
    return [(metric_date, by_date.get(metric_date)) for metric_date in CWV_DATES]


def load_bi_grouped() -> dict[str, dict[tuple[str, str], list[dict]]]:
    with BI_SERIES_PATH.open() as f:
        rows = list(csv.DictReader(f))
    grouped: dict[str, dict[tuple[str, str], list[dict]]] = defaultdict(lambda: defaultdict(list))
    for row in rows:
        grouped[row["section_key"]][(row["pilot_property_name"], row["sister_property_name"])].append(row)
    for section_key in grouped:
        for pair_key in grouped[section_key]:
            grouped[section_key][pair_key].sort(key=lambda r: r["snapshot_date"])
    return grouped


def style_chart(chart: LineChart, data_ws, top_row: int, bottom_row: int, num_dates: int, colors: list[str], widths: list[int], dashes: list[str | None]) -> None:
    chart.height = 2.15
    chart.width = 15.8
    chart.style = 2
    chart.legend = None
    chart.title = None
    chart.layout = Layout(
        manualLayout=ManualLayout(
            x=0.01,
            y=0.04,
            w=0.97,
            h=0.80,
            layoutTarget="inner",
        )
    )
    chart.graphical_properties = GraphicalProperties()
    chart.graphical_properties.noFill = True
    chart.graphical_properties.line.noFill = True
    chart.y_axis.delete = True
    chart.x_axis.majorTickMark = "none"
    chart.y_axis.majorTickMark = "none"
    chart.y_axis.spPr = None
    chart.y_axis.majorGridlines = None
    chart.x_axis.delete = False
    chart.x_axis.axPos = "b"
    chart.x_axis.tickLblPos = "low"
    chart.x_axis.spPr = GraphicalProperties()
    chart.x_axis.spPr.line.noFill = True
    chart.x_axis.majorGridlines = ChartLines()
    chart.x_axis.majorGridlines.spPr = GraphicalProperties()
    chart.x_axis.majorGridlines.spPr.line.noFill = True
    chart.plot_area.spPr = None

    max_col = 1 + num_dates
    data = Reference(data_ws, min_col=2, max_col=max_col, min_row=top_row, max_row=bottom_row)
    cats = Reference(data_ws, min_col=2, max_col=max_col, min_row=top_row - 1, max_row=top_row - 1)
    chart.add_data(data, titles_from_data=False, from_rows=True)
    chart.set_categories(cats)

    for idx, series in enumerate(chart.series):
        line = series.graphicalProperties.line
        line.solidFill = colors[idx]
        line.width = widths[idx]
        if dashes[idx]:
            line.prstDash = dashes[idx]
        marker = series.marker
        marker.symbol = "circle"
        marker.size = 3 if idx < 2 else 6
        marker.graphicalProperties.solidFill = colors[idx]
        marker.graphicalProperties.line.solidFill = colors[idx]


def write_value_labels(ws, top_row: int, labels: list[tuple[str, str, str]]) -> None:
    for idx, (label, value, color) in enumerate(labels):
        row = top_row + idx
        ws.cell(row, 10, label).font = Font(size=9, bold=True, color=color)
        ws.cell(row, 10).alignment = Alignment(horizontal="left")
        ws.cell(row, 11, value).font = Font(size=9, color=TEXT_BLACK)
        ws.cell(row, 11).alignment = Alignment(horizontal="right")


def fmt_num(value: float | None, is_ratio: bool = False) -> str:
    if value is None or value == "":
        return "n/a"
    return f"{value:.1f}" if not is_ratio else f"{value * 100:.1f}%"


def align_heap_series(
    pilot_series: list[tuple[str, float]],
    sister_series: list[tuple[str, float]],
) -> tuple[list[str], list[float | None], list[float | None]]:
    dates = sorted({d for d, _ in pilot_series} | {d for d, _ in sister_series}, key=lambda s: (int(s.split('/')[0]), int(s.split('/')[1])))
    pilot_map = {d: v for d, v in pilot_series}
    sister_map = {d: v for d, v in sister_series}
    pilot_values = [pilot_map.get(d) for d in dates]
    sister_values = [sister_map.get(d) for d in dates]

    def backfill_leading(values):
        first = next((v for v in values if v is not None), None)
        if first is None:
            return values
        out = values[:]
        for i, val in enumerate(out):
            if val is None:
                out[i] = first
            else:
                break
        return out

    return dates, backfill_leading(pilot_values), backfill_leading(sister_values)


def add_pair_title(ws, row: int, pilot_name: str, sister_name: str) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    cell = ws.cell(row, 1)
    cell.value = CellRichText(
        TextBlock(InlineFont(rFont="Calibri", sz=18, b=True, color=PILOT_COLOR), pilot_name),
        TextBlock(InlineFont(rFont="Calibri", sz=18, b=True, color=TEXT_BLACK), " vs "),
        TextBlock(InlineFont(rFont="Calibri", sz=18, b=True, color=SISTER_COLOR), sister_name),
    )
    cell.fill = PatternFill("solid", fgColor=TITLE_FILL)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def add_metric_label(ws, row: int, text: str, subtitle: str | None = None) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    cell = ws.cell(row, 1, text)
    cell.font = Font(size=10, bold=True, color=TEXT_BLACK)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    if subtitle:
        ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=12)
        sub = ws.cell(row + 1, 1, subtitle)
        sub.font = Font(size=8, color=MUTED_TEXT)
        sub.alignment = Alignment(horizontal="left", vertical="center")


def add_cwv_card(ws, data_ws, start_row: int, pair: dict, conn: sqlite3.Connection, data_row: int) -> tuple[int, int]:
    pilot = pair["pilot"]
    sister = pair["sister"]

    add_pair_title(ws, start_row, pilot["display_name"], sister["display_name"])
    for row in range(start_row + 1, start_row + 15):
        for col in range(1, 13):
            ws.cell(row, col).fill = PatternFill("solid", fgColor=CARD_FILL)

    add_metric_label(ws, start_row + 1, "PSI")
    pilot_psi = fetch_cwv_series(conn, pilot["property_id"], "psi")
    sister_psi = fetch_cwv_series(conn, sister["property_id"], "psi")
    dates = [datetime.fromisoformat(d).strftime("%-m/%d") for d, _ in pilot_psi]
    data_ws.cell(data_row, 1, "PSI")
    for idx, day in enumerate(dates, start=2):
        data_ws.cell(data_row, idx, day)
        data_ws.cell(data_row + 1, idx, PSI_BASELINE)
        data_ws.cell(data_row + 2, idx, PSI_FLOOR)
        data_ws.cell(data_row + 3, idx, pilot_psi[idx - 2][1])
        data_ws.cell(data_row + 4, idx, sister_psi[idx - 2][1])
    psi_chart = LineChart()
    style_chart(
        psi_chart,
        data_ws,
        data_row + 1,
        data_row + 4,
        len(dates),
        [BASELINE_COLOR, FLOOR_COLOR, PILOT_COLOR, SISTER_COLOR],
        [12000, 12000, 23000, 23000],
        ["sysDash", "sysDash", None, None],
    )
    ws.add_chart(psi_chart, f"A{start_row + 2}")
    write_value_labels(
        ws,
        start_row + 3,
        [
            ("Pilot", fmt_num(pilot_psi[-1][1]), PILOT_COLOR),
            ("Sister", fmt_num(sister_psi[-1][1]), SISTER_COLOR),
            ("Baseline", fmt_num(PSI_BASELINE), BASELINE_COLOR),
            ("Floor", fmt_num(PSI_FLOOR), FLOOR_COLOR),
        ],
    )
    data_row += 7

    add_metric_label(ws, start_row + 8, "GTMetrix")
    pilot_gtm = fetch_cwv_series(conn, pilot["property_id"], "gtm")
    sister_gtm = fetch_cwv_series(conn, sister["property_id"], "gtm")
    data_ws.cell(data_row, 1, "GTM")
    for idx, day in enumerate(dates, start=2):
        data_ws.cell(data_row, idx, day)
        data_ws.cell(data_row + 1, idx, GTM_BASELINE)
        data_ws.cell(data_row + 2, idx, GTM_FLOOR)
        data_ws.cell(data_row + 3, idx, pilot_gtm[idx - 2][1])
        data_ws.cell(data_row + 4, idx, sister_gtm[idx - 2][1])
    gtm_chart = LineChart()
    style_chart(
        gtm_chart,
        data_ws,
        data_row + 1,
        data_row + 4,
        len(dates),
        [BASELINE_COLOR, FLOOR_COLOR, PILOT_COLOR, SISTER_COLOR],
        [12000, 12000, 23000, 23000],
        ["sysDash", "sysDash", None, None],
    )
    ws.add_chart(gtm_chart, f"A{start_row + 9}")
    write_value_labels(
        ws,
        start_row + 10,
        [
            ("Pilot", fmt_num(pilot_gtm[-1][1]), PILOT_COLOR),
            ("Sister", fmt_num(sister_gtm[-1][1]), SISTER_COLOR),
            ("Baseline", fmt_num(GTM_BASELINE), BASELINE_COLOR),
            ("Floor", fmt_num(GTM_FLOOR), FLOOR_COLOR),
        ],
    )
    data_row += 7

    return start_row + 16, data_row


def add_bi_pending(ws, row: int) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    title = ws.cell(row, 1, "High Intent User Rate")
    title.font = Font(size=14, bold=True, color=TEXT_BLACK)
    title.fill = PatternFill("solid", fgColor=SECTION_FILL)
    title.alignment = Alignment(horizontal="left")
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 3, end_column=12)
    note = ws.cell(row + 1, 1, "Pending Heap export")
    note.font = Font(size=12, bold=True, color="9A3412")
    note.fill = PatternFill("solid", fgColor=PENDING_FILL)
    note.alignment = Alignment(horizontal="left", vertical="center")
    return row + 5


def add_heap_traffic_card(ws, data_ws, start_row: int, pair: dict, data_row: int) -> tuple[int, int]:
    pilot_name = pair["pilot"]["display_name"]
    sister_name = pair["sister"]["display_name"]
    add_pair_title(ws, start_row, pilot_name, sister_name)
    for row in range(start_row + 1, start_row + 8):
        for col in range(1, 13):
            ws.cell(row, col).fill = PatternFill("solid", fgColor=CARD_FILL)

    status = latest_heap_status()
    subtitle = f"Heap source | Latest available {status.latest_date_label} | Today's Heap drop pending"
    add_metric_label(ws, start_row + 1, "Organic Traffic as a % of Unique Users", subtitle)

    pilot_series = organic_heap_series(pilot_name)[-6:]
    sister_series = organic_heap_series(sister_name)[-6:]
    dates, pilot_values, sister_values = align_heap_series(pilot_series, sister_series)
    pilot_current, pilot_baseline = organic_heap_current_and_baseline(pilot_name)
    sister_current, sister_baseline = organic_heap_current_and_baseline(sister_name)

    data_ws.cell(data_row, 1, "Heap Organic %")
    for idx, day in enumerate(dates, start=2):
        data_ws.cell(data_row, idx, day)
        data_ws.cell(data_row + 1, idx, pilot_baseline)
        data_ws.cell(data_row + 2, idx, sister_baseline)
        data_ws.cell(data_row + 3, idx, pilot_values[idx - 2])
        data_ws.cell(data_row + 4, idx, sister_values[idx - 2])

    chart = LineChart()
    style_chart(
        chart,
        data_ws,
        data_row + 1,
        data_row + 4,
        len(dates),
        [BASELINE_COLOR, SISTER_BASELINE_COLOR, PILOT_COLOR, SISTER_COLOR],
        [12000, 12000, 23000, 23000],
        ["sysDash", "sysDash", None, None],
    )
    ws.add_chart(chart, f"A{start_row + 2}")
    write_value_labels(
        ws,
        start_row + 3,
        [
            ("Pilot", fmt_num(pilot_current, True), PILOT_COLOR),
            ("Sister", fmt_num(sister_current, True), SISTER_COLOR),
            ("Pilot BL", fmt_num(pilot_baseline, True), BASELINE_COLOR),
            ("Sister BL", fmt_num(sister_baseline, True), SISTER_BASELINE_COLOR),
        ],
    )
    return start_row + 9, data_row + 7


def add_bi_card(ws, data_ws, start_row: int, section_title: str, rows: list[dict], data_row: int) -> tuple[int, int]:
    pilot_name = rows[0]["pilot_property_name"]
    sister_name = rows[0]["sister_property_name"]
    add_pair_title(ws, start_row, pilot_name, sister_name)
    for row in range(start_row + 1, start_row + 8):
        for col in range(1, 13):
            ws.cell(row, col).fill = PatternFill("solid", fgColor=CARD_FILL)

    subtitle = f"{rows[0]['daily_window']} daily | Baseline seeded {rows[0]['baseline_source_date'] or 'n/a'}"
    add_metric_label(ws, start_row + 1, section_title, subtitle)

    dates = [datetime.fromisoformat(r["snapshot_date"]).strftime("%-m/%d") for r in rows]
    data_ws.cell(data_row, 1, section_title)
    for idx, day in enumerate(dates, start=2):
        row = rows[idx - 2]
        data_ws.cell(data_row, idx, day)
        data_ws.cell(data_row + 1, idx, float(row["pilot_baseline_value"]) if row["pilot_baseline_value"] else None)
        data_ws.cell(data_row + 2, idx, float(row["sister_baseline_value"]) if row["sister_baseline_value"] else None)
        data_ws.cell(data_row + 3, idx, float(row["pilot_daily_value"]) if row["pilot_daily_value"] else None)
        data_ws.cell(data_row + 4, idx, float(row["sister_daily_value"]) if row["sister_daily_value"] else None)

    chart = LineChart()
    style_chart(
        chart,
        data_ws,
        data_row + 1,
        data_row + 4,
        len(dates),
        [BASELINE_COLOR, SISTER_BASELINE_COLOR, PILOT_COLOR, SISTER_COLOR],
        [12000, 12000, 23000, 23000],
        ["sysDash", "sysDash", None, None],
    )
    ws.add_chart(chart, f"A{start_row + 2}")

    latest = rows[-1]
    write_value_labels(
        ws,
        start_row + 3,
        [
            ("Pilot", fmt_num(float(latest["pilot_daily_value"]) if latest["pilot_daily_value"] else None, True), PILOT_COLOR),
            ("Sister", fmt_num(float(latest["sister_daily_value"]) if latest["sister_daily_value"] else None, True), SISTER_COLOR),
            ("Pilot BL", fmt_num(float(latest["pilot_baseline_value"]) if latest["pilot_baseline_value"] else None, True), BASELINE_COLOR),
            ("Sister BL", fmt_num(float(latest["sister_baseline_value"]) if latest["sister_baseline_value"] else None, True), SISTER_BASELINE_COLOR),
        ],
    )

    return start_row + 9, data_row + 7


def build_workbook() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "KPI Prototype"
    data_ws = wb.create_sheet("Chart Data")
    data_ws.sheet_state = "hidden"

    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:L1")
    ws["A1"] = "Pilot KPI Combined Prototype"
    ws["A1"].font = Font(size=18, bold=True, color=TEXT_BLACK)
    ws["A1"].alignment = Alignment(horizontal="left")
    ws.merge_cells("A2:L2")
    ws["A2"] = "CWV at top, BI sections below. BI values use T7D daily snapshots with seeded T90D baselines."
    ws["A2"].font = Font(size=10, color=MUTED_TEXT)

    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].width = 9

    conn = sqlite3.connect(DB_PATH)
    row = 4
    data_row = 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    sec = ws.cell(row, 1, "Core Web Vitals")
    sec.font = Font(size=14, bold=True, color=TEXT_BLACK)
    sec.fill = PatternFill("solid", fgColor=SECTION_FILL)
    sec.alignment = Alignment(horizontal="left")
    row += 2

    for pair in load_pairs():
        row, data_row = add_cwv_card(ws, data_ws, row, pair, conn, data_row)
        row += 1
    conn.close()

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    sec = ws.cell(row, 1, "Traffic & Engagement")
    sec.font = Font(size=14, bold=True, color=TEXT_BLACK)
    sec.fill = PatternFill("solid", fgColor=SECTION_FILL)
    sec.alignment = Alignment(horizontal="left")
    row += 2

    for pair in load_pairs():
        row, data_row = add_heap_traffic_card(ws, data_ws, row, pair, data_row)
        row += 1

    row = add_bi_pending(ws, row)

    grouped_bi = load_bi_grouped()
    for section_key in BI_SECTION_ORDER:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
        sec = ws.cell(row, 1, BI_SECTION_TITLES[section_key])
        sec.font = Font(size=14, bold=True, color=TEXT_BLACK)
        sec.fill = PatternFill("solid", fgColor=SECTION_FILL)
        sec.alignment = Alignment(horizontal="left")
        row += 2
        for _, pair_rows in grouped_bi.get(section_key, {}).items():
            row, data_row = add_bi_card(ws, data_ws, row, BI_SECTION_TITLES[section_key], pair_rows, data_row)
            row += 1

    wb.save(OUTPUT_PATH)


if __name__ == "__main__":
    build_workbook()
    print(OUTPUT_PATH)
