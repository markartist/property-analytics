#!/usr/bin/env python3
"""
Generate the dedicated pilot/control daily CWV workbook and HTML summary.
"""

from __future__ import annotations

import argparse
import csv
import json
import sqlite3
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parents[1]
REPO_ROOT = BASE_DIR.parent
DEFAULT_CONFIG_PATH = BASE_DIR / "config" / "pilot_control_cwv_config.json"
OUTPUT_DIR = BASE_DIR / "reports"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


@dataclass
class CohortEntry:
    key: str
    display_name: str
    role: str
    property_id: Optional[str]
    site_url: str
    history_source: str
    sister_key: Optional[str]
    active: bool


def load_config(path: Path) -> Dict[str, object]:
    with path.open() as fh:
        config = json.load(fh)
    config.setdefault("report_start_date", config.get("launch_date"))
    config.setdefault("baseline_date", config.get("report_start_date"))
    return config


def display_metric_date(config: Dict[str, object], metric_date: str) -> str:
    launch_date = config.get("launch_date")
    if launch_date and metric_date < launch_date:
        return launch_date
    return metric_date


def active_entries(config: Dict[str, object]) -> List[CohortEntry]:
    rows = []
    for row in config.get("cohorts", []):
        if not row.get("active", True):
            continue
        rows.append(
            CohortEntry(
                key=row["key"],
                display_name=row["display_name"],
                role=row["role"],
                property_id=row.get("property_id"),
                site_url=row["site_url"],
                history_source=row.get("history_source", "dedicated"),
                sister_key=row.get("sister_key"),
                active=bool(row.get("active", True)),
            )
        )
    return rows


def fmt_num(value: Optional[float]) -> str:
    return "" if value is None else f"{value:.1f}"


def fmt_delta(value: Optional[float]) -> str:
    return "" if value is None else f"{value:+.1f}"


def is_delta_key(key: str) -> bool:
    return key.endswith("_vs_today") or key.endswith("_vs_sister_today")


def judgment(score: Optional[float], lcp: Optional[float], cls: Optional[float], tbt: Optional[float]) -> str:
    if score is None:
        return "No Data"
    if score >= 90 and (lcp is None or lcp <= 2.5) and (cls is None or cls <= 0.10) and (tbt is None or tbt <= 200):
        return "Strong"
    poor_flags = sum(
        int(val)
        for val in [
            lcp is not None and lcp > 4.0,
            cls is not None and cls > 0.25,
            tbt is not None and tbt > 600,
        ]
    )
    if score < 60 or poor_flags >= 1:
        return "At Risk"
    if score < 75:
        return "Needs Improvement"
    return "Healthy"


def get_dedicated_history(conn: sqlite3.Connection, entry: CohortEntry, as_of_date: str) -> List[sqlite3.Row]:
    return conn.execute(
        """
        SELECT metric_date, performance_score, lcp_value, cls_value, total_blocking_time
        FROM pilot_control_psi_metrics
        WHERE cohort_key = ?
          AND strategy = 'mobile'
          AND metric_date <= ?
        ORDER BY metric_date DESC
        """,
        (entry.key, as_of_date),
    ).fetchall()


def get_portfolio_history(conn: sqlite3.Connection, entry: CohortEntry, as_of_date: str) -> List[sqlite3.Row]:
    if not entry.property_id:
        return []
    return conn.execute(
        """
        SELECT metric_date, performance_score, lcp_value, cls_value, total_blocking_time
        FROM pagespeed_metrics
        WHERE property_id = ?
          AND strategy = 'mobile'
          AND metric_date <= ?
        ORDER BY metric_date DESC
        """,
        (entry.property_id, as_of_date),
    ).fetchall()


def get_history(conn: sqlite3.Connection, entry: CohortEntry, as_of_date: str) -> List[sqlite3.Row]:
    if entry.history_source == "portfolio_property":
        return get_portfolio_history(conn, entry, as_of_date)
    return get_dedicated_history(conn, entry, as_of_date)


def get_trend_history(conn: sqlite3.Connection, entry: CohortEntry, as_of_date: str) -> List[sqlite3.Row]:
    # Pilot vanity domains use their legacy property PSI history for trailing windows/baselines.
    if entry.role == "pilot" and entry.property_id:
        return get_portfolio_history(conn, entry, as_of_date)
    return get_history(conn, entry, as_of_date)


def prior_scores(history: List[sqlite3.Row], as_of_date: str) -> List[float]:
    return [
        row["performance_score"]
        for row in history
        if row["metric_date"] < as_of_date and row["performance_score"] is not None
    ]


def average_window(scores: List[float], start: int, length: int) -> Optional[float]:
    window = scores[start : start + length]
    if len(window) < length:
        return None
    return sum(window) / len(window)


def static_baseline(
    history: List[sqlite3.Row],
    baseline_date: str,
    days: int = 90,
    minimum_days: int = 30,
) -> Optional[float]:
    scores = prior_scores(history, baseline_date)
    if len(scores) < minimum_days:
        return None
    window = scores[:days]
    return sum(window) / len(window)


def yoy_score(history: List[sqlite3.Row], as_of_date: str) -> Optional[float]:
    target = (datetime.strptime(as_of_date, "%Y-%m-%d").date() - timedelta(days=365)).isoformat()
    for row in history:
        if row["metric_date"] == target:
            return row["performance_score"]
    return None


def current_mobile_row(conn: sqlite3.Connection, entry: CohortEntry, as_of_date: str) -> Optional[sqlite3.Row]:
    return conn.execute(
        """
        SELECT metric_date, performance_score, lcp_value, cls_value, total_blocking_time
        FROM pilot_control_psi_metrics
        WHERE cohort_key = ?
          AND strategy = 'mobile'
          AND metric_date = ?
        """,
        (entry.key, as_of_date),
    ).fetchone()


def all_dates(conn: sqlite3.Connection, report_start_date: str) -> List[str]:
    rows = conn.execute(
        """
        SELECT DISTINCT metric_date
        FROM pilot_control_psi_metrics
        WHERE metric_date >= ?
        ORDER BY metric_date ASC
        """,
        (report_start_date,),
    ).fetchall()
    return [row[0] for row in rows]


def complete_dates(conn: sqlite3.Connection, entries: List[CohortEntry], report_start_date: str) -> List[str]:
    expected_count = len(entries)
    rows = conn.execute(
        """
        SELECT metric_date
        FROM pilot_control_psi_metrics
        WHERE metric_date >= ?
          AND strategy = 'mobile'
        GROUP BY metric_date
        HAVING COUNT(DISTINCT cohort_key) >= ?
        ORDER BY metric_date ASC
        """,
        (report_start_date, expected_count),
    ).fetchall()
    return [row[0] for row in rows]


def raw_rows(conn: sqlite3.Connection, report_dates: List[str]) -> List[sqlite3.Row]:
    if not report_dates:
        return []
    placeholders = ",".join("?" for _ in report_dates)
    return conn.execute(
        """
        SELECT
            metric_date,
            cohort_key,
            display_name,
            role,
            site_url,
            strategy,
            performance_score,
            lcp_value,
            cls_value,
            total_blocking_time
        FROM pilot_control_psi_metrics
        WHERE metric_date IN ({placeholders})
        ORDER BY metric_date ASC, display_name ASC, strategy ASC
        """.format(placeholders=placeholders),
        report_dates,
    ).fetchall()


def build_dataset(
    conn: sqlite3.Connection,
    entries: List[CohortEntry],
    baseline_date: str,
    report_dates: List[str],
) -> Dict[str, Dict[str, Dict[str, Optional[float]]]]:
    index = {entry.key: entry for entry in entries}
    static_baselines: Dict[str, Optional[float]] = {}
    for entry in entries:
        baseline_history = get_trend_history(conn, entry, baseline_date)
        static_baselines[entry.key] = static_baseline(baseline_history, baseline_date)

    dataset: Dict[str, Dict[str, Dict[str, Optional[float]]]] = {}
    for metric_date in report_dates:
        dataset[metric_date] = {}
        for entry in entries:
            current = current_mobile_row(conn, entry, metric_date)
            history = get_trend_history(conn, entry, metric_date)
            score = current["performance_score"] if current else None
            scores = prior_scores(history, metric_date)
            t7 = average_window(scores, 0, 7)
            t30 = average_window(scores, 0, 30)
            yoy = yoy_score(history, metric_date)
            sister_score = None
            if entry.sister_key and entry.sister_key in index:
                sister = current_mobile_row(conn, index[entry.sister_key], metric_date)
                sister_score = sister["performance_score"] if sister else None
            baseline = static_baselines.get(entry.key)
            dataset[metric_date][entry.key] = {
                "score": score,
                "baseline_t90_static": baseline,
                "baseline_vs_today": (score - baseline) if score is not None and baseline is not None else None,
                "t7": t7,
                "t7_vs_today": (score - t7) if score is not None and t7 is not None else None,
                "t30": t30,
                "t30_vs_today": (score - t30) if score is not None and t30 is not None else None,
                "yoy": yoy,
                "variance_yoy": (score - yoy) if score is not None and yoy is not None else None,
                "today_vs_sister_today": (score - sister_score) if score is not None and sister_score is not None else None,
                "judgment": judgment(
                    score,
                    current["lcp_value"] if current else None,
                    current["cls_value"] if current else None,
                    current["total_blocking_time"] if current else None,
                ),
            }
    return dataset


def pilot_control_pairs(entries: List[CohortEntry]) -> List[tuple[CohortEntry, Optional[CohortEntry]]]:
    pilots = [entry for entry in entries if entry.role == "pilot"]
    entry_index = {entry.key: entry for entry in entries}
    return [(pilot, entry_index.get(pilot.sister_key) if pilot.sister_key else None) for pilot in pilots]


def matrix_columns() -> tuple[list[tuple[str, str]], list[tuple[str, str]]]:
    pilot_columns = [
        ("Today", "score"),
        ("Baseline (T90 static)", "baseline_t90_static"),
        ("T7", "t7"),
        ("T30", "t30"),
    ]
    control_columns = [
        ("Today", "score"),
        ("Baseline (T90 static)", "baseline_t90_static"),
        ("T7", "t7"),
        ("T30", "t30"),
    ]
    return pilot_columns, control_columns


def exact_paste_columns() -> tuple[list[tuple[str, Optional[str]]], list[tuple[str, Optional[str]]]]:
    pilot_columns = [
        ("Today", "score"),
        ("Baseline (T90 static)", "baseline_t90_static"),
        ("", None),
        ("T7", "t7"),
        ("", None),
        ("T30", "t30"),
        ("", None),
        ("", None),
    ]
    control_columns = [
        ("Today", "score"),
        ("Baseline (T90 static)", "baseline_t90_static"),
        ("", None),
        ("T7", "t7"),
        ("", None),
        ("T30", "t30"),
        ("", None),
    ]
    return pilot_columns, control_columns


def matrix_rows(
    config: Dict[str, object],
    entries: List[CohortEntry],
    dataset: Dict[str, Dict[str, Dict[str, Optional[float]]]],
) -> List[List[str]]:
    pilot_columns, control_columns = matrix_columns()
    pairs = pilot_control_pairs(entries)

    header_row_1 = [""]
    header_row_2 = [""]
    practice_row = ["PRACTICE ROW"]

    for pilot, sister in pairs:
        header_row_1.extend([pilot.display_name] + [""] * (len(pilot_columns) - 1))
        header_row_2.extend([label for label, _ in pilot_columns])
        practice_row.extend([""] * len(pilot_columns))
        if sister:
            header_row_1.extend([sister.display_name] + [""] * (len(control_columns) - 1))
            header_row_2.extend([label for label, _ in control_columns])
            practice_row.extend([""] * len(control_columns))

    rows = [header_row_1, header_row_2, practice_row]
    for metric_date in sorted(dataset.keys()):
        row = [display_metric_date(config, metric_date)]
        for pilot, sister in pairs:
            pilot_values = dataset[metric_date][pilot.key]
            for _, key in pilot_columns:
                value = pilot_values[key]
                row.append(fmt_delta(value) if is_delta_key(key) else fmt_num(value))
            if sister:
                sister_values = dataset[metric_date][sister.key]
                for _, key in control_columns:
                    value = sister_values[key]
                    row.append(fmt_delta(value) if is_delta_key(key) else fmt_num(value))
        rows.append(row)
    return rows


def sheet_title_for_pilot(pilot: CohortEntry) -> str:
    custom = {
        "pilot_district_universal": "The District",
    }
    return custom.get(pilot.key, pilot.display_name)[:31]


def pair_sheet_rows(
    config: Dict[str, object],
    pilot: CohortEntry,
    sister: Optional[CohortEntry],
    dataset: Dict[str, Dict[str, Dict[str, Optional[float]]]],
) -> List[List[str]]:
    pilot_columns, control_columns = matrix_columns()
    rows: List[List[str]] = []
    total_cols = 1 + len(pilot_columns) + (len(control_columns) if sister else 0)

    rows.append(["", "Core Web Vitals SCORE"] + [""] * (total_cols - 2))
    rows.append(
        [""]
        + [pilot.display_name]
        + [""] * (len(pilot_columns) - 1)
        + (([sister.display_name] + [""] * (len(control_columns) - 1)) if sister else [])
    )
    rows.append(
        [""]
        + [label for label, _ in pilot_columns]
        + ([label for label, _ in control_columns] if sister else [])
    )

    for metric_date in sorted(dataset.keys()):
        row = [display_metric_date(config, metric_date)]
        pilot_values = dataset[metric_date][pilot.key]
        for _, key in pilot_columns:
            value = pilot_values[key]
            row.append(fmt_delta(value) if is_delta_key(key) else fmt_num(value))
        if sister:
            sister_values = dataset[metric_date][sister.key]
            for _, key in control_columns:
                value = sister_values[key]
                row.append(fmt_delta(value) if is_delta_key(key) else fmt_num(value))
        rows.append(row)
    return rows


def pair_paste_headers(sister: Optional[CohortEntry]) -> List[str]:
    pilot_columns, control_columns = exact_paste_columns()
    headers = ["Date"] + [label for label, _ in pilot_columns]
    if sister:
        headers.extend(label for label, _ in control_columns)
    return headers


def pair_paste_rows(
    config: Dict[str, object],
    pilot: CohortEntry,
    sister: Optional[CohortEntry],
    dataset: Dict[str, Dict[str, Dict[str, Optional[float]]]],
) -> List[List[str]]:
    pilot_columns, control_columns = exact_paste_columns()
    rows: List[List[str]] = [pair_paste_headers(sister)]
    for metric_date in sorted(dataset.keys()):
        row = [display_metric_date(config, metric_date)]
        pilot_values = dataset[metric_date][pilot.key]
        for _, key in pilot_columns:
            if key is None:
                row.append("")
                continue
            value = pilot_values[key]
            row.append(fmt_delta(value) if is_delta_key(key) else fmt_num(value))
        if sister:
            sister_values = dataset[metric_date][sister.key]
            for _, key in control_columns:
                if key is None:
                    row.append("")
                    continue
                value = sister_values[key]
                row.append(fmt_delta(value) if is_delta_key(key) else fmt_num(value))
        rows.append(row)
    return rows


def copy_paste_headers() -> List[str]:
    return [
        "Pilot Property",
        "Sister Property",
        "Date",
        "Pilot Today",
        "Pilot Baseline (T90 static)",
        "Pilot T7",
        "Pilot T30",
        "Sister Today",
        "Sister Baseline (T90 static)",
        "Sister T7",
        "Sister T30",
    ]


def copy_paste_rows(
    entries: List[CohortEntry],
    dataset: Dict[str, Dict[str, Dict[str, Optional[float]]]],
) -> List[List[str]]:
    rows = [copy_paste_headers()]
    if not dataset:
        return rows

    latest_date = max(dataset.keys())
    for pilot, sister in pilot_control_pairs(entries):
        pilot_values = dataset[latest_date][pilot.key]
        sister_values = dataset[latest_date][sister.key] if sister else {}
        rows.append(
            [
                pilot.display_name,
                sister.display_name if sister else "",
                latest_date,
                fmt_num(pilot_values.get("score")),
                fmt_num(pilot_values.get("baseline_t90_static")),
                fmt_num(pilot_values.get("t7")),
                fmt_num(pilot_values.get("t30")),
                fmt_num(sister_values.get("score")),
                fmt_num(sister_values.get("baseline_t90_static")),
                fmt_num(sister_values.get("t7")),
                fmt_num(sister_values.get("t30")),
            ]
        )
    return rows


def milestone_snapshot(
    dataset: Dict[str, Dict[str, Dict[str, Optional[float]]]],
    entry_key: str,
    offset_days: int,
    report_start_date: str,
) -> Optional[Dict[str, Optional[float]]]:
    target_date = (datetime.strptime(report_start_date, "%Y-%m-%d").date() + timedelta(days=offset_days)).isoformat()
    return dataset.get(target_date, {}).get(entry_key)


def relative_snapshot(
    dataset: Dict[str, Dict[str, Dict[str, Optional[float]]]],
    entry_key: str,
    anchor_date: str,
    days_back: int,
) -> Optional[Dict[str, Optional[float]]]:
    target_date = (datetime.strptime(anchor_date, "%Y-%m-%d").date() - timedelta(days=days_back)).isoformat()
    return dataset.get(target_date, {}).get(entry_key)


def build_workbook(
    config: Dict[str, object],
    entries: List[CohortEntry],
    dataset: Dict[str, Dict[str, Dict[str, Optional[float]]]],
    raw_metric_rows: List[sqlite3.Row],
) -> Workbook:
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="D9E2F3")
    section_fill = PatternFill("solid", fgColor="DDEBF7")
    green_fill = PatternFill("solid", fgColor="E2F0D9")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_gray = Side(style="thin", color="D9DCE3")
    dark_gray = Side(style="thin", color="333333")
    light_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    header_border = Border(left=dark_gray, right=dark_gray, top=dark_gray, bottom=dark_gray)
    pilot_columns, control_columns = matrix_columns()
    pairs = pilot_control_pairs(entries)
    latest_report_date = max(dataset.keys()) if dataset else config["report_start_date"]

    overview = wb.active
    overview.title = "Overview"
    overview.sheet_view.showGridLines = False
    overview.freeze_panes = "B2"

    overview_headers = [
        "Metric",
        "Organic Traffic",
        "High Intent User Rate (HEAP CTR)",
        "Website-Sourced Lead Volume (Guest Cards)",
        "Lead (Guest Card) to Available Unit Rate",
        "Website Sales Funnel Conversions",
        "CWV - Mobile Pass Rate",
    ]
    overview_widths = [28, 18, 20, 20, 20, 18, 18]
    for idx, width in enumerate(overview_widths, start=1):
        overview.column_dimensions[get_column_letter(idx)].width = width
        cell = overview.cell(row=1, column=idx, value=overview_headers[idx - 1])
        cell.font = Font(bold=(idx == 1), size=11 if idx != 1 else 16)
        cell.alignment = center
        cell.border = header_border
        cell.fill = green_fill if idx in (4, 5, 6) else white_fill

    overview_row = 2
    for pilot, sister in pairs:
        title = overview.cell(row=overview_row, column=1, value=pilot.display_name)
        title.font = Font(bold=True, size=12)
        title.border = header_border
        for col in range(2, len(overview_headers) + 1):
            overview.cell(row=overview_row, column=col).border = header_border
        overview_row += 1

        for section_label, days_back in [("Today", 0), ("T7", 7), ("T15", 15), ("T21", 21), ("T30", 30)]:
            section = overview.cell(row=overview_row, column=1, value=section_label)
            section.font = Font(bold=True, size=11)
            section.fill = section_fill
            section.border = light_border
            for col in range(2, len(overview_headers) + 1):
                cell = overview.cell(row=overview_row, column=col)
                cell.fill = section_fill
                cell.border = light_border
            overview_row += 1

            snapshot = relative_snapshot(dataset, pilot.key, latest_report_date, days_back)
            sister_snapshot = relative_snapshot(dataset, sister.key, latest_report_date, days_back) if sister else None
            metric_rows = [
                ("Baseline (T90 static)", snapshot["baseline_t90_static"] if snapshot else None),
                ("Variance", snapshot["baseline_vs_today"] if snapshot else None),
                ("T7", snapshot["t7"] if snapshot else None),
                ("Variance", snapshot["t7_vs_today"] if snapshot else None),
                ("T30", snapshot["t30"] if snapshot else None),
                ("Variance", snapshot["t30_vs_today"] if snapshot else None),
                ("Sister Property", sister_snapshot["score"] if sister_snapshot else None),
                ("Variance", snapshot["today_vs_sister_today"] if snapshot else None),
            ]
            if section_label == "Today":
                metric_rows.insert(0, ("Today", snapshot["score"] if snapshot else None))

            for metric_label, metric_value in metric_rows:
                label_cell = overview.cell(row=overview_row, column=1, value=metric_label)
                label_cell.border = light_border
                label_cell.alignment = Alignment(horizontal="left", vertical="center")
                for col in range(2, len(overview_headers)):
                    overview.cell(row=overview_row, column=col).border = light_border
                value_cell = overview.cell(
                    row=overview_row,
                    column=len(overview_headers),
                    value=fmt_delta(metric_value) if metric_label == "Variance" and metric_value is not None else fmt_num(metric_value),
                )
                value_cell.border = light_border
                value_cell.alignment = center
                overview_row += 1

        overview_row += 1

    copy_sheet = wb.create_sheet("Copy_Paste")
    for row_idx, row in enumerate(copy_paste_rows(entries, dataset), start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = copy_sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.border = header_border if row_idx == 1 else light_border
            cell.alignment = center if row_idx == 1 else Alignment(horizontal="left", vertical="center")
            if row_idx == 1:
                cell.font = Font(bold=True)
                cell.fill = header_fill
            copy_sheet.column_dimensions[get_column_letter(col_idx)].width = 18

    for pilot, sister in pairs:
        ws = wb.create_sheet(sheet_title_for_pilot(pilot))
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "B4"
        ws.column_dimensions["A"].width = 16
        ws.row_dimensions[1].height = 36
        ws.row_dimensions[2].height = 40
        ws.row_dimensions[3].height = 42

        total_cols = 1 + len(pilot_columns) + (len(control_columns) if sister else 0)
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=total_cols)
        top = ws.cell(row=1, column=2, value="Core Web Vitals SCORE")
        top.font = Font(size=12)
        top.alignment = center
        top.fill = header_fill
        top.border = header_border

        ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=1 + len(pilot_columns))
        pilot_header = ws.cell(row=2, column=2, value=pilot.display_name)
        pilot_header.font = Font(size=12)
        pilot_header.alignment = center
        pilot_header.fill = header_fill
        pilot_header.border = header_border

        if sister:
            start = 2 + len(pilot_columns)
            end = start + len(control_columns) - 1
            ws.merge_cells(start_row=2, start_column=start, end_row=2, end_column=end)
            sister_header = ws.cell(row=2, column=start, value=sister.display_name)
            sister_header.font = Font(size=12)
            sister_header.alignment = center
            sister_header.fill = header_fill
            sister_header.border = header_border

        ws.cell(row=3, column=1, value="")
        col = 2
        for label, _ in pilot_columns:
            cell = ws.cell(row=3, column=col, value=label)
            cell.font = Font(size=11)
            cell.fill = white_fill
            cell.alignment = center
            cell.border = light_border
            ws.column_dimensions[get_column_letter(col)].width = 13
            col += 1
        if sister:
            for label, _ in control_columns:
                cell = ws.cell(row=3, column=col, value=label)
                cell.font = Font(size=11)
                cell.fill = white_fill
                cell.alignment = center
                cell.border = light_border
                ws.column_dimensions[get_column_letter(col)].width = 13
                col += 1

        row = 4
        for metric_date in sorted(dataset.keys()):
            day_label = ws.cell(row=row, column=1, value=metric_date)
            day_label.font = Font(size=11)
            day_label.alignment = center
            day_label.border = light_border

            col = 2
            values = dataset[metric_date][pilot.key]
            for _, key in pilot_columns:
                value = values[key]
                cell = ws.cell(row=row, column=col, value=fmt_delta(value) if is_delta_key(key) else fmt_num(value))
                cell.alignment = center
                cell.border = light_border
                col += 1
            if sister:
                values = dataset[metric_date][sister.key]
                for _, key in control_columns:
                    value = values[key]
                    cell = ws.cell(row=row, column=col, value=fmt_delta(value) if is_delta_key(key) else fmt_num(value))
                    cell.alignment = center
                    cell.border = light_border
                    col += 1
            row += 1

    raw = wb.create_sheet("Raw PSI")
    raw.append(
        [
            "metric_date",
            "cohort_key",
            "display_name",
            "role",
            "site_url",
            "strategy",
            "performance_score",
            "lcp_value",
            "cls_value",
            "total_blocking_time",
        ]
    )
    for row_data in raw_metric_rows:
        raw.append(
            [
                row_data["metric_date"],
                row_data["cohort_key"],
                row_data["display_name"],
                row_data["role"],
                row_data["site_url"],
                row_data["strategy"],
                row_data["performance_score"],
                row_data["lcp_value"],
                row_data["cls_value"],
                row_data["total_blocking_time"],
            ]
        )
    for column in range(1, 11):
        raw.column_dimensions[get_column_letter(column)].width = 20

    notes = wb.create_sheet("Notes")
    notes["A1"] = "Methodology"
    notes["A1"].font = Font(bold=True)
    notes["A3"] = "Score"
    notes["B3"] = "Actual PSI Mobile Performance Score collected for the configured site URL."
    notes["A4"] = "Baseline (T90 static)"
    notes["B4"] = "Fixed 90-day trailing average calculated once using history before the configured baseline date."
    notes["A5"] = "T7 / T30"
    notes["B5"] = "Trailing averages excluding the current day. New pilot vanity domains stay blank until enough direct history exists."
    notes["A6"] = "Variance Columns"
    notes["B6"] = "All variance columns are shown as Today minus the comparison value."
    notes["A7"] = "Today Vs. Sister Today"
    notes["B7"] = "Pilot current score minus assigned sister/control property's current score on the same date."
    notes["A8"] = "Date Inclusion"
    notes["B8"] = "Only dates with a complete required mobile cohort are included in the commissioned exports."

    return wb


def build_exact_paste_workbook(
    config: Dict[str, object],
    entries: List[CohortEntry],
    dataset: Dict[str, Dict[str, Dict[str, Optional[float]]]],
) -> Workbook:
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="D9E2F3")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_gray = Side(style="thin", color="D9DCE3")
    dark_gray = Side(style="thin", color="333333")
    light_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    header_border = Border(left=dark_gray, right=dark_gray, top=dark_gray, bottom=dark_gray)
    pilot_columns, control_columns = exact_paste_columns()

    first = True
    for pilot, sister in pilot_control_pairs(entries):
        ws = wb.active if first else wb.create_sheet()
        first = False
        ws.title = sheet_title_for_pilot(pilot)
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "B4"
        ws.column_dimensions["A"].width = 16
        ws.row_dimensions[1].height = 36
        ws.row_dimensions[2].height = 40
        ws.row_dimensions[3].height = 42

        total_cols = 1 + len(pilot_columns) + (len(control_columns) if sister else 0)
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=total_cols)
        top = ws.cell(row=1, column=2, value="Core Web Vitals SCORE")
        top.alignment = center
        top.fill = header_fill
        top.border = header_border
        top.font = Font(size=12)

        ws.cell(row=1, column=1, value=pilot.display_name).font = Font(size=16, bold=True)

        ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=1 + len(pilot_columns))
        pilot_header = ws.cell(row=2, column=2, value=pilot.display_name)
        pilot_header.alignment = center
        pilot_header.fill = header_fill
        pilot_header.border = header_border
        pilot_header.font = Font(size=12)

        if sister:
            start = 2 + len(pilot_columns)
            end = start + len(control_columns) - 1
            ws.merge_cells(start_row=2, start_column=start, end_row=2, end_column=end)
            sister_header = ws.cell(row=2, column=start, value=sister.display_name)
            sister_header.alignment = center
            sister_header.fill = header_fill
            sister_header.border = header_border
            sister_header.font = Font(size=12)

        ws.cell(row=3, column=1, value="")
        col = 2
        for label, _ in pilot_columns:
            cell = ws.cell(row=3, column=col, value=label)
            cell.alignment = center
            cell.fill = white_fill
            cell.border = light_border
            cell.font = Font(size=11)
            ws.column_dimensions[get_column_letter(col)].width = 13
            col += 1
        if sister:
            for label, _ in control_columns:
                cell = ws.cell(row=3, column=col, value=label)
                cell.alignment = center
                cell.fill = white_fill
                cell.border = light_border
                cell.font = Font(size=11)
                ws.column_dimensions[get_column_letter(col)].width = 13
                col += 1

        row_num = 4
        ws.cell(row=row_num, column=1, value="PRACTICE ROW").alignment = center
        ws.cell(row=row_num, column=1).border = light_border
        for col_num in range(2, total_cols + 1):
            ws.cell(row=row_num, column=col_num).border = light_border
        row_num += 1

        for metric_date in sorted(dataset.keys()):
            row = [display_metric_date(config, metric_date)]
            pilot_values = dataset[metric_date][pilot.key]
            for _, key in pilot_columns:
                if key is None:
                    row.append("")
                    continue
                value = pilot_values[key]
                row.append(fmt_delta(value) if is_delta_key(key) else fmt_num(value))
            if sister:
                sister_values = dataset[metric_date][sister.key]
                for _, key in control_columns:
                    if key is None:
                        row.append("")
                        continue
                    value = sister_values[key]
                    row.append(fmt_delta(value) if is_delta_key(key) else fmt_num(value))

            for col_num, value in enumerate(row, start=1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = center
                cell.border = light_border
            row_num += 1

    return wb


def write_matrix_csv(
    config: Dict[str, object],
    entries: List[CohortEntry],
    dataset: Dict[str, Dict[str, Dict[str, Optional[float]]]],
    out_path: Path,
) -> None:
    rows = matrix_rows(config, entries, dataset)
    with out_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerows(rows)


def write_sheet_csvs(
    config: Dict[str, object],
    entries: List[CohortEntry],
    dataset: Dict[str, Dict[str, Dict[str, Optional[float]]]],
    out_dir: Path,
) -> None:
    pairs = pilot_control_pairs(entries)
    latest_report_date = max(dataset.keys()) if dataset else config["report_start_date"]

    overview_path = out_dir / "Pilot_Control_CWV_Overview.csv"
    with overview_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(["Pilot Property", "Section", "Metric Row", "CWV - Mobile Pass Rate"])
        for pilot, sister in pairs:
            for section_label, days_back in [("Today", 0), ("T7", 7), ("T15", 15), ("T21", 21), ("T30", 30)]:
                snapshot = relative_snapshot(dataset, pilot.key, latest_report_date, days_back)
                sister_snapshot = relative_snapshot(dataset, sister.key, latest_report_date, days_back) if sister else None
                metric_rows = [
                    ("Baseline (T90 static)", snapshot["baseline_t90_static"] if snapshot else None),
                    ("Variance", snapshot["baseline_vs_today"] if snapshot else None),
                    ("T7", snapshot["t7"] if snapshot else None),
                    ("Variance", snapshot["t7_vs_today"] if snapshot else None),
                    ("T30", snapshot["t30"] if snapshot else None),
                    ("Variance", snapshot["t30_vs_today"] if snapshot else None),
                    ("Sister Property", sister_snapshot["score"] if sister_snapshot else None),
                    ("Variance", snapshot["today_vs_sister_today"] if snapshot else None),
                ]
                if section_label == "Today":
                    metric_rows.insert(0, ("Today", snapshot["score"] if snapshot else None))
                for metric_label, metric_value in metric_rows:
                    value = fmt_delta(metric_value) if metric_label == "Variance" and metric_value is not None else fmt_num(metric_value)
                    writer.writerow([pilot.display_name, section_label, metric_label, value])

    for pilot, sister in pairs:
        path = out_dir / f"{sheet_title_for_pilot(pilot).replace(' ', '_')}.csv"
        with path.open("w", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh)
            writer.writerows(pair_sheet_rows(config, pilot, sister, dataset))

        paste_path = out_dir / f"{sheet_title_for_pilot(pilot).replace(' ', '_')}_Paste_Lines.tsv"
        with paste_path.open("w", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh, delimiter="\t")
            writer.writerows(pair_paste_rows(config, pilot, sister, dataset))

    copy_paste_path = out_dir / "Pilot_Control_CWV_Copy_Paste.csv"
    with copy_paste_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerows(copy_paste_rows(entries, dataset))

    line_dump_path = out_dir / "Pilot_Control_CWV_Paste_Lines.txt"
    with line_dump_path.open("w", encoding="utf-8") as fh:
        for pilot, sister in pairs:
            title = sheet_title_for_pilot(pilot)
            fh.write(f"{title}\n")
            for row in pair_paste_rows(config, pilot, sister, dataset):
                fh.write("\t".join(row) + "\n")
            fh.write("\n")


def write_html_summary(config: Dict[str, object], entries: List[CohortEntry], dataset: Dict[str, Dict[str, Dict[str, Optional[float]]]], out_path: Path) -> None:
    latest_date = max(dataset.keys()) if dataset else None
    rows = []
    if latest_date:
        for entry in entries:
            vals = dataset[latest_date][entry.key]
            rows.append(
                f"""
                <tr>
                  <td>{entry.display_name}</td>
                  <td>{entry.role.title()}</td>
                  <td>{fmt_num(vals['score']) or '—'}</td>
                  <td>{fmt_num(vals['baseline_t90_static']) or '—'}</td>
                  <td>{fmt_delta(vals['baseline_vs_today']) or '—'}</td>
                  <td>{fmt_num(vals['t7']) or '—'}</td>
                  <td>{fmt_delta(vals['today_vs_sister_today']) or '—'}</td>
                </tr>
                """
            )

    html = f"""
    <html>
      <body style="font-family: Arial, sans-serif; color: #0f172a;">
        <h1>{config['report_name']}</h1>
        <p>Latest collection date: {latest_date or 'No data'}</p>
        <p>This dedicated report is isolated from portfolio PSI history. Pilots use dedicated vanity-domain history only; sister/control properties can be configured to compare against existing portfolio PSI history.</p>
        <table border="1" cellpadding="8" cellspacing="0" style="border-collapse: collapse;">
          <tr style="background: #d9e2f3; font-weight: 700;">
            <th>Property</th>
            <th>Role</th>
            <th>Today</th>
            <th>Baseline (T90 static)</th>
            <th>Baseline vs. Today</th>
            <th>T7</th>
            <th>Today Vs. Sister Today</th>
          </tr>
          {''.join(rows)}
        </table>
      </body>
    </html>
    """
    out_path.write_text(html, encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate the pilot/control CWV report")
    parser.add_argument("--config", default=str(DEFAULT_CONFIG_PATH), help="Path to config JSON")
    parser.add_argument("--date", default=None, help="Optional label date for output filenames")
    args = parser.parse_args()

    config = load_config(Path(args.config))
    entries = active_entries(config)
    db_path = Path(config["db_path"])
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    report_dates = complete_dates(conn, entries, config["report_start_date"])
    dataset = build_dataset(conn, entries, config["baseline_date"], report_dates)
    raw_metric_rows = raw_rows(conn, report_dates)
    conn.close()

    label_date = args.date or (max(dataset.keys()) if dataset else date.today().isoformat())
    workbook_path = OUTPUT_DIR / f"Pilot_Control_CWV_Report_{label_date}.xlsx"
    exact_paste_workbook_path = OUTPUT_DIR / f"Pilot_Control_CWV_Paste_Exact_{label_date}.xlsx"
    html_path = OUTPUT_DIR / f"Pilot_Control_CWV_Report_{label_date}.html"
    csv_path = OUTPUT_DIR / f"Pilot_Control_CWV_Report_{label_date}.csv"

    wb = build_workbook(config, entries, dataset, raw_metric_rows)
    wb.save(workbook_path)
    exact_wb = build_exact_paste_workbook(config, entries, dataset)
    exact_wb.save(exact_paste_workbook_path)
    write_html_summary(config, entries, dataset, html_path)
    write_matrix_csv(config, entries, dataset, csv_path)
    write_sheet_csvs(config, entries, dataset, OUTPUT_DIR)

    print(f"Saved workbook: {workbook_path}")
    print(f"Saved paste WB: {exact_paste_workbook_path}")
    print(f"Saved HTML:     {html_path}")
    print(f"Saved CSV:      {csv_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
