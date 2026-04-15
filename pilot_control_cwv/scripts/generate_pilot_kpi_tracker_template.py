#!/usr/bin/env python3
"""
Generate the commissioned pilot KPI tracker workbook structure.

This workbook is intentionally scaffold-first:
- tabs and Excel Tables are created now
- formulas are wired where practical
- placeholder rows remain for sources that are not fully available yet
"""

from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Iterable, List, Optional

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


BASE_DIR = Path(__file__).resolve().parents[1]
DEFAULT_CONFIG_PATH = BASE_DIR / "config" / "pilot_control_cwv_config.json"
OUTPUT_DIR = BASE_DIR / "reports"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


@dataclass
class CohortEntry:
    key: str
    display_name: str
    role: str
    property_id: Optional[str]
    site_url: str
    sister_key: Optional[str]
    active: bool


def load_config(path: Path) -> dict:
    with path.open() as fh:
        return json.load(fh)


def active_entries(config: dict) -> List[CohortEntry]:
    rows: List[CohortEntry] = []
    for row in config.get("cohorts", []):
        if not row.get("active", True):
            continue
        rows.append(
            CohortEntry(
                key=row["key"],
                display_name=row["display_name"],
                role=row["role"],
                property_id=row.get("property_id"),
                site_url=row["site_url"],
                sister_key=row.get("sister_key"),
                active=bool(row.get("active", True)),
            )
        )
    return rows


def pair_map(entries: Iterable[CohortEntry]) -> dict[str, Optional[str]]:
    return {entry.key: entry.sister_key for entry in entries}


def sibling_lookup(entries: Iterable[CohortEntry]) -> dict[str, Optional[CohortEntry]]:
    index = {entry.key: entry for entry in entries}
    return {entry.key: index.get(entry.sister_key) if entry.sister_key else None for entry in entries}


def style_table(ws, end_row: int, end_col: int) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=1, max_row=end_row, min_col=1, max_col=end_col):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.freeze_panes = "A2"


def add_excel_table(ws, name: str, end_row: int, end_col: int) -> None:
    table = Table(displayName=name, ref=f"A1:{column_letter(end_col)}{end_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def column_letter(col_num: int) -> str:
    result = ""
    while col_num:
        col_num, rem = divmod(col_num - 1, 26)
        result = chr(65 + rem) + result
    return result


def set_widths(ws, widths: List[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[column_letter(idx)].width = width


def create_executive_dashboard(ws) -> None:
    headers = [
        "Metric Name",
        "Current Value",
        "7-Day Rolling Average",
        "T30 Baseline",
        "Sister Property Avg",
        "YoY (if available)",
        "Variance vs T30 (%)",
        "Variance vs Sister (%)",
        "Status (1-7 Days)",
        "Status (7-15 Days)",
        "Status (15-30 Days)",
        "Status (30-60 Days)",
        "Status (60-90 Days)",
        "Source Status",
        "Notes / Placeholder Guidance",
    ]
    metrics = [
        ("Technical Health Index", "LIVE", "PSI + GTMetrix available now. CWV pass rate may need placeholder until final definition is locked."),
        ("Organic Traffic (True Organic Only)", "LIVE", "Use GA4 Organic Search channel only."),
        ("Organic % of Total Traffic", "LIVE", "Use GA4 traffic-source sessions / total sessions."),
        ("Weighted Engagement Score", "PLACEHOLDER", "Populate from Heap CTA events once naming is locked."),
        ("Guest Cards per Available Unit per Day", "PARTIAL", "Guest cards exist, available unit daily model may need final source alignment."),
        ("Early Funnel Conversion Rate", "PLACEHOLDER", "Fill from Snowflake / Heap / BI handoff once finalized."),
        ("Late Funnel Conversion Rate", "PLACEHOLDER", "Fill from Snowflake / BI leasing progression once finalized."),
    ]
    ws.append(headers)
    for metric_name, source_status, note in metrics:
        ws.append(
            [
                metric_name,
                "",
                "",
                "",
                "",
                "",
                '=IFERROR(([@[Current Value]]-[@[T30 Baseline]])/[@[T30 Baseline]],"")',
                '=IFERROR(([@[Current Value]]-[@[Sister Property Avg]])/[@[Sister Property Avg]],"")',
                "",
                "",
                "",
                "",
                "",
                source_status,
                note,
            ]
        )

    style_table(ws, ws.max_row, len(headers))
    add_excel_table(ws, "Executive_Dashboard", ws.max_row, len(headers))
    set_widths(ws, [30, 14, 16, 14, 16, 14, 16, 16, 16, 16, 16, 16, 16, 14, 44])


def create_property_map(ws, entries: List[CohortEntry]) -> None:
    sibling = sibling_lookup(entries)
    headers = [
        "Cohort Key",
        "Property ID",
        "Property Name",
        "Property Type",
        "Sister Cohort Key",
        "Sister Property ID",
        "Sister Property Name",
        "Site URL",
        "Market",
        "Notes",
    ]
    ws.append(headers)
    for entry in entries:
        sister = sibling[entry.key]
        ws.append(
            [
                entry.key,
                entry.property_id or "",
                entry.display_name,
                "Pilot" if entry.role == "pilot" else "Sister",
                sister.key if sister else "",
                sister.property_id if sister and sister.property_id else "",
                sister.display_name if sister else "",
                entry.site_url,
                "",
                "Populate market if stakeholder wants market-level rollups.",
            ]
        )
    style_table(ws, ws.max_row, len(headers))
    add_excel_table(ws, "Property_Map", ws.max_row, len(headers))
    set_widths(ws, [28, 14, 32, 14, 28, 14, 32, 42, 16, 36])


def create_property_performance(ws, entries: List[CohortEntry]) -> None:
    headers = [
        "Date",
        "Property ID",
        "Property Name",
        "Property Type",
        "Market",
        "Metric Category",
        "Metric Name",
        "Value",
        "7-Day Rolling Avg",
        "T30 Baseline",
        "YoY",
        "Sister Property ID",
        "Sister Value",
        "Variance vs T30 (%)",
        "Variance vs Sister (%)",
        "Source Status",
        "Source System",
        "Notes",
    ]
    metric_rows = [
        ("Technical", "PSI Score", "LIVE", "PSI"),
        ("Technical", "GTMetrix Score", "LIVE", "GTMetrix"),
        ("Technical", "Core Web Vitals Pass Rate", "PLACEHOLDER", "PSI / CrUX"),
        ("Technical", "Technical Health Index", "PARTIAL", "Derived"),
        ("Traffic", "Organic Traffic (True Organic Only)", "LIVE", "GA4"),
        ("Traffic", "Organic % of Total Traffic", "LIVE", "GA4"),
        ("Traffic", "YoY Organic Traffic", "LIVE", "GA4"),
        ("Engagement", "Weighted Engagement Score", "PLACEHOLDER", "Heap"),
        ("Funnel", "Early Funnel Conversion Rate", "PLACEHOLDER", "Heap / BI"),
        ("Funnel", "Late Funnel Conversion Rate", "PLACEHOLDER", "BI"),
        ("Operations", "Guest Cards per Available Unit per Day", "PARTIAL", "BI / Availability"),
    ]
    sibling = sibling_lookup(entries)

    ws.append(headers)
    for entry in entries:
        sister = sibling[entry.key]
        for category, metric_name, source_status, source_system in metric_rows:
            ws.append(
                [
                    "",
                    entry.property_id or "",
                    entry.display_name,
                    "Pilot" if entry.role == "pilot" else "Sister",
                    "",
                    category,
                    metric_name,
                    "",
                    '=IFERROR(AVERAGEIFS(Property_Performance[Value],Property_Performance[Property ID],[@[Property ID]],Property_Performance[Metric Name],[@[Metric Name]],Property_Performance[Date],">="&[@Date]-6,Property_Performance[Date],"<="&[@Date]),"")',
                    "",
                    "",
                    sister.property_id if sister and sister.property_id else "",
                    "",
                    '=IFERROR(([@Value]-[@[T30 Baseline]])/[@[T30 Baseline]],"")',
                    '=IFERROR(([@Value]-[@[Sister Value]])/[@[Sister Value]],"")',
                    source_status,
                    source_system,
                    "Populate via Power Query or scripted export.",
                ]
            )

    style_table(ws, ws.max_row, len(headers))
    add_excel_table(ws, "Property_Performance", ws.max_row, len(headers))
    set_widths(ws, [12, 14, 28, 14, 16, 18, 32, 12, 16, 14, 12, 14, 12, 16, 16, 14, 14, 32])


def create_engagement_detail(ws, entries: List[CohortEntry]) -> None:
    headers = [
        "Date",
        "Property ID",
        "Property Name",
        "Property Type",
        "CTA Type",
        "Users",
        "Clicks",
        "CTR (Clicks / Users)",
        "Weight",
        "Weighted Score",
        "Source Status",
        "Notes",
    ]
    cta_weights = {
        "Apply Now": 5,
        "Schedule Tour": 4,
        "Price Quote": 3,
        "Call": 2,
        "Text": 2,
        "Contact": 1,
    }
    ws.append(headers)
    for entry in entries:
        for cta_type, weight in cta_weights.items():
            ws.append(
                [
                    "",
                    entry.property_id or "",
                    entry.display_name,
                    "Pilot" if entry.role == "pilot" else "Sister",
                    cta_type,
                    "",
                    "",
                    '=IFERROR([@Clicks]/[@Users],"")',
                    weight,
                    '=IFERROR([@[CTR (Clicks / Users)]]*[@Weight],"")',
                    "PLACEHOLDER",
                    "Populate from Heap once CTA naming and extract are finalized.",
                ]
            )
    style_table(ws, ws.max_row, len(headers))
    add_excel_table(ws, "Engagement_Detail", ws.max_row, len(headers))
    set_widths(ws, [12, 14, 28, 14, 18, 12, 12, 16, 10, 14, 14, 38])

    start_row = ws.max_row + 3
    ws.cell(row=start_row, column=1, value="Daily Weighted Engagement Summary").font = Font(size=12, bold=True)
    summary_headers = ["Date", "Property ID", "Property Name", "Property Type", "Total Weighted Engagement Score", "Source Status"]
    for idx, header in enumerate(summary_headers, start=1):
        ws.cell(row=start_row + 1, column=idx, value=header)
    for entry in entries:
        ws.append([])
        row_num = ws.max_row + 1
        ws.cell(row=row_num, column=1, value="")
        ws.cell(row=row_num, column=2, value=entry.property_id or "")
        ws.cell(row=row_num, column=3, value=entry.display_name)
        ws.cell(row=row_num, column=4, value="Pilot" if entry.role == "pilot" else "Sister")
        ws.cell(
            row=row_num,
            column=5,
            value='=IFERROR(SUMIFS(Engagement_Detail[Weighted Score],Engagement_Detail[Property ID],B{0},Engagement_Detail[Date],A{0}),"")'.format(row_num),
        )
        ws.cell(row=row_num, column=6, value="PLACEHOLDER")


def create_funnel_performance(ws, entries: List[CohortEntry]) -> None:
    headers = [
        "Date",
        "Property ID",
        "Property Name",
        "Property Type",
        "Stage",
        "Count",
        "Conversion Rate (stage-to-stage)",
        "T30 Baseline",
        "YoY",
        "Sister Property Value",
        "Variance vs T30",
        "Variance vs Sister",
        "Source Status",
        "Notes",
    ]
    stages = ["Guest Card", "Price Quote", "Tour", "Application", "Move-in"]
    ws.append(headers)
    for entry in entries:
        for stage in stages:
            ws.append(
                [
                    "",
                    entry.property_id or "",
                    entry.display_name,
                    "Pilot" if entry.role == "pilot" else "Sister",
                    stage,
                    "",
                    "",
                    "",
                    "",
                    "",
                    '=IFERROR(([@Count]-[@[T30 Baseline]])/[@[T30 Baseline]],"")',
                    '=IFERROR(([@Count]-[@[Sister Property Value]])/[@[Sister Property Value]],"")',
                    "PLACEHOLDER",
                    "Populate from Snowflake / BI funnel output.",
                ]
            )

    style_table(ws, ws.max_row, len(headers))
    add_excel_table(ws, "Funnel_Performance", ws.max_row, len(headers))
    set_widths(ws, [12, 14, 28, 14, 16, 12, 18, 14, 12, 16, 14, 14, 14, 34])

    start_row = ws.max_row + 3
    ws.cell(row=start_row, column=1, value="Derived Funnel Metrics").font = Font(size=12, bold=True)
    for idx, header in enumerate(["Metric", "Formula / Placeholder Guidance"], start=1):
        ws.cell(row=start_row + 1, column=idx, value=header)
    ws.cell(row=start_row + 2, column=1, value="Early Funnel Conversion Rate")
    ws.cell(row=start_row + 2, column=2, value="(Price Quote + Tour) / Guest Card")
    ws.cell(row=start_row + 3, column=1, value="Late Funnel Conversion Rate")
    ws.cell(row=start_row + 3, column=2, value="Move-in / Application")


def create_operations_performance(ws, entries: List[CohortEntry]) -> None:
    headers = [
        "Date",
        "Property ID",
        "Property Name",
        "Property Type",
        "Available Units",
        "Guest Cards",
        "Guest Cards per Available Unit per Day",
        "Tours",
        "Applications",
        "Move-ins",
        "Lead-to-Lease Conversion %",
        "Sister Property Value",
        "Variance vs Sister",
        "Status Indicator",
        "Source Status",
        "Notes",
    ]
    ws.append(headers)
    for entry in entries:
        ws.append(
            [
                "",
                entry.property_id or "",
                entry.display_name,
                "Pilot" if entry.role == "pilot" else "Sister",
                "",
                "",
                '=IFERROR([@[Guest Cards]]/[@[Available Units]],"")',
                "",
                "",
                "",
                '=IFERROR([@[Move-ins]]/[@[Guest Cards]],"")',
                "",
                '=IFERROR(([@[Guest Cards per Available Unit per Day]]-[@[Sister Property Value]])/[@[Sister Property Value]],"")',
                '=IF([@[Variance vs Sister]]="","",IF([@[Variance vs Sister]]>=0.05,"Green",IF([@[Variance vs Sister]]<=-0.05,"Red","Yellow")))',
                "PARTIAL",
                "Guest cards are available. Tours / applications / move-ins may need final BI source alignment.",
            ]
        )
    style_table(ws, ws.max_row, len(headers))
    add_excel_table(ws, "Operations_Performance", ws.max_row, len(headers))
    set_widths(ws, [12, 14, 28, 14, 14, 12, 18, 12, 12, 12, 16, 16, 14, 14, 14, 40])


def create_technical_health(ws, entries: List[CohortEntry]) -> None:
    headers = [
        "Date",
        "Property ID",
        "Property Name",
        "Property Type",
        "PSI Score",
        "GTMetrix Score",
        "Core Web Vitals Pass Rate",
        "Technical Health Index",
        "Source Status",
        "Notes",
    ]
    ws.append(headers)
    for entry in entries:
        ws.append(
            [
                "",
                entry.property_id or "",
                entry.display_name,
                "Pilot" if entry.role == "pilot" else "Sister",
                "",
                "",
                "",
                '=IF(OR([@[PSI Score]]="",[@[GTMetrix Score]]="",[@[Core Web Vitals Pass Rate]]=""),"",([@[PSI Score]]*0.4)+([@[GTMetrix Score]]*0.4)+([@[Core Web Vitals Pass Rate]]*0.2))',
                "PARTIAL",
                "PSI and GTMetrix are live. CWV pass rate may remain placeholder until stakeholder definition is finalized.",
            ]
        )
    style_table(ws, ws.max_row, len(headers))
    add_excel_table(ws, "Technical_Health", ws.max_row, len(headers))
    set_widths(ws, [12, 14, 28, 14, 12, 14, 18, 18, 14, 40])

    green_fill = PatternFill("solid", fgColor="C6EFCE")
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    ws.conditional_formatting.add("H2:H500", CellIsRule(operator="greaterThanOrEqual", formula=["90"], fill=green_fill))
    ws.conditional_formatting.add("H2:H500", CellIsRule(operator="between", formula=["85", "89.999"], fill=yellow_fill))
    ws.conditional_formatting.add("H2:H500", CellIsRule(operator="lessThan", formula=["85"], fill=red_fill))


def create_metric_definitions(ws) -> None:
    headers = [
        "Metric Name",
        "Definition",
        "Current Source of Record",
        "Status Logic",
        "Placeholder / Decision Notes",
    ]
    rows = [
        (
            "Technical Health Index",
            "Composite score: (PSI * 0.4) + (GTMetrix * 0.4) + (CWV Pass Rate * 0.2).",
            "PSI + GTMetrix + CWV/CrUX",
            "Use staged threshold logic by time window; tighten expectations after the volatility phase.",
            "CWV pass rate source still needs final agreement.",
        ),
        (
            "Organic Traffic (True Organic Only)",
            "GA4 sessions from Organic Search only.",
            "GA4",
            "Compare current value vs T30 / sister / YoY.",
            "Do not blend Direct with Organic.",
        ),
        (
            "Weighted Engagement Score",
            "Weighted CTA interaction score using Apply / Tour / Quote / Call / Text / Contact weights.",
            "Heap",
            "Interpret as UX efficiency, not just traffic volume.",
            "Needs locked Heap event naming.",
        ),
        (
            "Guest Cards per Available Unit per Day",
            "Guest Cards divided by available inventory for the day.",
            "BI / Inventory",
            "Best used as revenue-pressure context signal.",
            "Inventory snapshot definition still needs to be finalized.",
        ),
        (
            "Early Funnel Conversion Rate",
            "(Price Quote + Tour) / Guest Cards.",
            "Heap / BI",
            "Should stabilize by day 30 if experience is healthy.",
            "Requires aligned funnel source.",
        ),
        (
            "Late Funnel Conversion Rate",
            "Move-ins / Applications.",
            "BI",
            "Used to detect lead-quality degradation.",
            "Requires downstream leasing data feed.",
        ),
    ]
    ws.append(headers)
    for row in rows:
        ws.append(list(row))
    style_table(ws, ws.max_row, len(headers))
    add_excel_table(ws, "Metric_Definitions", ws.max_row, len(headers))
    set_widths(ws, [28, 46, 24, 42, 42])


def create_thresholds(ws) -> None:
    headers = ["Metric", "Time Window", "Green Threshold", "Yellow Threshold", "Red Threshold", "Notes"]
    rows = [
        ("Variance-based metrics", "1-7 Days", ">= +5%", "-5% to +5%", "<= -5%", "Default dashboard variance rule from commissioned brief."),
        ("Variance-based metrics", "7-15 Days", ">= +5%", "-5% to +5%", "<= -5%", "Use until custom metric thresholds are finalized."),
        ("Variance-based metrics", "15-30 Days", ">= +5%", "-5% to +5%", "<= -5%", "Use until custom metric thresholds are finalized."),
        ("Variance-based metrics", "30-60 Days", ">= +5%", "-5% to +5%", "<= -5%", "Use until custom metric thresholds are finalized."),
        ("Variance-based metrics", "60-90 Days", ">= +5%", "-5% to +5%", "<= -5%", "Use until custom metric thresholds are finalized."),
        ("Technical Health Index", "1-7 Days", ">= 90", "85-89", "< 85", "From the status-threshold notes in the instructions."),
        ("Technical Health Index", "7-15 Days", ">= 92", "88-91", "< 88", "From the status-threshold notes in the instructions."),
        ("Technical Health Index", "15-30 Days", ">= 94", "90-93", "< 90", "From the status-threshold notes in the instructions."),
        ("Technical Health Index", "30-60 Days", ">= 95", "92-94", "< 92", "From the status-threshold notes in the instructions."),
        ("Technical Health Index", "60-90 Days", ">= 96", "93-95", "< 93", "From the status-threshold notes in the instructions."),
    ]
    ws.append(headers)
    for row in rows:
        ws.append(list(row))
    style_table(ws, ws.max_row, len(headers))
    add_excel_table(ws, "Thresholds", ws.max_row, len(headers))
    set_widths(ws, [28, 16, 16, 18, 16, 42])


def create_readme(ws) -> None:
    ws["A1"] = "Pilot KPI Tracker Template"
    ws["A1"].font = Font(size=16, bold=True)
    lines = [
        "Purpose: commissioned live workbook structure for pilot vs sister-property performance tracking.",
        "Built with full workbook structure now, including placeholders for sources still pending.",
        "Live-ready inputs today: PSI, GTMetrix, GA4 organic traffic, YoY GA4 traffic, pilot/sister mapping.",
        "Placeholder inputs pending final delivery: Heap CTA detail, some funnel data, some operations / inventory joins.",
        "Recommended ingestion path: Power Query or scripted exports into the Excel Tables on each source tab.",
        "All metric tables are linked by Property ID.",
    ]
    for idx, line in enumerate(lines, start=3):
        ws[f"A{idx}"] = line
    ws.column_dimensions["A"].width = 120


def build_workbook(entries: List[CohortEntry]) -> Workbook:
    wb = Workbook()
    readme = wb.active
    readme.title = "README"
    create_readme(readme)

    create_executive_dashboard(wb.create_sheet("Executive Dashboard"))
    create_property_map(wb.create_sheet("Property Map"), entries)
    create_property_performance(wb.create_sheet("Property Performance"), entries)
    create_engagement_detail(wb.create_sheet("Engagement Detail"), entries)
    create_funnel_performance(wb.create_sheet("Funnel Performance"), entries)
    create_operations_performance(wb.create_sheet("Operations Performance"), entries)
    create_technical_health(wb.create_sheet("Technical Health"), entries)
    create_metric_definitions(wb.create_sheet("Metric Definitions"))
    create_thresholds(wb.create_sheet("Thresholds"))
    return wb


def main() -> int:
    config = load_config(DEFAULT_CONFIG_PATH)
    entries = active_entries(config)
    workbook = build_workbook(entries)
    out_path = OUTPUT_DIR / f"Pilot_KPI_Tracker_Template_{date.today().isoformat()}.xlsx"
    workbook.save(out_path)
    print(f"Saved workbook: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
