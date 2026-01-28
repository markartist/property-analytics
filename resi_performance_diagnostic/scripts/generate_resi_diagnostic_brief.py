#!/usr/bin/env python3
"""
Resi Performance Diagnostic — Conversion Efficiency & Demand Ramp
Separates conversion quality from visibility/ramp constraints.
"""

import sqlite3
import json
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Paths
BASE_DIR = Path(__file__).parent.parent
DB_PATH = Path("/Users/mark/Property_Analytics/data/portfolio_analytics.db")
REGISTRY_PATH = Path("/Users/mark/Property_Analytics/config/venterra_properties_official.json")
OUTPUT_DIR = BASE_DIR / "reports" / "resi_diagnostic" / datetime.now().strftime("%Y-%m-%d")

# Create output directory
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Date ranges (30-day window)
TODAY = datetime.now().date()
GA4_END = TODAY - timedelta(days=1)  # 1-day lag
GA4_START = GA4_END - timedelta(days=30)
GSC_END = TODAY - timedelta(days=3)  # 3-day lag
GSC_START = GSC_END - timedelta(days=30)

print("=" * 70)
print("RESI PERFORMANCE DIAGNOSTIC")
print("Conversion Efficiency & Demand Ramp")
print("=" * 70)
print(f"\nRun Date: {TODAY}")
print(f"GA4 Window: {GA4_START} to {GA4_END}")
print(f"GSC Window: {GSC_START} to {GSC_END}")
print()

# Load property registry
print("📚 Loading Resi properties from registry...")
with open(REGISTRY_PATH, 'r') as f:
    registry_data = json.load(f)

resi_properties = []
for prop in registry_data.get('properties', []):
    if prop.get('site_type') == 'resi':
        resi_properties.append({
            'name': prop.get('name'),
            'ga4_id': prop.get('ga4_property_id'),
            'gsc_url': prop.get('gsc_url')
        })

print(f"✓ Found {len(resi_properties)} Resi properties")
for p in resi_properties:
    print(f"  - {p['name']}")
print()

# Connect to database
print("🗄️  Connecting to database...")
conn = sqlite3.connect(DB_PATH)
cursor = conn.cursor()

# Collect comprehensive data for all properties
print("\n" + "=" * 70)
print("DATA COLLECTION (ALL RESI PROPERTIES)")
print("=" * 70)

property_data = []

for prop in resi_properties:
    print(f"\n📊 Analyzing: {prop['name']}")
    
    data = {
        'name': prop['name'],
        'ga4_id': prop['ga4_id'],
        'gsc_url': prop['gsc_url']
    }
    
    # === GSC DATA ===
    gsc_query = """
    SELECT 
        MIN(metric_date) as first_date,
        MAX(metric_date) as last_date,
        SUM(clicks) as total_clicks,
        SUM(impressions) as total_impressions,
        AVG(average_position) as avg_position
    FROM gsc_daily_metrics
    WHERE property_id = ?
    AND metric_date BETWEEN ? AND ?
    """
    
    cursor.execute(gsc_query, (prop['gsc_url'], str(GSC_START), str(GSC_END)))
    gsc_row = cursor.fetchone()
    
    if gsc_row and gsc_row[2]:  # Has clicks
        first_date, last_date, clicks, impressions, position = gsc_row
        
        # Calculate days tracked (from first ever data to today)
        cursor.execute(
            "SELECT MIN(metric_date) FROM gsc_daily_metrics WHERE property_id = ?",
            (prop['gsc_url'],)
        )
        first_ever = cursor.fetchone()[0]
        
        if first_ever:
            days_tracked = (TODAY - datetime.strptime(first_ever, '%Y-%m-%d').date()).days
            click_velocity = clicks / 30  # Clicks per day in 30-day window
        else:
            days_tracked = None
            click_velocity = None
        
        ctr = (clicks / impressions * 100) if impressions else 0
        
        data['gsc'] = {
            'clicks_30d': clicks,
            'impressions_30d': impressions,
            'ctr': ctr,
            'avg_position': position,
            'days_tracked': days_tracked,
            'click_velocity': click_velocity,
            'first_tracked': first_ever
        }
        
        print(f"  GSC: {clicks} clicks, {impressions} impressions, {click_velocity:.1f} clicks/day")
        print(f"       Tracked for {days_tracked} days since {first_ever}")
    else:
        data['gsc'] = None
        print(f"  GSC: No data")
    
    # === GA4 DATA ===
    ga4_query = """
    SELECT 
        SUM(sessions) as total_sessions,
        SUM(engaged_sessions) as total_engaged_sessions
    FROM ga4_daily_metrics
    WHERE property_id = ?
    AND metric_date BETWEEN ? AND ?
    """
    
    cursor.execute(ga4_query, (prop['ga4_id'], str(GA4_START), str(GA4_END)))
    ga4_row = cursor.fetchone()
    
    if ga4_row and ga4_row[0]:
        sessions, engaged_sessions = ga4_row
        engagement_rate = (engaged_sessions / sessions * 100) if sessions else 0
        
        data['ga4'] = {
            'sessions': sessions,
            'engaged_sessions': engaged_sessions,
            'engagement_rate': engagement_rate
        }
        
        print(f"  GA4: {sessions} sessions, {engagement_rate:.1f}% engagement rate")
        
        # TODO: Add meaningful actions per session when event mapping is validated
        data['ga4']['meaningful_actions_per_session'] = None
    else:
        data['ga4'] = None
        print(f"  GA4: No data")
    
    # === CWV DATA ===
    cwv_query = """
    SELECT 
        lcp_value, fid_value, cls_value, strategy
    FROM pagespeed_metrics
    WHERE property_id = ?
    AND metric_date = (
        SELECT MAX(metric_date) 
        FROM pagespeed_metrics 
        WHERE property_id = ?
    )
    """
    
    cursor.execute(cwv_query, (prop['ga4_id'], prop['ga4_id']))
    cwv_rows = cursor.fetchall()
    
    if cwv_rows:
        mobile_cwv = None
        for row in cwv_rows:
            lcp, fid, cls, strategy = row
            if strategy == 'mobile':
                mobile_cwv = {'lcp': lcp, 'fid': fid, 'cls': cls}
                break
        
        if mobile_cwv:
            lcp_pass = mobile_cwv['lcp'] <= 2.5
            fid_pass = mobile_cwv['fid'] <= 100
            cls_pass = mobile_cwv['cls'] <= 0.1
            cwv_pass = lcp_pass and fid_pass and cls_pass
            
            data['cwv'] = {
                'lcp': mobile_cwv['lcp'],
                'fid': mobile_cwv['fid'],
                'cls': mobile_cwv['cls'],
                'passes_good': cwv_pass
            }
            
            status = "✓ Pass" if cwv_pass else "✗ Needs Improvement"
            print(f"  CWV: {status} (LCP: {mobile_cwv['lcp']:.2f}s, FID: {mobile_cwv['fid']:.0f}ms, CLS: {mobile_cwv['cls']:.3f})")
        else:
            data['cwv'] = None
            print(f"  CWV: No mobile data")
    else:
        data['cwv'] = None
        print(f"  CWV: No data")
    
    property_data.append(data)

conn.close()

# === CLASSIFICATION LOGIC ===
print("\n" + "=" * 70)
print("PROPERTY CLASSIFICATION")
print("=" * 70)

for prop in property_data:
    gsc = prop.get('gsc')
    
    if not gsc:
        prop['classification'] = 'No GSC Data'
        prop['classification_reason'] = 'No GSC tracking data available'
        continue
    
    clicks = gsc['clicks_30d']
    days_tracked = gsc['days_tracked']
    velocity = gsc['click_velocity']
    
    # Classification rules (deterministic)
    if clicks >= 300:
        prop['classification'] = 'Conversion-Ready'
        prop['classification_reason'] = f'Passed volume gate ({clicks} clicks ≥ 300)'
    elif days_tracked < 45:
        prop['classification'] = 'Ramp-Stage'
        prop['classification_reason'] = f'New property ({days_tracked} days tracked < 45)'
    elif days_tracked >= 60 and clicks < 300 and velocity < 5:
        prop['classification'] = 'Visibility-Constrained'
        prop['classification_reason'] = f'Low organic visibility ({clicks} clicks in {days_tracked} days, {velocity:.1f} clicks/day)'
    else:
        # Edge case: Between ramp and visibility-constrained
        prop['classification'] = 'Ramp-Stage'
        prop['classification_reason'] = f'Building visibility ({days_tracked} days tracked, {velocity:.1f} clicks/day)'
    
    print(f"\n{prop['name']}")
    print(f"  Classification: {prop['classification']}")
    print(f"  Reason: {prop['classification_reason']}")

# === SECTION A: CONVERSION EFFICIENCY ===
print("\n" + "=" * 70)
print("SECTION A — CONVERSION EFFICIENCY (HIGH CONFIDENCE)")
print("=" * 70)

conversion_ready = [p for p in property_data if p['classification'] == 'Conversion-Ready']

print(f"\nProperties with sufficient demand: {len(conversion_ready)}")
for prop in conversion_ready:
    print(f"  ✓ {prop['name']}")

if not conversion_ready:
    print("\n⚠️  No properties passed volume gate (≥300 clicks)")
    print("   Section A will note insufficient data for conversion analysis.")

# === SECTION B: VISIBILITY & RAMP HEALTH ===
print("\n" + "=" * 70)
print("SECTION B — VISIBILITY & RAMP HEALTH (ALL PROPERTIES)")
print("=" * 70)

print(f"\nAll {len(property_data)} Resi properties included:")
for prop in property_data:
    print(f"  • {prop['name']}: {prop['classification']}")

# === GENERATE HTML BRIEF ===
print("\n" + "=" * 70)
print("GENERATING HTML EXECUTIVE BRIEF")
print("=" * 70)

html = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>Resi Performance Diagnostic</title>
    <style>
        body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; max-width: 900px; margin: 40px auto; padding: 20px; background: #f5f5f5; }}
        .container {{ background: white; padding: 30px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
        h1 {{ color: #2c3e50; border-bottom: 3px solid #3498db; padding-bottom: 10px; }}
        h2 {{ color: #34495e; margin-top: 30px; background: #ecf0f1; padding: 10px; border-left: 4px solid #3498db; }}
        .summary-box {{ background: #e8f4f8; padding: 15px; border-left: 4px solid #3498db; margin: 20px 0; }}
        .warning-box {{ background: #fff3cd; border-left: 4px solid #ffc107; padding: 15px; margin: 20px 0; }}
        table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
        th, td {{ padding: 12px; text-align: left; border-bottom: 1px solid #ddd; }}
        th {{ background: #3498db; color: white; font-weight: 600; }}
        tr:hover {{ background: #f8f9fa; }}
        .metric-value {{ font-weight: 600; font-size: 1.1em; }}
        .conversion-ready {{ color: #27ae60; font-weight: 600; }}
        .ramp-stage {{ color: #f39c12; font-weight: 600; }}
        .visibility-constrained {{ color: #e74c3c; font-weight: 600; }}
        .pass {{ color: #27ae60; }}
        .fail {{ color: #e74c3c; }}
        .footer {{ text-align: center; margin-top: 40px; color: #7f8c8d; font-size: 0.9em; border-top: 1px solid #ddd; padding-top: 20px; }}
    </style>
</head>
<body>
    <div class="container">
        <h1>Resi Performance Diagnostic</h1>
        <h3>Conversion Efficiency & Demand Ramp (30-Day)</h3>
        
        <div class="summary-box">
            <strong>Run Date:</strong> {TODAY}<br>
            <strong>Data Window:</strong> GA4 ({GA4_START} to {GA4_END}), GSC ({GSC_START} to {GSC_END})<br>
            <strong>Properties Analyzed:</strong> {len(property_data)} Resi properties<br>
            <strong>Framework:</strong> Atlas Governed Diagnostic (separates conversion quality from visibility constraints)
        </div>
        
        <h2>Section A — Conversion Efficiency (When Demand Exists)</h2>
"""

if conversion_ready:
    html += f"""
        <p><strong>Properties with Sufficient Demand:</strong> {len(conversion_ready)} (≥300 organic clicks in 30 days)</p>
        
        <table>
            <tr>
                <th>Property</th>
                <th>SERP CTR</th>
                <th>Avg Position</th>
                <th>Engagement Rate</th>
                <th>Sessions</th>
            </tr>
"""
    
    for prop in conversion_ready:
        gsc = prop['gsc']
        ga4 = prop['ga4']
        
        html += f"""
            <tr>
                <td><strong>{prop['name']}</strong></td>
                <td class="metric-value">{gsc['ctr']:.2f}%</td>
                <td>{gsc['avg_position']:.1f}</td>
                <td class="metric-value">{ga4['engagement_rate']:.1f}%</td>
                <td>{ga4['sessions']:,}</td>
            </tr>
"""
    
    html += """
        </table>
        
        <p style="margin-top: 15px;">
            <strong>Interpretation:</strong> These properties have sufficient organic demand to assess conversion efficiency. 
            Results represent Resi performance when visibility constraints are not the limiting factor.
        </p>
"""
else:
    html += """
        <div class="warning-box">
            <strong>⚠️ Insufficient Data for Conversion Analysis</strong><br>
            No Resi properties passed the volume gate (≥300 organic clicks in 30 days). 
            Conversion efficiency cannot be assessed without sufficient demand sample size.
        </div>
        <p>See Section B for visibility and ramp health analysis.</p>
"""

html += f"""
        <h2>Section B — Visibility & Ramp Health (All Properties)</h2>
        
        <p>All {len(property_data)} Resi properties included, regardless of volume:</p>
        
        <table>
            <tr>
                <th>Property</th>
                <th>Days Tracked</th>
                <th>Clicks (30d)</th>
                <th>Click Velocity</th>
                <th>CWV Status</th>
                <th>Classification</th>
            </tr>
"""

for prop in property_data:
    gsc = prop.get('gsc')
    cwv = prop.get('cwv')
    
    if gsc:
        days_tracked = gsc['days_tracked']
        clicks = gsc['clicks_30d']
        velocity = gsc['click_velocity']
    else:
        days_tracked = 'N/A'
        clicks = 0
        velocity = 0
    
    if cwv:
        cwv_status = '<span class="pass">✓ Pass</span>' if cwv['passes_good'] else '<span class="fail">✗ Needs Improvement</span>'
    else:
        cwv_status = 'No data'
    
    classification = prop['classification']
    classification_class = classification.lower().replace(' ', '-').replace('-', '_')
    
    html += f"""
            <tr>
                <td><strong>{prop['name']}</strong></td>
                <td>{days_tracked}</td>
                <td class="metric-value">{clicks}</td>
                <td>{velocity:.1f} /day</td>
                <td>{cwv_status}</td>
                <td class="{classification_class}">{classification}</td>
            </tr>
"""

html += """
        </table>
        
        <h3>Classification Definitions</h3>
        <ul>
            <li><strong class="conversion-ready">Conversion-Ready:</strong> Passed volume gate (≥300 clicks). Sufficient data for conversion analysis.</li>
            <li><strong class="ramp-stage">Ramp-Stage:</strong> New property (&lt;45 days tracked) building organic visibility.</li>
            <li><strong class="visibility-constrained">Visibility-Constrained:</strong> Mature property (≥60 days) with low organic traffic (&lt;300 clicks, &lt;5 clicks/day).</li>
        </ul>
        
        <h3>Executive Interpretation</h3>
"""

# Generate deterministic narrative
conversion_count = len([p for p in property_data if p['classification'] == 'Conversion-Ready'])
ramp_count = len([p for p in property_data if p['classification'] == 'Ramp-Stage'])
visibility_count = len([p for p in property_data if p['classification'] == 'Visibility-Constrained'])

html += f"""
        <div class="summary-box">
"""

if conversion_count > 0:
    html += f"""
            <p><strong>Conversion Efficiency:</strong> {conversion_count} of {len(property_data)} Resi properties have sufficient demand for conversion analysis. 
            These properties demonstrate the platform's ability to convert organic traffic when visibility is not the constraint.</p>
"""
else:
    html += f"""
            <p><strong>Conversion Efficiency:</strong> None of the {len(property_data)} Resi properties have sufficient demand (≥300 clicks) for high-confidence conversion analysis. 
            This indicates visibility, not conversion failure, is the primary constraint.</p>
"""

if ramp_count > 0:
    html += f"""
            <p><strong>Ramp Health:</strong> {ramp_count} properties are in ramp-stage (&lt;45 days tracked), actively building organic visibility. 
            Performance will strengthen as these properties accumulate search presence.</p>
"""

if visibility_count > 0:
    html += f"""
            <p><strong>Visibility Constraints:</strong> {visibility_count} properties are visibility-constrained despite 60+ days of tracking. 
            These properties require SEO/visibility interventions, not conversion optimization.</p>
"""

html += """
        </div>
        
        <h3>Guardrails & Limitations</h3>
        <ul>
            <li><strong>Volume gates respected:</strong> ≥300 GSC clicks for Section A inclusion</li>
            <li><strong>No relaxed thresholds:</strong> Properties below gates excluded from conversion analysis</li>
            <li><strong>Deterministic classification:</strong> No subjective performance judgments</li>
            <li><strong>No external comparisons:</strong> This is an internal Resi diagnostic, not a competitive analysis</li>
        </ul>
        
        <div class="footer">
            Generated: """ + datetime.now().strftime("%Y-%m-%d %H:%M:%S") + """<br>
            Framework: Atlas Resi Performance Diagnostic v1.0
        </div>
    </div>
</body>
</html>"""

# Save HTML
html_path = OUTPUT_DIR / "resi_performance_diagnostic.html"
with open(html_path, 'w') as f:
    f.write(html)

print(f"\n✓ HTML brief saved: {html_path}")

# === GENERATE JSON ARTIFACT ===
json_data = {
    "meta": {
        "run_date": str(TODAY),
        "ga4_window": {"start": str(GA4_START), "end": str(GA4_END)},
        "gsc_window": {"start": str(GSC_START), "end": str(GSC_END)},
        "framework_version": "1.0",
        "total_properties": len(property_data)
    },
    "classifications": {
        "conversion_ready": conversion_count,
        "ramp_stage": ramp_count,
        "visibility_constrained": visibility_count
    },
    "properties": property_data
}

json_path = OUTPUT_DIR / "resi_performance_diagnostic.json"
with open(json_path, 'w') as f:
    json.dump(json_data, f, indent=2)

print(f"✓ JSON artifact saved: {json_path}")

# === GENERATE EXCEL APPENDIX ===
print("\n📊 Generating Excel appendix...")

wb = Workbook()
ws = wb.active
ws.title = "Resi Visibility & Ramp"

# Header styling
header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)

# Headers
headers = ["Property", "Classification", "Days Tracked", "Clicks (30d)", "Click Velocity", 
           "Impressions (30d)", "CTR", "Avg Position", "GA4 Sessions", "Engagement Rate", 
           "CWV Status", "LCP", "FID", "CLS"]

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_num, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

# Data rows
for row_num, prop in enumerate(property_data, 2):
    gsc = prop.get('gsc')
    ga4 = prop.get('ga4')
    cwv = prop.get('cwv')
    
    ws.cell(row=row_num, column=1, value=prop['name'])
    ws.cell(row=row_num, column=2, value=prop['classification'])
    
    if gsc:
        ws.cell(row=row_num, column=3, value=gsc['days_tracked'])
        ws.cell(row=row_num, column=4, value=gsc['clicks_30d'])
        ws.cell(row=row_num, column=5, value=f"{gsc['click_velocity']:.1f}")
        ws.cell(row=row_num, column=6, value=gsc['impressions_30d'])
        ws.cell(row=row_num, column=7, value=f"{gsc['ctr']:.2f}%")
        ws.cell(row=row_num, column=8, value=f"{gsc['avg_position']:.1f}")
    else:
        for col in range(3, 9):
            ws.cell(row=row_num, column=col, value="No data")
    
    if ga4:
        ws.cell(row=row_num, column=9, value=ga4['sessions'])
        ws.cell(row=row_num, column=10, value=f"{ga4['engagement_rate']:.1f}%")
    else:
        ws.cell(row=row_num, column=9, value="No data")
        ws.cell(row=row_num, column=10, value="No data")
    
    if cwv:
        ws.cell(row=row_num, column=11, value="Pass" if cwv['passes_good'] else "Needs Improvement")
        ws.cell(row=row_num, column=12, value=f"{cwv['lcp']:.2f}s")
        ws.cell(row=row_num, column=13, value=f"{cwv['fid']:.0f}ms")
        ws.cell(row=row_num, column=14, value=f"{cwv['cls']:.3f}")
    else:
        for col in range(11, 15):
            ws.cell(row=row_num, column=col, value="No data")

# Adjust column widths
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    ws.column_dimensions[column].width = min(max_length + 2, 50)

# Add Notes sheet
notes_ws = wb.create_sheet("Classification Logic")
notes_ws['A1'] = "Resi Performance Diagnostic — Classification Logic"
notes_ws['A1'].font = Font(bold=True, size=14)

notes_content = [
    "",
    "CONVERSION-READY",
    "• Passed volume gate (≥300 clicks in 30 days)",
    "• Sufficient data for conversion efficiency analysis",
    "",
    "RAMP-STAGE",
    "• < 45 days since tracking start",
    "• Building organic visibility",
    "• Performance will strengthen over time",
    "",
    "VISIBILITY-CONSTRAINED",
    "• ≥ 60 days tracked",
    "• Low clicks (< 300 in 30 days)",
    "• Low velocity (< 5 clicks/day)",
    "• Requires SEO/visibility interventions",
    "",
    "Data Sources:",
    "• GSC: Google Search Console (3-day lag)",
    "• GA4: Google Analytics 4 (1-day lag)",
    "• CWV: PageSpeed Insights (CrUX field data)",
    "",
    f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
]

for row_num, line in enumerate(notes_content, 3):
    notes_ws.cell(row=row_num, column=1, value=line)

excel_path = OUTPUT_DIR / "resi_performance_diagnostic.xlsx"
wb.save(excel_path)

print(f"✓ Excel appendix saved: {excel_path}")

print("\n" + "=" * 70)
print("✅ RESI PERFORMANCE DIAGNOSTIC COMPLETE")
print("=" * 70)
print(f"\nOutputs:")
print(f"  HTML: {html_path}")
print(f"  JSON: {json_path}")
print(f"  Excel: {excel_path}")
print(f"\nClassifications:")
print(f"  Conversion-Ready: {conversion_count}")
print(f"  Ramp-Stage: {ramp_count}")
print(f"  Visibility-Constrained: {visibility_count}")
