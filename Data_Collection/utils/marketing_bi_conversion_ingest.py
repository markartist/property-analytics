#!/usr/bin/env python3
"""Ingest selected Marketing BI conversion exports for Captain Brief diagnostics."""

from __future__ import annotations

import argparse
import json
import re
import sqlite3
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

from pypdf import PdfReader
from openpyxl import load_workbook

try:
    from Data_Collection.utils.property_identity import resolve_property_identity
except ModuleNotFoundError:
    import sys

    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
    from Data_Collection.utils.property_identity import resolve_property_identity

ROOT = Path("/Users/mark/Property_Analytics")
DB_PATH = ROOT / "data" / "portfolio_analytics.db"
CANCEL_DENIAL_SOURCE = Path("/Users/mark/Downloads/Property CancelDenial by Mktg Source.pdf")
TRAFFIC_CONVERSIONS_SOURCE = Path("/Users/mark/Downloads/Traffic Conversions T7D-T90D.pdf")
MIGRATION_SQL = ROOT / "apps" / "api" / "migrations" / "0030_create_marketing_bi_conversion_sources.sql"
CANCEL_DENIAL_XLSX_COLUMNS = {
    "Region",
    "Property",
    "Cnc Den Type",
    "Cnc Den Reason",
    "Mktg Src Desc",
    "C&Ds",
    "Apps",
    "GCards",
}

REGION_NAMES = {
    "Arkansas",
    "Atlanta, GA",
    "Austin, TX",
    "Dallas, TX",
    "Florida",
    "Houston, TX",
    "Kansas City",
    "Kentucky",
    "Killeen",
    "Nashville, TN",
    "Oklahoma",
    "Raleigh, NC",
    "San Antonio, TX",
    "Savannah, GA",
    "Total",
}


@dataclass
class MarketingBiIngestResult:
    report_date: str | None = None
    cancel_denial_rows: int = 0
    traffic_conversion_rows: int = 0
    source_files: list[str] = field(default_factory=list)


def parse_pdf_creation_date(raw: str | None) -> str | None:
    if not raw:
        return None
    match = re.search(r"D:(\d{4})(\d{2})(\d{2})", raw)
    if match:
        return f"{match.group(1)}-{match.group(2)}-{match.group(3)}"
    return None


def parse_number(value: str | None) -> float | None:
    if value is None:
        return None
    text = value.strip().replace(",", "").replace("%", "")
    if not text or text in {"\xa0", " "}:
        return None
    return float(text)


def parse_int(value: str | None) -> int | None:
    parsed = parse_number(value)
    if parsed is None:
        return None
    return int(round(parsed))


def parse_pct(value: str | None) -> float | None:
    parsed = parse_number(value)
    if parsed is None:
        return None
    return parsed / 100


def clean_line(line: str) -> str:
    return (
        line.replace("\uf166", " ")
        .replace("\uf164", " ")
        .replace("\uf165", " ")
        .replace("\ue102", " ")
        .replace("\xa0", " ")
        .strip()
    )


def split_line(line: str) -> list[str]:
    return [part.strip() for part in re.split(r"\s{2,}", clean_line(line)) if part.strip()]


def pdf_text(path: Path) -> tuple[str, str]:
    reader = PdfReader(str(path))
    report_date = parse_pdf_creation_date(str(reader.metadata.get("/CreationDate") if reader.metadata else "")) or datetime.now().date().isoformat()
    text = "\n".join(page.extract_text(extraction_mode="layout") or page.extract_text() or "" for page in reader.pages)
    return report_date, text


def slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")


def map_property(name: str) -> dict[str, Any]:
    identity = resolve_property_identity(name)
    if identity:
        return identity.as_mapping()
    return {"property_id": None, "ga4_property_id": None, "community_id": None, "canonical_name": name, "match_source": "unmapped"}


def parse_cancel_denial_pdf(path: Path = CANCEL_DENIAL_SOURCE) -> tuple[str, list[dict[str, Any]]]:
    report_date, text = pdf_text(path)
    rows: list[dict[str, Any]] = []
    for line in text.splitlines():
        parts = split_line(line)
        if len(parts) < 6 or parts[0] not in REGION_NAMES or parts[0] == "Total":
            continue
        region = parts[0]
        property_name = parts[1]
        if property_name in REGION_NAMES or property_name == "Total":
            continue
        mapped = map_property(property_name)

        if parts[2] in {"Cancel", "Denial"} and len(parts) >= 7:
            cancel_denial_type = parts[2]
            cancel_denial_reason = parts[3]
            marketing_source = parts[4]
            count = parse_int(parts[5])
            applications = parse_int(parts[6])
            guest_cards = parse_int(parts[7]) if len(parts) > 7 and re.fullmatch(r"[\d,]+", parts[7]) else None
        else:
            cancel_denial_type = None
            cancel_denial_reason = None
            marketing_source = parts[2]
            count = parse_int(parts[3])
            applications = parse_int(parts[4])
            guest_cards = parse_int(parts[5]) if len(parts) > 5 and re.fullmatch(r"[\d,]+", parts[5]) else None

        rows.append(
            {
                "id": f"cancel_denial_{report_date}_{slug(property_name)}_{slug(marketing_source)}_{slug(cancel_denial_type or 'summary')}_{slug(cancel_denial_reason or 'all')}",
                "report_date": report_date,
                "region": region,
                "property_name": property_name,
                "property_id": mapped.get("property_id"),
                "community_id": mapped.get("community_id"),
                "cancel_denial_type": cancel_denial_type,
                "cancel_denial_reason": cancel_denial_reason,
                "marketing_source": marketing_source,
                "cancel_denial_count": count,
                "applications": applications,
                "guest_cards": guest_cards,
                "source_file": str(path.resolve()),
                "evidence_json": json.dumps({"raw": parts, "mapping": mapped}, sort_keys=True),
            }
        )
    return report_date, rows


def parse_cancel_denial_xlsx(path: Path) -> tuple[str, list[dict[str, Any]]]:
    """Parse the native Power BI export for Property Cancel/Denial by Marketing Source."""
    report_date = datetime.fromtimestamp(path.stat().st_mtime).date().isoformat()
    workbook = load_workbook(str(path), read_only=True, data_only=True)
    worksheet = workbook["Export"] if "Export" in workbook.sheetnames else workbook.worksheets[0]
    iterator = worksheet.iter_rows(values_only=True)
    raw_header = next(iterator, None)
    if not raw_header:
        return report_date, []

    header = [str(value).strip() if value is not None else "" for value in raw_header]
    missing_columns = CANCEL_DENIAL_XLSX_COLUMNS - set(header)
    if missing_columns:
        raise ValueError(f"{path} is missing required columns: {sorted(missing_columns)}")

    rows: list[dict[str, Any]] = []
    for raw in iterator:
        if not raw or not any(value is not None for value in raw):
            continue
        values = dict(zip(header, raw))
        region = str(values.get("Region") or "").strip()
        property_name = str(values.get("Property") or "").strip()
        if not region or region == "Total" or region.startswith("Applied filters") or not property_name:
            continue

        marketing_source = str(values.get("Mktg Src Desc") or "").strip() or "Unattributed"
        cancel_denial_type = str(values.get("Cnc Den Type") or "").strip() or None
        cancel_denial_reason = str(values.get("Cnc Den Reason") or "").strip() or None
        mapped = map_property(property_name)
        rows.append(
            {
                "id": f"cancel_denial_{report_date}_{slug(property_name)}_{slug(marketing_source)}_{slug(cancel_denial_type or 'summary')}_{slug(cancel_denial_reason or 'all')}",
                "report_date": report_date,
                "region": region,
                "property_name": property_name,
                "property_id": mapped.get("property_id"),
                "community_id": mapped.get("community_id"),
                "cancel_denial_type": cancel_denial_type,
                "cancel_denial_reason": cancel_denial_reason,
                "marketing_source": marketing_source,
                "cancel_denial_count": parse_int(str(values.get("C&Ds"))) if values.get("C&Ds") is not None else None,
                "applications": parse_int(str(values.get("Apps"))) if values.get("Apps") is not None else None,
                "guest_cards": parse_int(str(values.get("GCards"))) if values.get("GCards") is not None else None,
                "source_file": str(path.resolve()),
                "evidence_json": json.dumps({"raw": values, "mapping": mapped}, sort_keys=True, default=str),
            }
        )
    return report_date, rows


def parse_cancel_denial_source(path: Path = CANCEL_DENIAL_SOURCE) -> tuple[str, list[dict[str, Any]]]:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return parse_cancel_denial_xlsx(path)
    return parse_cancel_denial_pdf(path)


def parse_traffic_conversions_pdf(path: Path = TRAFFIC_CONVERSIONS_SOURCE) -> tuple[str, list[dict[str, Any]]]:
    report_date, text = pdf_text(path)
    rows: list[dict[str, Any]] = []
    for line in text.splitlines():
        is_property_level = line.startswith(" \uf164") or line.startswith(" \uf166")
        if not is_property_level:
            continue
        parts = split_line(line)
        if len(parts) < 11:
            continue
        property_name = parts[0]
        if property_name in REGION_NAMES or property_name in {"Better Living Consultant", "Community Manager", "Community Support Specialist", "Former Employee", "Virtual Leasing Agent", "N/A"}:
            continue
        mapped = map_property(property_name)
        rows.append(
            {
                "id": f"traffic_conversions_{report_date}_{slug(property_name)}",
                "report_date": report_date,
                "region": None,
                "property_name": property_name,
                "property_id": mapped.get("property_id"),
                "community_id": mapped.get("community_id"),
                "assigned_pct_t7": parse_pct(parts[1]),
                "assigned_pct_t30": parse_pct(parts[2]),
                "guest_cards_t7": parse_int(parts[3]),
                "guest_cards_t7_py": parse_int(parts[4]),
                "guest_cards_t7_yoy": parse_pct(parts[5]),
                "guest_cards_t30": parse_int(parts[6]),
                "guest_cards_t30_py": parse_int(parts[7]),
                "guest_cards_t30_yoy": parse_pct(parts[8]),
                "guest_cards_t60": parse_int(parts[9]),
                "guest_cards_t60_py": parse_int(parts[10]),
                "source_file": str(path.resolve()),
                "evidence_json": json.dumps({"raw": parts, "mapping": mapped}, sort_keys=True),
            }
        )
    return report_date, rows


def ensure_tables(conn: sqlite3.Connection) -> None:
    conn.executescript(MIGRATION_SQL.read_text(encoding="utf-8"))


def upsert_cancel_denial(conn: sqlite3.Connection, rows: list[dict[str, Any]]) -> int:
    if not rows:
        return 0
    columns = [
        "id",
        "report_date",
        "region",
        "property_name",
        "property_id",
        "community_id",
        "cancel_denial_type",
        "cancel_denial_reason",
        "marketing_source",
        "cancel_denial_count",
        "applications",
        "guest_cards",
        "source_file",
        "evidence_json",
    ]
    placeholders = ", ".join("?" for _ in columns)
    conn.executemany(
        f"INSERT OR REPLACE INTO marketing_cancel_denial_by_source ({', '.join(columns)}) VALUES ({placeholders})",
        [[row.get(column) for column in columns] for row in rows],
    )
    return len(rows)


def upsert_traffic_conversions(conn: sqlite3.Connection, rows: list[dict[str, Any]]) -> int:
    if not rows:
        return 0
    columns = [
        "id",
        "report_date",
        "region",
        "property_name",
        "property_id",
        "community_id",
        "assigned_pct_t7",
        "assigned_pct_t30",
        "guest_cards_t7",
        "guest_cards_t7_py",
        "guest_cards_t7_yoy",
        "guest_cards_t30",
        "guest_cards_t30_py",
        "guest_cards_t30_yoy",
        "guest_cards_t60",
        "guest_cards_t60_py",
        "source_file",
        "evidence_json",
    ]
    placeholders = ", ".join("?" for _ in columns)
    conn.executemany(
        f"INSERT OR REPLACE INTO marketing_traffic_conversions ({', '.join(columns)}) VALUES ({placeholders})",
        [[row.get(column) for column in columns] for row in rows],
    )
    return len(rows)


def ingest_marketing_bi_conversion_pdfs(
    cancel_denial_source: Path = CANCEL_DENIAL_SOURCE,
    traffic_conversions_source: Path = TRAFFIC_CONVERSIONS_SOURCE,
    db_path: Path = DB_PATH,
    include_traffic_conversions: bool = True,
) -> MarketingBiIngestResult:
    cancel_date, cancel_rows = parse_cancel_denial_source(cancel_denial_source)
    traffic_date, traffic_rows = (
        parse_traffic_conversions_pdf(traffic_conversions_source) if include_traffic_conversions else (cancel_date, [])
    )
    with sqlite3.connect(str(db_path)) as conn:
        ensure_tables(conn)
        cancel_count = upsert_cancel_denial(conn, cancel_rows)
        traffic_count = upsert_traffic_conversions(conn, traffic_rows) if include_traffic_conversions else 0
        conn.commit()
    return MarketingBiIngestResult(
        report_date=max(cancel_date, traffic_date),
        cancel_denial_rows=cancel_count,
        traffic_conversion_rows=traffic_count,
        source_files=[str(cancel_denial_source), *([str(traffic_conversions_source)] if include_traffic_conversions else [])],
    )


def main() -> None:
    parser = argparse.ArgumentParser(description="Ingest Marketing BI conversion exports.")
    parser.add_argument("--cancel-denial-source", default=str(CANCEL_DENIAL_SOURCE))
    parser.add_argument("--traffic-conversions-source", default=str(TRAFFIC_CONVERSIONS_SOURCE))
    parser.add_argument("--db", default=str(DB_PATH))
    parser.add_argument("--skip-traffic-conversions", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    cancel_source = Path(args.cancel_denial_source).expanduser().resolve()
    traffic_source = Path(args.traffic_conversions_source).expanduser().resolve()
    cancel_date, cancel_rows = parse_cancel_denial_source(cancel_source)
    traffic_date, traffic_rows = (
        parse_traffic_conversions_pdf(traffic_source) if not args.skip_traffic_conversions else (cancel_date, [])
    )
    result = MarketingBiIngestResult(
        report_date=max(cancel_date, traffic_date),
        cancel_denial_rows=len(cancel_rows),
        traffic_conversion_rows=len(traffic_rows),
        source_files=[str(cancel_source), *([str(traffic_source)] if not args.skip_traffic_conversions else [])],
    )
    if not args.dry_run:
        with sqlite3.connect(str(Path(args.db).expanduser().resolve())) as conn:
            ensure_tables(conn)
            upsert_cancel_denial(conn, cancel_rows)
            if not args.skip_traffic_conversions:
                upsert_traffic_conversions(conn, traffic_rows)
            conn.commit()
    print(json.dumps(result.__dict__, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
