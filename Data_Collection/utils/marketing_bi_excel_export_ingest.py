#!/usr/bin/env python3
"""Ingest native Marketing BI Excel exports for Captain Brief grounding."""

from __future__ import annotations

import argparse
import json
import re
import sqlite3
import tempfile
import xml.etree.ElementTree as ET
from zipfile import ZipFile
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

try:
    from Data_Collection.utils.property_identity import resolve_property_identity
except ModuleNotFoundError:
    import sys

    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
    from Data_Collection.utils.property_identity import resolve_property_identity

ROOT = Path("/Users/mark/Property_Analytics")
DB_PATH = ROOT / "data" / "portfolio_analytics.db"
MIGRATION_SQL = ROOT / "apps" / "api" / "migrations" / "0035_create_marketing_bi_excel_exports.sql"
CONVERSION_DASHBOARD_MIGRATION_SQL = (
    ROOT / "apps" / "api" / "migrations" / "0036_create_marketing_bi_conversion_dashboard.sql"
)
RECOVERY_MIGRATION_SQL = ROOT / "apps" / "api" / "migrations" / "0037_create_marketing_bi_recovery_sources.sql"
COST_PER_CONVERSION_MIGRATION_SQL = (
    ROOT / "apps" / "api" / "migrations" / "0038_create_marketing_bi_cost_per_conversion.sql"
)
SOURCE_PERFORMANCE_MIGRATION_SQL = (
    ROOT / "apps" / "api" / "migrations" / "0045_create_marketing_bi_source_performance.sql"
)
AVAILABLE_UNIT_INTEREST_MIGRATION_SQL = (
    ROOT / "apps" / "api" / "migrations" / "0029_create_available_unit_interest_metrics.sql"
)
GAP_FILL_MIGRATION_SQL = ROOT / "apps" / "api" / "migrations" / "0046_create_marketing_bi_gap_fill_tables.sql"

DEFAULT_SOURCES = [
    Path("/Users/mark/Downloads/Ad Spend Total and Ad Spend Δ by Calendar Period (bins).xlsx"),
    Path("/Users/mark/Downloads/perf by src.xlsx"),
    Path("/Users/mark/Downloads/cancel-deny-reasons.xlsx"),
    Path("/Users/mark/Downloads/gift cards by source.xlsx"),
    Path("/Users/mark/Downloads/traffic performance.xlsx"),
    Path("/Users/mark/Downloads/Portfolio Summary.xlsx"),
    Path("/Users/mark/Downloads/Traffic Conversions.xlsx"),
    Path("/Users/mark/Downloads/conv per data.xlsx"),
    Path("/Users/mark/Downloads/converdsion by source.xlsx"),
    Path("/Users/mark/Downloads/conversion dashboard.xlsx"),
    Path("/Users/mark/Library/CloudStorage/OneDrive-VenterraRealty(Canada)Inc/Guest_Card_Reports/conversion-data.xlsx"),
    Path("/Users/mark/Library/CloudStorage/OneDrive-VenterraRealty(Canada)Inc/Guest_Card_Reports/converting-performance.xlsx"),
    Path("/Users/mark/Library/CloudStorage/OneDrive-VenterraRealty(Canada)Inc/Guest_Card_Reports/marketing-performance.xlsx"),
    Path("/Users/mark/Downloads/GC by Period.xlsx"),
    Path("/Users/mark/Downloads/Wow Spending.xlsx"),
    Path("/Users/mark/Downloads/Ad Spend Total, GCards, Visits, Leases and Ad Spend Δ by Calendar Period (bins).xlsx"),
    Path("/Users/mark/Downloads/Ad Spend Total and Ad Spend Δ by Region.xlsx"),
    Path("/Users/mark/Downloads/LeaseTermAvg and LeaseTermAvg Δ by Lease Type.xlsx"),
    Path("/Users/mark/Downloads/LeaseTermAvg and LeaseTermAvg Δ by Property.xlsx"),
    Path("/Users/mark/Downloads/Unfulfilled Lease Totals  Percentage of Move-Outs.xlsx"),
    Path("/Users/mark/Downloads/Skips and Evictions.xlsx"),
    Path("/Users/mark/Downloads/30-Day Live-It Love-It Guarantee.xlsx"),
    Path("/Users/mark/Downloads/Leases Paid by Mktg Src Grp.xlsx"),
    Path("/Users/mark/Downloads/Ad Spend Total by Mktg Src Grp.xlsx"),
    Path("/Users/mark/Downloads/Cost per Conversion by Ad Source.xlsx"),
    Path("/Users/mark/Downloads/Cost per Conversion - Trend.xlsx"),
    Path("/Users/mark/Downloads/Conversion by Month.xlsx"),
    Path("/Users/mark/Downloads/Top 10 Converting Sources (%).xlsx"),
    Path("/Users/mark/Downloads/Top 10 Marketing Sources (%).xlsx"),
    Path("/Users/mark/Downloads/Top 10 Marketing Sources.xlsx"),
    Path("/Users/mark/Downloads/Available Units by Floor Plan.xlsx"),
    Path("/Users/mark/Downloads/avail units.xlsx"),
    Path("/Users/mark/Downloads/VacDaysHistMax and VacDaysHistMax Δ by Bedroom Ct.xlsx"),
    Path("/Users/mark/Downloads/VacDaysCurrMax and VacDaysCurrMax Δ by Property.xlsx"),
    Path("/Users/mark/Downloads/VacDaysCurrAvg and VacDaysCurrAvg Δ by Property.xlsx"),
    Path("/Users/mark/Downloads/VacDaysHistMax and VacDaysHistMax Δ by Property.xlsx"),
    Path("/Users/mark/Downloads/VacDaysHistAvg and VacDaysHistAvg Δ by Property.xlsx"),
    Path("/Users/mark/Downloads/GC by time.xlsx"),
    Path("/Users/mark/Downloads/GC by Period2.xlsx"),
]

CONVERSION_DASHBOARD_FILENAMES = {
    "conversion dashboard.xlsx",
    "conversion-data.xlsx",
    "init contact  property  mktg src.xlsx",
}


def normalized_source_name(path: Path) -> str:
    """Normalize browser download suffixes without changing the stored source path."""
    name = path.name.lower()
    return re.sub(r"\s+\(\d+\)(?=\.xlsx$)", "", name)

TRAFFIC_COLUMN_MAP = {
    "Assigned Pct T7D": "assigned_pct_t7",
    "Assigned Pct T30D": "assigned_pct_t30",
    "GCards T7D": "guest_cards_t7",
    "GCards T7D PY": "guest_cards_t7_py",
    "GCards T7D YoY": "guest_cards_t7_yoy",
    "GCards T30D": "guest_cards_t30",
    "GCards T30D PY": "guest_cards_t30_py",
    "GCards T30D YoY": "guest_cards_t30_yoy",
    "GCards T60D": "guest_cards_t60",
    "GCards T60D PY": "guest_cards_t60_py",
    "GCards T60D YoY": "guest_cards_t60_yoy",
    "GCards T90D": "guest_cards_t90",
    "GCards T90D PY": "guest_cards_t90_py",
    "GCards T90D YoY": "guest_cards_t90_yoy",
    "Visits T7D": "visits_t7",
    "Visits T7D PY": "visits_t7_py",
    "Visits T7D YoY": "visits_t7_yoy",
    "Visits T30D": "visits_t30",
    "Visits T30D PY": "visits_t30_py",
    "Visits T30D YoY": "visits_t30_yoy",
    "Visits T60D": "visits_t60",
    "Visits T60D PY": "visits_t60_py",
    "Visits T60D YoY": "visits_t60_yoy",
    "Visits T90D": "visits_t90",
    "Visits T90D PY": "visits_t90_py",
    "Visits T90D YoY": "visits_t90_yoy",
    "Apps T7D": "apps_t7",
    "Apps T7D PY": "apps_t7_py",
    "Apps T7D YoY": "apps_t7_yoy",
    "Apps T30D": "apps_t30",
    "Apps T30D PY": "apps_t30_py",
    "Apps T30D YoY": "apps_t30_yoy",
    "Apps T60D": "apps_t60",
    "Apps T60D PY": "apps_t60_py",
    "Apps T60D YoY": "apps_t60_yoy",
    "Apps T90D": "apps_t90",
    "Apps T90D PY": "apps_t90_py",
    "Apps T90D YoY": "apps_t90_yoy",
    "RFP T7D": "rfp_t7",
    "RFP T7D PY": "rfp_t7_py",
    "RFP T7D YoY": "rfp_t7_yoy",
    "RFP T30D": "rfp_t30",
    "RFP T30D PY": "rfp_t30_py",
    "RFP T30D YoY": "rfp_t30_yoy",
    "RFP T60D": "rfp_t60",
    "RFP T60D PY": "rfp_t60_py",
    "RFP T60D YoY": "rfp_t60_yoy",
    "RFP T90D": "rfp_t90",
    "RFP T90D PY": "rfp_t90_py",
    "RFP T90D YoY": "rfp_t90_yoy",
    "Closing Ratio T7D": "closing_ratio_t7",
    "Closing Ratio T7D PY": "closing_ratio_t7_py",
    "Closing Ratio T7D YoY": "closing_ratio_t7_yoy",
    "Closing Ratio T30D": "closing_ratio_t30",
    "Closing Ratio T30D PY": "closing_ratio_t30_py",
    "Closing Ratio T30D YoY": "closing_ratio_t30_yoy",
    "Closing Ratio T45D": "closing_ratio_t45",
    "Closing Ratio T45D PY": "closing_ratio_t45_py",
    "Closing Ratio T45D YoY": "closing_ratio_t45_yoy",
    "Curr Apt Unit Ct": "current_apartment_unit_count",
    "ATR Avg T7D": "atr_avg_t7",
    "ATR Avg T30D": "atr_avg_t30",
    "ATR Avg T60D": "atr_avg_t60",
    "ATR Avg T90D": "atr_avg_t90",
}


@dataclass
class ExcelIngestResult:
    files_processed: int = 0
    files_failed: int = 0
    portfolio_summary_rows: int = 0
    ad_spend_rows: int = 0
    traffic_conversion_rows: int = 0
    conversion_dashboard_rows: int = 0
    vacancy_days_rows: int = 0
    lease_term_rows: int = 0
    wow_spending_rows: int = 0
    ad_spend_performance_rows: int = 0
    period_leakage_rows: int = 0
    cost_per_conversion_rows: int = 0
    source_performance_rows: int = 0
    move_ins_by_source_rows: int = 0
    available_unit_interest_rows: int = 0
    monthly_ad_spend_source_rows: int = 0
    portfolio_box_score_rows: int = 0
    service_delivery_rows: int = 0
    abandoned_application_rows: int = 0
    generic_rows: int = 0
    source_files: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


def slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")


def parse_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").replace("%", "").replace("$", "").replace(" ", "")
    if not text or text.lower() in {"-", "–", "—", "#value!", "#div/0!", "#n/a", "#ref!", "n/a", "na"}:
        return None
    return float(text)


def parse_int(value: Any) -> int | None:
    parsed = parse_number(value)
    if parsed is None:
        return None
    return int(round(parsed))


def iso_date(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, (int, float)) and value > 30000:
        return (datetime(1899, 12, 30) + timedelta(days=float(value))).date().isoformat()
    text = str(value).strip()
    return text or None


def report_date(path: Path) -> str:
    return datetime.fromtimestamp(path.stat().st_mtime).date().isoformat()


def map_property(name: str | None) -> dict[str, Any]:
    if not name:
        return {"property_id": None, "community_id": None, "match_source": "blank"}
    identity = resolve_property_identity(name)
    if identity:
        return identity.as_mapping()
    return {"property_id": None, "community_id": None, "canonical_name": name, "match_source": "unmapped"}


def source_performance_identity_from_filename(path: Path):
    stem = path.stem
    if stem.startswith("marketing-source-"):
        source_name = stem.replace("marketing-source-", "", 1)
    elif stem.startswith("perf-by-source-"):
        source_name = stem.replace("perf-by-source-", "", 1)
    else:
        return None
    return resolve_property_identity(source_name.replace("-", " "))


def load_rows(path: Path) -> tuple[list[str], list[tuple[int, dict[str, Any]]]]:
    # Some Power BI table exports expose table data only in normal workbook mode.
    try:
        wb = load_workbook(str(path), read_only=False, data_only=True)
    except ValueError:
        return load_rows_from_worksheet_xml(path)
    ws = wb["Export"] if "Export" in wb.sheetnames else wb.worksheets[0]
    header: list[str] | None = None
    rows: list[tuple[int, dict[str, Any]]] = []
    for row_number, raw in enumerate(ws.iter_rows(values_only=True), 1):
        if not raw or not any(value is not None for value in raw):
            continue
        values = [str(value).strip() if value is not None else "" for value in raw]
        if header is None:
            if len([value for value in values if value]) >= 2 and not values[0].startswith("Applied filters"):
                header = values
            continue
        row = dict(zip(header, raw))
        if str(next(iter(row.values()), "") or "").startswith("Applied filters"):
            continue
        rows.append((row_number, row))
    return header or [], rows


def load_rows_with_header_marker(path: Path, marker: str, sheet_name: str | None = None) -> tuple[list[str], list[tuple[int, dict[str, Any]]]]:
    wb = load_workbook(str(path), read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.worksheets[0]
    header: list[str] | None = None
    rows: list[tuple[int, dict[str, Any]]] = []
    for row_number, raw in enumerate(ws.iter_rows(values_only=True), 1):
        values = [str(value).strip() if value is not None else "" for value in raw]
        if header is None:
            if marker in values:
                header = values
            continue
        if not raw or not any(value is not None for value in raw):
            continue
        row = dict(zip(header, raw))
        rows.append((row_number, row))
    return header or [], rows


def load_rows_from_worksheet_xml(path: Path) -> tuple[list[str], list[tuple[int, dict[str, Any]]]]:
    """Parse simple Power BI worksheet XML when invalid NaN/Infinity values break openpyxl."""
    namespace = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    with ZipFile(path) as workbook_zip:
        sheet_name = "xl/worksheets/sheet1.xml"
        root = ET.fromstring(workbook_zip.read(sheet_name))
    header: list[str] | None = None
    rows: list[tuple[int, dict[str, Any]]] = []
    for row_el in root.findall(".//x:sheetData/x:row", namespace):
        row_number = int(row_el.attrib.get("r", "0") or 0)
        values = []
        for cell_el in row_el.findall("x:c", namespace):
            value_el = cell_el.find("x:v", namespace)
            value_text = value_el.text if value_el is not None else None
            if value_text in {None, ""}:
                values.append(None)
            elif cell_el.attrib.get("t") in {"str", "inlineStr"}:
                values.append(value_text)
            elif value_text in {"NaN", "-NaN", "Infinity", "-Infinity"}:
                values.append(None)
            else:
                try:
                    number = float(value_text)
                    values.append(int(number) if number.is_integer() else number)
                except ValueError:
                    values.append(value_text)
        if not any(value is not None for value in values):
            continue
        normalized = [str(value).strip() if value is not None else "" for value in values]
        if header is None:
            if len([value for value in normalized if value]) >= 2 and not normalized[0].startswith("Applied filters"):
                header = normalized
            continue
        row = dict(zip(header, values))
        if str(next(iter(row.values()), "") or "").startswith("Applied filters"):
            continue
        rows.append((row_number, row))
    return header or [], rows


def ensure_tables(conn: sqlite3.Connection) -> None:
    conn.executescript(MIGRATION_SQL.read_text(encoding="utf-8"))
    conn.executescript(CONVERSION_DASHBOARD_MIGRATION_SQL.read_text(encoding="utf-8"))
    conn.executescript(RECOVERY_MIGRATION_SQL.read_text(encoding="utf-8"))
    conn.executescript(COST_PER_CONVERSION_MIGRATION_SQL.read_text(encoding="utf-8"))
    conn.executescript(SOURCE_PERFORMANCE_MIGRATION_SQL.read_text(encoding="utf-8"))
    conn.executescript(AVAILABLE_UNIT_INTEREST_MIGRATION_SQL.read_text(encoding="utf-8"))
    conn.executescript(GAP_FILL_MIGRATION_SQL.read_text(encoding="utf-8"))


def upsert_rows(conn: sqlite3.Connection, table: str, rows: list[dict[str, Any]]) -> int:
    if not rows:
        return 0
    columns = list(rows[0].keys())
    placeholders = ", ".join("?" for _ in columns)
    conn.executemany(
        f"INSERT OR REPLACE INTO {table} ({', '.join(columns)}) VALUES ({placeholders})",
        [[row.get(column) for column in columns] for row in rows],
    )
    return len(rows)


def parse_portfolio_summary(path: Path) -> list[dict[str, Any]]:
    rd = report_date(path)
    _, rows = load_rows(path)
    parsed = []
    for _, row in rows:
        property_name = str(row.get("Short Name") or "").strip()
        if not property_name:
            continue
        mapped = map_property(property_name)
        parsed.append(
            {
                "id": f"portfolio_summary_{rd}_{slug(property_name)}",
                "report_date": rd,
                "encasa_region": row.get("Encasa Region"),
                "metro_region": row.get("Metro Region"),
                "property_name": property_name,
                "property_id": mapped.get("property_id"),
                "community_id": mapped.get("community_id"),
                "acquired_date": iso_date(row.get("Acquired")),
                "year_built": parse_int(row.get("Built")),
                "apartments": parse_int(row.get("Apts")),
                "retail": row.get("Retail"),
                "head_count_bom": parse_int(row.get("Head Ct (BoM)")),
                "residents": parse_int(row.get("Residents")),
                "leaseholders": parse_int(row.get("Leaseholders")),
                "occupants": parse_int(row.get("Occupants")),
                "adults": parse_int(row.get("Adults")),
                "minors": parse_int(row.get("Minors")),
                "age_unknown": parse_int(row.get("Age Unkn.")),
                "pets": parse_int(row.get("Pets")),
                "source_file": str(path.resolve()),
                "evidence_json": json.dumps({"raw": row, "mapping": mapped}, sort_keys=True, default=str),
            }
        )
    return parsed


def parse_ad_spend(path: Path) -> list[dict[str, Any]]:
    rd = report_date(path)
    _, rows = load_rows(path)
    parsed = []
    for _, row in rows:
        property_name = str(row.get("Property Hierarchy - Property") or "").strip()
        if not property_name:
            continue
        mapped = map_property(property_name)
        month = iso_date(row.get("Calendar Period (bins)"))
        parsed.append(
            {
                "id": f"ad_spend_{rd}_{month}_{slug(property_name)}",
                "report_date": rd,
                "calendar_month": month,
                "region": row.get("Property Hierarchy - Region"),
                "property_name": property_name,
                "property_id": mapped.get("property_id"),
                "community_id": mapped.get("community_id"),
                "ad_spend_total": parse_number(row.get("Ad Spend Total")),
                "ad_spend_delta": parse_number(row.get("Ad Spend \u0394")),
                "source_file": str(path.resolve()),
                "evidence_json": json.dumps({"raw": row, "mapping": mapped}, sort_keys=True, default=str),
            }
        )
    return parsed


def parse_traffic_conversions(path: Path) -> list[dict[str, Any]]:
    rd = report_date(path)
    _, rows = load_rows(path)
    parsed = []
    for _, row in rows:
        property_name = str(row.get("Property") or "").strip()
        if not property_name or property_name == "Total":
            continue
        mapped = map_property(property_name)
        item = {
            "id": f"traffic_conversions_full_{rd}_{slug(property_name)}_{slug(str(row.get('RA Job Function') or 'total'))}_{slug(str(row.get('Responsible Agent') or 'all'))}",
            "report_date": rd,
            "region": row.get("Region"),
            "property_name": property_name,
            "property_id": mapped.get("property_id"),
            "community_id": mapped.get("community_id"),
            "ra_job_function": row.get("RA Job Function"),
            "responsible_agent": row.get("Responsible Agent"),
            "source_file": str(path.resolve()),
            "evidence_json": json.dumps({"raw": row, "mapping": mapped}, sort_keys=True, default=str),
        }
        for source_col, db_col in TRAFFIC_COLUMN_MAP.items():
            if db_col == "current_apartment_unit_count":
                item[db_col] = parse_int(row.get(source_col))
            else:
                item[db_col] = parse_number(row.get(source_col))
        parsed.append(item)
    return parsed


def parse_generic(path: Path) -> list[dict[str, Any]]:
    rd = report_date(path)
    header, rows = load_rows(path)
    export_name = path.stem
    parsed = []
    current_property_name: str | None = None
    current_mapping: dict[str, Any] = {"property_id": None, "community_id": None}
    for row_number, row in rows:
        property_name = row.get("Property") or row.get("Property Hierarchy - Property")
        if property_name:
            current_property_name = str(property_name)
            current_mapping = map_property(str(property_name))
        elif row.get("Current Level") == "Bedrooms" and current_property_name:
            property_name = current_property_name
        row_label = row.get("Mktg Source - Mktg Src Grp") or row.get("marketing_src_desc") or row.get("Cnc Den Reason") or property_name
        mapped = map_property(str(property_name)) if property_name else current_mapping if row.get("Current Level") == "Bedrooms" else {"property_id": None, "community_id": None}
        metrics = {key: value for key, value in row.items() if key not in {"Scope", "Region", "Property", "Property Hierarchy - Property"}}
        parsed.append(
            {
                "id": f"excel_export_{rd}_{slug(export_name)}_{row_number}",
                "report_date": rd,
                "export_name": export_name,
                "source_file": str(path.resolve()),
                "row_number": row_number,
                "scope": row.get("Scope"),
                "region": row.get("Region") or row.get("Property Hierarchy - Region"),
                "property_name": property_name,
                "property_id": mapped.get("property_id"),
                "community_id": mapped.get("community_id"),
                "row_label": row_label,
                "metrics_json": json.dumps(metrics, sort_keys=True, default=str),
                "evidence_json": json.dumps({"header": header, "raw": row, "mapping": mapped}, sort_keys=True, default=str),
            }
        )
    return parsed


def parse_conversion_dashboard(path: Path) -> list[dict[str, Any]]:
    rd = report_date(path)
    _, rows = load_rows(path)
    parsed = []
    current_contact_type = None
    for _, row in rows:
        if row.get("Init Contact Type"):
            current_contact_type = row.get("Init Contact Type")
        property_name = str(row.get("Property") or "").strip()
        if not property_name or property_name == "Total":
            continue
        mapped = map_property(property_name)
        parsed.append(
            {
                "id": f"conversion_dashboard_{rd}_{slug(str(current_contact_type or 'unknown'))}_{slug(property_name)}",
                "report_date": rd,
                "init_contact_type": current_contact_type,
                "property_name": property_name,
                "property_id": mapped.get("property_id"),
                "community_id": mapped.get("community_id"),
                "conversions": parse_int(row.get("Conversions")),
                "conversion_compare": parse_int(row.get("Conv Compare")),
                "conversion_delta": parse_number(row.get("Conversion \u0394")),
                "atr_avg": parse_number(row.get("ATR Avg")),
                "atr_delta": parse_number(row.get("ATR \u0394")),
                "source_file": str(path.resolve()),
                "evidence_json": json.dumps({"raw": row, "mapping": mapped}, sort_keys=True, default=str),
            }
        )
    return parsed


def first_present(row: dict[str, Any], names: list[str]) -> Any:
    for name in names:
        if name in row:
            return row.get(name)
    return None


def parse_vacancy_days(path: Path) -> list[dict[str, Any]]:
    rd = report_date(path)
    _, rows = load_rows(path)
    parsed = []
    metric_name = next(
        (
            metric
            for metric in ["VacDaysCurrMax", "VacDaysCurrAvg", "VacDaysHistMax", "VacDaysHistAvg"]
            if metric.lower() in path.stem.lower()
        ),
        "VacancyDays",
    )
    value_col = metric_name
    delta_col = f"{metric_name} \u0394"
    for _, row in rows:
        property_name = str(row.get("Property Hierarchy - Property") or "").strip()
        unit_number = str(row.get("Unit Number") or "").strip()
        if not property_name or not unit_number:
            continue
        mapped = map_property(property_name)
        parsed.append(
            {
                "id": f"vacancy_days_{rd}_{slug(metric_name)}_{slug(property_name)}_{slug(unit_number)}_{slug(path.stem)}",
                "report_date": rd,
                "metric_name": metric_name,
                "region": row.get("Property Hierarchy - Region"),
                "property_name": property_name,
                "property_id": mapped.get("property_id"),
                "community_id": mapped.get("community_id"),
                "bedroom_count": parse_int(row.get("Bedroom Ct")),
                "floorplan_desc": row.get("Floorplan Desc"),
                "unit_number": unit_number,
                "vacancy_days": parse_number(row.get(value_col)),
                "vacancy_days_delta": parse_number(row.get(delta_col)),
                "source_file": str(path.resolve()),
                "evidence_json": json.dumps({"raw": row, "mapping": mapped}, sort_keys=True, default=str),
            }
        )
    return parsed


def parse_lease_terms(path: Path) -> list[dict[str, Any]]:
    rd = report_date(path)
    _, rows = load_rows(path)
    parsed = []
    for idx, row in rows:
        property_name = str(row.get("Property Hierarchy - Property") or "").strip() or None
        mapped = map_property(property_name) if property_name else {"property_id": None, "community_id": None}
        lease_type = row.get("Lease Type")
        floorplan = row.get("Floorplan Desc")
        if not property_name and not lease_type and not floorplan:
            continue
        parsed.append(
            {
                "id": f"lease_terms_{rd}_{idx}_{slug(path.stem)}",
                "report_date": rd,
                "region": row.get("Property Hierarchy - Region"),
                "property_name": property_name,
                "property_id": mapped.get("property_id"),
                "community_id": mapped.get("community_id"),
                "lease_type": lease_type,
                "bedroom_count": parse_int(row.get("Bedroom Ct")),
                "floorplan_desc": floorplan,
                "lease_term_avg": parse_number(row.get("LeaseTermAvg")),
                "lease_term_avg_delta": parse_number(row.get("LeaseTermAvg \u0394")),
                "source_file": str(path.resolve()),
                "evidence_json": json.dumps({"raw": row, "mapping": mapped}, sort_keys=True, default=str),
            }
        )
    return parsed


def parse_wow_spending(path: Path) -> list[dict[str, Any]]:
    rd = report_date(path)
    _, rows = load_rows(path)
    parsed = []
    for _, row in rows:
        property_name = str(row.get("Location") or "").strip()
        if not property_name:
            continue
        mapped = map_property(property_name)
        parsed.append(
            {
                "id": f"wow_spending_{rd}_{slug(property_name)}",
                "report_date": rd,
                "region": row.get("Encasa Region"),
                "property_name": property_name,
                "property_id": mapped.get("property_id"),
                "community_id": mapped.get("community_id"),
                "apartment_units": parse_int(row.get("Apt Units")),
                "acquired_date": iso_date(row.get("Acq Date")),
                "spent": parse_number(row.get("Spent")),
                "budgeted": parse_number(row.get("Budgeted")),
                "budget_per_unit_month": parse_number(row.get("Budget per Unit-Mo")),
                "left_to_spend": parse_number(row.get("Left to Spend")),
                "pct_remaining": parse_number(row.get("Pct Remaining")),
                "source_file": str(path.resolve()),
                "evidence_json": json.dumps({"raw": row, "mapping": mapped}, sort_keys=True, default=str),
            }
        )
    return parsed


def parse_ad_spend_performance(path: Path) -> list[dict[str, Any]]:
    rd = report_date(path)
    _, rows = load_rows(path)
    parsed = []
    for _, row in rows:
        property_name = str(row.get("Property Hierarchy - Property") or "").strip()
        month = iso_date(row.get("Calendar Period (bins)"))
        if not property_name or not month:
            continue
        mapped = map_property(property_name)
        parsed.append(
            {
                "id": f"ad_spend_performance_{rd}_{month}_{slug(property_name)}",
                "report_date": rd,
                "calendar_month": month,
                "region": row.get("Property Hierarchy - Region"),
                "property_name": property_name,
                "property_id": mapped.get("property_id"),
                "community_id": mapped.get("community_id"),
                "guest_cards": parse_int(row.get("GCards")),
                "visits": parse_int(row.get("Visits")),
                "leases": parse_int(row.get("Leases")),
                "ad_spend_total": parse_number(row.get("Ad Spend Total")),
                "ad_spend_delta": parse_number(row.get("Ad Spend \u0394")),
                "source_file": str(path.resolve()),
                "evidence_json": json.dumps({"raw": row, "mapping": mapped}, sort_keys=True, default=str),
            }
        )
    return parsed


def parse_available_unit_interest_excel(path: Path) -> list[dict[str, Any]]:
    rd = report_date(path)
    _, rows = load_rows(path)
    parsed = []
    current_property_name: str | None = None
    current_mapping: dict[str, Any] = {"property_id": None, "community_id": None}
    for row_number, row in rows:
        current_level = str(row.get("Current Level") or "").strip()
        raw_property = str(row.get("Property") or "").strip()
        bedrooms = row.get("Bedrooms")
        bedrooms_text = str(bedrooms).strip() if bedrooms is not None and str(bedrooms).strip() else None
        location = str(row.get("Property") or row.get("Location") or "").strip()
        if not location or location.startswith("Applied filters"):
            if current_level.lower() == "bedrooms" and current_property_name and bedrooms_text:
                location = f"{current_property_name} - Bedroom {bedrooms_text}"
            else:
                continue
        property_name = location if current_level.lower() == "property" else None
        if property_name:
            current_property_name = property_name
            current_mapping = map_property(property_name)
        mapped = (
            map_property(property_name)
            if property_name
            else current_mapping
            if current_level.lower() == "bedrooms" and current_property_name
            else {"property_id": None, "community_id": None}
        )
        parsed.append(
            {
                "id": f"available_interest_{rd}_{slug(location)}_{slug(current_level)}_{row_number}",
                "report_date": rd,
                "location": location,
                "current_level": current_level,
                "bedrooms": bedrooms_text,
                "property_id": mapped.get("property_id"),
                "community_id": mapped.get("community_id"),
                "unit_count": parse_int(row.get("Unit Count")),
                "available_units": parse_int(row.get("Available Units")),
                "pct_available_units_by_bedroom": parse_number(row.get("% Available Units by Bedroom")),
                "vacant_available_units": parse_int(row.get("Vacant Available Units")),
                "notice_available_units": parse_int(row.get("Notice Available Units")),
                "pct_available": parse_number(row.get("% Available")),
                "t7_guest_cards_vol": parse_int(row.get("T7D Guest Cards Vol")),
                "t7_guest_cards_per_available_unit": parse_number(row.get("T7D Guest Cards per Available Unit")),
                "t7_guest_cards_delta_pct": parse_number(row.get("T7D v Prev T7D Day GC Vol. Delta %")),
                "pct_t7_guest_cards_by_bedrooms": parse_number(row.get("% of T7D GCs by Bedrooms")),
                "t30_guest_cards_vol": parse_int(row.get("T30D Guest Cards Vol")),
                "t30_guest_cards_per_available_unit": parse_number(row.get("T30D Guest Cards per Available Unit")),
                "t30_guest_cards_delta_pct": parse_number(row.get("T30D v Prev T30D Day GC Vol. Delta %")),
                "pct_t30_guest_cards_by_bedrooms": parse_number(row.get("% of T30D GCs by Bedrooms")),
                "t7_prospect_quote_vol": parse_int(row.get("T7D Prospect Quote Vol")),
                "t7_quote_delta_pct": parse_number(row.get("T7D v Prev T7D Quote Delta %")),
                "t30_prospect_quote_vol": parse_int(row.get("T30D Prospect Quote Vol")),
                "t30_quote_delta_pct": parse_number(row.get("T30D v Prev T30D Quote Delta %")),
                "source_file": str(path.resolve()),
                "evidence_json": json.dumps({"raw": row, "mapping": mapped}, sort_keys=True, default=str),
            }
        )
    return parsed


def parse_period_metrics(path: Path, metric_set: str) -> list[dict[str, Any]]:
    rd = report_date(path)
    _, rows = load_rows(path)
    parsed = []
    for _, row in rows:
        month = iso_date(row.get("Calendar Period (bins)"))
        if not month:
            continue
        metrics = {key: value for key, value in row.items() if key != "Calendar Period (bins)"}
        parsed.append(
            {
                "id": f"period_metric_{rd}_{slug(metric_set)}_{month}",
                "report_date": rd,
                "metric_set": metric_set,
                "calendar_month": month,
                "metric_json": json.dumps(metrics, sort_keys=True, default=str),
                "source_file": str(path.resolve()),
                "evidence_json": json.dumps({"raw": row}, sort_keys=True, default=str),
            }
        )
    return parsed


def parse_cost_per_conversion(path: Path) -> list[dict[str, Any]]:
    rd = report_date(path)
    _, rows = load_rows(path)
    parsed = []
    for row_number, row in rows:
        property_name = row.get("Property Hierarchy - Property")
        mapped = map_property(str(property_name)) if property_name else {"property_id": None, "community_id": None}
        invalid_count = 0
        for column in ["Cost per GC", "Cost per Visit", "Cost per App", "Cost per Lease"]:
            if column in row and row.get(column) is None:
                invalid_count += 1
        month = iso_date(row.get("Calendar Period (bins)"))
        parsed.append(
            {
                "id": f"cost_per_conversion_{rd}_{slug(path.stem)}_{row_number}",
                "report_date": rd,
                "calendar_month": month,
                "region": row.get("Property Hierarchy - Region"),
                "property_name": property_name,
                "property_id": mapped.get("property_id"),
                "community_id": mapped.get("community_id"),
                "marketing_source_group": row.get("Mktg Source - Mktg Src Grp"),
                "marketing_source_desc": row.get("Mktg Source - Mktg Src Desc"),
                "cost_per_guest_card": parse_number(row.get("Cost per GC")),
                "cost_per_visit": parse_number(row.get("Cost per Visit")),
                "cost_per_application": parse_number(row.get("Cost per App")),
                "cost_per_lease": parse_number(row.get("Cost per Lease")),
                "invalid_value_count": invalid_count,
                "source_file": str(path.resolve()),
                "evidence_json": json.dumps({"raw": row, "mapping": mapped}, sort_keys=True, default=str),
            }
        )
    return parsed


def source_kind_for_row(row: dict[str, Any]) -> str:
    if row.get("Mktg Source - Mktg Src Grp") or row.get("Mktg Source - Mktg Src Desc"):
        return "marketing_source"
    if row.get("Conv Source - Conv Src Grp") or row.get("Conv Source - Conv Src Desc"):
        return "conversion_source"
    if row.get("Origin"):
        return "origin"
    return "property_total"


def parse_source_performance(path: Path) -> list[dict[str, Any]]:
    rd = report_date(path)
    _, rows = load_rows(path)
    parsed = []
    current_region: str | None = None
    current_property_name: str | None = None
    current_scope: str | None = None
    source_identity = source_performance_identity_from_filename(path)
    for row_number, row in rows:
        raw_scope = row.get("Scope")
        if raw_scope:
            current_scope = raw_scope
        raw_region = row.get("Region") or row.get("Property Hierarchy - Region")
        raw_property = row.get("Property") or row.get("Property Hierarchy - Property")
        if raw_region:
            current_region = raw_region
        if raw_property and raw_property != "Total":
            current_property_name = raw_property
        elif raw_property == "Total":
            current_property_name = None
        property_name = raw_property or current_property_name
        if not property_name and source_identity and current_scope == "Selection":
            property_name = source_identity.property_name
        if property_name == "Total":
            property_name = None
        mapped = map_property(str(property_name)) if property_name else {"property_id": None, "community_id": None}
        kind = source_kind_for_row(row)
        source_group = (
            row.get("Mktg Source - Mktg Src Grp")
            or row.get("Conv Source - Conv Src Grp")
            or row.get("Origin")
            or row.get("Conv Source - Conv Src Desc")
            or row.get("Mktg Source - Mktg Src Desc")
        )
        source_desc = (
            row.get("Mktg Source - Mktg Src Desc")
            or row.get("Conv Source - Conv Src Desc")
            or row.get("Origin")
            or source_group
        )
        if not any(row.get(column) is not None for column in ["GC Incl. Zero", "GCards", "Visits", "Apps", "Leases", "Move-Ins"]):
            continue
        parsed.append(
            {
                "id": f"source_perf_{rd}_{slug(path.stem)}_{row_number}",
                "report_date": rd,
                "source_file": str(path.resolve()),
                "export_name": path.stem,
                "row_number": row_number,
                "scope": raw_scope or current_scope,
                "region": raw_region or current_region,
                "property_name": property_name,
                "property_id": mapped.get("property_id"),
                "community_id": mapped.get("community_id"),
                "origin": row.get("Origin"),
                "source_kind": kind,
                "source_group": source_group,
                "source_desc": source_desc,
                "guest_cards": parse_int(row.get("GC Incl. Zero") if "GC Incl. Zero" in row else row.get("GCards")),
                "visits": parse_int(row.get("Visits")),
                "first_tours": parse_int(row.get("First Tours")),
                "applications": parse_int(row.get("Apps")),
                "leases": parse_int(row.get("Leases")),
                "cancel_denials": parse_int(row.get("C&Ds")),
                "move_ins": parse_int(row.get("Move-Ins")),
                "visit_guest_card_conversion": parse_number(row.get("V/GC Conv")),
                "app_guest_card_conversion": parse_number(row.get("A/GC Conv")),
                "lease_guest_card_conversion": parse_number(row.get("L/GC Conv")),
                "lease_visit_ratio": parse_number(row.get("L/V Ratio")),
                "cancel_denial_pct_of_guest_cards": parse_number(row.get("C&D Pct of GCs")),
                "move_in_guest_card_conversion": parse_number(row.get("MI/GC Conv")),
                "move_in_visit_ratio": parse_number(row.get("MI/V Ratio")),
                "guest_cards_delta": parse_number(row.get("GCards \u0394")),
                "visits_delta": parse_number(row.get("Visits \u0394")),
                "applications_delta": parse_number(row.get("Apps \u0394")),
                "leases_delta": parse_number(row.get("Leases \u0394")),
                "cancel_denials_delta": parse_number(row.get("C&Ds \u0394")),
                "move_ins_delta": parse_number(row.get("Move-Ins \u0394")),
                "visit_guest_card_conversion_delta": parse_number(row.get("V/GC Conv \u0394")),
                "app_guest_card_conversion_delta": parse_number(row.get("A/GC Conv \u0394")),
                "lease_guest_card_conversion_delta": parse_number(row.get("L/GC Conv \u0394")),
                "lease_visit_ratio_delta": parse_number(row.get("L/V Ratio \u0394")),
                "cancel_denial_pct_delta": parse_number(row.get("C&D Pct of GCs \u0394")),
                "move_in_guest_card_conversion_delta": parse_number(row.get("MI/GC Conv \u0394")),
                "move_in_visit_ratio_delta": parse_number(row.get("MI/V Ratio \u0394")),
                "evidence_json": json.dumps({"raw": row, "mapping": mapped}, sort_keys=True, default=str),
            }
        )
    return parsed


def parse_move_ins_by_source(path: Path) -> list[dict[str, Any]]:
    rd = report_date(path)
    _, rows = load_rows(path)
    parsed = []
    for row_number, row in rows:
        property_name = str(row.get("Property") or "").strip()
        if not property_name:
            continue
        mapped = map_property(property_name)
        parsed.append(
            {
                "id": f"move_ins_source_{rd}_{slug(path.stem)}_{row_number}",
                "report_date": rd,
                "property_name": property_name,
                "property_id": mapped.get("property_id"),
                "community_id": mapped.get("community_id"),
                "unit": row.get("Unit"),
                "unit_move_in_date": iso_date(row.get("UnitMoveInDt")),
                "reservation_status": row.get("ResyStatus"),
                "resident_type": row.get("ResType"),
                "resident_move_in_date": iso_date(row.get("ResMoveInDt")),
                "resident_move_out_date": iso_date(row.get("ResMoveOutDt")),
                "marketing_source": row.get("MktgSrc"),
                "conversion_source": row.get("ConvSrc"),
                "source_file": str(path.resolve()),
                "row_number": row_number,
                "evidence_json": json.dumps(
                    {
                        "mapping": mapped,
                        "pii_note": "Resident names were intentionally not stored in this source route.",
                    },
                    sort_keys=True,
                    default=str,
                ),
            }
        )
    return parsed


MONTHLY_AD_SPEND_SOURCE_COLUMNS = {
    1: {
        "ADC Jan 2026": "ADC",
        "Kurie/AptGeo fencing Jan 2026": "Kurie/AptGeo fencing",
        "ZIL Jan 2026": "Zillow",
        "GOA Jan 2026": "Google Ads",
        "Social Media Jan 2026": "Social Media",
        "total": "TOTAL Jan 2026",
        "budget": "Jan 2026 Budget",
        "delta": "Jan 2026 Actual v Budget \u0394",
    },
    2: {
        "ADC Feb 2026": "ADC",
        "Kurie/AptGeo fencing Feb 2026": "Kurie/AptGeo fencing",
        "ZIL Feb 2026": "Zillow",
        "GOA Feb 2026": "Google Ads",
        "Social media Feb 2026": "Social Media",
        "total": "Feb 2026 total",
        "budget": "Feb 2026 Budget",
        "delta": "Feb 2026 Actual v Budget \u0394",
    },
    3: {
        "ADC March 2026": "ADC",
        "Kurie/AptGeo fencing March 2026": "Kurie/AptGeo fencing",
        "ZIL March 2026": "Zillow",
        "GOA March 2026": "Google Ads",
        "Forthea Fees March 2026": "Forthea Fees",
        "Social Media March 2026": "Social Media",
        "total": "March 2026 total",
        "budget": "March 2026 Budget",
        "delta": "March 2026 Actual v Budget \u0394",
    },
    4: {
        "ADC April 2026": "ADC",
        "ZIL April 2026": "Zillow",
        "GOA April 2026": "Google Ads",
        "Forthea Fees April 2026": "Forthea Fees",
        "Social April 2026": "Social Media",
        "total": "TOTAL April 2026",
        "budget": "April 2026 Budget",
        "delta": "April 2026 Actual v Budget \u0394",
    },
    5: {
        "ADC May 2026": "ADC",
        "ZIL May 2026": "Zillow",
        "GOA May 2026": "Google Ads",
        "Forthea Fees May 2026": "Forthea Fees",
        "Social May 2026": "Social Media",
        "total": "Total May 2026",
        "budget": "May 2026 Budget",
        "delta": "May 2026 Actual v Budget \u0394",
    },
}


def parse_monthly_ad_spend_source(path: Path) -> list[dict[str, Any]]:
    rd = report_date(path)
    _, rows = load_rows_with_header_marker(path, "Prop Code", "2026 Monthly Breakdown")
    parsed = []
    for row_number, row in rows:
        property_name = str(row.get("Property") or "").strip()
        property_code = str(row.get("Prop Code") or "").strip() or None
        if not property_name or property_name.startswith("Applied filters"):
            continue
        mapped = map_property(property_code or property_name)
        for month_number, config in MONTHLY_AD_SPEND_SOURCE_COLUMNS.items():
            month = f"2026-{month_number:02d}-01"
            month_total = parse_number(row.get(config["total"]))
            month_budget = parse_number(row.get(config["budget"]))
            month_delta = parse_number(row.get(config["delta"]))
            for column, source_group in config.items():
                if column in {"total", "budget", "delta"}:
                    continue
                spend = parse_number(row.get(column))
                if spend is None:
                    continue
                parsed.append(
                    {
                        "id": f"monthly_ad_spend_source_{rd}_{slug(str(property_code or property_name))}_{month}_{slug(source_group)}",
                        "report_date": rd,
                        "property_id": mapped.get("property_id"),
                        "community_id": mapped.get("community_id"),
                        "property_code": property_code,
                        "property_name": property_name,
                        "property_short_name": row.get("Property short name"),
                        "region": row.get("Region"),
                        "market": row.get("Market"),
                        "calendar_month": month,
                        "source_group": source_group,
                        "ad_spend_total": spend,
                        "month_total": month_total,
                        "month_budget": month_budget,
                        "month_actual_vs_budget_delta": month_delta,
                        "annual_budget": parse_number(row.get("2026 Budget")),
                        "annual_spend_trend_to_date": parse_number(row.get("2026 Spend Trend to Date")),
                        "annual_trend_delta_to_date": parse_number(row.get("2026 Trend \u0394 to Date")),
                        "source_file": str(path.resolve()),
                        "evidence_json": json.dumps(
                            {
                                "raw_property_code": property_code,
                                "raw_property": property_name,
                                "raw_column": column,
                                "mapping": mapped,
                            },
                            sort_keys=True,
                            default=str,
                        ),
                    }
                )
    return parsed


BOX_SCORE_COLUMN_MAP = {
    4: ("units", parse_int),
    5: ("physical_occupancy_pct", parse_number),
    6: ("occupancy_30_pct", parse_number),
    7: ("occupancy_60_pct", parse_number),
    8: ("adjusted_physical_occupancy_pct", parse_number),
    9: ("adjusted_occupancy_30_pct", parse_number),
    10: ("adjusted_occupancy_60_pct", parse_number),
    11: ("current_occupancy_rent", parse_number),
    12: ("net_effective_rent", parse_number),
    13: ("co_vs_ner_pct", parse_number),
    14: ("rent_actual_effective_yoy_pct", parse_number),
    15: ("concession_months_t90", parse_number),
    16: ("concession_months_91_180", parse_number),
    17: ("concession_frequency_t90_pct", parse_number),
    18: ("concession_frequency_91_180_pct", parse_number),
    19: ("guest_cards_t30", parse_int),
    20: ("guest_cards_t30_vs_ly_pct", parse_number),
    21: ("guest_cards_t7", parse_int),
    22: ("guest_cards_t7_vs_ly_pct", parse_number),
    23: ("onsite_tours_t30", parse_int),
    24: ("onsite_tours_t30_vs_ly_pct", parse_number),
    25: ("onsite_tours_t7", parse_int),
    26: ("onsite_tours_t7_vs_ly_pct", parse_number),
    27: ("tour_guest_card_t30_pct", parse_number),
    28: ("application_tour_t30_pct", parse_number),
    29: ("application_guest_card_t30_pct", parse_number),
    30: ("lease_application_t30_pct", parse_number),
    31: ("applications", parse_int),
    32: ("cancel_denials", parse_int),
    33: ("ntv", parse_int),
    34: ("coverage", parse_int),
    35: ("ready_available", parse_int),
    36: ("make_ready_pct", parse_number),
    37: ("turn_time_t30", parse_number),
    38: ("deferred_t30_pct", parse_number),
    39: ("service_requests_t7", parse_int),
    40: ("service_requests_t7_units_pct", parse_number),
    41: ("kingsley_os_t90_pct", parse_number),
    42: ("kingsley_os_t90_vs_ly_pct", parse_number),
    43: ("kingsley_nps_t90_pct", parse_number),
    44: ("kingsley_nps_t90_vs_ly_pct", parse_number),
}


def parse_portfolio_box_score(path: Path) -> list[dict[str, Any]]:
    rd = report_date(path)
    wb = load_workbook(str(path), read_only=True, data_only=True)
    ws = wb["Portfolio Box Score"] if "Portfolio Box Score" in wb.sheetnames else wb.worksheets[0]
    parsed = []
    for row in ws.iter_rows(values_only=True):
        property_name = row[2] if len(row) >= 3 else None
        if not property_name or property_name in {"Property Name", "Region\n"}:
            continue
        if not isinstance(property_name, str):
            continue
        mapped = map_property(property_name)
        if not mapped.get("property_id"):
            continue
        item = {
            "id": f"portfolio_box_score_{rd}_{slug(property_name)}",
            "report_date": rd,
            "property_name": property_name,
            "property_id": mapped.get("property_id"),
            "community_id": mapped.get("community_id"),
            "source_file": str(path.resolve()),
            "evidence_json": json.dumps({"mapping": mapped}, sort_keys=True, default=str),
        }
        for one_based_index, (column_name, parser) in BOX_SCORE_COLUMN_MAP.items():
            value = row[one_based_index - 1] if len(row) >= one_based_index else None
            item[column_name] = parser(value)
        parsed.append(item)
    return parsed


def parse_service_delivery(path: Path) -> list[dict[str, Any]]:
    rd = report_date(path)
    _, rows = load_rows(path)
    parsed = []
    for _, row in rows:
        property_name = str(row.get("Property") or "").strip()
        if not property_name or property_name.startswith("Applied filters"):
            continue
        mapped = map_property(property_name)
        parsed.append(
            {
                "id": f"service_delivery_{rd}_{slug(property_name)}",
                "report_date": rd,
                "property_name": property_name,
                "property_id": mapped.get("property_id"),
                "community_id": mapped.get("community_id"),
                "no_response_count": parse_int(row.get("# No Response")),
                "no_response_aging": parse_int(row.get("No Response Aging")),
                "unresolved_volume_48h_plus": parse_int(row.get("Unresolved Volume 48h+")),
                "first_response_vs_region": row.get("First Response vs Region"),
                "first_response": parse_int(row.get("First Response")),
                "total_resolution": parse_int(row.get("Total Resolution")),
                "reopen_rate_vs_portfolio": row.get("Reopen Rate vs Portfolio"),
                "ticket_volume_vs_portfolio": row.get("Ticket Volume vs Portfolio"),
                "source_file": str(path.resolve()),
                "evidence_json": json.dumps({"raw": row, "mapping": mapped}, sort_keys=True, default=str),
            }
        )
    return parsed


def parse_abandoned_applications(path: Path) -> list[dict[str, Any]]:
    rd = report_date(path)
    _, rows = load_rows(path)
    parsed = []
    for row_number, row in rows:
        unit_code = row.get("Unit Code")
        if not unit_code:
            continue
        parsed.append(
            {
                "id": f"abandoned_application_{rd}_{slug(path.stem)}_{row_number}",
                "report_date": rd,
                "property_name": None,
                "property_id": None,
                "community_id": None,
                "unit_code": unit_code,
                "count_value": parse_int(row.get("Ct")),
                "beds_baths": row.get("Beds-baths"),
                "floorplan_type": row.get("Type"),
                "total_units": parse_int(row.get("Total Units")),
                "available_units": parse_int(row.get("Available Units")),
                "sqft": parse_int(row.get("SqFt")),
                "applied_date": iso_date(row.get("Applied")),
                "lease_type": row.get("Lease Type"),
                "lease_term": parse_int(row.get("Lease Term")),
                "contract_created_date": iso_date(row.get("Contract Created")),
                "lease_start_date": iso_date(row.get("Lease Start")),
                "expected_move_in_date": iso_date(row.get("Expected Move-in")),
                "lease_end_date": iso_date(row.get("Lease End")),
                "net_rent": parse_number(row.get("Net Rent")),
                "source_file": str(path.resolve()),
                "row_number": row_number,
                "evidence_json": json.dumps(
                    {
                        "source_note": "Source export has no property column; primary resident names intentionally not stored.",
                    },
                    sort_keys=True,
                ),
            }
        )
    return parsed


def ingest_excel_exports(sources: list[Path], db_path: Path = DB_PATH) -> ExcelIngestResult:
    result = ExcelIngestResult()
    with sqlite3.connect(str(db_path)) as conn:
        ensure_tables(conn)
        for source in sources:
            path = source.expanduser().resolve()
            if not path.exists():
                result.files_failed += 1
                result.errors.append(f"{path}: not found")
                continue
            try:
                name = normalized_source_name(path)
                if name == "portfolio summary.xlsx":
                    rows = parse_portfolio_summary(path)
                    result.portfolio_summary_rows += upsert_rows(conn, "marketing_bi_portfolio_summary", rows)
                elif name.startswith("ad spend total, gcards"):
                    rows = parse_ad_spend_performance(path)
                    result.ad_spend_performance_rows += upsert_rows(
                        conn,
                        "marketing_bi_ad_spend_performance_month",
                        rows,
                    )
                elif name in {
                    "available units.xlsx",
                    "guest-cards-per-unit.xlsx",
                    "guest cards per unit.xlsx",
                    "guest cards per unit type.xlsx",
                    "guset cards per unit type.xlsx",
                }:
                    rows = parse_available_unit_interest_excel(path)
                    result.available_unit_interest_rows += upsert_rows(conn, "available_unit_interest_metrics", rows)
                elif name.startswith("ad spend total"):
                    rows = parse_ad_spend(path)
                    result.ad_spend_rows += upsert_rows(conn, "marketing_bi_ad_spend_property_month", rows)
                elif name in {"traffic conversions.xlsx", "traffic-conversions.xlsx"}:
                    rows = parse_traffic_conversions(path)
                    result.traffic_conversion_rows += upsert_rows(conn, "marketing_bi_traffic_conversions_full", rows)
                elif name in CONVERSION_DASHBOARD_FILENAMES:
                    rows = parse_conversion_dashboard(path)
                    conn.execute(
                        "DELETE FROM marketing_bi_conversion_dashboard_rows WHERE source_file = ?",
                        (str(path.resolve()),),
                    )
                    result.conversion_dashboard_rows += upsert_rows(conn, "marketing_bi_conversion_dashboard_rows", rows)
                elif name.startswith("vacdays"):
                    rows = parse_vacancy_days(path)
                    result.vacancy_days_rows += upsert_rows(conn, "marketing_bi_vacancy_days_units", rows)
                elif name.startswith("leasetermavg"):
                    rows = parse_lease_terms(path)
                    result.lease_term_rows += upsert_rows(conn, "marketing_bi_lease_term_rows", rows)
                elif name == "wow spending.xlsx":
                    rows = parse_wow_spending(path)
                    result.wow_spending_rows += upsert_rows(conn, "marketing_bi_wow_spending", rows)
                elif name == "unfulfilled lease totals  percentage of move-outs.xlsx":
                    rows = parse_period_metrics(path, "unfulfilled_leases")
                    result.period_leakage_rows += upsert_rows(conn, "marketing_bi_period_leakage_metrics", rows)
                elif name == "skips and evictions.xlsx":
                    rows = parse_period_metrics(path, "skips_evictions")
                    result.period_leakage_rows += upsert_rows(conn, "marketing_bi_period_leakage_metrics", rows)
                elif name == "30-day live-it love-it guarantee.xlsx":
                    rows = parse_period_metrics(path, "live_it_love_it_guarantee")
                    result.period_leakage_rows += upsert_rows(conn, "marketing_bi_period_leakage_metrics", rows)
                elif name.startswith("cost per conversion"):
                    rows = parse_cost_per_conversion(path)
                    result.cost_per_conversion_rows += upsert_rows(
                        conn,
                        "marketing_bi_cost_per_conversion_rows",
                        rows,
                    )
                elif name == "t365.xlsx":
                    rows = parse_move_ins_by_source(path)
                    result.move_ins_by_source_rows += upsert_rows(
                        conn,
                        "marketing_bi_move_ins_by_source_rows",
                        rows,
                    )
                elif name == "month by month adv spend per property.xlsx":
                    rows = parse_monthly_ad_spend_source(path)
                    result.monthly_ad_spend_source_rows += upsert_rows(
                        conn,
                        "marketing_bi_monthly_ad_spend_source_rows",
                        rows,
                    )
                elif name.startswith("portfolio box score"):
                    rows = parse_portfolio_box_score(path)
                    result.portfolio_box_score_rows += upsert_rows(
                        conn,
                        "marketing_bi_portfolio_box_score_rows",
                        rows,
                    )
                elif name == "t90 service delivery.xlsx":
                    rows = parse_service_delivery(path)
                    result.service_delivery_rows += upsert_rows(
                        conn,
                        "marketing_bi_service_delivery_rows",
                        rows,
                    )
                elif name.startswith("abandoned"):
                    rows = parse_abandoned_applications(path)
                    result.abandoned_application_rows += upsert_rows(
                        conn,
                        "marketing_bi_abandoned_application_rows",
                        rows,
                    )
                source_header, _ = load_rows(path)
                if any(
                    column in source_header
                    for column in [
                        "GC Incl. Zero",
                        "GCards",
                        "Mktg Source - Mktg Src Grp",
                        "Conv Source - Conv Src Desc",
                    ]
                ):
                    rows = parse_source_performance(path)
                    result.source_performance_rows += upsert_rows(
                        conn,
                        "marketing_bi_source_performance_rows",
                        rows,
                    )
                generic_rows = parse_generic(path)
                result.generic_rows += upsert_rows(conn, "marketing_bi_excel_export_rows", generic_rows)
                result.files_processed += 1
                result.source_files.append(str(path))
            except Exception as exc:
                result.files_failed += 1
                result.errors.append(f"{path}: {exc}")
        conn.commit()
    return result


def main() -> None:
    parser = argparse.ArgumentParser(description="Ingest native Marketing BI Excel exports.")
    parser.add_argument("--source", action="append", help="Specific Excel source. May be passed more than once.")
    parser.add_argument("--db", default=str(DB_PATH))
    args = parser.parse_args()
    sources = [Path(value) for value in args.source] if args.source else DEFAULT_SOURCES
    result = ingest_excel_exports(sources, Path(args.db).expanduser().resolve())
    print(json.dumps(result.__dict__, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
