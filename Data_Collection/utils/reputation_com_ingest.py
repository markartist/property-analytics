#!/usr/bin/env python3
"""Ingest Reputation.com XLSX exports into the Data Pond."""

from __future__ import annotations

import argparse
import json
import re
import sqlite3
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

try:
    from Data_Collection.utils.property_identity import resolve_property_identity
except ModuleNotFoundError:
    import sys

    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
    from Data_Collection.utils.property_identity import resolve_property_identity

ROOT = Path("/Users/mark/Property_Analytics")
DB_PATH = ROOT / "data" / "portfolio_analytics.db"
DEFAULT_SOURCE_DIR = Path("/Users/mark/Downloads/Reputation")
MIGRATION_SQL = ROOT / "apps" / "api" / "migrations" / "0040_create_reputation_com_tables.sql"


@dataclass
class ReputationComIngestResult:
    files_processed: int = 0
    files_failed: int = 0
    location_leaderboard_rows: int = 0
    score_component_rows: int = 0
    score_time_series_rows: int = 0
    local_competition_rows: int = 0
    mapped_locations: set[str] = field(default_factory=set)
    unmapped_locations: set[str] = field(default_factory=set)
    source_files: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


def slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")


def parse_report_date(path: Path) -> str:
    match = re.search(r"(20\d{2})-(\d{2})-(\d{2})", path.name)
    if match:
        return f"{match.group(1)}-{match.group(2)}-{match.group(3)}"
    return datetime.fromtimestamp(path.stat().st_mtime).date().isoformat()


def normalize_reputation_location(value: Any) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    midpoint = len(text) // 2
    if len(text) % 2 == 0 and text[:midpoint] == text[midpoint:]:
        return text[:midpoint].strip()
    return text


def parse_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").replace("%", "")
    if not text:
        return None
    return float(text)


def parse_int(value: Any) -> int | None:
    parsed = parse_number(value)
    if parsed is None:
        return None
    return int(round(parsed))


def parse_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if re.fullmatch(r"\d+\.0", text):
        return text[:-2]
    return text


def map_property(raw_location: Any) -> dict[str, Any]:
    normalized = normalize_reputation_location(raw_location)
    if not normalized:
        return {"normalized_location": normalized, "property_id": None, "community_id": None, "match_source": "blank"}
    identity = resolve_property_identity(normalized)
    if identity:
        mapped = identity.as_mapping()
        mapped["normalized_location"] = normalized
        return mapped
    return {
        "normalized_location": normalized,
        "property_id": None,
        "community_id": None,
        "canonical_name": normalized,
        "match_source": "unmapped",
    }


def load_sheet_rows(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(str(path), read_only=False, data_only=True)
    worksheet = workbook.worksheets[0]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    parsed: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows[1:], 2):
        if not row or not any(value is not None for value in row):
            continue
        record = dict(zip(headers, row))
        record["_row_number"] = row_number
        parsed.append(record)
    return parsed


def ensure_tables(conn: sqlite3.Connection) -> None:
    conn.executescript(MIGRATION_SQL.read_text(encoding="utf-8"))


def track_mapping(result: ReputationComIngestResult, location: Any, property_id: str | None) -> None:
    normalized = normalize_reputation_location(location)
    if not normalized or normalized in {"Best In Class", "Industry Average"}:
        return
    if property_id:
        result.mapped_locations.add(normalized)
    else:
        result.unmapped_locations.add(normalized)


def ingest_location_leaderboard(conn: sqlite3.Connection, path: Path, rows: list[dict[str, Any]], result: ReputationComIngestResult) -> None:
    report_date = parse_report_date(path)
    payloads = []
    for row in rows:
        location = parse_text(row.get("Location"))
        if not location:
            continue
        mapped = map_property(location)
        track_mapping(result, location, mapped.get("property_id"))
        evidence = {"source": "Reputation.com Location Leaderboard", "row_number": row["_row_number"], "match_source": mapped.get("match_source")}
        payloads.append(
            (
                f"reputation_location_{report_date}_{slug(location)}_{slug(path.stem)}",
                report_date,
                location,
                mapped["normalized_location"],
                mapped.get("property_id"),
                mapped.get("community_id"),
                parse_int(row.get("Current Total Reviews")),
                parse_number(row.get("Current Positive Reviews (Percentage)")),
                parse_number(row.get("Current Neutral Reviews (Percentage)")),
                parse_number(row.get("Current Negative Reviews (Percentage)")),
                parse_number(row.get("Average Rating")),
                parse_number(row.get("Response Rate")),
                parse_number(row.get("Reputation Score")),
                parse_text(row.get("Current Period")),
                str(path.resolve()),
                json.dumps(evidence, sort_keys=True),
            )
        )
    conn.executemany(
        """
        INSERT INTO reputation_com_location_leaderboard (
          id, report_date, location, normalized_location, property_id, community_id,
          current_total_reviews, current_positive_reviews_pct, current_neutral_reviews_pct,
          current_negative_reviews_pct, average_rating, response_rate, reputation_score,
          current_period, source_file, evidence_json
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(report_date, location, source_file) DO UPDATE SET
          normalized_location=excluded.normalized_location,
          property_id=excluded.property_id,
          community_id=excluded.community_id,
          current_total_reviews=excluded.current_total_reviews,
          current_positive_reviews_pct=excluded.current_positive_reviews_pct,
          current_neutral_reviews_pct=excluded.current_neutral_reviews_pct,
          current_negative_reviews_pct=excluded.current_negative_reviews_pct,
          average_rating=excluded.average_rating,
          response_rate=excluded.response_rate,
          reputation_score=excluded.reputation_score,
          current_period=excluded.current_period,
          evidence_json=excluded.evidence_json,
          updated_at=datetime('now')
        """,
        payloads,
    )
    result.location_leaderboard_rows += len(payloads)


def ingest_score_components(conn: sqlite3.Connection, path: Path, rows: list[dict[str, Any]], result: ReputationComIngestResult) -> None:
    report_date = parse_report_date(path)
    payloads = []
    for row in rows:
        location = parse_text(row.get("Location"))
        if not location:
            continue
        entity_type = "benchmark" if location in {"Best In Class", "Industry Average"} else "property"
        mapped = map_property(location) if entity_type == "property" else {"normalized_location": location, "property_id": None, "community_id": None, "match_source": "benchmark"}
        track_mapping(result, location, mapped.get("property_id"))
        evidence = {"source": "Reputation.com Reputation Score By Location", "row_number": row["_row_number"], "match_source": mapped.get("match_source")}
        payloads.append(
            (
                f"reputation_components_{report_date}_{slug(location)}_{slug(path.stem)}",
                report_date,
                location,
                mapped["normalized_location"],
                entity_type,
                mapped.get("property_id"),
                mapped.get("community_id"),
                parse_number(row.get("Reputation Score")),
                parse_number(row.get("Review Sentiment")),
                parse_number(row.get("Review Volume")),
                parse_number(row.get("Review Recency")),
                parse_number(row.get("Review Quality")),
                parse_number(row.get("Review Spread")),
                parse_number(row.get("Review Response")),
                parse_number(row.get("Search Impressions")),
                parse_number(row.get("Listing Completeness")),
                parse_number(row.get("Social Score")),
                str(path.resolve()),
                json.dumps(evidence, sort_keys=True),
            )
        )
    conn.executemany(
        """
        INSERT INTO reputation_com_score_components (
          id, report_date, location, normalized_location, entity_type, property_id,
          community_id, reputation_score, review_sentiment, review_volume,
          review_recency, review_quality, review_spread, review_response,
          search_impressions, listing_completeness, social_score, source_file, evidence_json
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(report_date, location, source_file) DO UPDATE SET
          normalized_location=excluded.normalized_location,
          entity_type=excluded.entity_type,
          property_id=excluded.property_id,
          community_id=excluded.community_id,
          reputation_score=excluded.reputation_score,
          review_sentiment=excluded.review_sentiment,
          review_volume=excluded.review_volume,
          review_recency=excluded.review_recency,
          review_quality=excluded.review_quality,
          review_spread=excluded.review_spread,
          review_response=excluded.review_response,
          search_impressions=excluded.search_impressions,
          listing_completeness=excluded.listing_completeness,
          social_score=excluded.social_score,
          evidence_json=excluded.evidence_json,
          updated_at=datetime('now')
        """,
        payloads,
    )
    result.score_component_rows += len(payloads)


def ingest_time_series(conn: sqlite3.Connection, path: Path, rows: list[dict[str, Any]], result: ReputationComIngestResult) -> None:
    report_date = parse_report_date(path)
    payloads = []
    for row in rows:
        location = parse_text(row.get("Location"))
        if not location:
            continue
        mapped = map_property(location)
        track_mapping(result, location, mapped.get("property_id"))
        for column, value in row.items():
            if column in {"Location", "_row_number"}:
                continue
            score = parse_number(value)
            if score is None:
                continue
            month = datetime.strptime(str(column), "%b %Y").strftime("%Y-%m")
            evidence = {"source": "Reputation.com Score Time Series By Location", "row_number": row["_row_number"], "match_source": mapped.get("match_source")}
            payloads.append(
                (
                    f"reputation_timeseries_{slug(location)}_{month}_{slug(path.stem)}",
                    report_date,
                    location,
                    mapped["normalized_location"],
                    mapped.get("property_id"),
                    mapped.get("community_id"),
                    month,
                    score,
                    str(path.resolve()),
                    json.dumps(evidence, sort_keys=True),
                )
            )
    conn.executemany(
        """
        INSERT INTO reputation_com_score_time_series (
          id, report_date, location, normalized_location, property_id, community_id,
          score_month, reputation_score, source_file, evidence_json
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(location, score_month, source_file) DO UPDATE SET
          report_date=excluded.report_date,
          normalized_location=excluded.normalized_location,
          property_id=excluded.property_id,
          community_id=excluded.community_id,
          reputation_score=excluded.reputation_score,
          evidence_json=excluded.evidence_json,
          updated_at=datetime('now')
        """,
        payloads,
    )
    result.score_time_series_rows += len(payloads)


def ingest_local_competition(conn: sqlite3.Connection, path: Path, rows: list[dict[str, Any]], result: ReputationComIngestResult) -> None:
    report_date = parse_report_date(path)
    payloads = []
    for row in rows:
        location = parse_text(row.get("Locations"))
        if not location:
            continue
        mapped = map_property(location)
        track_mapping(result, location, mapped.get("property_id"))
        for rank in (1, 2, 3):
            competitor_location = parse_text(row.get(f"Competitor {rank} Locations"))
            evidence = {"source": "Reputation.com Local Competition Leaderboard", "row_number": row["_row_number"], "match_source": mapped.get("match_source")}
            payloads.append(
                (
                    f"reputation_competition_{report_date}_{slug(location)}_{rank}_{slug(path.stem)}",
                    report_date,
                    location,
                    mapped["normalized_location"],
                    mapped.get("property_id"),
                    mapped.get("community_id"),
                    parse_int(row.get("Total Reviews")),
                    parse_int(row.get("Positive Reviews")),
                    parse_int(row.get("Neutral Reviews")),
                    parse_int(row.get("Negative Reviews")),
                    parse_number(row.get("Average Rating")),
                    parse_number(row.get("Reputation Score")),
                    rank,
                    competitor_location,
                    parse_int(row.get(f"Competitor {rank} Total Reviews")),
                    parse_int(row.get(f"Competitor {rank} Positive Reviews")),
                    parse_int(row.get(f"Competitor {rank} Neutral Reviews")),
                    parse_int(row.get(f"Competitor {rank} Negative Reviews")),
                    parse_number(row.get(f"Competitor {rank} Average Rating")),
                    parse_number(row.get(f"Competitor {rank} Reputation Score")),
                    str(path.resolve()),
                    json.dumps(evidence, sort_keys=True),
                )
            )
    conn.executemany(
        """
        INSERT INTO reputation_com_local_competition (
          id, report_date, location, normalized_location, property_id, community_id,
          subject_total_reviews, subject_positive_reviews, subject_neutral_reviews,
          subject_negative_reviews, subject_average_rating, subject_reputation_score,
          competitor_rank, competitor_location, competitor_total_reviews,
          competitor_positive_reviews, competitor_neutral_reviews, competitor_negative_reviews,
          competitor_average_rating, competitor_reputation_score, source_file, evidence_json
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(report_date, location, competitor_rank, source_file) DO UPDATE SET
          normalized_location=excluded.normalized_location,
          property_id=excluded.property_id,
          community_id=excluded.community_id,
          subject_total_reviews=excluded.subject_total_reviews,
          subject_positive_reviews=excluded.subject_positive_reviews,
          subject_neutral_reviews=excluded.subject_neutral_reviews,
          subject_negative_reviews=excluded.subject_negative_reviews,
          subject_average_rating=excluded.subject_average_rating,
          subject_reputation_score=excluded.subject_reputation_score,
          competitor_location=excluded.competitor_location,
          competitor_total_reviews=excluded.competitor_total_reviews,
          competitor_positive_reviews=excluded.competitor_positive_reviews,
          competitor_neutral_reviews=excluded.competitor_neutral_reviews,
          competitor_negative_reviews=excluded.competitor_negative_reviews,
          competitor_average_rating=excluded.competitor_average_rating,
          competitor_reputation_score=excluded.competitor_reputation_score,
          evidence_json=excluded.evidence_json,
          updated_at=datetime('now')
        """,
        payloads,
    )
    result.local_competition_rows += len(payloads)


def ingest_reputation_exports(source_dir: Path = DEFAULT_SOURCE_DIR, db_path: Path = DB_PATH) -> ReputationComIngestResult:
    result = ReputationComIngestResult()
    files = sorted(source_dir.glob("*.xlsx"))
    with sqlite3.connect(str(db_path)) as conn:
        ensure_tables(conn)
        for path in files:
            result.source_files.append(str(path.resolve()))
            try:
                rows = load_sheet_rows(path)
                name = path.name.lower()
                if "location-leaderboard" in name:
                    ingest_location_leaderboard(conn, path, rows, result)
                elif "reputation-score-by-location" in name:
                    ingest_score_components(conn, path, rows, result)
                elif "score-time-series-by-location" in name:
                    ingest_time_series(conn, path, rows, result)
                elif "local-competition-leaderboard" in name:
                    ingest_local_competition(conn, path, rows, result)
                else:
                    raise ValueError(f"Unrecognized Reputation.com export name: {path.name}")
                result.files_processed += 1
            except Exception as exc:  # noqa: BLE001
                result.files_failed += 1
                result.errors.append(f"{path.name}: {exc}")
        conn.commit()
    return result


def main() -> None:
    parser = argparse.ArgumentParser(description="Ingest Reputation.com XLSX exports into the Data Pond.")
    parser.add_argument("--source-dir", default=str(DEFAULT_SOURCE_DIR))
    parser.add_argument("--db", default=str(DB_PATH))
    args = parser.parse_args()

    result = ingest_reputation_exports(Path(args.source_dir).expanduser().resolve(), Path(args.db).expanduser().resolve())
    print(
        json.dumps(
            {
                "files_processed": result.files_processed,
                "files_failed": result.files_failed,
                "location_leaderboard_rows": result.location_leaderboard_rows,
                "score_component_rows": result.score_component_rows,
                "score_time_series_rows": result.score_time_series_rows,
                "local_competition_rows": result.local_competition_rows,
                "mapped_location_count": len(result.mapped_locations),
                "unmapped_locations": sorted(result.unmapped_locations),
                "source_files": result.source_files,
                "errors": result.errors,
            },
            indent=2,
            sort_keys=True,
        )
    )


if __name__ == "__main__":
    main()
