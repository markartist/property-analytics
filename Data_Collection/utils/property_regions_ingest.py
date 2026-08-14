#!/usr/bin/env python3
"""Ingest governed property-region assignments from the regions workbook."""

from __future__ import annotations

import argparse
import json
import sqlite3
import sys
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

UTC = timezone.utc

ROOT = Path("/Users/mark/Property_Analytics")
DB_PATH = ROOT / "data" / "portfolio_analytics.db"
DEFAULT_SOURCE = Path("/Users/mark/Downloads/regions.xlsx")
DEFAULT_REGISTRY = ROOT / "config" / "venterra_properties_official.json"

if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from Data_Collection.utils.property_identity import resolve_property_identity  # noqa: E402


@dataclass
class RegionAssignment:
    region: str
    source_property_name: str
    property_name: str
    property_id: str
    property_code: str | None
    ga4_property_id: str | None


@dataclass
class PropertyRegionsIngestResult:
    source_file: str
    assignments_found: int = 0
    assignments_mapped: int = 0
    database_rows_updated: int = 0
    registry_rows_updated: int = 0
    regions: list[str] = field(default_factory=list)
    unmapped_properties: list[dict[str, str]] = field(default_factory=list)


def clean_text(value: Any) -> str:
    return str(value or "").strip()


def load_region_assignments(source_path: Path, sheet_name: str | None = None) -> tuple[list[RegionAssignment], list[dict[str, str]]]:
    workbook = load_workbook(source_path, data_only=True, read_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    if hasattr(worksheet, "reset_dimensions"):
        worksheet.reset_dimensions()

    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return [], []

    headers = [clean_text(value) for value in rows[0]]
    try:
        region_index = headers.index("Region") + 1
        property_index = headers.index("Property") + 1
    except ValueError as exc:
        raise ValueError("Workbook must include Region and Property columns.") from exc

    assignments: list[RegionAssignment] = []
    unmapped: list[dict[str, str]] = []
    seen: set[str] = set()

    for row_number, row in enumerate(rows[1:], start=2):
        region = clean_text(row[region_index - 1] if len(row) >= region_index else None)
        source_property_name = clean_text(row[property_index - 1] if len(row) >= property_index else None)
        if not region or not source_property_name or source_property_name.lower() == "total":
            continue

        identity = resolve_property_identity(source_property_name)
        if not identity:
            unmapped.append(
                {
                    "row": str(row_number),
                    "region": region,
                    "property": source_property_name,
                }
            )
            continue

        if identity.canonical_property_id in seen:
            continue
        seen.add(identity.canonical_property_id)
        assignments.append(
            RegionAssignment(
                region=region,
                source_property_name=source_property_name,
                property_name=identity.property_name,
                property_id=identity.ga4_property_id or identity.canonical_property_id,
                property_code=identity.property_code,
                ga4_property_id=identity.ga4_property_id,
            )
        )

    return assignments, unmapped


def update_database(assignments: list[RegionAssignment], db_path: Path) -> int:
    updated = 0
    now = datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    with sqlite3.connect(str(db_path)) as conn:
        for assignment in assignments:
            cursor = conn.execute(
                """
                UPDATE properties
                SET encasa_region = ?, updated_at = ?
                WHERE property_id = ?
                  AND COALESCE(encasa_region, '') != ?
                """,
                (assignment.region, now, assignment.property_id, assignment.region),
            )
            updated += cursor.rowcount
    return updated


def update_registry(assignments: list[RegionAssignment], registry_path: Path) -> int:
    raw = json.loads(registry_path.read_text(encoding="utf-8"))
    properties = raw.get("properties", raw) if isinstance(raw, dict) else raw
    by_ga4 = {str(assignment.ga4_property_id): assignment for assignment in assignments if assignment.ga4_property_id}
    by_name = {assignment.property_name: assignment for assignment in assignments}

    updated = 0
    for row in properties:
        assignment = by_ga4.get(str(row.get("ga4_property_id"))) or by_name.get(str(row.get("name")))
        if not assignment:
            continue
        if row.get("encasa_region") != assignment.region:
            row["encasa_region"] = assignment.region
            updated += 1

    if isinstance(raw, dict):
        raw["_last_updated"] = datetime.now(UTC).date().isoformat()
        raw["_region_source"] = "regions.xlsx via Data_Collection/utils/property_regions_ingest.py"
    registry_path.write_text(json.dumps(raw, indent=2) + "\n", encoding="utf-8")
    return updated


def ingest_property_regions(
    source_path: Path = DEFAULT_SOURCE,
    db_path: Path = DB_PATH,
    registry_path: Path = DEFAULT_REGISTRY,
    sheet_name: str | None = None,
) -> PropertyRegionsIngestResult:
    assignments, unmapped = load_region_assignments(source_path, sheet_name=sheet_name)
    result = PropertyRegionsIngestResult(source_file=str(source_path))
    result.assignments_found = len(assignments) + len(unmapped)
    result.assignments_mapped = len(assignments)
    result.unmapped_properties = unmapped
    result.regions = sorted({assignment.region for assignment in assignments})

    if unmapped:
        return result

    result.database_rows_updated = update_database(assignments, db_path)
    result.registry_rows_updated = update_registry(assignments, registry_path)
    return result


def main() -> None:
    parser = argparse.ArgumentParser(description="Ingest property-region assignments into Data Pond config and DB.")
    parser.add_argument("--source", default=str(DEFAULT_SOURCE))
    parser.add_argument("--db", default=str(DB_PATH))
    parser.add_argument("--registry", default=str(DEFAULT_REGISTRY))
    parser.add_argument("--sheet")
    args = parser.parse_args()

    result = ingest_property_regions(
        source_path=Path(args.source),
        db_path=Path(args.db),
        registry_path=Path(args.registry),
        sheet_name=args.sheet,
    )
    print(json.dumps(result.__dict__, indent=2, sort_keys=True))
    if result.unmapped_properties:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
