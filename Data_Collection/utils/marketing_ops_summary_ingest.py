#!/usr/bin/env python3
"""Ingest the Marketing Ops Summary workbook into the Data Pond."""

from __future__ import annotations

import argparse
import json
import re
import sqlite3
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

try:
    from Data_Collection.utils.property_identity import resolve_property_identity
except ModuleNotFoundError:
    import sys

    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
    from Data_Collection.utils.property_identity import resolve_property_identity

ROOT = Path("/Users/mark/Property_Analytics")
DB_PATH = ROOT / "data" / "portfolio_analytics.db"
DEFAULT_SOURCE = Path("/Users/mark/Downloads/Marketing Ops Summary.xlsx")
MIGRATION_SQL = ROOT / "apps" / "api" / "migrations" / "0041_create_marketing_ops_summary.sql"

TABLE_COLUMNS = [
    "id",
    "report_date",
    "source_as_of_date",
    "region",
    "property_name",
    "property_id",
    "community_id",
    "units",
    "occupancy",
    "atr30",
    "atr",
    "leads_t30_py",
    "leads_t30",
    "leads_t30_var",
    "leads_t7_py",
    "leads_t7",
    "leads_t7_var",
    "leads_trend",
    "lead_to_visit_t30",
    "lead_to_visit_t7",
    "visits_t30_py",
    "visits_t30",
    "visits_t30_var",
    "visits_t7_py",
    "visits_t7",
    "visits_t7_var",
    "visits_trend",
    "leases_t7_py",
    "leases_t7",
    "traffic_per_unit",
    "benchmark_t30_pf",
    "benchmark_rank",
    "benchmark_var",
    "close_ratio",
    "projected_traffic",
    "projected_traffic_gap",
    "current_month_expirations",
    "forward_3_month_expirations",
    "forward_3_month_expirations_pct",
    "expirations_var",
    "new_lease_tradeoff_dollars",
    "new_lease_tradeoff_pct",
    "renewal_tradeoff_dollars",
    "renewal_tradeoff_pct",
    "offered_vs_book",
    "offered_60d_delta",
    "offered_30d_delta",
    "offered_7d_delta",
    "ad_spend_t3_property",
    "ad_spend_t3_portfolio",
    "ad_spend_rank",
    "ad_spend_var",
    "ad_spend_t1_budget",
    "ad_spend_t1_actual",
    "ad_spend_t1_var",
    "ad_spend_t3_budget",
    "ad_spend_t3_actual",
    "ad_spend_t3_var",
    "ad_spend_t1_vs_t3",
    "ad_spend_t3_vs_t6",
    "kingsley_find_property_avg_t7",
    "kingsley_find_property_avg_t30",
    "kingsley_find_property_py_avg_t30",
    "kingsley_deposit_yes_pct_t7",
    "kingsley_deposit_yes_pct_t30",
    "kingsley_tour_yes_pct_t7",
    "kingsley_tour_yes_pct_t30",
    "source_file",
    "evidence_json",
]

CELL_MAP = {
    "region": 2,
    "property_name": 3,
    "units": 4,
    "occupancy": 5,
    "atr30": 6,
    "atr": 7,
    "leads_t30_py": 8,
    "leads_t30": 10,
    "leads_t30_var": 11,
    "leads_t7_py": 12,
    "leads_t7": 13,
    "leads_t7_var": 14,
    "leads_trend": 15,
    "lead_to_visit_t30": 16,
    "lead_to_visit_t7": 17,
    "visits_t30_py": 18,
    "visits_t30": 19,
    "visits_t30_var": 20,
    "visits_t7_py": 21,
    "visits_t7": 22,
    "visits_t7_var": 23,
    "visits_trend": 24,
    "leases_t7_py": 25,
    "leases_t7": 26,
    "traffic_per_unit": 27,
    "benchmark_t30_pf": 28,
    "benchmark_rank": 29,
    "benchmark_var": 30,
    "close_ratio": 31,
    "projected_traffic": 32,
    "projected_traffic_gap": 33,
    "current_month_expirations": 34,
    "forward_3_month_expirations": 35,
    "forward_3_month_expirations_pct": 36,
    "expirations_var": 37,
    "new_lease_tradeoff_dollars": 38,
    "new_lease_tradeoff_pct": 39,
    "renewal_tradeoff_dollars": 40,
    "renewal_tradeoff_pct": 41,
    "offered_vs_book": 42,
    "offered_60d_delta": 43,
    "offered_30d_delta": 44,
    "offered_7d_delta": 45,
    "ad_spend_t3_property": 46,
    "ad_spend_t3_portfolio": 47,
    "ad_spend_rank": 48,
    "ad_spend_var": 49,
    "ad_spend_t1_budget": 50,
    "ad_spend_t1_actual": 51,
    "ad_spend_t1_var": 52,
    "ad_spend_t3_budget": 53,
    "ad_spend_t3_actual": 54,
    "ad_spend_t3_var": 55,
    "ad_spend_t1_vs_t3": 56,
    "ad_spend_t3_vs_t6": 57,
    "kingsley_find_property_avg_t7": 58,
    "kingsley_find_property_avg_t30": 59,
    "kingsley_find_property_py_avg_t30": 60,
    "kingsley_deposit_yes_pct_t7": 61,
    "kingsley_deposit_yes_pct_t30": 63,
    "kingsley_tour_yes_pct_t7": 64,
    "kingsley_tour_yes_pct_t30": 65,
}

INTEGER_FIELDS = {
    "units",
    "leads_t30_py",
    "leads_t30",
    "leads_t7_py",
    "leads_t7",
    "visits_t30_py",
    "visits_t30",
    "visits_t7_py",
    "visits_t7",
    "leases_t7_py",
    "leases_t7",
    "benchmark_rank",
    "current_month_expirations",
    "forward_3_month_expirations",
    "ad_spend_rank",
}


@dataclass
class MarketingOpsSummaryIngestResult:
    report_date: str | None = None
    source_as_of_date: str | None = None
    rows_found: int = 0
    rows_upserted: int = 0
    mapped_properties: int = 0
    unmapped_properties: list[str] = field(default_factory=list)
    source_file: str | None = None


def slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")


def parse_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "").replace("%", "")
    if not text:
        return None
    return float(text)


def parse_int(value: Any) -> int | None:
    parsed = parse_number(value)
    if parsed is None:
        return None
    return int(round(parsed))


def report_date(path: Path) -> str:
    return datetime.fromtimestamp(path.stat().st_mtime).date().isoformat()


def source_as_of_date(sheet_title: str | None, workbook_path: Path) -> str:
    text = sheet_title or ""
    match = re.search(r"as of (\d{1,2})/(\d{1,2})/(\d{4})", text)
    if match:
        month, day, year = (int(part) for part in match.groups())
        return f"{year:04d}-{month:02d}-{day:02d}"
    return report_date(workbook_path)


def row_id(report_date_value: str, property_name: str) -> str:
    return f"marketing_ops_summary_{report_date_value}_{slug(property_name)}"


def map_property(name: str) -> dict[str, Any]:
    identity = resolve_property_identity(name)
    if identity:
        return identity.as_mapping()
    return {"property_id": None, "community_id": None, "canonical_name": name, "match_source": "unmapped"}


def normalize_value(field: str, value: Any) -> int | float | str | None:
    if field in {"region", "property_name"}:
        return str(value).strip() if value is not None and str(value).strip() else None
    if field in INTEGER_FIELDS:
        return parse_int(value)
    return parse_number(value)


def parse_workbook(path: Path) -> tuple[str, str, list[dict[str, Any]]]:
    workbook = load_workbook(str(path), read_only=False, data_only=True)
    worksheet = workbook["Marketing Ops Summary"] if "Marketing Ops Summary" in workbook.sheetnames else workbook.worksheets[0]
    as_of = source_as_of_date(str(worksheet.cell(7, 8).value or ""), path)
    rd = report_date(path)
    rows: list[dict[str, Any]] = []
    for row_number in range(10, worksheet.max_row + 1):
        property_name = worksheet.cell(row_number, CELL_MAP["property_name"]).value
        region = worksheet.cell(row_number, CELL_MAP["region"]).value
        if not property_name or str(property_name).strip() in {"Property", "Total", "Totals", "Averages"}:
            continue
        if str(region or "").strip() in {"Totals", "Averages"}:
            continue
        if str(property_name).startswith("Applied filters"):
            continue

        row: dict[str, Any] = {}
        for field, column_number in CELL_MAP.items():
            row[field] = normalize_value(field, worksheet.cell(row_number, column_number).value)
        mapped = map_property(str(row["property_name"]))
        row.update(
            {
                "id": row_id(rd, str(row["property_name"])),
                "report_date": rd,
                "source_as_of_date": as_of,
                "property_id": mapped.get("property_id"),
                "community_id": mapped.get("community_id"),
                "source_file": str(path.resolve()),
                "evidence_json": json.dumps(
                    {
                        "row_number": row_number,
                        "mapping": mapped,
                        "source_groups": {
                            "property_performance": "Property Performance / Current",
                            "traffic": f"Traffic (as of {as_of})",
                            "pricing": f"Pricing (as of {as_of})",
                            "financial": f"Financial (as of {as_of})",
                            "kingsley": "Kingsley Data",
                        },
                    },
                    sort_keys=True,
                    default=str,
                ),
            }
        )
        rows.append(row)
    return rd, as_of, rows


def ensure_table(conn: sqlite3.Connection) -> None:
    conn.executescript(MIGRATION_SQL.read_text(encoding="utf-8"))


def upsert_rows(conn: sqlite3.Connection, rows: list[dict[str, Any]]) -> int:
    ensure_table(conn)
    placeholders = ", ".join("?" for _ in TABLE_COLUMNS)
    updates = ", ".join(f"{column}=excluded.{column}" for column in TABLE_COLUMNS if column != "id")
    conn.executemany(
        f"""
        INSERT INTO marketing_ops_summary_rows ({", ".join(TABLE_COLUMNS)})
        VALUES ({placeholders})
        ON CONFLICT(report_date, property_name, source_file)
        DO UPDATE SET {updates}, updated_at=datetime('now')
        """,
        [[row.get(column) for column in TABLE_COLUMNS] for row in rows],
    )
    conn.commit()
    return len(rows)


def ingest_marketing_ops_summary(path: Path = DEFAULT_SOURCE, db_path: Path = DB_PATH) -> MarketingOpsSummaryIngestResult:
    rd, as_of, rows = parse_workbook(path)
    with sqlite3.connect(str(db_path)) as conn:
        rows_upserted = upsert_rows(conn, rows)
    unmapped = sorted({str(row["property_name"]) for row in rows if not row.get("property_id")})
    return MarketingOpsSummaryIngestResult(
        report_date=rd,
        source_as_of_date=as_of,
        rows_found=len(rows),
        rows_upserted=rows_upserted,
        mapped_properties=len(rows) - len(unmapped),
        unmapped_properties=unmapped,
        source_file=str(path),
    )


def main() -> None:
    parser = argparse.ArgumentParser(description="Ingest Marketing Ops Summary workbook into the Pond.")
    parser.add_argument("--source-file", default=str(DEFAULT_SOURCE))
    parser.add_argument("--db", default=str(DB_PATH))
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    source = Path(args.source_file).expanduser().resolve()
    rd, as_of, rows = parse_workbook(source)
    unmapped = sorted({str(row["property_name"]) for row in rows if not row.get("property_id")})
    result = MarketingOpsSummaryIngestResult(
        report_date=rd,
        source_as_of_date=as_of,
        rows_found=len(rows),
        rows_upserted=0,
        mapped_properties=len(rows) - len(unmapped),
        unmapped_properties=unmapped,
        source_file=str(source),
    )
    if not args.dry_run:
        with sqlite3.connect(str(Path(args.db).expanduser().resolve())) as conn:
            result.rows_upserted = upsert_rows(conn, rows)
    print(json.dumps(result.__dict__, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
