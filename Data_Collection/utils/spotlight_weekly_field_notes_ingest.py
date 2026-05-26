#!/usr/bin/env python3
"""Ingest weekly Spotlight field notes and action plans into the Data Pond."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sqlite3
import sys
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from docx import Document
from openpyxl import load_workbook

ROOT = Path("/Users/mark/Property_Analytics")
DB_PATH = ROOT / "data" / "portfolio_analytics.db"
DEFAULT_INPUT_DIR = Path("/Users/mark/Downloads/spotlight")
MIGRATION_SQL = ROOT / "apps" / "api" / "migrations" / "0042_create_spotlight_weekly_field_notes.sql"

if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from Data_Collection.utils.property_identity import PropertyIdentity, resolve_property_identity  # noqa: E402


SNAPSHOT_COLUMNS = [
    "id",
    "report_date",
    "week_ending",
    "property_name",
    "property_id",
    "community_id",
    "region",
    "occupancy",
    "trend_30",
    "trend_60",
    "adjusted_trend_60",
    "tours",
    "new_leads",
    "applications_received",
    "pending_applications",
    "approved_applications",
    "cancellations_denials",
    "renewals_completed",
    "notices",
    "social_posts",
    "recovery_goal",
    "narrative_summary",
    "narrative_text",
    "source_files_json",
    "metrics_json",
    "quality_flags_json",
    "evidence_json",
]

ACTION_COLUMNS = [
    "id",
    "snapshot_id",
    "report_date",
    "property_name",
    "property_id",
    "community_id",
    "region",
    "action_item",
    "action_area",
    "assigned_to",
    "deadline",
    "completed_status",
    "notes",
    "action_category",
    "is_open",
    "quality_flags_json",
    "source_file",
    "evidence_json",
]

PROPERTY_HINTS = {
    "botanic": "Botanic",
    "forest view": "Forest View",
    "pointe": "The Pointe",
    "retreat": "The Retreat",
    "steeple": "Steeplechase",
    "thomas": "The Reserves of Thomas Glen",
}

METRIC_PATTERNS = {
    "occupancy": (r"Current Occupancy:\s*([^\n]+)", r"Physical Occupancy:\s*([^\n]+)", r"Occupancy:\s*([^\n]+)"),
    "trend_30": (r"30-Day\s*Trend:\s*([^\n]+)",),
    "trend_60": (r"60-Day\s*Trend:\s*([^\n]+)",),
    "adjusted_trend_60": (r"Adjusted 60-Day\s*Trend:\s*([^\n]+)",),
    "leased_pct": (r"\(?\s*([0-9.]+%)\s+Leased\)?",),
    "goal_this_week": (r"Goal this week:\s*([^\n]+)",),
    "new_leads": (r"New Leads:\s*([^\n]+)",),
    "tours": (r"Tours:\s*([^\n]+)",),
    "applications_received": (r"Applications Received:\s*([^\n]+)", r"Applications received today:\s*([^\n]+)", r"Gross Applications:\s*([^\n]+)", r"Applications:\s*([^\n]+)"),
    "pending_applications": (r"Pending Applications:\s*([^\n]+)",),
    "approved_applications": (r"Approved Applications:?\s*([^\n]+)",),
    "reservation_fee_paid_today": (r"Reservation fee paid\s*today:\s*([^\n]+)",),
    "reservation_fee_week_to_date": (r"Reservation fee\s*Week To Date:\s*([^\n]+)",),
    "cancellations_denials": (r"Application Cancellations/Denials:\s*([^\n]+)", r"Cancel/Denials:\s*([^\n]+)", r"Cancellations/Denials:\s*([^\n]+)"),
    "renewals_completed": (r"Renewals Completed:\s*([^\n]+)", r"Renewal Completed:\s*([^\n]+)", r"Renewals signed:\s*([^\n]+)"),
    "notices": (r"Notices:\s*([^\n]+)",),
    "social_posts": (r"Social\s*Media Posts Completed:\s*([^\n]+)", r"Social Ads posted\s*week to date:\s*([^\n]+)", r"Social Media\s*-\s*([^\n]+)", r"Social Media:\s*([^\n]+)"),
    "outreach_marketing_visits": (r"Outreach marketing visits this week:\s*([^\n]+)",),
    "work_orders_received": (r"Work orders received:\s*([^\n]+)",),
    "work_orders_completed": (r"Work orders completed:\s*([^\n]+)",),
    "work_orders_open": (r"Work orders\s*remaining\s*open:\s*([^\n]+)",),
    "move_ins_today": (r"Move\s*in:\s*([^\n]+)",),
    "ntv": (r"NTV:\s*([^\n]+)",),
    "vacant_ready_leased": (r"Vacant Ready Units:\s*\nLeased:\s*([^\n]+)",),
    "vacant_ready_unleased": (r"Vacant Ready Units:\s*\nLeased:\s*[^\n]+\nUnleased:\s*([^\n]+)",),
    "vacant_unready_leased": (r"Vacant\s*Unready units:\s*\nLeased:\s*([^\n]+)",),
    "vacant_unready_unleased": (r"Vacant\s*Unready units:\s*\nLeased:\s*[^\n]+\nUnleased:\s*([^\n]+)",),
}


@dataclass
class ActionItem:
    action_item: str
    action_area: str | None
    assigned_to: str | None
    deadline: str | None
    completed_status: str | None
    notes: str | None
    source_file: str
    row_number: int


@dataclass
class PropertyPacket:
    identity: PropertyIdentity
    report_date: str
    week_ending: str | None = None
    recovery_goal: str | None = None
    narrative_texts: list[str] = field(default_factory=list)
    narrative_summaries: list[str] = field(default_factory=list)
    source_files: list[str] = field(default_factory=list)
    metrics: dict[str, Any] = field(default_factory=dict)
    actions: list[ActionItem] = field(default_factory=list)
    quality_flags: list[str] = field(default_factory=list)


@dataclass
class SpotlightWeeklyFieldNotesIngestResult:
    source_dir: str
    report_date: str
    files_seen: int = 0
    snapshots_upserted: int = 0
    action_items_upserted: int = 0
    properties_mapped: int = 0
    unmapped_files: list[str] = field(default_factory=list)
    quality_flags: list[str] = field(default_factory=list)


def clean_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip()


def slug(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")


def stable_hash(*parts: Any) -> str:
    raw = "|".join(str(part or "") for part in parts)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:12]


def parse_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = clean_text(value).replace(",", "")
    if not text:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    return float(match.group(0))


def parse_percent(value: Any) -> float | None:
    parsed = parse_number(value)
    if parsed is None:
        return None
    text = clean_text(value)
    if "%" in text or parsed > 1:
        return parsed / 100
    return parsed


def parse_int(value: Any) -> int | None:
    parsed = parse_number(value)
    if parsed is None:
        return None
    return int(round(parsed))


def parse_date_value(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = clean_text(value)
    if not text or text.lower() in {"ongoing", "weekly", "monthly", "quarterly", "completed", "asap"}:
        return text or None
    for pattern in (r"(\d{4})-(\d{1,2})-(\d{1,2})", r"(\d{1,2})/(\d{1,2})/(\d{4})", r"(\d{1,2})\.(\d{1,2})\.(\d{2,4})"):
        match = re.search(pattern, text)
        if not match:
            continue
        first, second, third = match.groups()
        if len(first) == 4:
            year, month, day = int(first), int(second), int(third)
        else:
            month, day, year = int(first), int(second), int(third)
            if year < 100:
                year += 2000
        return f"{year:04d}-{month:02d}-{day:02d}"
    return text


def report_date_from_path(path: Path) -> str:
    return datetime.fromtimestamp(path.stat().st_mtime).date().isoformat()


def infer_property_from_name(path: Path) -> str | None:
    lowered = path.name.lower()
    for key, value in PROPERTY_HINTS.items():
        if key in lowered:
            return value
    return None


def resolve_property(name: str | None) -> PropertyIdentity | None:
    if not name:
        return None
    return resolve_property_identity(name)


def packet_key(identity: PropertyIdentity, report_date_value: str) -> str:
    return f"{identity.marketing_bi_property_id}:{report_date_value}"


def snapshot_id(identity: PropertyIdentity, report_date_value: str) -> str:
    return f"spotlight_weekly_{report_date_value}_{slug(identity.marketing_bi_property_id)}"


def summarize_text(text: str) -> str | None:
    lines = [clean_text(line) for line in text.splitlines() if clean_text(line)]
    narrative = [line for line in lines if len(line) > 80 and not line.lower().startswith(("from:", "sent:", "to:", "subject:"))]
    return " ".join(narrative[:2])[:900] if narrative else None


def action_category(action_area: str | None, action_item: str, notes: str | None) -> str:
    text = f"{action_area or ''} {action_item} {notes or ''}".lower()
    if any(term in text for term in ("renewal", "resident", "back door", "referral")):
        return "retention"
    if any(term in text for term in ("anyone", "lead", "follow up", "guest card", "leasing")):
        return "lead_management"
    if any(term in text for term in ("google", "review", "social", "outreach", "smartcomm", "post", "pep")):
        return "promotion"
    if any(term in text for term in ("pricing", "special", "concession", "rate")):
        return "pricing"
    if any(term in text for term in ("curb", "stale", "mini model", "amenity", "maintenance", "painter")):
        return "product"
    if any(term in text for term in ("training", "1:1", "hire", "staff", "team")):
        return "people"
    return "general"


def action_quality_flags(item: ActionItem, report_date_value: str) -> list[str]:
    flags: list[str] = []
    if not item.assigned_to:
        flags.append("missing_owner")
    if not item.deadline:
        flags.append("missing_deadline")
    if not item.completed_status:
        flags.append("missing_status")
    if item.deadline and re.match(r"\d{4}-\d{2}-\d{2}$", item.deadline):
        try:
            if date.fromisoformat(item.deadline) < date.fromisoformat(report_date_value) and (item.completed_status or "").lower() not in {"yes", "completed"}:
                flags.append("past_due_open")
        except ValueError:
            pass
    if len(item.action_item.split()) < 3:
        flags.append("vague_action")
    return flags


def load_workbook_packet(path: Path) -> tuple[str | None, str, dict[str, Any], list[ActionItem], list[str]]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    plan = workbook.worksheets[0]
    if hasattr(plan, "reset_dimensions"):
        plan.reset_dimensions()
    rows = list(plan.iter_rows(values_only=True))
    property_name = clean_text(rows[0][1] if rows and len(rows[0]) > 1 else None) or infer_property_from_name(path)
    report_date_value = parse_date_value(rows[0][10] if rows and len(rows[0]) > 10 else None) or report_date_from_path(path)
    metrics: dict[str, Any] = {}
    flags: list[str] = []

    def labeled_row(label: str) -> list[Any]:
        for row in rows[:20]:
            if clean_text(row[0] if row else None).lower() == label.lower():
                return list(row)
        return []

    budgeted = labeled_row("Budgeted")
    actual = labeled_row("Actual")
    delta = labeled_row("Delta (%)")
    if len(actual) > 1:
        metrics["occupancy"] = parse_percent(actual[1])
    if len(actual) > 2:
        metrics["trend_30"] = parse_percent(actual[2])
    if len(actual) > 3:
        metrics["trend_60"] = parse_percent(actual[3])
    if len(delta) > 1:
        metrics["budget_delta_current"] = parse_percent(delta[1])
    if len(budgeted) > 1:
        metrics["budgeted_current"] = parse_percent(budgeted[1])
    recovery_goal = None
    for row in rows[:20]:
        if clean_text(row[0] if row else None).lower() == "goal:":
            recovery_goal = clean_text(row[1] if len(row) > 1 else None) or None
            break

    for field in ("trend_30", "trend_60"):
        if metrics.get(field) is not None and metrics[field] < -0.25:
            flags.append(f"{field}_blank_or_invalid_in_template")
            metrics[field] = None

    actions: list[ActionItem] = []
    for worksheet in workbook.worksheets[1:]:
        if hasattr(worksheet, "reset_dimensions"):
            worksheet.reset_dimensions()
        action_rows = list(worksheet.iter_rows(values_only=True))
        in_table = False
        for row_number, row in enumerate(action_rows, start=1):
            values = [clean_text(value) for value in row[:6]]
            if values[:6] == ["Action Item", "Action Area", "Assigned To", "Deadline", "Completed", "Notes"]:
                in_table = True
                continue
            if not in_table or not any(values):
                continue
            action = values[0]
            if not action or action.upper() == "ACTION ITEMS":
                continue
            actions.append(
                ActionItem(
                    action_item=action,
                    action_area=values[1] or None,
                    assigned_to=values[2] or None,
                    deadline=parse_date_value(values[3]),
                    completed_status=values[4] or None,
                    notes=values[5] or None,
                    source_file=str(path.resolve()),
                    row_number=row_number,
                )
            )
    return property_name, report_date_value, {"metrics": metrics, "recovery_goal": recovery_goal}, actions, flags


def load_text_packet(path: Path) -> tuple[str | None, str, str, dict[str, Any], list[str]]:
    if path.suffix.lower() == ".docx":
        document = Document(path)
        text = "\n".join(paragraph.text.strip() for paragraph in document.paragraphs if paragraph.text.strip())
    else:
        text = path.read_text(encoding="utf-8", errors="replace")
    property_name = infer_property_from_name(path)
    lines = [clean_text(line) for line in text.splitlines() if clean_text(line)]
    if not property_name and lines:
        property_name = re.split(r"\s+[–-]\s+| Action Plan| Weekly Update", lines[0], maxsplit=1)[0].strip()

    report_date_value = report_date_from_path(path)
    week_ending = None
    match = re.search(r"Week Ending\s+(\d{1,2})\.(\d{1,2})\.(\d{4})", text, re.IGNORECASE)
    if match:
        month, day, year = (int(part) for part in match.groups())
        week_ending = f"{year:04d}-{month:02d}-{day:02d}"
        report_date_value = (date.fromisoformat(week_ending) + timedelta(days=1)).isoformat()

    metrics: dict[str, Any] = {}
    raw_metrics: dict[str, str] = {}
    for field, patterns in METRIC_PATTERNS.items():
        for pattern in patterns:
            found = re.search(pattern, text, re.IGNORECASE)
            if found:
                raw = clean_text(found.group(1))
                raw_metrics[field] = raw
                if field in {"occupancy", "trend_30", "trend_60", "adjusted_trend_60", "leased_pct"}:
                    metrics[field] = parse_percent(raw)
                else:
                    metrics[field] = parse_int(raw)
                break
    metrics["raw_note_metrics"] = raw_metrics
    return property_name, report_date_value, text, metrics, ([f"week_ending:{week_ending}"] if week_ending else [])


def merge_metric(existing: Any, incoming: Any) -> Any:
    if incoming is None:
        return existing
    return incoming


def build_packets(input_dir: Path) -> tuple[dict[str, PropertyPacket], list[str], list[str], int]:
    packets: dict[str, PropertyPacket] = {}
    unmapped_files: list[str] = []
    all_quality_flags: list[str] = []
    files = sorted(path for path in input_dir.iterdir() if path.suffix.lower() in {".xlsx", ".docx", ".txt"})

    def get_packet(identity: PropertyIdentity, report_date_value: str) -> PropertyPacket:
        key = packet_key(identity, report_date_value)
        if key not in packets:
            packets[key] = PropertyPacket(identity=identity, report_date=report_date_value)
        return packets[key]

    for path in files:
        if path.name.startswith("~$"):
            all_quality_flags.append(f"{path.name}:skipped_office_temp_file")
            continue
        try:
            if path.suffix.lower() == ".xlsx":
                property_name, report_date_value, workbook_payload, actions, flags = load_workbook_packet(path)
                identity = resolve_property(property_name)
                if not identity:
                    unmapped_files.append(path.name)
                    continue
                packet = get_packet(identity, report_date_value)
                packet.source_files.append(str(path.resolve()))
                packet.recovery_goal = workbook_payload.get("recovery_goal") or packet.recovery_goal
                for field, value in workbook_payload.get("metrics", {}).items():
                    packet.metrics[field] = merge_metric(packet.metrics.get(field), value)
                packet.actions.extend(actions)
                packet.quality_flags.extend(flags)
            else:
                property_name, report_date_value, text, metrics, flags = load_text_packet(path)
                identity = resolve_property(property_name)
                if not identity:
                    unmapped_files.append(path.name)
                    continue
                packet = get_packet(identity, report_date_value)
                packet.source_files.append(str(path.resolve()))
                packet.narrative_texts.append(text)
                summary = summarize_text(text)
                if summary:
                    packet.narrative_summaries.append(summary)
                week_flag = next((flag for flag in flags if flag.startswith("week_ending:")), None)
                if week_flag:
                    packet.week_ending = week_flag.split(":", 1)[1]
                for field, value in metrics.items():
                    if field == "raw_note_metrics":
                        raw = packet.metrics.setdefault("raw_note_metrics", {})
                        raw.update(value)
                    elif field == "goal_this_week":
                        packet.recovery_goal = f"Goal this week: {value}"
                        packet.metrics[field] = merge_metric(packet.metrics.get(field), value)
                    else:
                        packet.metrics[field] = merge_metric(packet.metrics.get(field), value)
        except Exception as exc:
            unmapped_files.append(f"{path.name}: {exc}")
            all_quality_flags.append(f"{path.name}:parse_error")

    return packets, unmapped_files, all_quality_flags, len(files)


def snapshot_row(packet: PropertyPacket) -> dict[str, Any]:
    identity = packet.identity
    sid = snapshot_id(identity, packet.report_date)
    source_files = sorted(set(packet.source_files))
    quality_flags = list(dict.fromkeys(packet.quality_flags))
    narrative_text = "\n\n---\n\n".join(packet.narrative_texts)[:12000] if packet.narrative_texts else None
    narrative_summary = " ".join(packet.narrative_summaries)[:1500] if packet.narrative_summaries else None
    metrics = packet.metrics
    return {
        "id": sid,
        "report_date": packet.report_date,
        "week_ending": packet.week_ending,
        "property_name": identity.property_name,
        "property_id": identity.marketing_bi_property_id,
        "community_id": identity.community_id,
        "region": identity.encasa_region,
        "occupancy": metrics.get("occupancy"),
        "trend_30": metrics.get("trend_30"),
        "trend_60": metrics.get("trend_60"),
        "adjusted_trend_60": metrics.get("adjusted_trend_60"),
        "tours": metrics.get("tours"),
        "new_leads": metrics.get("new_leads"),
        "applications_received": metrics.get("applications_received"),
        "pending_applications": metrics.get("pending_applications"),
        "approved_applications": metrics.get("approved_applications"),
        "cancellations_denials": metrics.get("cancellations_denials"),
        "renewals_completed": metrics.get("renewals_completed"),
        "notices": metrics.get("notices"),
        "social_posts": metrics.get("social_posts"),
        "recovery_goal": packet.recovery_goal,
        "narrative_summary": narrative_summary,
        "narrative_text": narrative_text,
        "source_files_json": json.dumps(source_files, sort_keys=True),
        "metrics_json": json.dumps(metrics, sort_keys=True, default=str),
        "quality_flags_json": json.dumps(quality_flags, sort_keys=True),
        "evidence_json": json.dumps(
            {
                "source": "spotlight_weekly_field_notes",
                "identity": identity.as_mapping(),
                "action_item_count": len(packet.actions),
            },
            sort_keys=True,
            default=str,
        ),
    }


def action_rows(packet: PropertyPacket) -> list[dict[str, Any]]:
    identity = packet.identity
    sid = snapshot_id(identity, packet.report_date)
    rows: list[dict[str, Any]] = []
    for index, item in enumerate(packet.actions, start=1):
        flags = action_quality_flags(item, packet.report_date)
        status = clean_text(item.completed_status).lower()
        is_open = 0 if status in {"yes", "completed", "complete"} else 1
        category = action_category(item.action_area, item.action_item, item.notes)
        rows.append(
            {
                "id": f"{sid}_action_{index:03d}_{stable_hash(item.action_item, item.source_file, item.row_number)}",
                "snapshot_id": sid,
                "report_date": packet.report_date,
                "property_name": identity.property_name,
                "property_id": identity.marketing_bi_property_id,
                "community_id": identity.community_id,
                "region": identity.encasa_region,
                "action_item": item.action_item,
                "action_area": item.action_area,
                "assigned_to": item.assigned_to,
                "deadline": item.deadline,
                "completed_status": item.completed_status,
                "notes": item.notes,
                "action_category": category,
                "is_open": is_open,
                "quality_flags_json": json.dumps(flags, sort_keys=True),
                "source_file": item.source_file,
                "evidence_json": json.dumps({"row_number": item.row_number}, sort_keys=True),
            }
        )
    return rows


def ensure_tables(conn: sqlite3.Connection) -> None:
    conn.executescript(MIGRATION_SQL.read_text(encoding="utf-8"))


def upsert_packets(conn: sqlite3.Connection, packets: dict[str, PropertyPacket]) -> tuple[int, int]:
    ensure_tables(conn)
    snapshots = [snapshot_row(packet) for packet in packets.values()]
    actions = [row for packet in packets.values() for row in action_rows(packet)]
    snapshot_placeholders = ", ".join("?" for _ in SNAPSHOT_COLUMNS)
    snapshot_updates = ", ".join(f"{column}=excluded.{column}" for column in SNAPSHOT_COLUMNS if column != "id")
    conn.executemany(
        f"""
        INSERT INTO spotlight_weekly_field_snapshots ({", ".join(SNAPSHOT_COLUMNS)})
        VALUES ({snapshot_placeholders})
        ON CONFLICT(report_date, property_id)
        DO UPDATE SET {snapshot_updates}, updated_at=datetime('now')
        """,
        [[row.get(column) for column in SNAPSHOT_COLUMNS] for row in snapshots],
    )

    if actions:
        action_placeholders = ", ".join("?" for _ in ACTION_COLUMNS)
        action_updates = ", ".join(f"{column}=excluded.{column}" for column in ACTION_COLUMNS if column != "id")
        conn.executemany(
            f"""
            INSERT INTO spotlight_weekly_action_items ({", ".join(ACTION_COLUMNS)})
            VALUES ({action_placeholders})
            ON CONFLICT(id)
            DO UPDATE SET {action_updates}, updated_at=datetime('now')
            """,
            [[row.get(column) for column in ACTION_COLUMNS] for row in actions],
        )
    conn.commit()
    return len(snapshots), len(actions)


def ingest_spotlight_weekly_field_notes(
    input_dir: Path = DEFAULT_INPUT_DIR,
    db_path: Path = DB_PATH,
    dry_run: bool = False,
) -> SpotlightWeeklyFieldNotesIngestResult:
    packets, unmapped, quality_flags, files_seen = build_packets(input_dir)
    report_dates = sorted({packet.report_date for packet in packets.values()})
    result = SpotlightWeeklyFieldNotesIngestResult(
        source_dir=str(input_dir),
        report_date=report_dates[-1] if report_dates else datetime.now().date().isoformat(),
        files_seen=files_seen,
        properties_mapped=len({packet.identity.marketing_bi_property_id for packet in packets.values()}),
        unmapped_files=unmapped,
        quality_flags=quality_flags + [
            f"{packet.identity.marketing_bi_property_id}:{flag}"
            for packet in packets.values()
            for flag in packet.quality_flags
        ],
    )
    if not dry_run:
        with sqlite3.connect(str(db_path)) as conn:
            result.snapshots_upserted, result.action_items_upserted = upsert_packets(conn, packets)
    return result


def main() -> None:
    parser = argparse.ArgumentParser(description="Ingest weekly Spotlight field notes and action plans.")
    parser.add_argument("--input-dir", default=str(DEFAULT_INPUT_DIR))
    parser.add_argument("--db", default=str(DB_PATH))
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    result = ingest_spotlight_weekly_field_notes(
        input_dir=Path(args.input_dir).expanduser().resolve(),
        db_path=Path(args.db).expanduser().resolve(),
        dry_run=args.dry_run,
    )
    print(json.dumps(result.__dict__, indent=2, sort_keys=True, default=str))
    if result.unmapped_files:
        raise SystemExit(1)


if __name__ == "__main__":
    main()
