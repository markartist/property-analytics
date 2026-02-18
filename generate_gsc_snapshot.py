#!/usr/bin/env python3
"""
Portfolio Google Search Console Snapshot Report Generator v1.0
===============================================================

Complete portfolio listing sorted by organic clicks (high to low)
showing detailed GSC metrics for all properties over the last 30 days.

Author: Mark Laufhutte / Atlas
Date: 2026-01-28
Version: 1.0
"""

import sys
import sqlite3
import json
from datetime import datetime, timedelta
from pathlib import Path
import pandas as pd

# Add utils to path
sys.path.insert(0, str(Path(__file__).parent / "utils"))
from report_builder import ReportBuilder, Section, KPITile

# Configuration
DB_PATH = Path(__file__).parent / "data" / "portfolio_analytics.db"
REGISTRY_PATH = Path(__file__).parent / "config" / "venterra_properties_official.json"
OUTPUT_DIR = Path(__file__).parent / "reports" / "gsc_snapshot"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


class GSCSnapshotReport:
    """Portfolio Google Search Console Snapshot Report Generator"""
    
    def __init__(self, report_date: datetime):
        self.report_date = report_date
        self.date_str = report_date.strftime("%Y-%m-%d")
        
        self.conn = sqlite3.connect(DB_PATH)
        self.conn.row_factory = sqlite3.Row
        
        # Load property registry for name mapping
        with open(REGISTRY_PATH) as f:
            registry = json.load(f)
        self.property_name_map = {}
        for prop in registry['properties']:
            if prop.get('gsc_url'):
                self.property_name_map[prop['gsc_url']] = prop['name']
        
        self.properties = []
        self.stats = {}
        
    def _get_data(self):
        """Get all property data for last 30 days sorted by clicks"""
        
        # Calculate date range (last 30 days)
        end_date = self.report_date - timedelta(days=1)  # Yesterday
        start_date = end_date - timedelta(days=29)  # 30 days total
        
        # Get previous 30 days for comparison
        prev_end_date = start_date - timedelta(days=1)
        prev_start_date = prev_end_date - timedelta(days=29)
        
        # Get all properties with GSC metrics for current period
        properties = self.conn.execute("""
            SELECT 
                g.property_id,
                SUM(g.clicks) as clicks,
                SUM(g.impressions) as impressions,
                ROUND(AVG(g.ctr), 2) as avg_ctr,
                ROUND(AVG(g.average_position), 1) as avg_position
            FROM gsc_daily_metrics g
            WHERE g.metric_date BETWEEN ? AND ?
            GROUP BY g.property_id
            ORDER BY clicks DESC
        """, (start_date.strftime('%Y-%m-%d'), end_date.strftime('%Y-%m-%d'))).fetchall()
        
        # Get previous period data for trends
        prev_data = {}
        prev_properties = self.conn.execute("""
            SELECT 
                property_id,
                SUM(clicks) as prev_clicks,
                SUM(impressions) as prev_impressions,
                ROUND(AVG(ctr), 2) as prev_ctr,
                ROUND(AVG(average_position), 1) as prev_average_position
            FROM gsc_daily_metrics
            WHERE metric_date BETWEEN ? AND ?
            GROUP BY property_id
        """, (prev_start_date.strftime('%Y-%m-%d'), prev_end_date.strftime('%Y-%m-%d'))).fetchall()
        
        for row in prev_properties:
            prev_data[row['property_id']] = dict(row)
        
        # Combine current and previous data
        self.properties = []
        for row in properties:
            prop = dict(row)
            prop_id = prop['property_id']
            
            # Get actual property name from registry, fallback to URL
            if prop_id in self.property_name_map:
                prop['property_name'] = self.property_name_map[prop_id]
            elif ':' in prop_id:
                prop['property_name'] = prop_id.split(':', 1)[1]
            else:
                prop['property_name'] = prop_id
            
            # Add previous period data if available
            if prop_id in prev_data:
                prop['prev_clicks'] = prev_data[prop_id]['prev_clicks']
                prop['prev_impressions'] = prev_data[prop_id]['prev_impressions']
                prop['prev_ctr'] = prev_data[prop_id]['prev_ctr']
                prop['prev_average_position'] = prev_data[prop_id]['prev_average_position']
            else:
                prop['prev_clicks'] = None
                prop['prev_impressions'] = None
                prop['prev_ctr'] = None
                prop['prev_average_position'] = None
            
            self.properties.append(prop)
        
        # Calculate portfolio stats
        total = len(self.properties)
        if total > 0:
            total_clicks = sum(p['clicks'] for p in self.properties)
            total_impressions = sum(p['impressions'] for p in self.properties)
            avg_ctr = (total_clicks / total_impressions * 100) if total_impressions > 0 else 0
            avg_position = sum(p['avg_position'] for p in self.properties) / total
            
            # Previous period totals
            prev_total_clicks = sum(p['prev_clicks'] for p in self.properties if p['prev_clicks'])
            prev_total_impressions = sum(p['prev_impressions'] for p in self.properties if p['prev_impressions'])
            prev_avg_ctr = (prev_total_clicks / prev_total_impressions * 100) if prev_total_impressions > 0 else None
            prev_avg_position = None
            if any(p['prev_average_position'] for p in self.properties):
                prev_average_positions = [p['prev_average_position'] for p in self.properties if p['prev_average_position']]
                prev_avg_position = sum(prev_average_positions) / len(prev_average_positions) if prev_average_positions else None
            
            # CTR performance bands
            excellent_ctr = sum(1 for p in self.properties if p['avg_ctr'] >= 5.0)
            good_ctr = sum(1 for p in self.properties if 3.0 <= p['avg_ctr'] < 5.0)
            needs_improvement = sum(1 for p in self.properties if p['avg_ctr'] < 3.0)
            
            self.stats = {
                'total': total,
                'total_clicks': int(total_clicks),
                'total_impressions': int(total_impressions),
                'avg_ctr': round(avg_ctr, 2),
                'avg_position': round(avg_position, 1),
                'start_date': start_date.strftime('%Y-%m-%d'),
                'end_date': end_date.strftime('%Y-%m-%d'),
                'prev_total_clicks': int(prev_total_clicks) if prev_total_clicks else None,
                'prev_total_impressions': int(prev_total_impressions) if prev_total_impressions else None,
                'prev_avg_ctr': round(prev_avg_ctr, 2) if prev_avg_ctr else None,
                'prev_avg_position': round(prev_avg_position, 1) if prev_avg_position else None,
                'excellent_ctr': excellent_ctr,
                'good_ctr': good_ctr,
                'needs_improvement': needs_improvement
            }
    
    def _format_trend(self, delta, invert=False, decimals=0, suffix='') -> str:
        """Format trend indicator with color coding"""
        if delta is None or delta == 0:
            return ''
        
        is_improvement = (delta < 0) if invert else (delta > 0)
        color = '#28a745' if is_improvement else '#dc3545'
        arrow = '↓' if delta < 0 else '↑'
        value_str = f"{abs(delta):,.{decimals}f}"
        
        return f' <span style="font-size: 14px; color: {color}; margin-left: 8px;">{arrow}{value_str}{suffix}</span>'
    
    def _format_number(self, num):
        """Format large numbers with commas"""
        if num is None:
            return 'N/A'
        return f"{int(num):,}"
    
    def _get_ctr_badge(self, ctr):
        """Get color-coded CTR badge"""
        if ctr >= 5.0:
            return '#28a745', 'EXCELLENT'
        elif ctr >= 3.0:
            return '#ff8800', 'GOOD'
        else:
            return '#dc3545', 'NEEDS IMPROVEMENT'
    
    def _get_position_badge(self, position):
        """Get color-coded position badge"""
        if position <= 10:
            return '#28a745', 'EXCELLENT'
        elif position <= 20:
            return '#ff8800', 'GOOD'
        else:
            return '#dc3545', 'NEEDS IMPROVEMENT'
    
    def _generate_overview(self) -> str:
        """Generate portfolio overview section"""
        
        ctr_color, ctr_label = self._get_ctr_badge(self.stats['avg_ctr'])
        pos_color, pos_label = self._get_position_badge(self.stats['avg_position'])
        
        html = f"""
        <div style="background: #f8f9fa; padding: 20px; border-radius: 6px; margin-bottom: 30px;">
            <div style="text-align: center; margin-bottom: 15px;">
                <div style="font-size: 16px; font-weight: 600; margin-bottom: 5px;">
                    <span style="color: #007bff;">{self._format_number(self.stats['total_clicks'])} Clicks</span>
                    <span style="color: #6c757d; margin: 0 10px;">|</span>
                    <span style="color: #6c757d;">{self._format_number(self.stats['total_impressions'])} Impressions</span>
                </div>
                <div style="font-size: 13px; color: #6c757d;">
                    {self.stats['total']} properties analyzed | {self.stats['start_date']} to {self.stats['end_date']} (30 days)
                </div>
            </div>
            
            <table style="width: 100%; max-width: 600px; margin: 0 auto;">
                <tr>
                    <td style="width: 33%; text-align: center; padding: 15px; background: #f8d7da; border-radius: 4px;">
                        <div style="font-size: 32px; font-weight: 700; color: #dc3545;">{self.stats['needs_improvement']}</div>
                        <div style="font-size: 12px; color: #721c24; margin-top: 5px; font-weight: 600;">NEEDS IMPROVEMENT<br>CTR &lt;3%</div>
                    </td>
                    <td style="width: 2%;"></td>
                    <td style="width: 33%; text-align: center; padding: 15px; background: #fff3cd; border-radius: 4px;">
                        <div style="font-size: 32px; font-weight: 700; color: #ff8800;">{self.stats['good_ctr']}</div>
                        <div style="font-size: 12px; color: #856404; margin-top: 5px; font-weight: 600;">GOOD<br>CTR 3-5%</div>
                    </td>
                    <td style="width: 2%;"></td>
                    <td style="width: 33%; text-align: center; padding: 15px; background: #d4edda; border-radius: 4px;">
                        <div style="font-size: 32px; font-weight: 700; color: #28a745;">{self.stats['excellent_ctr']}</div>
                        <div style="font-size: 12px; color: #155724; margin-top: 5px; font-weight: 600;">EXCELLENT<br>CTR 5%+</div>
                    </td>
                </tr>
            </table>
        </div>
        """
        
        return html
    
    def _generate_property_list(self) -> str:
        """Generate complete property listing"""
        
        html = '<div style="margin-bottom: 30px;">'
        
        for i, prop in enumerate(self.properties, 1):
            # Determine border color based on CTR
            ctr = prop['avg_ctr']
            if ctr >= 5.0:
                border_color = '#28a745'
            elif ctr >= 3.0:
                border_color = '#ff8800'
            else:
                border_color = '#dc3545'
            
            # Calculate trends
            clicks_trend = ''
            impressions_trend = ''
            ctr_trend = ''
            position_trend = ''
            
            if prop.get('prev_clicks') is not None:
                clicks_delta = prop['clicks'] - prop['prev_clicks']
                if clicks_delta != 0:
                    clicks_trend = self._format_trend(clicks_delta, invert=False)
            
            if prop.get('prev_impressions') is not None:
                impressions_delta = prop['impressions'] - prop['prev_impressions']
                if impressions_delta != 0:
                    impressions_trend = self._format_trend(impressions_delta, invert=False)
            
            if prop.get('prev_ctr') is not None:
                ctr_delta = prop['avg_ctr'] - prop['prev_ctr']
                if abs(ctr_delta) >= 0.01:
                    ctr_trend = self._format_trend(ctr_delta, invert=False, decimals=2, suffix='%')
            
            if prop.get('prev_average_position') is not None:
                position_delta = prop['avg_position'] - prop['prev_average_position']
                if abs(position_delta) >= 0.1:
                    position_trend = self._format_trend(position_delta, invert=True, decimals=1)
            
            # Position color
            pos_color = self._get_position_badge(prop['avg_position'])[0]
            
            html += f"""
            <div style="padding: 12px; background: #fff; border-left: 4px solid {border_color}; margin-bottom: 12px; border-radius: 4px;">
                <div style="display: flex; justify-content: space-between; align-items: flex-start;">
                    <div style="flex: 1;">
                        <div style="font-weight: 600; color: #333; font-size: 14px; margin-bottom: 8px;">
                            #{i}. {prop['property_name']}
                        </div>
                        <table style="width: 100%; font-size: 11px;">
                            <tr>
                                <td style="padding: 2px 20px 2px 0; width: 30%;"><strong>Clicks:</strong> <span style="color: #007bff; font-size: 13px; font-weight: 600;">{self._format_number(prop['clicks'])}</span>{clicks_trend}</td>
                                <td style="padding: 2px 20px 2px 0; width: 35%;"><strong>Impressions:</strong> <span style="color: #6c757d;">{self._format_number(prop['impressions'])}</span>{impressions_trend}</td>
                                <td style="padding: 2px 0; width: 35%;"><strong>CTR:</strong> <span style="color: {border_color}; font-weight: 600;">{prop['avg_ctr']:.2f}%</span>{ctr_trend}</td>
                            </tr>
                            <tr>
                                <td style="padding: 2px 20px 2px 0;"><strong>Avg Position:</strong> <span style="color: {pos_color}; font-weight: 600;">{prop['avg_position']:.1f}</span>{position_trend}</td>
                                <td colspan="2" style="padding: 2px 0; color: #6c757d; font-size: 10px;">30-day total (vs. previous 30 days)</td>
                            </tr>
                        </table>
                    </div>
                    <div style="margin-left: 15px; text-align: center; min-width: 120px;">
                        <div style="font-size: 28px; font-weight: 700; color: #007bff; line-height: 1;">
                            {self._format_number(prop['clicks'])}
                        </div>
                        <div style="font-size: 11px; color: #6c757d; margin: 2px 0;">clicks</div>
                        <div style="font-size: 12px; color: #6c757d; margin-top: 8px;">{clicks_trend if clicks_trend else '—'}</div>
                    </div>
                </div>
            </div>
            """
        
        html += '</div>'
        return html
    
    def _generate_data_integrity_section(self) -> str:
        """Generate data integrity footer"""
        
        html = f"""
        <div style="margin-top: 40px; padding: 15px; background: #d4edda; border-left: 4px solid #28a745; border-radius: 4px;">
            <div style="font-size: 13px; font-weight: 600; color: #333; margin-bottom: 10px;">✅ Data Integrity</div>
            <div style="font-size: 12px; color: #333; line-height: 1.8;">
                <strong>Date Range:</strong> {self.stats['start_date']} to {self.stats['end_date']} (30 days)<br>
                <strong>Properties in Report:</strong> {self.stats['total']}<br>
                <strong>Data Source:</strong> Google Search Console API<br>
                <strong>Metrics:</strong> Clicks, Impressions, CTR, Average Position
            </div>
        </div>
        """
        
        return html
    
    def _generate_excel(self) -> str:
        """Generate Excel spreadsheet with all property data"""
        
        # Prepare data for Excel
        excel_data = []
        for i, prop in enumerate(self.properties, 1):
            # Calculate trends
            clicks_delta = None
            impressions_delta = None
            ctr_delta = None
            position_delta = None
            
            if prop.get('prev_clicks') is not None:
                clicks_delta = prop['clicks'] - prop['prev_clicks']
            if prop.get('prev_impressions') is not None:
                impressions_delta = prop['impressions'] - prop['prev_impressions']
            if prop.get('prev_ctr') is not None:
                ctr_delta = prop['avg_ctr'] - prop['prev_ctr']
            if prop.get('prev_average_position') is not None:
                position_delta = prop['avg_position'] - prop['prev_average_position']
            
            # Grade based on CTR
            if prop['avg_ctr'] >= 5.0:
                grade = 'EXCELLENT'
            elif prop['avg_ctr'] >= 3.0:
                grade = 'GOOD'
            else:
                grade = 'NEEDS IMPROVEMENT'
            
            excel_data.append({
                'Rank': i,
                'Property Name': prop['property_name'],
                'Grade': grade,
                'Clicks': prop['clicks'],
                'Clicks Change': clicks_delta if clicks_delta is not None else '',
                'Impressions': prop['impressions'],
                'Impressions Change': impressions_delta if impressions_delta is not None else '',
                'CTR (%)': round(prop['avg_ctr'], 2),
                'CTR Change (%)': round(ctr_delta, 2) if ctr_delta is not None else '',
                'Avg Position': round(prop['avg_position'], 1),
                'Position Change': round(position_delta, 1) if position_delta is not None else ''
            })
        
        # Create DataFrame
        df = pd.DataFrame(excel_data)
        
        # Save to Excel with formatting
        excel_file = OUTPUT_DIR / f"Portfolio_GSC_Snapshot_{self.date_str}.xlsx"
        
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Search Console Metrics', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Search Console Metrics']
            
            # Set column widths
            worksheet.column_dimensions['A'].width = 6   # Rank
            worksheet.column_dimensions['B'].width = 40  # Property Name
            worksheet.column_dimensions['C'].width = 20  # Grade
            worksheet.column_dimensions['D'].width = 12  # Clicks
            worksheet.column_dimensions['E'].width = 14  # Clicks Change
            worksheet.column_dimensions['F'].width = 14  # Impressions
            worksheet.column_dimensions['G'].width = 18  # Impressions Change
            worksheet.column_dimensions['H'].width = 10  # CTR
            worksheet.column_dimensions['I'].width = 14  # CTR Change
            worksheet.column_dimensions['J'].width = 12  # Avg Position
            worksheet.column_dimensions['K'].width = 15  # Position Change
            
            # Color code the Grade column
            from openpyxl.styles import PatternFill, Font
            
            for row in range(2, len(excel_data) + 2):
                grade_cell = worksheet[f'C{row}']
                
                if grade_cell.value == 'EXCELLENT':
                    grade_cell.fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
                    grade_cell.font = Font(color='155724', bold=True)
                elif grade_cell.value == 'GOOD':
                    grade_cell.fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
                    grade_cell.font = Font(color='856404', bold=True)
                elif grade_cell.value == 'NEEDS IMPROVEMENT':
                    grade_cell.fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
                    grade_cell.font = Font(color='721C24', bold=True)
                
                # Color code changes
                for col, is_inverse in [('E', False), ('G', False), ('I', False), ('K', True)]:
                    change_cell = worksheet[f'{col}{row}']
                    if change_cell.value != '':
                        is_positive = change_cell.value > 0
                        is_good = (not is_positive) if is_inverse else is_positive
                        if is_good:
                            change_cell.font = Font(color='28A745')
                        else:
                            change_cell.font = Font(color='DC3545')
            
            # Freeze the header row
            worksheet.freeze_panes = 'A2'
        
        print(f"   Excel: {excel_file}")
        return str(excel_file)
    
    def generate(self) -> tuple:
        """Generate complete report and Excel export"""
        
        print("Collecting GSC data for last 30 days...")
        self._get_data()
        
        print(f"Building report for {self.stats['total']} properties...")
        builder = ReportBuilder(
            title="Portfolio Google Search Console Snapshot",
            subtitle="Complete Property Listing Sorted by Organic Clicks (30 Days)",
            version="1.0",
            date_range=f"{self.stats['start_date']} to {self.stats['end_date']}"
        )
        
        # Calculate trends for KPI tiles
        clicks_trend = ''
        impressions_trend = ''
        ctr_trend = ''
        
        if self.stats.get('prev_total_clicks') is not None:
            delta = self.stats['total_clicks'] - self.stats['prev_total_clicks']
            if delta != 0:
                arrow = '↑' if delta > 0 else '↓'
                clicks_trend = f"{arrow}{abs(delta):,}"
        
        if self.stats.get('prev_total_impressions') is not None:
            delta = self.stats['total_impressions'] - self.stats['prev_total_impressions']
            if delta != 0:
                arrow = '↑' if delta > 0 else '↓'
                impressions_trend = f"{arrow}{abs(delta):,}"
        
        if self.stats.get('prev_avg_ctr') is not None:
            delta = self.stats['avg_ctr'] - self.stats['prev_avg_ctr']
            if abs(delta) >= 0.01:
                arrow = '↑' if delta > 0 else '↓'
                ctr_trend = f"{arrow}{abs(delta):.2f}%"
        
        # Add KPI tiles
        builder.add_kpi_tiles([
            KPITile(
                label="Total Clicks",
                value=f"{self.stats['total_clicks']:,}",
                sublabel="30-day total",
                trend=clicks_trend if clicks_trend else None,
                is_primary=True
            ),
            KPITile(
                label="Total Impressions",
                value=f"{self.stats['total_impressions']:,}",
                sublabel="30-day total",
                trend=impressions_trend if impressions_trend else None
            ),
            KPITile(
                label="Average CTR",
                value=f"{self.stats['avg_ctr']:.2f}%",
                sublabel="Portfolio average",
                trend=ctr_trend if ctr_trend else None
            )
        ], columns=3)
        
        # Portfolio Overview (no status badge)
        overview_html = self._generate_overview()
        
        # Custom section class without status badge
        class SectionNoStatus(Section):
            def to_html(self):
                html = f'''
                <!-- Section Header -->
                <table cellpadding="0" cellspacing="0" border="0" style="width: 100%; margin: 40px 0 0 0;">
                    <tr>
                        <td style="background: #15284B; padding: 15px 20px; border-radius: 6px 6px 0 0;">
                            <h2 style="font-size: 18px; color: #ffffff; margin: 0; font-weight: 600;">{self.title}</h2>
                        </td>
                    </tr>
                '''
                
                if self.description:
                    html += f'''
                    <tr>
                        <td style="background: #f8f9fa; padding: 12px 20px; border-radius: 0 0 6px 6px; margin-bottom: 20px;">
                            <p style="font-size: 13px; color: #6c757d; margin: 0; font-style: italic;">{self.description}</p>
                        </td>
                    </tr>
                    '''
                
                html += '''
                </table>
                
                <!-- Section Content -->
                '''
                html += self.content
                
                return html
        
        builder.add_section(SectionNoStatus(
            title="Portfolio Overview",
            content=overview_html,
            description=f"Organic search performance for {self.stats['total']} properties"
        ))
        
        # Complete Property Listing
        property_list_html = self._generate_property_list()
        builder.add_section(SectionNoStatus(
            title="Complete Property Ranking by Clicks",
            content=property_list_html,
            description="All properties sorted by organic clicks (highest to lowest)"
        ))
        
        # Data Integrity Section
        integrity_html = self._generate_data_integrity_section()
        builder.add_section(SectionNoStatus(
            title="Data Integrity",
            content=integrity_html,
            description=f"30-day collection period: {self.stats['start_date']} to {self.stats['end_date']}"
        ))
        
        # Generate HTML
        html = builder.generate()
        
        # Save report
        output_file = OUTPUT_DIR / f"Portfolio_GSC_Snapshot_{self.date_str}.html"
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(html)
        
        print(f"\n✅ Portfolio GSC Snapshot generated:")
        print(f"   HTML: {output_file}")
        
        # Generate Excel export
        excel_file = self._generate_excel()
        
        print(f"\n📊 Portfolio Summary:")
        print(f"   Total Properties: {self.stats['total']}")
        print(f"   Total Clicks: {self.stats['total_clicks']:,}")
        print(f"   Total Impressions: {self.stats['total_impressions']:,}")
        print(f"   Average CTR: {self.stats['avg_ctr']:.2f}%")
        print(f"   Average Position: {self.stats['avg_position']:.1f}")
        print(f"   Date Range: {self.stats['start_date']} to {self.stats['end_date']}")
        
        return str(output_file), str(excel_file)
    
    def __del__(self):
        if hasattr(self, 'conn'):
            self.conn.close()


def main():
    import argparse
    
    parser = argparse.ArgumentParser(description="Generate Portfolio GSC Snapshot")
    parser.add_argument('--date', help='Report date (YYYY-MM-DD), defaults to today')
    
    args = parser.parse_args()
    
    # Determine report date
    if args.date:
        report_date = datetime.strptime(args.date, "%Y-%m-%d")
    else:
        report_date = datetime.now()
    
    print("=" * 70)
    print("PORTFOLIO GOOGLE SEARCH CONSOLE SNAPSHOT v1.0")
    print("=" * 70)
    print(f"\nReport Date: {report_date.strftime('%Y-%m-%d')}\n")
    
    # Generate report
    reporter = GSCSnapshotReport(report_date)
    html_file, excel_file = reporter.generate()
    
    print(f"\n✅ Files ready for email attachment:")
    print(f"   {excel_file}")
    
    return 0


if __name__ == "__main__":
    sys.exit(main())
