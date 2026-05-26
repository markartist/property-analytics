#!/usr/bin/env python3
"""
Paid Media Performance Workbook Generator
==========================================
Generates Excel workbook with 30-day Google Ads performance for all Venterra properties.

Audience: Community Managers, Regional Managers
Purpose: Visibility into paid media spend, targeting, and inventory alignment
"""

import os
import sys
import json
import re
from datetime import datetime, timedelta
from typing import Dict, List
import requests
from google.ads.googleads.client import GoogleAdsClient
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from openpyxl.utils import get_column_letter
from utils.google_ads_ksm import materialize_google_ads_yaml

# Configuration
CUSTOMER_ID = "9089267423"
PROPERTY_REGISTRY = "/Users/mark/Property_Analytics/config/venterra_properties_official.json"
AVAILABILITY_FEED_URL = "https://online.venterraliving.com/encasa-external/ThirtyLines"
OUTPUT_DIR = "/Users/mark/Property_Analytics/paid_media_workbook/outputs"
TIME_WINDOW_DAYS = 30

# ============================================================================
# KEYWORD CLASSIFICATION (Same logic as PIB)
# ============================================================================

def classify_keyword(keyword: str) -> str:
    """
    Deterministic keyword classification for floor plan targeting

    Returns: 'Studio' | '1BR' | '2BR' | 'Unclassified'
    """
    keyword_lower = keyword.lower()

    # Studio / Efficiency
    if any(term in keyword_lower for term in ['studio', 'efficiency']):
        return 'Studio'

    # 1 Bedroom
    one_br_patterns = [r'\bone\s*bedroom', r'\b1\s*bedroom', r'\b1\s*br\b', r'\b1-bedroom']
    for pattern in one_br_patterns:
        if re.search(pattern, keyword_lower):
            return '1BR'

    # 2 Bedroom
    two_br_patterns = [r'\btwo\s*bedroom', r'\b2\s*bedroom', r'\b2\s*br\b', r'\b2-bedroom']
    for pattern in two_br_patterns:
        if re.search(pattern, keyword_lower):
            return '2BR'

    return 'Unclassified'


def classify_unclassified_keyword(keyword: str, property_name: str) -> str:
    """
    Classify unclassified (generic) keywords into subtypes

    Priority order:
    1. Brand - contains property name
    2. Competitor - contains competitor name (if list available)
    3. Local Generic - generic + geographic modifiers
    4. Other Generic - fallback

    Returns: 'Brand' | 'Competitor' | 'Local Generic' | 'Other Generic'
    """
    keyword_lower = keyword.lower()
    property_name_lower = property_name.lower()

    # Brand: Contains property name or significant parts of it
    # Remove common words and check if substantial part matches
    prop_words = set(property_name_lower.split()) - {'the', 'at', 'on', 'of', 'apartments'}
    if any(word in keyword_lower for word in prop_words if len(word) > 3):
        return 'Brand'

    # Competitor: Check against known competitor list (placeholder for now)
    # In future, load from config file
    competitor_keywords = ['maa', 'greystar', 'camden', 'equity', 'aimco', 'bozzuto']
    if any(comp in keyword_lower for comp in competitor_keywords):
        return 'Competitor'

    # Local Generic: Contains geographic modifiers + apartment intent
    geo_patterns = ['near me', 'in ', ' tx', ' fl', ' nc', ' ga', ' sc', ' al', ' tn',
                    'austin', 'dallas', 'houston', 'orlando', 'tampa', 'jacksonville',
                    'charlotte', 'raleigh', 'atlanta', 'nashville']
    apartment_terms = ['apartment', 'apt', 'rental', 'lease', 'housing', 'complex']

    has_geo = any(geo in keyword_lower for geo in geo_patterns)
    has_apartment = any(term in keyword_lower for term in apartment_terms)

    if has_geo and has_apartment:
        return 'Local Generic'

    # Default fallback
    return 'Other Generic'


# ============================================================================
# DATA COLLECTION
# ============================================================================

def load_property_registry() -> Dict:
    """Load property registry with market and region info"""
    with open(PROPERTY_REGISTRY, 'r') as f:
        registry = json.load(f)

    # Create lookup by property name (normalized for matching)
    properties = {}
    for prop in registry.get('properties', []):
        name = prop.get('name', '')
        if not name:  # Skip if no name
            continue

        # Attempt to populate market and region from registry
        market = prop.get('market', '')
        region = prop.get('region', '')

        # Use placeholders if not available
        if not market:
            market = 'TBD (Coming Soon)'
        if not region:
            region = 'TBD (Coming Soon)'

        # Determine source for tracking
        if prop.get('market') and prop.get('region'):
            data_source = 'Registry'
        elif prop.get('market') or prop.get('region'):
            data_source = 'Registry (Partial)'
        else:
            data_source = 'Placeholder'

        properties[name.lower()] = {
            'name': name,
            'market': market,
            'region': region,
            'ga4_id': prop.get('ga4_property_id', ''),
            'feed_id': prop.get('feed_id', ''),
            'market_region_source': data_source
        }

    return properties


def collect_google_ads_data(property_registry: Dict) -> Dict:
    """
    Collect Google Ads data for all properties over 30-day window

    Returns dict: {property_name: {spend_data, targeting_data, performance_data}}
    """
    client = GoogleAdsClient.load_from_storage(str(materialize_google_ads_yaml()))
    ga_service = client.get_service("GoogleAdsService")

    # Date range
    end_date = datetime.now()
    start_date = end_date - timedelta(days=TIME_WINDOW_DAYS)
    date_str_start = start_date.strftime("%Y-%m-%d")
    date_str_end = end_date.strftime("%Y-%m-%d")

    print(f"📅 Collecting Google Ads data: {date_str_start} to {date_str_end}")

    # Query all campaigns
    query = f"""
        SELECT
            campaign.name,
            ad_group_criterion.keyword.text,
            metrics.impressions,
            metrics.clicks,
            metrics.cost_micros,
            metrics.conversions
        FROM keyword_view
        WHERE segments.date BETWEEN '{date_str_start}' AND '{date_str_end}'
    """

    response = ga_service.search(customer_id=CUSTOMER_ID, query=query)

    # Initialize data structure for all properties
    property_data = {}
    for prop_name in property_registry.keys():
        property_data[prop_name] = {
            'total_spend': 0,
            'total_clicks': 0,
            'total_conversions': 0,
            'categories': {
                'Studio': {'spend': 0, 'clicks': 0},
                '1BR': {'spend': 0, 'clicks': 0},
                '2BR': {'spend': 0, 'clicks': 0},
                'Unclassified': {'spend': 0, 'clicks': 0}
            },
            'subtypes': {
                # Classified subtypes
                'Studio': {'spend': 0},
                '1BR': {'spend': 0},
                '2BR': {'spend': 0},
                # Unclassified subtypes
                'Brand': {'spend': 0},
                'Competitor': {'spend': 0},
                'Local Generic': {'spend': 0},
                'Other Generic': {'spend': 0}
            }
        }

    # Process keywords
    rows_processed = 0
    for row in response:
        campaign_name = row.campaign.name.lower()
        keyword = row.ad_group_criterion.keyword.text
        cost = row.metrics.cost_micros / 1_000_000
        clicks = row.metrics.clicks
        conversions = row.metrics.conversions

        # Match campaign to property
        matched_property = None
        for prop_name in property_registry.keys():
            if prop_name in campaign_name:
                matched_property = prop_name
                break

        if matched_property:
            # Classify keyword (floor plan level)
            category = classify_keyword(keyword)

            # Classify subtype (granular level)
            if category in ['Studio', '1BR', '2BR']:
                subtype = category  # Floor plan keywords map directly
            else:
                # Get property name for brand matching
                prop_name = property_registry[matched_property]['name']
                subtype = classify_unclassified_keyword(keyword, prop_name)

            # Aggregate data
            property_data[matched_property]['total_spend'] += cost
            property_data[matched_property]['total_clicks'] += clicks
            property_data[matched_property]['total_conversions'] += conversions
            property_data[matched_property]['categories'][category]['spend'] += cost
            property_data[matched_property]['categories'][category]['clicks'] += clicks
            property_data[matched_property]['subtypes'][subtype]['spend'] += cost

            rows_processed += 1

    print(f"✓ Processed {rows_processed} keyword rows")

    return property_data


def collect_availability_data() -> Dict:
    """
    Collect availability data from Venterra feed

    Returns dict: {property_name_lower: {occupancy, units_available, floor_plan_distribution}}
    """
    print("📊 Fetching availability data...")
    response = requests.get(AVAILABILITY_FEED_URL)
    data = response.json()

    availability = {}
    for property_data in data:
        prop_name = property_data.get('name', '')
        if not prop_name:
            continue

        total_units = property_data.get('apartmentCount', 0)

        # Calculate availability by floor plan
        floor_plans = {'Studio': 0, '1BR': 0, '2BR': 0}
        for fp in property_data.get('floorplans', []):
            bedrooms = fp.get('bedrooms', 0)
            units_now = fp.get('unitsAvailable', 0)
            units_30 = fp.get('unitsAvailable30', 0)
            total = units_now + units_30

            if bedrooms == 0:
                floor_plans['Studio'] += total
            elif bedrooms == 1:
                floor_plans['1BR'] += total
            elif bedrooms == 2:
                floor_plans['2BR'] += total

        total_available = sum(floor_plans.values())
        # Calculate as decimal (0-1) since Excel percentage format will multiply by 100
        occupancy = ((total_units - total_available) / total_units) if total_units > 0 else 0

        # Use property name (lowercase) as key for matching
        availability[prop_name.lower()] = {
            'total_units': total_units,
            'total_available': total_available,
            'occupancy': occupancy,
            'floor_plans': floor_plans
        }

    print(f"✓ Loaded availability for {len(availability)} properties")

    return availability


# ============================================================================
# ALIGNMENT CLASSIFICATION
# ============================================================================

def calculate_alignment_status(google_ads_data: Dict, availability_data: Dict) -> str:
    """
    Determine alignment status: 'Aligned' | 'Partially Aligned' | 'Not Targeted'

    Logic per contract:
    - Not Targeted: Classified spend = 0%
    - Aligned: Classified ≥ 15% AND max delta ≤ 15%
    - Partially Aligned: Classified ≥ 15% AND max delta > 15%
    """
    total_spend = google_ads_data['total_spend']
    if total_spend == 0:
        return 'Not Targeted'

    # Calculate classified spend percentage
    classified_spend = sum(
        google_ads_data['categories'][cat]['spend']
        for cat in ['Studio', '1BR', '2BR']
    )
    classified_pct = (classified_spend / total_spend * 100) if total_spend > 0 else 0

    # Not Targeted if no floor plan spend
    if classified_pct < 0.01:  # effectively 0%
        return 'Not Targeted'

    # If classified < 15%, also Not Targeted
    if classified_pct < 15:
        return 'Not Targeted'

    # Check availability data
    if not availability_data or availability_data['total_available'] == 0:
        # Can't determine alignment without availability
        return 'Partially Aligned'

    # Calculate floor plan distribution percentages
    spend_dist = {}
    avail_dist = {}

    for floor_plan in ['Studio', '1BR', '2BR']:
        # Spend distribution (as % of classified spend, not total)
        fp_spend = google_ads_data['categories'][floor_plan]['spend']
        spend_dist[floor_plan] = (fp_spend / classified_spend * 100) if classified_spend > 0 else 0

        # Availability distribution
        fp_avail = availability_data['floor_plans'].get(floor_plan, 0)
        avail_dist[floor_plan] = (fp_avail / availability_data['total_available'] * 100) if availability_data['total_available'] > 0 else 0

    # Calculate deltas
    deltas = []
    for floor_plan in ['Studio', '1BR', '2BR']:
        delta = abs(spend_dist[floor_plan] - avail_dist[floor_plan])
        deltas.append(delta)

    max_delta = max(deltas)

    # Alignment decision
    if max_delta <= 15:
        return 'Aligned'
    else:
        return 'Partially Aligned'


# ============================================================================
# EXCEL GENERATION
# ============================================================================

def generate_excel_workbook(property_registry: Dict, google_ads_data: Dict, availability_lookup: Dict):
    """Generate Excel workbook with Paid_Media_Overview worksheet"""

    wb = Workbook()
    ws = wb.active
    ws.title = "Paid_Media_Overview"

    # Define headers
    headers = [
        # Property Context
        "Property Name",
        "Market",
        "Region",
        # Spend Overview
        "Total Ad Spend ($)",
        "Classified Spend ($)",
        "Classified Spend (%)",
        "Generic Spend ($)",
        "Generic Spend (%)",
        # Targeting Distribution
        "Floor Plans Targeted",
        "% Spend on Studio",
        "% Spend on 1BR",
        "% Spend on 2BR",
        # Performance
        "Clicks",
        "CPC ($)",
        "Conversions",
        "Cost per Conversion ($)",
        # Inventory Context
        "Occupancy (%)",
        "Units Available",
        "% 1BR Available",
        "% 2BR Available",
        # Alignment
        "Targeting Status",
        # Data Quality
        "Market/Region Source"
    ]

    # Write headers
    ws.append(headers)

    # Format header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Write data rows
    row_num = 2
    for prop_name_lower, prop_info in sorted(property_registry.items()):
        ads_data = google_ads_data.get(prop_name_lower, {})
        # Match availability by property name (both lowercase)
        avail_data = availability_lookup.get(prop_name_lower, {})

        # Calculate metrics
        total_spend = ads_data.get('total_spend', 0)
        total_clicks = ads_data.get('total_clicks', 0)
        total_conversions = ads_data.get('total_conversions', 0)

        categories = ads_data.get('categories', {
            'Studio': {'spend': 0}, '1BR': {'spend': 0}, '2BR': {'spend': 0}, 'Unclassified': {'spend': 0}
        })

        classified_spend = sum(categories[cat]['spend'] for cat in ['Studio', '1BR', '2BR'])
        generic_spend = total_spend - classified_spend

        classified_pct = (classified_spend / total_spend) if total_spend > 0 else 0
        generic_pct = (generic_spend / total_spend) if total_spend > 0 else 0

        # Floor plans targeted
        targeted_plans = [fp for fp in ['Studio', '1BR', '2BR'] if categories[fp]['spend'] > 0]
        floor_plans_targeted = ', '.join(targeted_plans) if targeted_plans else 'None'

        # Spend distribution
        studio_pct = (categories['Studio']['spend'] / total_spend) if total_spend > 0 else None
        one_br_pct = (categories['1BR']['spend'] / total_spend) if total_spend > 0 else None
        two_br_pct = (categories['2BR']['spend'] / total_spend) if total_spend > 0 else None

        # Performance
        cpc = (total_spend / total_clicks) if total_clicks > 0 else None
        cost_per_conversion = (total_spend / total_conversions) if total_conversions > 0 else None

        # Availability
        occupancy = avail_data.get('occupancy', None)
        units_available = avail_data.get('total_available', None)

        total_avail = avail_data.get('total_available', 0)
        one_br_avail_pct = (avail_data.get('floor_plans', {}).get('1BR', 0) / total_avail) if total_avail > 0 else None
        two_br_avail_pct = (avail_data.get('floor_plans', {}).get('2BR', 0) / total_avail) if total_avail > 0 else None

        # Alignment
        targeting_status = calculate_alignment_status(ads_data, avail_data)

        # Build row
        row_data = [
            prop_info['name'],  # Property Name
            prop_info['market'],  # Market
            prop_info['region'],  # Region
            total_spend,  # Total Ad Spend
            classified_spend,  # Classified Spend
            classified_pct,  # Classified Spend %
            generic_spend,  # Generic Spend
            generic_pct,  # Generic Spend %
            floor_plans_targeted,  # Floor Plans Targeted
            studio_pct,  # % Spend on Studio
            one_br_pct,  # % Spend on 1BR
            two_br_pct,  # % Spend on 2BR
            total_clicks if total_clicks > 0 else None,  # Clicks
            cpc,  # CPC
            None,  # Conversions (leave blank per contract)
            None,  # Cost per Conversion (leave blank)
            occupancy,  # Occupancy %
            units_available,  # Units Available
            one_br_avail_pct,  # % 1BR Available
            two_br_avail_pct,  # % 2BR Available
            targeting_status,  # Targeting Status
            prop_info['market_region_source']  # Market/Region Source
        ]

        ws.append(row_data)
        row_num += 1

    # Apply number formatting
    for row in ws.iter_rows(min_row=2, max_row=row_num-1):
        # Currency columns: D, E, G, N
        for col_idx in [3, 4, 6, 13]:  # 0-indexed: 3=D, 4=E, 6=G, 13=N
            cell = row[col_idx]
            if cell.value is not None:
                cell.number_format = '$#,##0.00'

        # Percentage columns: F, H, J, K, L, Q, S, T
        for col_idx in [5, 7, 9, 10, 11, 16, 18, 19]:
            cell = row[col_idx]
            if cell.value is not None:
                cell.number_format = '0.0%'

        # Integer columns: M, R
        for col_idx in [12, 17]:
            cell = row[col_idx]
            if cell.value is not None:
                cell.number_format = '#,##0'

    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Prepare output file path (don't save yet, return wb for additional worksheets)
    today = datetime.now().strftime("%Y-%m-%d")
    output_file = os.path.join(OUTPUT_DIR, f"paid_media_workbook_{today}_v1.2.xlsx")

    print(f"\n✓ Paid_Media_Overview worksheet generated")

    return wb, output_file


def generate_spend_breakdown_worksheet(wb: Workbook, property_registry: Dict, google_ads_data: Dict):
    """
    Generate Spend_Breakdown worksheet with granular spend classification

    One row per Property × Spend Subtype
    """
    ws = wb.create_sheet("Spend_Breakdown")

    # Define headers
    headers = [
        "Property Name",
        "Spend Type",
        "Spend Subtype",
        "Spend ($)",
        "% of Total Spend",
        "% of Classified Spend",
        "Lookback Window"
    ]

    # Write headers
    ws.append(headers)

    # Format header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Write data rows
    row_num = 2
    for prop_name_lower, prop_info in sorted(property_registry.items()):
        ads_data = google_ads_data.get(prop_name_lower, {})
        total_spend = ads_data.get('total_spend', 0)

        if total_spend == 0:
            continue  # Skip properties with no spend

        subtypes_data = ads_data.get('subtypes', {})
        categories = ads_data.get('categories', {})

        # Calculate classified spend for percentage calculations
        classified_spend = sum(categories[cat]['spend'] for cat in ['Studio', '1BR', '2BR'])

        # Classified subtypes
        for subtype in ['Studio', '1BR', '2BR']:
            spend = subtypes_data.get(subtype, {}).get('spend', 0)
            if spend > 0:
                pct_of_total = (spend / total_spend) if total_spend > 0 else 0
                pct_of_classified = (spend / classified_spend) if classified_spend > 0 else 0

                row_data = [
                    prop_info['name'],
                    'Classified',
                    subtype,
                    spend,
                    pct_of_total,
                    pct_of_classified,
                    'Last 30 Days'
                ]
                ws.append(row_data)
                row_num += 1

        # Unclassified subtypes
        for subtype in ['Brand', 'Competitor', 'Local Generic', 'Other Generic']:
            spend = subtypes_data.get(subtype, {}).get('spend', 0)
            if spend > 0:
                pct_of_total = (spend / total_spend) if total_spend > 0 else 0

                row_data = [
                    prop_info['name'],
                    'Unclassified',
                    subtype,
                    spend,
                    pct_of_total,
                    None,  # Leave blank for unclassified
                    'Last 30 Days'
                ]
                ws.append(row_data)
                row_num += 1

    # Apply number formatting
    for row in ws.iter_rows(min_row=2, max_row=row_num-1):
        # Currency column: D
        cell = row[3]
        if cell.value is not None:
            cell.number_format = '$#,##0.00'

        # Percentage columns: E, F
        for col_idx in [4, 5]:
            cell = row[col_idx]
            if cell.value is not None:
                cell.number_format = '0.0%'

    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[column_letter].width = adjusted_width

    print("✓ Spend_Breakdown worksheet generated")


# ============================================================================
# MAIN
# ============================================================================

def main():
    print()
    print("=" * 80)
    print("PAID MEDIA PERFORMANCE WORKBOOK GENERATOR")
    print("30-Day Google Ads Performance for Venterra Properties")
    print("=" * 80)
    print()

    # Load property registry
    print("📚 Loading property registry...")
    property_registry = load_property_registry()
    print(f"✓ Loaded {len(property_registry)} properties")

    # Collect Google Ads data
    google_ads_data = collect_google_ads_data(property_registry)

    # Collect availability data
    availability_lookup = collect_availability_data()

    # Generate Excel workbook
    print()
    print("📊 Generating Excel workbook...")
    wb, output_file = generate_excel_workbook(property_registry, google_ads_data, availability_lookup)

    # Add Spend_Breakdown worksheet
    print("📊 Generating Spend_Breakdown worksheet...")
    generate_spend_breakdown_worksheet(wb, property_registry, google_ads_data)

    # Save final workbook
    wb.save(output_file)
    print(f"✓ Workbook finalized: {output_file}")

    print()
    print("=" * 80)
    print("✓ WORKBOOK GENERATION COMPLETE")
    print("=" * 80)
    print()
    print(f"Output: {output_file}")
    print()

    return 0


if __name__ == '__main__':
    sys.exit(main())
