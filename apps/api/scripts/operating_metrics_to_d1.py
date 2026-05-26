#!/usr/bin/env python3
"""
Ingest official property operating metrics into the Pond and optional remote D1.

This is the source-of-record lane for Captain Brief operating facts:
occupancy, leased percentage, lease/cancel counts, and booked concession dollars.
It intentionally does not infer those values from AptIQ or public unit-feed data.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sqlite3
import subprocess
import sys
import tempfile
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from wrangler_auth import build_runtime_env, npx_wrangler_prefix

SCRIPT_DIR = Path(__file__).resolve().parent
API_DIR = SCRIPT_DIR.parent
REPO_ROOT = API_DIR.parent.parent
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from Data_Collection.utils.property_identity import resolve_property_identity  # noqa: E402

CANONICAL_DB = REPO_ROOT / "data" / "portfolio_analytics.db"
MIGRATION_SQL = API_DIR / "migrations" / "0028_create_property_operating_metrics.sql"
WRANGLER_TOML = API_DIR / "wrangler.toml"
GENERATED_DIR = SCRIPT_DIR / "generated"

DEFAULT_PROPERTY_KEY = "AR4PB"
DEFAULT_SOURCE_SYSTEM = "official_operating_feed"

HEADER_ALIASES = {
    "property_id": {
        "propertyid",
        "propertycode",
        "propertycodeid",
        "property",
        "communitycode",
        "communityid",
        "code",
    },
    "property_name": {"propertyname", "communityname", "community", "asset", "name"},
    "community_id": {"pondcommunityid", "communityuuid", "d1communityid"},
    "metric_date": {"metricdate", "reportdate", "date", "asofdate", "snapshotdate", "activitydate"},
    "period_start": {"periodstart", "startdate", "fromdate", "beginningdate"},
    "period_end": {"periodend", "enddate", "todate", "throughdate", "endingdate"},
    "occupancy_rate": {"occupancy", "occupancyrate", "occupiedpercent", "physicaloccupancy"},
    "leased_rate": {"leased", "leasedrate", "leasedpercent", "percentleased", "economicoccupancy"},
    "occupied_units": {"occupiedunits", "occupiedunitcount", "occunits"},
    "leased_units": {"leasedunits", "leasedunitcount", "leasecountunits"},
    "available_units": {"availableunits", "availableunitcount", "vacantavailableunits"},
    "total_units": {"totalunits", "unitcount", "units", "numberofunits"},
    "leases_count": {"leases", "leasecount", "leasescount", "signedleases", "newleases"},
    "cancellations_count": {"cancellations", "cancellationcount", "cancelcount", "cancels"},
    "denials_count": {"denials", "denialcount", "denied"},
    "move_ins_count": {"moveins", "moveincount", "moveinscount"},
    "move_outs_count": {"moveouts", "moveoutcount", "moveoutscount"},
    "booked_concession_dollars": {
        "bookedconcessiondollars",
        "bookedconcessions",
        "concessiondollars",
        "concessionsdollars",
        "actualconcessiondollars",
        "leaseconcessiondollars",
    },
    "booked_concession_lease_count": {
        "bookedconcessionleasecount",
        "concessionleasecount",
        "leaseswithconcessions",
        "concessionleases",
    },
}

INT_FIELDS = {
    "occupied_units",
    "leased_units",
    "available_units",
    "total_units",
    "leases_count",
    "cancellations_count",
    "denials_count",
    "move_ins_count",
    "move_outs_count",
    "booked_concession_lease_count",
}
REAL_FIELDS = {"occupancy_rate", "leased_rate", "booked_concession_dollars"}
DATE_FIELDS = {"metric_date", "period_start", "period_end"}

TABLE_COLUMNS = [
    "id",
    "property_id",
    "community_id",
    "metric_date",
    "period_start",
    "period_end",
    "occupancy_rate",
    "leased_rate",
    "occupied_units",
    "leased_units",
    "available_units",
    "total_units",
    "leases_count",
    "cancellations_count",
    "denials_count",
    "move_ins_count",
    "move_outs_count",
    "booked_concession_dollars",
    "booked_concession_lease_count",
    "source_system",
    "source_file",
    "evidence_json",
]


def normalize_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def canonical_header(raw_header: Any) -> str | None:
    normalized = normalize_key(raw_header)
    if not normalized:
        return None
    for canonical, aliases in HEADER_ALIASES.items():
        if normalized == canonical.replace("_", "") or normalized in aliases:
            return canonical
    return None


def is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def parse_date(value: Any) -> str | None:
    if is_blank(value):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%Y/%m/%d", "%m-%d-%Y", "%m-%d-%y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).date().isoformat()
    except ValueError as exc:
        raise ValueError(f"Could not parse date value {text!r}") from exc


def parse_number(value: Any) -> float | None:
    if is_blank(value):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    negative = text.startswith("(") and text.endswith(")")
    text = text.strip("()").replace("$", "").replace(",", "").replace("%", "").strip()
    if not text:
        return None
    parsed = float(text)
    return -parsed if negative else parsed


def parse_rate(value: Any) -> float | None:
    if is_blank(value):
        return None
    has_percent = isinstance(value, str) and "%" in value
    parsed = parse_number(value)
    if parsed is None:
        return None
    if has_percent or parsed > 1:
        return parsed / 100
    return parsed


def parse_int(value: Any) -> int | None:
    parsed = parse_number(value)
    if parsed is None:
        return None
    return int(round(parsed))


def read_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as fp:
        return [dict(row) for row in csv.DictReader(fp)]


def read_xlsx(path: Path, sheet_name: str | None) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for XLSX input. Use CSV or run with the bundled workspace Python.") from exc

    workbook = load_workbook(path, data_only=True, read_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook[workbook.sheetnames[0]]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(cell or "").strip() for cell in rows[0]]
    records: list[dict[str, Any]] = []
    for row in rows[1:]:
        if not row or all(is_blank(cell) for cell in row):
            continue
        records.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
    return records


def read_source(path: Path, sheet_name: str | None) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv(path)
    if suffix in {".xlsx", ".xlsm"}:
        return read_xlsx(path, sheet_name)
    raise ValueError(f"Unsupported source file type: {path.suffix}. Use .csv, .xlsx, or .xlsm.")


def row_matches_property(row: dict[str, Any], property_code: str, property_name: str | None, aliases: set[str] | None = None) -> bool:
    candidates = []
    for key, value in row.items():
        canonical = canonical_header(key)
        if canonical in {"property_id", "property_name"} and not is_blank(value):
            candidates.append(str(value).strip().lower())
    if not candidates:
        return True
    expected = {property_code.lower()}
    if property_name:
        expected.add(property_name.lower())
    if aliases:
        expected.update(alias.lower() for alias in aliases if alias)
    return any(candidate in expected for candidate in candidates)


def normalize_row(
    row: dict[str, Any],
    *,
    property_code: str,
    property_name: str | None,
    community_id: str | None,
    aliases: set[str] | None,
    source_system: str,
    source_file: Path,
) -> dict[str, Any] | None:
    if not row_matches_property(row, property_code, property_name, aliases):
        return None

    normalized: dict[str, Any] = {
        "property_id": property_code,
        "community_id": community_id,
        "source_system": source_system,
        "source_file": str(source_file),
        "evidence_json": json.dumps({"raw": row}, sort_keys=True, default=str),
    }

    for key, value in row.items():
        canonical = canonical_header(key)
        if not canonical or canonical in {"property_id", "property_name", "community_id"}:
            if canonical == "community_id" and not is_blank(value):
                normalized["community_id"] = str(value).strip()
            continue
        if canonical in DATE_FIELDS:
            normalized[canonical] = parse_date(value)
        elif canonical in {"occupancy_rate", "leased_rate"}:
            normalized[canonical] = parse_rate(value)
        elif canonical in INT_FIELDS:
            normalized[canonical] = parse_int(value)
        elif canonical in REAL_FIELDS:
            normalized[canonical] = parse_number(value)

    metric_date = normalized.get("metric_date") or normalized.get("period_end")
    if not metric_date:
        raise ValueError(f"Matched row is missing metric/report date: {row}")
    normalized["metric_date"] = metric_date
    normalized["id"] = f"op_{property_code}_{metric_date}_{normalize_key(source_system) or 'source'}"

    for column in TABLE_COLUMNS:
        normalized.setdefault(column, None)
    return normalized


def load_metrics(args: argparse.Namespace) -> list[dict[str, Any]]:
    source = Path(args.source_file).expanduser().resolve()
    records = read_source(source, args.sheet)
    property_key = getattr(args, "property_key", None) or getattr(args, "property_code", None) or DEFAULT_PROPERTY_KEY
    identity = resolve_property_identity(property_key)
    property_code = getattr(args, "property_code", None) or (identity.marketing_bi_property_id if identity else property_key)
    property_name = getattr(args, "property_name", None) or (identity.property_name if identity else None)
    community_id = getattr(args, "community_id", None) or (identity.community_id if identity else None)
    aliases = set(identity.aliases) if identity else set()
    if identity:
        aliases.update(
            value
            for value in [
                identity.property_name,
                identity.community_name,
                identity.encasa_short_name,
                identity.ga4_property_id,
                identity.property_code,
            ]
            if value
        )
    metrics = [
        metric
        for row in records
        if (
            metric := normalize_row(
                row,
                property_code=property_code,
                property_name=property_name,
                community_id=community_id,
                aliases=aliases,
                source_system=args.source_system,
                source_file=source,
            )
        )
    ]
    if not metrics:
        raise ValueError(f"No rows matched property {property_code} in {source}")
    return metrics


def create_table_sql() -> str:
    return MIGRATION_SQL.read_text(encoding="utf-8")


def sql_literal(value: Any) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, (int, float)):
        return str(value)
    return "'" + str(value).replace("'", "''") + "'"


def insert_sql(metrics: Iterable[dict[str, Any]]) -> str:
    columns = ", ".join(TABLE_COLUMNS)
    statements = [create_table_sql().strip()]
    update_columns = [col for col in TABLE_COLUMNS if col != "id"]
    for metric in metrics:
        values = ", ".join(sql_literal(metric.get(column)) for column in TABLE_COLUMNS)
        updates = ", ".join(f"{column} = excluded.{column}" for column in update_columns)
        statements.append(
            f"INSERT INTO property_operating_metrics ({columns}) VALUES ({values}) "
            f"ON CONFLICT(property_id, metric_date, source_system) DO UPDATE SET {updates}, updated_at = datetime('now');"
        )
    return "\n".join(statements) + "\n"


def write_local(metrics: list[dict[str, Any]], db_path: Path) -> None:
    conn = sqlite3.connect(db_path)
    try:
        conn.executescript(create_table_sql())
        placeholders = ", ".join("?" for _ in TABLE_COLUMNS)
        updates = ", ".join(f"{column} = excluded.{column}" for column in TABLE_COLUMNS if column != "id")
        conn.executemany(
            f"""
            INSERT INTO property_operating_metrics ({", ".join(TABLE_COLUMNS)})
            VALUES ({placeholders})
            ON CONFLICT(property_id, metric_date, source_system)
            DO UPDATE SET {updates}, updated_at = datetime('now')
            """,
            [[metric.get(column) for column in TABLE_COLUMNS] for metric in metrics],
        )
        conn.commit()
    finally:
        conn.close()


def execute_remote(sql: str) -> None:
    GENERATED_DIR.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile("w", suffix=".sql", delete=False, dir=GENERATED_DIR) as fp:
        fp.write(sql)
        temp_path = Path(fp.name)
    env = build_runtime_env()
    cmd = [
        *npx_wrangler_prefix(env),
        "d1",
        "execute",
        "pop-brief-db",
        "--remote",
        "--file",
        str(temp_path),
        "--config",
        str(WRANGLER_TOML),
    ]
    result = subprocess.run(cmd, cwd=str(API_DIR), env=env, text=True, capture_output=True, timeout=1800)
    print(result.stdout)
    if result.stderr:
        print(result.stderr, file=sys.stderr)
    if result.returncode != 0:
        raise SystemExit(result.returncode)


def summarize(metrics: list[dict[str, Any]], *, dry_run: bool, remote: bool) -> dict[str, Any]:
    dates = sorted({str(metric["metric_date"]) for metric in metrics})
    return {
        "rows": len(metrics),
        "property_id": metrics[0]["property_id"],
        "dates": dates,
        "latest_metric_date": dates[-1] if dates else None,
        "source_system": metrics[0]["source_system"],
        "dry_run": dry_run,
        "remote": remote,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Ingest official property operating metrics into local Pond DB and optional remote D1.")
    parser.add_argument("--source-file", required=True, help="CSV/XLSX/XLSM file containing official operating metrics.")
    parser.add_argument("--sheet", help="Workbook sheet name for XLSX/XLSM input. Defaults to the first sheet.")
    parser.add_argument("--property-key", default=DEFAULT_PROPERTY_KEY, help="Any known property identity: property code, GA4 id, GSC URL, name, or alias.")
    parser.add_argument("--property-code", help="Optional override. Defaults from the identity matrix.")
    parser.add_argument("--property-name", help="Optional override. Defaults from the identity matrix.")
    parser.add_argument("--community-id", help="Optional override. Defaults from the identity matrix when known.")
    parser.add_argument("--source-system", default=DEFAULT_SOURCE_SYSTEM)
    parser.add_argument("--db", default=str(CANONICAL_DB), help="Local SQLite database path.")
    parser.add_argument("--remote", action="store_true", help="Also upsert matched rows into remote D1.")
    parser.add_argument("--dry-run", action="store_true", help="Parse and print summary without writing.")
    args = parser.parse_args()

    metrics = load_metrics(args)
    print(json.dumps(summarize(metrics, dry_run=args.dry_run, remote=args.remote), indent=2))
    if args.dry_run:
        return

    write_local(metrics, Path(args.db).expanduser().resolve())
    if args.remote:
        execute_remote(insert_sql(metrics))


if __name__ == "__main__":
    main()
