#!/usr/bin/env python3
"""Governed orchestration for ad hoc executive report generation and delivery."""

from __future__ import annotations

import json
import sys
from dataclasses import asdict
from datetime import date, datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from adhoc_report_sources import ReportRequest, build_report
from email_sender import EmailSender
from outlook_email_validator import validate_outlook_email_file, write_validation_result
from outlook_report_builder import BRAND, render_outlook_report


ROOT = Path("/Users/mark/Property_Analytics")
RUN_ROOT = ROOT / "reports" / "adhoc_executive"
DELIVERY_DIR = ROOT / "logs" / "email_delivery"


def slugify(value: str) -> str:
    cleaned = "".join(ch.lower() if ch.isalnum() else "-" for ch in value).strip("-")
    while "--" in cleaned:
        cleaned = cleaned.replace("--", "-")
    return cleaned[:80] or "adhoc-report"


def autosize(ws) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[letter].width = min(max(width + 2, 12), 48)


def write_workbook(path: Path, metadata: dict[str, object], sheets: dict[str, list[dict[str, object]]]) -> None:
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary["A1"] = str(metadata.get("title", "Ad Hoc Executive Report"))
    summary["A1"].font = Font(bold=True, size=16, color=BRAND["navy"].replace("#", ""))
    row_idx = 3
    for key, value in metadata.items():
        summary.cell(row=row_idx, column=1, value=key).font = Font(bold=True, color=BRAND["bay"].replace("#", ""))
        summary.cell(row=row_idx, column=2, value=json.dumps(value, ensure_ascii=True) if isinstance(value, (dict, list)) else value)
        row_idx += 1
    autosize(summary)

    header_fill = PatternFill("solid", fgColor=BRAND["navy"].replace("#", ""))
    header_font = Font(color=BRAND["white"].replace("#", ""), bold=True)
    for title, rows in sheets.items():
        ws = wb.create_sheet(title[:31])
        if not rows:
            ws["A1"] = "No rows"
            continue
        keys = list(rows[0].keys())
        for col_idx, key in enumerate(keys, start=1):
            cell = ws.cell(row=1, column=col_idx, value=key)
            cell.fill = header_fill
            cell.font = header_font
        for row_num, row in enumerate(rows, start=2):
            for col_idx, key in enumerate(keys, start=1):
                ws.cell(row=row_num, column=col_idx, value=row.get(key))
        ws.freeze_panes = "A2"
        autosize(ws)
    wb.save(path)


def write_sources(path: Path, spec: dict[str, object]) -> None:
    sources = spec.get("sources", [])
    lines = ["# Sources Used", ""]
    if isinstance(sources, list):
        for source in sources:
            lines.append(f"- {source}")
    lines.extend(["", "## Spec", "", "```json", json.dumps(spec, indent=2, ensure_ascii=True), "```", ""])
    path.write_text("\n".join(lines), encoding="utf-8")


def run_adhoc_report(
    request: ReportRequest,
    *,
    email: bool = False,
    recipients: Optional[list[str]] = None,
    provider: str = "aws_ses",
) -> dict[str, object]:
    """Generate, validate, optionally email, and archive an ad hoc report packet."""

    build = build_report(request)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    run_dir = RUN_ROOT / build.report_type / f"{timestamp}_{slugify(request.subject)}"
    run_dir.mkdir(parents=True, exist_ok=True)

    request_path = run_dir / "request.json"
    spec_path = run_dir / "report_spec.json"
    html_path = run_dir / "report.html"
    workbook_path = run_dir / "report.xlsx"
    validation_path = run_dir / "validation.json"
    delivery_path = run_dir / "delivery.json"
    sources_path = run_dir / "sources_used.md"

    request_path.write_text(json.dumps(asdict(request), indent=2, ensure_ascii=True) + "\n", encoding="utf-8")
    spec_path.write_text(json.dumps(build.spec, indent=2, ensure_ascii=True) + "\n", encoding="utf-8")
    html_path.write_text(render_outlook_report(build.report), encoding="utf-8")
    write_sources(sources_path, build.spec)

    if request.include_workbook:
        write_workbook(
            workbook_path,
            {
                "title": build.report.title,
                "report_type": build.report_type,
                "date_range": build.report.date_range,
                "generated_at": build.report.generated_at,
                "subject": request.subject,
            },
            build.workbook_sheets,
        )

    validation = validate_outlook_email_file(html_path)
    write_validation_result(validation_path, validation)
    if not validation.passed:
        delivery = {
            "success": False,
            "status": "blocked_by_validation",
            "errors": validation.errors,
            "created_at": datetime.now().isoformat(timespec="seconds"),
        }
        delivery_path.write_text(json.dumps(delivery, indent=2, ensure_ascii=True) + "\n", encoding="utf-8")
        return {
            "success": False,
            "run_dir": str(run_dir),
            "html_path": str(html_path),
            "workbook_path": str(workbook_path) if workbook_path.exists() else None,
            "validation": validation.to_dict(),
            "delivery": delivery,
        }

    delivery: dict[str, object] = {
        "success": True,
        "status": "not_sent",
        "created_at": datetime.now().isoformat(timespec="seconds"),
    }
    if email:
        sender = EmailSender(provider=provider, verbose=False)
        attachments = []
        if workbook_path.exists():
            attachments.append(
                (
                    workbook_path.name,
                    workbook_path.read_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            )
        metadata = sender.send_email_with_tracking(
            subject=build.report.title,
            html_body=html_path.read_text(encoding="utf-8"),
            plain_text=f"{build.report.title}\n\nGenerated {build.report.generated_at}.",
            recipients=recipients,
            attachments=attachments,
            log_path=DELIVERY_DIR / f"email_delivery_{date.today().isoformat()}.jsonl",
        )
        delivery = metadata

    delivery_path.write_text(json.dumps(delivery, indent=2, ensure_ascii=True) + "\n", encoding="utf-8")
    return {
        "success": bool(delivery.get("success")),
        "run_dir": str(run_dir),
        "html_path": str(html_path),
        "workbook_path": str(workbook_path) if workbook_path.exists() else None,
        "validation": validation.to_dict(),
        "delivery": delivery,
    }


def load_request_from_json(path: Path) -> ReportRequest:
    payload = json.loads(path.read_text(encoding="utf-8"))
    return ReportRequest(**payload)


if __name__ == "__main__":
    raise SystemExit("Use scripts/run_adhoc_report.py")
