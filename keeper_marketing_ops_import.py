#!/usr/bin/env python3
"""One-time spreadsheet-to-Keeper JSON migration utility.

This tool intentionally separates spreadsheet parsing, validation, Keeper JSON
generation, and optional Commander execution. It never logs password values.
"""

from __future__ import annotations

import argparse
import csv
import getpass
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple
from urllib.parse import urlparse

import yaml
from openpyxl import load_workbook


DEFAULT_TEAM_NAME = "Marketing Ops"
DEFAULT_SHARED_FOLDER_NAME = "Marketing Ops Shared Credentials"
DEFAULT_CONFIG = {
    "team_name": DEFAULT_TEAM_NAME,
    "shared_folder_name": DEFAULT_SHARED_FOLDER_NAME,
    "default_permissions": {
        "can_edit": False,
        "can_share": False,
        "manage_users": False,
        "manage_records": False,
    },
    "allow_blank_passwords": False,
    "duplicate_policy": "warn",
    "url_validation": "warn",
}

HEADER_ALIASES = {
    "title": {"title", "name", "record title", "credential name", "account", "service"},
    "username": {"username", "user name", "login", "email", "user", "account login"},
    "password": {"password", "pass", "secret"},
    "url": {"url", "website", "login url", "login_url", "site", "web site"},
    "notes": {"notes", "note", "comments", "comment", "description"},
    "folder": {"folder", "folder path", "path", "category"},
}
REQUIRED_CANONICAL_COLUMNS = {"title", "password"}


@dataclass(frozen=True)
class SheetMapping:
    title: str
    username: Optional[str] = None
    password: Optional[str] = None
    url: Optional[str] = None
    notes: Optional[str] = None
    record_kind: str = "credential"
    password_required: bool = True


VENTERRA_MARKETING_LOGIN_SHEETS: Dict[str, SheetMapping] = {
    "Venterra Master Log ins": SheetMapping(
        title="Site",
        username="Username",
        password="Password",
        url="URL",
    ),
    "Instagram Logins": SheetMapping(
        title="Property",
        username="Username",
        password="Password",
        url="Social URL",
    ),
    "Craigslist -Rooof Log ins": SheetMapping(
        title="Community",
        username="Login",
        password="Password",
    ),
    "Twitter Logins": SheetMapping(
        title="Property",
        username="Twitter Handle",
        password="Password",
    ),
    "Apartment Associations": SheetMapping(
        title="Organization",
        username="User Name",
        password="Password",
    ),
}


VENTERRA_MARKETING_REFERENCE_SHEETS: Dict[str, SheetMapping] = {
    "youtube urls": SheetMapping(
        title="Community",
        url="URL: Community's YouTube Channel",
        record_kind="reference_url",
        password_required=False,
    ),
    "FB URLS": SheetMapping(
        title="PROPERTY",
        url="FB PAGE URL",
        record_kind="reference_url",
        password_required=False,
    ),
    "Yelp Pages ": SheetMapping(
        title="Property Name",
        url="Yelp Page",
        notes="Notes",
        record_kind="reference_url",
        password_required=False,
    ),
}


class MigrationError(Exception):
    """Raised when validation or execution cannot continue safely."""


@dataclass
class ImportConfig:
    team_name: str = DEFAULT_TEAM_NAME
    shared_folder_name: str = DEFAULT_SHARED_FOLDER_NAME
    can_edit: bool = False
    can_share: bool = False
    manage_users: bool = False
    manage_records: bool = False
    allow_blank_passwords: bool = False
    duplicate_policy: str = "warn"
    url_validation: str = "warn"


@dataclass
class CredentialRecord:
    row_number: int
    title: str
    username: str = ""
    password: str = ""
    url: str = ""
    notes: str = ""
    source_folder: str = ""
    custom_fields: Dict[str, str] = field(default_factory=dict)
    password_required: bool = True


@dataclass
class ValidationMessage:
    level: str
    row_number: Optional[int]
    message: str


@dataclass
class ValidationResult:
    errors: List[ValidationMessage] = field(default_factory=list)
    warnings: List[ValidationMessage] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return not self.errors


@dataclass
class MigrationSummary:
    records_seen: int
    records_ready: int
    shared_folder_name: str
    team_name: str
    warnings_count: int
    errors_count: int
    unique_source_folders: int
    keeper_folders_created: int = 1


def normalize_header(header: Any) -> str:
    value = "" if header is None else str(header)
    value = value.strip().lower().replace("_", " ").replace("-", " ")
    value = re.sub(r"\s+", " ", value)
    return value


def canonical_for_header(header: Any) -> Optional[str]:
    normalized = normalize_header(header)
    for canonical, aliases in HEADER_ALIASES.items():
        if normalized in aliases:
            return canonical
    return None


def mask_username(username: str) -> str:
    if not username:
        return ""
    if "@" in username:
        local, domain = username.split("@", 1)
        masked_local = _mask_token(local)
        return f"{masked_local}@{domain}"
    return _mask_token(username)


def _mask_token(token: str) -> str:
    if len(token) <= 2:
        return "*" * len(token)
    return token[0] + ("*" * max(1, len(token) - 2)) + token[-1]


def load_config(path: Optional[Path]) -> ImportConfig:
    raw: Dict[str, Any] = dict(DEFAULT_CONFIG)
    if path:
        with path.open("r", encoding="utf-8") as handle:
            loaded = yaml.safe_load(handle) or {}
        raw = _deep_merge(raw, loaded)

    permissions = raw.get("default_permissions") or {}
    duplicate_policy = str(raw.get("duplicate_policy", "warn")).lower()
    url_validation = str(raw.get("url_validation", "warn")).lower()
    if duplicate_policy not in {"warn", "error"}:
        raise MigrationError("duplicate_policy must be 'warn' or 'error'.")
    if url_validation not in {"warn", "error", "ignore"}:
        raise MigrationError("url_validation must be 'warn', 'error', or 'ignore'.")

    return ImportConfig(
        team_name=str(raw.get("team_name") or DEFAULT_TEAM_NAME),
        shared_folder_name=str(raw.get("shared_folder_name") or DEFAULT_SHARED_FOLDER_NAME),
        can_edit=bool(permissions.get("can_edit", False)),
        can_share=bool(permissions.get("can_share", False)),
        manage_users=bool(permissions.get("manage_users", False)),
        manage_records=bool(permissions.get("manage_records", False)),
        allow_blank_passwords=bool(raw.get("allow_blank_passwords", False)),
        duplicate_policy=duplicate_policy,
        url_validation=url_validation,
    )


def _deep_merge(base: Dict[str, Any], override: Dict[str, Any]) -> Dict[str, Any]:
    merged = dict(base)
    for key, value in override.items():
        if isinstance(value, dict) and isinstance(merged.get(key), dict):
            merged[key] = _deep_merge(merged[key], value)
        else:
            merged[key] = value
    return merged


def read_spreadsheet(path: Path) -> Tuple[List[str], List[Dict[str, str]]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv(path)
    if suffix == ".xlsx":
        return read_xlsx(path)
    raise MigrationError("Input file must be .csv or .xlsx.")


def read_csv(path: Path) -> Tuple[List[str], List[Dict[str, str]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(4096)
        handle.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample) if sample.strip() else csv.excel
        except csv.Error:
            dialect = csv.excel
        reader = csv.DictReader(handle, dialect=dialect)
        headers = list(reader.fieldnames or [])
        rows = [
            {header: _stringify_cell(row.get(header)) for header in headers}
            for row in reader
        ]
    return headers, rows


def read_xlsx(path: Path) -> Tuple[List[str], List[Dict[str, str]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_values = next(rows_iter)
    except StopIteration:
        return [], []
    headers = [_stringify_cell(value) for value in header_values]
    rows = []
    for row_values in rows_iter:
        row = {}
        for index, header in enumerate(headers):
            row[header] = _stringify_cell(row_values[index] if index < len(row_values) else "")
        rows.append(row)
    workbook.close()
    return headers, rows


def read_venterra_marketing_logins_workbook(
    path: Path,
    include_reference_records: bool = False,
) -> Tuple[List[CredentialRecord], List[str]]:
    """Read the legacy multi-sheet Marketing logins workbook.

    The generic importer reads the active sheet and expects canonical headers.
    This workbook has several known sheet layouts, so this mode maps only those
    columns and preserves sheet/row provenance as Keeper custom fields.
    """
    workbook = load_workbook(path, read_only=True, data_only=True)
    mappings = dict(VENTERRA_MARKETING_LOGIN_SHEETS)
    if include_reference_records:
        mappings.update(VENTERRA_MARKETING_REFERENCE_SHEETS)

    records: List[CredentialRecord] = []
    try:
        for sheet_name, mapping in mappings.items():
            if sheet_name not in workbook.sheetnames:
                continue
            sheet = workbook[sheet_name]
            if sheet.sheet_state != "visible":
                continue
            records.extend(_records_from_mapped_sheet(sheet, mapping))
    finally:
        workbook.close()
    return records, ["password", "title"]


def _records_from_mapped_sheet(sheet: Any, mapping: SheetMapping) -> List[CredentialRecord]:
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_values = next(rows_iter)
    except StopIteration:
        return []

    headers = [_stringify_cell(value) for value in header_values]
    header_index = {header: index for index, header in enumerate(headers) if header}
    records: List[CredentialRecord] = []

    for row_number, row_values in enumerate(rows_iter, start=2):
        if not any(_stringify_cell(value) for value in row_values):
            continue

        title = _mapped_sheet_value(row_values, header_index, mapping.title)
        username = _mapped_sheet_value(row_values, header_index, mapping.username)
        password = _mapped_sheet_value(row_values, header_index, mapping.password)
        url = _mapped_sheet_value(row_values, header_index, mapping.url)
        notes = _mapped_sheet_value(row_values, header_index, mapping.notes)

        if mapping.record_kind == "reference_url" and not url:
            continue

        custom_fields: Dict[str, str] = {
            "Source Sheet": sheet.title,
            "Source Row": str(row_number),
            "Source Workbook Type": "Venterra Marketing Logins",
            "Record Kind": mapping.record_kind,
        }
        for header in headers:
            if not header:
                continue
            if header in {mapping.title, mapping.username, mapping.password, mapping.url, mapping.notes}:
                continue
            value = _mapped_sheet_value(row_values, header_index, header)
            if value:
                custom_fields[header] = value
        if mapping.password_required and not password:
            custom_fields["Password Status"] = "Missing in source workbook"

        records.append(
            CredentialRecord(
                row_number=row_number,
                title=title,
                username=username,
                password=password,
                url=url,
                notes=notes,
                source_folder=sheet.title,
                custom_fields=custom_fields,
                password_required=mapping.password_required,
            )
        )
    return records


def _mapped_sheet_value(row_values: Sequence[Any], header_index: Dict[str, int], header: Optional[str]) -> str:
    if not header:
        return ""
    index = header_index.get(header)
    if index is None or index >= len(row_values):
        return ""
    return _stringify_cell(row_values[index])


def _stringify_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def transform_rows(headers: Sequence[str], rows: Sequence[Dict[str, str]]) -> Tuple[List[CredentialRecord], List[str]]:
    header_map: Dict[str, str] = {}
    recognized_originals = set()
    for header in headers:
        canonical = canonical_for_header(header)
        if canonical and canonical not in header_map:
            header_map[canonical] = header
            recognized_originals.add(header)

    records = []
    for offset, row in enumerate(rows, start=2):
        if not any((value or "").strip() for value in row.values()):
            continue
        custom_fields = {}
        for header in headers:
            if header in recognized_originals:
                continue
            value = _stringify_cell(row.get(header))
            if value:
                custom_fields[str(header).strip()] = value

        source_folder = _get_mapped_value(row, header_map, "folder")
        if source_folder:
            custom_fields.setdefault("Source Folder", source_folder)

        records.append(
            CredentialRecord(
                row_number=offset,
                title=_get_mapped_value(row, header_map, "title"),
                username=_get_mapped_value(row, header_map, "username"),
                password=_get_mapped_value(row, header_map, "password"),
                url=_get_mapped_value(row, header_map, "url"),
                notes=_get_mapped_value(row, header_map, "notes"),
                source_folder=source_folder,
                custom_fields=custom_fields,
            )
        )
    return records, sorted(header_map)


def _get_mapped_value(row: Dict[str, str], header_map: Dict[str, str], canonical: str) -> str:
    header = header_map.get(canonical)
    return _stringify_cell(row.get(header)) if header else ""


def validate_records(
    records: Sequence[CredentialRecord],
    mapped_columns: Sequence[str],
    config: ImportConfig,
    verbose_sensitive: bool = False,
) -> ValidationResult:
    result = ValidationResult()
    mapped_set = set(mapped_columns)
    missing = sorted(REQUIRED_CANONICAL_COLUMNS - mapped_set)
    for canonical in missing:
        result.errors.append(
            ValidationMessage("error", None, f"Missing required column: {canonical}.")
        )

    seen: Dict[Tuple[str, str, str], int] = {}
    for record in records:
        if not record.title.strip():
            result.errors.append(ValidationMessage("error", record.row_number, "Title is blank."))
        if record.password_required and not config.allow_blank_passwords and not record.password:
            result.errors.append(
                ValidationMessage("error", record.row_number, "Password is blank.")
            )

        duplicate_key = (
            record.title.strip().lower(),
            record.username.strip().lower(),
            record.url.strip().lower(),
        )
        if duplicate_key in seen:
            username_display = record.username if verbose_sensitive else mask_username(record.username)
            message = (
                "Duplicate title + username + URL combination "
                f"matches row {seen[duplicate_key]}."
            )
            if username_display:
                message += f" Username: {username_display}."
            target = result.errors if config.duplicate_policy == "error" else result.warnings
            target.append(ValidationMessage(config.duplicate_policy, record.row_number, message))
        else:
            seen[duplicate_key] = record.row_number

        if record.url and config.url_validation != "ignore" and not is_probably_valid_url(record.url):
            target = result.errors if config.url_validation == "error" else result.warnings
            target.append(
                ValidationMessage(
                    config.url_validation,
                    record.row_number,
                    "URL appears malformed; it was not modified.",
                )
            )
    return result


def is_probably_valid_url(url: str) -> bool:
    parsed = urlparse(url)
    return parsed.scheme in {"http", "https"} and bool(parsed.netloc)


def build_keeper_import_json(records: Sequence[CredentialRecord], config: ImportConfig) -> Dict[str, Any]:
    shared_folder = {
        "path": config.shared_folder_name,
        "manage_users": config.manage_users,
        "manage_records": config.manage_records,
        "can_edit": config.can_edit,
        "can_share": config.can_share,
        "permissions": [
            {
                "name": config.team_name,
                "manage_users": config.manage_users,
                "manage_records": config.manage_records,
            }
        ],
    }
    keeper_records = []
    for record in records:
        keeper_record: Dict[str, Any] = {
            "title": record.title,
            "login": record.username,
            "password": record.password,
            "login_url": record.url,
            "notes": record.notes,
            "folders": [
                {
                    "shared_folder": config.shared_folder_name,
                    "can_edit": config.can_edit,
                    "can_share": config.can_share,
                }
            ],
        }
        if record.custom_fields:
            keeper_record["custom_fields"] = dict(record.custom_fields)
        keeper_records.append(keeper_record)
    return {"shared_folders": [shared_folder], "records": keeper_records}


def write_keeper_json(payload: Dict[str, Any], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, indent=2, ensure_ascii=False)
        handle.write("\n")
    try:
        os.chmod(path, 0o600)
    except OSError:
        pass


def build_summary(
    records: Sequence[CredentialRecord],
    validation: ValidationResult,
    config: ImportConfig,
) -> MigrationSummary:
    source_folders = {record.source_folder for record in records if record.source_folder}
    return MigrationSummary(
        records_seen=len(records),
        records_ready=0 if validation.errors else len(records),
        shared_folder_name=config.shared_folder_name,
        team_name=config.team_name,
        warnings_count=len(validation.warnings),
        errors_count=len(validation.errors),
        unique_source_folders=len(source_folders),
        keeper_folders_created=1,
    )


def print_validation(validation: ValidationResult) -> None:
    for message in validation.errors + validation.warnings:
        prefix = message.level.upper()
        location = f"row {message.row_number}: " if message.row_number else ""
        print(f"{prefix}: {location}{message.message}", file=sys.stderr)


def print_safe_summary(summary: MigrationSummary, rotate_after_import: bool = False) -> None:
    print("Keeper Marketing Ops migration summary")
    print(f"Records scanned: {summary.records_seen}")
    print(f"Records ready: {summary.records_ready}")
    print(f"Keeper folders created: {summary.keeper_folders_created}")
    print(f"Shared folder: {summary.shared_folder_name}")
    print(f"Team: {summary.team_name}")
    print(f"Warnings: {summary.warnings_count}")
    print(f"Errors: {summary.errors_count}")
    print(f"Source folder values preserved: {summary.unique_source_folders}")
    if rotate_after_import:
        print("Reminder: review and rotate high-risk shared passwords after import.")


def confirm_execution() -> bool:
    print("Type IMPORT to run Keeper Commander import. Anything else cancels.")
    return getpass.getpass("Confirmation: ") == "IMPORT"


def run_keeper_import(json_path: Path, keeper_binary: str = "keeper") -> None:
    executable = shutil.which(keeper_binary) or keeper_binary
    commands = [
        [executable, "import", "--format=json", str(json_path)],
        [executable, "import", "--format=json", "--users", str(json_path)],
    ]
    for command in commands:
        completed = subprocess.run(
            command,
            stdin=subprocess.DEVNULL,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            check=False,
        )
        if completed.returncode != 0:
            raise MigrationError(
                "Keeper Commander import failed. Raw Commander output was suppressed "
                f"to avoid exposing migration data. Exit code: {completed.returncode}."
            )


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate and optionally import Keeper JSON from a local credential spreadsheet."
    )
    parser.add_argument("--input", required=True, type=Path, help="Input .csv or .xlsx file.")
    parser.add_argument("--config", type=Path, help="YAML config file.")
    parser.add_argument("--output", type=Path, help="Path for generated Keeper JSON.")
    parser.add_argument("--dry-run", action="store_true", help="Validate and summarize without importing.")
    parser.add_argument("--execute", action="store_true", help="Run Keeper Commander after validation and confirmation.")
    parser.add_argument(
        "--venterra-marketing-logins-workbook",
        action="store_true",
        help="Read the legacy multi-sheet Venterra Marketing Logins workbook with sheet-specific mappings.",
    )
    parser.add_argument(
        "--include-reference-records",
        action="store_true",
        help="With --venterra-marketing-logins-workbook, include URL-only reference sheets as Keeper records.",
    )
    parser.add_argument(
        "--allow-blank-passwords",
        action="store_true",
        help="Allow records with blank passwords for complete legacy workbook archival.",
    )
    parser.add_argument(
        "--rotate-after-import",
        action="store_true",
        help="Print a safe post-import reminder to rotate high-risk shared passwords.",
    )
    parser.add_argument(
        "--verbose-sensitive",
        action="store_true",
        help="Allow usernames in validation messages. Passwords are never printed.",
    )
    parser.add_argument(
        "--keeper-binary",
        default=os.environ.get("KEEPER_COMMANDER_BINARY", "keeper"),
        help="Keeper Commander binary name/path. Non-secret runtime configuration only.",
    )
    return parser.parse_args(argv)


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = parse_args(argv)
    if args.execute and args.dry_run:
        print("ERROR: Use either --execute or --dry-run, not both.", file=sys.stderr)
        return 2

    try:
        config = load_config(args.config)
        if args.allow_blank_passwords:
            config.allow_blank_passwords = True
        if args.include_reference_records and not args.venterra_marketing_logins_workbook:
            raise MigrationError("--include-reference-records requires --venterra-marketing-logins-workbook.")
        if args.venterra_marketing_logins_workbook:
            records, mapped_columns = read_venterra_marketing_logins_workbook(
                args.input,
                include_reference_records=args.include_reference_records,
            )
        else:
            headers, rows = read_spreadsheet(args.input)
            records, mapped_columns = transform_rows(headers, rows)
        validation = validate_records(records, mapped_columns, config, args.verbose_sensitive)
        print_validation(validation)
        summary = build_summary(records, validation, config)
        print_safe_summary(summary, args.rotate_after_import)

        if not validation.ok:
            return 1

        payload = build_keeper_import_json(records, config)

        if args.output:
            write_keeper_json(payload, args.output)
            print(f"Generated Keeper JSON: {args.output}")

        if args.execute:
            if not confirm_execution():
                print("Import canceled.")
                return 1
            if args.output:
                json_path = args.output
            else:
                with tempfile.TemporaryDirectory(prefix="keeper-marketing-ops-") as temp_dir:
                    json_path = Path(temp_dir) / "keeper_import.json"
                    write_keeper_json(payload, json_path)
                    run_keeper_import(json_path, args.keeper_binary)
                    print_success_summary(summary, args.rotate_after_import)
                    return 0
            run_keeper_import(json_path, args.keeper_binary)
            print_success_summary(summary, args.rotate_after_import)
        else:
            print("Dry run only. No Keeper import was executed.")
        return 0
    except MigrationError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1


def print_success_summary(summary: MigrationSummary, rotate_after_import: bool = False) -> None:
    print("Keeper import completed")
    print(f"Records imported: {summary.records_ready}")
    print(f"Folders created: {summary.keeper_folders_created}")
    print(f"Shared folder: {summary.shared_folder_name}")
    print(f"Team: {summary.team_name}")
    print(f"Warnings: {summary.warnings_count}")
    if rotate_after_import:
        print("Reminder: review and rotate high-risk shared passwords after import.")


if __name__ == "__main__":
    raise SystemExit(main())
