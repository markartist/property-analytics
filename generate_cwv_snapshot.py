#!/usr/bin/env python3
"""
Portfolio Core Web Vitals Snapshot Report Generator v1.0
========================================================

Complete portfolio listing sorted by performance score (high to low)
showing detailed Core Web Vitals metrics for all properties.

Author: Mark Laufhutte / Atlas
Date: 2026-01-27
Version: 1.0
"""

import sys
import sqlite3
from datetime import datetime, timedelta
from pathlib import Path
import pandas as pd

# Add utils to path
sys.path.insert(0, str(Path(__file__).parent / "utils"))
from report_builder import ReportBuilder, Section

# Configuration
DB_PATH = Path(__file__).parent / "data" / "portfolio_analytics.db"
OUTPUT_DIR = Path(__file__).parent / "reports" / "cwv_snapshot"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


class CWVSnapshotReport:
    """Portfolio Core Web Vitals Snapshot Report Generator"""
    
    def __init__(self, report_date: datetime):
        self.report_date = report_date
        self.date_str = report_date.strftime("%Y-%m-%d")
        
        self.conn = sqlite3.connect(DB_PATH)
        self.conn.row_factory = sqlite3.Row
        
        self.properties = []
        self.stats = {}
        
    def _get_data(self):
        """Get all property data sorted by performance"""
        
        # Get latest date
        latest_date = self.conn.execute("""
            SELECT MAX(metric_date) FROM pagespeed_metrics WHERE strategy='mobile'
        """).fetchone()[0]
        
        # Get previous date for trends
        prev_date = self.conn.execute("""
            SELECT MAX(metric_date) 
            FROM pagespeed_metrics 
            WHERE strategy='mobile' AND metric_date < ?
        """, (latest_date,)).fetchone()[0]
        
        # Get all properties with mobile and desktop metrics
        properties = self.conn.execute(f"""
            SELECT 
                COALESCE(p.property_name, 'Property ' || pm_mobile.property_id) as property_name,
                pm_mobile.performance_score as mobile_performance,
                pm_desktop.performance_score as desktop_performance,
                pm_mobile.accessibility_score,
                pm_mobile.best_practices_score,
                pm_mobile.seo_score,
                pm_mobile.pwa_score,
                pm_mobile.lcp_value,
                pm_mobile.lcp_score,
                pm_mobile.cls_value,
                pm_mobile.cls_score,
                pm_mobile.fid_value,
                pm_mobile.fid_score,
                pm_mobile.fcp_value,
                pm_mobile.ttfb_value,
                pm_mobile.speed_index,
                pm_mobile.time_to_interactive,
                pm_mobile.total_blocking_time,
                pm_prev_mobile.performance_score as prev_mobile_score,
                pm_prev_desktop.performance_score as prev_desktop_score,
                pm_prev_mobile.lcp_value as prev_lcp
            FROM pagespeed_metrics pm_mobile
            LEFT JOIN property_metadata p ON pm_mobile.property_id = p.property_id
            LEFT JOIN pagespeed_metrics pm_desktop ON pm_mobile.property_id = pm_desktop.property_id
                AND pm_desktop.strategy = 'desktop'
                AND pm_desktop.metric_date = '{latest_date}'
            LEFT JOIN pagespeed_metrics pm_prev_mobile ON pm_mobile.property_id = pm_prev_mobile.property_id 
                AND pm_prev_mobile.strategy = 'mobile' 
                AND pm_prev_mobile.metric_date = '{prev_date if prev_date else latest_date}'
            LEFT JOIN pagespeed_metrics pm_prev_desktop ON pm_mobile.property_id = pm_prev_desktop.property_id 
                AND pm_prev_desktop.strategy = 'desktop' 
                AND pm_prev_desktop.metric_date = '{prev_date if prev_date else latest_date}'
            WHERE pm_mobile.strategy='mobile'
            AND pm_mobile.metric_date = '{latest_date}'
            AND pm_mobile.performance_score IS NOT NULL
            ORDER BY pm_mobile.performance_score DESC
        """).fetchall()
        
        self.properties = [dict(row) for row in properties]
        
        # Calculate portfolio stats
        total = len(self.properties)
        if total > 0:
            avg_mobile = sum(p['mobile_performance'] for p in self.properties) / total
            avg_desktop = sum(p['desktop_performance'] for p in self.properties if p.get('desktop_performance')) / len([p for p in self.properties if p.get('desktop_performance')])
            poor = sum(1 for p in self.properties if p['mobile_performance'] < 50)
            needs_improvement = sum(1 for p in self.properties if 50 <= p['mobile_performance'] < 90)
            good = sum(1 for p in self.properties if p['mobile_performance'] >= 90)
            
            # Calculate previous averages for trends
            prev_avg_mobile = None
            prev_avg_lcp = None
            prev_avg_cls = None
            
            if prev_date:
                prev_props = self.conn.execute(f"""
                    SELECT 
                        performance_score,
                        lcp_value,
                        cls_value
                    FROM pagespeed_metrics
                    WHERE strategy='mobile'
                    AND metric_date = '{prev_date}'
                    AND performance_score IS NOT NULL
                """).fetchall()
                
                if prev_props:
                    prev_avg_mobile = sum(p['performance_score'] for p in prev_props) / len(prev_props)
                    prev_avg_lcp = sum(p['lcp_value'] for p in prev_props) / len(prev_props)
                    prev_avg_cls = sum(p['cls_value'] for p in prev_props) / len(prev_props)
            
            self.stats = {
                'total': total,
                'avg_mobile': round(avg_mobile, 1),
                'avg_desktop': round(avg_desktop, 1),
                'latest_date': latest_date,
                'prev_date': prev_date,
                'poor': poor,
                'needs_improvement': needs_improvement,
                'good': good,
                'prev_avg_mobile': prev_avg_mobile,
                'prev_avg_lcp': prev_avg_lcp,
                'prev_avg_cls': prev_avg_cls
            }
    
    def _format_trend(self, delta, invert=False, decimals=1, suffix='') -> str:
        """Format trend indicator with color coding"""
        if delta is None or delta == 0:
            return ''
        
        is_improvement = (delta < 0) if invert else (delta > 0)
        color = '#28a745' if is_improvement else '#dc3545'
        arrow = '↓' if delta < 0 else '↑'
        value_str = f"{abs(delta):.{decimals}f}"
        
        return f' <span style="font-size: 14px; color: {color}; margin-left: 8px;">{arrow}{value_str}{suffix}</span>'
    
    def _get_score_badge(self, score):
        """Get color-coded score badge"""
        if score >= 90:
            return '#28a745', 'GOOD'
        elif score >= 50:
            return '#ff8800', 'NEEDS IMPROVEMENT'  # Orange matching screenshot
        else:
            return '#dc3545', 'POOR'
    
    def _generate_overview(self) -> str:
        """Generate portfolio overview section"""
        
        mobile_color, mobile_label = self._get_score_badge(self.stats['avg_mobile'])
        desktop_color, desktop_label = self._get_score_badge(self.stats['avg_desktop'])
        
        html = f"""
        <div style="background: #f8f9fa; padding: 20px; border-radius: 6px; margin-bottom: 30px;">
            <div style="text-align: center; margin-bottom: 15px;">
                <div style="font-size: 16px; font-weight: 600; margin-bottom: 5px;">
                    <span style="color: {mobile_color};">Mobile: {self.stats['avg_mobile']}</span>
                    <span style="color: #6c757d; margin: 0 10px;">|</span>
                    <span style="color: {desktop_color};">Desktop: {self.stats['avg_desktop']}</span>
                </div>
                <div style="font-size: 13px; color: #6c757d;">
                    {self.stats['total']} properties analyzed | Latest data: {self.stats['latest_date']}
                    {f" | vs. {self.stats['prev_date']}" if self.stats.get('prev_date') else ''}
                </div>
            </div>
            
            <table style="width: 100%; max-width: 600px; margin: 0 auto;">
                <tr>
                    <td style="width: 33%; text-align: center; padding: 15px; background: #f8d7da; border-radius: 4px;">
                        <div style="font-size: 32px; font-weight: 700; color: #dc3545;">{self.stats['poor']}</div>
                        <div style="font-size: 12px; color: #721c24; margin-top: 5px; font-weight: 600;">POOR (&lt;50)</div>
                    </td>
                    <td style="width: 2%;"></td>
                    <td style="width: 33%; text-align: center; padding: 15px; background: #fff3cd; border-radius: 4px;">
                        <div style="font-size: 32px; font-weight: 700; color: #ff8800;">{self.stats['needs_improvement']}</div>
                        <div style="font-size: 12px; color: #856404; margin-top: 5px; font-weight: 600;">NEEDS IMPROVEMENT (50-89)</div>
                    </td>
                    <td style="width: 2%;"></td>
                    <td style="width: 33%; text-align: center; padding: 15px; background: #d4edda; border-radius: 4px;">
                        <div style="font-size: 32px; font-weight: 700; color: #28a745;">{self.stats['good']}</div>
                        <div style="font-size: 12px; color: #155724; margin-top: 5px; font-weight: 600;">GOOD (90+)</div>
                    </td>
                </tr>
            </table>
        </div>
        """
        
        return html
    
    def _generate_data_integrity_section(self) -> str:
        """Generate data integrity footer"""
        
        latest_date = self.stats['latest_date']
        total_props = self.stats['total']
        
        html = f"""
        <div style="margin-top: 40px; padding: 15px; background: #d4edda; border-left: 4px solid #28a745; border-radius: 4px;">
            <div style="font-size: 13px; font-weight: 600; color: #333; margin-bottom: 10px;">✅ Data Integrity</div>
            <div style="font-size: 12px; color: #333; line-height: 1.8;">
                <strong>Data Collection Date:</strong> {latest_date}<br>
                <strong>Properties in Report:</strong> {total_props}<br>
                <strong>Mobile & Desktop Scores:</strong> Collected from PageSpeed Insights API<br>
                <strong>Metrics Included:</strong> Performance, Accessibility, Best Practices, SEO, PWA scores + Core Web Vitals
            </div>
        </div>
        """
        
        return html
    
    def _generate_property_list(self) -> str:
        """Generate complete property listing"""
        
        html = '<div style="margin-bottom: 30px;">'
        
        # Color coding functions
        def score_color(score):
            if not score: return '#999'
            if score >= 90: return '#28a745'
            if score >= 50: return '#ff8800'  # Orange color matching the 87 in screenshot
            return '#dc3545'
        
        def lcp_color(lcp):
            if lcp <= 2.5: return '#28a745'
            if lcp <= 4.0: return '#ffc107'
            return '#dc3545'
        
        def cls_color(cls):
            if cls <= 0.1: return '#28a745'
            if cls <= 0.25: return '#ffc107'
            return '#dc3545'
        
        def fid_color(fid):
            if fid <= 100: return '#28a745'
            if fid <= 300: return '#ffc107'
            return '#dc3545'
        
        def fcp_color(fcp):
            if fcp <= 1.8: return '#28a745'
            if fcp <= 3.0: return '#ffc107'
            return '#dc3545'
        
        def ttfb_color(ttfb):
            if ttfb <= 800: return '#28a745'
            if ttfb <= 1800: return '#ffc107'
            return '#dc3545'
        
        def si_color(si):
            if si <= 3.4: return '#28a745'
            if si <= 5.8: return '#ffc107'
            return '#dc3545'
        
        def tti_color(tti):
            if tti <= 3.8: return '#28a745'
            if tti <= 7.3: return '#ffc107'
            return '#dc3545'
        
        def tbt_color(tbt):
            if tbt <= 200: return '#28a745'
            if tbt <= 600: return '#ffc107'
            return '#dc3545'
        
        for prop in self.properties:
            # Determine border color based on mobile score
            mobile_score = prop['mobile_performance']
            if mobile_score >= 90:
                border_color = '#28a745'
            elif mobile_score >= 50:
                border_color = '#ff8800'  # Orange matching screenshot
            else:
                border_color = '#dc3545'
            
            # Calculate trends
            mobile_trend = ''
            desktop_trend = ''
            lcp_trend = ''
            
            if prop.get('prev_mobile_score') is not None:
                mobile_delta = prop['mobile_performance'] - prop['prev_mobile_score']
                if mobile_delta != 0:
                    mobile_trend = self._format_trend(mobile_delta, invert=False)
            
            if prop.get('prev_desktop_score') is not None and prop.get('desktop_performance'):
                desktop_delta = prop['desktop_performance'] - prop['prev_desktop_score']
                if desktop_delta != 0:
                    desktop_trend = self._format_trend(desktop_delta, invert=False)
            
            if prop.get('prev_lcp') is not None:
                lcp_delta = prop['lcp_value'] - prop['prev_lcp']
                if lcp_delta != 0:
                    lcp_trend = self._format_trend(lcp_delta, invert=True, decimals=1, suffix='s')
            
            html += f"""
            <div style="padding: 12px; background: #fff; border-left: 4px solid {border_color}; margin-bottom: 12px; border-radius: 4px;">
                <div style="display: flex; justify-content: space-between; align-items: flex-start;">
                    <div style="flex: 1;">
                        <div style="font-weight: 600; color: #333; font-size: 14px; margin-bottom: 8px;">
                            {prop['property_name']}
                        </div>
                        <table style="width: 100%; font-size: 11px;">
                            <tr>
                                <td style="padding: 2px 8px 2px 0;"><strong>Mob:</strong> <span style="color: {score_color(prop['mobile_performance'])};">{prop['mobile_performance']}</span></td>
                                <td style="padding: 2px 8px 2px 0;"><strong>Desk:</strong> <span style="color: {score_color(prop.get('desktop_performance')) if prop.get('desktop_performance') else '#999'};">{prop.get('desktop_performance') or 'N/A'}</span></td>
                                <td style="padding: 2px 8px 2px 0;"><strong>A11y:</strong> <span style="color: {score_color(prop['accessibility_score'])};">{prop['accessibility_score'] or 'N/A'}</span></td>
                                <td style="padding: 2px 0;"><strong>BP:</strong> <span style="color: {score_color(prop['best_practices_score'])};">{prop['best_practices_score'] or 'N/A'}</span></td>
                            </tr>
                            <tr>
                                <td style="padding: 2px 8px 2px 0;"><strong>LCP:</strong> <span style="color: {lcp_color(prop['lcp_value'])};">{prop['lcp_value']:.1f}s</span>{lcp_trend}</td>
                                <td style="padding: 2px 8px 2px 0;"><strong>CLS:</strong> <span style="color: {cls_color(prop['cls_value'])};">{prop['cls_value']:.3f}</span></td>
                                <td style="padding: 2px 8px 2px 0;"><strong>FID:</strong> <span style="color: {fid_color(prop['fid_value'])};">{prop['fid_value']:.0f}ms</span></td>
                                <td style="padding: 2px 0;"><strong>FCP:</strong> <span style="color: {fcp_color(prop['fcp_value'])};">{prop['fcp_value']:.1f}s</span></td>
                            </tr>
                            <tr>
                                <td style="padding: 2px 8px 2px 0;"><strong>TTFB:</strong> <span style="color: {ttfb_color(prop['ttfb_value'])};">{prop['ttfb_value']:.0f}ms</span></td>
                                <td style="padding: 2px 8px 2px 0;"><strong>SI:</strong> <span style="color: {si_color(prop['speed_index'])};">{prop['speed_index']:.1f}s</span></td>
                                <td style="padding: 2px 8px 2px 0;"><strong>TTI:</strong> <span style="color: {tti_color(prop['time_to_interactive'])};">{prop['time_to_interactive']:.1f}s</span></td>
                                <td style="padding: 2px 0;"><strong>TBT:</strong> <span style="color: {tbt_color(prop['total_blocking_time'])};">{prop['total_blocking_time']:.0f}ms</span></td>
                            </tr>
                        </table>
                    </div>
                    <div style="margin-left: 15px; text-align: center; min-width: 100px;">
                        <div style="font-size: 32px; font-weight: 700; color: {border_color}; line-height: 1;">
                            {prop['mobile_performance']}
                        </div>
                        <div style="font-size: 12px; color: #6c757d; margin: 4px 0 12px 0;">{mobile_trend}</div>
                        
                        <div style="font-size: 20px; font-weight: 600; color: {score_color(prop.get('desktop_performance')) if prop.get('desktop_performance') else '#999'}; line-height: 1;">
                            {prop.get('desktop_performance') or 'N/A'}
                        </div>
                        <div style="font-size: 11px; color: #6c757d; margin-top: 4px;">{desktop_trend}</div>
                    </div>
                </div>
            </div>
            """
        
        html += '</div>'
        return html
    
    def _generate_excel(self) -> str:
        """Generate Excel spreadsheet with all property data"""
        
        # Prepare data for Excel
        excel_data = []
        for i, prop in enumerate(self.properties, 1):
            # Calculate trend indicators
            mobile_delta = None
            desktop_delta = None
            lcp_delta = None
            if prop.get('prev_mobile_score') is not None:
                mobile_delta = prop['mobile_performance'] - prop['prev_mobile_score']
            if prop.get('prev_desktop_score') is not None and prop.get('desktop_performance'):
                desktop_delta = prop['desktop_performance'] - prop['prev_desktop_score']
            if prop.get('prev_lcp') is not None:
                lcp_delta = prop['lcp_value'] - prop['prev_lcp']
            
            # Grade based on mobile score
            if prop['mobile_performance'] >= 90:
                grade = 'GOOD'
            elif prop['mobile_performance'] >= 50:
                grade = 'NEEDS IMPROVEMENT'
            else:
                grade = 'POOR'
            
            excel_data.append({
                'Rank': i,
                'Property Name': prop['property_name'],
                'Grade': grade,
                'Mobile Score': prop['mobile_performance'],
                'Mobile Change': mobile_delta if mobile_delta is not None else '',
                'Desktop Score': prop.get('desktop_performance') if prop.get('desktop_performance') else 'N/A',
                'Desktop Change': desktop_delta if desktop_delta is not None else '',
                'Accessibility': prop['accessibility_score'] if prop['accessibility_score'] else 'N/A',
                'Best Practices': prop['best_practices_score'] if prop['best_practices_score'] else 'N/A',
                'SEO': prop['seo_score'] if prop['seo_score'] else 'N/A',
                'LCP (s)': round(prop['lcp_value'], 1),
                'LCP Change (s)': round(lcp_delta, 1) if lcp_delta is not None else '',
                'LCP Score': round(prop['lcp_score'], 2) if prop.get('lcp_score') is not None else 'N/A',
                'CLS': round(prop['cls_value'], 3),
                'CLS Score': round(prop['cls_score'], 2) if prop.get('cls_score') is not None else 'N/A',
                'FID (ms)': round(prop['fid_value'], 0),
                'FID Score': round(prop['fid_score'], 2) if prop.get('fid_score') is not None else 'N/A',
                'FCP (s)': round(prop['fcp_value'], 1),
                'TTFB (ms)': round(prop['ttfb_value'], 0),
                'Speed Index (s)': round(prop['speed_index'], 1),
                'TTI (s)': round(prop['time_to_interactive'], 1),
                'TBT (ms)': round(prop['total_blocking_time'], 0)
            })
        
        # Create DataFrame
        df = pd.DataFrame(excel_data)
        
        # Save to Excel with formatting
        excel_file = OUTPUT_DIR / f"Portfolio_CWV_Snapshot_{self.date_str}.xlsx"
        
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Core Web Vitals', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Core Web Vitals']
            
            # Set column widths
            worksheet.column_dimensions['A'].width = 6   # Rank
            worksheet.column_dimensions['B'].width = 35  # Property Name
            worksheet.column_dimensions['C'].width = 18  # Grade
            worksheet.column_dimensions['D'].width = 15  # Mobile Score
            worksheet.column_dimensions['E'].width = 12  # Mobile Change
            worksheet.column_dimensions['F'].width = 12  # Desktop Score
            worksheet.column_dimensions['G'].width = 14  # Desktop Change
            worksheet.column_dimensions['H'].width = 12  # Accessibility
            worksheet.column_dimensions['I'].width = 14  # Best Practices
            worksheet.column_dimensions['J'].width = 8   # SEO
            worksheet.column_dimensions['K'].width = 10  # LCP
            worksheet.column_dimensions['L'].width = 14  # LCP Change
            worksheet.column_dimensions['M'].width = 10  # LCP Score
            worksheet.column_dimensions['N'].width = 8   # CLS
            worksheet.column_dimensions['O'].width = 10  # CLS Score
            worksheet.column_dimensions['P'].width = 10  # FID
            worksheet.column_dimensions['Q'].width = 10  # FID Score
            worksheet.column_dimensions['R'].width = 10  # FCP
            worksheet.column_dimensions['S'].width = 11  # TTFB
            worksheet.column_dimensions['T'].width = 14  # Speed Index
            worksheet.column_dimensions['U'].width = 10  # TTI
            worksheet.column_dimensions['V'].width = 10  # TBT
            
            # Color code the Grade column
            from openpyxl.styles import PatternFill, Font
            
            for row in range(2, len(excel_data) + 2):
                grade_cell = worksheet[f'C{row}']
                score_cell = worksheet[f'D{row}']
                
                if grade_cell.value == 'GOOD':
                    grade_cell.fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
                    grade_cell.font = Font(color='155724', bold=True)
                elif grade_cell.value == 'NEEDS IMPROVEMENT':
                    grade_cell.fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')
                    grade_cell.font = Font(color='856404', bold=True)
                elif grade_cell.value == 'POOR':
                    grade_cell.fill = PatternFill(start_color='F8D7DA', end_color='F8D7DA', fill_type='solid')
                    grade_cell.font = Font(color='721C24', bold=True)
                
                # Color code mobile score changes
                score_change_cell = worksheet[f'E{row}']
                if score_change_cell.value != '':
                    if score_change_cell.value > 0:
                        score_change_cell.font = Font(color='28A745')
                    elif score_change_cell.value < 0:
                        score_change_cell.font = Font(color='DC3545')
                
                # Color code LCP changes (now column L instead of J)
                lcp_change_cell = worksheet[f'L{row}']
                if lcp_change_cell.value != '':
                    if lcp_change_cell.value < 0:  # Negative is good for LCP
                        lcp_change_cell.font = Font(color='28A745')
                    elif lcp_change_cell.value > 0:
                        lcp_change_cell.font = Font(color='DC3545')
            
            # Freeze the header row
            worksheet.freeze_panes = 'A2'
        
        print(f"   Excel: {excel_file}")
        return str(excel_file)
    
    def generate(self) -> tuple:
        """Generate complete report and Excel export"""
        
        print("Collecting portfolio data...")
        self._get_data()
        
        print(f"Building report for {self.stats['total']} properties...")
        builder = ReportBuilder(
            title="Portfolio Core Web Vitals Snapshot",
            subtitle="Complete Property Listing Sorted by Performance",
            version="1.0",
            date_range=self.date_str
        )
        
        # Add KPI tiles for key metrics
        from utils.report_builder import KPITile
        
        # Calculate average LCP and CLS
        avg_lcp = sum(p['lcp_value'] for p in self.properties) / len(self.properties)
        avg_cls = sum(p['cls_value'] for p in self.properties) / len(self.properties)
        
        # Determine LCP color
        if avg_lcp <= 2.5:
            lcp_color = '#28a745'
            lcp_label = 'Good'
        elif avg_lcp <= 4.0:
            lcp_color = '#ff8800'
            lcp_label = 'Needs Improvement'
        else:
            lcp_color = '#dc3545'
            lcp_label = 'Poor'
        
        # Determine CLS color
        if avg_cls <= 0.1:
            cls_color = '#28a745'
            cls_label = 'Good'
        elif avg_cls <= 0.25:
            cls_color = '#ff8800'
            cls_label = 'Needs Improvement'
        else:
            cls_color = '#dc3545'
            cls_label = 'Poor'
        
        # Calculate trends for KPI tiles
        mobile_trend_text = ''
        lcp_trend_text = ''
        cls_trend_text = ''
        
        if self.stats.get('prev_avg_mobile') is not None:
            delta = self.stats['avg_mobile'] - self.stats['prev_avg_mobile']
            if delta != 0:
                arrow = '↑' if delta > 0 else '↓'
                mobile_trend_text = f"{arrow}{abs(delta):.1f}"
        
        if self.stats.get('prev_avg_lcp') is not None:
            delta = avg_lcp - self.stats['prev_avg_lcp']
            if delta != 0:
                arrow = '↓' if delta < 0 else '↑'
                lcp_trend_text = f"{arrow}{abs(delta):.2f}s"
        
        if self.stats.get('prev_avg_cls') is not None:
            delta = avg_cls - self.stats['prev_avg_cls']
            if delta != 0:
                arrow = '↓' if delta < 0 else '↑'
                cls_trend_text = f"{arrow}{abs(delta):.3f}"
        
        mobile_color, mobile_label = self._get_score_badge(self.stats['avg_mobile'])
        
        builder.add_kpi_tiles([
            KPITile(
                label="Mobile Performance",
                value=str(self.stats['avg_mobile']),
                sublabel=mobile_label,
                trend=mobile_trend_text if mobile_trend_text else None,
                is_primary=True
            ),
            KPITile(
                label="Avg LCP (Largest Contentful Paint)",
                value=f"{avg_lcp:.2f}s",
                sublabel=lcp_label,
                trend=lcp_trend_text if lcp_trend_text else None,
                trend_inverse=True  # Lower LCP is better
            ),
            KPITile(
                label="Avg CLS (Cumulative Layout Shift)",
                value=f"{avg_cls:.3f}",
                sublabel=cls_label,
                trend=cls_trend_text if cls_trend_text else None,
                trend_inverse=True  # Lower CLS is better
            )
        ], columns=3)
        
        # Portfolio Overview (no status badge)
        overview_html = self._generate_overview()
        from utils.report_builder import Section
        
        # Create section without status by passing None
        class SectionNoStatus(Section):
            def to_html(self):
                html = f'''
                <!-- Section Header -->
                <table cellpadding="0" cellspacing="0" border="0" style="width: 100%; margin: 40px 0 0 0;">
                    <tr>
                        <td style="background: #15284B; padding: 15px 20px; border-radius: 6px 6px 0 0;">
                            <h2 style="font-size: 18px; color: #ffffff; margin: 0; font-weight: 600;">{self.title}</h2>
                        </td>
                    </tr>
                '''
                
                if self.description:
                    html += f'''
                    <tr>
                        <td style="background: #f8f9fa; padding: 12px 20px; border-radius: 0 0 6px 6px; margin-bottom: 20px;">
                            <p style="font-size: 13px; color: #6c757d; margin: 0; font-style: italic;">{self.description}</p>
                        </td>
                    </tr>
                    '''
                
                html += '''
                </table>
                
                <!-- Section Content -->
                '''
                html += self.content
                
                return html
        
        builder.add_section(SectionNoStatus(
            title="Portfolio Overview",
            content=overview_html,
            description=f"Current state of {self.stats['total']} properties"
        ))
        
        # Complete Property Listing (no status badge)
        property_list_html = self._generate_property_list()
        builder.add_section(SectionNoStatus(
            title="Complete Property Performance Ranking",
            content=property_list_html,
            description="All properties sorted by performance score (highest to lowest)"
        ))
        
        # Data Integrity Section (no status badge)
        integrity_html = self._generate_data_integrity_section()
        builder.add_section(SectionNoStatus(
            title="Data Integrity",
            content=integrity_html,
            description=f"Collection status for {self.stats['latest_date']}"
        ))
        
        # Generate HTML
        html = builder.generate()
        
        # Save report
        output_file = OUTPUT_DIR / f"Portfolio_CWV_Snapshot_{self.date_str}.html"
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(html)
        
        print(f"\n✅ Portfolio Core Web Vitals Snapshot generated:")
        print(f"   HTML: {output_file}")
        
        # Generate Excel export
        excel_file = self._generate_excel()
        
        print(f"\n📊 Portfolio Summary:")
        print(f"   Total Properties: {self.stats['total']}")
        print(f"   Mobile Average: {self.stats['avg_mobile']}")
        print(f"   Desktop Average: {self.stats['avg_desktop']}")
        print(f"   Poor (<50): {self.stats['poor']}")
        print(f"   Needs Improvement (50-89): {self.stats['needs_improvement']}")
        print(f"   Good (90+): {self.stats['good']}")
        
        return str(output_file), str(excel_file)
    
    def __del__(self):
        if hasattr(self, 'conn'):
            self.conn.close()


def main():
    import argparse
    
    parser = argparse.ArgumentParser(description="Generate Portfolio Core Web Vitals Snapshot")
    parser.add_argument('--date', help='Report date (YYYY-MM-DD), defaults to today')
    
    args = parser.parse_args()
    
    # Determine report date
    if args.date:
        report_date = datetime.strptime(args.date, "%Y-%m-%d")
    else:
        report_date = datetime.now()
    
    print("=" * 70)
    print("PORTFOLIO CORE WEB VITALS SNAPSHOT v1.0")
    print("=" * 70)
    print(f"\nReport Date: {report_date.strftime('%Y-%m-%d')}\n")
    
    # Generate report
    reporter = CWVSnapshotReport(report_date)
    html_file, excel_file = reporter.generate()
    
    print(f"\n✅ Files ready for email attachment:")
    print(f"   {excel_file}")
    
    return 0


if __name__ == "__main__":
    sys.exit(main())
