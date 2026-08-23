#!/usr/bin/env python3
"""Build a non-mutating Resi Edge Phase 2 preflight queue/report.

The report merges the rollout workbook, QA Pastel/Kinsta staging links, the
governed property identity matrix, the current Resi Edge contract, and the
latest local read-only Cloudflare zone inventory. It does not mutate Cloudflare,
DNS, WordPress, Zaraz, Ahrefs, GSC, Captain, Data Pond, R2, or cache.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import urllib.request
from collections import Counter
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from docx import Document
from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from Data_Collection.utils.property_identity import resolve_property_identity


DEFAULT_ROLLOUT_WORKBOOK = Path("/Users/mark/Downloads/Property Vanity URLs for RollOut.xlsx")
DEFAULT_QA_DOCX = Path("/Users/mark/Downloads/QA Pastel Links.docx")
CONTRACT_PATH = REPO_ROOT / "ops/cloudflare/shared/resi-edge-package/contract.json"
MANIFEST_DIR = REPO_ROOT / "config/portfolio_resi_edge_stabilization"
REPORT_ROOT = REPO_ROOT / "reports/resi_edge_performance/phase2-preflight"
LOCAL_TZ = ZoneInfo("America/Chicago")

TRACKING_SUFFIXES = {
    "Apartments.com - APT": "APT",
    "ApartmentList.com - APL": "APL",
    "ApartmentRatings.com - APR": "APR",
    "Bing.com - BNG": "BNG",
    "Google - GOO": "GOO",
    "Google Ads - GOA": "GOA",
    "Social - SOC": "SOC",
    "ThirtyLines - VWS": "VWS",
    "Video - VID": "VID",
    "Yahoo - YAH": "YAH",
    "Zillow - ZIL": "ZIL",
}

EXCLUDED_CODES = {
    "TX4DP": "Lease-up with its own template; user said Delta should not move through this rollout.",
    "TX4MV": "Lease-up with its own template; user said Monteverde should not move through this rollout.",
}


@dataclass
class RolloutRow:
    phase: int
    go_live: str | None
    property_name: str
    property_code: str | None
    canonical_name: str | None
    vanity_domain_raw: str
    vanity_domain: str
    region: str | None
    units: int | None
    workbook_status: str | None


def read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def normalize_domain(value: str | None) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"^https?://", "", text)
    return text.strip().strip("/")


def format_human_date(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        return value
    if hasattr(value, "strftime"):
        return value.strftime("%m/%d/%Y")
    return str(value)


def row_cell(row: tuple[Any, ...], index: int) -> Any:
    return row[index] if index < len(row) else None


def load_rollout_rows(path: Path, phase: int | None) -> tuple[list[RolloutRow], dict[int, str | None]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["Property Rollout URLs"]
    rows = list(ws.iter_rows(values_only=True))

    phase_dates: dict[int, str | None] = {}
    for row in rows[1:]:
        if row_cell(row, 9) is None:
            continue
        try:
            phase_dates[int(row_cell(row, 9))] = format_human_date(row_cell(row, 10))
        except (TypeError, ValueError):
            continue

    output: list[RolloutRow] = []
    for row in rows[1:]:
        if not row_cell(row, 0) or not row_cell(row, 1) or row_cell(row, 4) is None:
            continue
        try:
            row_phase = int(row_cell(row, 4))
        except (TypeError, ValueError):
            continue
        if phase is not None and row_phase != phase:
            continue
        identity = resolve_property_identity(str(row_cell(row, 1)))
        raw_domain = str(row_cell(row, 0))
        output.append(
            RolloutRow(
                phase=row_phase,
                go_live=phase_dates.get(row_phase),
                property_name=str(row_cell(row, 1)).strip(),
                property_code=identity.property_code if identity else None,
                canonical_name=identity.property_name if identity else None,
                vanity_domain_raw=raw_domain,
                vanity_domain=normalize_domain(raw_domain),
                region=str(row_cell(row, 2)).strip() if row_cell(row, 2) else None,
                units=int(row_cell(row, 3)) if isinstance(row_cell(row, 3), (int, float)) else None,
                workbook_status=str(row_cell(row, 5)).strip() if row_cell(row, 5) else None,
            )
        )
    return output, phase_dates


def load_tracking_rows(path: Path) -> dict[str, dict[str, Any]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["Tracking URLs"]
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    output: dict[str, dict[str, Any]] = {}
    for index, row in enumerate(rows[1:], start=2):
        record = dict(zip(headers, row))
        code = record.get("Property Code")
        if not code:
            continue
        record["_row"] = index
        output[str(code)] = record
    return output


def load_qa_rows(path: Path) -> dict[str, dict[str, Any]]:
    lines: list[str] = []
    for paragraph in Document(path).paragraphs:
        for line in paragraph.text.splitlines():
            line = line.strip()
            if line:
                lines.append(line)

    items: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None
    for line in lines:
        lower = line.lower()
        if lower.startswith("pastel:"):
            if current is not None:
                current["pastel_url"] = line.split(":", 1)[1].strip()
            continue
        if lower.startswith("staging:"):
            if current is not None:
                current["staging_kinsta_url"] = line.split(":", 1)[1].strip()
                items.append(current)
                current = None
            continue
        if "please see julie" in lower:
            if current is not None:
                current["note"] = line
                items.append(current)
                current = None
            continue
        if current is not None:
            items.append(current)
        current = {"doc_name": line, "pastel_url": None, "staging_kinsta_url": None, "note": None}
    if current:
        items.append(current)

    output: dict[str, dict[str, Any]] = {}
    for item in items:
        identity = resolve_property_identity(item["doc_name"])
        if not identity or not identity.property_code:
            continue
        item["property_code"] = identity.property_code
        item["canonical_name"] = identity.property_name
        output[identity.property_code] = item
    return output


def latest_cloudflare_inventory_path() -> Path | None:
    candidates = sorted((REPO_ROOT / "reports/domain_ops").glob("*/cloudflare_zones.json"))
    return candidates[-1] if candidates else None


def load_cloudflare_zones(path: Path | None) -> dict[str, dict[str, Any]]:
    if not path or not path.exists():
        return {}
    payload = read_json(path)
    if not isinstance(payload, list):
        return {}
    return {
        normalize_domain(item.get("name")): item
        for item in payload
        if isinstance(item, dict) and item.get("name")
    }


def resolve_input_path(path: Path | None) -> Path | None:
    if path is None:
        return None
    if path.is_absolute():
        return path
    return REPO_ROOT / path


def display_repo_path(path: Path | None) -> str | None:
    if path is None:
        return None
    resolved = resolve_input_path(path)
    if resolved is None:
        return None
    try:
        return str(resolved.relative_to(REPO_ROOT))
    except ValueError:
        return str(resolved)


def manifest_candidates(domain: str, code: str | None) -> list[str]:
    candidates: list[str] = []
    normalized_domain = normalize_domain(domain)
    for path in sorted(MANIFEST_DIR.glob("*.manifest.json")):
        try:
            payload = read_json(path)
        except json.JSONDecodeError:
            continue
        target = payload.get("target") if isinstance(payload, dict) else {}
        routing = payload.get("routing") if isinstance(payload, dict) else {}
        target_domain = normalize_domain((target or {}).get("domain"))
        zone_name = normalize_domain((routing or {}).get("cloudflare_zone_name"))
        target_codes = {
            str((target or {}).get("property_code") or "").strip().upper(),
            str((target or {}).get("source_property_code") or "").strip().upper(),
        }
        if (
            normalized_domain in {target_domain, zone_name}
            or (code and code.strip().upper() in target_codes)
        ):
            candidates.append(str(path.relative_to(REPO_ROOT)))
    return candidates


def check_tracking_urls(row: RolloutRow, tracking: dict[str, Any] | None) -> tuple[bool, list[str], list[str]]:
    if not tracking:
        return False, ["tracking row missing"], []
    issues: list[str] = []
    decisions: list[str] = []
    base = str(tracking.get("Vanity Domain") or "").rstrip("/")
    expected_base = f"https://{row.vanity_domain}"
    if normalize_domain(base) != row.vanity_domain:
        issues.append(f"tracking vanity domain mismatch: {base}")
    code = row.property_code
    for column, suffix in TRACKING_SUFFIXES.items():
        value = tracking.get(column)
        if not value:
            issues.append(f"{column} missing")
            continue
        text = str(value)
        if not text.startswith(f"{expected_base}/?id="):
            issues.append(f"{column} base mismatch")
        if code and f"id={code}{suffix}" not in text:
            if row.property_code == "GA4AB" and column == "Google Ads - GOA":
                decisions.append("Axial Buckhead GOA URL lacks normal GOA source suffix; confirm intentional.")
            else:
                issues.append(f"{column} id mismatch; expected {code}{suffix}")
    return not issues, issues, decisions


def check_staging_url(url: str | None) -> tuple[bool, list[str]]:
    if not url:
        return False, ["staging Kinsta URL missing"]
    issues: list[str] = []
    if "kinsta.cloud" not in url:
        issues.append("staging URL is not kinsta.cloud")
    if not url.startswith("https://"):
        issues.append("staging URL is not https")
    if not url.endswith("/"):
        issues.append("staging URL is missing trailing slash")
    return not issues, issues


def fetch_head_status(url: str | None, timeout: int) -> dict[str, Any]:
    if not url:
        return {"checked": False, "pass": False, "reason": "missing URL"}
    request = urllib.request.Request(
        url,
        method="HEAD",
        headers={
            "User-Agent": "PropertyAnalytics-ResiEdgePreflight/1.0",
            "Accept": "text/html,application/xhtml+xml",
        },
    )
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            return {
                "checked": True,
                "pass": 200 <= int(response.status) < 400,
                "status": int(response.status),
                "content_type": response.headers.get("content-type"),
                "final_url": response.geturl(),
            }
    except Exception as exc:  # noqa: BLE001 - report sanitized network failure
        return {"checked": True, "pass": False, "reason": type(exc).__name__}


def status_for(
    *,
    excluded: bool,
    blockers: list[str],
    decisions: list[str],
    manifest_paths: list[str],
    cloudflare_zone: dict[str, Any] | None,
) -> str:
    if excluded:
        return "excluded"
    if blockers:
        return "blocked"
    if decisions:
        return "needs_decision"
    if not manifest_paths:
        return "source_ready_manifest_needed"
    if not cloudflare_zone:
        return "source_ready_cloudflare_zone_needed"
    return "source_ready"


def summarize_rows(rows: list[dict[str, Any]]) -> dict[str, Any]:
    counts = Counter(row["dashboard_status"] for row in rows)
    return {
        "total": len(rows),
        "excluded": counts.get("excluded", 0),
        "blocked": counts.get("blocked", 0),
        "needs_decision": counts.get("needs_decision", 0),
        "source_ready": counts.get("source_ready", 0),
        "source_ready_manifest_needed": counts.get("source_ready_manifest_needed", 0),
        "source_ready_cloudflare_zone_needed": counts.get("source_ready_cloudflare_zone_needed", 0),
    }


def markdown_report(payload: dict[str, Any]) -> str:
    properties = payload["properties"]
    first_property = properties[0] if properties else {}
    contract_gate_count = first_property.get("current_contract_gate_count") or 0
    staging_checked = sum(1 for row in properties if (row.get("staging_probe") or {}).get("checked"))
    staging_passed = sum(1 for row in properties if (row.get("staging_probe") or {}).get("pass"))
    lines = [
        "# Resi Edge Phase 2 Preflight",
        "",
        f"Generated: {payload['generated_at_human']}",
        f"Go-live target: {payload.get('go_live') or 'not provided'}",
        f"Current contract required gates: {contract_gate_count}",
        f"Cloudflare inventory: {payload['sources'].get('cloudflare_inventory') or 'not available'}",
        f"Staging probes: {staging_passed}/{staging_checked} passed" if staging_checked else "Staging probes: not run",
        "",
        "## Summary",
        "",
    ]
    summary = payload["summary"]
    for key in [
        "total",
        "source_ready",
        "source_ready_manifest_needed",
        "source_ready_cloudflare_zone_needed",
        "needs_decision",
        "blocked",
        "excluded",
    ]:
        lines.append(f"- {key}: {summary.get(key, 0)}")
    lines.extend(
        [
            "",
            "## Phase 2 Queue",
            "",
            "| Property | Code | Vanity | Staging | HTTP | Cloudflare | Status | Blockers / Decisions | Prep Notes |",
            "| --- | --- | --- | --- | --- | --- | --- | --- | --- |",
        ]
    )
    for row in payload["properties"]:
        blockers = "; ".join(row["blockers"] + row["decisions"]) or "none"
        warnings = "; ".join(row["warnings"]) or "none"
        probe = row.get("staging_probe") or {}
        if probe.get("checked"):
            staging_http = str(probe.get("status") or probe.get("reason") or "checked")
        else:
            staging_http = "not run"
        lines.append(
            "| {property} | `{code}` | `{vanity}` | {staging} | {http} | {cf} | `{status}` | {blockers} | {warnings} |".format(
                property=row["property_name"],
                code=row.get("property_code") or "",
                vanity=row["vanity_domain"],
                staging=row.get("staging_kinsta_url") or "",
                http=staging_http,
                cf=row.get("cloudflare_zone_status") or "missing",
                status=row["dashboard_status"],
                blockers=blockers,
                warnings=warnings,
            )
        )
    lines.extend(
        [
            "",
            "## Boundaries",
            "",
            "- This report is non-mutating. It does not change Cloudflare, DNS, WordPress, Zaraz, Ahrefs, GSC, Captain, Data Pond, R2, or cache.",
            "- `source_ready` and related states are preparation states only. They are not live production approval.",
            "- Final proof still requires vanity-domain live gates after explicit approval.",
        ]
    )
    return "\n".join(lines) + "\n"


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    fieldnames = [
        "phase",
        "go_live",
        "dashboard_status",
        "property_code",
        "property_name",
        "canonical_name",
        "vanity_domain",
        "staging_kinsta_url",
        "pastel_url",
        "cloudflare_zone_status",
        "cloudflare_zone_present",
        "staging_probe_checked",
        "staging_probe_pass",
        "staging_probe_status",
        "staging_probe_final_url",
        "manifest_count",
        "blockers",
        "decisions",
        "warnings",
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            staging_probe = row.get("staging_probe") or {}
            writer.writerow(
                {
                    **{key: row.get(key) for key in fieldnames},
                    "blockers": "; ".join(row.get("blockers") or []),
                    "decisions": "; ".join(row.get("decisions") or []),
                    "warnings": "; ".join(row.get("warnings") or []),
                    "cloudflare_zone_present": int(bool(row.get("cloudflare_zone_present"))),
                    "staging_probe_checked": int(bool(staging_probe.get("checked"))),
                    "staging_probe_pass": int(bool(staging_probe.get("pass"))),
                    "staging_probe_status": staging_probe.get("status") or staging_probe.get("reason"),
                    "staging_probe_final_url": staging_probe.get("final_url"),
                    "manifest_count": len(row.get("manifest_paths") or []),
                }
            )


def main() -> int:
    parser = argparse.ArgumentParser(description="Build non-mutating Resi Edge phase preflight queue/report.")
    parser.add_argument("--phase", type=int, default=2, help="Rollout phase to report. Default: 2.")
    parser.add_argument("--rollout-workbook", type=Path, default=DEFAULT_ROLLOUT_WORKBOOK)
    parser.add_argument("--qa-docx", type=Path, default=DEFAULT_QA_DOCX)
    parser.add_argument("--cloudflare-inventory", type=Path, default=None, help="Path to cloudflare_zones.json. Defaults to latest local inventory.")
    parser.add_argument("--output-dir", type=Path, default=None)
    parser.add_argument("--check-staging", action="store_true", help="Run non-mutating HEAD checks against staging Kinsta URLs.")
    parser.add_argument("--staging-timeout", type=int, default=8)
    args = parser.parse_args()

    rollout_rows, phase_dates = load_rollout_rows(args.rollout_workbook, args.phase)
    tracking_rows = load_tracking_rows(args.rollout_workbook)
    qa_rows = load_qa_rows(args.qa_docx)
    cloudflare_inventory_path = resolve_input_path(args.cloudflare_inventory) or latest_cloudflare_inventory_path()
    cloudflare_zones = load_cloudflare_zones(cloudflare_inventory_path)
    contract = read_json(CONTRACT_PATH)

    output_rows: list[dict[str, Any]] = []
    for rollout in rollout_rows:
        tracking = tracking_rows.get(rollout.property_code or "")
        qa = qa_rows.get(rollout.property_code or "")
        zone = cloudflare_zones.get(rollout.vanity_domain)
        manifest_paths = manifest_candidates(rollout.vanity_domain, rollout.property_code)
        blockers: list[str] = []
        decisions: list[str] = []
        warnings: list[str] = []
        excluded_reason = EXCLUDED_CODES.get(rollout.property_code or "")

        if not rollout.property_code:
            blockers.append("property identity did not resolve")
        if rollout.vanity_domain_raw != rollout.vanity_domain:
            warnings.append("vanity domain normalized from workbook source")
        staging_ok, staging_issues = check_staging_url((qa or {}).get("staging_kinsta_url"))
        if not staging_ok:
            blockers.extend(staging_issues)
        tracking_ok, tracking_issues, tracking_decisions = check_tracking_urls(rollout, tracking)
        if not tracking_ok:
            blockers.extend(tracking_issues)
        decisions.extend(tracking_decisions)
        if not zone:
            blockers.append("Cloudflare zone not found in latest local read-only inventory")
        elif zone.get("status") != "active":
            blockers.append(f"Cloudflare zone is not active: {zone.get('status')}")
        if not manifest_paths:
            warnings.append("canonical Resi Edge manifest not generated yet")
        if excluded_reason:
            blockers = []
            decisions = []
            warnings.append(excluded_reason)

        staging_probe = fetch_head_status((qa or {}).get("staging_kinsta_url"), args.staging_timeout) if args.check_staging else {"checked": False}
        if staging_probe.get("checked") and not staging_probe.get("pass") and not excluded_reason:
            blockers.append(f"staging Kinsta HEAD check failed: {staging_probe.get('reason') or staging_probe.get('status')}")

        output_rows.append(
            {
                "phase": rollout.phase,
                "go_live": rollout.go_live,
                "property_code": rollout.property_code,
                "property_name": rollout.property_name,
                "canonical_name": rollout.canonical_name,
                "region": rollout.region,
                "units": rollout.units,
                "workbook_status": rollout.workbook_status,
                "vanity_domain": rollout.vanity_domain,
                "vanity_domain_raw": rollout.vanity_domain_raw,
                "staging_kinsta_url": (qa or {}).get("staging_kinsta_url"),
                "pastel_url": (qa or {}).get("pastel_url"),
                "qa_doc_name": (qa or {}).get("doc_name"),
                "tracking_row": tracking.get("_row") if tracking else None,
                "cloudflare_zone_present": bool(zone),
                "cloudflare_zone_status": zone.get("status") if zone else None,
                "cloudflare_plan": zone.get("plan_name") if zone else None,
                "cloudflare_inventory_source": display_repo_path(cloudflare_inventory_path),
                "manifest_paths": manifest_paths,
                "current_contract_gate_count": len(contract.get("required_gates") or []),
                "staging_probe": staging_probe,
                "blockers": blockers,
                "decisions": decisions,
                "warnings": warnings,
                "excluded_reason": excluded_reason or None,
                "dashboard_status": status_for(
                    excluded=bool(excluded_reason),
                    blockers=blockers,
                    decisions=decisions,
                    manifest_paths=manifest_paths,
                    cloudflare_zone=zone,
                ),
            }
        )

    generated_at = datetime.now(timezone.utc)
    stamp = generated_at.strftime("%Y%m%dT%H%M%SZ")
    output_dir = args.output_dir or REPORT_ROOT / f"phase-{args.phase}-preflight-{stamp}"
    if not output_dir.is_absolute():
        output_dir = REPO_ROOT / output_dir
    output_dir.mkdir(parents=True, exist_ok=True)
    payload = {
        "schema": "resi_edge_phase_preflight_v1",
        "phase": args.phase,
        "go_live": phase_dates.get(args.phase),
        "generated_at": generated_at.isoformat(),
        "generated_at_human": generated_at.astimezone(LOCAL_TZ).strftime("%m/%d/%Y %I:%M %p %Z"),
        "mutations_performed": False,
        "sources": {
            "rollout_workbook": str(args.rollout_workbook),
            "qa_docx": str(args.qa_docx),
            "identity_matrix": str((REPO_ROOT / "config/property_identity_matrix.json")),
            "contract": str(CONTRACT_PATH),
            "cloudflare_inventory": str(cloudflare_inventory_path) if cloudflare_inventory_path else None,
        },
        "summary": summarize_rows(output_rows),
        "properties": output_rows,
    }
    write_json(output_dir / "phase-preflight.json", payload)
    write_csv(output_dir / "phase-preflight.csv", output_rows)
    (output_dir / "PHASE_PREFLIGHT_READOUT.md").write_text(markdown_report(payload), encoding="utf-8")
    print(
        json.dumps(
            {
                "status": "passed",
                "output_dir": str(output_dir.relative_to(REPO_ROOT)),
                "summary": payload["summary"],
            },
            indent=2,
            sort_keys=True,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
