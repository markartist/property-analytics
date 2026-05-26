#!/usr/bin/env python3
"""
Export a lifetime PageSpeed Insights report from the local master database.

Outputs an Excel workbook with:
- Summary: portfolio-level PSI coverage and recency
- Property Summary: per-property lifetime coverage by strategy
- Latest by Property: latest available row for each property/strategy with all PSI metrics
- Metric Coverage: non-null metric counts by property across lifetime records
- PSI Raw Data: full pagespeed_metrics history joined to property names
"""

from __future__ import annotations

import sqlite3
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


ROOT = Path("/Users/mark/Property_Analytics")
DB_PATH = ROOT / "data" / "portfolio_analytics.db"
OUTPUT_DIR = ROOT / "outputs"


RAW_QUERY = """
SELECT
    pm.property_id,
    COALESCE(p.property_name, pm.property_id) AS property_name,
    pm.metric_date,
    pm.strategy,
    pm.performance_score,
    pm.accessibility_score,
    pm.best_practices_score,
    pm.seo_score,
    pm.pwa_score,
    pm.lcp_value,
    pm.lcp_score,
    pm.fid_value,
    pm.fid_score,
    pm.cls_value,
    pm.cls_score,
    pm.fcp_value,
    pm.ttfb_value,
    pm.speed_index,
    pm.time_to_interactive,
    pm.total_blocking_time,
    pm.collection_id,
    pm.collected_at
FROM pagespeed_metrics pm
LEFT JOIN properties p
    ON p.property_id = pm.property_id
ORDER BY pm.metric_date ASC, property_name ASC, pm.strategy ASC
"""


PROPERTY_SUMMARY_QUERY = """
WITH strategy_counts AS (
    SELECT
        pm.property_id,
        COALESCE(p.property_name, pm.property_id) AS property_name,
        COUNT(*) AS total_rows,
        MIN(pm.metric_date) AS first_metric_date,
        MAX(pm.metric_date) AS latest_metric_date,
        COUNT(DISTINCT pm.metric_date) AS distinct_dates,
        SUM(CASE WHEN pm.strategy = 'mobile' THEN 1 ELSE 0 END) AS mobile_rows,
        SUM(CASE WHEN pm.strategy = 'desktop' THEN 1 ELSE 0 END) AS desktop_rows,
        AVG(CASE WHEN pm.strategy = 'mobile' THEN pm.performance_score END) AS avg_mobile_score,
        AVG(CASE WHEN pm.strategy = 'desktop' THEN pm.performance_score END) AS avg_desktop_score,
        AVG(CASE WHEN pm.strategy = 'mobile' THEN pm.lcp_value END) AS avg_mobile_lcp,
        AVG(CASE WHEN pm.strategy = 'desktop' THEN pm.lcp_value END) AS avg_desktop_lcp
    FROM pagespeed_metrics pm
    LEFT JOIN properties p
        ON p.property_id = pm.property_id
    GROUP BY pm.property_id, COALESCE(p.property_name, pm.property_id)
)
SELECT
    property_id,
    property_name,
    total_rows,
    distinct_dates,
    first_metric_date,
    latest_metric_date,
    mobile_rows,
    desktop_rows,
    ROUND(avg_mobile_score, 2) AS avg_mobile_score,
    ROUND(avg_desktop_score, 2) AS avg_desktop_score,
    ROUND(avg_mobile_lcp, 3) AS avg_mobile_lcp,
    ROUND(avg_desktop_lcp, 3) AS avg_desktop_lcp
FROM strategy_counts
ORDER BY property_name ASC
"""


LATEST_BY_PROPERTY_QUERY = """
WITH latest_per_strategy AS (
    SELECT
        property_id,
        strategy,
        MAX(metric_date) AS latest_metric_date
    FROM pagespeed_metrics
    GROUP BY property_id, strategy
)
SELECT
    pm.property_id,
    COALESCE(p.property_name, pm.property_id) AS property_name,
    pm.strategy,
    pm.metric_date,
    pm.performance_score,
    pm.accessibility_score,
    pm.best_practices_score,
    pm.seo_score,
    pm.pwa_score,
    pm.lcp_value,
    pm.lcp_score,
    pm.fid_value,
    pm.fid_score,
    pm.cls_value,
    pm.cls_score,
    pm.fcp_value,
    pm.ttfb_value,
    pm.speed_index,
    pm.time_to_interactive,
    pm.total_blocking_time,
    pm.collection_id,
    pm.collected_at
FROM pagespeed_metrics pm
JOIN latest_per_strategy lps
    ON lps.property_id = pm.property_id
   AND lps.strategy = pm.strategy
   AND lps.latest_metric_date = pm.metric_date
LEFT JOIN properties p
    ON p.property_id = pm.property_id
ORDER BY property_name ASC, pm.strategy ASC
"""


METRIC_COVERAGE_QUERY = """
SELECT
    pm.property_id,
    COALESCE(p.property_name, pm.property_id) AS property_name,
    COUNT(*) AS total_rows,
    SUM(CASE WHEN pm.strategy = 'mobile' THEN 1 ELSE 0 END) AS mobile_rows,
    SUM(CASE WHEN pm.strategy = 'desktop' THEN 1 ELSE 0 END) AS desktop_rows,
    SUM(CASE WHEN pm.performance_score IS NOT NULL THEN 1 ELSE 0 END) AS performance_score_rows,
    SUM(CASE WHEN pm.accessibility_score IS NOT NULL THEN 1 ELSE 0 END) AS accessibility_score_rows,
    SUM(CASE WHEN pm.best_practices_score IS NOT NULL THEN 1 ELSE 0 END) AS best_practices_score_rows,
    SUM(CASE WHEN pm.seo_score IS NOT NULL THEN 1 ELSE 0 END) AS seo_score_rows,
    SUM(CASE WHEN pm.pwa_score IS NOT NULL THEN 1 ELSE 0 END) AS pwa_score_rows,
    SUM(CASE WHEN pm.lcp_value IS NOT NULL THEN 1 ELSE 0 END) AS lcp_value_rows,
    SUM(CASE WHEN pm.lcp_score IS NOT NULL THEN 1 ELSE 0 END) AS lcp_score_rows,
    SUM(CASE WHEN pm.fid_value IS NOT NULL THEN 1 ELSE 0 END) AS fid_value_rows,
    SUM(CASE WHEN pm.fid_score IS NOT NULL THEN 1 ELSE 0 END) AS fid_score_rows,
    SUM(CASE WHEN pm.cls_value IS NOT NULL THEN 1 ELSE 0 END) AS cls_value_rows,
    SUM(CASE WHEN pm.cls_score IS NOT NULL THEN 1 ELSE 0 END) AS cls_score_rows,
    SUM(CASE WHEN pm.fcp_value IS NOT NULL THEN 1 ELSE 0 END) AS fcp_value_rows,
    SUM(CASE WHEN pm.ttfb_value IS NOT NULL THEN 1 ELSE 0 END) AS ttfb_value_rows,
    SUM(CASE WHEN pm.speed_index IS NOT NULL THEN 1 ELSE 0 END) AS speed_index_rows,
    SUM(CASE WHEN pm.time_to_interactive IS NOT NULL THEN 1 ELSE 0 END) AS time_to_interactive_rows,
    SUM(CASE WHEN pm.total_blocking_time IS NOT NULL THEN 1 ELSE 0 END) AS total_blocking_time_rows
FROM pagespeed_metrics pm
LEFT JOIN properties p
    ON p.property_id = pm.property_id
GROUP BY pm.property_id, COALESCE(p.property_name, pm.property_id)
ORDER BY property_name ASC
"""


PORTFOLIO_SUMMARY_QUERY = """
SELECT
    COUNT(*) AS total_rows,
    COUNT(DISTINCT property_id) AS property_count,
    COUNT(DISTINCT metric_date) AS distinct_dates,
    MIN(metric_date) AS first_metric_date,
    MAX(metric_date) AS latest_metric_date,
    SUM(CASE WHEN strategy = 'mobile' THEN 1 ELSE 0 END) AS mobile_rows,
    SUM(CASE WHEN strategy = 'desktop' THEN 1 ELSE 0 END) AS desktop_rows,
    ROUND(AVG(CASE WHEN strategy = 'mobile' THEN performance_score END), 2) AS avg_mobile_score,
    ROUND(AVG(CASE WHEN strategy = 'desktop' THEN performance_score END), 2) AS avg_desktop_score,
    ROUND(AVG(CASE WHEN strategy = 'mobile' THEN lcp_value END), 3) AS avg_mobile_lcp,
    ROUND(AVG(CASE WHEN strategy = 'desktop' THEN lcp_value END), 3) AS avg_desktop_lcp
FROM pagespeed_metrics
"""


def autosize_worksheet(ws) -> None:
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 28)


def append_query_results(ws, cursor: sqlite3.Cursor, query: str) -> int:
    cursor.execute(query)
    headers = [desc[0] for desc in cursor.description]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    row_count = 0
    for row in cursor.fetchall():
        ws.append(list(row))
        row_count += 1
    autosize_worksheet(ws)
    ws.freeze_panes = "A2"
    return row_count


def build_workbook() -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = OUTPUT_DIR / f"psi_lifetime_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    wb = Workbook()

    summary_ws = wb.active
    summary_ws.title = "Summary"
    summary_ws["A1"] = "PSI Lifetime Report"
    summary_ws["A1"].font = Font(bold=True, size=14)
    summary_ws["A3"] = "Generated At"
    summary_ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    summary_ws["A4"] = "Database"
    summary_ws["B4"] = str(DB_PATH)

    cursor.execute(PORTFOLIO_SUMMARY_QUERY)
    portfolio = cursor.fetchone()
    metric_labels = [
        "Total Rows",
        "Properties",
        "Distinct Dates",
        "First Metric Date",
        "Latest Metric Date",
        "Mobile Rows",
        "Desktop Rows",
        "Avg Mobile Score",
        "Avg Desktop Score",
        "Avg Mobile LCP",
        "Avg Desktop LCP",
    ]
    for idx, label in enumerate(metric_labels, start=6):
        summary_ws[f"A{idx}"] = label
        summary_ws[f"A{idx}"].font = Font(bold=True)
        summary_ws[f"B{idx}"] = portfolio[idx - 6]

    summary_ws["A18"] = "Sheet Guide"
    summary_ws["A18"].font = Font(bold=True)
    summary_ws["A19"] = "Property Summary"
    summary_ws["B19"] = "One row per property with lifetime coverage and averages"
    summary_ws["A20"] = "Latest by Property"
    summary_ws["B20"] = "Latest available full PSI metrics for each property and strategy"
    summary_ws["A21"] = "Metric Coverage"
    summary_ws["B21"] = "Per-property non-null counts for each PSI metric across lifetime records"
    summary_ws["A22"] = "PSI Raw Data"
    summary_ws["B22"] = "Full pagespeed_metrics history joined to property names"
    autosize_worksheet(summary_ws)

    property_ws = wb.create_sheet("Property Summary")
    append_query_results(property_ws, cursor, PROPERTY_SUMMARY_QUERY)

    latest_ws = wb.create_sheet("Latest by Property")
    append_query_results(latest_ws, cursor, LATEST_BY_PROPERTY_QUERY)

    coverage_ws = wb.create_sheet("Metric Coverage")
    append_query_results(coverage_ws, cursor, METRIC_COVERAGE_QUERY)

    raw_ws = wb.create_sheet("PSI Raw Data")
    append_query_results(raw_ws, cursor, RAW_QUERY)

    conn.close()
    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    output = build_workbook()
    print(output)
