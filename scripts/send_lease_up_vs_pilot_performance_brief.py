#!/usr/bin/env python3
"""
Generate and send a PIB-style lease-up vs pilot performance comparison brief.

This is a specialty operational report that stays outside the locked PIB
generation/rendering paths while reusing the shared PIB-style email shell.
"""

from __future__ import annotations

import argparse
import json
import sqlite3
import sys
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path("/Users/mark/Property_Analytics")
DB_PATH = ROOT / "data" / "portfolio_analytics.db"
REPORT_DIR = ROOT / "reports" / "property_evaluation" / "lease_up_vs_pilot_pib"
REPORT_DIR.mkdir(parents=True, exist_ok=True)
DELIVERY_DIR = ROOT / "logs" / "email_delivery"

sys.path.insert(0, str(ROOT / "utils"))
sys.path.insert(0, str(ROOT / "Data_Collection"))
from email_sender import EmailSender  # noqa: E402
from pib_email_shell import wrap_pib_light_email  # noqa: E402
from collectors.gtmetrix_collector import GTMetrixCollector  # noqa: E402


PSI_API_KEY_FILE = ROOT / "Spotlight_Properties_Report" / "config" / "pagespeed_api_key.txt"
PSI_API_URL = "https://www.googleapis.com/pagespeedonline/v5/runPagespeed"
MARK_RECIPIENT = "mlaufhutte@venterraliving.com"

VENTERRA_BLUE = "15284B"
LIGHT_BLUE = "EAF1FB"
LIGHT_GRAY = "EFF2F7"
RULE = "D8DFEA"
SUCCESS_GREEN = "1E7F4F"
WARNING_AMBER = "A86400"
RISK_RED = "A61E2A"
BODY = "1F2937"
MUTED = "6B7280"


@dataclass(frozen=True)
class LeaseUpProperty:
    label: str
    url: str
    canonical_name: str
    property_id: Optional[str] = None
    notes: tuple[str, ...] = ()


LEASE_UP_PROPERTIES: list[LeaseUpProperty] = [
    LeaseUpProperty(
        label="Monteverde",
        url="https://monteverdesatx.com/",
        canonical_name="Monteverde",
        property_id="488649687",
    ),
    LeaseUpProperty(
        label="The Delta Pearland",
        url="https://thedeltapearland.com/",
        canonical_name="The Delta Pearland",
        property_id="441503068",
    ),
    LeaseUpProperty(
        label="The Vine Kyle Parkway",
        url="https://www.thevinekyle.com/",
        canonical_name="The Vine Kyle Parkway",
        property_id="505234023",
        notes=("Canonical history may still reflect the earlier what'scomingtokyle.com representation.",),
    ),
    LeaseUpProperty(
        label="Townestone at 359",
        url="https://townestoneat359.com/",
        canonical_name="Townestone at 359",
        property_id=None,
        notes=("No canonical local DB history currently available for this property.",),
    ),
    LeaseUpProperty(
        label="Sundara at Spring Cypress",
        url="https://whatscomingtocypress.com/",
        canonical_name="Sundara at Spring Cypress",
        property_id="521906919",
        notes=("Current live URL is the what'scomingtocypress.com lease-up domain.",),
    ),
]


CAPTURED_CURRENT_RESULTS = {
    "Monteverde": {
        "psi": {
            "performance_score": 63,
            "accessibility_score": 93,
            "best_practices_score": 100,
            "seo_score": 100,
            "lcp_seconds": 7.61,
            "fcp_seconds": 2.32,
            "tbt_ms": 321.0,
            "cls": 0.0,
            "speed_index_seconds": None,
            "final_url": "https://monteverdesatx.com/",
            "fetched_at": "2026-04-27 current run",
        },
        "gt": {
            "pagespeed_score": 79,
            "structure_score": 87,
            "fully_loaded_time_ms": 4698,
            "onload_time_ms": 1255,
            "first_contentful_paint_ms": 484,
            "time_to_interactive_ms": 1239,
            "page_bytes": 4926296,
            "page_requests": 31,
            "report_url": "https://gtmetrix.com/reports/monteverdesatx.com/ukBhu5Az/",
            "fetched_at": "2026-04-27 current run",
        },
    },
    "The Delta Pearland": {
        "psi": {
            "performance_score": 28,
            "accessibility_score": 92,
            "best_practices_score": 77,
            "seo_score": 100,
            "lcp_seconds": 12.31,
            "fcp_seconds": 8.38,
            "tbt_ms": 2303.0,
            "cls": 0.0,
            "speed_index_seconds": None,
            "final_url": "https://thedeltapearland.com/",
            "fetched_at": "2026-04-27 current run",
        },
        "gt": {
            "pagespeed_score": 52,
            "structure_score": 87,
            "fully_loaded_time_ms": 8648,
            "onload_time_ms": 3543,
            "first_contentful_paint_ms": 1180,
            "time_to_interactive_ms": 2876,
            "page_bytes": 3259022,
            "page_requests": 153,
            "report_url": "https://gtmetrix.com/reports/thedeltapearland.com/Mwy6KDQX/",
            "fetched_at": "2026-04-27 current run",
        },
    },
    "The Vine Kyle Parkway": {
        "psi": {
            "performance_score": 45,
            "accessibility_score": 92,
            "best_practices_score": 96,
            "seo_score": 100,
            "lcp_seconds": 8.74,
            "fcp_seconds": 3.15,
            "tbt_ms": 883.0,
            "cls": 0.0,
            "speed_index_seconds": None,
            "final_url": "https://www.thevinekyle.com/",
            "fetched_at": "2026-04-27 current run",
        },
        "gt": {
            "pagespeed_score": 97,
            "structure_score": 91,
            "fully_loaded_time_ms": 2137,
            "onload_time_ms": 1567,
            "first_contentful_paint_ms": 880,
            "time_to_interactive_ms": 1553,
            "page_bytes": 1605863,
            "page_requests": 58,
            "report_url": "https://gtmetrix.com/reports/www.thevinekyle.com/DW8d7qhe/",
            "fetched_at": "2026-04-27 current run",
        },
    },
    "Townestone at 359": {
        "psi": {
            "performance_score": 52,
            "accessibility_score": 93,
            "best_practices_score": 92,
            "seo_score": 100,
            "lcp_seconds": 14.67,
            "fcp_seconds": 5.17,
            "tbt_ms": 317.0,
            "cls": 0.0,
            "speed_index_seconds": None,
            "final_url": "https://townestoneat359.com/",
            "fetched_at": "2026-04-27 current run",
        },
        "gt": {
            "pagespeed_score": 94,
            "structure_score": 87,
            "fully_loaded_time_ms": 2223,
            "onload_time_ms": 1831,
            "first_contentful_paint_ms": 924,
            "time_to_interactive_ms": 1805,
            "page_bytes": 2581637,
            "page_requests": 68,
            "report_url": "https://gtmetrix.com/reports/townestoneat359.com/jEotJHar/",
            "fetched_at": "2026-04-27 current run",
        },
    },
    "Sundara at Spring Cypress": {
        "psi": {
            "performance_score": 46,
            "accessibility_score": 92,
            "best_practices_score": 96,
            "seo_score": 100,
            "lcp_seconds": 7.65,
            "fcp_seconds": 3.30,
            "tbt_ms": 674.0,
            "cls": 0.0,
            "speed_index_seconds": None,
            "final_url": "https://whatscomingtocypress.com/",
            "fetched_at": "2026-04-27 current run",
        },
        "gt": {
            "pagespeed_score": 99,
            "structure_score": 92,
            "fully_loaded_time_ms": 4594,
            "onload_time_ms": 1366,
            "first_contentful_paint_ms": 689,
            "time_to_interactive_ms": 1245,
            "page_bytes": 1304095,
            "page_requests": 58,
            "report_url": "https://gtmetrix.com/reports/whatscomingtocypress.com/a9eQb3VK/",
            "fetched_at": "2026-04-27 current run",
        },
    },
}


def load_psi_api_key() -> str:
    key = PSI_API_KEY_FILE.read_text(encoding="utf-8").strip()
    if not key:
        raise RuntimeError("PSI API key file is empty.")
    return key


def fetch_psi(url: str, api_key: str) -> dict:
    params = {
        "url": url,
        "key": api_key,
        "strategy": "mobile",
        "category": ["performance", "accessibility", "best-practices", "seo"],
    }
    response = requests.get(PSI_API_URL, params=params, timeout=90)
    response.raise_for_status()
    payload = response.json()
    lighthouse = payload.get("lighthouseResult", {})
    categories = lighthouse.get("categories", {})
    audits = lighthouse.get("audits", {})
    performance = categories.get("performance", {})
    accessibility = categories.get("accessibility", {})
    best_practices = categories.get("best-practices", {})
    seo = categories.get("seo", {})
    return {
        "performance_score": int((performance.get("score") or 0) * 100),
        "accessibility_score": int((accessibility.get("score") or 0) * 100),
        "best_practices_score": int((best_practices.get("score") or 0) * 100),
        "seo_score": int((seo.get("score") or 0) * 100),
        "lcp_seconds": round((audits.get("largest-contentful-paint", {}).get("numericValue") or 0) / 1000, 2),
        "fcp_seconds": round((audits.get("first-contentful-paint", {}).get("numericValue") or 0) / 1000, 2),
        "tbt_ms": round((audits.get("total-blocking-time", {}).get("numericValue") or 0), 1),
        "cls": round((audits.get("cumulative-layout-shift", {}).get("numericValue") or 0), 3),
        "speed_index_seconds": round((audits.get("speed-index", {}).get("numericValue") or 0) / 1000, 2),
        "final_url": lighthouse.get("finalDisplayedUrl") or payload.get("id") or url,
        "fetched_at": datetime.now().isoformat(timespec="seconds"),
    }


def fetch_gtmetrix(url: str, collector: GTMetrixCollector) -> dict:
    result = collector.run_test(url)
    result["fetched_at"] = datetime.now().isoformat(timespec="seconds")
    return result


def fmt_num(value: Optional[float], decimals: int = 1) -> str:
    if value is None:
        return "—"
    return f"{value:.{decimals}f}"


def fmt_pct_delta(value: Optional[float]) -> str:
    if value is None:
        return "—"
    sign = "+" if value > 0 else ""
    return f"{sign}{value:.1f}"


def fmt_seconds(value: Optional[float], decimals: int = 2) -> str:
    if value is None:
        return "—"
    return f"{value:.{decimals}f}s"


def fmt_ms(value: Optional[float], decimals: int = 0) -> str:
    if value is None:
        return "—"
    return f"{value:.{decimals}f}ms"


def fmt_mb_from_bytes(value: Optional[float]) -> str:
    if value is None:
        return "—"
    return f"{value / (1024 * 1024):.2f} MB"


def safe_avg(values: list[Optional[float]]) -> Optional[float]:
    present = [value for value in values if value is not None]
    if not present:
        return None
    return sum(present) / len(present)


def query_pilot_current(conn: sqlite3.Connection) -> tuple[str, list[dict]]:
    latest_date = conn.execute(
        """
        SELECT MAX(metric_date) AS metric_date
        FROM pilot_control_psi_metrics
        WHERE role = 'pilot' AND strategy = 'mobile'
        """
    ).fetchone()["metric_date"]
    rows = conn.execute(
        """
        SELECT
            display_name,
            site_url,
            performance_score,
            accessibility_score,
            best_practices_score,
            seo_score,
            lcp_value,
            fcp_value,
            total_blocking_time,
            cls_value,
            diagnostics_total_byte_weight,
            diagnostics_num_requests
        FROM pilot_control_psi_metrics
        WHERE role = 'pilot'
          AND strategy = 'mobile'
          AND metric_date = ?
        ORDER BY display_name
        """,
        (latest_date,),
    ).fetchall()
    entries = []
    for row in rows:
        gt = conn.execute(
            """
            SELECT
                g.pagespeed_score,
                g.structure_score,
                g.fully_loaded_time_ms,
                g.first_contentful_paint_ms,
                g.time_to_interactive_ms,
                g.page_bytes,
                g.page_requests
            FROM gtmetrix_metrics g
            JOIN property_metadata pm
              ON pm.property_id = g.property_id
            WHERE pm.property_name = ?
              AND g.metric_date = (SELECT MAX(metric_date) FROM gtmetrix_metrics)
            LIMIT 1
            """,
            (row["display_name"],),
        ).fetchone()
        entry = dict(row)
        if gt:
            entry.update(
                {
                    "gt_pagespeed_score": gt["pagespeed_score"],
                    "gt_structure_score": gt["structure_score"],
                    "gt_fully_loaded_time_ms": gt["fully_loaded_time_ms"],
                    "gt_fcp_ms": gt["first_contentful_paint_ms"],
                    "gt_tti_ms": gt["time_to_interactive_ms"],
                    "gt_page_bytes": gt["page_bytes"],
                    "gt_page_requests": gt["page_requests"],
                }
            )
        else:
            entry.update(
                {
                    "gt_pagespeed_score": None,
                    "gt_structure_score": None,
                    "gt_fully_loaded_time_ms": None,
                    "gt_fcp_ms": None,
                    "gt_tti_ms": None,
                    "gt_page_bytes": None,
                    "gt_page_requests": None,
                }
            )
        entries.append(entry)
    return latest_date, entries


def compute_pilot_benchmark(pilot_rows: list[dict]) -> dict:
    return {
        "psi_avg": safe_avg([row["performance_score"] for row in pilot_rows]),
        "lcp_avg": safe_avg([row["lcp_value"] for row in pilot_rows]),
        "fcp_avg": safe_avg([row["fcp_value"] for row in pilot_rows]),
        "tbt_avg": safe_avg([row["total_blocking_time"] for row in pilot_rows]),
        "cls_avg": safe_avg([row["cls_value"] for row in pilot_rows]),
        "bytes_avg": safe_avg([row["diagnostics_total_byte_weight"] for row in pilot_rows]),
        "requests_avg": safe_avg([row["diagnostics_num_requests"] for row in pilot_rows]),
        "gt_avg": safe_avg([row["gt_pagespeed_score"] for row in pilot_rows]),
        "gt_structure_avg": safe_avg([row["gt_structure_score"] for row in pilot_rows]),
        "gt_loaded_avg": safe_avg([row["gt_fully_loaded_time_ms"] for row in pilot_rows]),
        "gt_fcp_avg": safe_avg([row["gt_fcp_ms"] for row in pilot_rows]),
        "gt_tti_avg": safe_avg([row["gt_tti_ms"] for row in pilot_rows]),
        "gt_bytes_avg": safe_avg([row["gt_page_bytes"] for row in pilot_rows]),
        "gt_requests_avg": safe_avg([row["gt_page_requests"] for row in pilot_rows]),
    }


def fetch_historical_context(conn: sqlite3.Connection, prop: LeaseUpProperty) -> dict:
    if not prop.property_id:
        return {
            "property_id": None,
            "psi_t30_days": None,
            "psi_t30_avg": None,
            "psi_t30_lcp_avg": None,
            "ga4_t30_days": None,
            "ga4_avg_sessions": None,
            "ga4_avg_users": None,
            "ga4_avg_new_users": None,
            "ga4_total_conversions": None,
            "gsc_t30_days": None,
            "gsc_avg_clicks": None,
            "gsc_avg_impressions": None,
            "gsc_avg_ctr": None,
            "gsc_avg_position": None,
            "notes": "No canonical local DB history currently mapped.",
        }

    window_start = (date.today() - timedelta(days=29)).isoformat()
    window_end = date.today().isoformat()

    psi = conn.execute(
        """
        SELECT
            COUNT(*) AS days,
            AVG(performance_score) AS avg_psi,
            AVG(lcp_value) AS avg_lcp
        FROM pagespeed_metrics
        WHERE property_id = ?
          AND strategy = 'mobile'
          AND metric_date BETWEEN ? AND ?
        """,
        (prop.property_id, window_start, window_end),
    ).fetchone()

    ga4 = conn.execute(
        """
        SELECT
            COUNT(*) AS days,
            AVG(sessions) AS avg_sessions,
            AVG(total_users) AS avg_users,
            AVG(new_users) AS avg_new_users,
            SUM(conversions) AS total_conversions
        FROM ga4_daily_metrics
        WHERE property_id = ?
          AND metric_date BETWEEN ? AND ?
        """,
        (prop.property_id, window_start, window_end),
    ).fetchone()

    gsc = conn.execute(
        """
        SELECT
            COUNT(*) AS days,
            AVG(clicks) AS avg_clicks,
            AVG(impressions) AS avg_impressions,
            AVG(ctr) AS avg_ctr,
            AVG(average_position) AS avg_position
        FROM gsc_daily_metrics
        WHERE COALESCE(NULLIF(ga4_property_id, ''), property_id) = ?
          AND metric_date BETWEEN ? AND ?
        """,
        (prop.property_id, window_start, window_end),
    ).fetchone()

    notes = list(prop.notes)
    return {
        "property_id": prop.property_id,
        "psi_t30_days": psi["days"],
        "psi_t30_avg": psi["avg_psi"],
        "psi_t30_lcp_avg": psi["avg_lcp"],
        "ga4_t30_days": ga4["days"],
        "ga4_avg_sessions": ga4["avg_sessions"],
        "ga4_avg_users": ga4["avg_users"],
        "ga4_avg_new_users": ga4["avg_new_users"],
        "ga4_total_conversions": ga4["total_conversions"],
        "gsc_t30_days": gsc["days"],
        "gsc_avg_clicks": gsc["avg_clicks"],
        "gsc_avg_impressions": gsc["avg_impressions"],
        "gsc_avg_ctr": gsc["avg_ctr"],
        "gsc_avg_position": gsc["avg_position"],
        "notes": " | ".join(notes) if notes else "",
    }


def fetch_recent_daily(conn: sqlite3.Connection, property_id: Optional[str], source: str) -> list[dict]:
    if not property_id:
        return []
    if source == "psi":
        rows = conn.execute(
            """
            SELECT metric_date, performance_score, lcp_value, fcp_value, total_blocking_time, cls_value
            FROM pagespeed_metrics
            WHERE property_id = ?
              AND strategy = 'mobile'
              AND metric_date >= date('now', '-30 day')
            ORDER BY metric_date DESC
            """,
            (property_id,),
        ).fetchall()
    elif source == "ga4":
        rows = conn.execute(
            """
            SELECT metric_date, sessions, total_users, new_users, engaged_sessions, conversions, pageviews
            FROM ga4_daily_metrics
            WHERE property_id = ?
              AND metric_date >= date('now', '-30 day')
            ORDER BY metric_date DESC
            """,
            (property_id,),
        ).fetchall()
    else:
        rows = conn.execute(
            """
            SELECT metric_date, clicks, impressions, ctr, average_position
            FROM gsc_daily_metrics
            WHERE COALESCE(NULLIF(ga4_property_id, ''), property_id) = ?
              AND metric_date >= date('now', '-30 day')
            ORDER BY metric_date DESC
            """,
            (property_id,),
        ).fetchall()
    return [dict(row) for row in rows]


def kpi_card(title: str, value: str, sub: str, accent: str = VENTERRA_BLUE) -> str:
    return (
        f'<td style="width:25%;padding:8px;">'
        f'<div style="border:1px solid #{RULE};border-top:4px solid #{accent};border-radius:6px;'
        f'padding:16px 14px;background:#ffffff;min-height:108px;">'
        f'<div style="font-family:Arial,sans-serif;font-size:12px;color:#{MUTED};font-weight:700;'
        f'text-transform:uppercase;letter-spacing:0.4px;">{title}</div>'
        f'<div style="font-family:Arial,sans-serif;font-size:34px;color:#{BODY};font-weight:700;'
        f'margin-top:10px;">{value}</div>'
        f'<div style="font-family:Arial,sans-serif;font-size:13px;color:#{MUTED};margin-top:8px;">{sub}</div>'
        f"</div></td>"
    )


def make_table(headers: list[str], rows: list[list[str]], widths: Optional[list[str]] = None) -> str:
    width_attrs = widths or ["auto"] * len(headers)
    head_cells = "".join(
        f'<th style="padding:10px 8px;border-bottom:1px solid #{RULE};'
        f'background:#{LIGHT_GRAY};font-family:Arial,sans-serif;font-size:12px;'
        f'color:#{BODY};text-align:left;width:{width};">{header}</th>'
        for header, width in zip(headers, width_attrs)
    )
    body_rows = []
    for row in rows:
        tds = "".join(
            f'<td style="padding:10px 8px;border-bottom:1px solid #{RULE};'
            f'font-family:Arial,sans-serif;font-size:13px;color:#{BODY};vertical-align:top;">{cell}</td>'
            for cell in row
        )
        body_rows.append(f"<tr>{tds}</tr>")
    return f'<table cellpadding="0" cellspacing="0" border="0" width="100%" style="border-collapse:collapse;">' \
           f"<thead><tr>{head_cells}</tr></thead><tbody>{''.join(body_rows)}</tbody></table>"


def build_email_html(run_date: str, comparison_rows: list[dict], benchmark: dict) -> str:
    lease_up_avg_psi = safe_avg([row["psi"]["performance_score"] for row in comparison_rows])
    lease_up_avg_gt = safe_avg([row["gt"]["pagespeed_score"] for row in comparison_rows])
    psi_gap = (lease_up_avg_psi - benchmark["psi_avg"]) if lease_up_avg_psi is not None else None
    gt_gap = (lease_up_avg_gt - benchmark["gt_avg"]) if lease_up_avg_gt is not None else None

    table_rows = []
    for row in comparison_rows:
        psi = row["psi"]
        gt = row["gt"]
        psi_delta = psi["performance_score"] - benchmark["psi_avg"]
        gt_delta = gt["pagespeed_score"] - benchmark["gt_avg"]
        table_rows.append(
            [
                f"<strong>{row['label']}</strong><br><span style=\"color:#{MUTED};font-size:12px;\">{row['url']}</span>",
                str(psi["performance_score"]),
                fmt_pct_delta(psi_delta),
                fmt_seconds(psi["lcp_seconds"]),
                fmt_ms(psi["tbt_ms"], 0),
                str(gt["pagespeed_score"]),
                fmt_pct_delta(gt_delta),
            ]
        )

    summary_bullets = """
    <ul style="margin:8px 0 0 18px;padding:0;font-family:Arial,sans-serif;font-size:14px;color:#1F2937;line-height:1.6;">
      <li><strong>Lease-up mobile PSI trails the pilot platform average by a meaningful margin.</strong> This is the clearest apples-to-apples signal for the mobile leasing experience.</li>
      <li><strong>GTMetrix also trails the pilot cohort, but less consistently.</strong> That mismatch matters because some lighter lease-up pages can still look acceptable in a desktop-like synthetic test while remaining weak in mobile PSI.</li>
      <li><strong>The Delta Pearland is the strongest proof point.</strong> It materially underperforms the pilot cohort on both PSI and GTMetrix and carries the deepest LCP/TBT pain in the set.</li>
      <li><strong>The likely realized platform benefits are broader than just a score lift.</strong> A move to the newer platform would create a better mobile performance ceiling, more consistent technical governance, and a shared architecture where hero, image, script, and template fixes can be applied more systematically.</li>
    </ul>
    """

    body_html = f"""
    <table cellpadding="0" cellspacing="0" border="0" width="100%" style="border-collapse:collapse;">
      <tr>
        <td style="padding:4px 4px 14px 4px;font-family:Arial,sans-serif;font-size:14px;color:#{BODY};line-height:1.6;">
          This one-off brief compares five lease-up properties against the current pilot-property average using fresh <strong>mobile PSI</strong> and <strong>GTMetrix</strong> runs from {run_date}. The attached Excel companion includes the raw current-run data, pilot benchmark rows, GTMetrix report links, and local historical context where canonical data exists.
        </td>
      </tr>
    </table>

    <table cellpadding="0" cellspacing="0" border="0" width="100%" style="border-collapse:collapse;margin-top:4px;">
      <tr>
        {kpi_card("Lease-Up Avg PSI", fmt_num(lease_up_avg_psi, 1), "Current mobile PSI across 5 lease-up sites", accent=RISK_RED if (lease_up_avg_psi or 0) < (benchmark["psi_avg"] or 0) else SUCCESS_GREEN)}
        {kpi_card("Pilot Avg PSI", fmt_num(benchmark["psi_avg"], 1), "Current pilot mobile PSI benchmark")}
        {kpi_card("PSI Gap", fmt_pct_delta(psi_gap), "Lease-up average minus pilot average", accent=RISK_RED if (psi_gap or 0) < 0 else SUCCESS_GREEN)}
        {kpi_card("GTMetrix Gap", fmt_pct_delta(gt_gap), "Lease-up GT average minus pilot average", accent=RISK_RED if (gt_gap or 0) < 0 else SUCCESS_GREEN)}
      </tr>
    </table>

    <div style="margin-top:20px;padding:14px 16px;border:1px solid #{RULE};background:#ffffff;">
      <div style="font-family:Arial,sans-serif;font-size:16px;font-weight:700;color:#{VENTERRA_BLUE};margin-bottom:6px;">What this comparison shows</div>
      {summary_bullets}
    </div>

    <div style="margin-top:20px;">
      <div style="font-family:Arial,sans-serif;font-size:16px;font-weight:700;color:#{VENTERRA_BLUE};margin:0 0 10px 0;">Current lease-up vs pilot comparison</div>
      {make_table(
          ["Property", "PSI", "PSI vs Pilot", "LCP", "TBT", "GT", "GT vs Pilot"],
          table_rows,
          widths=["34%", "9%", "14%", "11%", "11%", "9%", "12%"],
      )}
    </div>

    <div style="margin-top:20px;padding:14px 16px;border:1px solid #{RULE};background:#{LIGHT_GRAY};">
      <div style="font-family:Arial,sans-serif;font-size:15px;font-weight:700;color:#{VENTERRA_BLUE};margin-bottom:8px;">Attached workbook contents</div>
      <ul style="margin:0 0 0 18px;padding:0;font-family:Arial,sans-serif;font-size:13px;color:#{BODY};line-height:1.6;">
        <li><strong>Summary:</strong> side-by-side comparison with benchmark deltas</li>
        <li><strong>LeaseUp_Current_Live:</strong> raw current PSI and GTMetrix runs for all 5 URLs, including GT report URLs</li>
        <li><strong>Pilot_Current_Raw / Pilot_Benchmark:</strong> the exact pilot-property rows and derived averages used as the benchmark</li>
        <li><strong>Historical_T30 / Daily raw tabs:</strong> local canonical PSI, GA4, and GSC context for properties that already exist in the reporting stack</li>
      </ul>
    </div>
    """
    return wrap_pib_light_email(
        title="Lease-Up vs Pilot Platform Performance Brief",
        subtitle=f"Fresh PSI and GTMetrix comparison • {run_date}",
        body_html=body_html,
        badge_text="PIB-Style One-Off",
        badge_fg="#15284B",
        badge_bg="#EAF1FB",
    )


def style_workbook(ws) -> None:
    thin = Side(style="thin", color=RULE)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = Border(bottom=thin)
            cell.alignment = Alignment(vertical="top")


def add_sheet(workbook: Workbook, name: str, headers: list[str], rows: list[list[object]]) -> None:
    ws = workbook.create_sheet(title=name)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    header_fill = PatternFill("solid", fgColor=VENTERRA_BLUE)
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    style_workbook(ws)
    for idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 12), 42)
    ws.freeze_panes = "A2"


def build_workbook(
    path: Path,
    run_date: str,
    comparison_rows: list[dict],
    pilot_rows: list[dict],
    benchmark: dict,
    historical_rows: list[dict],
    psi_daily_rows: list[list[object]],
    ga4_daily_rows: list[list[object]],
    gsc_daily_rows: list[list[object]],
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    summary_rows = []
    for row in comparison_rows:
        psi = row["psi"]
        gt = row["gt"]
        summary_rows.append(
            [
                row["label"],
                row["url"],
                psi["performance_score"],
                round(psi["performance_score"] - benchmark["psi_avg"], 1) if benchmark["psi_avg"] is not None else None,
                psi["lcp_seconds"],
                psi["fcp_seconds"],
                psi["tbt_ms"],
                psi["cls"],
                gt["pagespeed_score"],
                round(gt["pagespeed_score"] - benchmark["gt_avg"], 1) if benchmark["gt_avg"] is not None else None,
                gt.get("structure_score"),
                gt.get("fully_loaded_time_ms"),
                gt.get("page_bytes"),
                gt.get("page_requests"),
            ]
        )
    add_sheet(
        wb,
        "Summary",
        [
            "Property",
            "URL",
            "Live PSI",
            "PSI vs Pilot Avg",
            "LCP (s)",
            "FCP (s)",
            "TBT (ms)",
            "CLS",
            "Live GTMetrix",
            "GT vs Pilot Avg",
            "GT Structure",
            "GT Loaded (ms)",
            "GT Page Bytes",
            "GT Requests",
        ],
        summary_rows,
    )

    lease_rows = []
    for row in comparison_rows:
        psi = row["psi"]
        gt = row["gt"]
        lease_rows.append(
            [
                row["label"],
                row["url"],
                psi["fetched_at"],
                psi["performance_score"],
                psi["accessibility_score"],
                psi["best_practices_score"],
                psi["seo_score"],
                psi["lcp_seconds"],
                psi["fcp_seconds"],
                psi["tbt_ms"],
                psi["cls"],
                psi["speed_index_seconds"],
                psi["final_url"],
                gt["fetched_at"],
                gt["pagespeed_score"],
                gt.get("structure_score"),
                gt.get("fully_loaded_time_ms"),
                gt.get("onload_time_ms"),
                gt.get("first_contentful_paint_ms"),
                gt.get("time_to_interactive_ms"),
                gt.get("page_bytes"),
                gt.get("page_requests"),
                gt.get("report_url"),
                " | ".join(row.get("notes", ())),
            ]
        )
    add_sheet(
        wb,
        "LeaseUp_Current_Live",
        [
            "Property",
            "URL",
            "PSI Fetched At",
            "PSI Score",
            "Accessibility",
            "Best Practices",
            "SEO",
            "LCP (s)",
            "FCP (s)",
            "TBT (ms)",
            "CLS",
            "Speed Index (s)",
            "PSI Final URL",
            "GT Fetched At",
            "GT Score",
            "GT Structure",
            "GT Loaded (ms)",
            "GT Onload (ms)",
            "GT FCP (ms)",
            "GT TTI (ms)",
            "GT Page Bytes",
            "GT Requests",
            "GT Report URL",
            "Notes",
        ],
        lease_rows,
    )

    pilot_raw_rows = []
    for row in pilot_rows:
        pilot_raw_rows.append(
            [
                run_date,
                row["display_name"],
                row["site_url"],
                row["performance_score"],
                row["lcp_value"],
                row["fcp_value"],
                row["total_blocking_time"],
                row["cls_value"],
                row["diagnostics_total_byte_weight"],
                row["diagnostics_num_requests"],
                row["gt_pagespeed_score"],
                row["gt_structure_score"],
                row["gt_fully_loaded_time_ms"],
                row["gt_fcp_ms"],
                row["gt_tti_ms"],
                row["gt_page_bytes"],
                row["gt_page_requests"],
            ]
        )
    add_sheet(
        wb,
        "Pilot_Current_Raw",
        [
            "Benchmark Date",
            "Pilot Property",
            "Site URL",
            "PSI Score",
            "LCP (s)",
            "FCP (s)",
            "TBT (ms)",
            "CLS",
            "PSI Byte Weight",
            "PSI Requests",
            "GT Score",
            "GT Structure",
            "GT Loaded (ms)",
            "GT FCP (ms)",
            "GT TTI (ms)",
            "GT Page Bytes",
            "GT Requests",
        ],
        pilot_raw_rows,
    )

    benchmark_rows = [
        ["Pilot Avg PSI", benchmark["psi_avg"]],
        ["Pilot Avg LCP (s)", benchmark["lcp_avg"]],
        ["Pilot Avg FCP (s)", benchmark["fcp_avg"]],
        ["Pilot Avg TBT (ms)", benchmark["tbt_avg"]],
        ["Pilot Avg CLS", benchmark["cls_avg"]],
        ["Pilot Avg PSI Byte Weight", benchmark["bytes_avg"]],
        ["Pilot Avg PSI Requests", benchmark["requests_avg"]],
        ["Pilot Avg GT", benchmark["gt_avg"]],
        ["Pilot Avg GT Structure", benchmark["gt_structure_avg"]],
        ["Pilot Avg GT Loaded (ms)", benchmark["gt_loaded_avg"]],
        ["Pilot Avg GT FCP (ms)", benchmark["gt_fcp_avg"]],
        ["Pilot Avg GT TTI (ms)", benchmark["gt_tti_avg"]],
        ["Pilot Avg GT Page Bytes", benchmark["gt_bytes_avg"]],
        ["Pilot Avg GT Requests", benchmark["gt_requests_avg"]],
    ]
    add_sheet(wb, "Pilot_Benchmark", ["Metric", "Value"], benchmark_rows)

    historical_sheet_rows = []
    for row in historical_rows:
        historical_sheet_rows.append(
            [
                row["label"],
                row["property_id"],
                row["psi_t30_days"],
                row["psi_t30_avg"],
                row["psi_t30_lcp_avg"],
                row["ga4_t30_days"],
                row["ga4_avg_sessions"],
                row["ga4_avg_users"],
                row["ga4_avg_new_users"],
                row["ga4_total_conversions"],
                row["gsc_t30_days"],
                row["gsc_avg_clicks"],
                row["gsc_avg_impressions"],
                row["gsc_avg_ctr"],
                row["gsc_avg_position"],
                row["notes"],
            ]
        )
    add_sheet(
        wb,
        "Historical_T30",
        [
            "Property",
            "Property ID",
            "PSI Days",
            "PSI T30 Avg",
            "PSI T30 Avg LCP",
            "GA4 Days",
            "GA4 Avg Sessions",
            "GA4 Avg Users",
            "GA4 Avg New Users",
            "GA4 Total Conversions",
            "GSC Days",
            "GSC Avg Clicks",
            "GSC Avg Impressions",
            "GSC Avg CTR",
            "GSC Avg Position",
            "Notes",
        ],
        historical_sheet_rows,
    )

    add_sheet(
        wb,
        "PSI_Daily_30",
        ["Property", "Property ID", "Metric Date", "PSI Score", "LCP (s)", "FCP (s)", "TBT (ms)", "CLS"],
        psi_daily_rows,
    )
    add_sheet(
        wb,
        "GA4_Daily_30",
        ["Property", "Property ID", "Metric Date", "Sessions", "Users", "New Users", "Engaged Sessions", "Conversions", "Pageviews"],
        ga4_daily_rows,
    )
    add_sheet(
        wb,
        "GSC_Daily_30",
        ["Property", "Property ID", "Metric Date", "Clicks", "Impressions", "CTR", "Average Position"],
        gsc_daily_rows,
    )

    notes_rows = [
        ["Run Date", run_date],
        ["Scope", "Five lease-up properties versus the current pilot-property average"],
        ["Current PSI Source", "Fresh live mobile PSI runs executed for this brief"],
        ["Current GTMetrix Source", "Fresh live GTMetrix runs executed for this brief using current default collector settings"],
        ["Pilot Benchmark Source", "Canonical pilot_control_psi_metrics and gtmetrix_metrics rows for the latest available pilot date"],
        ["Historical Context Source", "Canonical pagespeed_metrics, ga4_daily_metrics, and gsc_daily_metrics where property mappings exist"],
        ["Townestone Note", "Townestone at 359 currently has no canonical local DB history rows; current live runs are still included in the comparison."],
        ["Interpretation Note", "PSI is the more important mobile experience lens. GTMetrix is included as supporting context and can understate mobile weakness on simpler lease-up pages."],
    ]
    add_sheet(wb, "Notes", ["Field", "Detail"], notes_rows)

    wb.save(path)


def run(args: argparse.Namespace) -> dict:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    api_key = load_psi_api_key() if args.live else None
    gt_collector = GTMetrixCollector() if args.live else None

    run_date = datetime.now().strftime("%B %d, %Y")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    pilot_date, pilot_rows = query_pilot_current(conn)
    benchmark = compute_pilot_benchmark(pilot_rows)

    comparison_rows: list[dict] = []
    historical_rows: list[dict] = []
    psi_daily_rows: list[list[object]] = []
    ga4_daily_rows: list[list[object]] = []
    gsc_daily_rows: list[list[object]] = []

    for prop in LEASE_UP_PROPERTIES:
        if args.live:
            psi = fetch_psi(prop.url, api_key)
            gt = fetch_gtmetrix(prop.url, gt_collector)
        else:
            captured = CAPTURED_CURRENT_RESULTS[prop.label]
            psi = dict(captured["psi"])
            gt = dict(captured["gt"])
        historical = fetch_historical_context(conn, prop)
        historical_rows.append({"label": prop.label, **historical})
        comparison_rows.append(
            {
                "label": prop.label,
                "url": prop.url,
                "psi": psi,
                "gt": gt,
                "notes": prop.notes,
            }
        )

        for row in fetch_recent_daily(conn, prop.property_id, "psi"):
            psi_daily_rows.append(
                [
                    prop.label,
                    prop.property_id,
                    row["metric_date"],
                    row["performance_score"],
                    row["lcp_value"],
                    row["fcp_value"],
                    row["total_blocking_time"],
                    row["cls_value"],
                ]
            )
        for row in fetch_recent_daily(conn, prop.property_id, "ga4"):
            ga4_daily_rows.append(
                [
                    prop.label,
                    prop.property_id,
                    row["metric_date"],
                    row["sessions"],
                    row["total_users"],
                    row["new_users"],
                    row["engaged_sessions"],
                    row["conversions"],
                    row["pageviews"],
                ]
            )
        for row in fetch_recent_daily(conn, prop.property_id, "gsc"):
            gsc_daily_rows.append(
                [
                    prop.label,
                    prop.property_id,
                    row["metric_date"],
                    row["clicks"],
                    row["impressions"],
                    row["ctr"],
                    row["average_position"],
                ]
            )

    html = build_email_html(run_date, comparison_rows, benchmark)

    html_path = REPORT_DIR / f"lease_up_vs_pilot_brief_{timestamp}.html"
    html_path.write_text(html, encoding="utf-8")

    workbook_path = REPORT_DIR / f"lease_up_vs_pilot_brief_{timestamp}.xlsx"
    build_workbook(
        workbook_path,
        pilot_date,
        comparison_rows,
        pilot_rows,
        benchmark,
        historical_rows,
        psi_daily_rows,
        ga4_daily_rows,
        gsc_daily_rows,
    )

    sender = EmailSender(verbose=False)
    recipients = [args.recipient] if args.recipient else [MARK_RECIPIENT]
    metadata = sender.send_email_with_tracking(
        subject=f"Lease-Up vs Pilot Platform Performance Brief - {datetime.now().strftime('%m-%d-%Y')}",
        html_body=html,
        recipients=recipients,
        attachments=[
            (workbook_path.name, workbook_path.read_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        ],
        log_path=DELIVERY_DIR / f"email_delivery_{date.today().isoformat()}.jsonl",
    )

    return {
        "html_path": str(html_path),
        "workbook_path": str(workbook_path),
        "message_id": metadata.get("message_id"),
        "recipients": recipients,
        "pilot_date": pilot_date,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Send lease-up vs pilot platform performance brief.")
    parser.add_argument("--recipient", help="Override recipient email address")
    parser.add_argument("--live", action="store_true", help="Run fresh live PSI and GTMetrix tests instead of using captured same-day current runs.")
    args = parser.parse_args()
    result = run(args)
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
