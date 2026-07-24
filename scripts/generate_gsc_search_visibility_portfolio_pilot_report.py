#!/usr/bin/env python3
"""Generate portfolio organic search performance and five-site pilot comparison report."""

from __future__ import annotations

import csv
import html
import json
import sqlite3
import sys
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path("/Users/mark/Property_Analytics")
DB_PATH = ROOT / "data" / "portfolio_analytics.db"
REGISTRY_PATH = ROOT / "config" / "venterra_properties_official.json"
PILOT_CONFIG_PATH = ROOT / "pilot_control_cwv" / "config" / "pilot_control_cwv_config.json"
OUT_ROOT = ROOT / "reports" / "organic_search_performance_portfolio_pilot"
REQUESTED_START = date(2024, 6, 25)
PILOT_LAUNCH_DATE = date(2026, 3, 24)

sys.path.insert(0, str(ROOT))
from Data_Collection.utils.property_identity import resolve_property_identity  # noqa: E402


BRAND = {
    "navy": "15284B",
    "san_marino": "3D66B9",
    "bay": "294782",
    "indigo": "5A81CF",
    "monte_carlo": "7DCAC2",
    "pink": "E02472",
    "white_smoke": "F6F6F5",
    "terra_cotta": "BD4830",
    "quill": "D6D6D2",
    "blue_chill": "3B9189",
    "delta": "9B9B96",
    "black": "000000",
    "white": "FFFFFF",
}


@dataclass(frozen=True)
class Property:
    name: str
    property_id: str
    property_code: str | None
    gsc_url: str | None


@dataclass(frozen=True)
class PilotPair:
    pair_key: str
    pilot_name: str
    pilot_id: str
    control_name: str
    control_id: str


def iso(value: date) -> str:
    return value.isoformat()


def parse_date(value: str) -> date:
    return datetime.strptime(value, "%Y-%m-%d").date()


def fmt_num(value: Any, decimals: int = 0) -> str:
    if value is None:
        return "-"
    return f"{float(value):,.{decimals}f}"


def fmt_pct(value: Any, decimals: int = 1, signed: bool = False) -> str:
    if value is None:
        return "-"
    sign = "+" if signed else ""
    return f"{float(value) * 100:{sign},.{decimals}f}%"


def fmt_pos(value: Any) -> str:
    if value is None:
        return "-"
    return f"{float(value):,.1f}"


def pct_change(current: float | int | None, prior: float | int | None) -> float | None:
    if current is None or prior in (None, 0):
        return None
    return (float(current) - float(prior)) / float(prior)


def placeholders(values: Iterable[str]) -> str:
    return ",".join("?" for _ in values)


def load_portfolio_properties() -> list[Property]:
    payload = json.loads(REGISTRY_PATH.read_text(encoding="utf-8"))
    properties: list[Property] = []
    for row in payload.get("properties", []):
        ga4 = row.get("ga4_property_id")
        if not ga4 or row.get("property_type") == "new_development":
            continue
        identity = resolve_property_identity(str(ga4)) or resolve_property_identity(str(row.get("name", "")))
        properties.append(
            Property(
                name=(identity.property_name if identity else row.get("name")) or str(ga4),
                property_id=str(ga4),
                property_code=identity.property_code if identity else None,
                gsc_url=identity.gsc_url if identity else row.get("gsc_url"),
            )
        )
    return sorted(properties, key=lambda item: item.name)


def load_pilot_pairs() -> list[PilotPair]:
    payload = json.loads(PILOT_CONFIG_PATH.read_text(encoding="utf-8"))
    cohorts = {item["key"]: item for item in payload.get("cohorts", []) if item.get("active")}
    pairs: list[PilotPair] = []
    for pilot in [item for item in cohorts.values() if item.get("role") == "pilot"]:
        control = cohorts.get(str(pilot.get("sister_key")))
        if not control:
            continue
        pilot_identity = resolve_property_identity(str(pilot["property_id"])) or resolve_property_identity(str(pilot["display_name"]))
        control_identity = resolve_property_identity(str(control["property_id"])) or resolve_property_identity(str(control["display_name"]))
        pairs.append(
            PilotPair(
                pair_key=str(pilot["key"]).replace("pilot_", ""),
                pilot_name=pilot_identity.property_name if pilot_identity else str(pilot["display_name"]),
                pilot_id=str(pilot["property_id"]),
                control_name=control_identity.property_name if control_identity else str(control["display_name"]),
                control_id=str(control["property_id"]),
            )
        )
    return sorted(pairs, key=lambda item: item.pilot_name)


def fetch_one(conn: sqlite3.Connection, sql: str, params: tuple[Any, ...] = ()) -> sqlite3.Row:
    row = conn.execute(sql, params).fetchone()
    if row is None:
        raise RuntimeError("Query returned no row")
    return row


def fetch_gsc_summary(conn: sqlite3.Connection, property_ids: list[str], start: date, end: date) -> dict[str, Any]:
    ph = placeholders(property_ids)
    row = fetch_one(
        conn,
        f"""
        SELECT
            COUNT(DISTINCT metric_date) AS days,
            SUM(clicks) AS clicks,
            SUM(impressions) AS impressions,
            CASE WHEN SUM(impressions) > 0 THEN CAST(SUM(clicks) AS REAL) / SUM(impressions) END AS ctr,
            CASE WHEN SUM(impressions) > 0 THEN SUM(average_position * impressions) / SUM(impressions) END AS avg_position
        FROM gsc_daily_metrics
        WHERE COALESCE(ga4_property_id, property_id) IN ({ph})
          AND metric_date BETWEEN ? AND ?
        """,
        tuple([*property_ids, iso(start), iso(end)]),
    )
    return dict(row)


def fetch_monthly(conn: sqlite3.Connection, property_ids: list[str], start: date, end: date) -> list[dict[str, Any]]:
    ph = placeholders(property_ids)
    rows = conn.execute(
        f"""
        SELECT
            substr(metric_date, 1, 7) AS month,
            SUM(clicks) AS clicks,
            SUM(impressions) AS impressions,
            CASE WHEN SUM(impressions) > 0 THEN CAST(SUM(clicks) AS REAL) / SUM(impressions) END AS ctr,
            CASE WHEN SUM(impressions) > 0 THEN SUM(average_position * impressions) / SUM(impressions) END AS avg_position,
            COUNT(DISTINCT COALESCE(ga4_property_id, property_id)) AS properties_with_data,
            COUNT(DISTINCT metric_date) AS days
        FROM gsc_daily_metrics
        WHERE COALESCE(ga4_property_id, property_id) IN ({ph})
          AND metric_date BETWEEN ? AND ?
        GROUP BY month
        ORDER BY month
        """,
        tuple([*property_ids, iso(start), iso(end)]),
    ).fetchall()
    return [dict(row) for row in rows]


def fetch_property_rankings(conn: sqlite3.Connection, properties: list[Property], start: date, end: date) -> list[dict[str, Any]]:
    property_ids = [item.property_id for item in properties]
    ph = placeholders(property_ids)
    name_map = {item.property_id: item for item in properties}
    rows = conn.execute(
        f"""
        SELECT
            COALESCE(ga4_property_id, property_id) AS property_id,
            SUM(clicks) AS clicks,
            SUM(impressions) AS impressions,
            CASE WHEN SUM(impressions) > 0 THEN CAST(SUM(clicks) AS REAL) / SUM(impressions) END AS ctr,
            CASE WHEN SUM(impressions) > 0 THEN SUM(average_position * impressions) / SUM(impressions) END AS avg_position,
            MIN(metric_date) AS first_date,
            MAX(metric_date) AS last_date,
            COUNT(DISTINCT metric_date) AS days
        FROM gsc_daily_metrics
        WHERE COALESCE(ga4_property_id, property_id) IN ({ph})
          AND metric_date BETWEEN ? AND ?
        GROUP BY COALESCE(ga4_property_id, property_id)
        ORDER BY clicks DESC
        """,
        tuple([*property_ids, iso(start), iso(end)]),
    ).fetchall()
    rankings: list[dict[str, Any]] = []
    for idx, row in enumerate(rows, start=1):
        prop = name_map[str(row["property_id"])]
        item = dict(row)
        item["rank"] = idx
        item["property_name"] = prop.name
        item["property_code"] = prop.property_code
        item["gsc_url"] = prop.gsc_url
        rankings.append(item)
    return rankings


def fetch_pilot_post_launch(conn: sqlite3.Connection, pairs: list[PilotPair], start: date, end: date) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for pair in pairs:
        pilot = fetch_gsc_summary(conn, [pair.pilot_id], start, end)
        control = fetch_gsc_summary(conn, [pair.control_id], start, end)
        rows.append(
            {
                "pair_key": pair.pair_key,
                "pilot_property": pair.pilot_name,
                "pilot_property_id": pair.pilot_id,
                "control_property": pair.control_name,
                "control_property_id": pair.control_id,
                "pilot_clicks": pilot.get("clicks") or 0,
                "control_clicks": control.get("clicks") or 0,
                "click_gap": (pilot.get("clicks") or 0) - (control.get("clicks") or 0),
                "click_gap_pct": pct_change(pilot.get("clicks") or 0, control.get("clicks") or 0),
                "pilot_impressions": pilot.get("impressions") or 0,
                "control_impressions": control.get("impressions") or 0,
                "impression_gap_pct": pct_change(pilot.get("impressions") or 0, control.get("impressions") or 0),
                "pilot_ctr": pilot.get("ctr"),
                "control_ctr": control.get("ctr"),
                "pilot_avg_position": pilot.get("avg_position"),
                "control_avg_position": control.get("avg_position"),
            }
        )
    return sorted(rows, key=lambda item: item["pilot_clicks"], reverse=True)


def fetch_pilot_t30(conn: sqlite3.Connection, pairs: list[PilotPair], current_end: date) -> list[dict[str, Any]]:
    current_start = current_end - timedelta(days=29)
    prior_end = current_start - timedelta(days=1)
    prior_start = prior_end - timedelta(days=29)
    rows: list[dict[str, Any]] = []
    for pair in pairs:
        current = fetch_gsc_summary(conn, [pair.pilot_id], current_start, current_end)
        prior = fetch_gsc_summary(conn, [pair.pilot_id], prior_start, prior_end)
        rows.append(
            {
                "pilot_property": pair.pilot_name,
                "clicks_t30": current.get("clicks") or 0,
                "clicks_prior30": prior.get("clicks") or 0,
                "click_change_pct": pct_change(current.get("clicks") or 0, prior.get("clicks") or 0),
                "impressions_t30": current.get("impressions") or 0,
                "impressions_prior30": prior.get("impressions") or 0,
                "impression_change_pct": pct_change(current.get("impressions") or 0, prior.get("impressions") or 0),
                "ctr_t30": current.get("ctr"),
                "ctr_prior30": prior.get("ctr"),
                "avg_position_t30": current.get("avg_position"),
                "avg_position_prior30": prior.get("avg_position"),
            }
        )
    return sorted(rows, key=lambda item: item["clicks_t30"], reverse=True)


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        return
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def autosize(ws) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 11), 42)


def write_sheet(ws, rows: list[dict[str, Any]]) -> None:
    if not rows:
        return
    header_fill = PatternFill("solid", fgColor=BRAND["navy"])
    header_font = Font(color=BRAND["white"], bold=True)
    for col_idx, key in enumerate(rows[0].keys(), start=1):
        cell = ws.cell(row=1, column=col_idx, value=key)
        cell.fill = header_fill
        cell.font = header_font
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, key in enumerate(row.keys(), start=1):
            ws.cell(row=row_idx, column=col_idx, value=row[key])
    ws.freeze_panes = "A2"
    autosize(ws)


def build_workbook(path: Path, sheets: dict[str, list[dict[str, Any]]], metadata: list[list[Any]]) -> None:
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary["A1"] = "Organic Search Performance Portfolio And Pilot Report"
    summary["A1"].font = Font(bold=True, size=16, color=BRAND["navy"])
    for row_idx, row in enumerate(metadata, start=3):
        summary.cell(row=row_idx, column=1, value=row[0]).font = Font(bold=True, color=BRAND["bay"])
        summary.cell(row=row_idx, column=2, value=row[1])
    autosize(summary)
    for title, rows in sheets.items():
        ws = wb.create_sheet(title[:31])
        write_sheet(ws, rows)
        if title == "Monthly Portfolio" and len(rows) >= 2:
            chart = LineChart()
            chart.title = "Portfolio GSC Clicks By Month"
            chart.y_axis.title = "GSC clicks"
            chart.x_axis.title = "Month"
            data = Reference(ws, min_col=2, min_row=1, max_row=len(rows) + 1)
            cats = Reference(ws, min_col=1, min_row=2, max_row=len(rows) + 1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 8
            chart.width = 18
            ws.add_chart(chart, "J2")
    wb.save(path)


def html_table(rows: list[dict[str, Any]], columns: list[tuple[str, str]], limit: int | None = None) -> str:
    visible = rows[:limit] if limit else rows
    header = "".join(f"<th>{html.escape(label)}</th>" for _, label in columns)
    body_rows = []
    for row in visible:
        cells = "".join(f"<td>{html.escape(str(row.get(key, '-')))}</td>" for key, _ in columns)
        body_rows.append(f"<tr>{cells}</tr>")
    return f"<table><thead><tr>{header}</tr></thead><tbody>{''.join(body_rows)}</tbody></table>"


def display_monthly(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "month": row["month"],
            "clicks": fmt_num(row.get("clicks")),
            "impressions": fmt_num(row.get("impressions")),
            "ctr": fmt_pct(row.get("ctr")),
            "avg_position": fmt_pos(row.get("avg_position")),
        }
        for row in rows
    ]


def display_rankings(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "rank": row["rank"],
            "property_name": row["property_name"],
            "property_code": row["property_code"] or "",
            "clicks": fmt_num(row.get("clicks")),
            "impressions": fmt_num(row.get("impressions")),
            "ctr": fmt_pct(row.get("ctr")),
            "avg_position": fmt_pos(row.get("avg_position")),
        }
        for row in rows
    ]


def display_pilot_post(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "pilot_property": row["pilot_property"],
            "control_property": row["control_property"],
            "pilot_clicks": fmt_num(row["pilot_clicks"]),
            "control_clicks": fmt_num(row["control_clicks"]),
            "click_gap_pct": fmt_pct(row["click_gap_pct"], signed=True),
            "pilot_impressions": fmt_num(row["pilot_impressions"]),
            "pilot_ctr": fmt_pct(row["pilot_ctr"]),
            "pilot_avg_position": fmt_pos(row["pilot_avg_position"]),
        }
        for row in rows
    ]


def display_pilot_t30(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "pilot_property": row["pilot_property"],
            "clicks_t30": fmt_num(row["clicks_t30"]),
            "click_change_pct": fmt_pct(row["click_change_pct"], signed=True),
            "impressions_t30": fmt_num(row["impressions_t30"]),
            "impression_change_pct": fmt_pct(row["impression_change_pct"], signed=True),
            "ctr_t30": fmt_pct(row["ctr_t30"]),
            "avg_position_t30": fmt_pos(row["avg_position_t30"]),
        }
        for row in rows
    ]


def build_html(
    path: Path,
    summary: dict[str, Any],
    monthly: list[dict[str, Any]],
    rankings: list[dict[str, Any]],
    pilot_post: list[dict[str, Any]],
    pilot_t30: list[dict[str, Any]],
) -> None:
    css = f"""
    body {{ margin:0; font-family: Arial, sans-serif; color:#{BRAND['navy']}; background:#{BRAND['white_smoke']}; }}
    main {{ max-width:1180px; margin:0 auto; padding:34px 24px 48px; }}
    h1 {{ margin:0; font-size:32px; line-height:1.1; color:#{BRAND['navy']}; }}
    h2 {{ margin:30px 0 10px; font-size:20px; color:#{BRAND['bay']}; }}
    p {{ line-height:1.5; color:#{BRAND['black']}; }}
    .sub {{ margin-top:8px; color:#{BRAND['bay']}; }}
    .kpis {{ display:grid; grid-template-columns:repeat(4, minmax(0,1fr)); gap:12px; margin-top:24px; }}
    .kpi {{ background:#{BRAND['white']}; border:1px solid #{BRAND['quill']}; border-top:4px solid #{BRAND['blue_chill']}; padding:14px; }}
    .label {{ font-size:12px; text-transform:uppercase; letter-spacing:.04em; color:#{BRAND['bay']}; font-weight:700; }}
    .value {{ margin-top:8px; font-size:28px; font-weight:800; color:#{BRAND['navy']}; }}
    .note {{ background:#{BRAND['white']}; border-left:5px solid #{BRAND['terra_cotta']}; padding:14px 16px; }}
    table {{ width:100%; border-collapse:collapse; background:#{BRAND['white']}; margin-top:12px; font-size:13px; }}
    th {{ background:#{BRAND['navy']}; color:#{BRAND['white']}; text-align:left; padding:9px; }}
    td {{ border:1px solid #{BRAND['quill']}; padding:8px; color:#{BRAND['black']}; }}
    .footer {{ margin-top:28px; font-size:12px; color:#{BRAND['delta']}; }}
    """
    content = f"""<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <title>Organic Search Performance Portfolio And Pilot Report</title>
  <style>{css}</style>
</head>
<body>
<main>
  <h1>Organic Search Performance Portfolio And Pilot Report</h1>
  <div class="sub">Generated {html.escape(summary['generated_at'])}; source: Data Pond Google Search Console organic search performance metrics.</div>
  <div class="kpis">
    <div class="kpi"><div class="label">Portfolio GSC Clicks</div><div class="value">{html.escape(summary['portfolio_clicks'])}</div></div>
    <div class="kpi"><div class="label">Portfolio Impressions</div><div class="value">{html.escape(summary['portfolio_impressions'])}</div></div>
    <div class="kpi"><div class="label">Portfolio CTR</div><div class="value">{html.escape(summary['portfolio_ctr'])}</div></div>
    <div class="kpi"><div class="label">Avg Position</div><div class="value">{html.escape(summary['portfolio_avg_position'])}</div></div>
  </div>
  <h2>Executive Read</h2>
  <p>{html.escape(summary['executive_read'])}</p>
  <div class="note">{html.escape(summary['coverage_note'])}</div>
  <h2>Latest 30-Day Movement</h2>
  <p>{html.escape(summary['t30_note'])}</p>
  <h2>Five New Pilot Sites: Post-Launch Organic Search Comparison</h2>
  {html_table(pilot_post, [
      ('pilot_property', 'Pilot'),
      ('control_property', 'Control'),
      ('pilot_clicks', 'Pilot clicks'),
      ('control_clicks', 'Control clicks'),
      ('click_gap_pct', 'Click gap'),
      ('pilot_impressions', 'Pilot impressions'),
      ('pilot_ctr', 'Pilot CTR'),
      ('pilot_avg_position', 'Pilot avg position')
  ])}
  <h2>Pilot Latest T30</h2>
  {html_table(pilot_t30, [
      ('pilot_property', 'Pilot'),
      ('clicks_t30', 'Clicks T30'),
      ('click_change_pct', 'Clicks vs prior 30'),
      ('impressions_t30', 'Impressions T30'),
      ('impression_change_pct', 'Impressions vs prior 30'),
      ('ctr_t30', 'CTR'),
      ('avg_position_t30', 'Avg position')
  ])}
  <h2>Monthly Portfolio Trend</h2>
  {html_table(monthly, [
      ('month', 'Month'),
      ('clicks', 'Clicks'),
      ('impressions', 'Impressions'),
      ('ctr', 'CTR'),
      ('avg_position', 'Avg position')
  ])}
  <h2>Top Portfolio Properties By GSC Clicks</h2>
  {html_table(rankings, [
      ('rank', 'Rank'),
      ('property_name', 'Property'),
      ('property_code', 'Property code'),
      ('clicks', 'Clicks'),
      ('impressions', 'Impressions'),
      ('ctr', 'CTR'),
      ('avg_position', 'Avg position')
  ], limit=20)}
  <div class="footer">Method: Google Search Console daily organic search performance rows joined by governed GA4 identity. Pilot launch/post-launch window starts {PILOT_LAUNCH_DATE.isoformat()}.</div>
</main>
</body>
</html>"""
    path.write_text(content, encoding="utf-8")


def main() -> int:
    generated_at = datetime.now().astimezone()
    run_dir = OUT_ROOT / generated_at.strftime("%Y%m%d_%H%M%S")
    run_dir.mkdir(parents=True, exist_ok=True)

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    properties = load_portfolio_properties()
    property_ids = [item.property_id for item in properties]
    pairs = load_pilot_pairs()

    bounds = fetch_one(
        conn,
        f"""
        SELECT MIN(metric_date) AS min_date, MAX(metric_date) AS max_date
        FROM gsc_daily_metrics
        WHERE COALESCE(ga4_property_id, property_id) IN ({placeholders(property_ids)})
        """,
        tuple(property_ids),
    )
    start = max(REQUESTED_START, parse_date(bounds["min_date"]))
    end = parse_date(bounds["max_date"])

    portfolio = fetch_gsc_summary(conn, property_ids, start, end)
    current_start = end - timedelta(days=29)
    prior_end = current_start - timedelta(days=1)
    prior_start = prior_end - timedelta(days=29)
    current = fetch_gsc_summary(conn, property_ids, current_start, end)
    prior = fetch_gsc_summary(conn, property_ids, prior_start, prior_end)

    monthly = fetch_monthly(conn, property_ids, start, end)
    rankings = fetch_property_rankings(conn, properties, start, end)
    pilot_post = fetch_pilot_post_launch(conn, pairs, PILOT_LAUNCH_DATE, end)
    pilot_t30 = fetch_pilot_t30(conn, pairs, end)

    pilot_ids = [pair.pilot_id for pair in pairs]
    control_ids = [pair.control_id for pair in pairs]
    pilot_rollup = fetch_gsc_summary(conn, pilot_ids, PILOT_LAUNCH_DATE, end)
    control_rollup = fetch_gsc_summary(conn, control_ids, PILOT_LAUNCH_DATE, end)

    click_change = pct_change(current.get("clicks") or 0, prior.get("clicks") or 0)
    impression_change = pct_change(current.get("impressions") or 0, prior.get("impressions") or 0)
    pilot_gap = pct_change(pilot_rollup.get("clicks") or 0, control_rollup.get("clicks") or 0)

    summary = {
        "generated_at": generated_at.strftime("%Y-%m-%d %H:%M %Z"),
        "requested_window": f"{REQUESTED_START.isoformat()} through {end.isoformat()}",
        "effective_gsc_window": f"{start.isoformat()} through {end.isoformat()}",
        "portfolio_property_count": len(property_ids),
        "portfolio_clicks": fmt_num(portfolio.get("clicks")),
        "portfolio_impressions": fmt_num(portfolio.get("impressions")),
        "portfolio_ctr": fmt_pct(portfolio.get("ctr")),
        "portfolio_avg_position": fmt_pos(portfolio.get("avg_position")),
        "coverage_note": (
            f"Requested last-two-years window starts {REQUESTED_START.isoformat()}; canonical local GSC coverage begins "
            f"{start.isoformat()}. This report does not infer missing {REQUESTED_START.isoformat()} to "
            f"{(start - timedelta(days=1)).isoformat()} Search Console organic search performance."
        ),
        "executive_read": (
            f"Across {len(property_ids)} governed portfolio properties, available GSC data produced "
            f"{fmt_num(portfolio.get('clicks'))} clicks and {fmt_num(portfolio.get('impressions'))} impressions "
            f"from {start.isoformat()} to {end.isoformat()}, with {fmt_pct(portfolio.get('ctr'))} CTR and "
            f"{fmt_pos(portfolio.get('avg_position'))} weighted average position."
        ),
        "t30_note": (
            f"Latest T30 GSC clicks were {fmt_num(current.get('clicks'))} ({fmt_pct(click_change, signed=True)} vs prior 30); "
            f"impressions were {fmt_num(current.get('impressions'))} ({fmt_pct(impression_change, signed=True)}). "
            f"CTR was {fmt_pct(current.get('ctr'))} and weighted average position was {fmt_pos(current.get('avg_position'))}."
        ),
        "pilot_summary": (
            f"Post-launch, the five pilot sites produced {fmt_num(pilot_rollup.get('clicks'))} GSC clicks versus "
            f"{fmt_num(control_rollup.get('clicks'))} for the five controls ({fmt_pct(pilot_gap, signed=True)})."
        ),
    }

    write_csv(run_dir / "monthly_portfolio_gsc.csv", monthly)
    write_csv(run_dir / "property_gsc_rankings.csv", rankings)
    write_csv(run_dir / "pilot_post_launch_gsc_comparison.csv", pilot_post)
    write_csv(run_dir / "pilot_latest_t30_gsc.csv", pilot_t30)

    build_workbook(
        run_dir / "organic_search_performance_portfolio_pilot_report.xlsx",
        {
            "Monthly Portfolio": monthly,
            "Property Rankings": rankings,
            "Pilot Post Launch": pilot_post,
            "Pilot Latest T30": pilot_t30,
        },
        [
            ["Generated", summary["generated_at"]],
            ["Requested Window", summary["requested_window"]],
            ["Effective GSC Window", summary["effective_gsc_window"]],
            ["Portfolio Properties", summary["portfolio_property_count"]],
            ["Method", "Google Search Console organic search performance metrics resolved through governed GA4 identity."],
            ["Coverage Note", summary["coverage_note"]],
            ["Pilot Summary", summary["pilot_summary"]],
        ],
    )

    build_html(
        run_dir / "organic_search_performance_portfolio_pilot_report.html",
        summary,
        display_monthly(monthly),
        display_rankings(rankings),
        display_pilot_post(pilot_post),
        display_pilot_t30(pilot_t30),
    )

    summary_path = run_dir / "organic_search_performance_portfolio_pilot_summary.json"
    summary_path.write_text(
        json.dumps(
            {
                "summary": summary,
                "artifacts": {
                    "html": str(run_dir / "organic_search_performance_portfolio_pilot_report.html"),
                    "workbook": str(run_dir / "organic_search_performance_portfolio_pilot_report.xlsx"),
                    "monthly_portfolio_csv": str(run_dir / "monthly_portfolio_gsc.csv"),
                    "property_rankings_csv": str(run_dir / "property_gsc_rankings.csv"),
                    "pilot_post_launch_comparison_csv": str(run_dir / "pilot_post_launch_gsc_comparison.csv"),
                    "pilot_latest_t30_csv": str(run_dir / "pilot_latest_t30_gsc.csv"),
                },
            },
            indent=2,
        ),
        encoding="utf-8",
    )
    print(json.dumps({"run_dir": str(run_dir), "summary": summary}, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
