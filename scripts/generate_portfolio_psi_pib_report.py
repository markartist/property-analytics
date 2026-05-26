#!/usr/bin/env python3
"""
Generate and optionally email a PIB-style portfolio PSI / CWV report.

This report is portfolio-scoped and intentionally does not modify canonical PIB files.
"""

from __future__ import annotations

import argparse
import sqlite3
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional

ROOT = Path("/Users/mark/Property_Analytics")
DB_PATH = ROOT / "data" / "portfolio_analytics.db"
OUT_DIR = ROOT / "reports" / "portfolio_psi_pib"
OUT_DIR.mkdir(parents=True, exist_ok=True)
DELIVERY_DIR = ROOT / "logs" / "email_delivery"

sys.path.insert(0, str(ROOT / "utils"))
from email_sender import EmailSender  # noqa: E402
from report_builder import ReportBuilder, Section  # noqa: E402
from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

SHOW_BOTTOM_10_SECTION = False


@dataclass
class Trend:
    current: Optional[float]
    t7_delta: Optional[float]
    t30_delta: Optional[float]
    series: List[Optional[float]]


def fmt_num(value: Optional[float], decimals: int = 1, suffix: str = "") -> str:
    if value is None:
        return "—"
    return f"{value:.{decimals}f}{suffix}"


def fmt_score(value: Optional[float]) -> str:
    if value is None:
        return "—"
    return f"{int(round(value))}"


def health_label(score: Optional[float]) -> str:
    if score is None:
        return "NO DATA"
    if score >= 90:
        return "GOOD"
    if score >= 50:
        return "NEEDS IMPROVEMENT"
    return "POOR"


def metric_color(metric: str, value: Optional[float]) -> str:
    if value is None:
        return "#94a3b8"
    if metric in {"mobile_score", "desktop_score"}:
        if value >= 90:
            return "#15803d"
        if value >= 50:
            return "#b45309"
        return "#b91c1c"
    if metric == "lcp":
        if value <= 2.5:
            return "#15803d"
        if value <= 4.0:
            return "#b45309"
        return "#b91c1c"
    if metric == "cls":
        if value <= 0.1:
            return "#15803d"
        if value <= 0.25:
            return "#b45309"
        return "#b91c1c"
    if metric == "fid":
        if value <= 100:
            return "#15803d"
        if value <= 300:
            return "#b45309"
        return "#b91c1c"
    if metric == "fcp":
        if value <= 1.8:
            return "#15803d"
        if value <= 3.0:
            return "#b45309"
        return "#b91c1c"
    return "#0f172a"


def trend_color(metric: str, delta: Optional[float]) -> str:
    if delta is None or delta == 0:
        return "#64748b"
    lower_is_better = metric in {"lcp", "cls", "fid", "fcp"}
    improved = delta < 0 if lower_is_better else delta > 0
    return "#15803d" if improved else "#b91c1c"


def trend_line(metric: str, t7: Optional[float], t30: Optional[float], decimals: int = 1, suffix: str = "") -> str:
    def part(label: str, value: Optional[float]) -> str:
        if value is None:
            return f"{label} —"
        arrow = "↓" if value < 0 else "↑" if value > 0 else "→"
        color = trend_color(metric, value)
        return f'<span style="color:{color};">{label} {arrow}{abs(value):.{decimals}f}{suffix}</span>'

    return f"{part('T7', t7)} • {part('T30', t30)}"


def make_sparkline(series: List[Optional[float]], stroke: str = "#0f766e") -> str:
    values = [v for v in series if v is not None]
    if not values:
        return ""
    width, height, pad = 220, 64, 6
    min_v = min(values)
    max_v = max(values)
    span = max(max_v - min_v, 1e-9)
    step = (width - 2 * pad) / max(len(series) - 1, 1)

    points: List[str] = []
    for idx, value in enumerate(series):
        if value is None:
            continue
        x = pad + idx * step
        y = height - pad - ((value - min_v) / span) * (height - 2 * pad)
        points.append(f"{x:.1f},{y:.1f}")

    return f"""
    <svg width="100%" viewBox="0 0 {width} {height}" role="img" aria-label="metric trend">
      <line x1="{pad}" y1="{height-pad}" x2="{width-pad}" y2="{height-pad}" stroke="#cbd5e1" stroke-width="1" />
      <polyline points="{' '.join(points)}" fill="none" stroke="{stroke}" stroke-width="2.5" />
    </svg>
    """


def recent_average(rows: Iterable[sqlite3.Row], field: str, days: int, current_date: str) -> Optional[float]:
    vals = [row[field] for row in rows if row["metric_date"] < current_date and row[field] is not None][:days]
    if not vals:
        return None
    return sum(vals) / len(vals)


def query_portfolio_daily(conn: sqlite3.Connection) -> List[sqlite3.Row]:
    return conn.execute(
        """
        WITH daily AS (
            SELECT
                metric_date,
                ROUND(AVG(CASE WHEN strategy='mobile' THEN performance_score END), 2) AS mobile_score,
                ROUND(AVG(CASE WHEN strategy='desktop' THEN performance_score END), 2) AS desktop_score,
                ROUND(AVG(CASE WHEN strategy='mobile' THEN lcp_value END), 3) AS lcp,
                ROUND(AVG(CASE WHEN strategy='mobile' THEN cls_value END), 4) AS cls,
                ROUND(AVG(CASE WHEN strategy='mobile' THEN fid_value END), 2) AS fid,
                ROUND(AVG(CASE WHEN strategy='mobile' THEN fcp_value END), 3) AS fcp
            FROM pagespeed_metrics
            GROUP BY metric_date
        )
        SELECT *
        FROM daily
        ORDER BY metric_date DESC
        LIMIT 31
        """
    ).fetchall()


def build_portfolio_trend(rows: List[sqlite3.Row], field: str) -> Trend:
    ordered = sorted(rows, key=lambda row: row["metric_date"], reverse=True)
    current_date = ordered[0]["metric_date"]
    current = ordered[0][field]
    t7_avg = recent_average(ordered, field, 7, current_date)
    t30_avg = recent_average(ordered, field, 30, current_date)
    series = [row[field] for row in reversed(ordered[:30])]
    return Trend(
        current=current,
        t7_delta=(current - t7_avg) if current is not None and t7_avg is not None else None,
        t30_delta=(current - t30_avg) if current is not None and t30_avg is not None else None,
        series=series,
    )


def query_properties(conn: sqlite3.Connection) -> List[Dict[str, object]]:
    latest_rows = conn.execute(
        """
        WITH latest_mobile AS (
            SELECT property_id, MAX(metric_date) AS latest_mobile_date
            FROM pagespeed_metrics
            WHERE strategy = 'mobile'
            GROUP BY property_id
        )
        SELECT
            lm.property_id,
            COALESCE(p.property_name, lm.property_id) AS property_name,
            lm.latest_mobile_date,
            m.performance_score AS mobile_score,
            d.performance_score AS desktop_score,
            m.lcp_value AS lcp,
            m.cls_value AS cls,
            m.fid_value AS fid,
            m.fcp_value AS fcp
        FROM latest_mobile lm
        JOIN pagespeed_metrics m
            ON m.property_id = lm.property_id
           AND m.metric_date = lm.latest_mobile_date
           AND m.strategy = 'mobile'
        LEFT JOIN pagespeed_metrics d
            ON d.property_id = lm.property_id
           AND d.metric_date = lm.latest_mobile_date
           AND d.strategy = 'desktop'
        LEFT JOIN properties p
            ON p.property_id = lm.property_id
        ORDER BY m.performance_score ASC, COALESCE(p.property_name, lm.property_id) ASC
        """
    ).fetchall()

    result: List[Dict[str, object]] = []
    for row in latest_rows:
        history = conn.execute(
            """
            SELECT
                metric_date,
                MAX(CASE WHEN strategy='mobile' THEN performance_score END) AS mobile_score,
                MAX(CASE WHEN strategy='desktop' THEN performance_score END) AS desktop_score,
                MAX(CASE WHEN strategy='mobile' THEN lcp_value END) AS lcp,
                MAX(CASE WHEN strategy='mobile' THEN cls_value END) AS cls,
                MAX(CASE WHEN strategy='mobile' THEN fid_value END) AS fid,
                MAX(CASE WHEN strategy='mobile' THEN fcp_value END) AS fcp
            FROM pagespeed_metrics
            WHERE property_id = ?
            GROUP BY metric_date
            ORDER BY metric_date DESC
            LIMIT 31
            """,
            (row["property_id"],),
        ).fetchall()
        current_date = row["latest_mobile_date"]

        entry: Dict[str, object] = dict(row)
        for field in ("mobile_score", "desktop_score", "lcp", "cls", "fid", "fcp"):
            current = row[field]
            t7_avg = recent_average(history, field, 7, current_date)
            t30_avg = recent_average(history, field, 30, current_date)
            entry[f"{field}_t7"] = (current - t7_avg) if current is not None and t7_avg is not None else None
            entry[f"{field}_t30"] = (current - t30_avg) if current is not None and t30_avg is not None else None

        result.append(entry)
    return result


def build_stats(portfolio_daily: List[sqlite3.Row], properties: List[Dict[str, object]]) -> Dict[str, object]:
    latest_date = portfolio_daily[0]["metric_date"]
    cards = {
        "mobile_score": build_portfolio_trend(portfolio_daily, "mobile_score"),
        "desktop_score": build_portfolio_trend(portfolio_daily, "desktop_score"),
        "lcp": build_portfolio_trend(portfolio_daily, "lcp"),
        "cls": build_portfolio_trend(portfolio_daily, "cls"),
        "fid": build_portfolio_trend(portfolio_daily, "fid"),
        "fcp": build_portfolio_trend(portfolio_daily, "fcp"),
    }
    mobile_values = [p["mobile_score"] for p in properties if p["mobile_score"] is not None]
    poor = sum(1 for score in mobile_values if score < 50)
    needs = sum(1 for score in mobile_values if 50 <= score < 90)
    good = sum(1 for score in mobile_values if score >= 90)
    return {
        "latest_date": latest_date,
        "cards": cards,
        "total_properties": len(properties),
        "min_mobile_score": min(mobile_values) if mobile_values else None,
        "max_mobile_score": max(mobile_values) if mobile_values else None,
        "poor_count": poor,
        "needs_count": needs,
        "good_count": good,
    }


def _format_delta(delta: Optional[float], invert: bool = False, decimals: int = 1, suffix: str = "", show_sign: bool = False) -> str:
    if delta is None:
        return ""
    if delta == 0:
        arrow = "→"
        color = "#6c757d"
    else:
        improved = (delta < 0) if invert else (delta > 0)
        color = "#28a745" if improved else "#dc3545"
        arrow = "↓" if delta < 0 else "↑"
    value = f"{delta:+.{decimals}f}" if show_sign else f"{abs(delta):.{decimals}f}"
    return f'<span style="font-size:14px;color:{color};margin-left:8px;">{arrow}{value}{suffix}</span>'


def render_overview(stats: Dict[str, object]) -> str:
    cards = stats["cards"]
    avg_mobile = cards["mobile_score"].current
    label = health_label(avg_mobile)
    label_color = metric_color("mobile_score", avg_mobile)
    avg_lcp = cards["lcp"].current
    avg_cls = cards["cls"].current
    avg_fid = cards["fid"].current
    avg_fcp = cards["fcp"].current

    def t30_only(metric: str, delta: Optional[float], decimals: int, suffix: str) -> str:
        if delta is None:
            return "T30 —"
        arrow = "↓" if delta < 0 else "↑" if delta > 0 else "→"
        color = trend_color(metric, delta)
        return f'T30 <span style="color:{color};">{arrow}{abs(delta):.{decimals}f}{suffix}</span>'

    def t7_only(metric: str, delta: Optional[float], decimals: int, suffix: str) -> str:
        if delta is None:
            return "T7 —"
        arrow = "↓" if delta < 0 else "↑" if delta > 0 else "→"
        color = trend_color(metric, delta)
        return f'T7 <span style="color:{color};">{arrow}{abs(delta):.{decimals}f}{suffix}</span>'

    return f"""
    <div style="margin-bottom:30px;">
      <div style="text-align:center;padding:20px;background:#f8f9fa;border-radius:6px;margin-bottom:20px;">
        <div style="font-size:16px;font-weight:600;color:{label_color};margin-bottom:10px;">
          Portfolio Average: {fmt_num(avg_mobile, 1)} ({label})
          {_format_delta(cards['mobile_score'].t7_delta, invert=False, decimals=1)}
        </div>
        <div style="font-size:13px;color:#6c757d;">
          {stats['total_properties']} properties analyzed | Latest data: {stats['latest_date']} | T7 and T30 trends embedded below
        </div>
      </div>

      <table cellpadding="0" cellspacing="0" border="0" style="width:100%;margin-bottom:25px;">
        <tr>
          <td style="width:32%;vertical-align:top;">
            <div style="padding:20px;background:white;border:1px solid #e9ecef;border-radius:6px;text-align:center;">
              <div style="font-size:11px;color:#868e96;text-transform:uppercase;letter-spacing:0.5px;margin-bottom:10px;font-weight:600;">Avg Mobile Score</div>
              <div style="font-size:28px;font-weight:700;color:{metric_color('mobile_score', avg_mobile)};margin:8px 0;">
                {fmt_num(avg_mobile, 1)}
              </div>
              <div style="font-size:12px;color:#6c757d;margin-top:5px;">{t7_only('mobile_score', cards['mobile_score'].t7_delta, 1, '')}</div>
              <div style="font-size:12px;color:#6c757d;margin-top:2px;">{t30_only('mobile_score', cards['mobile_score'].t30_delta, 1, '')}</div>
            </div>
          </td>
          <td style="width:2%;"></td>
          <td style="width:32%;vertical-align:top;">
            <div style="padding:20px;background:white;border:1px solid #e9ecef;border-radius:6px;text-align:center;">
              <div style="font-size:11px;color:#868e96;text-transform:uppercase;letter-spacing:0.5px;margin-bottom:10px;font-weight:600;">Avg Desktop Score</div>
              <div style="font-size:28px;font-weight:700;color:{metric_color('desktop_score', cards['desktop_score'].current)};margin:8px 0;">
                {fmt_score(cards['desktop_score'].current)}
              </div>
              <div style="font-size:12px;color:#6c757d;margin-top:5px;">{t7_only('desktop_score', cards['desktop_score'].t7_delta, 1, '')}</div>
              <div style="font-size:12px;color:#6c757d;margin-top:2px;">{t30_only('desktop_score', cards['desktop_score'].t30_delta, 1, '')}</div>
            </div>
          </td>
          <td style="width:2%;"></td>
          <td style="width:32%;vertical-align:top;">
            <div style="padding:20px;background:white;border:1px solid #e9ecef;border-radius:6px;text-align:center;">
              <div style="font-size:11px;color:#868e96;text-transform:uppercase;letter-spacing:0.5px;margin-bottom:10px;font-weight:600;">Avg FCP</div>
              <div style="font-size:28px;font-weight:700;color:{metric_color('fcp', avg_fcp)};margin:8px 0;">
                {fmt_num(avg_fcp, 2, 's')}
              </div>
              <div style="font-size:12px;color:#6c757d;margin-top:5px;">{t7_only('fcp', cards['fcp'].t7_delta, 2, 's')}</div>
              <div style="font-size:12px;color:#6c757d;margin-top:2px;">{t30_only('fcp', cards['fcp'].t30_delta, 2, 's')}</div>
            </div>
          </td>
        </tr>
      </table>

      <table cellpadding="0" cellspacing="0" border="0" style="width:100%;margin-bottom:25px;">
        <tr>
          <td style="width:32%;vertical-align:top;">
            <div style="padding:20px;background:white;border:1px solid #e9ecef;border-radius:6px;text-align:center;">
              <div style="font-size:11px;color:#868e96;text-transform:uppercase;letter-spacing:0.5px;margin-bottom:10px;font-weight:600;">Avg LCP</div>
              <div style="font-size:28px;font-weight:700;color:{metric_color('lcp', avg_lcp)};margin:8px 0;">
                {fmt_num(avg_lcp, 2, 's')}
              </div>
              <div style="font-size:12px;color:#6c757d;margin-top:5px;">{t7_only('lcp', cards['lcp'].t7_delta, 2, 's')}</div>
              <div style="font-size:12px;color:#6c757d;margin-top:2px;">{t30_only('lcp', cards['lcp'].t30_delta, 2, 's')}</div>
            </div>
          </td>
          <td style="width:2%;"></td>
          <td style="width:32%;vertical-align:top;">
            <div style="padding:20px;background:white;border:1px solid #e9ecef;border-radius:6px;text-align:center;">
              <div style="font-size:11px;color:#868e96;text-transform:uppercase;letter-spacing:0.5px;margin-bottom:10px;font-weight:600;">Avg FID</div>
              <div style="font-size:28px;font-weight:700;color:{metric_color('fid', avg_fid)};margin:8px 0;">
                {fmt_num(avg_fid, 0, 'ms')}
              </div>
              <div style="font-size:12px;color:#6c757d;margin-top:5px;">{t7_only('fid', cards['fid'].t7_delta, 0, 'ms')}</div>
              <div style="font-size:12px;color:#6c757d;margin-top:2px;">{t30_only('fid', cards['fid'].t30_delta, 0, 'ms')}</div>
            </div>
          </td>
          <td style="width:2%;"></td>
          <td style="width:32%;vertical-align:top;">
            <div style="padding:20px;background:white;border:1px solid #e9ecef;border-radius:6px;text-align:center;">
              <div style="font-size:11px;color:#868e96;text-transform:uppercase;letter-spacing:0.5px;margin-bottom:10px;font-weight:600;">Avg CLS</div>
              <div style="font-size:28px;font-weight:700;color:{metric_color('cls', avg_cls)};margin:8px 0;">
                {fmt_num(avg_cls, 3)}
              </div>
              <div style="font-size:12px;color:#6c757d;margin-top:5px;">{t7_only('cls', cards['cls'].t7_delta, 3, '')}</div>
              <div style="font-size:12px;color:#6c757d;margin-top:2px;">{t30_only('cls', cards['cls'].t30_delta, 3, '')}</div>
            </div>
          </td>
        </tr>
      </table>

      <div style="background:white;border:1px solid #e9ecef;border-radius:6px;padding:20px;">
        <div style="font-size:14px;font-weight:600;margin-bottom:15px;">Score Distribution</div>
        <table style="width:100%;">
          <tr>
            <td style="width:33%;text-align:center;padding:15px;background:#f8d7da;border-radius:4px;">
              <div style="font-size:32px;font-weight:700;color:#dc3545;">{stats['poor_count']}</div>
              <div style="font-size:12px;color:#721c24;margin-top:5px;font-weight:600;">POOR (&lt;50)</div>
            </td>
            <td style="width:2%;"></td>
            <td style="width:33%;text-align:center;padding:15px;background:#fff3cd;border-radius:4px;">
              <div style="font-size:32px;font-weight:700;color:#fd7e14;">{stats['needs_count']}</div>
              <div style="font-size:12px;color:#856404;margin-top:5px;font-weight:600;">NEEDS IMPROVEMENT (50-89)</div>
            </td>
            <td style="width:2%;"></td>
            <td style="width:33%;text-align:center;padding:15px;background:#d4edda;border-radius:4px;">
              <div style="font-size:32px;font-weight:700;color:#28a745;">{stats['good_count']}</div>
              <div style="font-size:12px;color:#155724;margin-top:5px;font-weight:600;">GOOD (90+)</div>
            </td>
          </tr>
        </table>
      </div>
    </div>
    """


def render_property_block(prop: Dict[str, object]) -> str:
    border = "#28a745" if (prop["mobile_score"] or 0) >= 90 else "#dc3545" if (prop["mobile_score"] or 0) < 50 else "#fd7e14"
    score_color = metric_color("mobile_score", prop["mobile_score"])

    def metric_html(label: str, key: str, value: Optional[float], t7: Optional[float], t30: Optional[float], decimals: int, suffix: str) -> str:
        display = fmt_score(value) if key in {"mobile_score", "desktop_score"} else fmt_num(value, decimals, suffix)
        return (
            f'<td style="padding:2px 14px 2px 0;vertical-align:top;">'
            f'<div style="font-size:11px;color:#333;"><strong>{label}:</strong> '
            f'<span style="color:{metric_color(key, value)};">{display}</span></div>'
            f'<div style="font-size:10px;color:#6c757d;">{trend_line(key, t7, t30, decimals, suffix)}</div>'
            f'</td>'
        )

    return f"""
    <div style="padding:12px;background:#fff;border-left:4px solid {border};margin-bottom:12px;border-radius:4px;">
      <div style="display:flex;justify-content:space-between;align-items:flex-start;">
        <div style="flex:1;">
          <div style="font-weight:600;color:#333;font-size:14px;margin-bottom:8px;">
            {prop['property_name']}
          </div>
          <div style="font-size:11px;color:#6c757d;margin:0 0 8px 0;">
            {trend_line('mobile_score', prop.get('mobile_score_t7'), prop.get('mobile_score_t30'), 1, '')}
          </div>
          <table style="width:100%;font-size:11px;">
            <tr>
              {metric_html('Mobile', 'mobile_score', prop['mobile_score'], prop['mobile_score_t7'], prop['mobile_score_t30'], 0, '')}
              {metric_html('Desktop', 'desktop_score', prop['desktop_score'], prop['desktop_score_t7'], prop['desktop_score_t30'], 0, '')}
            </tr>
            <tr>
              {metric_html('LCP', 'lcp', prop['lcp'], prop['lcp_t7'], prop['lcp_t30'], 2, 's')}
              {metric_html('CLS', 'cls', prop['cls'], prop['cls_t7'], prop['cls_t30'], 3, '')}
              {metric_html('FID', 'fid', prop['fid'], prop['fid_t7'], prop['fid_t30'], 0, 'ms')}
              {metric_html('FCP', 'fcp', prop['fcp'], prop['fcp_t7'], prop['fcp_t30'], 2, 's')}
            </tr>
          </table>
        </div>
        <div style="font-size:32px;font-weight:700;color:{score_color};margin-left:15px;">
          {fmt_score(prop['mobile_score'])}
        </div>
      </div>
    </div>
    """


def render_property_section(properties: List[Dict[str, object]]) -> str:
    bottom_ten = properties[:10]
    all_rows = ''.join(render_property_block(prop) for prop in properties)
    sections = ['<div style="margin-bottom:30px;">']
    # Keep the alert-style Bottom 10 section disabled for the standard report.
    # Only re-enable this if the user explicitly requests it for a future send.
    if SHOW_BOTTOM_10_SECTION:
        bottom_rows = ''.join(render_property_block(prop) for prop in bottom_ten)
        sections.extend(
            [
                '<div style="margin-bottom:25px;">',
                '<h3 style="font-size:15px;color:#dc3545;margin:0 0 15px 0;font-weight:600;">',
                '⚠️ Bottom 10 Performers (Immediate Attention Needed)',
                '</h3>',
                bottom_rows,
                '</div>',
            ]
        )
    sections.extend(
        [
            '<div style="margin-bottom:25px;">',
            '<h3 style="font-size:15px;color:#495057;margin:0 0 15px 0;font-weight:600;">',
            'Full Portfolio Listing',
            '</h3>',
            all_rows,
            '</div>',
            '<div style="margin-top:14px;padding:12px 14px;background:#f8fafc;border:1px solid #dbe4ef;border-radius:10px;'
            'font-size:12px;color:#475569;line-height:1.6;">',
            '<strong>Interpretation notes:</strong> This report uses the local master database only. '
            'Metrics shown per property are mobile score, desktop score, and mobile PSI lab metrics for '
            '<strong>LCP, CLS, FID, and FCP</strong>. Trend lines compare current values to trailing 7-day and 30-day averages.',
            '</div></div>',
        ]
    )
    return ''.join(sections)


def build_html(portfolio_daily: List[sqlite3.Row], properties: List[Dict[str, object]]) -> str:
    stats = build_stats(portfolio_daily, properties)
    builder = ReportBuilder(
        title="Portfolio CWV Report",
        subtitle="Portfolio State & Performance Overview",
        version="2.0",
        date_range=stats["latest_date"],
    )
    builder.add_section(
        Section(
            title="Portfolio Overview",
            content=render_overview(stats),
            status="action_needed" if (stats["cards"]["mobile_score"].current or 0) < 75 else "healthy",
            description=f"Current state of {stats['total_properties']} properties",
        )
    )
    builder.add_section(
        Section(
            title="Individual Property Performance",
            content=render_property_section(properties),
            status="action_needed",
            description="All properties listed with mobile/desktop scores and LCP, CLS, FID, FCP trend lines",
        )
    )
    html = builder.generate()
    tagline = (
        '<div style="text-align:center;margin:-8px 0 18px 0;">'
        '<span style="font-size:9px;color:#adb5bd;vertical-align:middle;">Powered by</span> '
        '<span style="font-size:12px;color:#495057;font-weight:600;vertical-align:middle;">MarketingOps</span>'
        '</div>'
    )
    subtitle_html = '<div style="font-size: 16px; color: #0066cc; margin: 10px 0 15px 0; font-weight: 600; letter-spacing: 0.3px; text-transform: uppercase;">Portfolio State & Performance Overview</div>'
    html = html.replace(subtitle_html, subtitle_html + tagline, 1)
    return html.replace(
        '<div style="font-size: 11px; color: #adb5bd; margin: 10px 0;">v2.0</div>',
        '<div style="font-size: 8px; color: #adb5bd; margin: 6px 0;">v2.0</div>',
        1,
    )


def autosize_worksheet(ws) -> None:
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 26)


def write_excel(output_path: Path, portfolio_daily: List[sqlite3.Row], properties: List[Dict[str, object]]) -> None:
    stats = build_stats(portfolio_daily, properties)
    wb = Workbook()

    summary_ws = wb.active
    summary_ws.title = "Summary"
    summary_ws["A1"] = "Portfolio CWV Report"
    summary_ws["A1"].font = Font(bold=True, size=14)
    summary_ws["A3"] = "Latest portfolio date"
    summary_ws["B3"] = stats["latest_date"]
    summary_ws["A4"] = "Properties in report"
    summary_ws["B4"] = stats["total_properties"]
    summary_ws["A5"] = "Poor (<50)"
    summary_ws["B5"] = stats["poor_count"]
    summary_ws["A6"] = "Needs Improvement (50-89)"
    summary_ws["B6"] = stats["needs_count"]
    summary_ws["A7"] = "Good (90+)"
    summary_ws["B7"] = stats["good_count"]

    cards = stats["cards"]
    metrics = [
        ("Avg Mobile Score", cards["mobile_score"].current, cards["mobile_score"].t7_delta, cards["mobile_score"].t30_delta),
        ("Avg Desktop Score", cards["desktop_score"].current, cards["desktop_score"].t7_delta, cards["desktop_score"].t30_delta),
        ("Avg FCP", cards["fcp"].current, cards["fcp"].t7_delta, cards["fcp"].t30_delta),
        ("Avg LCP", cards["lcp"].current, cards["lcp"].t7_delta, cards["lcp"].t30_delta),
        ("Avg FID", cards["fid"].current, cards["fid"].t7_delta, cards["fid"].t30_delta),
        ("Avg CLS", cards["cls"].current, cards["cls"].t7_delta, cards["cls"].t30_delta),
    ]
    start_row = 10
    for idx, (label, current, t7, t30) in enumerate(metrics, start=start_row):
        summary_ws[f"A{idx}"] = label
        summary_ws[f"B{idx}"] = current
        summary_ws[f"C{idx}"] = t7
        summary_ws[f"D{idx}"] = t30
    summary_ws["A9"] = "Metric"
    summary_ws["B9"] = "Current"
    summary_ws["C9"] = "T7 Delta"
    summary_ws["D9"] = "T30 Delta"
    for cell in summary_ws[9]:
        cell.font = Font(bold=True)
    autosize_worksheet(summary_ws)

    props_ws = wb.create_sheet("Property Detail")
    headers = [
        "Property Name",
        "Latest Mobile Date",
        "Mobile Score", "Mobile T7", "Mobile T30",
        "Desktop Score", "Desktop T7", "Desktop T30",
        "LCP", "LCP T7", "LCP T30",
        "CLS", "CLS T7", "CLS T30",
        "FID", "FID T7", "FID T30",
        "FCP", "FCP T7", "FCP T30",
    ]
    props_ws.append(headers)
    for cell in props_ws[1]:
        cell.font = Font(bold=True)

    for prop in properties:
        props_ws.append([
            prop["property_name"],
            prop["latest_mobile_date"],
            prop["mobile_score"], prop["mobile_score_t7"], prop["mobile_score_t30"],
            prop["desktop_score"], prop["desktop_score_t7"], prop["desktop_score_t30"],
            prop["lcp"], prop["lcp_t7"], prop["lcp_t30"],
            prop["cls"], prop["cls_t7"], prop["cls_t30"],
            prop["fid"], prop["fid_t7"], prop["fid_t30"],
            prop["fcp"], prop["fcp_t7"], prop["fcp_t30"],
        ])

    for row in range(2, props_ws.max_row + 1):
        score = props_ws[f"C{row}"].value
        score_cell = props_ws[f"C{row}"]
        if score is None:
            continue
        if score < 50:
            score_cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        elif score < 90:
            score_cell.fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        else:
            score_cell.fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    props_ws.freeze_panes = "A2"
    autosize_worksheet(props_ws)

    wb.save(output_path)


def generate_report(send_email: bool) -> Path:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    portfolio_daily = query_portfolio_daily(conn)
    properties = query_properties(conn)
    conn.close()

    if not portfolio_daily:
        raise RuntimeError("No pagespeed_metrics data available.")

    html = build_html(portfolio_daily, properties)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = OUT_DIR / f"portfolio_psi_pib_{stamp}.html"
    output_path.write_text(html, encoding="utf-8")
    excel_path = OUT_DIR / f"portfolio_psi_pib_{stamp}.xlsx"
    write_excel(excel_path, portfolio_daily, properties)

    if send_email:
        sender = EmailSender(verbose=False)
        log_path = DELIVERY_DIR / f"email_delivery_{datetime.now().strftime('%Y-%m-%d')}.jsonl"
        attachments = [
            (
                excel_path.name,
                excel_path.read_bytes(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        ]
        sender.send_email_with_tracking(
            subject=f"Venterra Portfolio CWV Report - {datetime.strptime(portfolio_daily[0]['metric_date'], '%Y-%m-%d').strftime('%m-%d-%Y')}",
            html_body=html,
            attachments=attachments,
            log_path=log_path,
        )

    return output_path


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate PIB-style portfolio PSI / CWV report")
    parser.add_argument("--email", action="store_true", help="Send the report using configured default recipients")
    args = parser.parse_args()

    out = generate_report(send_email=args.email)
    print(f"REPORT_HTML: {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
