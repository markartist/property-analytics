#!/usr/bin/env python3
from __future__ import annotations

import argparse
import sqlite3
import sys
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path("/Users/mark/Property_Analytics")
DB_PATH = ROOT / "data" / "portfolio_analytics.db"
OUT_DIR = ROOT / "reports" / "selected_cwv_t30"
OUT_DIR.mkdir(parents=True, exist_ok=True)

sys.path.insert(0, str(ROOT / "utils"))
from email_sender import EmailSender  # noqa: E402
from report_builder import KPITile, ReportBuilder, Section  # noqa: E402


ALIAS_MAP: Dict[str, str] = {
    "apex": "Apex West Midtown",
    "retreat": "The Retreat",
    "northbridge": "Northbridge at Millenia Lake",
    "maddox": "The Maddox",
    "stonecreek": "Stonecreek Ranch",
    "bradford mills": "Bradford Mills Lofts",
    "fairways": "Fairways at South Shore",
    "colony lakes": "Preserve at Colony Lakes",
    "spring branch": "Avasa Spring Branch",
    "cobblestone": "Cobblestone at Eagle Harbor",
    "townhomes": "Townhomes at Lake Park",
    "oakleaf": "The Villages at Oakleaf",
    "baywood": "Preserve at Baywood",
}

DEFAULT_REQUESTED_PROPERTIES: List[str] = [
    "Shadowbrooke",
    "Villa Lago",
    "Silverbrooke",
    "Gateway North",
    "Apex",
    "Retreat",
    "Champions Green",
    "Northbridge",
    "Maddox",
    "Belterra",
    "Stonecreek",
    "Canton Mill Lofts",
    "Bradford Mills",
    "Cendana District West",
    "French Place",
    "Clearwater Heights",
    "Fairways",
    "Colony Lakes",
    "Spring Branch",
    "Keystone",
    "Cobblestone",
    "Townhomes",
    "Ventana",
    "Oakleaf",
    "Baywood",
]


@dataclass
class PropertyMetricRow:
    requested_name: str
    property_name: str
    property_id: str
    latest_date: str
    coverage_days: int
    current_score: Optional[float]
    t30_score_avg: Optional[float]
    score_delta: Optional[float]
    current_lcp: Optional[float]
    t30_lcp_avg: Optional[float]
    lcp_delta: Optional[float]
    current_cls: Optional[float]
    t30_cls_avg: Optional[float]
    cls_delta: Optional[float]
    current_tbt: Optional[float]
    t30_tbt_avg: Optional[float]
    tbt_delta: Optional[float]


def normalize_name(value: str) -> str:
    return " ".join(value.strip().lower().split())


def resolve_requested_properties(conn: sqlite3.Connection, requested: Sequence[str]) -> Tuple[List[Tuple[str, str]], List[str]]:
    property_names = {
        row["property_name"]: row["property_name"]
        for row in conn.execute("SELECT property_name FROM properties")
    }
    lookup = {normalize_name(name): name for name in property_names}

    resolved: List[Tuple[str, str]] = []
    unresolved: List[str] = []
    for name in requested:
        key = normalize_name(name)
        canonical = lookup.get(key)
        if canonical is None:
            canonical = ALIAS_MAP.get(key)
        if canonical is None:
            unresolved.append(name)
            continue
        resolved.append((name, canonical))
    return resolved, unresolved


def fmt_num(value: Optional[float], decimals: int = 1) -> str:
    if value is None:
        return "—"
    return f"{value:.{decimals}f}"


def fmt_score(value: Optional[float]) -> str:
    if value is None:
        return "—"
    return f"{round(value):.0f}"


def fmt_delta(value: Optional[float], decimals: int = 1, invert: bool = False) -> str:
    if value is None:
        return "—"
    if value == 0:
        arrow = "→"
    else:
        arrow = "↓" if value < 0 else "↑"
    if invert:
        color = "#15803d" if value < 0 else "#b91c1c" if value > 0 else "#64748b"
    else:
        color = "#15803d" if value > 0 else "#b91c1c" if value < 0 else "#64748b"
    return f'<span style="color:{color}; font-weight:600;">{arrow}{abs(value):.{decimals}f}</span>'


def score_fill(value: Optional[float]) -> str:
    if value is None:
        return "#f8fafc"
    if value >= 90:
        return "#dcfce7"
    if value >= 50:
        return "#ffedd5"
    return "#fee2e2"


def build_rows(conn: sqlite3.Connection, requested_pairs: Sequence[Tuple[str, str]], report_date: str) -> List[PropertyMetricRow]:
    window_start = (datetime.strptime(report_date, "%Y-%m-%d").date() - timedelta(days=29)).isoformat()
    rows: List[PropertyMetricRow] = []
    for requested_name, property_name in requested_pairs:
        metrics = conn.execute(
            """
            SELECT
                p.property_id,
                p.property_name,
                pm.metric_date,
                pm.performance_score,
                pm.lcp_value,
                pm.cls_value,
                pm.total_blocking_time
            FROM pagespeed_metrics pm
            JOIN properties p ON p.property_id = pm.property_id
            WHERE pm.strategy = 'mobile'
              AND p.property_name = ?
              AND pm.metric_date BETWEEN ? AND ?
            ORDER BY pm.metric_date DESC
            """,
            (property_name, window_start, report_date),
        ).fetchall()
        if not metrics:
            continue
        current = metrics[0]

        def avg(field: str) -> Optional[float]:
            vals = [row[field] for row in metrics if row[field] is not None]
            return (sum(vals) / len(vals)) if vals else None

        t30_score = avg("performance_score")
        t30_lcp = avg("lcp_value")
        t30_cls = avg("cls_value")
        t30_tbt = avg("total_blocking_time")

        rows.append(
            PropertyMetricRow(
                requested_name=requested_name,
                property_name=property_name,
                property_id=current["property_id"],
                latest_date=current["metric_date"],
                coverage_days=len({row["metric_date"] for row in metrics}),
                current_score=current["performance_score"],
                t30_score_avg=t30_score,
                score_delta=(current["performance_score"] - t30_score) if current["performance_score"] is not None and t30_score is not None else None,
                current_lcp=current["lcp_value"],
                t30_lcp_avg=t30_lcp,
                lcp_delta=(current["lcp_value"] - t30_lcp) if current["lcp_value"] is not None and t30_lcp is not None else None,
                current_cls=current["cls_value"],
                t30_cls_avg=t30_cls,
                cls_delta=(current["cls_value"] - t30_cls) if current["cls_value"] is not None and t30_cls is not None else None,
                current_tbt=current["total_blocking_time"],
                t30_tbt_avg=t30_tbt,
                tbt_delta=(current["total_blocking_time"] - t30_tbt) if current["total_blocking_time"] is not None and t30_tbt is not None else None,
            )
        )
    return sorted(rows, key=lambda r: (r.current_score is None, r.current_score if r.current_score is not None else 999, r.property_name))


def build_summary(rows: Sequence[PropertyMetricRow], report_date: str) -> Tuple[List[KPITile], List[str]]:
    score_values = [r.current_score for r in rows if r.current_score is not None]
    t30_values = [r.t30_score_avg for r in rows if r.t30_score_avg is not None]
    lcp_values = [r.current_lcp for r in rows if r.current_lcp is not None]
    stale = [r for r in rows if r.latest_date != report_date]

    avg_current = sum(score_values) / len(score_values) if score_values else None
    avg_t30 = sum(t30_values) / len(t30_values) if t30_values else None
    delta = (avg_current - avg_t30) if avg_current is not None and avg_t30 is not None else None
    avg_lcp = sum(lcp_values) / len(lcp_values) if lcp_values else None

    tiles = [
        KPITile("Current Avg PSI", fmt_num(avg_current, 1), trend=(f"Δ {delta:+.1f} vs T30" if delta is not None else None), is_primary=True),
        KPITile("T30 Avg PSI", fmt_num(avg_t30, 1), sublabel="Average of available mobile PSI runs in the last 30 days"),
        KPITile("Avg Current LCP", f"{fmt_num(avg_lcp, 2)}s", sublabel="Latest mobile PSI run"),
    ]
    stale_lines = [f"{r.property_name}: latest PSI {r.latest_date}" for r in stale]
    return tiles, stale_lines


def table_html(rows: Sequence[PropertyMetricRow], report_date: str, unresolved: Sequence[str]) -> str:
    body_rows = []
    for row in rows:
        stale_note = "" if row.latest_date == report_date else f"<div style='font-size:11px;color:#b45309;'>latest {row.latest_date}</div>"
        body_rows.append(
            f"""
            <tr>
              <td style="padding:10px 12px; border-bottom:1px solid #e5e7eb; font-weight:600;">{row.property_name}{stale_note}</td>
              <td style="padding:10px 12px; border-bottom:1px solid #e5e7eb; text-align:center; background:{score_fill(row.current_score)};">{fmt_score(row.current_score)}</td>
              <td style="padding:10px 12px; border-bottom:1px solid #e5e7eb; text-align:center;">{fmt_num(row.t30_score_avg,1)}</td>
              <td style="padding:10px 12px; border-bottom:1px solid #e5e7eb; text-align:center;">{fmt_delta(row.score_delta,1, invert=False)}</td>
              <td style="padding:10px 12px; border-bottom:1px solid #e5e7eb; text-align:center;">{fmt_num(row.current_lcp,2)}s</td>
              <td style="padding:10px 12px; border-bottom:1px solid #e5e7eb; text-align:center;">{fmt_num(row.t30_lcp_avg,2)}s</td>
              <td style="padding:10px 12px; border-bottom:1px solid #e5e7eb; text-align:center;">{fmt_delta(row.lcp_delta,2, invert=True)}</td>
            </tr>
            """
        )

    notes = [
        "<li>T30 values are averages of available mobile PSI rows from the last 30 calendar days ending on the portfolio report date.</li>",
        "<li>GTMetrix is intentionally excluded here because canonical GT coverage exists only for the 10-property pilot cohort, not this 25-property set.</li>",
    ]
    if unresolved:
        notes.append("<li>Unresolved requested names: " + ", ".join(unresolved) + ".</li>")

    return f"""
    <table cellpadding="0" cellspacing="0" border="0" style="width:100%; border:1px solid #e5e7eb; border-radius:8px; overflow:hidden;">
      <thead>
        <tr style="background:#f8fafc;">
          <th style="padding:10px 12px; text-align:left; border-bottom:1px solid #e5e7eb;">Property</th>
          <th style="padding:10px 12px; text-align:center; border-bottom:1px solid #e5e7eb;">Current PSI</th>
          <th style="padding:10px 12px; text-align:center; border-bottom:1px solid #e5e7eb;">T30 PSI</th>
          <th style="padding:10px 12px; text-align:center; border-bottom:1px solid #e5e7eb;">Current vs T30</th>
          <th style="padding:10px 12px; text-align:center; border-bottom:1px solid #e5e7eb;">Current LCP</th>
          <th style="padding:10px 12px; text-align:center; border-bottom:1px solid #e5e7eb;">T30 LCP</th>
          <th style="padding:10px 12px; text-align:center; border-bottom:1px solid #e5e7eb;">LCP vs T30</th>
        </tr>
      </thead>
      <tbody>
        {''.join(body_rows)}
      </tbody>
    </table>
    <div style="margin-top:16px;">
      <ul style="margin:0; padding-left:20px; color:#475569; font-size:13px; line-height:1.7;">
        {''.join(notes)}
      </ul>
    </div>
    """


def write_excel(path: Path, rows: Sequence[PropertyMetricRow]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "CWV T30"
    headers = [
        "Canonical Property",
        "Latest Date",
        "Current PSI",
        "T30 PSI",
        "Current vs T30 PSI",
        "Current LCP (s)",
        "T30 LCP (s)",
        "Current vs T30 LCP (s)",
    ]
    ws.append(headers)
    header_fill = PatternFill(fill_type="solid", fgColor="15284B")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for row in rows:
        ws.append([
            row.property_name,
            row.latest_date,
            row.current_score,
            row.t30_score_avg,
            row.score_delta,
            row.current_lcp,
            row.t30_lcp_avg,
            row.lcp_delta,
        ])

    widths = [30, 12, 12, 10, 16, 15, 12, 16]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"
    wb.save(path)


def plain_text(rows: Sequence[PropertyMetricRow], report_date: str) -> str:
    lines = [
        f"Selected Property CWV T30 Report",
        f"Report date: {report_date}",
        "",
        "Columns: Current PSI | T30 PSI | Current LCP | T30 LCP | Coverage Days",
        "",
    ]
    for row in rows:
        lines.append(
            f"{row.property_name}: {fmt_score(row.current_score)} | {fmt_num(row.t30_score_avg,1)} | "
            f"{fmt_num(row.current_lcp,2)}s | {fmt_num(row.t30_lcp_avg,2)}s | {row.coverage_days}"
        )
    return "\n".join(lines)


def generate_html(rows: Sequence[PropertyMetricRow], report_date: str, unresolved: Sequence[str]) -> str:
    builder = ReportBuilder(
        title="Selected Property CWV T30 Report",
        subtitle="Mobile PSI / Core Web Vitals summary for requested communities",
        date_range=datetime.strptime(report_date, "%Y-%m-%d").strftime("%B %d, %Y"),
    )
    tiles, stale_lines = build_summary(rows, report_date)
    builder.add_kpi_tiles(tiles, columns=3)

    note_bits = []
    if stale_lines:
        note_bits.append("Stale latest-date exceptions: " + "; ".join(stale_lines))
    if "Retreat" in [r.requested_name for r in rows]:
        note_bits.append("Shorthand assumption applied: Retreat → The Retreat.")
    description = " | ".join(note_bits) if note_bits else "T30 anchored to the latest portfolio PSI date."

    builder.add_section(
        Section(
            title="Selected Property Detail",
            status=None,
            description=description,
            content=table_html(rows, report_date, unresolved),
        )
    )
    return builder.generate()


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate and email a selected-property CWV T30 report.")
    parser.add_argument("--recipient", action="append", dest="recipients", help="Email recipient (can be repeated)")
    parser.add_argument("--subject", help="Email subject override")
    parser.add_argument("properties", nargs="*", help="Requested property names; defaults to the current requested set")
    args = parser.parse_args()

    requested = args.properties or DEFAULT_REQUESTED_PROPERTIES
    recipients = args.recipients or ["mlaufhutte@venterraliving.com"]

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    report_date = conn.execute(
        "SELECT MAX(metric_date) FROM pagespeed_metrics WHERE strategy='mobile'"
    ).fetchone()[0]
    if not report_date:
        raise RuntimeError("No mobile PSI data available.")

    resolved, unresolved = resolve_requested_properties(conn, requested)
    rows = build_rows(conn, resolved, report_date)
    if not rows:
        raise RuntimeError("No selected properties resolved to PSI data.")

    stamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    html_path = OUT_DIR / f"Selected_Property_CWV_T30_{report_date}_{stamp}.html"
    xlsx_path = OUT_DIR / f"Selected_Property_CWV_T30_{report_date}_{stamp}.xlsx"
    html = generate_html(rows, report_date, unresolved)
    html_path.write_text(html)
    write_excel(xlsx_path, rows)

    sender = EmailSender(verbose=False)
    subject = args.subject or f"Selected Property CWV T30 Report - {report_date}"
    delivery = sender.send_email_with_tracking(
        subject=subject,
        html_body=html,
        plain_text=plain_text(rows, report_date),
        recipients=recipients,
        attachments=[
            (xlsx_path.name, xlsx_path.read_bytes(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        ],
    )
    print(f"HTML report: {html_path}")
    print(f"Excel report: {xlsx_path}")
    print(f"Recipients: {', '.join(recipients)}")
    print(f"Message ID: {delivery.get('message_id')}")
    if unresolved:
        print(f"Unresolved requested names: {', '.join(unresolved)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
