#!/usr/bin/env python3
"""Generate organic search share of traffic report for portfolio and pilot sites."""

from __future__ import annotations

import csv
import html
import json
import sqlite3
import sys
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path("/Users/mark/Property_Analytics")
DB_PATH = ROOT / "data" / "portfolio_analytics.db"
REGISTRY_PATH = ROOT / "config" / "venterra_properties_official.json"
PILOT_CONFIG_PATH = ROOT / "pilot_control_cwv" / "config" / "pilot_control_cwv_config.json"
OUT_ROOT = ROOT / "reports" / "organic_search_traffic_share_portfolio_pilot"
REQUESTED_START = date(2024, 6, 25)
PILOT_LAUNCH_DATE = date(2026, 3, 24)

sys.path.insert(0, str(ROOT))
from Data_Collection.utils.property_identity import resolve_property_identity  # noqa: E402
from utils.report_builder import VENTERRA_LOGO_BASE64  # noqa: E402


BRAND = {
    "navy": "15284B",
    "san_marino": "3D66B9",
    "bay": "294782",
    "white_smoke": "F6F6F5",
    "terra_cotta": "BD4830",
    "quill": "D6D6D2",
    "blue_chill": "3B9189",
    "delta": "9B9B96",
    "black": "000000",
    "white": "FFFFFF",
}


@dataclass(frozen=True)
class Property:
    name: str
    property_id: str
    property_code: str | None


@dataclass(frozen=True)
class PilotPair:
    pair_key: str
    pilot_name: str
    pilot_id: str
    control_name: str
    control_id: str


def iso(value: date) -> str:
    return value.isoformat()


def parse_date(value: str) -> date:
    return datetime.strptime(value, "%Y-%m-%d").date()


def placeholders(values: Iterable[str]) -> str:
    return ",".join("?" for _ in values)


def fmt_num(value: Any, decimals: int = 0) -> str:
    if value is None:
        return "-"
    return f"{float(value):,.{decimals}f}"


def fmt_pct(value: Any, decimals: int = 1, signed: bool = False) -> str:
    if value is None:
        return "-"
    sign = "+" if signed else ""
    return f"{float(value) * 100:{sign},.{decimals}f}%"


def fmt_pp(value: Any, decimals: int = 1) -> str:
    if value is None:
        return "-"
    return f"{float(value) * 100:+,.{decimals}f} pts"


def pct_change(current: float | int | None, prior: float | int | None) -> float | None:
    if current is None or prior in (None, 0):
        return None
    return (float(current) - float(prior)) / float(prior)


def point_change(current: float | None, prior: float | None) -> float | None:
    if current is None or prior is None:
        return None
    return float(current) - float(prior)


def load_portfolio_properties() -> list[Property]:
    payload = json.loads(REGISTRY_PATH.read_text(encoding="utf-8"))
    properties: list[Property] = []
    for row in payload.get("properties", []):
        ga4 = row.get("ga4_property_id")
        if not ga4 or row.get("property_type") == "new_development":
            continue
        identity = resolve_property_identity(str(ga4)) or resolve_property_identity(str(row.get("name", "")))
        properties.append(
            Property(
                name=(identity.property_name if identity else row.get("name")) or str(ga4),
                property_id=str(ga4),
                property_code=identity.property_code if identity else None,
            )
        )
    return sorted(properties, key=lambda item: item.name)


def load_pilot_pairs() -> list[PilotPair]:
    payload = json.loads(PILOT_CONFIG_PATH.read_text(encoding="utf-8"))
    cohorts = {item["key"]: item for item in payload.get("cohorts", []) if item.get("active")}
    pairs: list[PilotPair] = []
    for pilot in [item for item in cohorts.values() if item.get("role") == "pilot"]:
        control = cohorts.get(str(pilot.get("sister_key")))
        if not control:
            continue
        pilot_identity = resolve_property_identity(str(pilot["property_id"])) or resolve_property_identity(str(pilot["display_name"]))
        control_identity = resolve_property_identity(str(control["property_id"])) or resolve_property_identity(str(control["display_name"]))
        pairs.append(
            PilotPair(
                pair_key=str(pilot["key"]).replace("pilot_", ""),
                pilot_name=pilot_identity.property_name if pilot_identity else str(pilot["display_name"]),
                pilot_id=str(pilot["property_id"]),
                control_name=control_identity.property_name if control_identity else str(control["display_name"]),
                control_id=str(control["property_id"]),
            )
        )
    return sorted(pairs, key=lambda item: item.pilot_name)


def fetch_one(conn: sqlite3.Connection, sql: str, params: tuple[Any, ...] = ()) -> sqlite3.Row:
    row = conn.execute(sql, params).fetchone()
    if row is None:
        raise RuntimeError("Query returned no row")
    return row


def fetch_share_summary(conn: sqlite3.Connection, property_ids: list[str], start: date, end: date) -> dict[str, Any]:
    ph = placeholders(property_ids)
    row = fetch_one(
        conn,
        f"""
        WITH total AS (
            SELECT SUM(sessions) AS total_sessions, SUM(new_users) AS total_new_users, COUNT(DISTINCT metric_date) AS days
            FROM ga4_daily_metrics
            WHERE property_id IN ({ph})
              AND metric_date BETWEEN ? AND ?
        ),
        organic AS (
            SELECT SUM(sessions) AS organic_sessions, SUM(new_users) AS organic_new_users
            FROM ga4_traffic_sources
            WHERE property_id IN ({ph})
              AND channel_group = 'Organic Search'
              AND metric_date BETWEEN ? AND ?
        )
        SELECT
            total.days,
            total.total_sessions,
            total.total_new_users,
            organic.organic_sessions,
            organic.organic_new_users,
            CASE WHEN total.total_sessions > 0 THEN CAST(organic.organic_sessions AS REAL) / total.total_sessions END AS organic_session_share,
            CASE WHEN total.total_new_users > 0 THEN CAST(organic.organic_new_users AS REAL) / total.total_new_users END AS organic_new_user_share
        FROM total CROSS JOIN organic
        """,
        tuple([*property_ids, iso(start), iso(end), *property_ids, iso(start), iso(end)]),
    )
    return dict(row)


def fetch_monthly(conn: sqlite3.Connection, property_ids: list[str], start: date, end: date) -> list[dict[str, Any]]:
    ph = placeholders(property_ids)
    rows = conn.execute(
        f"""
        WITH total AS (
            SELECT substr(metric_date, 1, 7) AS month, SUM(sessions) AS total_sessions, SUM(new_users) AS total_new_users
            FROM ga4_daily_metrics
            WHERE property_id IN ({ph})
              AND metric_date BETWEEN ? AND ?
            GROUP BY month
        ),
        organic AS (
            SELECT substr(metric_date, 1, 7) AS month, SUM(sessions) AS organic_sessions, SUM(new_users) AS organic_new_users
            FROM ga4_traffic_sources
            WHERE property_id IN ({ph})
              AND channel_group = 'Organic Search'
              AND metric_date BETWEEN ? AND ?
            GROUP BY month
        )
        SELECT
            total.month,
            total.total_sessions,
            organic.organic_sessions,
            CASE WHEN total.total_sessions > 0 THEN CAST(organic.organic_sessions AS REAL) / total.total_sessions END AS organic_session_share,
            total.total_new_users,
            organic.organic_new_users,
            CASE WHEN total.total_new_users > 0 THEN CAST(organic.organic_new_users AS REAL) / total.total_new_users END AS organic_new_user_share
        FROM total
        LEFT JOIN organic ON organic.month = total.month
        ORDER BY total.month
        """,
        tuple([*property_ids, iso(start), iso(end), *property_ids, iso(start), iso(end)]),
    ).fetchall()
    return [dict(row) for row in rows]


def fetch_property_rows(conn: sqlite3.Connection, properties: list[Property], start: date, end: date) -> list[dict[str, Any]]:
    property_ids = [item.property_id for item in properties]
    name_map = {item.property_id: item for item in properties}
    ph = placeholders(property_ids)
    rows = conn.execute(
        f"""
        WITH total AS (
            SELECT property_id, SUM(sessions) AS total_sessions, SUM(new_users) AS total_new_users
            FROM ga4_daily_metrics
            WHERE property_id IN ({ph})
              AND metric_date BETWEEN ? AND ?
            GROUP BY property_id
        ),
        organic AS (
            SELECT property_id, SUM(sessions) AS organic_sessions, SUM(new_users) AS organic_new_users
            FROM ga4_traffic_sources
            WHERE property_id IN ({ph})
              AND channel_group = 'Organic Search'
              AND metric_date BETWEEN ? AND ?
            GROUP BY property_id
        )
        SELECT
            total.property_id,
            total.total_sessions,
            COALESCE(organic.organic_sessions, 0) AS organic_sessions,
            CASE WHEN total.total_sessions > 0 THEN CAST(COALESCE(organic.organic_sessions, 0) AS REAL) / total.total_sessions END AS organic_session_share,
            total.total_new_users,
            COALESCE(organic.organic_new_users, 0) AS organic_new_users,
            CASE WHEN total.total_new_users > 0 THEN CAST(COALESCE(organic.organic_new_users, 0) AS REAL) / total.total_new_users END AS organic_new_user_share
        FROM total
        LEFT JOIN organic ON organic.property_id = total.property_id
        ORDER BY organic_session_share DESC
        """,
        tuple([*property_ids, iso(start), iso(end), *property_ids, iso(start), iso(end)]),
    ).fetchall()
    output: list[dict[str, Any]] = []
    for idx, row in enumerate(rows, start=1):
        prop = name_map[str(row["property_id"])]
        item = dict(row)
        item["rank"] = idx
        item["property_name"] = prop.name
        item["property_code"] = prop.property_code
        output.append(item)
    return output


def fetch_pilot_rows(conn: sqlite3.Connection, pairs: list[PilotPair], start: date, end: date) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for pair in pairs:
        pilot = fetch_share_summary(conn, [pair.pilot_id], start, end)
        control = fetch_share_summary(conn, [pair.control_id], start, end)
        rows.append(
            {
                "pair_key": pair.pair_key,
                "pilot_property": pair.pilot_name,
                "control_property": pair.control_name,
                "pilot_organic_traffic_share": pilot.get("organic_session_share"),
                "control_organic_traffic_share": control.get("organic_session_share"),
                "share_point_gap": point_change(pilot.get("organic_session_share"), control.get("organic_session_share")),
                "pilot_organic_sessions": pilot.get("organic_sessions") or 0,
                "control_organic_sessions": control.get("organic_sessions") or 0,
                "pilot_total_sessions": pilot.get("total_sessions") or 0,
                "control_total_sessions": control.get("total_sessions") or 0,
                "pilot_organic_new_user_share": pilot.get("organic_new_user_share"),
                "control_organic_new_user_share": control.get("organic_new_user_share"),
            }
        )
    return sorted(rows, key=lambda item: item["pilot_organic_traffic_share"] or 0, reverse=True)


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        return
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def autosize(ws) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 11), 42)


def write_sheet(ws, rows: list[dict[str, Any]]) -> None:
    if not rows:
        return
    header_fill = PatternFill("solid", fgColor=BRAND["navy"])
    header_font = Font(color=BRAND["white"], bold=True)
    for col_idx, key in enumerate(rows[0].keys(), start=1):
        cell = ws.cell(row=1, column=col_idx, value=key)
        cell.fill = header_fill
        cell.font = header_font
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, key in enumerate(row.keys(), start=1):
            ws.cell(row=row_idx, column=col_idx, value=row[key])
    ws.freeze_panes = "A2"
    autosize(ws)


def build_workbook(path: Path, sheets: dict[str, list[dict[str, Any]]], metadata: list[list[Any]]) -> None:
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary["A1"] = "Organic Search Share Of Traffic Report"
    summary["A1"].font = Font(bold=True, size=16, color=BRAND["navy"])
    for row_idx, row in enumerate(metadata, start=3):
        summary.cell(row=row_idx, column=1, value=row[0]).font = Font(bold=True, color=BRAND["bay"])
        summary.cell(row=row_idx, column=2, value=row[1])
    autosize(summary)
    for title, rows in sheets.items():
        ws = wb.create_sheet(title[:31])
        write_sheet(ws, rows)
    wb.save(path)


def html_table(rows: list[dict[str, Any]], columns: list[tuple[str, str]], limit: int | None = None) -> str:
    visible = rows[:limit] if limit else rows
    header = "".join(
        (
            '<th style="text-align:left; padding:9px 8px; background:#F6F6F5; '
            'border:1px solid #D6D6D2; font-size:11px; color:#15284B; '
            f'font-weight:700;">{html.escape(label)}</th>'
        )
        for _, label in columns
    )
    body_rows = []
    for idx, row in enumerate(visible):
        bg = "#F6F6F5" if idx % 2 else "#FFFFFF"
        cells = "".join(
            (
                f'<td style="padding:8px; border:1px solid #D6D6D2; '
                f'font-size:11px; color:#000000; vertical-align:top; background:{bg};">'
                f"{html.escape(str(row.get(key, '-')))}</td>"
            )
            for key, _ in columns
        )
        body_rows.append(f"<tr>{cells}</tr>")
    return (
        '<table cellpadding="0" cellspacing="0" border="0" '
        'style="width:100%; border-collapse:collapse; margin:0;">'
        f"<tr>{header}</tr>{''.join(body_rows)}</table>"
    )


def display_monthly(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "month": row["month"],
            "organic_session_share": fmt_pct(row.get("organic_session_share")),
            "organic_sessions": fmt_num(row.get("organic_sessions")),
            "total_sessions": fmt_num(row.get("total_sessions")),
            "organic_new_user_share": fmt_pct(row.get("organic_new_user_share")),
        }
        for row in rows
    ]


def display_properties(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "rank": row["rank"],
            "property_name": row["property_name"],
            "property_code": row["property_code"] or "",
            "organic_session_share": fmt_pct(row.get("organic_session_share")),
            "organic_sessions": fmt_num(row.get("organic_sessions")),
            "total_sessions": fmt_num(row.get("total_sessions")),
            "organic_new_user_share": fmt_pct(row.get("organic_new_user_share")),
        }
        for row in rows
    ]


def display_pilots(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "pilot_property": row["pilot_property"],
            "control_property": row["control_property"],
            "pilot_organic_traffic_share": fmt_pct(row.get("pilot_organic_traffic_share")),
            "control_organic_traffic_share": fmt_pct(row.get("control_organic_traffic_share")),
            "share_point_gap": fmt_pp(row.get("share_point_gap")),
            "pilot_organic_sessions": fmt_num(row.get("pilot_organic_sessions")),
            "pilot_total_sessions": fmt_num(row.get("pilot_total_sessions")),
            "pilot_organic_new_user_share": fmt_pct(row.get("pilot_organic_new_user_share")),
        }
        for row in rows
    ]


def kpi_tile(label: str, value: str, primary: bool = False) -> str:
    border = f"2px solid #{BRAND['san_marino']}" if primary else f"1px solid #{BRAND['quill']}"
    label_color = f"#{BRAND['san_marino']}" if primary else f"#{BRAND['delta']}"
    return f"""
        <table cellpadding="0" cellspacing="0" border="0" style="width:100%; border:{border}; background:#{BRAND['white']};">
            <tr>
                <td style="padding:22px 16px; text-align:center; height:118px; vertical-align:middle;">
                    <div style="font-size:11px; color:{label_color}; text-transform:uppercase; letter-spacing:0.4px; font-weight:700; margin-bottom:12px;">{html.escape(label)}</div>
                    <div style="font-size:34px; line-height:1; color:#{BRAND['black']}; font-weight:700;">{html.escape(value)}</div>
                </td>
            </tr>
        </table>
    """


def section_block(title: str, content: str) -> str:
    return f"""
        <table cellpadding="0" cellspacing="0" border="0" style="width:100%; margin:26px 0 0 0;">
            <tr>
                <td style="background:#{BRAND['navy']}; padding:14px 18px;">
                    <h2 style="font-size:20px; color:#{BRAND['white']}; margin:0; font-weight:700;">{html.escape(title)}</h2>
                </td>
            </tr>
            <tr>
                <td style="background:#{BRAND['white']}; border:1px solid #{BRAND['quill']}; border-top:0; padding:18px;">
                    {content}
                </td>
            </tr>
        </table>
    """


def build_html(path: Path, summary: dict[str, Any], properties: list[dict[str, Any]], pilots: list[dict[str, Any]]) -> None:
    executive_content = f"""
        <table cellpadding="0" cellspacing="0" border="0" style="width:100%; background:#{BRAND['white_smoke']}; border:1px solid #{BRAND['quill']}; border-left:5px solid #{BRAND['blue_chill']}; margin:0 0 16px 0;">
            <tr>
                <td style="padding:13px 14px; font-size:13px; line-height:1.45; color:#{BRAND['black']};">
                    <strong style="color:#{BRAND['navy']};">Question answered:</strong> What percent of website traffic comes from Organic Search?<br>
                    <strong style="color:#{BRAND['navy']};">Portfolio:</strong> {html.escape(summary['portfolio_organic_share'])} of sessions are Organic Search.<br>
                    <strong style="color:#{BRAND['navy']};">Five pilot sites:</strong> {html.escape(summary['pilot_organic_share'])} of sessions are Organic Search, versus {html.escape(summary['control_organic_share'])} for controls.
                </td>
            </tr>
        </table>
        <p style="font-size:13px; line-height:1.45; color:#{BRAND['black']}; margin:0 0 12px 0;">{html.escape(summary['executive_read'])}</p>
        <p style="font-size:13px; line-height:1.45; color:#{BRAND['black']}; margin:0 0 12px 0;">{html.escape(summary['t30_note'])}</p>
        <table cellpadding="0" cellspacing="0" border="0" style="width:100%; background:#{BRAND['white']}; border:1px solid #{BRAND['quill']}; border-left:4px solid #{BRAND['terra_cotta']};">
            <tr>
                <td style="padding:12px 14px; font-size:12px; line-height:1.45; color:#{BRAND['bay']};">{html.escape(summary['coverage_note'])}</td>
            </tr>
        </table>
    """
    pilot_content = html_table(
        pilots,
        [
            ("pilot_property", "Pilot"),
            ("control_property", "Control"),
            ("pilot_organic_traffic_share", "Pilot organic %"),
            ("control_organic_traffic_share", "Control organic %"),
            ("share_point_gap", "Point gap"),
            ("pilot_organic_sessions", "Pilot organic sessions"),
            ("pilot_total_sessions", "Pilot total sessions"),
            ("pilot_organic_new_user_share", "Pilot organic new-user %"),
        ],
    )
    property_content = f"""
        <p style="font-size:13px; line-height:1.45; color:#{BRAND['black']}; margin:0 0 14px 0;">All properties with governed GA4 coverage are listed below, ranked by Organic Search share of traffic.</p>
        {html_table(properties, [
            ('rank', 'Rank'),
            ('property_name', 'Property'),
            ('property_code', 'Property code'),
            ('organic_session_share', 'Organic % of traffic'),
            ('organic_sessions', 'Organic sessions'),
            ('total_sessions', 'Total sessions'),
            ('organic_new_user_share', 'Organic % of new users')
        ])}
    """
    content = f"""<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>Organic Search Share Of Traffic Report</title>
</head>
<body style="margin:0; padding:20px; background:#{BRAND['white_smoke']}; font-family:Arial, Helvetica, sans-serif; color:#{BRAND['black']};">
  <table cellpadding="0" cellspacing="0" border="0" align="center" width="720" style="width:720px; max-width:720px; margin:0 auto; background:#{BRAND['white']};">
    <tr>
      <td style="padding:30px 40px 42px 40px;">
        <table cellpadding="0" cellspacing="0" border="0" style="width:100%; text-align:center; border-bottom:2px solid #{BRAND['quill']}; margin:0 0 28px 0;">
          <tr>
            <td style="padding:0 0 24px 0; text-align:center;">
              <img src="data:image/png;base64,{VENTERRA_LOGO_BASE64}" alt="Venterra" style="height:15px; width:auto; display:inline-block; border:0; margin:0 0 22px 0;">
              <div style="font-size:16px; color:#{BRAND['san_marino']}; font-weight:700; text-transform:uppercase; margin:0 0 20px 0;">Property Intelligence Brief</div>
              <div style="font-size:34px; line-height:1.12; color:#{BRAND['black']}; font-weight:700; margin:0;">Organic Search Share Of Traffic Report</div>
              <div style="font-size:11px; color:#{BRAND['delta']}; margin:20px 0 0 0; font-weight:700;">v1.0.0</div>
              <div style="font-size:14px; color:#{BRAND['delta']}; margin:14px 0 0 0; font-weight:700;">{html.escape(summary['effective_ga4_window'])} &middot; Generated {html.escape(summary['generated_at'])}</div>
            </td>
          </tr>
        </table>
        <table cellpadding="0" cellspacing="0" border="0" style="width:100%; margin:0 0 30px 0;">
          <tr>
            <td style="width:32%; vertical-align:top;">{kpi_tile("Organic % Of Traffic", summary["portfolio_organic_share"])}</td>
            <td style="width:2%; font-size:1px; line-height:1px;">&nbsp;</td>
            <td style="width:32%; vertical-align:top;">{kpi_tile("Organic Sessions", summary["portfolio_organic_sessions"], primary=True)}</td>
            <td style="width:2%; font-size:1px; line-height:1px;">&nbsp;</td>
            <td style="width:32%; vertical-align:top;">{kpi_tile("Total Sessions", summary["portfolio_total_sessions"])}</td>
          </tr>
        </table>
        {section_block("Executive Read", executive_content)}
        {section_block("Pilot And Control Organic Share", pilot_content)}
        {section_block("Exhaustive Property Performance", property_content)}
        <table cellpadding="0" cellspacing="0" border="0" style="width:100%; margin:24px 0 0 0;">
          <tr>
            <td style="font-size:11px; color:#{BRAND['bay']}; text-align:center;">Primary formula: GA4 Organic Search sessions / GA4 total sessions. Pilot launch/post-launch window starts {PILOT_LAUNCH_DATE.isoformat()}.</td>
          </tr>
        </table>
      </td>
    </tr>
  </table>
</body>
</html>"""
    path.write_text(content, encoding="utf-8")


def main() -> int:
    generated_at = datetime.now().astimezone()
    run_dir = OUT_ROOT / generated_at.strftime("%Y%m%d_%H%M%S")
    run_dir.mkdir(parents=True, exist_ok=True)

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    properties = load_portfolio_properties()
    property_ids = [item.property_id for item in properties]
    pairs = load_pilot_pairs()

    bounds = fetch_one(
        conn,
        f"""
        SELECT MIN(metric_date) AS min_date, MAX(metric_date) AS max_date
        FROM ga4_traffic_sources
        WHERE channel_group = 'Organic Search'
          AND property_id IN ({placeholders(property_ids)})
        """,
        tuple(property_ids),
    )
    start = max(REQUESTED_START, parse_date(bounds["min_date"]))
    end = parse_date(bounds["max_date"])

    portfolio = fetch_share_summary(conn, property_ids, start, end)
    current_start = end - timedelta(days=29)
    prior_end = current_start - timedelta(days=1)
    prior_start = prior_end - timedelta(days=29)
    current = fetch_share_summary(conn, property_ids, current_start, end)
    prior = fetch_share_summary(conn, property_ids, prior_start, prior_end)
    property_rows = fetch_property_rows(conn, properties, start, end)
    pilot_rows = fetch_pilot_rows(conn, pairs, PILOT_LAUNCH_DATE, end)
    pilot_ids = [pair.pilot_id for pair in pairs]
    control_ids = [pair.control_id for pair in pairs]
    pilot_rollup = fetch_share_summary(conn, pilot_ids, PILOT_LAUNCH_DATE, end)
    control_rollup = fetch_share_summary(conn, control_ids, PILOT_LAUNCH_DATE, end)

    share_change = point_change(current.get("organic_session_share"), prior.get("organic_session_share"))
    pilot_gap = point_change(pilot_rollup.get("organic_session_share"), control_rollup.get("organic_session_share"))
    summary = {
        "generated_at": generated_at.strftime("%Y-%m-%d %H:%M %Z"),
        "requested_window": f"{REQUESTED_START.isoformat()} through {end.isoformat()}",
        "effective_ga4_window": f"{start.isoformat()} through {end.isoformat()}",
        "portfolio_property_count": len(property_ids),
        "portfolio_organic_share": fmt_pct(portfolio.get("organic_session_share")),
        "portfolio_organic_sessions": fmt_num(portfolio.get("organic_sessions")),
        "portfolio_total_sessions": fmt_num(portfolio.get("total_sessions")),
        "pilot_organic_share": fmt_pct(pilot_rollup.get("organic_session_share")),
        "control_organic_share": fmt_pct(control_rollup.get("organic_session_share")),
        "coverage_note": (
            f"Requested last-two-years window starts {REQUESTED_START.isoformat()}; canonical local GA4 Organic Search coverage begins "
            f"{start.isoformat()}. This report does not infer missing {REQUESTED_START.isoformat()} to "
            f"{(start - timedelta(days=1)).isoformat()} traffic share."
        ),
        "executive_read": (
            f"Across {len(property_ids)} governed portfolio properties, Organic Search represented "
            f"{fmt_pct(portfolio.get('organic_session_share'))} of total traffic from {start.isoformat()} to {end.isoformat()} "
            f"({fmt_num(portfolio.get('organic_sessions'))} Organic Search sessions out of {fmt_num(portfolio.get('total_sessions'))} total sessions). "
            f"Post-launch, the five pilot sites were {fmt_pct(pilot_rollup.get('organic_session_share'))} organic versus "
            f"{fmt_pct(control_rollup.get('organic_session_share'))} for controls, a {fmt_pp(pilot_gap)} gap."
        ),
        "t30_note": (
            f"Latest T30 Organic Search share was {fmt_pct(current.get('organic_session_share'))}, "
            f"{fmt_pp(share_change)} vs the prior 30 days. Latest T30 Organic Search sessions were "
            f"{fmt_num(current.get('organic_sessions'))} out of {fmt_num(current.get('total_sessions'))} total sessions."
        ),
    }

    write_csv(run_dir / "property_organic_share_rankings.csv", property_rows)
    write_csv(run_dir / "pilot_post_launch_organic_share.csv", pilot_rows)

    build_workbook(
        run_dir / "organic_search_traffic_share_portfolio_pilot_report.xlsx",
        {
            "Property Rankings": property_rows,
            "Pilot Post Launch": pilot_rows,
        },
        [
            ["Generated", summary["generated_at"]],
            ["Requested Window", summary["requested_window"]],
            ["Effective GA4 Window", summary["effective_ga4_window"]],
            ["Portfolio Properties", summary["portfolio_property_count"]],
            ["Primary Formula", "GA4 Organic Search sessions / GA4 total sessions."],
            ["Coverage Note", summary["coverage_note"]],
        ],
    )
    build_html(
        run_dir / "organic_search_traffic_share_portfolio_pilot_report.html",
        summary,
        display_properties(property_rows),
        display_pilots(pilot_rows),
    )

    summary_path = run_dir / "organic_search_traffic_share_portfolio_pilot_summary.json"
    summary_path.write_text(
        json.dumps(
            {
                "summary": summary,
                "artifacts": {
                    "html": str(run_dir / "organic_search_traffic_share_portfolio_pilot_report.html"),
                    "workbook": str(run_dir / "organic_search_traffic_share_portfolio_pilot_report.xlsx"),
                    "property_rankings_csv": str(run_dir / "property_organic_share_rankings.csv"),
                    "pilot_post_launch_csv": str(run_dir / "pilot_post_launch_organic_share.csv"),
                },
            },
            indent=2,
        ),
        encoding="utf-8",
    )
    print(json.dumps({"run_dir": str(run_dir), "summary": summary}, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
