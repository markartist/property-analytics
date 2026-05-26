#!/usr/bin/env python3
"""Import the portfolio QA workbook into a governed EVS QA contract.

The workbook remains the human-maintained checklist. This importer turns the
Functionality and Data Integrity rows into a machine-readable contract that
BrowserStack/EVS can execute, defer, or route to another QA owner without
losing row-level lineage.
"""

from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_WORKBOOK = Path("/Users/mark/Downloads/_QA_Round 1_Property_Websites.xlsx")
DEFAULT_OUTPUT = Path("evs/config/portfolio-functionality-qa-contract.json")

CONTRACT_ID = "portfolio_functionality_qa_v1"
CONTRACT_VERSION = "2026-05-13"
INCLUDED_ELEMENTS = {"functionality", "data integrity"}


@dataclass(frozen=True)
class OwnerRule:
    owner: str
    automation_status: str
    side_effect_policy: str
    reason: str


def slugify(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")
    return slug or "item"


def detect_owner(page: str, section: str, description: str) -> OwnerRule:
    raw = f"{page} {section} {description}".lower()
    if "guest card" in raw or "ah & eai" in raw:
        return OwnerRule(
            owner="lead_attribution_qa",
            automation_status="blocked_governed_synthetic_lead",
            side_effect_policy="synthetic_lead_required",
            reason="Downstream AH/EAI proof needs a governed synthetic-lead workflow before unattended automation.",
        )
    if any(term in raw for term in ("camera icon", "photos open", "matterport", "virtual tour")):
        return OwnerRule(
            owner="evs",
            automation_status="ready_for_runner_mapping",
            side_effect_policy="no_submit",
            reason="Owned by EVS as browser-observable media interaction with no intentional form submission.",
        )
    if "images correct" in raw:
        return OwnerRule(
            owner="evs",
            automation_status="ready_for_runner_mapping",
            side_effect_policy="no_submit",
            reason="EVS can verify image/modal presence and render integrity; property-specific media correctness remains a review item.",
        )
    if any(term in raw for term in ("image", "photo")):
        return OwnerRule(
            owner="media_qa",
            automation_status="external_team",
            side_effect_policy="read_only",
            reason="Image/media correctness is owned outside EVS, while EVS can still preserve row lineage.",
        )
    if "required field validation" in raw:
        return OwnerRule(
            owner="forms_qa",
            automation_status="ready_for_runner_mapping",
            side_effect_policy="no_submit_validation_only",
            reason="Required-field validation can be exercised without submitting a form.",
        )
    if any(term in raw for term in ("contact form submit", "complete form")):
        return OwnerRule(
            owner="forms_qa",
            automation_status="external_team",
            side_effect_policy="form_submission_required",
            reason="Form submission ownership is outside the default EVS functionality lane and requires governed synthetic-submit controls.",
        )
    return OwnerRule(
        owner="evs",
        automation_status="ready_for_runner_mapping",
        side_effect_policy="no_submit",
        reason="Owned by EVS as browser-observable functionality with no intentional form submission.",
    )


def detect_assertion_type(page: str, section: str, description: str) -> str:
    raw = f"{page} {section} {description}".lower()
    if "unit types" in raw or "layouts correct" in raw:
        return "unit_types_and_layouts_match_pond"
    if "pricing accurate" in raw:
        return "pricing_matches_pond"
    if "guest card" in raw or "ah & eai" in raw:
        return "lead_attribution_guest_card_proof"
    if "arrow open and close" in raw or "opens and closes" in raw:
        return "toggle_open_close"
    if "see availability" in raw or "see available homes" in raw:
        return "route_to_apartments_pricing"
    if "contact us" in raw and "contact page" in raw:
        return "route_to_contact"
    if "apply now" in raw or "pipeline app" in raw or "pipeline application" in raw:
        return "external_handoff_pipeline_application"
    if "schedule a tour" in raw:
        return "external_handoff_schedule_tour"
    if "auto-scroll" in raw or "auto rotate" in raw or "carousel" in raw:
        return "carousel_behavior"
    if "filter" in raw:
        return "filter_behavior"
    if "availability displaying" in raw:
        return "rendered_availability_matches_pond"
    if "sorted" in raw:
        return "sort_order"
    if "map view" in raw or "change the floor" in raw:
        return "map_floor_unit_filter"
    if "unit number" in raw or "same unit" in raw:
        return "unit_detail_context_continuity"
    if "matterport" in raw or "virtual tour" in raw:
        return "external_handoff_matterport"
    if "sightmap" in raw or "apartment location" in raw:
        return "external_handoff_sightmap_unit"
    if "photos open" in raw or "camera icon" in raw or "images correct" in raw:
        return "media_modal_or_correctness"
    if "renting made simple" in raw:
        return "expanding_content_toggle"
    if "three buttons" in raw:
        return "required_cta_set_present"
    if "all-in pricing" in raw or "price quote" in raw or "get a quote" in raw:
        return "external_handoff_price_quote"
    if "map pin" in raw or "property location" in raw:
        return "map_pin_coordinate_match"
    if "reviews sorted" in raw:
        return "review_sort_order"
    if "contact form submit" in raw:
        return "form_submission"
    if "required field validation" in raw:
        return "required_field_validation"
    return "browser_functionality"


def detect_truth_sources(assertion_type: str) -> list[str]:
    sources = ["rendered_dom"]
    if assertion_type in {
        "rendered_availability_matches_pond",
        "unit_types_and_layouts_match_pond",
        "pricing_matches_pond",
    }:
        sources.append("pond_unit_availability")
    if assertion_type == "map_pin_coordinate_match":
        sources.append("thirtylines_feed_geo")
    if assertion_type in {
        "external_handoff_pipeline_application",
        "external_handoff_schedule_tour",
        "external_handoff_price_quote",
        "external_handoff_matterport",
        "external_handoff_sightmap_unit",
    }:
        sources.append("external_destination")
    if assertion_type == "lead_attribution_guest_card_proof":
        sources.extend(["synthetic_lead_identity", "ah_eai_guest_card"])
    if assertion_type == "review_sort_order":
        sources.append("review_dates")
    return sources


def detect_runner_profile(assertion_type: str, page: str, section: str) -> str:
    if assertion_type in {"form_submission", "required_field_validation"}:
        return "contact_form_checks"
    if page == "Apartments & Pricing" and section == "Unit Detail Page":
        return "apartments_pricing_deep_journey"
    if assertion_type in {
        "rendered_availability_matches_pond",
        "unit_types_and_layouts_match_pond",
        "pricing_matches_pond",
        "filter_behavior",
        "sort_order",
        "map_floor_unit_filter",
        "unit_detail_context_continuity",
        "external_handoff_sightmap_unit",
        "external_handoff_price_quote",
    }:
        return "apartments_pricing_deep_journey"
    if assertion_type == "lead_attribution_guest_card_proof":
        return "lead_attribution_e2e"
    return "portfolio_functionality_regression"


def load_contract_rows(workbook_path: Path, sheet_name: str) -> list[dict[str, Any]]:
    wb = load_workbook(workbook_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet_name!r} not found in {workbook_path}")
    ws = wb[sheet_name]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    required = ["Page", "Section", "Element", "Platform", "Description"]
    missing = [header for header in required if header not in headers]
    if missing:
        raise ValueError(f"Workbook is missing required columns: {missing}")

    rows: list[dict[str, Any]] = []
    for row_number in range(2, ws.max_row + 1):
        record = {headers[c - 1]: ws.cell(row_number, c).value for c in range(1, ws.max_column + 1)}
        element = str(record.get("Element") or "").strip()
        if element.lower() not in INCLUDED_ELEMENTS:
            continue
        page = str(record.get("Page") or "").strip()
        section = str(record.get("Section") or "").strip()
        platform = str(record.get("Platform") or "").strip()
        description = str(record.get("Description") or "").strip()
        if not page or not section or not description:
            continue
        assertion_type = detect_assertion_type(page, section, description)
        owner = detect_owner(page, section, description)
        check_id = f"qa_{row_number}_{slugify(page)}_{slugify(section)}_{slugify(assertion_type)}"
        rows.append(
            {
                "check_id": check_id,
                "source": {
                    "workbook": workbook_path.name,
                    "sheet": sheet_name,
                    "row": row_number,
                },
                "page": page,
                "section": section,
                "element": element,
                "platform": platform,
                "description": description,
                "owner": owner.owner,
                "owner_reason": owner.reason,
                "runner_profile": detect_runner_profile(assertion_type, page, section),
                "assertion_type": assertion_type,
                "truth_sources": detect_truth_sources(assertion_type),
                "device_profiles": ["desktop_chrome", "iphone_safari"],
                "side_effect_policy": owner.side_effect_policy,
                "automation_status": owner.automation_status,
                "severity": "high" if page in {"Apartments & Pricing", "Lead Attribution", "Contact"} else "medium",
            }
        )
    return rows


def build_contract(workbook_path: Path, sheet_name: str) -> dict[str, Any]:
    checks = load_contract_rows(workbook_path, sheet_name)
    owner_counts: dict[str, int] = {}
    for check in checks:
        owner_counts[check["owner"]] = owner_counts.get(check["owner"], 0) + 1
    return {
        "schema_version": "1.0",
        "contract_id": CONTRACT_ID,
        "contract_version": CONTRACT_VERSION,
        "name": "Portfolio Functionality and Data Integrity QA Contract",
        "status": "active_round_1_seed",
        "source": {
            "workbook": workbook_path.name,
            "sheet": sheet_name,
            "imported_at": date.today().isoformat(),
            "filter": "Element in (Functionality, Data Integrity)",
        },
        "default_devices": ["desktop_chrome", "iphone_safari"],
        "default_batch": "pilot_production_functionality",
        "ownership_model": {
            "evs": "Browser-observable functionality, routing, CTAs, availability/pricing/unit-data comparison, map coordinate checks, and no-submit vendor handoffs.",
            "media_qa": "Image, photo, virtual-tour, and media correctness checks retained in the contract but owned outside EVS.",
            "forms_qa": "Contact form submission and validation checks retained in the contract but owned outside EVS.",
            "lead_attribution_qa": "Guest-card proof in AH/EAI requires a governed synthetic-lead workflow before unattended automation.",
        },
        "summary": {
            "total_contract_checks": len(checks),
            "element_counts": {
                element: sum(1 for check in checks if check["element"] == element)
                for element in sorted({check["element"] for check in checks})
            },
            "owner_counts": owner_counts,
        },
        "checks": checks,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--sheet", default="Website QA Checklist")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    contract = build_contract(args.workbook, args.sheet)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(contract, indent=2) + "\n", encoding="utf-8")
    print(
        json.dumps(
            {
                "output": str(args.output),
                "checks": contract["summary"]["total_contract_checks"],
                "element_counts": contract["summary"]["element_counts"],
                "owner_counts": contract["summary"]["owner_counts"],
            },
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
