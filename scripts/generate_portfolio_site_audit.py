#!/usr/bin/env python3
"""
Generate Portfolio Site Audit Report
=====================================
Crawls all property websites (Resi + Legacy) for front-end quality issues:
  - Broken images
  - Unloaded virtual tours
  - JavaScript errors
  - Broken links
  - Missing meta / OG tags

Outputs a PIB-styled HTML report + XLSX, optionally emails via --email flag.

Usage:
    # Full portfolio audit
    python3 scripts/generate_portfolio_site_audit.py

    # Single property
    python3 scripts/generate_portfolio_site_audit.py --property "Camber Ridge"

    # Audit and email results
    python3 scripts/generate_portfolio_site_audit.py --email

    # Limit concurrency
    python3 scripts/generate_portfolio_site_audit.py --concurrency 5
"""

from __future__ import annotations

import argparse
import asyncio
import json
import sqlite3
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

ROOT = Path("/Users/mark/Property_Analytics")
REGISTRY_PATH = ROOT / "config" / "venterra_properties_official.json"
DB_PATH = ROOT / "data" / "portfolio_analytics.db"
OUT_DIR = ROOT / "reports" / "portfolio_site_audit"
OUT_DIR.mkdir(parents=True, exist_ok=True)
DELIVERY_DIR = ROOT / "logs" / "email_delivery"

sys.path.insert(0, str(ROOT / "utils"))
sys.path.insert(0, str(ROOT / "scripts"))

from email_sender import EmailSender  # noqa: E402
from report_builder import ReportBuilder, KPITile, Section, create_data_table  # noqa: E402
from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

from site_audit.crawler import audit_portfolio  # noqa: E402
from site_audit.models import PropertyAuditResult, AuditIssue  # noqa: E402

# Category display labels and colors
CATEGORY_META = {
    "broken_image": {"label": "Broken Images", "emoji": "🖼️", "color": "#dc3545"},
    "unloaded_tour": {"label": "Unloaded Tours", "emoji": "🎥", "color": "#9c27b0"},
    "js_error": {"label": "JS Errors", "emoji": "⚠️", "color": "#fd7e14"},
    "broken_link": {"label": "Broken Links", "emoji": "🔗", "color": "#0d6efd"},
    "missing_meta": {"label": "Missing Meta", "emoji": "🏷️", "color": "#6c757d"},
}

SEVERITY_STYLES = {
    "critical": {"bg": "#dc3545", "color": "#fff", "label": "Critical"},
    "warning": {"bg": "#ffc107", "color": "#1a1a1a", "label": "Warning"},
    "info": {"bg": "#17a2b8", "color": "#fff", "label": "Info"},
}


# ---------------------------------------------------------------------------
# Registry loading
# ---------------------------------------------------------------------------

def load_properties(filter_name: Optional[str] = None) -> List[dict]:
    """Load properties from the official registry, excluding new_development."""
    with open(REGISTRY_PATH) as f:
        registry = json.load(f)

    properties = []
    for prop in registry["properties"]:
        # Skip new_development properties (prelaunch, no live site to audit)
        if prop.get("property_type") == "new_development":
            continue
        if not prop.get("full_url"):
            continue
        if filter_name and prop["name"].lower() != filter_name.lower():
            continue
        properties.append(prop)

    return properties


# ---------------------------------------------------------------------------
# DB storage
# ---------------------------------------------------------------------------

def ensure_db_table():
    """Create the site_audit_results table if it doesn't exist."""
    conn = sqlite3.connect(DB_PATH)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS site_audit_results (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            audit_date TEXT NOT NULL,
            property_name TEXT NOT NULL,
            property_url TEXT NOT NULL,
            site_type TEXT NOT NULL,
            page_url TEXT NOT NULL,
            category TEXT NOT NULL,
            severity TEXT NOT NULL,
            description TEXT NOT NULL,
            detail TEXT DEFAULT '',
            location TEXT DEFAULT ''
        )
    """)
    conn.execute("""
        CREATE INDEX IF NOT EXISTS idx_site_audit_date
        ON site_audit_results (audit_date)
    """)
    conn.commit()
    conn.close()


def store_results(results: List[PropertyAuditResult], audit_date: str):
    """Store audit results in the database."""
    conn = sqlite3.connect(DB_PATH)
    rows = []
    for prop in results:
        for issue in prop.all_issues:
            rows.append((
                audit_date,
                prop.name,
                prop.base_url,
                prop.site_type,
                issue.page_url,
                issue.category,
                issue.severity,
                issue.description,
                issue.detail,
                issue.location,
            ))

    if rows:
        conn.executemany("""
            INSERT INTO site_audit_results
            (audit_date, property_name, property_url, site_type, page_url,
             category, severity, description, detail, location)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, rows)
        conn.commit()
    conn.close()
    return len(rows)


# ---------------------------------------------------------------------------
# HTML report rendering
# ---------------------------------------------------------------------------

def _severity_badge(severity: str) -> str:
    s = SEVERITY_STYLES.get(severity, SEVERITY_STYLES["info"])
    return (
        f'<span style="display:inline-block;padding:2px 8px;border-radius:3px;'
        f'font-size:11px;font-weight:600;background:{s["bg"]};color:{s["color"]};">'
        f'{s["label"]}</span>'
    )


def _category_bar(results: List[PropertyAuditResult]) -> str:
    """Render issue-count-by-category bar."""
    totals: Dict[str, int] = {}
    for r in results:
        for cat, count in r.issues_by_category().items():
            totals[cat] = totals.get(cat, 0) + count

    if not totals:
        return '<div style="padding:15px;color:#28a745;font-weight:600;">✅ No issues found across the portfolio!</div>'

    rows = ""
    for cat, meta in CATEGORY_META.items():
        count = totals.get(cat, 0)
        if count == 0:
            continue
        rows += f"""
        <tr>
            <td style="padding:8px 12px;font-size:13px;border-bottom:1px solid #e9ecef;">
                {meta['emoji']} {meta['label']}
            </td>
            <td style="padding:8px 12px;font-size:13px;border-bottom:1px solid #e9ecef;text-align:right;">
                <span style="font-weight:700;color:{meta['color']};">{count}</span>
            </td>
        </tr>
        """

    grand_total = sum(totals.values())
    return f"""
    <table cellpadding="0" cellspacing="0" border="0" style="width:100%;margin:15px 0;">
        <thead>
            <tr>
                <th style="text-align:left;padding:10px 12px;background:#f8f9fa;border-bottom:2px solid #dee2e6;
                    font-size:12px;color:#6c757d;font-weight:600;text-transform:uppercase;">Category</th>
                <th style="text-align:right;padding:10px 12px;background:#f8f9fa;border-bottom:2px solid #dee2e6;
                    font-size:12px;color:#6c757d;font-weight:600;text-transform:uppercase;">Count</th>
            </tr>
        </thead>
        <tbody>
            {rows}
            <tr>
                <td style="padding:10px 12px;font-size:13px;font-weight:700;">Total</td>
                <td style="padding:10px 12px;font-size:13px;font-weight:700;text-align:right;">{grand_total}</td>
            </tr>
        </tbody>
    </table>
    """


def _top_issues_table(results: List[PropertyAuditResult], limit: int = 10) -> str:
    """Top N properties by issue count."""
    sorted_props = sorted(results, key=lambda r: r.total_issues, reverse=True)
    top = [r for r in sorted_props if r.total_issues > 0][:limit]

    if not top:
        return ""

    rows = ""
    for r in top:
        status_color = "#dc3545" if r.critical_count > 0 else "#fd7e14" if r.warning_count > 0 else "#28a745"
        rows += f"""
        <tr>
            <td style="padding:8px 12px;font-size:13px;border-bottom:1px solid #e9ecef;">
                <span style="color:{status_color};font-weight:600;">●</span> {r.name}
                <span style="font-size:11px;color:#6c757d;margin-left:4px;">({r.site_type})</span>
            </td>
            <td style="padding:8px 12px;font-size:13px;border-bottom:1px solid #e9ecef;text-align:center;">
                {len(r.pages)}
            </td>
            <td style="padding:8px 12px;font-size:13px;border-bottom:1px solid #e9ecef;text-align:center;
                color:#dc3545;font-weight:600;">
                {r.critical_count}
            </td>
            <td style="padding:8px 12px;font-size:13px;border-bottom:1px solid #e9ecef;text-align:center;
                color:#fd7e14;font-weight:600;">
                {r.warning_count}
            </td>
            <td style="padding:8px 12px;font-size:13px;border-bottom:1px solid #e9ecef;text-align:right;
                font-weight:700;">
                {r.total_issues}
            </td>
        </tr>
        """

    return f"""
    <div style="margin-top:20px;">
        <div style="font-size:14px;font-weight:600;margin-bottom:10px;">Top Properties by Issue Count</div>
        <table cellpadding="0" cellspacing="0" border="0" style="width:100%;">
            <thead>
                <tr>
                    <th style="text-align:left;padding:10px 12px;background:#f8f9fa;border-bottom:2px solid #dee2e6;
                        font-size:12px;color:#6c757d;font-weight:600;">Property</th>
                    <th style="text-align:center;padding:10px 12px;background:#f8f9fa;border-bottom:2px solid #dee2e6;
                        font-size:12px;color:#6c757d;font-weight:600;">Pages</th>
                    <th style="text-align:center;padding:10px 12px;background:#f8f9fa;border-bottom:2px solid #dee2e6;
                        font-size:12px;color:#dc3545;font-weight:600;">Critical</th>
                    <th style="text-align:center;padding:10px 12px;background:#f8f9fa;border-bottom:2px solid #dee2e6;
                        font-size:12px;color:#fd7e14;font-weight:600;">Warning</th>
                    <th style="text-align:right;padding:10px 12px;background:#f8f9fa;border-bottom:2px solid #dee2e6;
                        font-size:12px;color:#6c757d;font-weight:600;">Total</th>
                </tr>
            </thead>
            <tbody>{rows}</tbody>
        </table>
    </div>
    """


def _critical_issues_detail(results: List[PropertyAuditResult]) -> str:
    """Render detailed critical issues grouped by property."""
    critical_props = [r for r in results if r.critical_count > 0]
    critical_props.sort(key=lambda r: r.critical_count, reverse=True)

    if not critical_props:
        return '<div style="padding:15px;color:#28a745;font-weight:600;">✅ No critical issues detected.</div>'

    html = ""
    for prop in critical_props:
        critical_issues = [i for i in prop.all_issues if i.severity == "critical"]
        issue_rows = ""
        for issue in critical_issues:
            cat_meta = CATEGORY_META.get(issue.category, {"label": issue.category, "emoji": "❓"})
            loc_badge = (
                f'<span style="display:inline-block;padding:1px 6px;border-radius:3px;'
                f'font-size:10px;font-weight:600;background:#e9ecef;color:#495057;'
                f'margin-right:4px;">{issue.location}</span>'
            ) if issue.location else ""
            issue_rows += f"""
            <tr>
                <td style="padding:6px 10px;font-size:12px;border-bottom:1px solid #f0f0f0;white-space:nowrap;">
                    {cat_meta['emoji']} {cat_meta['label']}
                </td>
                <td style="padding:6px 10px;font-size:12px;border-bottom:1px solid #f0f0f0;">
                    {loc_badge}{issue.description}
                </td>
                <td style="padding:6px 10px;font-size:11px;color:#6c757d;border-bottom:1px solid #f0f0f0;
                    word-break:break-all;">
                    {issue.detail}
                </td>
            </tr>
            """

        html += f"""
        <div style="margin-bottom:16px;padding:12px;background:#fff;border-left:4px solid #dc3545;border-radius:4px;">
            <div style="font-weight:600;font-size:14px;margin-bottom:8px;">
                {prop.name}
                <span style="font-size:11px;color:#6c757d;margin-left:6px;">
                    ({prop.site_type}) — {prop.critical_count} critical issue(s)
                </span>
            </div>
            <table cellpadding="0" cellspacing="0" border="0" style="width:100%;font-size:12px;">
                {issue_rows}
            </table>
        </div>
        """

    return html


def _property_detail_section(results: List[PropertyAuditResult]) -> str:
    """Full property listing with pass/fail and per-category counts."""
    sorted_results = sorted(results, key=lambda r: (-r.critical_count, -r.total_issues, r.name))

    html = ""
    for prop in sorted_results:
        if prop.total_issues == 0:
            border_color = "#28a745"
            score_html = '<span style="font-size:28px;font-weight:700;color:#28a745;">✓</span>'
        elif prop.critical_count > 0:
            border_color = "#dc3545"
            score_html = f'<span style="font-size:28px;font-weight:700;color:#dc3545;">{prop.total_issues}</span>'
        else:
            border_color = "#fd7e14"
            score_html = f'<span style="font-size:28px;font-weight:700;color:#fd7e14;">{prop.total_issues}</span>'

        cats = prop.issues_by_category()
        cat_chips = ""
        for cat, meta in CATEGORY_META.items():
            count = cats.get(cat, 0)
            if count > 0:
                cat_chips += (
                    f'<span style="display:inline-block;padding:2px 8px;margin:2px 4px 2px 0;'
                    f'border-radius:3px;font-size:11px;background:#f8f9fa;border:1px solid #dee2e6;">'
                    f'{meta["emoji"]} {count}</span>'
                )

        pages_scanned = len(prop.pages)
        html += f"""
        <div style="padding:12px;background:#fff;border-left:4px solid {border_color};margin-bottom:10px;border-radius:4px;">
            <div style="display:flex;justify-content:space-between;align-items:center;">
                <div style="flex:1;">
                    <div style="font-weight:600;font-size:14px;color:#333;margin-bottom:4px;">{prop.name}</div>
                    <div style="font-size:11px;color:#6c757d;margin-bottom:6px;">
                        {prop.site_type.upper()} · {pages_scanned} pages · {prop.base_url}
                    </div>
                    <div>{cat_chips}</div>
                </div>
                <div style="margin-left:15px;text-align:center;">{score_html}</div>
            </div>
        </div>
        """

    return html


def build_html(results: List[PropertyAuditResult], elapsed_seconds: float) -> str:
    """Build the full PIB-styled HTML report."""
    total_properties = len(results)
    total_issues = sum(r.total_issues for r in results)
    total_critical = sum(r.critical_count for r in results)
    total_pages = sum(len(r.pages) for r in results)
    pass_count = sum(1 for r in results if r.critical_count == 0)
    pass_rate = (pass_count / total_properties * 100) if total_properties else 0

    builder = ReportBuilder(
        title="Portfolio Site Audit",
        subtitle="Front-End Quality Assessment",
        version="1.0",
        date_range=datetime.now().strftime("%Y-%m-%d"),
    )

    # KPI tiles
    builder.add_kpi_tiles([
        KPITile(
            label="Properties Scanned",
            value=str(total_properties),
            sublabel=f"{total_pages} total pages",
        ),
        KPITile(
            label="Issues Found",
            value=str(total_issues),
            trend=f"↑ {total_critical} critical" if total_critical > 0 else None,
            sublabel=f"Scan time: {elapsed_seconds:.0f}s",
        ),
        KPITile(
            label="Pass Rate",
            value=f"{pass_rate:.0f}%",
            sublabel=f"{pass_count}/{total_properties} with 0 critical issues",
            is_primary=True,
        ),
    ])

    # Section 1: Portfolio Overview
    overview_html = _category_bar(results) + _top_issues_table(results)
    has_critical = total_critical > 0
    builder.add_section(Section(
        title="Portfolio Overview",
        content=overview_html,
        status="action_needed" if has_critical else "healthy",
        description=f"Issue distribution across {total_properties} properties ({total_pages} pages scanned)",
    ))

    # Section 2: Critical Issues
    builder.add_section(Section(
        title="Critical Issues",
        content=_critical_issues_detail(results),
        status="action_needed" if has_critical else "healthy",
        description="Properties with critical-severity findings requiring immediate attention",
    ))

    # Section 3: Full Property Detail
    builder.add_section(Section(
        title="Full Property Detail",
        content=_property_detail_section(results),
        status="action_needed" if has_critical else "healthy",
        description="All properties sorted by severity, with per-category issue counts",
    ))

    html = builder.generate()

    # Add MarketingOps tagline (same pattern as PSI/PIB report)
    tagline = (
        '<div style="text-align:center;margin:-8px 0 18px 0;">'
        '<span style="font-size:9px;color:#adb5bd;vertical-align:middle;">Powered by</span> '
        '<span style="font-size:12px;color:#495057;font-weight:600;vertical-align:middle;">MarketingOps</span>'
        '</div>'
    )
    subtitle_marker = "Front-End Quality Assessment</div>"
    html = html.replace(subtitle_marker, subtitle_marker + tagline, 1)

    return html


# ---------------------------------------------------------------------------
# XLSX generation
# ---------------------------------------------------------------------------

def _autosize(ws):
    for column in ws.columns:
        max_len = 0
        col_letter = get_column_letter(column[0].column)
        for cell in column:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)


def write_excel(path: Path, results: List[PropertyAuditResult]):
    """Generate XLSX with Summary, Property Detail, and All Issues sheets."""
    wb = Workbook()

    # --- Sheet 1: Summary ---
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Portfolio Site Audit"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Audit Date"
    ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws["A4"] = "Properties Scanned"
    ws["B4"] = len(results)
    ws["A5"] = "Total Pages"
    ws["B5"] = sum(len(r.pages) for r in results)
    ws["A6"] = "Total Issues"
    ws["B6"] = sum(r.total_issues for r in results)

    row = 8
    ws[f"A{row}"] = "Category"
    ws[f"B{row}"] = "Count"
    ws[f"A{row}"].font = Font(bold=True)
    ws[f"B{row}"].font = Font(bold=True)
    totals: Dict[str, int] = {}
    for r in results:
        for cat, count in r.issues_by_category().items():
            totals[cat] = totals.get(cat, 0) + count
    for cat, meta in CATEGORY_META.items():
        row += 1
        ws[f"A{row}"] = meta["label"]
        ws[f"B{row}"] = totals.get(cat, 0)
    _autosize(ws)

    # --- Sheet 2: Property Detail ---
    ws2 = wb.create_sheet("Property Detail")
    headers = ["Property", "Site Type", "URL", "Pages", "Critical", "Warning", "Info", "Total"]
    ws2.append(headers)
    for cell in ws2[1]:
        cell.font = Font(bold=True)

    for r in sorted(results, key=lambda x: (-x.critical_count, -x.total_issues)):
        ws2.append([
            r.name, r.site_type, r.base_url, len(r.pages),
            r.critical_count, r.warning_count, r.info_count, r.total_issues,
        ])
    # Conditional formatting on critical column
    for row_idx in range(2, ws2.max_row + 1):
        cell = ws2[f"E{row_idx}"]
        if cell.value and cell.value > 0:
            cell.fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
    ws2.freeze_panes = "A2"
    _autosize(ws2)

    # --- Sheet 3: All Issues ---
    ws3 = wb.create_sheet("All Issues")
    headers = ["Property", "Site Type", "Page URL", "Location", "Category", "Severity", "Description", "Detail"]
    ws3.append(headers)
    for cell in ws3[1]:
        cell.font = Font(bold=True)

    for r in results:
        for issue in r.all_issues:
            cat_label = CATEGORY_META.get(issue.category, {}).get("label", issue.category)
            ws3.append([
                r.name, r.site_type, issue.page_url, issue.location,
                cat_label, issue.severity, issue.description, issue.detail,
            ])
    ws3.freeze_panes = "A2"
    _autosize(ws3)

    wb.save(path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def generate_report(
    send_email: bool = False,
    property_filter: Optional[str] = None,
    max_concurrent: int = 3,
) -> Path:
    """Run the full audit pipeline and generate report."""
    print("=" * 80)
    print("PORTFOLIO SITE AUDIT")
    print("=" * 80)
    print()

    # Load properties
    properties = load_properties(filter_name=property_filter)
    if not properties:
        if property_filter:
            print(f"❌ No property found matching \"{property_filter}\"")
        else:
            print("❌ No properties loaded from registry")
        sys.exit(1)

    resi_count = sum(1 for p in properties if p.get("site_type") == "resi")
    legacy_count = len(properties) - resi_count
    print(f"📋 Loaded {len(properties)} properties ({resi_count} Resi, {legacy_count} Legacy)")
    print(f"⚡ Concurrency: {max_concurrent}")
    print()

    # Ensure DB table
    ensure_db_table()

    # Run audit
    print("🔍 Starting crawl...")
    start_time = time.monotonic()
    results = asyncio.run(audit_portfolio(properties, max_concurrent=max_concurrent))
    elapsed = time.monotonic() - start_time
    print()
    print(f"✅ Crawl complete in {elapsed:.1f}s")

    total_issues = sum(r.total_issues for r in results)
    total_critical = sum(r.critical_count for r in results)
    print(f"   Issues found: {total_issues} ({total_critical} critical)")
    print()

    # Store in DB
    audit_date = datetime.now().strftime("%Y-%m-%d")
    stored = store_results(results, audit_date)
    print(f"💾 Stored {stored} issue records in database")

    # Generate HTML
    print("📝 Generating HTML report...")
    html = build_html(results, elapsed)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    html_path = OUT_DIR / f"portfolio_site_audit_{stamp}.html"
    html_path.write_text(html, encoding="utf-8")
    print(f"   → {html_path}")

    # Generate XLSX
    print("📊 Generating XLSX...")
    xlsx_path = OUT_DIR / f"portfolio_site_audit_{stamp}.xlsx"
    write_excel(xlsx_path, results)
    print(f"   → {xlsx_path}")

    # Email
    if send_email:
        print()
        print("📧 Sending email...")
        sender = EmailSender(verbose=False)
        log_path = DELIVERY_DIR / f"email_delivery_{datetime.now().strftime('%Y-%m-%d')}.jsonl"
        attachments = [(
            xlsx_path.name,
            xlsx_path.read_bytes(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )]
        sender.send_email_with_tracking(
            subject=f"Venterra Portfolio Site Audit — {datetime.now().strftime('%m-%d-%Y')}",
            html_body=html,
            attachments=attachments,
            log_path=log_path,
        )

    print()
    print("✅ Done!")
    return html_path


def main() -> int:
    parser = argparse.ArgumentParser(description="Portfolio Site Audit — detect front-end quality issues")
    parser.add_argument("--email", action="store_true", help="Email the report after generation")
    parser.add_argument("--property", type=str, default=None, help="Audit a single property by name")
    parser.add_argument("--concurrency", type=int, default=3, help="Max concurrent browser contexts (default: 3)")
    args = parser.parse_args()

    generate_report(
        send_email=args.email,
        property_filter=args.property,
        max_concurrent=args.concurrency,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
