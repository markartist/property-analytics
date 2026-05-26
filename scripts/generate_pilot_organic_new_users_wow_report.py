#!/usr/bin/env python3
"""Generate a pilot vs sister weekly organic new-user traffic report."""

from __future__ import annotations

import json
import sqlite3
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Dict, Iterable, List, Sequence

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path("/Users/mark/Property_Analytics")
DB_PATH = ROOT / "data" / "portfolio_analytics.db"
PAIR_PATH = ROOT / "apps" / "web" / "public" / "pilot-kpi" / "latest" / "properties.json"
OUT_DIR = ROOT / "reports" / "pilot_organic_new_users_wow"
START_DATE = "2026-01-01"

import sys

sys.path.insert(0, str(ROOT / "utils"))
from pib_email_shell import wrap_pib_light_email  # noqa: E402
from report_builder import KPITile, ReportBuilder, Section, create_data_table  # noqa: E402


@dataclass(frozen=True)
class PairMember:
    pair_name: str
    role: str
    property_id: str
    property_name: str


def load_pair_members() -> List[PairMember]:
    payload = json.loads(PAIR_PATH.read_text())
    members: List[PairMember] = []
    for pair in payload["pairs"]:
        pilot = pair["identity"]["pilot"]
        sister = pair["identity"]["sister"]
        pair_name = f"{pilot['name']} vs {sister['name']}"
        members.append(PairMember(pair_name, "pilot", pilot["property_id"], pilot["name"]))
        members.append(PairMember(pair_name, "sister", sister["property_id"], sister["name"]))
    return members


def cohort_property_labels(members: Sequence[PairMember]) -> Dict[str, str]:
    grouped: Dict[str, List[str]] = {"pilot": [], "sister": []}
    for member in members:
        grouped[member.role].append(member.property_name)
    return {
        role: ", ".join(grouped[role])
        for role in grouped
    }


def build_member_values_sql(members: Sequence[PairMember]) -> str:
    values = []
    for member in members:
        escaped_pair = member.pair_name.replace("'", "''")
        escaped_role = member.role.replace("'", "''")
        escaped_id = member.property_id.replace("'", "''")
        escaped_name = member.property_name.replace("'", "''")
        values.append(f"('{escaped_pair}','{escaped_role}','{escaped_id}','{escaped_name}')")
    return ",\n                ".join(values)


def get_latest_complete_week_end(conn: sqlite3.Connection, members: Sequence[PairMember]) -> str:
    property_ids = tuple(member.property_id for member in members)
    placeholders = ",".join("?" for _ in property_ids)
    max_metric_date = conn.execute(
        f"""
        SELECT MAX(metric_date)
        FROM ga4_daily_metrics
        WHERE property_id IN ({placeholders})
          AND metric_date >= ?
        """,
        property_ids + (START_DATE,),
    ).fetchone()[0]
    if not max_metric_date:
        raise SystemExit("No GA4 rows found for the requested period.")

    max_date = datetime.strptime(str(max_metric_date), "%Y-%m-%d").date()
    latest_complete_week_end = max_date - timedelta(days=(max_date.weekday() + 1) % 7)
    return latest_complete_week_end.isoformat()


def fetch_weekly_rows(
    conn: sqlite3.Connection,
    members: Sequence[PairMember],
    latest_complete_week_end: str,
) -> List[sqlite3.Row]:
    member_values = build_member_values_sql(members)
    sql = f"""
        WITH pair_members(pair_name, role, property_id, property_name) AS (
            VALUES
                {member_values}
        ),
        daily AS (
            SELECT
                pm.pair_name,
                pm.role,
                pm.property_id,
                pm.property_name,
                d.metric_date,
                date(d.metric_date, '-' || ((CAST(strftime('%w', d.metric_date) AS INTEGER) + 6) % 7) || ' days') AS week_start,
                d.sessions AS total_sessions,
                d.new_users AS total_new_users,
                COALESCE(org.organic_sessions, 0) AS organic_sessions,
                COALESCE(org.organic_total_users, 0) AS organic_total_users,
                COALESCE(org.organic_new_users, 0) AS organic_new_users
            FROM pair_members pm
            JOIN ga4_daily_metrics d
              ON d.property_id = pm.property_id
            LEFT JOIN (
                SELECT
                    property_id,
                    metric_date,
                    SUM(sessions) AS organic_sessions,
                    SUM(total_users) AS organic_total_users,
                    SUM(new_users) AS organic_new_users
                FROM ga4_traffic_sources
                WHERE channel_group = 'Organic Search'
                GROUP BY property_id, metric_date
            ) org
              ON org.property_id = d.property_id
             AND org.metric_date = d.metric_date
            WHERE d.metric_date >= ?
              AND d.metric_date <= ?
        )
        SELECT
            pair_name,
            role,
            property_id,
            property_name,
            week_start,
            MIN(metric_date) AS period_start,
            MAX(metric_date) AS period_end,
            COUNT(*) AS days_in_period,
            SUM(total_sessions) AS total_sessions,
            SUM(total_new_users) AS total_new_users,
            SUM(organic_sessions) AS organic_sessions,
            SUM(organic_total_users) AS organic_total_users,
            SUM(organic_new_users) AS organic_new_users,
            ROUND(
                CASE
                    WHEN SUM(total_sessions) > 0 THEN CAST(SUM(organic_sessions) AS REAL) / SUM(total_sessions)
                    ELSE 0
                END,
                4
            ) AS organic_session_share
        FROM daily
        GROUP BY pair_name, role, property_id, property_name, week_start
        ORDER BY week_start, pair_name, role
    """
    return conn.execute(sql, (START_DATE, latest_complete_week_end)).fetchall()


def add_wow(rows: List[Dict[str, object]], value_key: str) -> None:
    grouped: Dict[str, List[Dict[str, object]]] = defaultdict(list)
    for row in rows:
        grouped[str(row["series_key"])].append(row)

    for series_rows in grouped.values():
        series_rows.sort(key=lambda item: str(item["week_start"]))
        previous = None
        for row in series_rows:
            current = float(row[value_key]) if row[value_key] is not None else None
            row["prev_week_value"] = previous
            row["wow_abs"] = None if current is None or previous in (None, 0) and previous is None else (current - previous if previous is not None else None)
            row["wow_pct"] = None
            if current is not None and previous not in (None, 0):
                row["wow_pct"] = (current - previous) / previous
            previous = current


def build_property_weekly_lookup(weekly_rows: Iterable[sqlite3.Row]) -> Dict[tuple[str, str], Dict[str, object]]:
    rows: List[Dict[str, object]] = []
    for row in weekly_rows:
        rows.append(
            {
                "series_key": str(row["property_id"]),
                "property_id": str(row["property_id"]),
                "week_start": str(row["week_start"]),
                "organic_new_users": float(row["organic_new_users"] or 0),
            }
        )
    add_wow(rows, "organic_new_users")
    return {(row["property_id"], row["week_start"]): row for row in rows}


def build_pair_rows(weekly_rows: Iterable[sqlite3.Row]) -> List[Dict[str, object]]:
    grouped: Dict[tuple[str, str], Dict[str, object]] = {}
    for row in weekly_rows:
        key = (row["pair_name"], row["week_start"])
        bucket = grouped.setdefault(
            key,
            {
                "pair_name": row["pair_name"],
                "week_start": row["week_start"],
                "period_start": row["period_start"],
                "period_end": row["period_end"],
                "days_in_period": row["days_in_period"],
            },
        )
        prefix = row["role"]
        bucket[f"{prefix}_property_name"] = row["property_name"]
        bucket[f"{prefix}_property_id"] = row["property_id"]
        bucket[f"{prefix}_organic_new_users"] = row["organic_new_users"]
        bucket[f"{prefix}_organic_session_share"] = row["organic_session_share"]
        bucket[f"{prefix}_organic_sessions"] = row["organic_sessions"]
        bucket[f"{prefix}_organic_total_users"] = row["organic_total_users"]
        bucket[f"{prefix}_total_new_users"] = row["total_new_users"]
        bucket[f"{prefix}_total_sessions"] = row["total_sessions"]

    pair_rows: List[Dict[str, object]] = []
    for row in grouped.values():
        pilot_val = float(row.get("pilot_organic_new_users", 0) or 0)
        sister_val = float(row.get("sister_organic_new_users", 0) or 0)
        row["gap_abs"] = round(pilot_val - sister_val, 2)
        row["gap_pct_vs_sister"] = None if sister_val == 0 else (pilot_val - sister_val) / sister_val
        row["series_key"] = row["pair_name"]
        row["organic_new_users"] = pilot_val + sister_val
        pair_rows.append(row)
    add_wow(pair_rows, "organic_new_users")
    return sorted(pair_rows, key=lambda item: (str(item["week_start"]), str(item["pair_name"])))


def build_rollup_rows(weekly_rows: Iterable[sqlite3.Row]) -> List[Dict[str, object]]:
    grouped: Dict[tuple[str, str], Dict[str, object]] = {}
    for row in weekly_rows:
        key = (row["role"], row["week_start"])
        bucket = grouped.setdefault(
            key,
            {
                "role": row["role"],
                "week_start": row["week_start"],
                "period_start": row["period_start"],
                "period_end": row["period_end"],
                "days_in_period": row["days_in_period"],
                "pair_count": 0,
                "organic_new_users": 0.0,
                "organic_sessions": 0,
                "organic_total_users": 0,
                "total_new_users": 0,
                "total_sessions": 0,
            },
        )
        bucket["pair_count"] += 1
        bucket["organic_new_users"] += float(row["organic_new_users"] or 0)
        bucket["organic_sessions"] += int(row["organic_sessions"] or 0)
        bucket["organic_total_users"] += int(row["organic_total_users"] or 0)
        bucket["total_new_users"] += int(row["total_new_users"] or 0)
        bucket["total_sessions"] += int(row["total_sessions"] or 0)

    rollup_rows = sorted(grouped.values(), key=lambda item: (str(item["role"]), str(item["week_start"])))
    for row in rollup_rows:
        row["organic_new_users"] = round(float(row["organic_new_users"]), 2)
        row["organic_session_share"] = round(
            (float(row["organic_sessions"]) / float(row["total_sessions"])) if row["total_sessions"] else 0,
            4,
        )
        row["series_key"] = row["role"]
    add_wow(rollup_rows, "organic_new_users")
    return rollup_rows


def latest_rollup_by_role(rollup_rows: Sequence[Dict[str, object]]) -> Dict[str, Dict[str, object]]:
    latest: Dict[str, Dict[str, object]] = {}
    for row in sorted(rollup_rows, key=lambda item: str(item["week_start"])):
        latest[str(row["role"])] = row
    return latest


def latest_pairs(pair_rows: Sequence[Dict[str, object]]) -> List[Dict[str, object]]:
    latest_week = max(str(row["week_start"]) for row in pair_rows)
    latest = [row for row in pair_rows if str(row["week_start"]) == latest_week]
    latest.sort(key=lambda item: float(item["gap_abs"]), reverse=True)
    return latest


def build_cohort_context_rows(rollup_rows: Sequence[Dict[str, object]], limit: int = 4) -> List[List[object]]:
    by_week: Dict[str, Dict[str, Dict[str, object]]] = defaultdict(dict)
    for row in rollup_rows:
        by_week[str(row["week_start"])][str(row["role"])] = row

    rows: List[List[object]] = []
    for week_start in sorted(by_week.keys(), reverse=True)[:limit]:
        pilot = by_week[week_start].get("pilot", {})
        sister = by_week[week_start].get("sister", {})
        pilot_total = float(pilot.get("organic_new_users", 0) or 0)
        sister_total = float(sister.get("organic_new_users", 0) or 0)
        gap_abs = pilot_total - sister_total
        gap_pct = None if sister_total == 0 else gap_abs / sister_total
        rows.append(
            [
                week_start,
                fmt_num(pilot_total, 1),
                trend_html(pilot.get("wow_pct"), "WoW"),
                fmt_num(sister_total, 1),
                trend_html(sister.get("wow_pct"), "WoW"),
                fmt_num(gap_abs, 1),
                trend_html(gap_pct, "vs sister"),
            ]
        )
    return rows


def fmt_num(value: object, decimals: int = 0) -> str:
    if value is None:
        return "—"
    return f"{float(value):,.{decimals}f}"


def fmt_pct(value: object, decimals: int = 1) -> str:
    if value is None:
        return "—"
    return f"{float(value) * 100:+,.{decimals}f}%"


def fmt_pct_plain(value: object, decimals: int = 1) -> str:
    if value is None:
        return "—"
    return f"{float(value) * 100:,.{decimals}f}%"


def trend_html(value: object, suffix: str = "", decimals: int = 1) -> str:
    if value is None:
        return '<span style="color:#94a3b8;font-weight:600;">→ —</span>'
    numeric = float(value)
    if numeric > 0:
        arrow = "↑"
        color = "#16a34a"
    elif numeric < 0:
        arrow = "↓"
        color = "#dc2626"
    else:
        arrow = "→"
        color = "#64748b"
    label = fmt_pct(numeric, decimals)
    suffix_html = f" {suffix}" if suffix else ""
    return f'<span style="color:{color};font-weight:700;white-space:nowrap;">{arrow} {label}{suffix_html}</span>'


def autosize(ws) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 38)


def write_sheet(ws, headers: Sequence[str], rows: Sequence[Sequence[object]]) -> None:
    header_fill = PatternFill("solid", fgColor="15284B")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    ws.freeze_panes = "A2"
    autosize(ws)


def build_workbook(
    weekly_rows: Sequence[sqlite3.Row],
    pair_rows: Sequence[Dict[str, object]],
    rollup_rows: Sequence[Dict[str, object]],
    latest_complete_week_end: str,
    out_path: Path,
) -> None:
    wb = Workbook()

    latest_rollups = latest_rollup_by_role(rollup_rows)
    latest_pair_rows = latest_pairs(pair_rows)
    latest_week = latest_rollups["pilot"]["week_start"]
    property_weekly_lookup = build_property_weekly_lookup(weekly_rows)

    summary = wb.active
    summary.title = "Latest Week View"
    summary["A1"] = "Pilot Organic New Users WoW"
    summary["A1"].font = Font(bold=True, size=15)
    summary["A2"] = "Pilot and sister property pairs"
    summary["A3"] = "Date Range"
    summary["B3"] = f"{START_DATE} through {latest_complete_week_end}"
    summary["A4"] = "Latest Complete Week"
    summary["B4"] = f"{latest_week} to {latest_rollups['pilot']['period_end']}"
    summary["A5"] = "Metric"
    summary["B5"] = "Exact GA4 Organic Search new users"

    for cell in ("A3", "A4", "A5"):
        summary[cell].font = Font(bold=True)

    header_row = 8
    executive_headers = [
        "Pair",
        "Pilot Property",
        "Pilot Organic New Users",
        "Pilot WoW %",
        "Sister Property",
        "Sister Organic New Users",
        "Sister WoW %",
        "Pilot Minus Sister",
        "Gap vs Sister %",
        "Pair WoW %",
    ]
    for col_idx, header in enumerate(executive_headers, start=1):
        cell = summary.cell(row=header_row, column=col_idx, value=header)
        cell.fill = PatternFill("solid", fgColor="15284B")
        cell.font = Font(color="FFFFFF", bold=True)

    pair_lookup = {
        (str(row["pair_name"]), str(row["week_start"])): row
        for row in pair_rows
    }
    for row_idx, row in enumerate(latest_pair_rows, start=header_row + 1):
        prev_row = pair_lookup.get((str(row["pair_name"]), "2026-03-23"))
        # The pair-level WoW is already on row; pilot/sister WoW lives in weekly_rows-derived data below.
        summary.cell(row=row_idx, column=1, value=row["pair_name"])
        summary.cell(row=row_idx, column=2, value=row["pilot_property_name"])
        summary.cell(row=row_idx, column=3, value=float(row["pilot_organic_new_users"]))
        summary.cell(
            row=row_idx,
            column=4,
            value=(
                None
                if property_weekly_lookup[(str(row["pilot_property_id"]), str(row["week_start"]))]["wow_pct"] is None
                else float(property_weekly_lookup[(str(row["pilot_property_id"]), str(row["week_start"]))]["wow_pct"])
            ),
        )
        summary.cell(row=row_idx, column=5, value=row["sister_property_name"])
        summary.cell(row=row_idx, column=6, value=float(row["sister_organic_new_users"]))
        summary.cell(
            row=row_idx,
            column=7,
            value=(
                None
                if property_weekly_lookup[(str(row["sister_property_id"]), str(row["week_start"]))]["wow_pct"] is None
                else float(property_weekly_lookup[(str(row["sister_property_id"]), str(row["week_start"]))]["wow_pct"])
            ),
        )
        summary.cell(row=row_idx, column=8, value=float(row["gap_abs"]))
        summary.cell(row=row_idx, column=9, value=None if row["gap_pct_vs_sister"] is None else float(row["gap_pct_vs_sister"]))
        summary.cell(row=row_idx, column=10, value=None if row["wow_pct"] is None else float(row["wow_pct"]))

    summary.freeze_panes = "A8"
    autosize(summary)

    rollup_ws = wb.create_sheet("Cohort Weekly Rollup")
    summary_rows = [
        [
            row["role"].title(),
            row["week_start"],
            row["period_start"],
            row["period_end"],
            row["days_in_period"],
            row["pair_count"],
            row["organic_new_users"],
            row["prev_week_value"],
            row["wow_abs"],
            row["wow_pct"],
            row["organic_session_share"],
            row["organic_sessions"],
            row["organic_total_users"],
            row["total_new_users"],
            row["total_sessions"],
        ]
        for row in rollup_rows
    ]
    write_sheet(
        rollup_ws,
        [
            "Role",
            "Week Start",
            "Period Start",
            "Period End",
            "Days In Period",
            "Pair Count",
            "Organic New Users",
            "Prior Week",
            "WoW Abs",
            "WoW %",
            "Organic Session Share",
            "Organic Sessions",
            "Organic Total Users",
            "Total New Users",
            "Total Sessions",
        ],
        summary_rows,
    )

    pair_ws = wb.create_sheet("Pair Detail")
    pair_sheet_rows = [
        [
            row["pair_name"],
            row["week_start"],
            row["period_start"],
            row["period_end"],
            row["pilot_property_name"],
            row["pilot_organic_new_users"],
            row["sister_property_name"],
            row["sister_organic_new_users"],
            row["gap_abs"],
            row["gap_pct_vs_sister"],
            row["organic_new_users"],
            row["prev_week_value"],
            row["wow_abs"],
            row["wow_pct"],
        ]
        for row in pair_rows
    ]
    write_sheet(
        pair_ws,
        [
            "Pair",
            "Week Start",
            "Period Start",
            "Period End",
            "Pilot Property",
            "Pilot Organic New Users",
            "Sister Property",
            "Sister Organic New Users",
            "Pilot Minus Sister",
            "Gap vs Sister %",
            "Pair Total",
            "Prior Week Pair Total",
            "WoW Abs",
            "WoW %",
        ],
        pair_sheet_rows,
    )

    detail_ws = wb.create_sheet("Property Weekly Detail")
    detail_rows = [
        [
            row["pair_name"],
            row["role"].title(),
            row["property_name"],
            row["property_id"],
            row["week_start"],
            row["period_start"],
            row["period_end"],
            row["days_in_period"],
            row["organic_new_users"],
            row["organic_session_share"],
            row["organic_sessions"],
            row["organic_total_users"],
            row["total_new_users"],
            row["total_sessions"],
        ]
        for row in weekly_rows
    ]
    write_sheet(
        detail_ws,
        [
            "Pair",
            "Role",
            "Property",
            "Property ID",
            "Week Start",
            "Period Start",
            "Period End",
            "Days In Period",
            "Organic New Users",
            "Organic Session Share",
            "Organic Sessions",
            "Organic Total Users",
            "Total New Users",
            "Total Sessions",
        ],
        detail_rows,
    )

    notes_ws = wb.create_sheet("Methodology")
    notes = [
        ["Metric", "Exact GA4 Organic Search new users"],
        [
            "Definition",
            "Daily GA4 newUsers for channel_group = Organic Search, summed into Monday-start weekly windows.",
        ],
        ["Formula", "organic_new_users = SUM(ga4_traffic_sources.new_users WHERE channel_group = 'Organic Search')"],
        ["Date Range", f"{START_DATE} through latest complete week ending {latest_complete_week_end}"],
        ["Week Logic", "Weeks are grouped by Monday start. The first week is partial because the request starts on 2026-01-01."],
        ["Source Tables", "ga4_daily_metrics, ga4_traffic_sources, apps/web/public/pilot-kpi/latest/properties.json"],
    ]
    write_sheet(notes_ws, ["Field", "Value"], notes)

    for ws in (summary, rollup_ws, pair_ws, detail_ws, notes_ws):
        autosize(ws)

    wb.save(out_path)


def build_email_html(
    rollup_rows: Sequence[Dict[str, object]],
    pair_rows: Sequence[Dict[str, object]],
    generated_at: datetime,
    latest_complete_week_end: str,
    cohort_labels: Dict[str, str],
) -> str:
    latest = latest_rollup_by_role(rollup_rows)
    pilot = latest["pilot"]
    sister = latest["sister"]
    latest_week = str(pilot["week_start"])
    gap_abs = float(pilot["organic_new_users"]) - float(sister["organic_new_users"])
    gap_pct = None if float(sister["organic_new_users"]) == 0 else gap_abs / float(sister["organic_new_users"])
    pair_items = latest_pairs(pair_rows)

    builder = ReportBuilder(
        title="Pilot Organic New Users WoW",
        subtitle="Pilot and sister property pairs",
        version="1.0",
        date_range=f"{START_DATE} to {latest_complete_week_end}",
    )
    builder.add_kpi_tiles(
        [
            KPITile(
                label="Pilot Cohort",
                value=fmt_num(pilot["organic_new_users"], 1),
                trend=f"{'↑' if (pilot['wow_pct'] or 0) > 0 else '↓' if (pilot['wow_pct'] or 0) < 0 else '→'} {fmt_pct(pilot['wow_pct'])} WoW",
                sublabel="Organic Search new users",
                value_color_override="#1a1a1a",
                trend_color_override="#16a34a" if (pilot["wow_pct"] or 0) > 0 else "#dc2626" if (pilot["wow_pct"] or 0) < 0 else "#64748b",
            ),
            KPITile(
                label="Sister Cohort",
                value=fmt_num(sister["organic_new_users"], 1),
                trend=f"{'↑' if (sister['wow_pct'] or 0) > 0 else '↓' if (sister['wow_pct'] or 0) < 0 else '→'} {fmt_pct(sister['wow_pct'])} WoW",
                sublabel="Organic Search new users",
                value_color_override="#1a1a1a",
                trend_color_override="#16a34a" if (sister["wow_pct"] or 0) > 0 else "#dc2626" if (sister["wow_pct"] or 0) < 0 else "#64748b",
            ),
            KPITile(
                label="Pilot vs Sister Gap",
                value=fmt_num(gap_abs, 1),
                trend=(
                    f"{'↑' if gap_pct > 0 else '↓' if gap_pct < 0 else '→'} {fmt_pct(gap_pct)} vs sister"
                    if gap_pct is not None
                    else None
                ),
                sublabel="Positive means pilot leads",
                value_color_override="#1a1a1a",
                trend_color_override="#16a34a" if gap_pct is not None and gap_pct > 0 else "#dc2626" if gap_pct is not None and gap_pct < 0 else "#64748b",
                is_primary=True,
            ),
        ],
        columns=3,
    )

    latest_table_rows = [
        [
            row["pair_name"],
            row["pilot_property_name"],
            fmt_num(row["pilot_organic_new_users"], 1),
            row["sister_property_name"],
            fmt_num(row["sister_organic_new_users"], 1),
            fmt_num(row["gap_abs"], 1),
            trend_html(row["gap_pct_vs_sister"]),
            trend_html(row["wow_pct"], "WoW"),
        ]
        for row in pair_items
    ]
    builder.add_section(
        Section(
            title="Latest Week Pair Detail",
            content=create_data_table(
                [
                    "Pair",
                    "Pilot Property",
                    "Pilot",
                    "Sister Property",
                    "Sister",
                    "Gap",
                    "Gap vs Sister",
                    "Pair WoW",
                ],
                latest_table_rows,
            ),
            status=None,
            description="Pair-by-pair readout for the latest complete week, with both property names shown in-line.",
        )
    )

    cohort_rows = build_cohort_context_rows(rollup_rows, limit=4)
    methodology_html = create_data_table(
        ["Week Start", "Pilot Cohort", "Pilot WoW", "Sister Cohort", "Sister WoW", "Gap", "Gap vs Sister"],
        cohort_rows,
    ) + (
        f'<p style="font-size:12px;color:#64748b;margin-top:14px;">'
        f'Method note: this uses exact GA4 Organic Search newUsers from ga4_traffic_sources, summed into Monday-start weekly windows through '
        f'{latest_complete_week_end}. Generated {generated_at.strftime("%Y-%m-%d %H:%M %Z").strip()}.'
        f"</p>"
    )
    builder.add_section(
        Section(
            title="Cohort Trend Context",
            content=methodology_html,
            status=None,
            description=(
                f"Pilot Cohort: {cohort_labels['pilot']}. "
                f"Sister Cohort: {cohort_labels['sister']}."
            ),
        )
    )

    html = builder.generate()
    html = html.replace(
        "<body ",
        "<body ",
        1,
    )
    return html


def build_email_text(
    rollup_rows: Sequence[Dict[str, object]],
    pair_rows: Sequence[Dict[str, object]],
    latest_complete_week_end: str,
) -> str:
    latest = latest_rollup_by_role(rollup_rows)
    latest_week = latest["pilot"]["week_start"]
    lines = [
        "Subject: Pilot Organic New Users WoW | Week of "
        f"{latest_week}",
        "",
        f"Week {latest_week} to {latest['pilot']['period_end']}:",
        f"- Pilot total Organic Search new users: {fmt_num(latest['pilot']['organic_new_users'], 1)} "
        f"(WoW {fmt_pct(latest['pilot']['wow_pct'])})",
        f"- Sister total Organic Search new users: {fmt_num(latest['sister']['organic_new_users'], 1)} "
        f"(WoW {fmt_pct(latest['sister']['wow_pct'])})",
        "",
        "Latest pair detail:",
    ]
    for row in latest_pairs(pair_rows):
        lines.append(
            f"- {row['pair_name']}: pilot {fmt_num(row['pilot_organic_new_users'], 1)}, "
            f"sister {fmt_num(row['sister_organic_new_users'], 1)}, gap {fmt_num(row['gap_abs'], 1)}, "
            f"pair WoW {fmt_pct(row['wow_pct'])}"
        )
    lines.extend(
        [
            "",
            f"Method note: exact GA4 Organic Search newUsers from ga4_traffic_sources, through {latest_complete_week_end}.",
        ]
    )
    return "\n".join(lines)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    generated_at = datetime.now().astimezone()
    timestamp = generated_at.strftime("%Y%m%d_%H%M%S")
    run_dir = OUT_DIR / timestamp
    run_dir.mkdir(parents=True, exist_ok=True)

    members = load_pair_members()
    cohort_labels = cohort_property_labels(members)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    latest_complete_week_end = get_latest_complete_week_end(conn, members)
    weekly_rows = fetch_weekly_rows(conn, members, latest_complete_week_end)
    if not weekly_rows:
        raise SystemExit("No GA4 rows found for the requested period.")

    pair_rows = build_pair_rows(weekly_rows)
    rollup_rows = build_rollup_rows(weekly_rows)

    workbook_path = run_dir / "pilot_organic_new_users_wow.xlsx"
    email_html_path = run_dir / "pilot_organic_new_users_wow_email_preview.html"
    email_text_path = run_dir / "pilot_organic_new_users_wow_email.txt"
    metadata_path = run_dir / "pilot_organic_new_users_wow_metadata.json"

    build_workbook(weekly_rows, pair_rows, rollup_rows, latest_complete_week_end, workbook_path)
    email_html_path.write_text(
        build_email_html(rollup_rows, pair_rows, generated_at, latest_complete_week_end, cohort_labels)
    )
    email_text_path.write_text(build_email_text(rollup_rows, pair_rows, latest_complete_week_end))
    metadata_path.write_text(
        json.dumps(
            {
                "generated_at": generated_at.isoformat(),
                "start_date": START_DATE,
                "latest_complete_week_end": latest_complete_week_end,
                "database": str(DB_PATH),
                "pair_source": str(PAIR_PATH),
                "latest_metric_date": max(str(row["period_end"]) for row in weekly_rows),
                "output_files": [
                    str(workbook_path),
                    str(email_html_path),
                    str(email_text_path),
                ],
            },
            indent=2,
        )
    )

    print(json.dumps({"run_dir": str(run_dir), "workbook": str(workbook_path), "email_html": str(email_html_path)}))


if __name__ == "__main__":
    main()
