#!/usr/bin/env python3
"""Export portfolio weekly Organic Search new users by property."""

from __future__ import annotations

import json
import re
import sqlite3
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path("/Users/mark/Property_Analytics")
DB_PATH = ROOT / "data" / "portfolio_analytics.db"
REGISTRY_PATH = ROOT / "config" / "venterra_properties_official.json"
COMMUNITY_SQL_PATHS = [
    ROOT / "apps" / "api" / "scripts" / "generated" / "03_update_communities.sql",
    ROOT / "apps" / "api" / "scripts" / "generated" / "04_insert_communities.sql",
]
OUT_ROOT = ROOT / "reports" / "portfolio_organic_new_users_weekly"
START_DATE = "2026-01-01"


def load_properties():
    payload = json.loads(REGISTRY_PATH.read_text())
    return [
        prop
        for prop in payload["properties"]
        if prop.get("ga4_property_id") and prop.get("property_type") != "new_development"
    ]


def load_property_codes() -> dict[str, str]:
    pattern = re.compile(r"ga4_property_id = '([^']+)'.*?encasa_property_code = '([^']+)'")
    mapping: dict[str, str] = {}
    for path in COMMUNITY_SQL_PATHS:
        if not path.exists():
            continue
        text = path.read_text()
        for ga4_id, code in pattern.findall(text):
            mapping[ga4_id] = code
        if "INSERT INTO communities" in text:
            for line in text.splitlines():
                if "INSERT INTO communities" not in line:
                    continue
                ga4_match = re.search(r", '([^']+)', 'https?://", line)
                code_match = re.search(r"'[^']*', '([^']+)', NULL, NULL,", line)
                if ga4_match and code_match:
                    mapping[ga4_match.group(1)] = code_match.group(1)
    return mapping


def latest_complete_week_end(conn: sqlite3.Connection, property_ids: list[str]) -> str:
    placeholders = ",".join("?" for _ in property_ids)
    max_metric_date = conn.execute(
        f"""
        SELECT MAX(metric_date)
        FROM ga4_daily_metrics
        WHERE property_id IN ({placeholders})
          AND metric_date >= ?
        """,
        property_ids + [START_DATE],
    ).fetchone()[0]
    if not max_metric_date:
        raise SystemExit("No GA4 daily metrics found for requested export.")

    max_date = datetime.strptime(str(max_metric_date), "%Y-%m-%d").date()
    week_end = max_date - timedelta(days=(max_date.weekday() + 1) % 7)
    return week_end.isoformat()


def fetch_rows(conn: sqlite3.Connection, properties: list[dict], week_end: str):
    property_codes = load_property_codes()
    values = []
    for prop in properties:
        name = prop["name"].replace("'", "''")
        ga4_id = str(prop["ga4_property_id"])
        code = property_codes.get(ga4_id, prop["url_slug"]).replace("'", "''")
        ga4_id = ga4_id.replace("'", "''")
        values.append(f"('{name}','{code}','{ga4_id}')")

    sql = f"""
        WITH property_map(property_name, property_code, property_id) AS (
            VALUES
                {", ".join(values)}
        )
        SELECT
            pm.property_name,
            pm.property_code,
            date(ts.metric_date, '-' || ((CAST(strftime('%w', ts.metric_date) AS INTEGER) + 6) % 7) || ' days') AS week_start,
            date(
                date(ts.metric_date, '-' || ((CAST(strftime('%w', ts.metric_date) AS INTEGER) + 6) % 7) || ' days'),
                '+6 days'
            ) AS week_end,
            SUM(ts.new_users) AS organic_new_users
        FROM property_map pm
        JOIN ga4_traffic_sources ts
          ON ts.property_id = pm.property_id
        WHERE ts.channel_group = 'Organic Search'
          AND ts.metric_date >= ?
          AND ts.metric_date <= ?
        GROUP BY pm.property_name, pm.property_code, week_start, week_end
        ORDER BY pm.property_name, week_start
    """
    return conn.execute(sql, (START_DATE, week_end)).fetchall()


def autosize(ws):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 36)


def build_workbook(rows, out_path: Path, week_end: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Organic New Users Weekly"

    ws["A1"] = "Portfolio Organic New Users Weekly Export"
    ws["A1"].font = Font(bold=True, size=15)
    ws["A2"] = f"Date range: {START_DATE} through {week_end}"
    ws["A3"] = "Property code uses encasa_property_code from the Property Analytics community config."

    headers = ["property name", "property code", "week start", "week end", "Organic new users"]
    header_fill = PatternFill("solid", fgColor="15284B")
    header_font = Font(color="FFFFFF", bold=True)
    start_row = 5
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=idx, value=header)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, row in enumerate(rows, start=start_row + 1):
        ws.cell(row=row_idx, column=1, value=row["property_name"])
        ws.cell(row=row_idx, column=2, value=row["property_code"])
        ws.cell(row=row_idx, column=3, value=row["week_start"])
        ws.cell(row=row_idx, column=4, value=row["week_end"])
        ws.cell(row=row_idx, column=5, value=float(row["organic_new_users"] or 0))

    ws.freeze_panes = f"A{start_row + 1}"
    autosize(ws)
    wb.save(out_path)


def main():
    OUT_ROOT.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    run_dir = OUT_ROOT / stamp
    run_dir.mkdir(parents=True, exist_ok=True)

    properties = load_properties()
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    week_end = latest_complete_week_end(conn, [str(prop["ga4_property_id"]) for prop in properties])
    rows = fetch_rows(conn, properties, week_end)
    if not rows:
        raise SystemExit("No Organic Search new-user rows found.")

    out_path = run_dir / "portfolio_organic_new_users_weekly.xlsx"
    build_workbook(rows, out_path, week_end)
    print(json.dumps({"run_dir": str(run_dir), "workbook": str(out_path), "rows": len(rows), "week_end": week_end}))


if __name__ == "__main__":
    main()
