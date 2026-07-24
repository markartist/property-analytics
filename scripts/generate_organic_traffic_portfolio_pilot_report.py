#!/usr/bin/env python3
"""Generate portfolio organic traffic and five-site pilot comparison report."""

from __future__ import annotations

import csv
import html
import json
import sqlite3
import sys
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path("/Users/mark/Property_Analytics")
DB_PATH = ROOT / "data" / "portfolio_analytics.db"
REGISTRY_PATH = ROOT / "config" / "venterra_properties_official.json"
PILOT_CONFIG_PATH = ROOT / "pilot_control_cwv" / "config" / "pilot_control_cwv_config.json"
OUT_ROOT = ROOT / "reports" / "organic_traffic_portfolio_pilot"
REQUESTED_START = date(2024, 6, 25)
PILOT_LAUNCH_DATE = date(2026, 3, 24)

sys.path.insert(0, str(ROOT))
from Data_Collection.utils.property_identity import resolve_property_identity  # noqa: E402


BRAND = {
    "navy": "15284B",
    "san_marino": "3D66B9",
    "bay": "294782",
    "indigo": "5A81CF",
    "monte_carlo": "7DCAC2",
    "pink": "E02472",
    "white_smoke": "F6F6F5",
    "terra_cotta": "BD4830",
    "quill": "D6D6D2",
    "blue_chill": "3B9189",
    "delta": "9B9B96",
    "black": "000000",
    "white": "FFFFFF",
}


@dataclass(frozen=True)
class Property:
    name: str
    property_id: str
    property_code: str | None
    gsc_url: str | None
    lifecycle: str | None
    property_type: str | None


@dataclass(frozen=True)
class PilotPair:
    pair_key: str
    pilot_name: str
    pilot_id: str
    control_name: str
    control_id: str


def iso(d: date) -> str:
    return d.isoformat()


def parse_date(value: str) -> date:
    return datetime.strptime(value, "%Y-%m-%d").date()


def fmt_num(value: Any, decimals: int = 0) -> str:
    if value is None:
        return "-"
    return f"{float(value):,.{decimals}f}"


def fmt_pct(value: Any, decimals: int = 1, signed: bool = False) -> str:
    if value is None:
        return "-"
    sign = "+" if signed else ""
    return f"{float(value) * 100:{sign},.{decimals}f}%"


def pct_change(current: float | int | None, prior: float | int | None) -> float | None:
    if current is None or prior in (None, 0):
        return None
    return (float(current) - float(prior)) / float(prior)


def pct_gap(a: float | int | None, b: float | int | None) -> float | None:
    if a is None or b in (None, 0):
        return None
    return (float(a) - float(b)) / float(b)


def load_portfolio_properties() -> list[Property]:
    payload = json.loads(REGISTRY_PATH.read_text(encoding="utf-8"))
    properties: list[Property] = []
    for row in payload.get("properties", []):
        ga4 = row.get("ga4_property_id")
        if not ga4 or row.get("property_type") == "new_development":
            continue
        identity = resolve_property_identity(str(ga4)) or resolve_property_identity(str(row.get("name", "")))
        properties.append(
            Property(
                name=(identity.property_name if identity else row.get("name")) or str(ga4),
                property_id=str(ga4),
                property_code=identity.property_code if identity else None,
                gsc_url=(identity.gsc_url if identity else row.get("gsc_url")),
                lifecycle=row.get("lifecycle"),
                property_type=row.get("property_type"),
            )
        )
    return sorted(properties, key=lambda item: item.name)


def load_pilot_pairs() -> list[PilotPair]:
    payload = json.loads(PILOT_CONFIG_PATH.read_text(encoding="utf-8"))
    cohorts = {item["key"]: item for item in payload.get("cohorts", []) if item.get("active")}
    pairs: list[PilotPair] = []
    for pilot in [item for item in cohorts.values() if item.get("role") == "pilot"]:
        control = cohorts.get(str(pilot.get("sister_key")))
        if not control:
            continue
        pilot_identity = resolve_property_identity(str(pilot["property_id"])) or resolve_property_identity(str(pilot["display_name"]))
        control_identity = resolve_property_identity(str(control["property_id"])) or resolve_property_identity(str(control["display_name"]))
        pairs.append(
            PilotPair(
                pair_key=str(pilot["key"]).replace("pilot_", ""),
                pilot_name=pilot_identity.property_name if pilot_identity else str(pilot["display_name"]),
                pilot_id=str(pilot["property_id"]),
                control_name=control_identity.property_name if control_identity else str(control["display_name"]),
                control_id=str(control["property_id"]),
            )
        )
    return sorted(pairs, key=lambda item: item.pilot_name)


def placeholders(values: Iterable[str]) -> str:
    return ",".join("?" for _ in values)


def fetch_one(conn: sqlite3.Connection, sql: str, params: tuple[Any, ...] = ()) -> sqlite3.Row:
    row = conn.execute(sql, params).fetchone()
    if row is None:
        raise RuntimeError("Query returned no row")
    return row


def fetch_ga4_summary(conn: sqlite3.Connection, property_ids: list[str], start: date, end: date) -> dict[str, Any]:
    ph = placeholders(property_ids)
    params = [*property_ids, iso(start), iso(end), *property_ids, iso(start), iso(end)]
    row = fetch_one(
        conn,
        f"""
        WITH total AS (
            SELECT
                COUNT(DISTINCT metric_date) AS days,
                SUM(sessions) AS total_sessions,
                SUM(total_users) AS total_users,
                SUM(new_users) AS total_new_users
            FROM ga4_daily_metrics
            WHERE property_id IN ({ph})
              AND metric_date BETWEEN ? AND ?
        ),
        organic AS (
            SELECT
                COUNT(DISTINCT metric_date) AS organic_days,
                SUM(sessions) AS organic_sessions,
                SUM(total_users) AS organic_total_users,
                SUM(new_users) AS organic_new_users,
                SUM(engaged_sessions) AS organic_engaged_sessions,
                SUM(conversions) AS organic_conversions
            FROM ga4_traffic_sources
            WHERE property_id IN ({ph})
              AND channel_group = 'Organic Search'
              AND metric_date BETWEEN ? AND ?
        )
        SELECT
            total.days,
            organic.organic_days,
            total.total_sessions,
            total.total_users,
            total.total_new_users,
            organic.organic_sessions,
            organic.organic_total_users,
            organic.organic_new_users,
            organic.organic_engaged_sessions,
            organic.organic_conversions
        FROM total CROSS JOIN organic
        """,
        tuple(params),
    )
    data = dict(row)
    total_sessions = data.get("total_sessions") or 0
    data["organic_session_share"] = (data.get("organic_sessions") or 0) / total_sessions if total_sessions else None
    return data


def fetch_gsc_summary(conn: sqlite3.Connection, property_ids: list[str], start: date, end: date) -> dict[str, Any]:
    ph = placeholders(property_ids)
    row = fetch_one(
        conn,
        f"""
        SELECT
            COUNT(DISTINCT metric_date) AS days,
            SUM(clicks) AS gsc_clicks,
            SUM(impressions) AS gsc_impressions,
            CASE WHEN SUM(impressions) > 0 THEN CAST(SUM(clicks) AS REAL) / SUM(impressions) END AS gsc_ctr,
            CASE WHEN SUM(impressions) > 0 THEN SUM(average_position * impressions) / SUM(impressions) END AS avg_position
        FROM gsc_daily_metrics
        WHERE COALESCE(ga4_property_id, property_id) IN ({ph})
          AND metric_date BETWEEN ? AND ?
        """,
        tuple([*property_ids, iso(start), iso(end)]),
    )
    return dict(row)


def fetch_monthly_portfolio(conn: sqlite3.Connection, property_ids: list[str], start: date, end: date) -> list[dict[str, Any]]:
    ph = placeholders(property_ids)
    rows = conn.execute(
        f"""
        WITH total AS (
            SELECT
                substr(metric_date, 1, 7) AS month,
                SUM(sessions) AS total_sessions,
                SUM(total_users) AS total_users,
                SUM(new_users) AS total_new_users
            FROM ga4_daily_metrics
            WHERE property_id IN ({ph})
              AND metric_date BETWEEN ? AND ?
            GROUP BY month
        ),
        organic AS (
            SELECT
                substr(metric_date, 1, 7) AS month,
                SUM(sessions) AS organic_sessions,
                SUM(total_users) AS organic_total_users,
                SUM(new_users) AS organic_new_users
            FROM ga4_traffic_sources
            WHERE property_id IN ({ph})
              AND channel_group = 'Organic Search'
              AND metric_date BETWEEN ? AND ?
            GROUP BY month
        ),
        gsc AS (
            SELECT
                substr(metric_date, 1, 7) AS month,
                SUM(clicks) AS gsc_clicks,
                SUM(impressions) AS gsc_impressions
            FROM gsc_daily_metrics
            WHERE COALESCE(ga4_property_id, property_id) IN ({ph})
              AND metric_date BETWEEN ? AND ?
            GROUP BY month
        )
        SELECT
            total.month,
            total.total_sessions,
            organic.organic_sessions,
            organic.organic_new_users,
            organic.organic_total_users,
            CASE WHEN total.total_sessions > 0 THEN CAST(organic.organic_sessions AS REAL) / total.total_sessions END AS organic_session_share,
            gsc.gsc_clicks,
            gsc.gsc_impressions
        FROM total
        LEFT JOIN organic ON organic.month = total.month
        LEFT JOIN gsc ON gsc.month = total.month
        ORDER BY total.month
        """,
        tuple([*property_ids, iso(start), iso(end), *property_ids, iso(start), iso(end), *property_ids, iso(start), iso(end)]),
    ).fetchall()
    return [dict(row) for row in rows]


def fetch_property_rankings(conn: sqlite3.Connection, properties: list[Property], start: date, end: date) -> list[dict[str, Any]]:
    property_ids = [item.property_id for item in properties]
    ph = placeholders(property_ids)
    name_map = {item.property_id: item for item in properties}
    rows = conn.execute(
        f"""
        WITH total AS (
            SELECT property_id, SUM(sessions) AS total_sessions
            FROM ga4_daily_metrics
            WHERE property_id IN ({ph})
              AND metric_date BETWEEN ? AND ?
            GROUP BY property_id
        ),
        organic AS (
            SELECT
                property_id,
                SUM(sessions) AS organic_sessions,
                SUM(total_users) AS organic_total_users,
                SUM(new_users) AS organic_new_users
            FROM ga4_traffic_sources
            WHERE property_id IN ({ph})
              AND channel_group = 'Organic Search'
              AND metric_date BETWEEN ? AND ?
            GROUP BY property_id
        ),
        gsc AS (
            SELECT
                COALESCE(ga4_property_id, property_id) AS property_id,
                SUM(clicks) AS gsc_clicks,
                SUM(impressions) AS gsc_impressions
            FROM gsc_daily_metrics
            WHERE COALESCE(ga4_property_id, property_id) IN ({ph})
              AND metric_date BETWEEN ? AND ?
            GROUP BY COALESCE(ga4_property_id, property_id)
        )
        SELECT
            organic.property_id,
            total.total_sessions,
            organic.organic_sessions,
            organic.organic_total_users,
            organic.organic_new_users,
            CASE WHEN total.total_sessions > 0 THEN CAST(organic.organic_sessions AS REAL) / total.total_sessions END AS organic_session_share,
            gsc.gsc_clicks,
            gsc.gsc_impressions
        FROM organic
        LEFT JOIN total ON total.property_id = organic.property_id
        LEFT JOIN gsc ON gsc.property_id = organic.property_id
        ORDER BY organic.organic_new_users DESC
        """,
        tuple([*property_ids, iso(start), iso(end), *property_ids, iso(start), iso(end), *property_ids, iso(start), iso(end)]),
    ).fetchall()
    ranking: list[dict[str, Any]] = []
    for idx, row in enumerate(rows, start=1):
        prop = name_map[str(row["property_id"])]
        item = dict(row)
        item["rank"] = idx
        item["property_name"] = prop.name
        item["property_code"] = prop.property_code
        ranking.append(item)
    return ranking


def fetch_pilot_comparison(
    conn: sqlite3.Connection,
    pairs: list[PilotPair],
    start: date,
    ga4_end: date,
    gsc_end: date,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for pair in pairs:
        pilot_ga4 = fetch_ga4_summary(conn, [pair.pilot_id], start, ga4_end)
        control_ga4 = fetch_ga4_summary(conn, [pair.control_id], start, ga4_end)
        pilot_gsc = fetch_gsc_summary(conn, [pair.pilot_id], start, gsc_end)
        control_gsc = fetch_gsc_summary(conn, [pair.control_id], start, gsc_end)
        rows.append(
            {
                "pair_key": pair.pair_key,
                "pilot_property": pair.pilot_name,
                "pilot_property_id": pair.pilot_id,
                "control_property": pair.control_name,
                "control_property_id": pair.control_id,
                "pilot_organic_sessions": pilot_ga4.get("organic_sessions") or 0,
                "control_organic_sessions": control_ga4.get("organic_sessions") or 0,
                "organic_sessions_gap": (pilot_ga4.get("organic_sessions") or 0) - (control_ga4.get("organic_sessions") or 0),
                "organic_sessions_gap_pct": pct_gap(pilot_ga4.get("organic_sessions") or 0, control_ga4.get("organic_sessions") or 0),
                "pilot_organic_new_users": pilot_ga4.get("organic_new_users") or 0,
                "control_organic_new_users": control_ga4.get("organic_new_users") or 0,
                "organic_new_users_gap": (pilot_ga4.get("organic_new_users") or 0) - (control_ga4.get("organic_new_users") or 0),
                "organic_new_users_gap_pct": pct_gap(pilot_ga4.get("organic_new_users") or 0, control_ga4.get("organic_new_users") or 0),
                "pilot_organic_share": pilot_ga4.get("organic_session_share"),
                "control_organic_share": control_ga4.get("organic_session_share"),
                "pilot_gsc_clicks": pilot_gsc.get("gsc_clicks") or 0,
                "control_gsc_clicks": control_gsc.get("gsc_clicks") or 0,
                "gsc_click_gap": (pilot_gsc.get("gsc_clicks") or 0) - (control_gsc.get("gsc_clicks") or 0),
                "gsc_click_gap_pct": pct_gap(pilot_gsc.get("gsc_clicks") or 0, control_gsc.get("gsc_clicks") or 0),
                "pilot_gsc_impressions": pilot_gsc.get("gsc_impressions") or 0,
                "control_gsc_impressions": control_gsc.get("gsc_impressions") or 0,
            }
        )
    return sorted(rows, key=lambda item: item["pilot_organic_new_users"], reverse=True)


def fetch_pilot_t30(
    conn: sqlite3.Connection,
    pairs: list[PilotPair],
    current_start: date,
    current_end: date,
    prior_start: date,
    prior_end: date,
    gsc_end: date,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    gsc_current_end = min(current_end, gsc_end)
    gsc_current_start = gsc_current_end - timedelta(days=29)
    gsc_prior_end = gsc_current_start - timedelta(days=1)
    gsc_prior_start = gsc_prior_end - timedelta(days=29)
    for pair in pairs:
        cur = fetch_ga4_summary(conn, [pair.pilot_id], current_start, current_end)
        prior = fetch_ga4_summary(conn, [pair.pilot_id], prior_start, prior_end)
        gsc_cur = fetch_gsc_summary(conn, [pair.pilot_id], gsc_current_start, gsc_current_end)
        gsc_prior = fetch_gsc_summary(conn, [pair.pilot_id], gsc_prior_start, gsc_prior_end)
        rows.append(
            {
                "pilot_property": pair.pilot_name,
                "organic_sessions_t30": cur.get("organic_sessions") or 0,
                "organic_sessions_prior30": prior.get("organic_sessions") or 0,
                "organic_sessions_change_pct": pct_change(cur.get("organic_sessions") or 0, prior.get("organic_sessions") or 0),
                "organic_new_users_t30": cur.get("organic_new_users") or 0,
                "organic_new_users_prior30": prior.get("organic_new_users") or 0,
                "organic_new_users_change_pct": pct_change(cur.get("organic_new_users") or 0, prior.get("organic_new_users") or 0),
                "organic_share_t30": cur.get("organic_session_share"),
                "organic_share_prior30": prior.get("organic_session_share"),
                "gsc_clicks_t30": gsc_cur.get("gsc_clicks") or 0,
                "gsc_clicks_prior30": gsc_prior.get("gsc_clicks") or 0,
                "gsc_clicks_change_pct": pct_change(gsc_cur.get("gsc_clicks") or 0, gsc_prior.get("gsc_clicks") or 0),
            }
        )
    return sorted(rows, key=lambda item: item["organic_new_users_t30"], reverse=True)


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        return
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def autosize(ws) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[letter].width = min(max(max_len + 2, 11), 42)


def write_sheet(ws, rows: list[dict[str, Any]]) -> None:
    if not rows:
        return
    header_fill = PatternFill("solid", fgColor=BRAND["navy"])
    header_font = Font(color=BRAND["white"], bold=True)
    for col_idx, key in enumerate(rows[0].keys(), start=1):
        cell = ws.cell(row=1, column=col_idx, value=key)
        cell.fill = header_fill
        cell.font = header_font
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, key in enumerate(row.keys(), start=1):
            ws.cell(row=row_idx, column=col_idx, value=row[key])
    ws.freeze_panes = "A2"
    autosize(ws)


def build_workbook(path: Path, sheets: dict[str, list[dict[str, Any]]], metadata: list[list[Any]]) -> None:
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary["A1"] = "Organic Traffic Portfolio And Pilot Report"
    summary["A1"].font = Font(bold=True, size=16, color=BRAND["navy"])
    for row_idx, row in enumerate(metadata, start=3):
        summary.cell(row=row_idx, column=1, value=row[0]).font = Font(bold=True, color=BRAND["bay"])
        summary.cell(row=row_idx, column=2, value=row[1])
    autosize(summary)
    for title, rows in sheets.items():
        ws = wb.create_sheet(title[:31])
        write_sheet(ws, rows)
        if title == "Monthly Portfolio" and len(rows) >= 2:
            chart = LineChart()
            chart.title = "Portfolio Organic New Users By Month"
            chart.y_axis.title = "Organic new users"
            chart.x_axis.title = "Month"
            data = Reference(ws, min_col=4, min_row=1, max_row=len(rows) + 1)
            cats = Reference(ws, min_col=1, min_row=2, max_row=len(rows) + 1)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 8
            chart.width = 18
            ws.add_chart(chart, "J2")
    wb.save(path)


def html_table(rows: list[dict[str, Any]], columns: list[tuple[str, str]], limit: int | None = None) -> str:
    visible = rows[:limit] if limit else rows
    header = "".join(f"<th>{html.escape(label)}</th>" for _, label in columns)
    body_rows = []
    for row in visible:
        cells = "".join(f"<td>{html.escape(str(row.get(key, '-')))}</td>" for key, _ in columns)
        body_rows.append(f"<tr>{cells}</tr>")
    return f"<table><thead><tr>{header}</tr></thead><tbody>{''.join(body_rows)}</tbody></table>"


def build_html(
    path: Path,
    summary: dict[str, Any],
    monthly: list[dict[str, Any]],
    pilot_compare_display: list[dict[str, Any]],
    pilot_t30_display: list[dict[str, Any]],
    rankings_display: list[dict[str, Any]],
) -> None:
    css = f"""
    body {{ margin:0; font-family: Arial, sans-serif; color:#{BRAND['navy']}; background:#{BRAND['white_smoke']}; }}
    main {{ max-width:1180px; margin:0 auto; padding:34px 24px 48px; }}
    h1 {{ margin:0; font-size:32px; line-height:1.1; color:#{BRAND['navy']}; }}
    h2 {{ margin:30px 0 10px; font-size:20px; color:#{BRAND['bay']}; }}
    p {{ line-height:1.5; color:#{BRAND['black']}; }}
    .sub {{ margin-top:8px; color:#{BRAND['bay']}; }}
    .kpis {{ display:grid; grid-template-columns:repeat(4, minmax(0,1fr)); gap:12px; margin-top:24px; }}
    .kpi {{ background:#{BRAND['white']}; border:1px solid #{BRAND['quill']}; border-top:4px solid #{BRAND['san_marino']}; padding:14px; }}
    .label {{ font-size:12px; text-transform:uppercase; letter-spacing:.04em; color:#{BRAND['bay']}; font-weight:700; }}
    .value {{ margin-top:8px; font-size:28px; font-weight:800; color:#{BRAND['navy']}; }}
    .note {{ background:#{BRAND['white']}; border-left:5px solid #{BRAND['terra_cotta']}; padding:14px 16px; }}
    table {{ width:100%; border-collapse:collapse; background:#{BRAND['white']}; margin-top:12px; font-size:13px; }}
    th {{ background:#{BRAND['navy']}; color:#{BRAND['white']}; text-align:left; padding:9px; }}
    td {{ border:1px solid #{BRAND['quill']}; padding:8px; color:#{BRAND['black']}; }}
    tr:nth-child(even) td {{ background:#FFFFFF; }}
    .footer {{ margin-top:28px; font-size:12px; color:#{BRAND['delta']}; }}
    """
    contents = f"""<!doctype html>
<html>
<head>
  <meta charset="utf-8">
  <title>Organic Traffic Portfolio And Pilot Report</title>
  <style>{css}</style>
</head>
<body>
<main>
  <h1>Organic Traffic Portfolio And Pilot Report</h1>
  <div class="sub">Generated {html.escape(summary['generated_at'])}; source: Data Pond GA4 Organic Search plus GSC daily metrics.</div>
  <div class="kpis">
    <div class="kpi"><div class="label">Portfolio Organic New Users</div><div class="value">{html.escape(summary['portfolio_organic_new_users'])}</div></div>
    <div class="kpi"><div class="label">Portfolio Organic Sessions</div><div class="value">{html.escape(summary['portfolio_organic_sessions'])}</div></div>
    <div class="kpi"><div class="label">Organic Session Share</div><div class="value">{html.escape(summary['portfolio_organic_share'])}</div></div>
    <div class="kpi"><div class="label">GSC Clicks</div><div class="value">{html.escape(summary['portfolio_gsc_clicks'])}</div></div>
  </div>
  <h2>Executive Read</h2>
  <p>{html.escape(summary['executive_read'])}</p>
  <div class="note">{html.escape(summary['coverage_note'])}</div>
  <h2>Latest 30-Day Movement</h2>
  <p>{html.escape(summary['t30_note'])}</p>
  <h2>Five New Pilot Sites: Post-Launch Comparison</h2>
  {html_table(pilot_compare_display, [
      ('pilot_property', 'Pilot'),
      ('control_property', 'Control'),
      ('pilot_organic_new_users', 'Pilot organic new users'),
      ('control_organic_new_users', 'Control organic new users'),
      ('organic_new_users_gap_pct', 'Gap vs control'),
      ('pilot_organic_share', 'Pilot organic share'),
      ('pilot_gsc_clicks', 'Pilot GSC clicks'),
      ('gsc_click_gap_pct', 'GSC click gap')
  ])}
  <h2>Pilot Latest T30</h2>
  {html_table(pilot_t30_display, [
      ('pilot_property', 'Pilot'),
      ('organic_new_users_t30', 'Organic new users T30'),
      ('organic_new_users_change_pct', 'New users vs prior 30'),
      ('organic_sessions_t30', 'Organic sessions T30'),
      ('organic_sessions_change_pct', 'Sessions vs prior 30'),
      ('organic_share_t30', 'Organic share'),
      ('gsc_clicks_t30', 'GSC clicks T30'),
      ('gsc_clicks_change_pct', 'GSC clicks vs prior 30')
  ])}
  <h2>Monthly Portfolio Trend</h2>
  {html_table([
      {
          'month': row['month'],
          'organic_new_users': fmt_num(row.get('organic_new_users')),
          'organic_sessions': fmt_num(row.get('organic_sessions')),
          'organic_session_share': fmt_pct(row.get('organic_session_share')),
          'gsc_clicks': fmt_num(row.get('gsc_clicks')),
      } for row in monthly
  ], [
      ('month', 'Month'),
      ('organic_new_users', 'Organic new users'),
      ('organic_sessions', 'Organic sessions'),
      ('organic_session_share', 'Organic share'),
      ('gsc_clicks', 'GSC clicks')
  ])}
  <h2>Top Portfolio Properties By Organic New Users</h2>
  {html_table(rankings_display, [
      ('rank', 'Rank'),
      ('property_name', 'Property'),
      ('organic_new_users', 'Organic new users'),
      ('organic_sessions', 'Organic sessions'),
      ('organic_session_share', 'Organic share'),
      ('gsc_clicks', 'GSC clicks')
  ], limit=20)}
  <div class="footer">Method: GA4 traffic source rows where channel_group = Organic Search; GSC rows joined by governed GA4 identity. Pilot launch/post-launch window starts {PILOT_LAUNCH_DATE.isoformat()}.</div>
</main>
</body>
</html>"""
    path.write_text(contents, encoding="utf-8")


def display_pilot_compare(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "pilot_property": row["pilot_property"],
            "control_property": row["control_property"],
            "pilot_organic_new_users": fmt_num(row["pilot_organic_new_users"]),
            "control_organic_new_users": fmt_num(row["control_organic_new_users"]),
            "organic_new_users_gap_pct": fmt_pct(row["organic_new_users_gap_pct"], signed=True),
            "pilot_organic_share": fmt_pct(row["pilot_organic_share"]),
            "pilot_gsc_clicks": fmt_num(row["pilot_gsc_clicks"]),
            "gsc_click_gap_pct": fmt_pct(row["gsc_click_gap_pct"], signed=True),
        }
        for row in rows
    ]


def display_t30(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "pilot_property": row["pilot_property"],
            "organic_sessions_t30": fmt_num(row["organic_sessions_t30"]),
            "organic_sessions_change_pct": fmt_pct(row["organic_sessions_change_pct"], signed=True),
            "organic_new_users_t30": fmt_num(row["organic_new_users_t30"]),
            "organic_new_users_change_pct": fmt_pct(row["organic_new_users_change_pct"], signed=True),
            "organic_share_t30": fmt_pct(row["organic_share_t30"]),
            "gsc_clicks_t30": fmt_num(row["gsc_clicks_t30"]),
            "gsc_clicks_change_pct": fmt_pct(row["gsc_clicks_change_pct"], signed=True),
        }
        for row in rows
    ]


def display_rankings(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "rank": row["rank"],
            "property_name": row["property_name"],
            "property_code": row["property_code"] or "",
            "organic_new_users": fmt_num(row["organic_new_users"]),
            "organic_sessions": fmt_num(row["organic_sessions"]),
            "organic_session_share": fmt_pct(row["organic_session_share"]),
            "gsc_clicks": fmt_num(row["gsc_clicks"]),
        }
        for row in rows
    ]


def main() -> int:
    generated_at = datetime.now().astimezone()
    run_dir = OUT_ROOT / generated_at.strftime("%Y%m%d_%H%M%S")
    run_dir.mkdir(parents=True, exist_ok=True)

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    properties = load_portfolio_properties()
    property_ids = [item.property_id for item in properties]
    pairs = load_pilot_pairs()

    ga4_bounds = fetch_one(
        conn,
        f"""
        SELECT MIN(metric_date) AS min_date, MAX(metric_date) AS max_date
        FROM ga4_traffic_sources
        WHERE channel_group = 'Organic Search'
          AND property_id IN ({placeholders(property_ids)})
        """,
        tuple(property_ids),
    )
    gsc_bounds = fetch_one(
        conn,
        f"""
        SELECT MIN(metric_date) AS min_date, MAX(metric_date) AS max_date
        FROM gsc_daily_metrics
        WHERE COALESCE(ga4_property_id, property_id) IN ({placeholders(property_ids)})
        """,
        tuple(property_ids),
    )
    ga4_start = max(REQUESTED_START, parse_date(ga4_bounds["min_date"]))
    ga4_end = parse_date(ga4_bounds["max_date"])
    gsc_start = max(REQUESTED_START, parse_date(gsc_bounds["min_date"]))
    gsc_end = parse_date(gsc_bounds["max_date"])

    portfolio_all = fetch_ga4_summary(conn, property_ids, ga4_start, ga4_end)
    portfolio_gsc_all = fetch_gsc_summary(conn, property_ids, gsc_start, gsc_end)
    t30_end = ga4_end
    t30_start = t30_end - timedelta(days=29)
    prior30_end = t30_start - timedelta(days=1)
    prior30_start = prior30_end - timedelta(days=29)
    portfolio_t30 = fetch_ga4_summary(conn, property_ids, t30_start, t30_end)
    portfolio_prior30 = fetch_ga4_summary(conn, property_ids, prior30_start, prior30_end)
    gsc_t30_end = gsc_end
    gsc_t30_start = gsc_t30_end - timedelta(days=29)
    gsc_prior30_end = gsc_t30_start - timedelta(days=1)
    gsc_prior30_start = gsc_prior30_end - timedelta(days=29)
    gsc_t30 = fetch_gsc_summary(conn, property_ids, gsc_t30_start, gsc_t30_end)
    gsc_prior30 = fetch_gsc_summary(conn, property_ids, gsc_prior30_start, gsc_prior30_end)

    monthly = fetch_monthly_portfolio(conn, property_ids, ga4_start, ga4_end)
    rankings = fetch_property_rankings(conn, properties, ga4_start, ga4_end)
    pilot_compare = fetch_pilot_comparison(conn, pairs, PILOT_LAUNCH_DATE, ga4_end, gsc_end)
    pilot_t30 = fetch_pilot_t30(conn, pairs, t30_start, t30_end, prior30_start, prior30_end, gsc_end)

    pilot_ids = [pair.pilot_id for pair in pairs]
    control_ids = [pair.control_id for pair in pairs]
    pilot_rollup = fetch_ga4_summary(conn, pilot_ids, PILOT_LAUNCH_DATE, ga4_end)
    control_rollup = fetch_ga4_summary(conn, control_ids, PILOT_LAUNCH_DATE, ga4_end)
    pilot_gsc_rollup = fetch_gsc_summary(conn, pilot_ids, PILOT_LAUNCH_DATE, gsc_end)
    control_gsc_rollup = fetch_gsc_summary(conn, control_ids, PILOT_LAUNCH_DATE, gsc_end)

    portfolio_new_users_change = pct_change(portfolio_t30.get("organic_new_users") or 0, portfolio_prior30.get("organic_new_users") or 0)
    portfolio_sessions_change = pct_change(portfolio_t30.get("organic_sessions") or 0, portfolio_prior30.get("organic_sessions") or 0)
    gsc_click_change = pct_change(gsc_t30.get("gsc_clicks") or 0, gsc_prior30.get("gsc_clicks") or 0)
    pilot_gap = pct_gap(pilot_rollup.get("organic_new_users") or 0, control_rollup.get("organic_new_users") or 0)
    pilot_gsc_gap = pct_gap(pilot_gsc_rollup.get("gsc_clicks") or 0, control_gsc_rollup.get("gsc_clicks") or 0)

    summary = {
        "generated_at": generated_at.strftime("%Y-%m-%d %H:%M %Z"),
        "requested_window": f"{REQUESTED_START.isoformat()} through {ga4_end.isoformat()}",
        "effective_ga4_window": f"{ga4_start.isoformat()} through {ga4_end.isoformat()}",
        "effective_gsc_window": f"{gsc_start.isoformat()} through {gsc_end.isoformat()}",
        "portfolio_property_count": len(property_ids),
        "portfolio_organic_new_users": fmt_num(portfolio_all.get("organic_new_users")),
        "portfolio_organic_sessions": fmt_num(portfolio_all.get("organic_sessions")),
        "portfolio_organic_share": fmt_pct(portfolio_all.get("organic_session_share")),
        "portfolio_gsc_clicks": fmt_num(portfolio_gsc_all.get("gsc_clicks")),
        "coverage_note": (
            f"Requested last-two-years window starts {REQUESTED_START.isoformat()}; canonical local GA4 Organic Search coverage begins "
            f"{ga4_start.isoformat()} and GSC coverage begins {gsc_start.isoformat()}. This report does not infer missing 2024-06-25 to "
            f"{(ga4_start - timedelta(days=1)).isoformat()} GA4 organic traffic."
        ),
        "executive_read": (
            f"Across {len(property_ids)} governed portfolio properties, available GA4 Organic Search produced "
            f"{fmt_num(portfolio_all.get('organic_new_users'))} new users and {fmt_num(portfolio_all.get('organic_sessions'))} sessions "
            f"from {ga4_start.isoformat()} to {ga4_end.isoformat()}, representing {fmt_pct(portfolio_all.get('organic_session_share'))} "
            f"of sessions. GSC visibility over its available window contributed {fmt_num(portfolio_gsc_all.get('gsc_clicks'))} clicks."
        ),
        "t30_note": (
            f"Latest T30 GA4 Organic Search new users were {fmt_num(portfolio_t30.get('organic_new_users'))} "
            f"({fmt_pct(portfolio_new_users_change, signed=True)} vs prior 30); organic sessions were "
            f"{fmt_num(portfolio_t30.get('organic_sessions'))} ({fmt_pct(portfolio_sessions_change, signed=True)}). "
            f"GSC clicks were {fmt_num(gsc_t30.get('gsc_clicks'))} ({fmt_pct(gsc_click_change, signed=True)} vs prior 30)."
        ),
        "pilot_summary": (
            f"Post-launch, the five pilot sites produced {fmt_num(pilot_rollup.get('organic_new_users'))} organic new users versus "
            f"{fmt_num(control_rollup.get('organic_new_users'))} for the five controls ({fmt_pct(pilot_gap, signed=True)}). "
            f"GSC clicks were {fmt_num(pilot_gsc_rollup.get('gsc_clicks'))} for pilots versus {fmt_num(control_gsc_rollup.get('gsc_clicks'))} "
            f"for controls ({fmt_pct(pilot_gsc_gap, signed=True)})."
        ),
    }

    write_csv(run_dir / "monthly_portfolio.csv", monthly)
    write_csv(run_dir / "property_rankings.csv", rankings)
    write_csv(run_dir / "pilot_post_launch_comparison.csv", pilot_compare)
    write_csv(run_dir / "pilot_latest_t30.csv", pilot_t30)

    sheets = {
        "Monthly Portfolio": monthly,
        "Property Rankings": rankings,
        "Pilot Post Launch": pilot_compare,
        "Pilot Latest T30": pilot_t30,
    }
    metadata = [
        ["Generated", summary["generated_at"]],
        ["Requested Window", summary["requested_window"]],
        ["Effective GA4 Window", summary["effective_ga4_window"]],
        ["Effective GSC Window", summary["effective_gsc_window"]],
        ["Portfolio Properties", summary["portfolio_property_count"]],
        ["Method", "GA4 Organic Search channel rows plus GSC daily metrics resolved through governed GA4 identity."],
        ["Coverage Note", summary["coverage_note"]],
        ["Pilot Summary", summary["pilot_summary"]],
    ]
    build_workbook(run_dir / "organic_traffic_portfolio_pilot_report.xlsx", sheets, metadata)

    build_html(
        run_dir / "organic_traffic_portfolio_pilot_report.html",
        summary,
        monthly,
        display_pilot_compare(pilot_compare),
        display_t30(pilot_t30),
        display_rankings(rankings),
    )

    summary_path = run_dir / "organic_traffic_portfolio_pilot_summary.json"
    summary_path.write_text(
        json.dumps(
            {
                "summary": summary,
                "artifacts": {
                    "html": str(run_dir / "organic_traffic_portfolio_pilot_report.html"),
                    "workbook": str(run_dir / "organic_traffic_portfolio_pilot_report.xlsx"),
                    "monthly_portfolio_csv": str(run_dir / "monthly_portfolio.csv"),
                    "property_rankings_csv": str(run_dir / "property_rankings.csv"),
                    "pilot_post_launch_comparison_csv": str(run_dir / "pilot_post_launch_comparison.csv"),
                    "pilot_latest_t30_csv": str(run_dir / "pilot_latest_t30.csv"),
                },
            },
            indent=2,
        ),
        encoding="utf-8",
    )
    print(json.dumps({"run_dir": str(run_dir), "summary": summary}, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
