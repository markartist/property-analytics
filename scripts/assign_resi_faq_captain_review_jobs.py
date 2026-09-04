#!/usr/bin/env python3
"""Create Captain job-board actions for a Resi FAQ review workbook."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import sqlite3
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path("/Users/mark/Property_Analytics")
sys.path.insert(0, str(ROOT))

from Data_Collection.utils.property_identity import load_property_identities  # noqa: E402
from scripts.captain_fleet_support import remote_d1_query  # noqa: E402


DEFAULT_DB = ROOT / "data" / "portfolio_analytics.db"
DEFAULT_OUTPUT_ROOT = ROOT / "reports" / "captains_log" / "faq_review_jobs"
DEFAULT_ASSIGNMENT_MESSAGE = (
    ROOT
    / "outputs"
    / "resi-faq-captain-review-20260901"
    / "captain_faq_review_assignment_message_20260901.md"
)


@dataclass(frozen=True)
class PropertyJob:
    property_code: str
    property_name: str
    community_id: str | None
    captain: str
    property_url: str
    total_faqs: int
    ready_for_review: int
    needs_fact_review: int
    blank_proposed: int
    issue_flagged: int
    due_date: str
    workbook_path: str
    assignment_message_path: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", required=True, help="Simple captain FAQ review workbook.")
    parser.add_argument("--due-date", required=True, help="Human deadline in MM/DD/YYYY or machine YYYY-MM-DD.")
    parser.add_argument("--batch-id", default="20260901", help="Stable batch id for action/watch keys.")
    parser.add_argument("--assignment-message", default=str(DEFAULT_ASSIGNMENT_MESSAGE))
    parser.add_argument("--output-root", default=str(DEFAULT_OUTPUT_ROOT))
    parser.add_argument("--apply", action="store_true", help="Apply job-board rows to remote Captain D1.")
    return parser.parse_args()


def stable_id(*parts: str) -> str:
    digest = hashlib.sha1(":".join(parts).encode("utf-8")).hexdigest()[:16]
    return f"{parts[0]}_{digest}"


def sql(value: Any) -> str:
    if value is None:
        return "NULL"
    return "'" + str(value).replace("'", "''") + "'"


def parse_due_date(value: str) -> str:
    text = value.strip()
    for fmt in ("%m/%d/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    raise ValueError("due date must be MM/DD/YYYY or YYYY-MM-DD")


def latest_inventory_run() -> str:
    with sqlite3.connect(DEFAULT_DB) as conn:
        row = conn.execute(
            "SELECT run_id FROM resi_content_inventory_runs ORDER BY fetched_at DESC LIMIT 1"
        ).fetchone()
    return row[0] if row else ""


def load_captain_names() -> dict[str, str]:
    names: dict[str, str] = {}
    try:
        rows = remote_d1_query(
            """
            SELECT
              property_id,
              community_id,
              json_extract(source_scope_json, '$.captain') AS captain
            FROM captain_support_agents
            WHERE status = 'active'
              AND json_extract(source_scope_json, '$.captain') IS NOT NULL
            GROUP BY property_id, community_id, captain
            """
        )
        for row in rows:
            name = row.get("captain")
            if not name:
                continue
            property_id = row.get("property_id")
            community_id = row.get("community_id")
            if property_id:
                names[str(property_id).upper()] = str(name)
            if community_id:
                names[str(community_id).lower()] = str(name)
    except Exception:
        names = {}

    candidates = sorted((ROOT / "reports" / "captains_log" / "activation").glob("captain_activation_roster_*.json"))
    if not candidates:
        return names
    data = json.loads(candidates[-1].read_text(encoding="utf-8"))
    for row in data.get("properties", []):
        name = row.get("captain")
        if not name:
            continue
        for key in ("property_code", "ga4_property_id"):
            value = row.get(key)
            if value and str(value).upper() not in names:
                names[str(value).upper()] = name
        community_id = row.get("community_id")
        if community_id and str(community_id).lower() not in names:
            names[str(community_id).lower()] = name
    return names


def property_identity_maps() -> tuple[dict[str, str | None], dict[str, str]]:
    community_by_code: dict[str, str | None] = {}
    canonical_by_any: dict[str, str] = {}
    for identity in load_property_identities():
        canonical = identity.property_code or identity.canonical_property_id or identity.marketing_bi_property_id
        if not canonical:
            continue
        canonical = str(canonical).upper()
        community_by_code[canonical] = identity.community_id
        for candidate in (
            identity.property_code,
            identity.canonical_property_id,
            identity.marketing_bi_property_id,
            identity.ga4_property_id,
            identity.display_property_id,
        ):
            if candidate:
                canonical_by_any[str(candidate).upper()] = canonical
    return community_by_code, canonical_by_any


def load_jobs(workbook: Path, due_date: str, batch_id: str, assignment_message: Path) -> list[PropertyJob]:
    wb = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
    if "Captain FAQ Review" not in wb.sheetnames:
        raise ValueError("Workbook must contain a 'Captain FAQ Review' sheet.")
    sheet = wb["Captain FAQ Review"]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(rows)]
    idx = {header: i for i, header in enumerate(headers)}
    required = [
        "Property Code",
        "Property Name",
        "Captain",
        "Property URL",
        "Proposed Answer",
        "Original Answer Issue",
        "Status",
    ]
    missing = [header for header in required if header not in idx]
    if missing:
        raise ValueError(f"Workbook missing required columns: {', '.join(missing)}")

    community_by_code, canonical_by_any = property_identity_maps()
    captain_names = load_captain_names()
    grouped: dict[str, dict[str, Any]] = {}
    for values in rows:
        raw_code = str(values[idx["Property Code"]] or "").strip().upper()
        if not raw_code:
            continue
        code = canonical_by_any.get(raw_code, raw_code)
        group = grouped.setdefault(
            code,
            {
                "property_code": code,
                "property_name": str(values[idx["Property Name"]] or "").strip(),
                "captain": "",
                "property_url": str(values[idx["Property URL"]] or "").strip(),
                "total_faqs": 0,
                "ready_for_review": 0,
                "needs_fact_review": 0,
                "blank_proposed": 0,
                "issue_flagged": 0,
            },
        )
        if not group["captain"]:
            community_id = community_by_code.get(code)
            group["captain"] = (
                captain_names.get(code)
                or captain_names.get(raw_code)
                or (captain_names.get(str(community_id).lower()) if community_id else None)
                or str(values[idx["Captain"]] or "").strip()
            )
        group["total_faqs"] += 1
        status = str(values[idx["Status"]] or "").strip()
        if status == "Ready for Review":
            group["ready_for_review"] += 1
        elif status == "Needs Fact Review":
            group["needs_fact_review"] += 1
        if not str(values[idx["Proposed Answer"]] or "").strip():
            group["blank_proposed"] += 1
        if str(values[idx["Original Answer Issue"]] or "").strip():
            group["issue_flagged"] += 1

    return [
        PropertyJob(
            community_id=community_by_code.get(code),
            due_date=due_date,
            workbook_path=str(workbook),
            assignment_message_path=str(assignment_message),
            **group,
        )
        for code, group in sorted(grouped.items())
    ]


def evidence(job: PropertyJob, batch_id: str, latest_run: str) -> dict[str, Any]:
    return {
        "batch_id": batch_id,
        "assignment": "Resi FAQ captain review",
        "deadline": "09/02/2026 12:00 PM CT" if job.due_date == "2026-09-02" else job.due_date,
        "workbook_path": job.workbook_path,
        "assignment_message_path": job.assignment_message_path,
        "latest_resi_inventory_run": latest_run,
        "faq_counts": {
            "total": job.total_faqs,
            "ready_for_review": job.ready_for_review,
            "needs_fact_review": job.needs_fact_review,
            "blank_proposed": job.blank_proposed,
            "original_issue_flagged": job.issue_flagged,
        },
    }


def render_sql(jobs: list[PropertyJob], batch_id: str, latest_run: str, as_of: datetime) -> str:
    now = as_of.isoformat().replace("+00:00", "Z")
    action_key = f"resi_faq_review_{batch_id}"
    watch_key = f"resi_faq_review_due_{batch_id}"
    lines: list[str] = []
    for job in jobs:
        ev = json.dumps(evidence(job, batch_id, latest_run), sort_keys=True, separators=(",", ":"))
        priority = "high" if job.needs_fact_review or job.blank_proposed else "medium"
        current_state = (
            f"{job.total_faqs} FAQ rows assigned; {job.ready_for_review} proposed answers ready for review; "
            f"{job.needs_fact_review} need fact review; {job.blank_proposed} blank proposed answers."
        )
        next_move = (
            "Review property FAQ rows in the workbook, update Proposed Answer where needed, add Captain Logic / "
            "Rationale for meaningful changes, and set Status by 09/02/2026 at 12:00 PM CT."
        )
        lines.append(
            "INSERT INTO captain_watch_items "
            "(id, property_id, community_id, watch_key, title, category, severity, status, current_state, evidence_json, next_move, owner_role, due_date, source_agent_key, first_seen_at, last_seen_at, resolved_at, created_at, updated_at, updated_by) VALUES "
            f"({sql(stable_id('watch', job.property_code, watch_key))}, {sql(job.property_code)}, {sql(job.community_id)}, {sql(watch_key)}, "
            f"{sql('FAQ answer review due by noon')}, 'website_content', 'high', 'open', {sql(current_state)}, {sql(ev)}, {sql(next_move)}, "
            f"{sql(job.captain or 'Property Captain')}, {sql(job.due_date)}, 'faq_review_job_board', {sql(now)}, {sql(now)}, NULL, {sql(now)}, {sql(now)}, 'faq-review-job-board') "
            "ON CONFLICT(property_id, watch_key) DO UPDATE SET status=excluded.status, current_state=excluded.current_state, evidence_json=excluded.evidence_json, next_move=excluded.next_move, owner_role=excluded.owner_role, due_date=excluded.due_date, last_seen_at=excluded.last_seen_at, updated_at=excluded.updated_at, updated_by=excluded.updated_by;"
        )
        lines.append(
            "INSERT INTO captain_actions "
            "(id, property_id, community_id, action_key, title, owner_role, due_date, status, priority, evidence_json, source_agent_key, created_from_run_id, created_at, updated_at, updated_by) VALUES "
            f"({sql(stable_id('action', job.property_code, action_key))}, {sql(job.property_code)}, {sql(job.community_id)}, {sql(action_key)}, "
            f"{sql('Review and validate Resi FAQ proposed answers')}, {sql(job.captain or 'Property Captain')}, {sql(job.due_date)}, 'open', {sql(priority)}, {sql(ev)}, "
            f"'faq_review_job_board', NULL, {sql(now)}, {sql(now)}, 'faq-review-job-board') "
            "ON CONFLICT(property_id, action_key) DO UPDATE SET status=excluded.status, priority=excluded.priority, evidence_json=excluded.evidence_json, owner_role=excluded.owner_role, due_date=excluded.due_date, updated_at=excluded.updated_at, updated_by=excluded.updated_by;"
        )
    lines.append("")
    return "\n".join(lines)


def write_outputs(run_dir: Path, jobs: list[PropertyJob], sql_text: str) -> None:
    run_dir.mkdir(parents=True, exist_ok=True)
    (run_dir / "captain_faq_review_job_board.sql").write_text(sql_text, encoding="utf-8")
    (run_dir / "captain_faq_review_job_board.json").write_text(
        json.dumps([job.__dict__ for job in jobs], indent=2),
        encoding="utf-8",
    )
    with (run_dir / "captain_faq_review_job_board.csv").open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(jobs[0].__dict__.keys()))
        writer.writeheader()
        for job in jobs:
            writer.writerow(job.__dict__)
    lines = [
        "# Captain FAQ Review Job Board",
        "",
        "Deadline: 09/02/2026 12:00 PM CT",
        "",
        "| Property | Captain | Total FAQs | Ready Drafts | Needs Fact Review | Blank Proposed |",
        "| --- | --- | ---: | ---: | ---: | ---: |",
    ]
    for job in jobs:
        lines.append(
            f"| {job.property_code} - {job.property_name} | {job.captain} | {job.total_faqs} | "
            f"{job.ready_for_review} | {job.needs_fact_review} | {job.blank_proposed} |"
        )
    (run_dir / "captain_faq_review_job_board.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def apply_sql(jobs: list[PropertyJob], batch_id: str, latest_run: str, as_of: datetime) -> dict[str, Any]:
    for job in jobs:
        remote_d1_query(render_sql([job], batch_id, latest_run, as_of))
    action_key = f"resi_faq_review_{batch_id}"
    watch_key = f"resi_faq_review_due_{batch_id}"
    quoted_codes = ",".join(sql(job.property_code) for job in jobs)
    actions = remote_d1_query(
        f"SELECT COUNT(*) AS count FROM captain_actions WHERE action_key = {sql(action_key)} AND property_id IN ({quoted_codes});"
    )[0]["count"]
    watches = remote_d1_query(
        f"SELECT COUNT(*) AS count FROM captain_watch_items WHERE watch_key = {sql(watch_key)} AND property_id IN ({quoted_codes});"
    )[0]["count"]
    return {"actions": actions, "watch_items": watches}


def main() -> int:
    args = parse_args()
    workbook = Path(args.workbook).resolve()
    assignment_message = Path(args.assignment_message).resolve()
    due_date = parse_due_date(args.due_date)
    as_of = datetime.now(timezone.utc)
    run_id = as_of.strftime("%Y%m%dT%H%M%SZ")
    run_dir = Path(args.output_root) / run_id
    latest_run = latest_inventory_run()
    jobs = load_jobs(workbook, due_date, args.batch_id, assignment_message)
    if not jobs:
        raise RuntimeError("No property jobs were generated.")
    sql_text = render_sql(jobs, args.batch_id, latest_run, as_of)
    write_outputs(run_dir, jobs, sql_text)
    result = {
        "run_dir": str(run_dir),
        "workbook": str(workbook),
        "properties": len(jobs),
        "total_faqs": sum(job.total_faqs for job in jobs),
        "ready_for_review": sum(job.ready_for_review for job in jobs),
        "needs_fact_review": sum(job.needs_fact_review for job in jobs),
        "blank_proposed": sum(job.blank_proposed for job in jobs),
        "applied": False,
        "remote_counts": None,
    }
    if args.apply:
        result["remote_counts"] = apply_sql(jobs, args.batch_id, latest_run, as_of)
        result["applied"] = True
    (run_dir / "run_summary.json").write_text(json.dumps(result, indent=2), encoding="utf-8")
    print(json.dumps(result, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
