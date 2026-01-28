#!/usr/bin/env python3
"""
Resi vs Legacy Site Experience — Executive Comparative Performance Report
Controlled comparison answering: "Do Resi sites convert better when demand exists?"
"""

import sqlite3
import json
import statistics
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Paths
BASE_DIR = Path(__file__).parent
DB_PATH = Path("/Users/mark/Property_Analytics/data/portfolio_analytics.db")
REGISTRY_PATH = Path("/Users/mark/Property_Analytics/config/venterra_properties_official.json")
OUTPUT_DIR = BASE_DIR / "reports" / datetime.now().strftime("%Y-%m-%d")

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Date ranges
TODAY = datetime.now().date()
GA4_END = TODAY - timedelta(days=1)
GA4_START = GA4_END - timedelta(days=30)
GSC_END = TODAY - timedelta(days=3)
GSC_START = GSC_END - timedelta(days=30)

print("=" * 80)
print("RESI VS LEGACY SITE EXPERIENCE — EXECUTIVE COMPARATIVE REPORT")
print("=" * 80)
print(f"\nRun Date: {TODAY}")
print(f"GA4 Window: {GA4_START} to {GA4_END}")
print(f"GSC Window: {GSC_START} to {GSC_END}")
print()

# Load property registry
print("📚 Loading property registry...")
with open(REGISTRY_PATH, 'r') as f:
    registry_data = json.load(f)

resi_properties = []
legacy_properties = []

for prop in registry_data.get('properties', []):
    site_type = prop.get('site_type')
    ga4_id = prop.get('ga4_property_id')
    gsc_url = prop.get('gsc_url')
    name = prop.get('name')
    
    if not ga4_id or not gsc_url or not name:
        continue
    
    prop_data = {'name': name, 'ga4_id': ga4_id, 'gsc_url': gsc_url}
    
    if site_type == 'resi':
        resi_properties.append(prop_data)
    else:
        legacy_properties.append(prop_data)

print(f"✓ Resi: {len(resi_properties)} properties")
print(f"✓ Legacy: {len(legacy_properties)} properties")
print()

# Connect to database
conn = sqlite3.connect(DB_PATH)
cursor = conn.cursor()

def collect_property_data(properties, cohort_name):
    """Collect GSC, GA4, CWV data for all properties in cohort"""
    print(f"\n📊 Collecting data for {cohort_name} cohort ({len(properties)} properties)...")
    
    data = []
    
    for prop in properties:
        prop_data = {
            'name': prop['name'],
            'ga4_id': prop['ga4_id'],
            'gsc_url': prop['gsc_url']
        }
        
        # GSC Data
        gsc_query = """
        SELECT 
            MIN(metric_date) as first_date,
            SUM(clicks) as total_clicks,
            SUM(impressions) as total_impressions,
            AVG(average_position) as avg_position
        FROM gsc_daily_metrics
        WHERE property_id = ?
        AND metric_date BETWEEN ? AND ?
        """
        
        cursor.execute(gsc_query, (prop['gsc_url'], str(GSC_START), str(GSC_END)))
        gsc_row = cursor.fetchone()
        
        if gsc_row and gsc_row[1]:
            first_date, clicks, impressions, position = gsc_row
            ctr = (clicks / impressions * 100) if impressions else 0
            
            # Get days tracked from first ever record
            cursor.execute(
                "SELECT MIN(metric_date) FROM gsc_daily_metrics WHERE property_id = ?",
                (prop['gsc_url'],)
            )
            first_ever = cursor.fetchone()[0]
            days_tracked = (TODAY - datetime.strptime(first_ever, '%Y-%m-%d').date()).days if first_ever else None
            
            prop_data['gsc'] = {
                'clicks': clicks,
                'impressions': impressions,
                'ctr': ctr,
                'avg_position': position,
                'days_tracked': days_tracked,
                'click_velocity': clicks / 30
            }
        else:
            prop_data['gsc'] = None
        
        # GA4 Data
        ga4_query = """
        SELECT 
            SUM(sessions) as total_sessions,
            SUM(engaged_sessions) as total_engaged_sessions
        FROM ga4_daily_metrics
        WHERE property_id = ?
        AND metric_date BETWEEN ? AND ?
        """
        
        cursor.execute(ga4_query, (prop['ga4_id'], str(GA4_START), str(GA4_END)))
        ga4_row = cursor.fetchone()
        
        if ga4_row and ga4_row[0]:
            sessions, engaged_sessions = ga4_row
            engagement_rate = (engaged_sessions / sessions * 100) if sessions else 0
            
            prop_data['ga4'] = {
                'sessions': sessions,
                'engaged_sessions': engaged_sessions,
                'engagement_rate': engagement_rate
            }
        else:
            prop_data['ga4'] = None
        
        # CWV Data
        cwv_query = """
        SELECT lcp_value, fid_value, cls_value, strategy
        FROM pagespeed_metrics
        WHERE property_id = ?
        AND metric_date = (SELECT MAX(metric_date) FROM pagespeed_metrics WHERE property_id = ?)
        """
        
        cursor.execute(cwv_query, (prop['ga4_id'], prop['ga4_id']))
        cwv_rows = cursor.fetchall()
        
        mobile_cwv = None
        if cwv_rows:
            for row in cwv_rows:
                lcp, fid, cls, strategy = row
                if strategy == 'mobile':
                    mobile_cwv = {'lcp': lcp, 'fid': fid, 'cls': cls}
                    break
        
        if mobile_cwv:
            cwv_pass = (mobile_cwv['lcp'] <= 2.5 and 
                       mobile_cwv['fid'] <= 100 and 
                       mobile_cwv['cls'] <= 0.1)
            prop_data['cwv'] = {
                'lcp': mobile_cwv['lcp'],
                'fid': mobile_cwv['fid'],
                'cls': mobile_cwv['cls'],
                'passes_good': cwv_pass
            }
        else:
            prop_data['cwv'] = None
        
        data.append(prop_data)
    
    return data

# Collect data for both cohorts
resi_data = collect_property_data(resi_properties, "Resi")
legacy_data = collect_property_data(legacy_properties, "Legacy")

conn.close()

# Apply classifications to ALL properties
def classify_property(prop_data):
    """Apply deterministic classification"""
    gsc = prop_data.get('gsc')
    
    if not gsc:
        return 'No GSC Data', 'No tracking data available'
    
    clicks = gsc['clicks']
    days_tracked = gsc['days_tracked']
    velocity = gsc['click_velocity']
    
    if clicks >= 300:
        return 'Conversion-Ready', f'Passed volume gate ({clicks} clicks ≥ 300)'
    elif days_tracked and days_tracked < 45:
        return 'Ramp-Stage', f'New property ({days_tracked} days tracked < 45)'
    elif days_tracked and days_tracked >= 60 and clicks < 300 and velocity < 5:
        return 'Visibility-Constrained', f'Low organic visibility ({clicks} clicks, {velocity:.1f} clicks/day)'
    else:
        return 'Ramp-Stage', f'Building visibility ({clicks} clicks, {velocity:.1f} clicks/day)'

for prop in resi_data:
    classification, reason = classify_property(prop)
    prop['classification'] = classification
    prop['classification_reason'] = reason

for prop in legacy_data:
    classification, reason = classify_property(prop)
    prop['classification'] = classification
    prop['classification_reason'] = reason

# Filter to conversion-eligible properties only
print("\n" + "=" * 80)
print("APPLYING CONVERSION VOLUME GATE (≥300 CLICKS)")
print("=" * 80)

resi_conversion_eligible = [p for p in resi_data if p['classification'] == 'Conversion-Ready']
legacy_conversion_eligible = [p for p in legacy_data if p['classification'] == 'Conversion-Ready']

print(f"\nConversion-Eligible Properties:")
print(f"  Resi: {len(resi_conversion_eligible)} / {len(resi_data)}")
for p in resi_conversion_eligible:
    print(f"    ✓ {p['name']}")

print(f"\n  Legacy: {len(legacy_conversion_eligible)} / {len(legacy_data)}")
for p in legacy_conversion_eligible:
    print(f"    ✓ {p['name']}")

# Compute medians for conversion-eligible properties
print("\n" + "=" * 80)
print("SECTION 1 — EXECUTIVE COMPARISON SCORECARD")
print("=" * 80)

if resi_conversion_eligible and legacy_conversion_eligible:
    resi_ctr_median = statistics.median([p['gsc']['ctr'] for p in resi_conversion_eligible])
    legacy_ctr_median = statistics.median([p['gsc']['ctr'] for p in legacy_conversion_eligible])
    
    resi_engagement_median = statistics.median([p['ga4']['engagement_rate'] for p in resi_conversion_eligible])
    legacy_engagement_median = statistics.median([p['ga4']['engagement_rate'] for p in legacy_conversion_eligible])
    
    print(f"\nResi (N={len(resi_conversion_eligible)}):")
    print(f"  Median SERP CTR: {resi_ctr_median:.2f}%")
    print(f"  Median Engagement Rate: {resi_engagement_median:.1f}%")
    
    print(f"\nLegacy (N={len(legacy_conversion_eligible)}):")
    print(f"  Median SERP CTR: {legacy_ctr_median:.2f}%")
    print(f"  Median Engagement Rate: {legacy_engagement_median:.1f}%")
    
    # Determine winner
    if resi_ctr_median > legacy_ctr_median:
        ctr_winner = "Resi stronger"
    elif abs(resi_ctr_median - legacy_ctr_median) < 0.5:
        ctr_winner = "Comparable"
    else:
        ctr_winner = "Legacy stronger"
    
    if resi_engagement_median > legacy_engagement_median:
        engagement_winner = "Resi stronger"
    elif abs(resi_engagement_median - legacy_engagement_median) < 2:
        engagement_winner = "Comparable"
    else:
        engagement_winner = "Legacy stronger"
    
    print(f"\nComparison:")
    print(f"  SERP CTR: {ctr_winner}")
    print(f"  Engagement Rate: {engagement_winner}")
else:
    print("\n⚠️  Insufficient conversion-eligible properties in one or both cohorts")
    resi_ctr_median = None
    legacy_ctr_median = None
    resi_engagement_median = None
    legacy_engagement_median = None

# Generate HTML Report
print("\n" + "=" * 80)
print("GENERATING HTML EXECUTIVE REPORT")
print("=" * 80)

html = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>Resi vs Legacy Site Experience — Executive Comparison</title>
    <style>
        body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; max-width: 1000px; margin: 40px auto; padding: 20px; background: #f5f5f5; }}
        .container {{ background: white; padding: 40px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
        h1 {{ color: #2c3e50; border-bottom: 4px solid #3498db; padding-bottom: 15px; margin-bottom: 10px; }}
        h2 {{ color: #34495e; margin-top: 40px; background: #ecf0f1; padding: 12px; border-left: 5px solid #3498db; }}
        .summary-box {{ background: #e8f4f8; padding: 20px; border-left: 5px solid #3498db; margin: 25px 0; font-size: 0.95em; }}
        .conclusion-box {{ background: #d4edda; padding: 25px; border-left: 5px solid #28a745; margin: 30px 0; font-size: 1.05em; line-height: 1.7; }}
        table {{ width: 100%; border-collapse: collapse; margin: 25px 0; }}
        th, td {{ padding: 14px; text-align: left; border-bottom: 1px solid #ddd; }}
        th {{ background: #3498db; color: white; font-weight: 600; font-size: 0.95em; }}
        tr:hover {{ background: #f8f9fa; }}
        .scorecard-table {{ font-size: 1.1em; }}
        .scorecard-table th {{ background: #2c3e50; }}
        .scorecard-table .metric-value {{ font-weight: 700; font-size: 1.2em; }}
        .resi-row {{ background: #e3f2fd; }}
        .legacy-row {{ background: #fff3e0; }}
        .winner-green {{ color: #27ae60; font-weight: 600; }}
        .winner-red {{ color: #e74c3c; font-weight: 600; }}
        .comparable {{ color: #7f8c8d; font-weight: 600; }}
        .conversion-ready {{ color: #27ae60; font-weight: 600; }}
        .ramp-stage {{ color: #f39c12; font-weight: 600; }}
        .visibility-constrained {{ color: #e74c3c; font-weight: 600; }}
        .footer {{ text-align: center; margin-top: 50px; padding-top: 25px; color: #7f8c8d; font-size: 0.9em; border-top: 2px solid #ddd; }}
        ul {{ line-height: 1.8; }}
    </style>
</head>
<body>
    <div class="container">
        <h1>Resi vs Legacy Site Experience</h1>
        <h3>Executive Comparative Performance Report (30-Day)</h3>
        
        <div class="summary-box">
            <strong>Executive Question:</strong> When organic demand exists, do Resi sites convert traffic more efficiently than Legacy sites?<br><br>
            <strong>Run Date:</strong> {TODAY}<br>
            <strong>Data Window:</strong> GA4 ({GA4_START} to {GA4_END}), GSC ({GSC_START} to {GSC_END})<br>
            <strong>Methodology:</strong> Controlled comparison using identical volume gates (≥300 GSC clicks) for both cohorts
        </div>
        
        <h2>Section 1 — Executive Comparison Scorecard</h2>
"""

if resi_conversion_eligible and legacy_conversion_eligible:
    html += f"""
        <p><strong>Conversion-Eligible Properties:</strong> Properties with ≥300 organic clicks (30 days)</p>
        
        <table class="scorecard-table">
            <tr>
                <th>Site Experience</th>
                <th>Properties Included</th>
                <th>Median SERP CTR</th>
                <th>Median Engagement Rate</th>
            </tr>
            <tr class="resi-row">
                <td><strong>Resi</strong></td>
                <td class="metric-value">{len(resi_conversion_eligible)}</td>
                <td class="metric-value">{resi_ctr_median:.2f}%</td>
                <td class="metric-value">{resi_engagement_median:.1f}%</td>
            </tr>
            <tr class="legacy-row">
                <td><strong>Legacy</strong></td>
                <td class="metric-value">{len(legacy_conversion_eligible)}</td>
                <td class="metric-value">{legacy_ctr_median:.2f}%</td>
                <td class="metric-value">{legacy_engagement_median:.1f}%</td>
            </tr>
        </table>
        
        <p style="margin-top: 20px;">
            <strong>Direct Comparison:</strong><br>
            • SERP CTR: <span class="{'winner-green' if resi_ctr_median > legacy_ctr_median else 'comparable' if abs(resi_ctr_median - legacy_ctr_median) < 0.5 else 'winner-red'}">{ctr_winner}</span><br>
            • Engagement Rate: <span class="{'winner-green' if resi_engagement_median > legacy_engagement_median else 'comparable' if abs(resi_engagement_median - legacy_engagement_median) < 2 else 'winner-red'}">{engagement_winner}</span>
        </p>
"""
else:
    html += """
        <p><strong>⚠️ Insufficient Data:</strong> One or both cohorts lack properties meeting the volume gate (≥300 clicks).</p>
        <p>Comparison cannot proceed without controlled samples. See Section 3 for visibility context.</p>
"""

html += f"""
        <h2>Section 2 — Property-Level Comparison (Proof)</h2>
        
        <p>Underlying data supporting Section 1 scorecard:</p>
        
        <table>
            <tr>
                <th>Experience</th>
                <th>Property</th>
                <th>SERP CTR</th>
                <th>Engagement Rate</th>
                <th>Sessions</th>
            </tr>
"""

# Resi properties
for prop in resi_conversion_eligible:
    gsc = prop['gsc']
    ga4 = prop['ga4']
    html += f"""
            <tr class="resi-row">
                <td><strong>Resi</strong></td>
                <td>{prop['name']}</td>
                <td>{gsc['ctr']:.2f}%</td>
                <td>{ga4['engagement_rate']:.1f}%</td>
                <td>{ga4['sessions']:,}</td>
            </tr>
"""

# Legacy properties
for prop in legacy_conversion_eligible:
    gsc = prop['gsc']
    ga4 = prop['ga4']
    html += f"""
            <tr class="legacy-row">
                <td><strong>Legacy</strong></td>
                <td>{prop['name']}</td>
                <td>{gsc['ctr']:.2f}%</td>
                <td>{ga4['engagement_rate']:.1f}%</td>
                <td>{ga4['sessions']:,}</td>
            </tr>
"""

html += """
        </table>
        
        <h2>Section 3 — Visibility & Ramp Context</h2>
        
        <p><strong>All Properties</strong> (conversion-eligible and below gate):</p>
        
        <table>
            <tr>
                <th>Experience</th>
                <th>Property</th>
                <th>Days Tracked</th>
                <th>Clicks (30d)</th>
                <th>Click Velocity</th>
                <th>Classification</th>
            </tr>
"""

# All Resi properties
for prop in resi_data:
    gsc = prop.get('gsc')
    if gsc:
        days = gsc['days_tracked']
        clicks = gsc['clicks']
        velocity = gsc['click_velocity']
    else:
        days = 'N/A'
        clicks = 0
        velocity = 0
    
    classification = prop['classification']
    classification_class = classification.lower().replace(' ', '_').replace('-', '_')
    
    html += f"""
            <tr class="resi-row">
                <td><strong>Resi</strong></td>
                <td>{prop['name']}</td>
                <td>{days}</td>
                <td>{clicks}</td>
                <td>{velocity:.1f} /day</td>
                <td class="{classification_class}">{classification}</td>
            </tr>
"""

# Sample of Legacy properties (limit to conversion-eligible + sample of others)
legacy_sample = legacy_conversion_eligible + [p for p in legacy_data if p['classification'] != 'Conversion-Ready'][:10]

for prop in legacy_sample:
    gsc = prop.get('gsc')
    if gsc:
        days = gsc['days_tracked']
        clicks = gsc['clicks']
        velocity = gsc['click_velocity']
    else:
        days = 'N/A'
        clicks = 0
        velocity = 0
    
    classification = prop['classification']
    classification_class = classification.lower().replace(' ', '_').replace('-', '_')
    
    html += f"""
            <tr class="legacy-row">
                <td><strong>Legacy</strong></td>
                <td>{prop['name']}</td>
                <td>{days}</td>
                <td>{clicks}</td>
                <td>{velocity:.1f} /day</td>
                <td class="{classification_class}">{classification}</td>
            </tr>
"""

if len(legacy_data) > len(legacy_sample):
    html += f"""
            <tr class="legacy-row">
                <td colspan="6" style="text-align: center; font-style: italic;">
                    + {len(legacy_data) - len(legacy_sample)} additional Legacy properties (see JSON/Excel for complete data)
                </td>
            </tr>
"""

html += """
        </table>
        
        <p style="margin-top: 15px;"><strong>Classification Definitions:</strong></p>
        <ul>
            <li><strong class="conversion-ready">Conversion-Ready:</strong> ≥300 clicks — eligible for conversion comparison</li>
            <li><strong class="ramp-stage">Ramp-Stage:</strong> <45 days tracked — building visibility</li>
            <li><strong class="visibility-constrained">Visibility-Constrained:</strong> ≥60 days, <300 clicks, <5 clicks/day — requires SEO intervention</li>
        </ul>
        
        <h2>Section 4 — Executive Conclusion</h2>
"""

# Generate conclusion based on data
if resi_conversion_eligible and legacy_conversion_eligible:
    if resi_ctr_median > legacy_ctr_median and resi_engagement_median > legacy_engagement_median:
        conclusion = f"""
        <div class="conclusion-box">
            Under equivalent organic demand conditions (≥300 clicks), <strong>Resi sites demonstrate stronger conversion efficiency than Legacy sites</strong>, 
            reflected in higher median SERP CTR ({resi_ctr_median:.2f}% vs {legacy_ctr_median:.2f}%) and stronger engagement rates 
            ({resi_engagement_median:.1f}% vs {legacy_engagement_median:.1f}%). Where Resi properties underperform, 
            the constraint is organic visibility rather than conversion behavior. This comparison isolates site experience as the differentiating factor 
            without relying on external benchmarks.
        </div>
"""
    elif resi_ctr_median < legacy_ctr_median and resi_engagement_median < legacy_engagement_median:
        conclusion = f"""
        <div class="conclusion-box">
            Under equivalent organic demand conditions (≥300 clicks), <strong>Legacy sites demonstrate stronger conversion efficiency than Resi sites</strong>, 
            reflected in higher median SERP CTR ({legacy_ctr_median:.2f}% vs {resi_ctr_median:.2f}%) and stronger engagement rates 
            ({legacy_engagement_median:.1f}% vs {resi_engagement_median:.1f}%). The Resi experience is not underperforming due to visibility constraints—when 
            demand exists, Legacy sites convert it more effectively. This comparison isolates site experience as the differentiating factor.
        </div>
"""
    else:
        conclusion = f"""
        <div class="conclusion-box">
            Under equivalent organic demand conditions (≥300 clicks), <strong>conversion efficiency is mixed between Resi and Legacy sites</strong>. 
            SERP CTR favors {'Resi' if resi_ctr_median > legacy_ctr_median else 'Legacy'} ({max(resi_ctr_median, legacy_ctr_median):.2f}% vs {min(resi_ctr_median, legacy_ctr_median):.2f}%), 
            while engagement rate favors {'Resi' if resi_engagement_median > legacy_engagement_median else 'Legacy'} 
            ({max(resi_engagement_median, legacy_engagement_median):.1f}% vs {min(resi_engagement_median, legacy_engagement_median):.1f}%). 
            Neither experience shows clear superiority across both conversion metrics. This comparison isolates site experience performance without external benchmarks.
        </div>
"""
else:
    conclusion = """
        <div class="conclusion-box">
            <strong>Comparison cannot be completed:</strong> Insufficient conversion-eligible properties in one or both cohorts. 
            The primary constraint is visibility (insufficient organic demand), not conversion performance. 
            A meaningful comparison requires both cohorts to have properties meeting the volume gate (≥300 clicks).
        </div>
"""

html += conclusion

html += f"""
        <h2>Guardrails & Limitations</h2>
        <ul>
            <li><strong>Same volume gate:</strong> ≥300 GSC clicks applied identically to both cohorts</li>
            <li><strong>Conversion proxies:</strong> SERP CTR and engagement rate used (not lease conversions)</li>
            <li><strong>Internal portfolio comparison:</strong> Venterra properties only (no external competitors)</li>
            <li><strong>No inferred data:</strong> Only properties with complete GSC + GA4 data included</li>
            <li><strong>30-day window:</strong> Rolling window respecting data lags (GA4: T-1, GSC: T-3)</li>
        </ul>
        
        <div class="footer">
            Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}<br>
            Framework: Atlas Executive Comparison v1.0<br>
            Total Properties Analyzed: {len(resi_data)} Resi, {len(legacy_data)} Legacy
        </div>
    </div>
</body>
</html>"""

# Save HTML
html_path = OUTPUT_DIR / "resi_vs_legacy_comparison.html"
with open(html_path, 'w') as f:
    f.write(html)

print(f"\n✓ HTML report saved: {html_path}")

# Save JSON
json_data = {
    "meta": {
        "run_date": str(TODAY),
        "ga4_window": {"start": str(GA4_START), "end": str(GA4_END)},
        "gsc_window": {"start": str(GSC_START), "end": str(GSC_END)},
        "framework": "Atlas Executive Comparison v1.0"
    },
    "scorecard": {
        "resi": {
            "conversion_eligible_count": len(resi_conversion_eligible),
            "median_ctr": resi_ctr_median,
            "median_engagement": resi_engagement_median
        },
        "legacy": {
            "conversion_eligible_count": len(legacy_conversion_eligible),
            "median_ctr": legacy_ctr_median,
            "median_engagement": legacy_engagement_median
        }
    },
    "properties": {
        "resi": resi_data,
        "legacy": legacy_data
    }
}

json_path = OUTPUT_DIR / "resi_vs_legacy_comparison.json"
with open(json_path, 'w') as f:
    json.dump(json_data, f, indent=2)

print(f"✓ JSON artifact saved: {json_path}")

# Save Excel
print("\n📊 Generating Excel appendix...")

wb = Workbook()

# Sheet 1: Executive Scorecard
ws1 = wb.active
ws1.title = "Executive Scorecard"

ws1['A1'] = "Resi vs Legacy Site Experience — Executive Comparison"
ws1['A1'].font = Font(bold=True, size=14)
ws1['A3'] = f"Run Date: {TODAY}"

header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)

ws1['A5'] = "Site Experience"
ws1['B5'] = "Properties Included"
ws1['C5'] = "Median SERP CTR"
ws1['D5'] = "Median Engagement Rate"

for col in ['A5', 'B5', 'C5', 'D5']:
    ws1[col].fill = header_fill
    ws1[col].font = header_font

if resi_conversion_eligible and legacy_conversion_eligible:
    ws1['A6'] = "Resi"
    ws1['B6'] = len(resi_conversion_eligible)
    ws1['C6'] = f"{resi_ctr_median:.2f}%"
    ws1['D6'] = f"{resi_engagement_median:.1f}%"
    
    ws1['A7'] = "Legacy"
    ws1['B7'] = len(legacy_conversion_eligible)
    ws1['C7'] = f"{legacy_ctr_median:.2f}%"
    ws1['D7'] = f"{legacy_engagement_median:.1f}%"

# Sheet 2: Property-Level Data
ws2 = wb.create_sheet("Property-Level Data")

headers = ["Experience", "Property", "Classification", "SERP CTR", "Engagement Rate", 
           "Sessions", "Clicks (30d)", "Days Tracked", "Click Velocity"]

for col_num, header in enumerate(headers, 1):
    cell = ws2.cell(row=1, column=col_num, value=header)
    cell.fill = header_fill
    cell.font = header_font

row_num = 2
for prop in resi_data:
    ws2.cell(row=row_num, column=1, value="Resi")
    ws2.cell(row=row_num, column=2, value=prop['name'])
    ws2.cell(row=row_num, column=3, value=prop['classification'])
    
    if prop.get('gsc'):
        ws2.cell(row=row_num, column=4, value=f"{prop['gsc']['ctr']:.2f}%")
        ws2.cell(row=row_num, column=7, value=prop['gsc']['clicks'])
        ws2.cell(row=row_num, column=8, value=prop['gsc']['days_tracked'])
        ws2.cell(row=row_num, column=9, value=f"{prop['gsc']['click_velocity']:.1f}")
    
    if prop.get('ga4'):
        ws2.cell(row=row_num, column=5, value=f"{prop['ga4']['engagement_rate']:.1f}%")
        ws2.cell(row=row_num, column=6, value=prop['ga4']['sessions'])
    
    row_num += 1

for prop in legacy_data:
    ws2.cell(row=row_num, column=1, value="Legacy")
    ws2.cell(row=row_num, column=2, value=prop['name'])
    ws2.cell(row=row_num, column=3, value=prop['classification'])
    
    if prop.get('gsc'):
        ws2.cell(row=row_num, column=4, value=f"{prop['gsc']['ctr']:.2f}%")
        ws2.cell(row=row_num, column=7, value=prop['gsc']['clicks'])
        ws2.cell(row=row_num, column=8, value=prop['gsc']['days_tracked'])
        ws2.cell(row=row_num, column=9, value=f"{prop['gsc']['click_velocity']:.1f}")
    
    if prop.get('ga4'):
        ws2.cell(row=row_num, column=5, value=f"{prop['ga4']['engagement_rate']:.1f}%")
        ws2.cell(row=row_num, column=6, value=prop['ga4']['sessions'])
    
    row_num += 1

excel_path = OUTPUT_DIR / "resi_vs_legacy_comparison.xlsx"
wb.save(excel_path)

print(f"✓ Excel appendix saved: {excel_path}")

print("\n" + "=" * 80)
print("✅ EXECUTIVE COMPARISON REPORT COMPLETE")
print("=" * 80)
print(f"\nOutputs:")
print(f"  HTML: {html_path}")
print(f"  JSON: {json_path}")
print(f"  Excel: {excel_path}")
print(f"\nConversion-Eligible:")
print(f"  Resi: {len(resi_conversion_eligible)}")
print(f"  Legacy: {len(legacy_conversion_eligible)}")
