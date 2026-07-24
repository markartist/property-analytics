import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from keeper_marketing_ops_import import (
    ImportConfig,
    build_keeper_import_json,
    load_config,
    read_csv,
    read_venterra_marketing_logins_workbook,
    transform_rows,
    validate_records,
)


class KeeperMarketingOpsImportTests(unittest.TestCase):
    def test_csv_transforms_headers_and_custom_fields(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            csv_path = Path(temp_dir) / "fake_credentials.csv"
            csv_path.write_text(
                "Name,Email,Password,Website,Notes,Folder,Owner\n"
                "Demo Portal,demo@example.com,fake-pass,https://example.com,Only fake data,Ops,Marketing\n",
                encoding="utf-8",
            )

            headers, rows = read_csv(csv_path)
            records, mapped = transform_rows(headers, rows)

        self.assertEqual(mapped, ["folder", "notes", "password", "title", "url", "username"])
        self.assertEqual(len(records), 1)
        self.assertEqual(records[0].title, "Demo Portal")
        self.assertEqual(records[0].username, "demo@example.com")
        self.assertEqual(records[0].url, "https://example.com")
        self.assertEqual(records[0].custom_fields["Owner"], "Marketing")
        self.assertEqual(records[0].custom_fields["Source Folder"], "Ops")

    def test_validation_flags_blank_password_duplicate_and_bad_url(self):
        headers = ["Title", "Username", "Password", "URL"]
        rows = [
            {"Title": "Demo", "Username": "demo@example.com", "Password": "", "URL": "not-a-url"},
            {"Title": "Demo", "Username": "demo@example.com", "Password": "fake-pass", "URL": "not-a-url"},
        ]
        records, mapped = transform_rows(headers, rows)
        result = validate_records(records, mapped, ImportConfig())

        self.assertEqual(len(result.errors), 1)
        self.assertIn("Password is blank", result.errors[0].message)
        self.assertEqual(len(result.warnings), 3)
        self.assertTrue(any("Duplicate" in warning.message for warning in result.warnings))
        self.assertTrue(all("demo@example.com" not in warning.message for warning in result.warnings))

    def test_keeper_json_uses_verified_shared_folder_shape(self):
        headers = ["Title", "Login", "Password", "URL", "Department"]
        rows = [
            {
                "Title": "Demo",
                "Login": "demo@example.com",
                "Password": "fake-pass",
                "URL": "https://example.com",
                "Department": "Marketing",
            }
        ]
        records, mapped = transform_rows(headers, rows)
        result = validate_records(records, mapped, ImportConfig())
        self.assertTrue(result.ok)

        payload = build_keeper_import_json(records, ImportConfig())

        self.assertEqual(payload["shared_folders"][0]["path"], "Marketing Ops Shared Credentials")
        self.assertEqual(payload["shared_folders"][0]["permissions"][0]["name"], "Marketing Ops")
        self.assertFalse(payload["shared_folders"][0]["can_edit"])
        self.assertFalse(payload["shared_folders"][0]["can_share"])
        self.assertFalse(payload["shared_folders"][0]["permissions"][0]["manage_users"])
        self.assertFalse(payload["shared_folders"][0]["permissions"][0]["manage_records"])
        self.assertEqual(
            payload["records"][0]["folders"][0]["shared_folder"],
            "Marketing Ops Shared Credentials",
        )
        self.assertEqual(payload["records"][0]["custom_fields"]["Department"], "Marketing")
        json.dumps(payload)

    def test_config_example_loads_defaults(self):
        config = load_config(Path("config.example.yaml"))

        self.assertEqual(config.team_name, "Marketing Ops")
        self.assertEqual(config.shared_folder_name, "Marketing Ops Shared Credentials")
        self.assertFalse(config.can_edit)
        self.assertFalse(config.can_share)
        self.assertFalse(config.manage_users)
        self.assertFalse(config.manage_records)

    def test_venterra_marketing_logins_workbook_maps_known_sheets(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "venterra_logins.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Venterra Master Log ins"
            ws.append(["Site", "URL", "Username", "Password", "Owner"])
            ws.append(["Demo Portal", "https://example.com", "demo@example.com", "fake-pass", "Marketing"])
            instagram = wb.create_sheet("Instagram Logins")
            instagram.append(["Property", "Username", "Password", "Social URL", "Managed By:"])
            instagram.append(["Demo Property", "demo_ig", "fake-pass-2", "https://instagram.com/demo", "Agency"])
            wb.save(workbook_path)

            records, mapped = read_venterra_marketing_logins_workbook(workbook_path)

        self.assertEqual(mapped, ["password", "title"])
        self.assertEqual(len(records), 2)
        self.assertEqual(records[0].title, "Demo Portal")
        self.assertEqual(records[0].url, "https://example.com")
        self.assertEqual(records[0].custom_fields["Source Sheet"], "Venterra Master Log ins")
        self.assertEqual(records[0].custom_fields["Owner"], "Marketing")
        self.assertEqual(records[1].custom_fields["Managed By:"], "Agency")

    def test_venterra_reference_records_can_have_blank_passwords(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / "venterra_logins.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Venterra Master Log ins"
            ws.append(["Site", "URL", "Username", "Password"])
            ws.append(["Demo Portal", "https://example.com", "demo@example.com", "fake-pass"])
            youtube = wb.create_sheet("youtube urls")
            youtube.append(["Community", "URL: Community's YouTube Channel"])
            youtube.append(["Demo Property", "https://youtube.com/@demo"])
            wb.save(workbook_path)

            records, mapped = read_venterra_marketing_logins_workbook(
                workbook_path,
                include_reference_records=True,
            )
            result = validate_records(records, mapped, ImportConfig())

        self.assertTrue(result.ok)
        reference = [record for record in records if record.custom_fields["Record Kind"] == "reference_url"]
        self.assertEqual(len(reference), 1)
        self.assertFalse(reference[0].password_required)
        self.assertEqual(reference[0].password, "")


if __name__ == "__main__":
    unittest.main()
